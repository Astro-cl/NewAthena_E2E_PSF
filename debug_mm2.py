#!/usr/bin/env python3
"""Debug: Check what's actually in the cells."""

import openpyxl

wb = openpyxl.load_workbook('Distributions/test10kev.xlsx')
aeff = wb['A_eff']

print("Checking cells in A_eff sheet:")
print("\nRows 1-3, Columns A-Q:")

for row_num in range(1, 4):
    print(f"\nRow {row_num}:")
    for col_idx in range(1, 18):
        cell = aeff.cell(row_num, col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if cell.data_type == 'f':
            print(f"  {col_letter}{row_num}: [FORMULA] {str(cell.value)[:60]}")
        else:
            print(f"  {col_letter}{row_num}: {cell.value} (type: {cell.data_type})")

# Check lookup table
print("\n\nLookup table (S24:AA38):")
print("Row | Col S | Col T-AA...")
ws = wb['A_eff']  # The lookup table is on A_eff sheet
for row_num in range(24, 27):
    key_cell = ws.cell(row_num, 19)  # Column S
    print(f"{row_num:3} | {key_cell.value}")
