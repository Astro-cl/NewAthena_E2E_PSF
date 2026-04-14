import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Distributions/test10kev.xlsx', data_only=False)
ws = wb['A_eff']

print("Columns D-E (standard presets area):")
print("Row  | Column D (Standard) | Column E (Values)")
print("-" * 60)
for r in range(1, 21):
    cell_d = ws.cell(r, 4).value
    cell_e = ws.cell(r, 5).value
    d_str = str(cell_d)[:30] if cell_d else ""
    e_str = str(cell_e)[:30] if cell_e else ""
    print(f"{r:3}  | {d_str:20} | {e_str}")
