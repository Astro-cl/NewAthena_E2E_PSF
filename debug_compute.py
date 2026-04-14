#!/usr/bin/env python3
"""Debug the compute process."""

import openpyxl

wb = openpyxl.load_workbook('Distributions/test10kev.xlsx')
aeff = wb['A_eff']
mm_config = wb['MM configuration']

# Check MM to Row mapping
mm_to_row = {}
for row in mm_config.iter_rows(min_row=2, max_row=10, min_col=3, max_col=4):
    row_cell = row[0]
    mm_cell = row[1]
    if mm_cell.value is not None and row_cell.value is not None:
        try:
            mm = int(float(mm_cell.value))
            row_num = int(float(row_cell.value))
            mm_to_row[mm] = row_num
            print(f"MM {mm} -> Row {row_num}")
        except Exception as e:
            print(f"Error: {e}")

# Check lookup data
print("\nSource data rows 5-6:")
print(f"Row 5, Col S (margin label): {aeff.cell(5, 19).value}")
for col_idx in range(20, 25):
    print(f"  Row 5, Col {openpyxl.utils.get_column_letter(col_idx)}: {aeff.cell(5, col_idx).value}")

print(f"\nRow 6, Col S (focal length): {aeff.cell(6, 19).value}")
for col_idx in range(20, 25):
    print(f"  Row 6, Col {openpyxl.utils.get_column_letter(col_idx)} (energy data): {aeff.cell(6, col_idx).value}")

# What value would be computed?
focal_len = aeff.cell(6, 19).value
base_val = aeff.cell(6, 20).value  # T6
margin = aeff.cell(5, 20).value    # T5

print(f"\nCompute example for row 6, col T:")
print(f"  Focal length (S6): {focal_len}")
print(f"  Base value (T6): {base_val}")
print(f"  System margin (T5): {margin}")

if focal_len and base_val and margin:
    try:
        adjusted = float(base_val) * (1.0 - float(margin)) / float(focal_len)
        print(f"  Calculated: {base_val} * (1 - {margin}) / {focal_len} = {adjusted}")
    except Exception as e:
        print(f"  Error: {e}")
