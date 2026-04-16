from openpyxl import load_workbook
import os
agg_fp='Exports/Export_reallistic_20260416_113827/Aggregated_results_reallistic.xlsx'
wb=load_workbook(agg_fp,data_only=False)
ws=wb['Aggregated']
# find row index for configuration_name
target='2__Ang__20__E__7__Def__0'
row_idx=None
for r in range(2, ws.max_row+1):
    if ws.cell(row=r, column=2).value == target:
        row_idx=r
        break
print('row_idx',row_idx)
# compute sums from package workbook
pkg_xls='Exports/Export_reallistic_20260416_113827/2__Ang__20__E__7__Def__0_reallistic_20260416_113954/2__Ang__20__E__7__Def__0_reallistic.xlsx'
if os.path.exists(pkg_xls):
    wb2=load_workbook(pkg_xls,data_only=True)
    ws2=wb2['A_eff']
    sumb=sumc=0.0
    for r in range(2, ws2.max_row+1):
        mm=ws2.cell(row=r, column=1).value
        if mm is None: continue
        try: float(mm)
        except: continue
        vB=ws2.cell(row=r, column=2).value
        vC=ws2.cell(row=r, column=3).value
        try: sumb += float(vB) if vB is not None else 0.0
        except: pass
        try: sumc += float(vC) if vC is not None else 0.0
        except: pass
    print('computed sumb,sumc', sumb, sumc)
    # patch aggregated workbook
    if row_idx:
        ws.cell(row=row_idx, column=6).value = sumb
        ws.cell(row=row_idx, column=7).value = sumc
        ws.cell(row=row_idx, column=8).value = (None if sumb==0 else 1.0-(sumc/sumb))
        wb.save(agg_fp)
        print('patched aggregated')
else:
    print('package workbook not found', pkg_xls)
