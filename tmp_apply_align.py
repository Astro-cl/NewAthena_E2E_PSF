from sensivitiy.sensitivity_run import load_standard_alignment_defs, _parse_standard_dist_spec
from gui_distributions import DATA_TYPES, generate_data_from_distributions
from pathlib import Path
from openpyxl import load_workbook
p=Path('Distributions/Test_Distribution.xlsx')
std=load_standard_alignment_defs(p)
specs=std.get('Standard medialario')
params_align={}
for p_label in DATA_TYPES['Alignment']['params']:
    spec_str = specs.get(p_label)
    kind,a,b = _parse_standard_dist_spec(spec_str)
    if kind=='fixed': params_align[p_label]=('fixed',a,0.0)
    elif kind=='gaussian': params_align[p_label]=('gaussian',a,b)
    else: params_align[p_label]=('uniform',a,b)
print('params_align built')
df_align_gen=generate_data_from_distributions(params_align, 8, DATA_TYPES['Alignment'])
print('df sample:\n', df_align_gen.head())
files=list(Path('sensivitiy/input').glob('*Standard_medialario*.xlsx'))
print('found', len(files), 'files')
if files:
    f=files[-1]
    print('writing to', f.name)
    wb=load_workbook(f)
    ws=wb['Alignment']
    for idx in range(8):
        for j in range(2,6):
            val=float(df_align_gen.iloc[idx, j-2])
            ws.cell(row=2+idx, column=j, value=val)
    wb.save(f)
    # read back
    wb2=load_workbook(f)
    ws2=wb2['Alignment']
    for row in ws2.iter_rows(min_row=2, max_row=6, min_col=2, max_col=5):
        print([c.value for c in row])
else:
    print('no files')
