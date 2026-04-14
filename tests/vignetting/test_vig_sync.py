import os
import pytest
from openpyxl import load_workbook
from main import load_gaussians_from_excel


def test_vignette_sync_with_test1():
    """Verify that for Distributions/test1.xlsx the returned DataFrame's
    aeff_adjusted equals aeff_base * (rotazi_B * rotrad_B) for every MM with
    a valid base value.
    """
    p = os.path.join('Distributions', 'test1.xlsx')
    df = load_gaussians_from_excel(p)

    wb = load_workbook(p, data_only=True)
    assert 'Vignetting rotazi' in wb.sheetnames
    assert 'Vignetting rotrad' in wb.sheetnames
    ws_azi = wb['Vignetting rotazi']
    ws_rad = wb['Vignetting rotrad']

    vig_azi = {}
    vig_rad = {}
    for r in range(1, ws_azi.max_row + 1):
        a = ws_azi.cell(row=r, column=1).value
        b = ws_azi.cell(row=r, column=2).value
        if a is None or b is None:
            continue
        try:
            key = int(a) if isinstance(a, (int, float)) or (isinstance(a, str) and str(a).strip().isdigit()) else None
        except Exception:
            key = None
        if key is not None:
            vig_azi[key] = float(b)
    for r in range(1, ws_rad.max_row + 1):
        a = ws_rad.cell(row=r, column=1).value
        b = ws_rad.cell(row=r, column=2).value
        if a is None or b is None:
            continue
        try:
            key = int(a) if isinstance(a, (int, float)) or (isinstance(a, str) and str(a).strip().isdigit()) else None
        except Exception:
            key = None
        if key is not None:
            vig_rad[key] = float(b)

    # Read MM configuration to map MM # -> Position #
    mmcfg = None
    try:
        import pandas as pd
        mmcfg = pd.read_excel(p, sheet_name='MM configuration', engine='openpyxl')
    except Exception:
        pytest.skip('MM configuration sheet not available')

    mm_to_pos = {}
    if 'MM #' in mmcfg.columns:
        for order_i, (_, row) in enumerate(mmcfg.iterrows()):
            mm_num = row.get('MM #')
            if mm_num is None or pytest.helpers.isna(mm_num) if hasattr(pytest, 'helpers') else False:
                continue
            mm_num_i = int(mm_num)
            if 'Position #' in mmcfg.columns:
                pos_val = row.get('Position #')
                if pos_val is not None:
                    try:
                        mm_to_pos[mm_num_i] = int(float(pos_val))
                    except Exception:
                        pass
                if mm_num_i not in mm_to_pos:
                    mm_to_pos[mm_num_i] = int(order_i) + 1
            else:
                mm_to_pos[mm_num_i] = int(order_i) + 1

    # Check each MM
    for _, row in df.iterrows():
        try:
            mm = int(row['MM #'])
        except Exception:
            continue
        base = float(row.get('aeff_base', 0.0) or 0.0)
        adj = float(row.get('aeff_adjusted', 0.0) or 0.0)
        if base <= 0:
            continue
        pos = mm_to_pos.get(mm)
        if pos is None:
            continue
        r = vig_rad.get(pos)
        a = vig_azi.get(pos)
        if r is None or a is None:
            continue
        prod = float(r) * float(a)
        assert abs((adj / base) - prod) < 1e-9, f"Mismatch for MM {mm}: adj/base={adj/base} prod={prod}"
