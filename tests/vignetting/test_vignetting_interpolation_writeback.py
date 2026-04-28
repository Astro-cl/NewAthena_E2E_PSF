import os
import sys
import tempfile
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook

# Ensure repository root is importable when pytest is invoked from subfolders.
ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from main import load_gaussians_from_excel


def _set_headers(ws, headers):
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=value)


def _build_minimal_workbook(path):
    wb = Workbook()

    ws = wb.active
    ws.title = "MM_PSF"
    _set_headers(
        ws,
        [
            "MM #",
            "m_rad [arcsec]",
            "m_azi [arcsec]",
            "sigma_rad [arcsec]",
            "sigma_azi [arcsec]",
        ],
    )
    ws.append([1, 0.0, 0.0, 4.3, 4.3])

    ws = wb.create_sheet("A_eff")
    _set_headers(ws, ["MM #", "7 keV"])
    ws.append([1, 1.0])
    ws.cell(row=2, column=4, value=7.0)  # D2 selected energy marker

    ws = wb.create_sheet("MM configuration")
    _set_headers(ws, ["MM #", "Position #", "Row #", "x_MM [m]", "y_MM [m]", "r_MM [m]"])
    ws.append([1, 1, 1, 1.0, 0.0, 1.0])

    ws = wb.create_sheet("Alignment")
    _set_headers(ws, ["Position #", "d_align_rotazi [arcsec]", "d_align_rotrad [arcsec]"])
    ws.append([1, -5.0, -5.0])

    # New per-row/energy vignetting layout uses H/I/J/K columns.
    # Keep column A as Position # so column B writeback can be validated.
    ws = wb.create_sheet("MM vignetting rotazi")
    _set_headers(ws, ["Position #", "vig_factor", "Selected energy [keV]", "D", "E", "F", "G", "cfg_row", "rot_arcmin", "energy", "factor"])
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=3, value=7.0)
    ws.cell(row=2, column=8, value=1)
    ws.cell(row=2, column=9, value=0.0)
    ws.cell(row=2, column=10, value=7.0)
    ws.cell(row=2, column=11, value=1.0)
    ws.cell(row=3, column=8, value=1)
    ws.cell(row=3, column=9, value=10.0 / 60.0)
    ws.cell(row=3, column=10, value=7.0)
    ws.cell(row=3, column=11, value=0.8)

    ws = wb.create_sheet("MM vignetting rotrad")
    _set_headers(ws, ["Position #", "vig_factor", "Selected energy [keV]", "D", "E", "F", "G", "cfg_row", "rot_arcmin", "energy", "factor"])
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=3, value=7.0)
    for row_idx, x_arcsec, y_factor in (
        (2, -10.0, 0.5),
        (3, 0.0, 1.0),
        (4, 10.0, 1.5),
    ):
        ws.cell(row=row_idx, column=8, value=1)
        ws.cell(row=row_idx, column=9, value=x_arcsec / 60.0)
        ws.cell(row=row_idx, column=10, value=7.0)
        ws.cell(row=row_idx, column=11, value=y_factor)

    wb.save(path)


def test_vignetting_per_row_energy_interpolation_and_writeback():
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    try:
        _build_minimal_workbook(tmp.name)

        df = load_gaussians_from_excel(tmp.name)

        expected_azi = float(np.interp(abs(-5.0), np.array([0.0, 10.0]), np.array([1.0, 0.8])))
        expected_rad = float(np.interp(abs(-5.0), np.array([-10.0, 0.0, 10.0]), np.array([0.5, 1.0, 1.5])))
        expected_combined = expected_azi * expected_rad

        assert "aeff_vig_factor_azi" in df.columns
        assert "aeff_vig_factor_rad" in df.columns
        assert "aeff_vig_factor" in df.columns
        assert "weight" in df.columns

        got_azi = float(df.loc[0, "aeff_vig_factor_azi"])
        got_rad = float(df.loc[0, "aeff_vig_factor_rad"])
        got_combined = float(df.loc[0, "aeff_vig_factor"])
        got_weight = float(df.loc[0, "weight"])

        assert np.isclose(got_azi, expected_azi, atol=1e-9)
        assert np.isclose(got_rad, expected_rad, atol=1e-9)
        assert np.isclose(got_combined, expected_combined, atol=1e-9)
        assert np.isclose(got_weight, expected_combined, atol=1e-9)

        # rotrad interpolation uses abs(total) at interpolation-time.
        assert np.isclose(got_rad, 1.25, atol=1e-9)

        wb = load_workbook(tmp.name, data_only=True)
        try:
            ws_azi = wb["MM vignetting rotazi"]
            ws_rad = wb["MM vignetting rotrad"]
            written_azi = float(ws_azi.cell(row=2, column=2).value)
            written_rad = float(ws_rad.cell(row=2, column=2).value)
            written_azi_total = float(ws_azi.cell(row=2, column=4).value)
            written_rad_total = float(ws_rad.cell(row=2, column=4).value)
            header_azi_total = ws_azi.cell(row=1, column=4).value
            header_rad_total = ws_rad.cell(row=1, column=4).value
        finally:
            wb.close()

        assert np.isclose(written_azi, expected_azi, atol=1e-9)
        assert np.isclose(written_rad, expected_rad, atol=1e-9)
        assert written_azi_total == -5.0
        assert written_rad_total == -5.0
        assert header_azi_total == "rotazi_total"
        assert header_rad_total == "rotrad_total"
    finally:
        try:
            os.remove(tmp.name)
        except Exception:
            pass
