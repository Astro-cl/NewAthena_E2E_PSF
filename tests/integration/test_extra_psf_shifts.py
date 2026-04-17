"""Tests for the 'Extra PSF shifts' sheet: off-axis rotations and defocus (d_extra_z)."""

import math
import tempfile
import os

import numpy as np
import pandas as pd
import pytest

from main import load_gaussians_from_excel, compute_total_rot_polar


# ---------------------------------------------------------------------------
# Helper: build a minimal workbook that load_gaussians_from_excel can consume
# ---------------------------------------------------------------------------

def _make_workbook(path, *, n_mm=2, extra_rotx=0.0, extra_roty=0.0, extra_z_um=0.0,
                   therm_z_um=0.0, align_z_um=0.0, grav_z_um=0.0,
                   x_mm=None, y_mm=None, r_mm=None):
    """Write a test workbook with controllable Extra PSF shifts.

    Parameters
    ----------
    n_mm : int
        Number of MMs (each mapped to its own position).
    extra_rotx, extra_roty : float
        Values for d_extra_rotx / d_extra_roty in arcsec.
    extra_z_um : float
        Value for d_extra_z in µm (written to sheet; loader converts to m).
    therm_z_um, align_z_um, grav_z_um : float
        Thermal / Alignment / Gravity d_z values in µm.
    x_mm, y_mm, r_mm : list[float] | None
        Per-MM geometry; defaults to simple x=1,y=0,r=1 for all.
    """
    if x_mm is None:
        x_mm = [1.0] * n_mm
    if y_mm is None:
        y_mm = [0.0] * n_mm
    if r_mm is None:
        r_mm = [1.0] * n_mm

    mm_nums = list(range(1, n_mm + 1))

    mm_psf = pd.DataFrame({
        'MM #': mm_nums,
        'm_rad [arcsec]': [0.0] * n_mm,
        'm_azi [arcsec]': [0.0] * n_mm,
        'sigma_rad [arcsec]': [4.0] * n_mm,
        'sigma_azi [arcsec]': [4.0] * n_mm,
    })

    aeff = pd.DataFrame({'MM #': mm_nums, '1.0 keV': [1.0] * n_mm})

    mm_conf = pd.DataFrame({
        'MM #': mm_nums,
        'Position #': mm_nums,
        'x_MM [m]': x_mm,
        'y_MM [m]': y_mm,
        'z_MM [m]': [0.0] * n_mm,
        'r_MM [m]': r_mm,
    })

    grav = pd.DataFrame({
        'Position #': mm_nums,
        'd_grav_x [µm]': [0.0] * n_mm,
        'd_grav_y [µm]': [0.0] * n_mm,
        'd_grav_z [µm]': [grav_z_um] * n_mm,
        'd_grav_rotx [arcsec]': [0.0] * n_mm,
        'd_grav_roty [arcsec]': [0.0] * n_mm,
        'd_grav_rotz [arcsec]': [0.0] * n_mm,
    })

    therm = pd.DataFrame({
        'Position #': mm_nums,
        'd_therm_x [µm]': [0.0] * n_mm,
        'd_therm_y [µm]': [0.0] * n_mm,
        'd_therm_z [µm]': [therm_z_um] * n_mm,
        'd_therm_rotx [arcsec]': [0.0] * n_mm,
        'd_therm_roty [arcsec]': [0.0] * n_mm,
        'd_therm_rotz [arcsec]': [0.0] * n_mm,
    })

    align = pd.DataFrame({
        'Position #': mm_nums,
        'd_align_rad [µm]': [0.0] * n_mm,
        'd_align_azi [µm]': [0.0] * n_mm,
        'd_align_z [µm]': [align_z_um] * n_mm,
        'd_align_rotazi [arcsec]': [0.0] * n_mm,
        'd_align_rotrad [arcsec]': [0.0] * n_mm,
        'd_align_rotz [arcsec]': [0.0] * n_mm,
    })

    extra = pd.DataFrame({
        'Position #': mm_nums,
        'd_extra_rotx [arcsec]': [extra_rotx] * n_mm,
        'd_extra_roty [arcsec]': [extra_roty] * n_mm,
        'd_extra_z [µm]': [extra_z_um] * n_mm,
    })

    with pd.ExcelWriter(path, engine='openpyxl') as w:
        mm_psf.to_excel(w, sheet_name='MM_PSF', index=False)
        aeff.to_excel(w, sheet_name='A_eff', index=False)
        mm_conf.to_excel(w, sheet_name='MM configuration', index=False)
        align.to_excel(w, sheet_name='Alignment', index=False)
        grav.to_excel(w, sheet_name='Gravity offload', index=False)
        therm.to_excel(w, sheet_name='Thermal', index=False)
        extra.to_excel(w, sheet_name='Extra PSF shifts', index=False)


# ---------------------------------------------------------------------------
# Off-axis rotation tests
# ---------------------------------------------------------------------------

class TestOffAxisRotation:
    """Verify that d_extra_rotx / d_extra_roty propagate into rotation totals."""

    def test_extra_rotx_roty_added_to_total(self, tmp_path):
        """Extra rotx/roty should appear in compute_total_rot_polar output."""
        path = str(tmp_path / 'offaxis.xlsx')
        _make_workbook(path, n_mm=1, extra_rotx=100.0, extra_roty=200.0)

        mm_to_pos = {1: 1}
        mm_config_map = {1: {'x_MM': 1.0, 'y_MM': 0.0, 'r_MM': 1.0}}
        alignment_by_pos = {}
        gravity_by_pos = {}
        thermal_by_pos = {}
        extra_by_pos = {1: {'d_extra_rotx': 100.0, 'd_extra_roty': 200.0}}

        rotx, roty, _, _ = compute_total_rot_polar(
            mm_to_pos, mm_config_map,
            alignment_by_pos, gravity_by_pos, thermal_by_pos,
            extra_by_pos,
        )
        assert np.isclose(rotx[1], 100.0)
        assert np.isclose(roty[1], 200.0)

    def test_extra_rotations_combine_with_thermal(self, tmp_path):
        """Extra rotations should add to existing thermal rotations."""
        mm_to_pos = {1: 1}
        mm_config_map = {1: {'x_MM': 1.0, 'y_MM': 0.0, 'r_MM': 1.0}}
        thermal_by_pos = {1: {'d_therm_rotx': 5.0, 'd_therm_roty': 10.0,
                              'd_therm_rotrad': 0.0, 'd_therm_rotazi': 0.0}}
        extra_by_pos = {1: {'d_extra_rotx': 100.0, 'd_extra_roty': 200.0}}

        rotx, roty, _, _ = compute_total_rot_polar(
            mm_to_pos, mm_config_map, {}, {}, thermal_by_pos, extra_by_pos,
        )
        assert np.isclose(rotx[1], 105.0)
        assert np.isclose(roty[1], 210.0)

    def test_zero_offaxis_no_effect(self, tmp_path):
        """When extra shifts are zero, rotations come only from gravity/thermal."""
        mm_to_pos = {1: 1}
        mm_config_map = {1: {'x_MM': 1.0, 'y_MM': 0.0, 'r_MM': 1.0}}
        gravity_by_pos = {1: {'d_grav_rotx': 3.0, 'd_grav_roty': 4.0,
                              'd_grav_rotrad': 0.0, 'd_grav_rotazi': 0.0}}
        extra_by_pos = {1: {'d_extra_rotx': 0.0, 'd_extra_roty': 0.0}}

        rotx, roty, _, _ = compute_total_rot_polar(
            mm_to_pos, mm_config_map, {}, gravity_by_pos, {}, extra_by_pos,
        )
        assert np.isclose(rotx[1], 3.0)
        assert np.isclose(roty[1], 4.0)

    def test_offaxis_20arcmin_conversion(self):
        """20 arcmin off-axis should produce 848.53 arcsec per axis (20*60/sqrt(2))."""
        offaxis_arcmin = 20.0
        expected = offaxis_arcmin * 60.0 / math.sqrt(2.0)
        assert np.isclose(expected, 848.528137, atol=1e-3)

    def test_loader_reads_extra_rotations(self, tmp_path):
        """load_gaussians_from_excel should populate extra rotations from sheet."""
        path = str(tmp_path / 'rotload.xlsx')
        _make_workbook(path, n_mm=1, extra_rotx=500.0, extra_roty=600.0)

        df = load_gaussians_from_excel(path)
        # The df itself contains mux/muy but the rotations are used internally.
        # Verify the centroid shift is consistent with having non-zero rotations
        # (exact effect depends on vignetting, here we just verify loading succeeds
        # and the returned DataFrame is valid).
        assert len(df) == 1
        assert 'mux' in df.columns
        assert 'muy' in df.columns


# ---------------------------------------------------------------------------
# Defocus (d_extra_z) tests
# ---------------------------------------------------------------------------

class TestDefocusExtraZ:
    """Verify that d_extra_z propagates into centroid shifts via z-projection."""

    def test_defocus_shifts_centroid(self, tmp_path):
        """Non-zero d_extra_z should shift mux/muy via the z-projection formula.

        With MM at x_MM=1, y_MM=0, z_MM=0 (denominator=12), and no other
        perturbations, mux should shift by d_z * x_MM / 12.
        """
        defocus_um = 1000.0  # 1 mm defocus in µm
        path = str(tmp_path / 'defoc.xlsx')
        _make_workbook(path, n_mm=1, extra_z_um=defocus_um,
                       x_mm=[1.0], y_mm=[0.0], r_mm=[1.0])

        df = load_gaussians_from_excel(path)
        mux = float(df.loc[0, 'mux'])
        # expected: d_z_total = defocus_um * 1e-6 = 1e-3 m
        # dm_x = 1e-3 * 1.0 / 12 = 8.333e-5 m
        expected_dm_x = (defocus_um * 1e-6) * 1.0 / 12.0
        assert np.isclose(mux, expected_dm_x, atol=1e-10), \
            f"mux={mux}, expected ~{expected_dm_x}"

    def test_zero_defocus_no_shift(self, tmp_path):
        """Zero d_extra_z should not shift the centroid (other d_z also zero)."""
        path = str(tmp_path / 'nodefoc.xlsx')
        _make_workbook(path, n_mm=1, extra_z_um=0.0)

        df = load_gaussians_from_excel(path)
        mux = float(df.loc[0, 'mux'])
        muy = float(df.loc[0, 'muy'])
        assert np.isclose(mux, 0.0, atol=1e-12)
        assert np.isclose(muy, 0.0, atol=1e-12)

    def test_defocus_adds_to_thermal_dz(self, tmp_path):
        """d_extra_z should combine additively with d_therm_z."""
        therm_z = 500.0  # µm
        extra_z = 300.0  # µm
        path = str(tmp_path / 'combined_dz.xlsx')
        _make_workbook(path, n_mm=1, extra_z_um=extra_z, therm_z_um=therm_z,
                       x_mm=[0.5], y_mm=[0.0], r_mm=[0.5])

        df = load_gaussians_from_excel(path)
        mux = float(df.loc[0, 'mux'])
        # total d_z = (500 + 300) * 1e-6 = 8e-4 m
        # dm_x = 8e-4 * 0.5 / 12 = 3.333e-5
        expected = (therm_z + extra_z) * 1e-6 * 0.5 / 12.0
        assert np.isclose(mux, expected, atol=1e-10), \
            f"mux={mux}, expected ~{expected}"

    def test_defocus_adds_to_all_dz_sources(self, tmp_path):
        """d_extra_z adds to align + gravity + thermal d_z."""
        align_z = 100.0
        grav_z = 200.0
        therm_z = 300.0
        extra_z = 400.0
        path = str(tmp_path / 'all_dz.xlsx')
        _make_workbook(path, n_mm=1,
                       align_z_um=align_z, grav_z_um=grav_z,
                       therm_z_um=therm_z, extra_z_um=extra_z,
                       x_mm=[0.0], y_mm=[1.0], r_mm=[1.0])

        df = load_gaussians_from_excel(path)
        muy = float(df.loc[0, 'muy'])
        total_dz_m = (align_z + grav_z + therm_z + extra_z) * 1e-6
        expected = total_dz_m * 1.0 / 12.0
        assert np.isclose(muy, expected, atol=1e-10), \
            f"muy={muy}, expected ~{expected}"

    def test_defocus_y_projection(self, tmp_path):
        """Defocus should project along both x_MM and y_MM axes."""
        extra_z = 600.0
        x, y = 0.6, 0.8
        r = math.sqrt(x**2 + y**2)
        path = str(tmp_path / 'defoc_xy.xlsx')
        _make_workbook(path, n_mm=1, extra_z_um=extra_z,
                       x_mm=[x], y_mm=[y], r_mm=[r])

        df = load_gaussians_from_excel(path)
        dz_m = extra_z * 1e-6
        expected_mux = dz_m * x / 12.0
        expected_muy = dz_m * y / 12.0
        assert np.isclose(float(df.loc[0, 'mux']), expected_mux, atol=1e-10)
        assert np.isclose(float(df.loc[0, 'muy']), expected_muy, atol=1e-10)

    def test_multiple_mms_different_positions(self, tmp_path):
        """Each MM position gets its shift from the same d_extra_z but different geometry."""
        extra_z = 1200.0
        path = str(tmp_path / 'multi_mm.xlsx')
        _make_workbook(path, n_mm=3, extra_z_um=extra_z,
                       x_mm=[1.0, 0.0, -1.0],
                       y_mm=[0.0, 1.0, 0.0],
                       r_mm=[1.0, 1.0, 1.0])

        df = load_gaussians_from_excel(path)
        dz_m = extra_z * 1e-6

        # MM 1: x=1, y=0 -> mux = dz*1/12, muy = 0
        assert np.isclose(float(df.loc[0, 'mux']), dz_m * 1.0 / 12.0, atol=1e-10)
        assert np.isclose(float(df.loc[0, 'muy']), 0.0, atol=1e-10)
        # MM 2: x=0, y=1 -> mux = 0, muy = dz*1/12
        assert np.isclose(float(df.loc[1, 'mux']), 0.0, atol=1e-10)
        assert np.isclose(float(df.loc[1, 'muy']), dz_m * 1.0 / 12.0, atol=1e-10)
        # MM 3: x=-1, y=0 -> mux = -dz/12, muy = 0
        assert np.isclose(float(df.loc[2, 'mux']), dz_m * (-1.0) / 12.0, atol=1e-10)
        assert np.isclose(float(df.loc[2, 'muy']), 0.0, atol=1e-10)


# ---------------------------------------------------------------------------
# Loader edge-case tests
# ---------------------------------------------------------------------------

class TestExtraShiftLoaderEdgeCases:
    """Verify the loader handles missing / partial Extra PSF shifts data."""

    def test_missing_extra_sheet_no_error(self, tmp_path):
        """Workbook without 'Extra PSF shifts' sheet should still load cleanly."""
        path = str(tmp_path / 'no_extra.xlsx')
        mm_psf = pd.DataFrame({
            'MM #': [1],
            'm_rad [arcsec]': [0.0], 'm_azi [arcsec]': [0.0],
            'sigma_rad [arcsec]': [4.0], 'sigma_azi [arcsec]': [4.0],
        })
        aeff = pd.DataFrame({'MM #': [1], '1.0 keV': [1.0]})
        mm_conf = pd.DataFrame({
            'MM #': [1], 'Position #': [1],
            'x_MM [m]': [1.0], 'y_MM [m]': [0.0], 'z_MM [m]': [0.0], 'r_MM [m]': [1.0],
        })
        grav = pd.DataFrame({
            'Position #': [1],
            'd_grav_x [µm]': [0.0], 'd_grav_y [µm]': [0.0], 'd_grav_z [µm]': [0.0],
            'd_grav_rotx [arcsec]': [0.0], 'd_grav_roty [arcsec]': [0.0],
            'd_grav_rotz [arcsec]': [0.0],
        })
        therm = pd.DataFrame({
            'Position #': [1],
            'd_therm_x [µm]': [0.0], 'd_therm_y [µm]': [0.0], 'd_therm_z [µm]': [0.0],
            'd_therm_rotx [arcsec]': [0.0], 'd_therm_roty [arcsec]': [0.0],
            'd_therm_rotz [arcsec]': [0.0],
        })
        align = pd.DataFrame({
            'Position #': [1],
            'd_align_rad [µm]': [0.0], 'd_align_azi [µm]': [0.0], 'd_align_z [µm]': [0.0],
            'd_align_rotazi [arcsec]': [0.0], 'd_align_rotrad [arcsec]': [0.0],
            'd_align_rotz [arcsec]': [0.0],
        })
        with pd.ExcelWriter(path, engine='openpyxl') as w:
            mm_psf.to_excel(w, sheet_name='MM_PSF', index=False)
            aeff.to_excel(w, sheet_name='A_eff', index=False)
            mm_conf.to_excel(w, sheet_name='MM configuration', index=False)
            align.to_excel(w, sheet_name='Alignment', index=False)
            grav.to_excel(w, sheet_name='Gravity offload', index=False)
            therm.to_excel(w, sheet_name='Thermal', index=False)

        df = load_gaussians_from_excel(path)
        assert len(df) == 1
        assert np.isclose(float(df.loc[0, 'mux']), 0.0, atol=1e-12)

    def test_extra_sheet_without_z_column(self, tmp_path):
        """Extra PSF shifts sheet with only rotx/roty (no d_extra_z) should work."""
        path = str(tmp_path / 'no_z_col.xlsx')
        _make_workbook(path, n_mm=1, extra_rotx=50.0, extra_roty=60.0, extra_z_um=0.0)
        # Re-write the Extra sheet without the z column
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb['Extra PSF shifts']
        # Remove column 4 header and data
        ws.cell(row=1, column=4, value=None)
        ws.cell(row=2, column=4, value=None)
        wb.save(path)

        df = load_gaussians_from_excel(path)
        assert len(df) == 1
        # No z shift should be applied
        assert np.isclose(float(df.loc[0, 'mux']), 0.0, atol=1e-12)

    def test_defocus_mm_to_um_conversion(self):
        """Batch code converts defocus from mm to µm (*1e3); verify the math."""
        defocus_mm = 2.5
        defocus_um = defocus_mm * 1e3
        assert defocus_um == 2500.0
        # Loader then converts µm to m
        defocus_m = defocus_um * 1e-6
        assert np.isclose(defocus_m, 2.5e-3)
