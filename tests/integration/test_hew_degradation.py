"""Tests for HEW degradation sheets: interpolation, column B write-back, and sigma broadening."""

import math
import tempfile
import os

import numpy as np
import pandas as pd
import openpyxl
import pytest

from main import load_gaussians_from_excel


# ---------------------------------------------------------------------------
# Helper: build a minimal workbook with HEW degradation sheets
# ---------------------------------------------------------------------------

_FWHM_TO_SIGMA = 2.0 * math.sqrt(2.0 * math.log(2.0))  # ~2.3548
_ARCSEC_TO_M = 12.0 * math.pi / 180.0 / 3600.0


def _make_workbook(
    path,
    *,
    n_mm=2,
    sigma_rad=4.0,
    sigma_azi=4.0,
    extra_rotx=0.0,
    extra_roty=0.0,
    hew_azi_table=None,
    hew_rad_table=None,
    sel_energy=1.0,
    row_nums=None,
    energy_scaling_table=None,
):
    """Write a test workbook with controllable HEW degradation sheets.

    Parameters
    ----------
    hew_azi_table / hew_rad_table : list[tuple] | None
        Each entry is (cfg_row, angle_arcmin, energy_keV, hew_arcsec).
        If None the corresponding sheet is omitted.
    row_nums : list[int] | None
        Row # per position in MM configuration (defaults to 1 for all).
    energy_scaling_table : list[tuple] | None
        Each entry is (energy_keV, scaling_factor) for the
        "MM HEW degradation energy" sheet.  If None the sheet is omitted.
    """
    mm_nums = list(range(1, n_mm + 1))
    if row_nums is None:
        row_nums = [1] * n_mm

    mm_psf = pd.DataFrame({
        'MM #': mm_nums,
        'm_rad [arcsec]': [0.0] * n_mm,
        'm_azi [arcsec]': [0.0] * n_mm,
        'sigma_rad [arcsec]': [sigma_rad] * n_mm,
        'sigma_azi [arcsec]': [sigma_azi] * n_mm,
    })
    aeff = pd.DataFrame({'MM #': mm_nums, '1.0 keV': [1.0] * n_mm})
    mm_conf = pd.DataFrame({
        'MM #': mm_nums,
        'Position #': mm_nums,
        'Row #': row_nums,
        'x_MM [m]': [1.0] * n_mm,
        'y_MM [m]': [0.0] * n_mm,
        'z_MM [m]': [0.0] * n_mm,
        'r_MM [m]': [1.0] * n_mm,
    })
    grav = pd.DataFrame({
        'Position #': mm_nums,
        'd_grav_x [µm]': [0.0] * n_mm,
        'd_grav_y [µm]': [0.0] * n_mm,
        'd_grav_z [µm]': [0.0] * n_mm,
        'd_grav_rotx [arcsec]': [0.0] * n_mm,
        'd_grav_roty [arcsec]': [0.0] * n_mm,
        'd_grav_rotz [arcsec]': [0.0] * n_mm,
    })
    therm = pd.DataFrame({
        'Position #': mm_nums,
        'd_therm_x [µm]': [0.0] * n_mm,
        'd_therm_y [µm]': [0.0] * n_mm,
        'd_therm_z [µm]': [0.0] * n_mm,
        'd_therm_rotx [arcsec]': [extra_rotx] * n_mm,
        'd_therm_roty [arcsec]': [extra_roty] * n_mm,
        'd_therm_rotz [arcsec]': [0.0] * n_mm,
    })
    align = pd.DataFrame({
        'Position #': mm_nums,
        'd_align_rad [µm]': [0.0] * n_mm,
        'd_align_azi [µm]': [0.0] * n_mm,
        'd_align_z [µm]': [0.0] * n_mm,
        'd_align_rotazi [arcsec]': [0.0] * n_mm,
        'd_align_rotrad [arcsec]': [0.0] * n_mm,
        'd_align_rotz [arcsec]': [0.0] * n_mm,
    })
    extra = pd.DataFrame({
        'Position #': mm_nums,
        'd_extra_rotx [arcsec]': [0.0] * n_mm,
        'd_extra_roty [arcsec]': [0.0] * n_mm,
        'd_extra_z [µm]': [0.0] * n_mm,
    })

    with pd.ExcelWriter(path, engine='openpyxl') as w:
        mm_psf.to_excel(w, sheet_name='MM_PSF', index=False)
        aeff.to_excel(w, sheet_name='A_eff', index=False)
        mm_conf.to_excel(w, sheet_name='MM configuration', index=False)
        align.to_excel(w, sheet_name='Alignment', index=False)
        grav.to_excel(w, sheet_name='Gravity offload', index=False)
        therm.to_excel(w, sheet_name='Thermal', index=False)
        extra.to_excel(w, sheet_name='Extra PSF shifts', index=False)

    # Add HEW degradation sheets with openpyxl (needs specific column layout)
    wb = openpyxl.load_workbook(path)

    # Write sel_energy to A_eff D2 so main.py detects it as authoritative source
    ws_aeff = wb['A_eff']
    ws_aeff.cell(row=1, column=4, value='Selected energy')
    ws_aeff.cell(row=2, column=4, value=f'{sel_energy} keV')

    # Minimal vignetting sheets (required for sel_energy detection path in main.py)
    for vig_name in ('MM vignetting rotazi', 'MM vignetting rotrad'):
        ws_vig = wb.create_sheet(vig_name)
        ws_vig.cell(row=1, column=1, value='rotazi [arcmin]' if 'rotazi' in vig_name else 'rotrad [arcmin]')
        ws_vig.cell(row=1, column=2, value='factor')
        ws_vig.cell(row=2, column=1, value=0.0)
        ws_vig.cell(row=2, column=2, value=1.0)
        ws_vig.cell(row=3, column=1, value=10.0)
        ws_vig.cell(row=3, column=2, value=1.0)

    # "MM HEW degradation energy" sheet (energy vs sigma scaling factor)
    if energy_scaling_table is not None:
        ws_esc = wb.create_sheet('MM HEW degradation energy')
        ws_esc.cell(row=1, column=1, value='Energy (Kev)')
        ws_esc.cell(row=1, column=2, value="σ's scaling factor (wrt 1keV)")
        for i, (e_kev, s_factor) in enumerate(energy_scaling_table):
            ws_esc.cell(row=i + 2, column=1, value=e_kev)
            ws_esc.cell(row=i + 2, column=2, value=s_factor)

    for sname, table in (
        ('MM HEW degradation rotazi', hew_azi_table),
        ('MM HEW degradation rotrad', hew_rad_table),
    ):
        if table is None:
            continue
        ws = wb.create_sheet(sname)
        # Headers
        ws.cell(row=1, column=1, value='Position #')
        ws.cell(row=1, column=2, value='HEW degradation (arcsec)')
        ws.cell(row=1, column=3, value='Selected energy [keV]')
        ws.cell(row=1, column=8, value='Row')
        angle_label = 'rotazi [arcmin] ' if 'rotazi' in sname else 'rotrad [arcmin] '
        ws.cell(row=1, column=9, value=angle_label)
        ws.cell(row=1, column=10, value='energy [keV]')
        ws.cell(row=1, column=11, value='HEW degradation [arcsec]')
        # Column A: positions, column C row2: energy
        for i, pos in enumerate(mm_nums):
            ws.cell(row=i + 2, column=1, value=pos)
        ws.cell(row=2, column=3, value=sel_energy)
        # Lookup table in H-K
        for i, (cfg_row, angle_am, energy, hew_val) in enumerate(table):
            ws.cell(row=i + 2, column=8, value=cfg_row)
            ws.cell(row=i + 2, column=9, value=angle_am)
            ws.cell(row=i + 2, column=10, value=energy)
            ws.cell(row=i + 2, column=11, value=hew_val)
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Tests: Column B write-back (interpolation)
# ---------------------------------------------------------------------------

class TestHEWDegradationWriteBack:
    """Verify interpolated HEW degradation values are written to column B."""

    def test_writes_column_b_for_both_sheets(self, tmp_path):
        """Both rotazi and rotrad sheets should get column B populated."""
        table = [
            # cfg_row, angle_arcmin, energy_keV, hew_arcsec
            (1, 0, 1.0, 0.0),
            (1, 1, 1.0, 0.5),
            (1, 5, 1.0, 2.0),
        ]
        path = str(tmp_path / 'hew_wb.xlsx')
        _make_workbook(path, n_mm=2, hew_azi_table=table, hew_rad_table=table)
        load_gaussians_from_excel(path)

        wb = openpyxl.load_workbook(path, data_only=True)
        for sname in ('MM HEW degradation rotazi', 'MM HEW degradation rotrad'):
            ws = wb[sname]
            vals = [ws.cell(row=r, column=2).value for r in range(2, 4)]
            assert all(v is not None for v in vals), f"{sname} col B has None values: {vals}"
        wb.close()

    def test_interpolation_at_zero_angle(self, tmp_path):
        """When rotation angle is 0 the HEW degradation from the table at 0 should be used."""
        table = [
            (1, 0, 1.0, 0.0),
            (1, 1, 1.0, 1.0),
        ]
        path = str(tmp_path / 'hew_zero.xlsx')
        # No thermal rotx/roty => rotation angles are ~0
        _make_workbook(path, n_mm=1, hew_azi_table=table, hew_rad_table=table,
                       extra_rotx=0.0, extra_roty=0.0)
        load_gaussians_from_excel(path)

        wb = openpyxl.load_workbook(path, data_only=True)
        for sname in ('MM HEW degradation rotazi', 'MM HEW degradation rotrad'):
            ws = wb[sname]
            val = ws.cell(row=2, column=2).value
            assert val is not None
            assert abs(val - 0.0) < 0.05, f"{sname}: expected ~0 at zero angle, got {val}"
        wb.close()

    def test_interpolation_midpoint(self, tmp_path):
        """A known angle should produce correct linear interpolation."""
        # Table: 0 arcmin -> 0, 1 arcmin (60 arcsec) -> 1.2 arcsec
        table = [
            (1, 0, 1.0, 0.0),
            (1, 1, 1.0, 1.2),
        ]
        path = str(tmp_path / 'hew_mid.xlsx')
        # rotx=30 arcsec with x_MM=1,y_MM=0,r_MM=1 gives rot_rad ~ 30 arcsec
        _make_workbook(path, n_mm=1, hew_rad_table=table, extra_rotx=30.0)
        load_gaussians_from_excel(path)

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb['MM HEW degradation rotrad']
        val = ws.cell(row=2, column=2).value
        assert val is not None
        # At 30 arcsec, linear interp between (0,0) and (60,1.2) -> 0.6
        assert abs(val - 0.6) < 0.15, f"Expected ~0.6, got {val}"
        wb.close()

    def test_no_hew_sheet_no_crash(self, tmp_path):
        """Workbook without HEW degradation sheets should load without error."""
        path = str(tmp_path / 'no_hew.xlsx')
        _make_workbook(path, n_mm=2)
        df = load_gaussians_from_excel(path)
        assert len(df) == 2

    def test_energy_selection_from_c2(self, tmp_path):
        """The selected energy in C2 should pick the matching series."""
        table = [
            (1, 0, 1.0, 0.0),
            (1, 1, 1.0, 1.0),
            (1, 0, 7.0, 0.0),
            (1, 1, 7.0, 5.0),  # much larger at 7 keV
        ]
        path = str(tmp_path / 'hew_energy.xlsx')
        _make_workbook(path, n_mm=1, hew_azi_table=table, sel_energy=7.0,
                       extra_roty=30.0)
        load_gaussians_from_excel(path)

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb['MM HEW degradation rotazi']
        val = ws.cell(row=2, column=2).value
        assert val is not None
        # At 7 keV, 30 arcsec ~ halfway between 0 and 60 arcsec -> ~2.5
        assert val > 1.0, f"Expected 7 keV series (large values), got {val}"
        wb.close()

    def test_multiple_cfg_rows(self, tmp_path):
        """Positions with different Row # should use different lookup series."""
        table = [
            (1, 0, 1.0, 0.0),
            (1, 1, 1.0, 1.0),
            (2, 0, 1.0, 0.0),
            (2, 1, 1.0, 3.0),  # Row 2 has steeper slope
        ]
        path = str(tmp_path / 'hew_rows.xlsx')
        _make_workbook(path, n_mm=2, hew_rad_table=table,
                       row_nums=[1, 2], extra_rotx=30.0)
        load_gaussians_from_excel(path)

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb['MM HEW degradation rotrad']
        v1 = ws.cell(row=2, column=2).value  # pos 1, Row #1
        v2 = ws.cell(row=3, column=2).value  # pos 2, Row #2
        assert v1 is not None and v2 is not None
        assert v2 > v1, f"Row 2 slope is steeper, so pos 2 ({v2}) should exceed pos 1 ({v1})"
        wb.close()


# ---------------------------------------------------------------------------
# Tests: Sigma broadening
# ---------------------------------------------------------------------------

class TestHEWSigmaBroadening:
    """Verify that positive HEW degradation broadens sigma_rad / sigma_azi."""

    def _expected_sigma(self, sigma_arcsec, hew_arcsec):
        """Compute the expected broadened sigma in meters."""
        sigma_m = sigma_arcsec * _ARCSEC_TO_M
        sigma_extra_m = (hew_arcsec / _FWHM_TO_SIGMA) * _ARCSEC_TO_M
        return math.sqrt(sigma_m**2 + sigma_extra_m**2)

    def test_positive_hew_broadens_sigma_rad(self, tmp_path):
        """Positive rotrad HEW degradation should increase sigma_rad."""
        sigma_0 = 4.0  # arcsec
        table = [
            (1, 0, 1.0, 2.0),  # constant 2 arcsec at all angles
            (1, 5, 1.0, 2.0),
        ]
        path = str(tmp_path / 'broad_rad.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0, hew_rad_table=table)
        df = load_gaussians_from_excel(path)
        expected = self._expected_sigma(sigma_0, 2.0)
        actual = float(df['sigma_rad'].iloc[0])
        assert abs(actual - expected) / expected < 1e-4, \
            f"sigma_rad: expected {expected:.6e}, got {actual:.6e}"

    def test_positive_hew_broadens_sigma_azi(self, tmp_path):
        """Positive rotazi HEW degradation should increase sigma_azi."""
        sigma_0 = 4.0
        table = [
            (1, 0, 1.0, 1.5),
            (1, 5, 1.0, 1.5),
        ]
        path = str(tmp_path / 'broad_azi.xlsx')
        _make_workbook(path, n_mm=1, sigma_azi=sigma_0, hew_azi_table=table)
        df = load_gaussians_from_excel(path)
        expected = self._expected_sigma(sigma_0, 1.5)
        actual = float(df['sigma_azi'].iloc[0])
        assert abs(actual - expected) / expected < 1e-4, \
            f"sigma_azi: expected {expected:.6e}, got {actual:.6e}"

    def test_negative_hew_no_broadening(self, tmp_path):
        """Negative HEW degradation values should NOT modify sigma."""
        sigma_0 = 4.0
        table = [
            (1, 0, 1.0, -0.5),
            (1, 5, 1.0, -0.5),
        ]
        path = str(tmp_path / 'neg_hew.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       hew_rad_table=table, hew_azi_table=table)
        df = load_gaussians_from_excel(path)
        expected = sigma_0 * _ARCSEC_TO_M
        assert abs(float(df['sigma_rad'].iloc[0]) - expected) < 1e-10
        assert abs(float(df['sigma_azi'].iloc[0]) - expected) < 1e-10

    def test_zero_hew_no_broadening(self, tmp_path):
        """Zero HEW degradation should leave sigma unchanged."""
        sigma_0 = 4.0
        table = [
            (1, 0, 1.0, 0.0),
            (1, 5, 1.0, 0.0),
        ]
        path = str(tmp_path / 'zero_hew.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       hew_rad_table=table, hew_azi_table=table)
        df = load_gaussians_from_excel(path)
        expected = sigma_0 * _ARCSEC_TO_M
        assert abs(float(df['sigma_rad'].iloc[0]) - expected) < 1e-10
        assert abs(float(df['sigma_azi'].iloc[0]) - expected) < 1e-10

    def test_sigmax_sigmay_match_broadened(self, tmp_path):
        """sigmax/sigmay must reflect the broadened sigma_rad/sigma_azi."""
        table = [
            (1, 0, 1.0, 2.0),
            (1, 5, 1.0, 2.0),
        ]
        path = str(tmp_path / 'sigxy.xlsx')
        _make_workbook(path, n_mm=1, hew_rad_table=table, hew_azi_table=table)
        df = load_gaussians_from_excel(path)
        assert float(df['sigmax'].iloc[0]) == float(df['sigma_rad'].iloc[0])
        assert float(df['sigmay'].iloc[0]) == float(df['sigma_azi'].iloc[0])

    def test_per_position_mapping(self, tmp_path):
        """Different positions (different Row #) get different broadening."""
        table = [
            (1, 0, 1.0, 0.0),
            (1, 5, 1.0, 0.0),  # Row 1 => 0 HEW degradation everywhere
            (2, 0, 1.0, 3.0),
            (2, 5, 1.0, 3.0),  # Row 2 => 3 arcsec
        ]
        sigma_0 = 4.0
        path = str(tmp_path / 'per_pos.xlsx')
        _make_workbook(path, n_mm=2, sigma_rad=sigma_0,
                       hew_rad_table=table, row_nums=[1, 2])
        df = load_gaussians_from_excel(path)
        # MM 1 -> pos 1 -> row 1 -> 0 HEW -> sigma unchanged
        expected_unchanged = sigma_0 * _ARCSEC_TO_M
        actual_1 = float(df.loc[df['MM #'] == 1, 'sigma_rad'].iloc[0])
        assert abs(actual_1 - expected_unchanged) < 1e-10, \
            f"Pos 1 should be unbroadened, got {actual_1}"
        # MM 2 -> pos 2 -> row 2 -> 3 arcsec HEW -> sigma broadened
        expected_broad = self._expected_sigma(sigma_0, 3.0)
        actual_2 = float(df.loc[df['MM #'] == 2, 'sigma_rad'].iloc[0])
        assert abs(actual_2 - expected_broad) / expected_broad < 1e-4, \
            f"Pos 2 should be broadened to {expected_broad:.6e}, got {actual_2:.6e}"

    def test_only_azi_sheet_broadens_azi_only(self, tmp_path):
        """When only the rotazi HEW sheet exists, only sigma_azi is broadened."""
        sigma_0 = 4.0
        table = [
            (1, 0, 1.0, 2.0),
            (1, 5, 1.0, 2.0),
        ]
        path = str(tmp_path / 'azi_only.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       hew_azi_table=table, hew_rad_table=None)
        df = load_gaussians_from_excel(path)
        expected_unchanged = sigma_0 * _ARCSEC_TO_M
        assert abs(float(df['sigma_rad'].iloc[0]) - expected_unchanged) < 1e-10
        expected_broad = self._expected_sigma(sigma_0, 2.0)
        assert abs(float(df['sigma_azi'].iloc[0]) - expected_broad) / expected_broad < 1e-4

    def test_broadening_formula_exact(self, tmp_path):
        """Verify the exact formula: sqrt(sigma^2 + (hew/(2*sqrt(2*ln2)))^2)."""
        sigma_0 = 5.0  # arcsec
        hew_deg = 3.5   # arcsec
        table = [
            (1, 0, 1.0, hew_deg),
            (1, 10, 1.0, hew_deg),
        ]
        path = str(tmp_path / 'exact.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0, hew_rad_table=table)
        df = load_gaussians_from_excel(path)
        sigma_m = sigma_0 * _ARCSEC_TO_M
        sigma_extra_m = (hew_deg / _FWHM_TO_SIGMA) * _ARCSEC_TO_M
        expected = math.sqrt(sigma_m**2 + sigma_extra_m**2)
        actual = float(df['sigma_rad'].iloc[0])
        assert abs(actual - expected) / expected < 1e-4, \
            f"Expected {expected:.10e}, got {actual:.10e}"


# ---------------------------------------------------------------------------
# Tests: Energy-dependent sigma scaling ("MM HEW degradation energy" sheet)
# ---------------------------------------------------------------------------

# Reference scaling table (subset of real data)
_ENERGY_SCALING_TABLE = [
    (0.2,  0.98),
    (0.35, 0.98),
    (1.0,  1.00),
    (2.0,  1.03),
    (4.0,  1.09),
    (7.0,  1.18),
    (10.0, 1.23),
    (12.0, 1.28),
]


class TestHEWEnergyScaling:
    """Verify energy-dependent sigma scaling from 'MM HEW degradation energy' sheet."""

    def test_factor_applied_at_7kev(self, tmp_path):
        """At 7 keV the factor is 1.18 – sigma should be scaled accordingly."""
        sigma_0 = 4.0  # arcsec
        path = str(tmp_path / 'energy_7.xlsx')
        _make_workbook(path, n_mm=2, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       sel_energy=7.0, energy_scaling_table=_ENERGY_SCALING_TABLE)
        df = load_gaussians_from_excel(path)
        expected = sigma_0 * _ARCSEC_TO_M * 1.18
        for col in ('sigma_rad', 'sigma_azi'):
            actual = float(df[col].iloc[0])
            assert abs(actual - expected) / expected < 1e-4, \
                f"{col}: expected {expected:.6e}, got {actual:.6e}"

    def test_factor_unity_at_1kev(self, tmp_path):
        """At 1 keV the factor is 1.0 – sigma should be unchanged."""
        sigma_0 = 4.0
        path = str(tmp_path / 'energy_1.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       sel_energy=1.0, energy_scaling_table=_ENERGY_SCALING_TABLE)
        df = load_gaussians_from_excel(path)
        expected = sigma_0 * _ARCSEC_TO_M
        for col in ('sigma_rad', 'sigma_azi'):
            actual = float(df[col].iloc[0])
            assert abs(actual - expected) < 1e-10, \
                f"{col}: expected {expected:.6e} (no scaling), got {actual:.6e}"

    def test_interpolation_at_intermediate_energy(self, tmp_path):
        """At 5.5 keV the factor should be linearly interpolated between 4 and 7 keV."""
        sigma_0 = 4.0
        path = str(tmp_path / 'energy_5_5.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0,
                       sel_energy=5.5, energy_scaling_table=_ENERGY_SCALING_TABLE)
        df = load_gaussians_from_excel(path)
        # np.interp between (4, 1.09) and (7, 1.18): factor = 1.09 + (5.5-4)/(7-4) * (1.18-1.09)
        expected_factor = 1.09 + (5.5 - 4.0) / (7.0 - 4.0) * (1.18 - 1.09)
        expected = sigma_0 * _ARCSEC_TO_M * expected_factor
        actual = float(df['sigma_rad'].iloc[0])
        assert abs(actual - expected) / expected < 1e-4, \
            f"Expected factor={expected_factor:.4f}, sigma={expected:.6e}, got {actual:.6e}"

    def test_no_energy_sheet_no_scaling(self, tmp_path):
        """Without 'MM HEW degradation energy' sheet, sigma should be unscaled."""
        sigma_0 = 4.0
        path = str(tmp_path / 'no_energy_sheet.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       sel_energy=7.0, energy_scaling_table=None)
        df = load_gaussians_from_excel(path)
        expected = sigma_0 * _ARCSEC_TO_M
        for col in ('sigma_rad', 'sigma_azi'):
            actual = float(df[col].iloc[0])
            assert abs(actual - expected) < 1e-10, \
                f"{col}: without energy sheet sigma should be unchanged, got {actual:.6e}"

    def test_energy_scaling_applies_without_hew_broadening(self, tmp_path):
        """Energy scaling must apply even when no rotazi/rotrad HEW sheets exist."""
        sigma_0 = 5.0
        path = str(tmp_path / 'energy_only.xlsx')
        _make_workbook(path, n_mm=2, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       sel_energy=7.0, energy_scaling_table=_ENERGY_SCALING_TABLE,
                       hew_azi_table=None, hew_rad_table=None)
        df = load_gaussians_from_excel(path)
        expected = sigma_0 * _ARCSEC_TO_M * 1.18
        for col in ('sigma_rad', 'sigma_azi'):
            actual = float(df[col].iloc[0])
            assert abs(actual - expected) / expected < 1e-4, \
                f"{col}: energy scaling should apply without HEW sheets, got {actual:.6e}"

    def test_energy_scaling_combined_with_hew_broadening(self, tmp_path):
        """Energy scaling should multiply on top of HEW broadening."""
        sigma_0 = 4.0
        hew_deg = 2.0  # constant HEW degradation
        hew_table = [
            (1, 0, 7.0, hew_deg),
            (1, 5, 7.0, hew_deg),
        ]
        path = str(tmp_path / 'combined.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0,
                       sel_energy=7.0,
                       hew_rad_table=hew_table,
                       energy_scaling_table=_ENERGY_SCALING_TABLE)
        df = load_gaussians_from_excel(path)
        # First HEW broadening: sqrt(sigma^2 + (hew/FWHM_to_sigma)^2)
        sigma_m = sigma_0 * _ARCSEC_TO_M
        sigma_extra_m = (hew_deg / _FWHM_TO_SIGMA) * _ARCSEC_TO_M
        broadened = math.sqrt(sigma_m**2 + sigma_extra_m**2)
        # Then energy scaling
        expected = broadened * 1.18
        actual = float(df['sigma_rad'].iloc[0])
        assert abs(actual - expected) / expected < 1e-4, \
            f"Expected broadened*1.18={expected:.6e}, got {actual:.6e}"

    def test_sigmax_sigmay_include_energy_scaling(self, tmp_path):
        """sigmax/sigmay (used by aggregation) must include energy scaling."""
        sigma_0 = 4.0
        path = str(tmp_path / 'sigxy_energy.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       sel_energy=7.0, energy_scaling_table=_ENERGY_SCALING_TABLE)
        df = load_gaussians_from_excel(path)
        assert float(df['sigmax'].iloc[0]) == float(df['sigma_rad'].iloc[0])
        assert float(df['sigmay'].iloc[0]) == float(df['sigma_azi'].iloc[0])
        # And they should be scaled
        expected = sigma_0 * _ARCSEC_TO_M * 1.18
        assert abs(float(df['sigmax'].iloc[0]) - expected) / expected < 1e-4

    def test_ij_columns_match_dataframe(self, tmp_path):
        """Workbook I/J columns must contain the same energy-scaled sigmas."""
        sigma_0 = 4.0
        path = str(tmp_path / 'ij_match.xlsx')
        _make_workbook(path, n_mm=2, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       sel_energy=7.0, energy_scaling_table=_ENERGY_SCALING_TABLE)
        df = load_gaussians_from_excel(path)

        m_to_arcsec = 1.0 / _ARCSEC_TO_M
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb['MM_PSF']
        for idx, row in df.iterrows():
            mm = int(row['MM #'])
            # Find the workbook row for this MM
            for r in range(2, (ws.max_row or 0) + 1):
                if ws.cell(r, 1).value is not None and int(float(ws.cell(r, 1).value)) == mm:
                    i_val = float(ws.cell(r, 9).value)
                    j_val = float(ws.cell(r, 10).value)
                    expected_i = float(row['sigma_rad']) * m_to_arcsec
                    expected_j = float(row['sigma_azi']) * m_to_arcsec
                    assert abs(i_val - expected_i) / expected_i < 1e-6, \
                        f"MM {mm}: I={i_val}, expected {expected_i}"
                    assert abs(j_val - expected_j) / expected_j < 1e-6, \
                        f"MM {mm}: J={j_val}, expected {expected_j}"
                    break
        wb.close()

    def test_extrapolation_below_range(self, tmp_path):
        """Energy below table range uses the lowest factor (flat extrapolation)."""
        sigma_0 = 4.0
        path = str(tmp_path / 'energy_low.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0,
                       sel_energy=0.1, energy_scaling_table=_ENERGY_SCALING_TABLE)
        df = load_gaussians_from_excel(path)
        # np.interp clamps below range to first value: 0.98
        expected = sigma_0 * _ARCSEC_TO_M * 0.98
        actual = float(df['sigma_rad'].iloc[0])
        assert abs(actual - expected) / expected < 1e-4, \
            f"Below range: expected factor=0.98, got sigma={actual:.6e}"

    def test_extrapolation_above_range(self, tmp_path):
        """Energy above table range uses the highest factor (flat extrapolation)."""
        sigma_0 = 4.0
        path = str(tmp_path / 'energy_high.xlsx')
        _make_workbook(path, n_mm=1, sigma_rad=sigma_0,
                       sel_energy=15.0, energy_scaling_table=_ENERGY_SCALING_TABLE)
        df = load_gaussians_from_excel(path)
        # np.interp clamps above range to last value: 1.28
        expected = sigma_0 * _ARCSEC_TO_M * 1.28
        actual = float(df['sigma_rad'].iloc[0])
        assert abs(actual - expected) / expected < 1e-4, \
            f"Above range: expected factor=1.28, got sigma={actual:.6e}"

    def test_scaling_all_mms(self, tmp_path):
        """Energy scaling should apply uniformly to all MMs."""
        sigma_0 = 4.0
        n = 5
        path = str(tmp_path / 'all_mms.xlsx')
        _make_workbook(path, n_mm=n, sigma_rad=sigma_0, sigma_azi=sigma_0,
                       sel_energy=7.0, energy_scaling_table=_ENERGY_SCALING_TABLE)
        df = load_gaussians_from_excel(path)
        expected = sigma_0 * _ARCSEC_TO_M * 1.18
        for i in range(n):
            for col in ('sigma_rad', 'sigma_azi'):
                actual = float(df[col].iloc[i])
                assert abs(actual - expected) / expected < 1e-4, \
                    f"MM {i+1} {col}: expected {expected:.6e}, got {actual:.6e}"
