import subprocess
import sys
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
import shutil


def test_batch_combinations_creates_zips(tmp_path):
    """Create sample input and combinations Excel, run batch, and verify zips."""
    repo_root = Path(__file__).resolve().parents[1]
    main_py = repo_root / 'main.py'

    # Workspace for test: tmp_path
    workdir = tmp_path
    ddir = workdir / 'Distributions'
    ddir.mkdir()

    # Create a minimal sample_input.xlsx with Thermal and Vignetting sheets
    sample_input = ddir / 'sample_input.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Thermal'
    ws.append(['Position #', 'd_therm_rotx', 'd_therm_roty', 'd_therm_z'])
    ws.append([1, 0.0, 0.0, 0.0])

    ws2 = wb.create_sheet('Vignetting rotrad')
    ws2.append(['col1', 'col2', 'col3'])
    ws2.append([0, 0, 0])

    ws3 = wb.create_sheet('Vignetting rotazi')
    ws3.append(['col1', 'col2', 'col3'])
    ws3.append([0, 0, 0])

    wb.save(sample_input)

    # Add a minimal MM_PSF sheet and MM configuration so main.py can load the workbook
    ws4 = wb.create_sheet('MM_PSF')
    ws4.append(['MM #', 'm_rad [arcsec]', 'm_azi [arcsec]', 'sigma_rad [arcsec]', 'sigma_azi [arcsec]'])
    ws4.append([1, 0.0, 0.0, 1.0, 1.0])

    ws_cfg = wb.create_sheet('MM configuration')
    ws_cfg.append(['MM #', 'x_MM [m]', 'r_MM [m]'])
    ws_cfg.append([1, 0.0, 0.0])
    # Add A_eff sheet with base A_eff for MM #1
    ws_a = wb.create_sheet('A_eff')
    ws_a.append([1, 1.0])

    # Create combinations Excel (columns: A,B,C,D,E -> we use B-E)
    combos = pd.DataFrame(
        [
            [1, 'cfgA', 1.0, 5.0, 0.1],
            [2, 'cfgB', 2.5, 6.0, 0.2],
            [3, 'cfgC', 0.5, 7.0, 0.0],
        ],
        columns=['A', 'B', 'C', 'D', 'E'],
    )
    combos_path = ddir / 'combinations.xlsx'
    combos.to_excel(combos_path, index=False)

    # Run the batch CLI pointing at test files; run from tmp workdir so Exports is created there
    cmd = [sys.executable, str(main_py), '--batch-combinations', str(combos_path), '--file', str(sample_input)]
    proc = subprocess.run(cmd, cwd=str(workdir), stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)

    assert proc.returncode == 0, f"Batch run failed: stdout:\n{proc.stdout}\nstderr:\n{proc.stderr}"

    exports = workdir / 'Exports'
    assert exports.exists() and exports.is_dir(), f"Exports directory not created: {exports}"

    # Check that a zip was created for each config with expected basename
    for prefix in ['cfgA', 'cfgB', 'cfgC']:
        # New naming: <prefix>_<input_stem>_<YYYYMMDD_HHMMSS>.zip
        matches = list(exports.glob(f"{prefix}_{sample_input.stem}_*.zip"))
        assert matches, f"Expected package not found for prefix {prefix} (pattern {prefix}_{sample_input.stem}_*.zip)"

    # Cleanup: remove Exports to avoid polluting workspace (tmp_path is ephemeral anyway)
    try:
        shutil.rmtree(exports)
    except Exception:
        pass
