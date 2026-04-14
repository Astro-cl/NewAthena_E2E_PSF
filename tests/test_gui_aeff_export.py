import tempfile
from openpyxl import Workbook, load_workbook
import pandas as pd

from gui_distributions import sync_aeff_column_b_in_workbook


def make_wb(path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'A_eff'
    ws.cell(row=1, column=1, value='MM #')
    ws.cell(row=1, column=2, value='A_eff')
    # two MM rows
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value=0.1)
    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=2, value=0.2)
    wb.save(path)


def test_sync_aeff_column_b_prefers_mapping_then_raw(tmp_path):
    p = tmp_path / 'test_aeff.xlsx'
    make_wb(str(p))
    wb = load_workbook(str(p))

    # Build mapping that overrides MM 1 only; MM2 should fall back to raw
    mapping = {1: 9.9}

    # Build raw df that has base values in second column
    raw = pd.DataFrame([[1, 0.11], [2, 0.22]])

    written = sync_aeff_column_b_in_workbook(wb, aeff_weights=mapping, aeff_raw_df=raw)
    # should have written for both rows (one via mapping, one via raw)
    assert written >= 2

    # Save and re-open to assert cell values
    wb.save(str(p))
    wb2 = load_workbook(str(p))
    ws2 = wb2['A_eff']
    assert ws2.cell(row=2, column=2).value == 9.9
    assert abs(float(ws2.cell(row=3, column=2).value) - 0.22) < 1e-9
