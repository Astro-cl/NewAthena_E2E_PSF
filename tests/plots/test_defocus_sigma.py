"""Unit tests for defocusing (dz) PSF sigma shape adjustment.

When a mirror module has a non-zero net axial displacement *dz*
(d_align_z + d_grav_z + d_therm_z + d_extra_z, all in metres) its PSF widths are
adjusted according to::

    sigma_rad_adjusted = sigma_rad_initial
                          + (MM_height - sigma_rad_initial) / 12 * abs(dz)
    sigma_azi_adjusted = sigma_azi_initial
                          + (MM_width  - sigma_azi_initial) / 12 * abs(dz)

``MM_height`` is taken from column I (0-based index 8) and ``MM_width``
from column J (0-based index 9) of the ``MM configuration`` sheet.
All quantities are in metres.

The adjusted sigma values are:
* Stored in ``df['sigma_rad']`` / ``df['sigma_azi']`` (metres) returned by
  ``load_gaussians_from_excel``.
* Written back to columns I / J of the ``MM_PSF`` sheet as arcsec.
"""

import math
import os
import tempfile

import numpy as np
import pandas as pd
import openpyxl
import pytest

from main import load_gaussians_from_excel

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_ARCSEC_TO_M = 12.0 * math.pi / 180.0 / 3600.0
_M_TO_ARCSEC = 1.0 / _ARCSEC_TO_M


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def _make_workbook(
    path: str,
    *,
    n_mm: int = 2,
    sigma_rad_arcsec: float = 4.0,
    sigma_azi_arcsec: float = 3.0,
    mm_height_m: float = 0.200,
    mm_width_m: float = 0.100,
    d_align_z_um: float = 0.0,
    include_height_width: bool = True,
) -> str:
    """Write a minimal workbook for defocusing sigma tests.

    Parameters
    ----------
    include_height_width:
        When ``False``, columns I / J of ``MM configuration`` are omitted so
        the graceful-fallback path (no sigma change) can be exercised.
    """
    mm_nums = list(range(1, n_mm + 1))

    # MM_PSF sheet (cols A-E)
    mm_psf = pd.DataFrame({
        'MM #': mm_nums,
        'm_rad [arcsec]': [0.0] * n_mm,
        'm_azi [arcsec]': [0.0] * n_mm,
        'sigma_rad [arcsec]': [sigma_rad_arcsec] * n_mm,
        'sigma_azi [arcsec]': [sigma_azi_arcsec] * n_mm,
    })

    # MM configuration sheet
    # Columns A-H are named; cols I / J carry MM_height / MM_width (metres).
    cfg_data = {
        'MM #':        mm_nums,           # col A (index 0)
        'Position #':  mm_nums,           # col B (index 1)
        'Row #':       [1] * n_mm,        # col C (index 2)
        'x_MM [m]':    [1.0] * n_mm,      # col D (index 3)
        'y_MM [m]':    [0.0] * n_mm,      # col E (index 4)
        'z_MM [m]':    [0.0] * n_mm,      # col F (index 5)
        'r_MM [m]':    [1.0] * n_mm,      # col G (index 6)
        'filler_H':    [0.0] * n_mm,      # col H (index 7) – placeholder
    }
    if include_height_width:
        cfg_data['MM_height [m]'] = [mm_height_m] * n_mm  # col I (index 8)
        cfg_data['MM_width [m]']  = [mm_width_m]  * n_mm  # col J (index 9)
    mm_conf = pd.DataFrame(cfg_data)

    # Alignment sheet – only d_align_z is non-zero in most tests
    align = pd.DataFrame({
        'Position #':          mm_nums,
        'd_align_rad [µm]':    [0.0] * n_mm,
        'd_align_azi [µm]':    [0.0] * n_mm,
        'd_align_x [µm]':      [0.0] * n_mm,
        'd_align_y [µm]':      [0.0] * n_mm,
        'd_align_z [µm]':      [d_align_z_um] * n_mm,
        'd_align_rotz [arcsec]': [0.0] * n_mm,
    })

    # A_eff sheet (required for weight loading)
    aeff = pd.DataFrame({'MM #': mm_nums, '1.0 keV': [1.0] * n_mm})

    # Gravity / Thermal sheets (all-zero)
    grav = pd.DataFrame({
        'Position #':          mm_nums,
        'd_grav_x [µm]':       [0.0] * n_mm,
        'd_grav_y [µm]':       [0.0] * n_mm,
        'd_grav_z [µm]':       [0.0] * n_mm,
        'd_grav_rotx [arcsec]': [0.0] * n_mm,
        'd_grav_roty [arcsec]': [0.0] * n_mm,
        'd_grav_rotz [arcsec]': [0.0] * n_mm,
    })
    therm = pd.DataFrame({
        'Position #':           mm_nums,
        'd_therm_x [µm]':       [0.0] * n_mm,
        'd_therm_y [µm]':       [0.0] * n_mm,
        'd_therm_z [µm]':       [0.0] * n_mm,
        'd_therm_rotx [arcsec]': [0.0] * n_mm,
        'd_therm_roty [arcsec]': [0.0] * n_mm,
        'd_therm_rotz [arcsec]': [0.0] * n_mm,
    })

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        mm_psf.to_excel(writer, sheet_name='MM_PSF',          index=False)
        mm_conf.to_excel(writer, sheet_name='MM configuration', index=False)
        align.to_excel(writer, sheet_name='Alignment',         index=False)
        aeff.to_excel(writer, sheet_name='A_eff',              index=False)
        grav.to_excel(writer, sheet_name='Gravity',            index=False)
        therm.to_excel(writer, sheet_name='Thermal',           index=False)

    return path


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _expected_sigma(sigma_initial_arcsec: float, mm_dim_m: float, dz_m: float) -> float:
    """Return the expected adjusted sigma in metres.

    Formula: sigma_adjusted = sigma_initial + (MM_dim - 6*sigma_initial) / 12 * dz / 6
    """
    s0 = sigma_initial_arcsec * _ARCSEC_TO_M
    return s0 + (mm_dim_m - 6.0 * s0) / 12.0 * dz_m / 6.0


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestDefocusSigmaFormula:
    """Verify the defocusing sigma formula applied inside load_gaussians_from_excel."""

    def test_sigma_unchanged_when_dz_zero(self, tmp_path):
        """With dz = 0 the sigma values must not change."""
        wb = str(tmp_path / 'test_dz0.xlsx')
        _make_workbook(wb, sigma_rad_arcsec=4.0, sigma_azi_arcsec=3.0,
                       mm_height_m=0.2, mm_width_m=0.1, d_align_z_um=0.0)
        df = load_gaussians_from_excel(wb)
        expected_rad = 4.0 * _ARCSEC_TO_M
        expected_azi = 3.0 * _ARCSEC_TO_M
        for _, row in df.iterrows():
            assert float(row['sigma_rad']) == pytest.approx(expected_rad, rel=1e-6)
            assert float(row['sigma_azi']) == pytest.approx(expected_azi, rel=1e-6)

    def test_sigma_adjusted_positive_dz(self, tmp_path):
        """Positive dz broadens sigma toward the MM physical dimension."""
        sigma_rad_arc = 4.0
        sigma_azi_arc = 3.0
        mm_h = 0.200   # metres
        mm_w = 0.100   # metres
        dz_um = 1_000_000.0   # 1 µm * 1e6 = 1 m axial displacement
        dz_m = dz_um * 1e-6   # → 1.0 m

        wb = str(tmp_path / 'test_dz_pos.xlsx')
        _make_workbook(wb, sigma_rad_arcsec=sigma_rad_arc,
                       sigma_azi_arcsec=sigma_azi_arc,
                       mm_height_m=mm_h, mm_width_m=mm_w,
                       d_align_z_um=dz_um)
        df = load_gaussians_from_excel(wb)

        exp_rad = _expected_sigma(sigma_rad_arc, mm_h, dz_m)
        exp_azi = _expected_sigma(sigma_azi_arc, mm_w, dz_m)
        for _, row in df.iterrows():
            assert float(row['sigma_rad']) == pytest.approx(exp_rad, rel=1e-6)
            assert float(row['sigma_azi']) == pytest.approx(exp_azi, rel=1e-6)

    def test_sigma_adjusted_negative_dz(self, tmp_path):
        """Negative dz reduces sigma (signed formula: sigma may decrease)."""
        sigma_rad_arc = 4.0
        sigma_azi_arc = 3.0
        mm_h = 0.200
        mm_w = 0.100
        dz_um = -500_000.0   # −0.5 m axial displacement
        dz_m = dz_um * 1e-6

        wb = str(tmp_path / 'test_dz_neg.xlsx')
        _make_workbook(wb, sigma_rad_arcsec=sigma_rad_arc,
                       sigma_azi_arcsec=sigma_azi_arc,
                       mm_height_m=mm_h, mm_width_m=mm_w,
                       d_align_z_um=dz_um)
        df = load_gaussians_from_excel(wb)

        exp_rad = _expected_sigma(sigma_rad_arc, mm_h, dz_m)
        exp_azi = _expected_sigma(sigma_azi_arc, mm_w, dz_m)
        for _, row in df.iterrows():
            assert float(row['sigma_rad']) == pytest.approx(exp_rad, rel=1e-6)
            assert float(row['sigma_azi']) == pytest.approx(exp_azi, rel=1e-6)

    def test_sigma_unchanged_when_height_width_absent(self, tmp_path):
        """When cols I / J are absent the adjustment is skipped gracefully."""
        sigma_rad_arc = 4.0
        sigma_azi_arc = 3.0
        dz_um = 500_000.0   # 0.5 m – non-zero dz, but no height/width data

        wb = str(tmp_path / 'test_no_hw.xlsx')
        _make_workbook(wb, sigma_rad_arcsec=sigma_rad_arc,
                       sigma_azi_arcsec=sigma_azi_arc,
                       d_align_z_um=dz_um,
                       include_height_width=False)
        df = load_gaussians_from_excel(wb)

        expected_rad = sigma_rad_arc * _ARCSEC_TO_M
        expected_azi = sigma_azi_arc * _ARCSEC_TO_M
        for _, row in df.iterrows():
            assert float(row['sigma_rad']) == pytest.approx(expected_rad, rel=1e-6)
            assert float(row['sigma_azi']) == pytest.approx(expected_azi, rel=1e-6)

    def test_each_mm_uses_its_own_dz(self, tmp_path):
        """Different d_align_z per position produce independently adjusted sigmas."""
        n_mm = 3
        sigma_rad_arc = 4.0
        sigma_azi_arc = 3.0
        mm_h = 0.200
        mm_w = 0.100
        dz_values_um = [0.0, 500_000.0, -300_000.0]  # 0, +0.5 m, −0.3 m

        mm_nums = list(range(1, n_mm + 1))
        mm_psf = pd.DataFrame({
            'MM #': mm_nums,
            'm_rad [arcsec]': [0.0] * n_mm,
            'm_azi [arcsec]': [0.0] * n_mm,
            'sigma_rad [arcsec]': [sigma_rad_arc] * n_mm,
            'sigma_azi [arcsec]': [sigma_azi_arc] * n_mm,
        })
        cfg = pd.DataFrame({
            'MM #':        mm_nums,
            'Position #':  mm_nums,
            'Row #':       [1] * n_mm,
            'x_MM [m]':    [1.0] * n_mm,
            'y_MM [m]':    [0.0] * n_mm,
            'z_MM [m]':    [0.0] * n_mm,
            'r_MM [m]':    [1.0] * n_mm,
            'filler_H':    [0.0] * n_mm,
            'MM_height [m]': [mm_h] * n_mm,
            'MM_width [m]':  [mm_w] * n_mm,
        })
        align = pd.DataFrame({
            'Position #':          mm_nums,
            'd_align_rad [µm]':    [0.0] * n_mm,
            'd_align_azi [µm]':    [0.0] * n_mm,
            'd_align_x [µm]':      [0.0] * n_mm,
            'd_align_y [µm]':      [0.0] * n_mm,
            'd_align_z [µm]':      dz_values_um,
            'd_align_rotz [arcsec]': [0.0] * n_mm,
        })
        aeff  = pd.DataFrame({'MM #': mm_nums, '1.0 keV': [1.0] * n_mm})
        grav  = pd.DataFrame({'Position #': mm_nums,
                               'd_grav_x [µm]': [0.0]*n_mm, 'd_grav_y [µm]': [0.0]*n_mm,
                               'd_grav_z [µm]': [0.0]*n_mm, 'd_grav_rotx [arcsec]': [0.0]*n_mm,
                               'd_grav_roty [arcsec]': [0.0]*n_mm, 'd_grav_rotz [arcsec]': [0.0]*n_mm})
        therm = pd.DataFrame({'Position #': mm_nums,
                               'd_therm_x [µm]': [0.0]*n_mm, 'd_therm_y [µm]': [0.0]*n_mm,
                               'd_therm_z [µm]': [0.0]*n_mm, 'd_therm_rotx [arcsec]': [0.0]*n_mm,
                               'd_therm_roty [arcsec]': [0.0]*n_mm, 'd_therm_rotz [arcsec]': [0.0]*n_mm})

        wb = str(tmp_path / 'test_per_mm.xlsx')
        with pd.ExcelWriter(wb, engine='openpyxl') as writer:
            mm_psf.to_excel(writer, sheet_name='MM_PSF',          index=False)
            cfg.to_excel(writer,   sheet_name='MM configuration', index=False)
            align.to_excel(writer, sheet_name='Alignment',         index=False)
            aeff.to_excel(writer,  sheet_name='A_eff',             index=False)
            grav.to_excel(writer,  sheet_name='Gravity',           index=False)
            therm.to_excel(writer, sheet_name='Thermal',           index=False)

        df = load_gaussians_from_excel(wb)
        df = df.sort_values('MM #').reset_index(drop=True)

        for i, dz_um in enumerate(dz_values_um):
            dz_m = dz_um * 1e-6
            exp_rad = _expected_sigma(sigma_rad_arc, mm_h, dz_m)
            exp_azi = _expected_sigma(sigma_azi_arc, mm_w, dz_m)
            got_rad = float(df.at[i, 'sigma_rad'])
            got_azi = float(df.at[i, 'sigma_azi'])
            assert got_rad == pytest.approx(exp_rad, rel=1e-6), (
                f"MM #{i+1}: sigma_rad={got_rad:.6e} expected {exp_rad:.6e} (dz={dz_m} m)"
            )
            assert got_azi == pytest.approx(exp_azi, rel=1e-6), (
                f"MM #{i+1}: sigma_azi={got_azi:.6e} expected {exp_azi:.6e} (dz={dz_m} m)"
            )


class TestDefocusSigmaWriteback:
    """Verify that the adjusted sigma is written to columns I / J of MM_PSF."""

    def test_workbook_ij_columns_written_after_dz(self, tmp_path):
        """Columns I / J of MM_PSF in the saved workbook must hold the adjusted sigma."""
        sigma_rad_arc = 4.0
        sigma_azi_arc = 3.0
        mm_h = 0.200
        mm_w = 0.100
        dz_um = 1_000_000.0   # 1 m
        dz_m = dz_um * 1e-6

        wb_path = str(tmp_path / 'test_writeback.xlsx')
        _make_workbook(wb_path, sigma_rad_arcsec=sigma_rad_arc,
                       sigma_azi_arcsec=sigma_azi_arc,
                       mm_height_m=mm_h, mm_width_m=mm_w,
                       d_align_z_um=dz_um)
        load_gaussians_from_excel(wb_path)   # triggers write-back

        exp_rad_arcsec = _expected_sigma(sigma_rad_arc, mm_h, dz_m) * _M_TO_ARCSEC
        exp_azi_arcsec = _expected_sigma(sigma_azi_arc, mm_w, dz_m) * _M_TO_ARCSEC

        wb = openpyxl.load_workbook(wb_path, data_only=True)
        ws = wb['MM_PSF']
        # Row 1 is the header; data starts at row 2.
        for r in range(2, ws.max_row + 1):
            mm_cell = ws.cell(row=r, column=1).value
            if mm_cell is None:
                continue
            col_i = ws.cell(row=r, column=9).value   # sigma_rad_deg [arcsec]
            col_j = ws.cell(row=r, column=10).value  # sigma_azi_deg [arcsec]
            assert col_i is not None, f"Row {r}: column I is empty"
            assert col_j is not None, f"Row {r}: column J is empty"
            assert float(col_i) == pytest.approx(exp_rad_arcsec, rel=1e-5), (
                f"Row {r} col I: got {col_i:.6f} expected {exp_rad_arcsec:.6f} arcsec"
            )
            assert float(col_j) == pytest.approx(exp_azi_arcsec, rel=1e-5), (
                f"Row {r} col J: got {col_j:.6f} expected {exp_azi_arcsec:.6f} arcsec"
            )
        wb.close()

    def test_workbook_ij_unchanged_when_dz_zero(self, tmp_path):
        """When dz = 0, column I / J reflect the original (unadjusted) sigma."""
        sigma_rad_arc = 4.0
        sigma_azi_arc = 3.0

        wb_path = str(tmp_path / 'test_wb_nodz.xlsx')
        _make_workbook(wb_path, sigma_rad_arcsec=sigma_rad_arc,
                       sigma_azi_arcsec=sigma_azi_arc,
                       mm_height_m=0.2, mm_width_m=0.1,
                       d_align_z_um=0.0)
        load_gaussians_from_excel(wb_path)

        wb = openpyxl.load_workbook(wb_path, data_only=True)
        ws = wb['MM_PSF']
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value is None:
                continue
            col_i = ws.cell(row=r, column=9).value
            col_j = ws.cell(row=r, column=10).value
            assert float(col_i) == pytest.approx(sigma_rad_arc, rel=1e-5)
            assert float(col_j) == pytest.approx(sigma_azi_arc, rel=1e-5)
        wb.close()


class TestDefocusSigmaFormulaMath:
    """Pure-math checks of the expected formula (no file I/O)."""

    @pytest.mark.parametrize("sigma0,dim,dz,expected", [
        # dz = 0 → no change
        (1e-4, 0.200,  0.0,  1e-4),
        # sigma0 == dim → no change regardless of dz
        (0.200, 0.200, 1.0,  0.200),
        # dz = 12 → sigma0 + (dim - sigma0) = dim
        (1e-4, 0.200,  12.0, 0.200),
        # dz = −12 → same result as +12 (abs)
        (1e-4, 0.200, -12.0, 0.200),
        # dz = 6 → midpoint
        (0.0,  0.100,  6.0,  0.100 / 2.0),
        # dz = −6 → same midpoint (abs)
        (0.0,  0.100, -6.0,  0.100 / 2.0),
    ])
    def test_formula_edge_cases(self, sigma0, dim, dz, expected):
        result = sigma0 + (dim - sigma0) / 12.0 * abs(dz)
        assert result == pytest.approx(expected, rel=1e-10, abs=1e-15)
