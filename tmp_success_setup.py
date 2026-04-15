from openpyxl import Workbook
import pandas as pd
from pathlib import Path
p=Path('tmp_success/Distributions/sample_input.xlsx')
p.parent.mkdir(parents=True, exist_ok=True)
wb=Workbook()
# Thermal
ws=wb.active; ws.title='Thermal'
ws.append(['Position #','d_therm_rotx','d_therm_roty','d_therm_z'])
ws.append([1,0.0,0.0,0.0])
# Vignetting sheets (simple)
ws2=wb.create_sheet('Vignetting rotrad'); ws2.append(['col1','col2','col3']); ws2.append([0,0,0])
ws3=wb.create_sheet('Vignetting rotazi'); ws3.append(['col1','col2','col3']); ws3.append([0,0,0])
# MM_PSF with non-zero sigmas
ws4=wb.create_sheet('MM_PSF')
ws4.append(['MM #','m_rad [arcsec]','m_azi [arcsec]','sigma_rad [arcsec]','sigma_azi [arcsec]','distribution'])
ws4.append([1,0.0,0.0,0.6,0.6,'gaussian'])
# MM configuration with non-zero r_MM
ws_cfg=wb.create_sheet('MM configuration')
ws_cfg.append(['MM #','x_MM [m]','y_MM [m]','r_MM [m]'])
ws_cfg.append([1,0.05,0.05,0.07])
# A_eff
ws_a=wb.create_sheet('A_eff'); ws_a.append([1,1.0])
wb.save(p)
# combinations.xlsx
combos=pd.DataFrame([[1,'cfgSuccess',1.0,5.0,0.1]], columns=['A','B','C','D','E'])
combos.to_excel('tmp_success/Distributions/combinations.xlsx', index=False)
print('created tmp_success inputs')
