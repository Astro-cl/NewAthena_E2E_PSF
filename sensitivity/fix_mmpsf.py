from pathlib import Path
import pandas as pd
import sys
sys.path.insert(0, '.')
from sensitivity.sensitivity_run import load_standard_mm_psf_defs
from gui_distributions import generate_data_from_distributions, DATA_TYPES
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

base = Path('Distributions/TestDistribution.xlsx')
std = load_standard_mm_psf_defs(base)
key = '10% Variable Sym Gaussian 4.3"'
if key not in std:
    print('Preset not found; available keys:', list(std.keys())[:20])
    raise SystemExit(1)

sd = std[key]
print('Parsed preset for', key, sd)

# build params
params = {'m_rad [arcsec]': ('fixed', 0.0, 0.0), 'm_azi [arcsec]': ('fixed', 0.0, 0.0)}
if sd.get('sigma_rad'):
    s = sd['sigma_rad']
    if s.get('dist') == 'fixed':
        params['sigma_rad [arcsec]'] = ('fixed', s.get('value', 0.0), 0.0)
    elif s.get('dist') == 'gaussian':
        params['sigma_rad [arcsec]'] = ('gaussian', s.get('mean', 0.0), s.get('sigma', 0.0))
if sd.get('sigma_azi'):
    s = sd['sigma_azi']
    if s.get('dist') == 'fixed':
        params['sigma_azi [arcsec]'] = ('fixed', s.get('value', 0.0), 0.0)
    elif s.get('dist') == 'gaussian':
        params['sigma_azi [arcsec]'] = ('gaussian', s.get('mean', 0.0), s.get('sigma', 0.0))

print('Generation params:', params)

mm_cfg = pd.read_excel(base, sheet_name='MM configuration', engine='openpyxl')
num_mm = mm_cfg.shape[0]
mm_list = mm_cfg['MM #'].dropna().astype(int).tolist() if 'MM #' in mm_cfg.columns else list(range(1, num_mm+1))

df_gen = generate_data_from_distributions(params, num_mm, DATA_TYPES['MM_PSF'])
df_gen.insert(0, 'MM #', mm_list[:len(df_gen)])

pbase = Path('sensitivity/input')
name = '20260123T221943Z_3_MM_PSF10_Variable_Sym_Gaussian_4.3_Alignment0.0_Gravity_offload0.0_Thermal0.0'
file1 = pbase / f'{name}.xlsx'
file2 = pbase / f'{name}_placed.xlsx'

print('Will write to:', file1.exists(), file1)
if file2.exists():
    print('and placed exists:', file2)


def write_mmpsf(fn, df):
    wb = load_workbook(fn)
    if 'MM_PSF' in wb.sheetnames:
        ws = wb['MM_PSF']
        wb.remove(ws)
    ws = wb.create_sheet('MM_PSF')
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(fn)
    print('Wrote MM_PSF to', fn)

if file1.exists():
    write_mmpsf(file1, df_gen)
else:
    print('Input file missing:', file1)

if file2.exists():
    write_mmpsf(file2, df_gen)
else:
    print('Placed file missing (not overwritten):', file2)

# inspect placed
if file2.exists():
    dfp = pd.read_excel(file2, sheet_name='MM_PSF', engine='openpyxl')
    print('\nPlaced file sigma_rad [arcsec] unique:', pd.to_numeric(dfp['sigma_rad [arcsec]'], errors='coerce').unique()[:10])
    print(dfp[['sigma_rad [arcsec]','sigma_azi [arcsec]']].head().to_string(index=False))
else:
    print('No placed file to inspect')
