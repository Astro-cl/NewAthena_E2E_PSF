"""
Minimal sensitivity run template.

Usage examples:
    python3 sensitivity/sensitivity_run.py --workers 64 --persist --baseline Distributions/TestDistribution.xlsx

This template does the following:
- Ensures the trial folders exist (`input`, `workbooks`, `results`, `figures`)
- Optionally copies a baseline workbook into `input`
- Invokes the existing driver `tools/run_sensitivity.py` with configurable env vars

Extend this script to prepare combos, manage placed-workbook caching, and post-process results.
"""

import argparse
import os
import shutil
import tempfile
import subprocess
import json
from pathlib import Path
from datetime import datetime
from itertools import product
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import re
import logging

ROOT = Path(__file__).resolve().parents[1]
SENS_DIR = Path(__file__).resolve().parent

try:
    from gui_distributions import DATA_TYPES, generate_data_from_distributions
except Exception:
    DATA_TYPES = None
    generate_data_from_distributions = None

def ensure_dirs():
    for d in (SENS_DIR / "input", SENS_DIR / "workbooks", SENS_DIR / "results", SENS_DIR / "figures"):
        d.mkdir(parents=True, exist_ok=True)


def _cleanup_input_dir(input_dir: Path, max_files: int = 100):
    """Keep only the newest `max_files` files in `input_dir` (by mtime).

    This is called after creating input workbooks so the folder does not
    grow unbounded. It's a best-effort cleanup: failures are ignored.
    """
    try:
        if not input_dir.exists() or not input_dir.is_dir():
            return
        files = [p for p in input_dir.iterdir() if p.is_file()]
        if len(files) <= int(max_files):
            return
        # sort by modification time, oldest first
        files_sorted = sorted(files, key=lambda p: p.stat().st_mtime)
        # remove oldest leaving the newest `max_files`
        to_remove = files_sorted[: max(0, len(files_sorted) - int(max_files))]
        for p in to_remove:
            try:
                p.unlink()
            except Exception:
                continue
    except Exception:
        return


def write_multisheet_csv(sheets: dict, out_path: Path):
    """Write a single CSV file containing multiple named sheets.

    Each sheet is written preceded by a marker line:
      # sheet: Sheet Name

    This format is readable by `parse_multisheet_csv` in `main.py`.
    """
    with open(out_path, 'w', encoding='utf-8', newline='') as fh:
        first = True
        for name, df in sheets.items():
            if not first:
                fh.write('\n')
            first = False
            fh.write(f"# sheet: {name}\n")
            # Use pandas to_csv into a buffer then write
            try:
                csv_text = df.to_csv(index=False)
            except Exception:
                csv_text = df.to_csv(index=False, header=True)
            fh.write(csv_text)


def copy_baseline(baseline_path: Path, dest_dir: Path | None = None) -> Path:
    baseline = Path(baseline_path)
    if not baseline.exists():
        raise FileNotFoundError(f"Baseline not found: {baseline}")
    if dest_dir is None:
        dest = SENS_DIR / "input" / baseline.name
    else:
        dest = Path(dest_dir) / baseline.name
    shutil.copy2(baseline, dest)
    return dest


def run_driver(workers: int, persist: bool, extra_env: dict):
    env = os.environ.copy()
    env.update(extra_env)
    env["SENS_MAX_WORKERS"] = str(workers)
    env["SENS_PERSIST_TMP"] = "1" if persist else "0"

    cmd = ["python3", str(ROOT / "tools" / "run_sensitivity.py")]

    print(f"Running: {' '.join(cmd)} (workers={workers} persist={persist})")
    completed = subprocess.run(cmd, cwd=ROOT, env=env)
    if completed.returncode != 0:
        raise RuntimeError(f"Driver exited with code {completed.returncode}")


def _safe_eval_numeric_expr(expr: str) -> float:
    import ast
    # sanitize expression from Excel cells: strip whitespace and remove newlines/carriage returns
    if expr is None:
        raise ValueError('Empty expression')
    expr = str(expr).strip().replace('\r', '').replace('\n', ' ')
    expr = expr.strip()

    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Num):
            return float(node.n)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            v = _eval(node.operand)
            return +v if isinstance(node.op, ast.UAdd) else -v
        if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
            a = _eval(node.left)
            b = _eval(node.right)
            if isinstance(node.op, ast.Add):
                return a + b
            if isinstance(node.op, ast.Sub):
                return a - b
            if isinstance(node.op, ast.Mult):
                return a * b
            if isinstance(node.op, ast.Div):
                return a / b
        raise ValueError(f'Unsupported expression: {expr!r}')

    node = ast.parse(expr, mode='eval')
    return float(_eval(node.body))


def _parse_standard_dist_spec(spec: str):
    """Parse specs like 'gaussian(0,12/3)' or 'uniform(1.5,9)' into (dist,a,b)."""
    import re
    if spec is None or (isinstance(spec, float) and pd.isna(spec)):
        raise ValueError('Empty spec')
    s = str(spec).strip()
    if not s:
        raise ValueError('Empty spec')

    compact = s.replace(' ', '')
    if re.fullmatch(r'[-+]?\d+(?:\.\d+)?', compact):
        return ('fixed', float(compact), 0.0)

    # match gaussian|normal|uniform
    m = re.match(r'^\s*(gaussian|normal|uniform)\s*\(\s*(.+)\s*\)\s*$', s, re.IGNORECASE)
    if not m:
        # try to match pseudo-voigt / voigt style specs like 'pseudo-voigt(1.2,0.3,0.5)'
        m2 = re.match(r'^\s*(?:pseudo[- ]?voigt|voigt|pv|pvoigt)\s*\(\s*(.+)\s*\)\s*$', s, re.IGNORECASE)
        if not m2:
            raise ValueError(f'Unsupported distribution spec: {spec!r}')
        kind = 'voigt'
        inner = m2.group(1)
    else:
        kind = m.group(1).lower()
        inner = m.group(2)
    parts = [p.strip() for p in inner.split(',') if p.strip()]
    if len(parts) < 1:
        raise ValueError(f'Expected parameters in spec: {spec!r}')
    # For gaussian/uniform expect two params; for voigt allow 2 or 3
    if kind in ('gaussian', 'normal', 'uniform') and len(parts) < 2:
        raise ValueError(f'Expected two parameters in spec: {spec!r}')

    left = parts[0]
    right = parts[1] if len(parts) > 1 else ''
    third = parts[2] if len(parts) > 2 else None

    # Normalize percent tokens in the left and right expressions so forms like
    # '-9*150%' and '12/3*150%' evaluate correctly.
    try:
        left_norm = re.sub(r'([+-]?\d+(?:\.\d+)?)\s*%', r'(\1/100)', left)
    except Exception:
        left_norm = left
    try:
        a = _safe_eval_numeric_expr(left_norm)
    except Exception:
        a = _safe_eval_numeric_expr(left)

    rstr = right.strip()
    # Normalize percent tokens everywhere: convert '150%' -> '(150/100)'.
    # This handles forms like '12/3*150%', '150%* -9' or '-9*150%'.
    try:
        rstr = re.sub(r'([+-]?\d+(?:\.\d+)?)\s*%', r'(\1/100)', rstr)
    except Exception:
        # fallback to original string if substitution fails
        rstr = right.strip()

    # After normalizing percent tokens, evaluate the right-hand expression safely.
    try:
        b = _safe_eval_numeric_expr(rstr) if rstr else 0.0
    except Exception:
        raise ValueError(f'Unsupported or unparseable numeric expression on right side: {right!r}')

    # If a third parameter provided (voigt), parse it as numeric alpha
    c = None
    if third is not None:
        try:
            third_norm = re.sub(r'([+-]?\d+(?:\.\d+)?)\s*%', r'(\1/100)', third)
            c = _safe_eval_numeric_expr(third_norm)
        except Exception:
            try:
                c = _safe_eval_numeric_expr(third)
            except Exception:
                c = None

    if kind in {'gaussian', 'normal'}:
        return ('gaussian', a, abs(b))
    if kind == 'voigt':
        # return third param if present as c
        return ('voigt', a, abs(b), c)
    if kind == 'uniform':
        return ('uniform', a, b)

    raise ValueError(f'Unsupported distribution spec: {spec!r}')


def load_standard_mm_psf_defs(path: Path) -> dict:
    """Load standard MM_PSF definitions from the baseline workbook (headerless table near K1)."""
    try:
        df = pd.read_excel(path, sheet_name='MM_PSF', header=None, engine='openpyxl')
    except Exception:
        return {}
    std = {}
    # start at row 0, col 10 (K)
    start_row = 0
    start_col = 10
    if df.shape[0] <= start_row or df.shape[1] <= start_col:
        return {}
    # headers are at start_col..start_col+4
    headers = [df.iloc[start_row, start_col + i] for i in range(5) if start_col + i < df.shape[1]]
    row_idx = start_row + 1
    while row_idx < df.shape[0]:
        name = df.iloc[row_idx, start_col]
        if pd.isna(name) or str(name).strip() == '':
            break
        key_name = str(name).strip()
        dist_def = {'name': key_name}
        # parse next 4 cells sigma_rad, sigma_azi, alpha_rad, alpha_azi
        for i, param in enumerate(['sigma_rad', 'sigma_azi', 'alpha_rad', 'alpha_azi']):
            cell_value = None
            if start_col + 1 + i < df.shape[1]:
                cell_value = df.iloc[row_idx, start_col + 1 + i]
            if pd.isna(cell_value) or cell_value == '':
                dist_def[param] = None
                continue
            cell_str = str(cell_value).strip()
            try:
                res = _parse_standard_dist_spec(cell_str)
                # handle voigt which may return (kind,a,b,c)
                if isinstance(res, tuple) and len(res) >= 1:
                    kind = res[0]
                else:
                    raise ValueError('Unexpected parse result')

                if kind == 'fixed':
                    _, a, _ = res
                    dist_def[param] = {'dist': 'fixed', 'value': float(a)}
                elif kind == 'gaussian':
                    _, a, b = res
                    dist_def[param] = {'dist': 'gaussian', 'mean': float(a), 'sigma': float(abs(b))}
                elif kind == 'uniform':
                    _, a, b = res
                    lo = float(min(a, b))
                    hi = float(max(a, b))
                    dist_def[param] = {'dist': 'uniform', 'min': lo, 'max': hi}
                elif kind == 'voigt':
                    # voigt returns (voigt, a, b, c?) where c is optional alpha
                    if param.startswith('sigma'):
                        # treat voigt sigma as gaussian sample for per-mm sampling
                        _, a, b, *rest = res
                        dist_def[param] = {'dist': 'gaussian', 'mean': float(a), 'sigma': float(abs(b))}
                    elif param.startswith('alpha'):
                        # if alpha provided, set as fixed value; else leave None
                        if len(res) >= 4 and res[3] is not None:
                            _, a, b, c = res
                            try:
                                dist_def[param] = {'dist': 'fixed', 'value': float(c)}
                            except Exception:
                                dist_def[param] = None
                        else:
                            dist_def[param] = None
                    else:
                        # generic fallback
                        _, a, b, c = res if len(res) >= 4 else (res[0], res[1], res[2], None)
                        dist_def[param] = {'dist': 'gaussian', 'mean': float(a), 'sigma': float(abs(b))}
                else:
                    dist_def[param] = None
            except Exception:
                # fallback: support explicit 'gamma(mean,XX%*mean)' strings present
                try:
                    import re as _re
                    m = _re.search(r'gamma\s*\(\s*([0-9.]+)\s*,\s*([0-9.]+)\s*%?', cell_str, flags=_re.IGNORECASE)
                    if m:
                        mu = float(m.group(1))
                        pct = float(m.group(2)) / 100.0
                        sigma = mu * pct
                        dist_def[param] = {'dist': 'gamma', 'mean': float(mu), 'sigma': float(sigma)}
                    else:
                        dist_def[param] = None
                except Exception:
                    dist_def[param] = None
            except Exception:
                dist_def[param] = None
        std[key_name] = dist_def
        row_idx += 1
    return std


def load_standard_alignment_defs(path: Path) -> dict:
    """Load standard Alignment presets from the baseline workbook.

    Tries to follow the same layout as `gui_distributions.load_standard_alignment_presets`:
    preset names in column G (index 6) and variable specs in subsequent columns.
    Returns mapping preset_name -> {param_label: spec_str}
    """
    try:
        df = pd.read_excel(path, sheet_name='Alignment', header=None, engine='openpyxl')
    except Exception:
        return {}

    std = {}
    start_row = 0
    name_col = 6
    first_var_col = 7
    if df.shape[0] <= start_row or df.shape[1] <= first_var_col:
        return {}

    headers = {}
    for c in range(first_var_col, min(df.shape[1], first_var_col + 16)):
        h = df.iloc[start_row, c]
        if pd.isna(h) or str(h).strip() == '':
            continue
        headers[c] = str(h).strip()

    # Map base variable names (like d_align_rad_) to the full GUI param labels
    param_map = {}
    try:
        for p in DATA_TYPES['Alignment']['params']:
            base = str(p).split(' ')[0].strip()
            param_map[base] = p
    except Exception:
        param_map = {}

    row_idx = start_row + 1
    while row_idx < df.shape[0]:
        preset_name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
        if pd.isna(preset_name) or str(preset_name).strip() == '':
            break
        preset_name = str(preset_name).strip()
        preset_specs = {}
        for c, h in headers.items():
            if c >= df.shape[1]:
                continue
            raw_spec = df.iloc[row_idx, c]
            if pd.isna(raw_spec) or str(raw_spec).strip() == '':
                continue
            var = str(h).strip()
            if var.endswith('_'):
                var = var[:-1]
            # Store the raw spec under multiple keys to be resilient when
            # `DATA_TYPES` was not importable (e.g. headless environments).
            clean_spec = str(raw_spec).strip()
            preset_specs[var] = clean_spec
            # also store under base name (without units) to allow lookups
            base_name = var.split(' ')[0] if ' ' in var else var
            preset_specs[base_name] = clean_spec
            if var in param_map:
                preset_specs[param_map[var]] = clean_spec

        if preset_specs:
            std[preset_name] = preset_specs
        row_idx += 1

    return std


def _find_std_mm_psf_key(name: str, std_map: dict) -> str:
    """Try to match a combo MM_PSF string to a key in `std_map`.

    Returns the matched key (exact or fuzzy) or None if not found.
    """
    if not name or not std_map:
        return None
    try:
        import re
        def _norm(s):
            s = str(s).lower()
            s = s.replace('"', '')
            s = re.sub(r'mm_psf\d*', '', s)
            s = re.sub(r'[\(\)\[\]]', ' ', s)
            s = re.sub(r'[^a-z0-9%\s]', ' ', s)
            s = re.sub(r'\s+', ' ', s).strip()
            return s

        target = _norm(name)
        # exact match
        if name in std_map:
            return name
        # try case-insensitive exact
        for k in std_map.keys():
            if k.lower() == name.lower():
                return k

        # fuzzy contains / contained
        for k in std_map.keys():
            nk = _norm(k)
            if nk and (nk in target or target in nk):
                return k
        return None
    except Exception:
        return None


def load_standard_thermal_defs(path: Path) -> dict:
    """Load standard Thermal presets from the baseline workbook.

    Mirrors `load_standard_alignment_defs` but for the `Thermal` sheet.
    """
    try:
        df = pd.read_excel(path, sheet_name='Thermal', header=None, engine='openpyxl')
    except Exception:
        return {}

    std = {}
    start_row = 0
    name_col = 6
    first_var_col = 7
    if df.shape[0] <= start_row or df.shape[1] <= first_var_col:
        return {}

    headers = {}
    for c in range(first_var_col, min(df.shape[1], first_var_col + 16)):
        h = df.iloc[start_row, c]
        if pd.isna(h) or str(h).strip() == '':
            continue
        headers[c] = str(h).strip()

    param_map = {}
    try:
        for p in DATA_TYPES['Thermal']['params']:
            base = str(p).split(' ')[0].strip()
            param_map[base] = p
    except Exception:
        param_map = {}

    row_idx = start_row + 1
    while row_idx < df.shape[0]:
        preset_name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
        if pd.isna(preset_name) or str(preset_name).strip() == '':
            break
        preset_name = str(preset_name).strip()
        preset_specs = {}
        for c, h in headers.items():
            if c >= df.shape[1]:
                continue
            raw_spec = df.iloc[row_idx, c]
            if pd.isna(raw_spec) or str(raw_spec).strip() == '':
                continue
            var = str(h).strip()
            if var.endswith('_'):
                var = var[:-1]
            clean_spec = str(raw_spec).strip()
            preset_specs[var] = clean_spec
            base_name = var.split(' ')[0] if ' ' in var else var
            preset_specs[base_name] = clean_spec
            if var in param_map:
                preset_specs[param_map[var]] = clean_spec

        if preset_specs:
            std[preset_name] = preset_specs
        row_idx += 1

    return std


def load_standard_gravity_defs(path: Path) -> dict:
    """Load standard Gravity offload presets from the baseline workbook.

    Mirrors `load_standard_alignment_defs` but for the `Gravity offload` sheet.
    """
    try:
        df = pd.read_excel(path, sheet_name='Gravity offload', header=None, engine='openpyxl')
    except Exception:
        return {}

    std = {}
    start_row = 0
    name_col = 6
    first_var_col = 7
    if df.shape[0] <= start_row or df.shape[1] <= first_var_col:
        return {}

    headers = {}
    for c in range(first_var_col, min(df.shape[1], first_var_col + 16)):
        h = df.iloc[start_row, c]
        if pd.isna(h) or str(h).strip() == '':
            continue
        headers[c] = str(h).strip()

    param_map = {}
    try:
        for p in DATA_TYPES['Gravity offload']['params']:
            base = str(p).split(' ')[0].strip()
            param_map[base] = p
    except Exception:
        param_map = {}

    row_idx = start_row + 1
    while row_idx < df.shape[0]:
        preset_name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
        if pd.isna(preset_name) or str(preset_name).strip() == '':
            break
        preset_name = str(preset_name).strip()
        preset_specs = {}
        for c, h in headers.items():
            if c >= df.shape[1]:
                continue
            raw_spec = df.iloc[row_idx, c]
            if pd.isna(raw_spec) or str(raw_spec).strip() == '':
                continue
            var = str(h).strip()
            if var.endswith('_'):
                var = var[:-1]
            clean_spec = str(raw_spec).strip()
            preset_specs[var] = clean_spec
            base_name = var.split(' ')[0] if ' ' in var else var
            preset_specs[base_name] = clean_spec
            if var in param_map:
                preset_specs[param_map[var]] = clean_spec

        if preset_specs:
            std[preset_name] = preset_specs
        row_idx += 1

    return std


def _sanitize_filename(s: str) -> str:
    import re
    s = str(s)
    s = s.replace('"', '').replace("'", '')
    s = re.sub(r'\s+', '_', s)
    s = re.sub(r'[^A-Za-z0-9._-]', '', s)
    return s


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--workers", type=int, default=64, help="SENS_MAX_WORKERS to use")
    p.add_argument("--persist", action="store_true", help="Set SENS_PERSIST_TMP=1 to keep tmp workbooks")
    p.add_argument("--baseline", type=str, default=None, help="Path to baseline workbook to copy into input")
    p.add_argument("--check-and-fix", type=str, default=None, help="Path to a single workbook to check/fix MM_PSF (or 'auto' to auto-detect latest 4.3 placed file)")
    p.add_argument("--dry-run", action="store_true", help="Create folders and copy baseline but don't execute driver")
    p.add_argument("--generate-only", action="store_true", help="Only generate input workbooks, do not run jobs")
    p.add_argument("--no-excel", action="store_true", help="Do not create Excel workbooks per combo; use pickled DataFrame inputs instead")
    p.add_argument("--csv-only", action="store_true", default=False, help="Do not create Excel workbooks; generate per-combo MM_PSF CSVs and call main.py on them")
    p.add_argument("--no-csv", dest='csv_only', action='store_false', help="Disable CSV-only fast path and create Excel workbooks instead")
    p.add_argument("--force-timestamp", type=str, default=None, help="Force the UTC timestamp used in generated filenames (e.g. 20260125T000945Z)")
    # jitter removed: deterministic fixed presets remain fixed
    args = p.parse_args()

    # Force excel-workbook mode to ensure per-combo presets are fully applied
    # (CSV-only path has been flaky for applying workbook presets reliably).
    args.csv_only = False

    ensure_dirs()

    # Determine effective persistence: consider args and SENS_PERSIST_TMP env var.
    env_persist = os.environ.get('SENS_PERSIST_TMP')
    persist_disabled_by_env = env_persist is not None and str(env_persist).strip() in ('0', 'false', 'False')
    non_persistent = (not args.persist) or persist_disabled_by_env
    if non_persistent:
        if not args.baseline:
            args.baseline = str(ROOT / 'Distributions' / 'TestDistribution.xlsx')
        # force worker count for non-persistent runs
        args.workers = 64
        print(f"Non-persistent run (persist flag={args.persist}, SENS_PERSIST_TMP={env_persist}): using baseline {args.baseline}, forcing workers={args.workers}")

    # Prepare input directory: ephemeral tempdir when non-persistent, otherwise sensitivity/input
    # Always persist input workbooks in the `sensitivity/input` folder so per-combo
    # presets and partial results are available after the run.
    input_dir = SENS_DIR / 'input'
    input_dir.mkdir(parents=True, exist_ok=True)
    print(f"Using input directory: {input_dir}")
    if args.baseline:
        try:
            if not getattr(args, 'csv_only', False):
                out = copy_baseline(Path(args.baseline), dest_dir=input_dir)
                print(f"Copied baseline to: {out}")
            else:
                # For csv-only mode we do not need to copy the baseline workbook
                print(f"csv-only mode: skipping baseline workbook copy ({args.baseline})")
        except Exception as e:
            print(f"Failed to copy baseline: {e}")
            return

    if args.dry_run:
        print("Dry run complete — folders created and baseline copied (if provided)")
        return

    # Read sensitivity specification from sensitivity/sensitivity_input.xlsx
    sens_path = SENS_DIR / 'sensitivity_input.xlsx'
    if not sens_path.exists():
        print(f"Sensitivity file not found: {sens_path}")
        return

    try:
        sens_df = pd.read_excel(sens_path, engine='openpyxl')
    except Exception as e:
        print(f"Failed to read sensitivity spreadsheet: {e}")
        return

    # Build choices per column: collect non-null values per column
    choices_map = {}
    for col in sens_df.columns:
        vals = []
        for v in sens_df[col].dropna().tolist():
            if isinstance(v, str) and (',' in v or '\n' in v):
                parts = [p.strip() for p in re.split('[,\n]', v) if p.strip()]
                vals.extend(parts)
            else:
                vals.append(v)
        # unique while preserving order
        seen = set()
        uniq = []
        for v in vals:
            if v not in seen:
                uniq.append(v)
                seen.add(v)
        if uniq:
            choices_map[col] = uniq

    if not choices_map:
        print("No choices found in sensitivity_input.xlsx. Nothing to run.")
        return

    # Cartesian product of choices
    cols = list(choices_map.keys())
    combos = []
    for prod in product(*(choices_map[c] for c in cols)):
        combos.append(dict(zip(cols, prod)))

    print(f"Prepared {len(combos)} combos from sensitivity_input.xlsx")

    # Prepare baseline workbook path
    baseline = Path(args.baseline) if args.baseline else (ROOT / 'Distributions' / 'TestDistribution.xlsx')
    if not baseline.exists():
        print(f"Baseline workbook not found: {baseline}")
        return

    # Minimal startup summary for runs
    try:
        print(f"Will run {len(combos)} combos; baseline={baseline}; persist={args.persist}; workers={args.workers}")
    except Exception:
        print(f"Will run {len(combos)} combos; persist={args.persist}; workers={args.workers}")

    # Load MM config to obtain MM # list and num_mm
    try:
        mm_cfg = pd.read_excel(baseline, sheet_name='MM configuration', engine='openpyxl')
        if 'MM #' in mm_cfg.columns:
            mm_list = mm_cfg['MM #'].dropna().tolist()
        else:
            mm_list = list(range(1, len(mm_cfg) + 1))
        num_mm = len(mm_list)
    except Exception:
        mm_list = list(range(1, 9))
        num_mm = len(mm_list)

    # Load standard MM_PSF defs from baseline
    std_mm_psf = load_standard_mm_psf_defs(baseline)
    # Load standard Alignment presets from baseline (if any)
    std_alignment = load_standard_alignment_defs(baseline)

    # If sensitivity_input.xlsx contains an A_eff placeholder like '1 keV [row#]',
    # expand that combo once per distinct Row # defined in the `MM configuration` sheet.
    try:
        # attempt to read row numbers from mm_cfg (flexible header)
        row_col = None
        if 'mm_cfg' in locals() and mm_cfg is not None:
            for c in mm_cfg.columns:
                if isinstance(c, str) and 'row' in c.lower():
                    row_col = c
                    break
            if row_col is None and len(mm_cfg.columns) >= 3:
                row_col = mm_cfg.columns[2]
        row_nums = []
        if row_col is not None:
            row_nums = sorted(set([int(x) for x in pd.to_numeric(mm_cfg[row_col], errors='coerce').dropna().tolist()]))
        # expand combos if needed
        expanded = []
        for combo in combos:
            a = combo.get('A_eff')
            if isinstance(a, str):
                m_row = re.match(r'^(.*?)\s*\[\s*row\s*#\s*\]$', a, flags=re.IGNORECASE)
                if m_row and row_nums:
                    base = m_row.group(1).strip()
                    for rn in row_nums:
                        nc = dict(combo)
                        nc['A_eff'] = f"{base} [row{rn}]"
                        expanded.append(nc)
                    continue
            expanded.append(combo)
        combos = expanded
        if len(expanded) != len(combos):
            print(f"Expanded A_eff [row#] into {len(expanded)} combos")
    except Exception:
        pass

    # helper to write dataframe to excel sheet (overwrite)
    def write_sheet(workbook_path: Path, sheet_name: str, df_to_write: pd.DataFrame):
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        if not workbook_path.exists():
            # create new workbook
            from openpyxl import Workbook
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        else:
            wb = load_workbook(workbook_path)

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)

        ws = wb.create_sheet(sheet_name)
        # write header + rows
        for r_idx, row in enumerate(dataframe_to_rows(df_to_write, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(workbook_path)

    def write_mmpsf_preserve_template(baseline_path: Path, workbook_path: Path, df_gen: pd.DataFrame):
        """Write an MM_PSF sheet that preserves the baseline workbook's template columns.

        The baseline `MM_PSF` sheet often contains a right-hand table of standard
        distributions (headerless) starting at a later column. To keep those
        template columns available for downstream special-casing (e.g. 4.3" cases),
        we merge the generated per-MM columns with the baseline tail columns and
        append any template rows that follow the per-MM rows.
        """
        import pandas as _pd
        from openpyxl import load_workbook

        # Read raw baseline sheet without headers so we can preserve layout
        try:
            raw = _pd.read_excel(baseline_path, sheet_name='MM_PSF', engine='openpyxl', header=None)
        except Exception:
            # Fallback: write simple sheet
            write_sheet(workbook_path, 'MM_PSF', df_gen)
            return

        # Find the header row in the raw baseline sheet. Some templates include
        # several preamble rows before the MM_PSF header (e.g. metadata), so
        # scan for the header row by looking for expected column name fragments.
        raw_header = []
        header_idx = 0
        if raw.shape[0] > 0:
            # detect header by looking for presence of key labels in a row
            key_frags = ['m_rad', 'm_azi', 'sigma_rad', 'sigma_azi', 'mm #']
            max_scan = min(raw.shape[0], 200)
            found = False
            for r in range(max_scan):
                row_vals = [str(x).strip().lower() for x in raw.iloc[r].tolist()]
                matches = sum(1 for k in key_frags if any(k in v for v in row_vals))
                if matches >= 2:
                    header_idx = r
                    raw_header = raw.iloc[r].tolist()
                    found = True
                    break
            if not found:
                # fallback to first row
                header_idx = 0
                raw_header = raw.iloc[0].tolist()

        # Prepare generated column names and count before we manipulate headers.
        gen_cols = df_gen.columns.tolist()
        n_gen = len(gen_cols)

        # Build the tail header directly from the raw sheet's columns that
        # follow the left-most `n_gen` columns. This avoids slicing a
        # potentially truncated or mis-detected `raw_header` list and keeps
        # the preserved template columns aligned with the original layout.
        try:
            if raw.shape[1] > n_gen:
                tail_header = raw.iloc[header_idx, n_gen:].tolist()
            else:
                tail_header = []
            raw_header = gen_cols + tail_header
        except Exception:
            # On any error, fall back to using generated columns as header
            raw_header = list(gen_cols)

        # Ensure header length at least n_gen (safeguard)
        if len(raw_header) < n_gen:
            raw_header = gen_cols + []

        rows = []
        # Preserve any preamble rows before the detected header
        if header_idx > 0:
            for r in range(header_idx):
                rows.append(raw.iloc[r].tolist())
        rows.append(raw_header)

        # Number of per-MM rows to write
        num_mm_rows = len(df_gen)


        # For each MM row, take generated values for the left columns.
        # Do NOT copy tail/template values from baseline per-row (these often
        # represent a separate right-hand table). Instead, leave tail columns
        # empty for per-MM rows and append any baseline template rows after
        # all per-MM rows below.
        tail_len = max(0, raw.shape[1] - n_gen) if raw.shape[0] > 0 else 0
        for i in range(num_mm_rows):
            gen_row = [df_gen.iloc[i].get(c) for c in gen_cols]
            tail = [None] * tail_len
            rows.append(gen_row + tail)

        # After writing per-MM rows, append any baseline rows that follow the
        # per-MM block (these are template rows that should not be merged
        # into per-MM rows).
        template_start = header_idx + 1 + num_mm_rows
        if raw.shape[0] > template_start:
            for r in range(template_start, raw.shape[0]):
                raw_row = list(raw.iloc[r].tolist())
                # ensure row length matches header
                if len(raw_row) < len(raw_header):
                    raw_row = raw_row + [None] * (len(raw_header) - len(raw_row))
                elif len(raw_row) > len(raw_header):
                    raw_row = raw_row[:len(raw_header)]
                rows.append(raw_row)

        # Write assembled rows into the workbook
        wb = load_workbook(workbook_path)
        if 'MM_PSF' in wb.sheetnames:
            ws = wb['MM_PSF']
            wb.remove(ws)
        ws = wb.create_sheet('MM_PSF')
        for r in rows:
            ws.append(r)
        wb.save(workbook_path)

    def write_mmpsf_preserve_template_and_expand(baseline_path: Path, workbook_path: Path, df_gen: pd.DataFrame, std_defs: dict, chosen_name: str | None = None):
        """Write MM_PSF preserving template and expand the chosen preset into numeric template cells.

        If `chosen_name` corresponds to a standard preset in `std_defs`, replace the
        template row for that preset with numeric mean/sigma/alpha values so downstream
        code that reads `sigma_rad_`/`sigma_azi_` sees numeric values instead of
        percent-expression strings.
        """
        from openpyxl import load_workbook

        # First assemble using the existing preservative writer
        try:
            write_mmpsf_preserve_template(baseline_path, workbook_path, df_gen)
        except Exception:
            # fallback to direct write
            write_sheet(workbook_path, 'MM_PSF', df_gen)
            return

        if not chosen_name or chosen_name not in (std_defs or {}):
            return

        # Load workbook and locate the template row for chosen_name
        wb = load_workbook(workbook_path)
        if 'MM_PSF' not in wb.sheetnames:
            wb.save(workbook_path)
            return
        ws = wb['MM_PSF']

        n_gen = len(df_gen.columns)
        # Template columns expected after generated columns: name, sigma_rad_, sigma_azi_, alpha_rad_, alpha_azi_
        name_col = n_gen + 1
        sigma_rad_col = n_gen + 2
        sigma_azi_col = n_gen + 3
        alpha_rad_col = n_gen + 4
        alpha_azi_col = n_gen + 5

        entry = std_defs.get(chosen_name)
        if not entry:
            wb.save(workbook_path)
            return

        # numeric values to write (arcsec for sigma_* templates)
        def _maybe_val(param):
            v = entry.get(param)
            if isinstance(v, dict):
                if v.get('dist') == 'fixed':
                    return float(v.get('value', 0.0))
                if v.get('dist') == 'gaussian':
                    return float(v.get('mean', 0.0))
                if v.get('dist') == 'uniform':
                    # represent uniform as 'lo-hi' string
                    return f"{v.get('min',0.0)}-{v.get('max',0.0)}"
            return None

        def _norm(s):
            if s is None:
                return ''
            s2 = str(s).strip().lower()
            s2 = s2.replace('"', '').replace("'", '')
            s2 = s2.replace('%', '')
            s2 = ''.join(s2.split())
            return s2

        chosen_norm = _norm(chosen_name)
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=name_col).value
            if cell is None:
                continue
            if _norm(cell) == chosen_norm:
                # write numeric replacements where available
                v_sr = _maybe_val('sigma_rad')
                v_sa = _maybe_val('sigma_azi')
                v_ar = _maybe_val('alpha_rad')
                v_aa = _maybe_val('alpha_azi')
                if v_sr is not None:
                    ws.cell(row=r, column=sigma_rad_col, value=v_sr)
                if v_sa is not None:
                    ws.cell(row=r, column=sigma_azi_col, value=v_sa)
                # write alpha values: numeric if fixed, otherwise write '-' to
                # indicate not-applicable for standard gaussian presets
                if v_ar is not None:
                    ws.cell(row=r, column=alpha_rad_col, value=v_ar)
                else:
                    ws.cell(row=r, column=alpha_rad_col, value='-')
                if v_aa is not None:
                    ws.cell(row=r, column=alpha_azi_col, value=v_aa)
                else:
                    ws.cell(row=r, column=alpha_azi_col, value='-')
                # Update left-side 'distribution' column if present to reflect voigt/gauss
                try:
                    distrib_col = None
                    for cidx in range(1, n_gen + 1):
                        hv = ws.cell(row=1, column=cidx).value
                        if isinstance(hv, str) and 'distrib' in hv.lower():
                            distrib_col = cidx
                            break
                    if distrib_col is not None:
                        lname = str(chosen_name).lower()
                        if 'voigt' in lname or 'pseudo' in lname:
                            ws.cell(row=r, column=distrib_col, value='pseudo-voigt')
                        elif 'gauss' in lname or 'gaussian' in lname:
                            ws.cell(row=r, column=distrib_col, value='gaussian')
                        else:
                            # fallback: if any alpha is fixed, set as pseudo-voigt
                            if (entry.get('alpha_rad') and isinstance(entry.get('alpha_rad'), dict) and entry.get('alpha_rad').get('dist') == 'fixed') or (entry.get('alpha_azi') and isinstance(entry.get('alpha_azi'), dict) and entry.get('alpha_azi').get('dist') == 'fixed'):
                                ws.cell(row=r, column=distrib_col, value='pseudo-voigt')
                            else:
                                ws.cell(row=r, column=distrib_col, value='gaussian')
                except Exception:
                    pass
                break

        # If the chosen preset corresponds to a standard definition that
        # provides fixed values for sigma/alpha, enforce those fixed values
        # in the first `num_mm` per-MM rows so generated workbooks always
        # contain canonical numeric entries for fixed presets.
        try:
            if chosen_name and (std_defs or {}).get(chosen_name):
                entry = std_defs.get(chosen_name)
                # compute number of MM rows to enforce: prefer df_gen length
                try:
                    num_enforce = int(len(df_gen)) if df_gen is not None else int(num_mm)
                except Exception:
                    num_enforce = int(num_mm)

                # find header row (assume header at row 1) and locate sigma/alpha columns
                # tolerant matching for names like 'sigma_rad', 'sigma_rad [arcsec]', 'sigma_rad_'
                header_cells = [str(ws.cell(row=1, column=c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
                def _find_col(name_frag):
                    for idx, hv in enumerate(header_cells, start=1):
                        if name_frag in hv:
                            return idx
                    # try underscore-suffixed or prefix matches
                    for idx, hv in enumerate(header_cells, start=1):
                        if hv.replace(' ', '').startswith(name_frag.replace(' ', '')):
                            return idx
                    return None

                sr_col_idx = _find_col('sigma_rad')
                sa_col_idx = _find_col('sigma_azi')
                ar_col_idx = _find_col('alpha_rad')
                aa_col_idx = _find_col('alpha_azi')

                def _val_for(param_key):
                    v = entry.get(param_key)
                    if isinstance(v, dict):
                        if v.get('dist') == 'fixed':
                            return float(v.get('value', 0.0))
                        if v.get('dist') == 'gaussian':
                            return float(v.get('mean', 0.0))
                    return None

                v_sr = _val_for('sigma_rad')
                v_sa = _val_for('sigma_azi')
                v_ar = _val_for('alpha_rad')
                v_aa = _val_for('alpha_azi')

                # write fixed values into the first num_enforce rows (rows 2..1+num_enforce)
                for rr in range(2, min(ws.max_row, 1 + num_enforce) + 1):
                    try:
                        if sr_col_idx and v_sr is not None:
                            ws.cell(row=rr, column=sr_col_idx, value=float(v_sr))
                        if sa_col_idx and v_sa is not None:
                            ws.cell(row=rr, column=sa_col_idx, value=float(v_sa))
                        if ar_col_idx and v_ar is not None:
                            ws.cell(row=rr, column=ar_col_idx, value=float(v_ar))
                        if aa_col_idx and v_aa is not None:
                            ws.cell(row=rr, column=aa_col_idx, value=float(v_aa))
                    except Exception:
                        continue
                # Also ensure the per-MM "distribution" column (if present) is set
                try:
                    distrib_col = None
                    for cidx in range(1, n_gen + 1):
                        hv = ws.cell(row=1, column=cidx).value
                        if isinstance(hv, str) and 'distrib' in hv.lower():
                            distrib_col = cidx
                            break
                    if distrib_col is not None:
                        lname = str(chosen_name).lower()
                        if 'voigt' in lname or 'pseudo' in lname:
                            dlabel = 'pseudo-voigt'
                        elif 'gauss' in lname or 'gaussian' in lname:
                            dlabel = 'gaussian'
                        else:
                            if (entry.get('alpha_rad') and isinstance(entry.get('alpha_rad'), dict) and entry.get('alpha_rad').get('dist') == 'fixed') or (entry.get('alpha_azi') and isinstance(entry.get('alpha_azi'), dict) and entry.get('alpha_azi').get('dist') == 'fixed'):
                                dlabel = 'pseudo-voigt'
                            else:
                                dlabel = 'gaussian'
                        for rr in range(2, min(ws.max_row, 1 + num_enforce) + 1):
                            try:
                                ws.cell(row=rr, column=distrib_col, value=dlabel)
                            except Exception:
                                continue
                except Exception:
                    pass

                # If any sigma/alpha definitions are non-fixed (variable preset),
                # sample per-MM values now so the first num_enforce rows reflect
                # the preset's distribution (gamma/gaussian/uniform) rather than
                # leaving template placeholders or non-numeric strings.
                try:
                    need_sample = False
                    for k in ('sigma_rad', 'sigma_azi', 'alpha_rad', 'alpha_azi'):
                        vvv = entry.get(k)
                        if isinstance(vvv, dict) and vvv.get('dist') != 'fixed':
                            need_sample = True
                            break
                    # also treat preset names containing 'variable' or '%' as variable
                    if not need_sample and isinstance(chosen_name, str) and ('variable' in chosen_name.lower() or '%' in chosen_name):
                        need_sample = True
                    if need_sample:
                        _sample_per_mm_sigmas_and_write(workbook_path, std_defs, chosen_name, num_enforce)
                except Exception:
                    pass
        except Exception:
            pass

        wb.save(workbook_path)

    def _sample_per_mm_sigmas_and_write(workbook_path: Path, std_defs: dict, chosen_name: str | None, num_mm: int):
        """Ensure per-MM `sigma_rad`/`sigma_azi` columns contain numeric values.

        If any of the first `num_mm` rows have non-numeric or non-positive values,
        sample numeric values from the standard preset definition and overwrite
        the per-MM sigma columns in-place.
        """
        try:
            import numpy as _np
            fp = Path(workbook_path)
            mm_df = pd.read_excel(fp, sheet_name='MM_PSF', engine='openpyxl')
        except Exception:
            return

        if not chosen_name and not std_defs:
            return

        entry = std_defs.get(chosen_name) if std_defs else None
        # Fallback: attempt tolerant matching if exact key not found
        if entry is None and std_defs and chosen_name:
            try:
                key_low = str(chosen_name).strip().lower()
                def _norm(s):
                    return str(s).strip().lower().replace('"', '').replace('\u201c', '').replace('\u201d', '').replace(' ', '')
                n_ch = _norm(chosen_name)
                for k in std_defs.keys():
                    if not k:
                        continue
                    if _norm(k) == n_ch or key_low in str(k).lower() or str(k).lower() in key_low:
                        entry = std_defs.get(k)
                        break
            except Exception:
                entry = None

        # find per-MM sigma columns
        sigma_rad_cols = [c for c in mm_df.columns if isinstance(c, str) and 'sigma_rad' in c.lower()]
        sigma_azi_cols = [c for c in mm_df.columns if isinstance(c, str) and 'sigma_azi' in c.lower()]
        # find per-MM alpha columns if present
        alpha_rad_cols = [c for c in mm_df.columns if isinstance(c, str) and 'alpha_rad' in c.lower()]
        alpha_azi_cols = [c for c in mm_df.columns if isinstance(c, str) and 'alpha_azi' in c.lower()]
        if not sigma_rad_cols or not sigma_azi_cols:
            return

        sr_col = sigma_rad_cols[0]
        sa_col = sigma_azi_cols[0]
        # Ensure columns are float dtype to avoid pandas dtype-mismatch warnings
        try:
            mm_df[sr_col] = pd.to_numeric(mm_df[sr_col], errors='coerce').astype(float)
        except Exception:
            mm_df[sr_col] = mm_df[sr_col].astype('float64', errors='ignore')
        try:
            mm_df[sa_col] = pd.to_numeric(mm_df[sa_col], errors='coerce').astype(float)
        except Exception:
            mm_df[sa_col] = mm_df[sa_col].astype('float64', errors='ignore')
        # Determine distribution defs from preset (preferred) or infer from
        # template tail placeholders (e.g. 'gamma(...)') when available.
        sr_def = entry.get('sigma_rad') if entry else None
        sa_def = entry.get('sigma_azi') if entry else None

        def _infer_from_template(col_key):
            try:
                if col_key == 'sigma_rad' and template_sr_col and template_sr_col in mm_df.columns:
                    ser = mm_df[template_sr_col].astype(str).dropna()
                    if len(ser):
                        return _parse_gamma_placeholder(ser.iloc[0])
                if col_key == 'sigma_azi' and template_sa_col and template_sa_col in mm_df.columns:
                    ser = mm_df[template_sa_col].astype(str).dropna()
                    if len(ser):
                        return _parse_gamma_placeholder(ser.iloc[0])
            except Exception:
                return None
            return None

        if not isinstance(sr_def, dict):
            inf = _infer_from_template('sigma_rad')
            if isinstance(inf, dict):
                sr_def = inf
        if not isinstance(sa_def, dict):
            inf = _infer_from_template('sigma_azi')
            if isinstance(inf, dict):
                sa_def = inf

        # check existing values for first num_mm rows
        sr_vals = pd.to_numeric(mm_df.loc[:num_mm-1, sr_col], errors='coerce')
        sa_vals = pd.to_numeric(mm_df.loc[:num_mm-1, sa_col], errors='coerce')

        # Force sampling when the preset itself defines a non-fixed distribution
        force_sample = False
        try:
            if isinstance(sr_def, dict) and sr_def.get('dist') != 'fixed':
                force_sample = True
            if isinstance(sa_def, dict) and sa_def.get('dist') != 'fixed':
                force_sample = True
        except Exception:
            force_sample = False

        # Also trigger sampling if the template tail columns (e.g. 'sigma_rad_')
        # contain placeholder strings like 'gamma(...)' which need to be expanded.
        template_sr_col = None
        template_sa_col = None
        for c in mm_df.columns:
            if isinstance(c, str) and c.strip().lower().startswith('sigma_rad') and c.strip().endswith('_'):
                template_sr_col = c
            if isinstance(c, str) and c.strip().lower().startswith('sigma_azi') and c.strip().endswith('_'):
                template_sa_col = c

        template_placeholders = False
        try:
            if template_sr_col is not None and mm_df[template_sr_col].astype(str).str.contains('gamma|gauss|gaussian', case=False, na=False).any():
                template_placeholders = True
            if template_sa_col is not None and mm_df[template_sa_col].astype(str).str.contains('gamma|gauss|gaussian', case=False, na=False).any():
                template_placeholders = True
        except Exception:
            template_placeholders = False

        # Determine if the chosen preset should use gamma sampling even when
        # the standard table lists a Gaussian (e.g. preset names like
        # '50% Variable ...'). In those cases we treat the mean/std as
        # gamma parameters (via shape/scale conversion) to produce positive,
        # skewed samples.
        force_gamma = False
        try:
            if chosen_name and isinstance(chosen_name, str) and ('variable' in chosen_name.lower() or '%' in chosen_name):
                force_gamma = True
        except Exception:
            force_gamma = False

        # If force_gamma is requested and we can parse a variability percent
        # or an explicit alpha token from the preset name, extract them now
        # so sampling decisions below can unconditionally trigger.
        var_pct = None
        parsed_alpha = None
        try:
            if force_gamma and isinstance(chosen_name, str):
                import re as _re
                m = _re.search(r'(\d+)\s*%\s*variable', chosen_name, flags=_re.IGNORECASE)
                if not m:
                    m = _re.search(r'^(\d+)\s*%\b', chosen_name)
                if m:
                    var_pct = float(m.group(1)) / 100.0
                m2 = _re.search(r'alpha\s*[:(]?\s*(\d+)\s*%?', chosen_name, flags=_re.IGNORECASE)
                if m2:
                    parsed_alpha = float(m2.group(1)) / 100.0
        except Exception:
            var_pct = None
            parsed_alpha = None

        # If the preset name indicates Variable(...) semantics, unconditionally
        # force sampling for all MM rows so the preset distributions dominate
        # any existing per-row numeric canonical values.
        need_sampling = force_gamma or force_sample or sr_vals.isnull().any() or sa_vals.isnull().any() or (sr_vals <= 0).any() or (sa_vals <= 0).any() or template_placeholders
        if not need_sampling:
            return

        # deterministic RNG seed base from filename + chosen_name
        import hashlib
        h = int(hashlib.sha256((fp.name + str(chosen_name)).encode('utf-8')).hexdigest()[:8], 16)

        # (force_gamma, var_pct, parsed_alpha already determined above)

        # When force_gamma is requested and we have a standard preset entry,
        # build explicit gamma defs that will be used for sampling across
        # all MM indices. This enforces that every MM follows the gamma
        # described by the preset mean and variability percent (e.g. 50%).
        sr_gamma_def = None
        sa_gamma_def = None
        ar_gamma_def = None
        aa_gamma_def = None
        try:
            if force_gamma and entry is not None:
                # sigma_rad
                srd = entry.get('sigma_rad') if isinstance(entry, dict) else None
                if isinstance(srd, dict) and srd.get('mean') is not None:
                    mu = float(srd.get('mean'))
                    pct = float(var_pct) if var_pct is not None else (float(srd.get('sigma')) / mu if srd.get('sigma') else 0.5)
                    sigma = max(abs(mu * pct), 1e-12)
                    sr_gamma_def = {'dist': 'gamma', 'mean': mu, 'sigma': float(sigma)}
                # sigma_azi
                sad = entry.get('sigma_azi') if isinstance(entry, dict) else None
                if isinstance(sad, dict) and sad.get('mean') is not None:
                    mu = float(sad.get('mean'))
                    pct = float(var_pct) if var_pct is not None else (float(sad.get('sigma')) / mu if sad.get('sigma') else 0.5)
                    sigma = max(abs(mu * pct), 1e-12)
                    sa_gamma_def = {'dist': 'gamma', 'mean': mu, 'sigma': float(sigma)}
                # alpha_rad/alpha_azi: if parsed_alpha provided, treat as percent-of-mean
                ard = entry.get('alpha_rad') if isinstance(entry, dict) else None
                aad = entry.get('alpha_azi') if isinstance(entry, dict) else None
                if parsed_alpha is not None:
                    if isinstance(ard, dict) and ard.get('mean') is not None:
                        mu = float(ard.get('mean'))
                        sigma = max(abs(mu * float(parsed_alpha)), 1e-12)
                        ar_gamma_def = {'dist': 'gamma', 'mean': mu, 'sigma': float(sigma)}
                    if isinstance(aad, dict) and aad.get('mean') is not None:
                        mu = float(aad.get('mean'))
                        sigma = max(abs(mu * float(parsed_alpha)), 1e-12)
                        aa_gamma_def = {'dist': 'gamma', 'mean': mu, 'sigma': float(sigma)}
        except Exception:
            sr_gamma_def = sa_gamma_def = ar_gamma_def = aa_gamma_def = None

        def _sample_with_index(defn, idx):
            """Sample a single value from defn using a per-index deterministic RNG."""
            if defn is None:
                return 0.0
            dist = defn.get('dist')
            # use a unique seed per index for reproducibility and per-MM variability
            rng = _np.random.default_rng(h + int(idx))
            if dist == 'fixed':
                return float(defn.get('value', 0.0))
            if dist == 'gaussian':
                mu = float(defn.get('mean', 0.0))
                sigma = float(defn.get('sigma', 0.0))
                if force_gamma and sigma > 0 and mu > 0:
                    # convert mean/std to gamma shape/scale and sample
                    k = (mu / sigma) ** 2
                    theta = (sigma ** 2) / mu
                    out = float(rng.gamma(shape=k, scale=theta, size=1)[0])
                    if out <= 0:
                        out = 1e-6
                    return out
                if sigma <= 0 or mu <= 0:
                    return float(mu)
                # truncated-at-zero normal via rejection sampling
                out = rng.normal(loc=mu, scale=sigma, size=1)[0]
                attempts = 0
                while out <= 0 and attempts < 100:
                    out = rng.normal(loc=mu, scale=sigma, size=1)[0]
                    attempts += 1
                if out <= 0:
                    out = 1e-6
                return float(out)
            if dist == 'gamma':
                mu = float(defn.get('mean', 0.0))
                sigma = float(defn.get('sigma', 0.0))
                if sigma <= 0 or mu <= 0:
                    return float(mu if mu > 0 else 1e-6)
                k = (mu / sigma) ** 2
                theta = (sigma ** 2) / mu
                out = float(rng.gamma(shape=k, scale=theta, size=1)[0])
                if out <= 0:
                    out = 1e-6
                return out
            if dist == 'uniform':
                lo = float(defn.get('min', 0.0))
                hi = float(defn.get('max', lo))
                return float(rng.uniform(lo, hi, size=1)[0])
            return 0.0

        # If the chosen preset is absent from std_defs but the baseline/template
        # includes explicit per-row placeholder strings (e.g. 'gamma(...)') in
        # tail columns like 'sigma_rad_'/'sigma_azi_', expand those now into
        # numeric per-MM values.
        def _parse_gamma_placeholder(s):
            try:
                if not isinstance(s, str):
                    return None
                import re as _re
                m = _re.search(r'gamma\s*\(\s*([0-9.]+)\s*,\s*([0-9.]+)%', s, flags=_re.IGNORECASE)
                if m:
                    mu = float(m.group(1)); pct = float(m.group(2))
                    sigma = (pct / 100.0) * mu
                    return {'dist': 'gamma', 'mean': mu, 'sigma': sigma}
            except Exception:
                return None
            return None

        # (sr_def/sa_def already determined above; don't overwrite)

        n = int(min(num_mm, mm_df.shape[0]))

        # fallback means from existing canonical columns (if present)
        fallback_sr = None
        fallback_sa = None
        try:
            tmp = sr_vals[sr_vals > 0]
            if tmp.size > 0:
                fallback_sr = float(tmp.mean())
        except Exception:
            fallback_sr = None
        # also try to infer a fallback from numeric entries in the template tail column
        try:
            if fallback_sr is None and template_sr_col and template_sr_col in mm_df.columns:
                tmp2 = pd.to_numeric(mm_df[template_sr_col], errors='coerce')
                tmp2 = tmp2[tmp2 > 0]
                if tmp2.size > 0:
                    fallback_sr = float(tmp2.mean())
        except Exception:
            pass
        try:
            tmp = sa_vals[sa_vals > 0]
            if tmp.size > 0:
                fallback_sa = float(tmp.mean())
        except Exception:
            fallback_sa = None
        try:
            if fallback_sa is None and template_sa_col and template_sa_col in mm_df.columns:
                tmp2 = pd.to_numeric(mm_df[template_sa_col], errors='coerce')
                tmp2 = tmp2[tmp2 > 0]
                if tmp2.size > 0:
                    fallback_sa = float(tmp2.mean())
        except Exception:
            pass

        # If the template contains per-row placeholders (e.g. 'gamma(...)')
        # expand them here into numeric per-MM samples. Otherwise use the
        # preset definitions (vectorized sampling) via _sample.
        eps = 1e-6
        # (debugging records removed)
        if template_placeholders and (template_sr_col or template_sa_col):
            new_sr = _np.zeros(n)
            new_sa = _np.zeros(n)
            for i in range(n):
                # Determine sigma_rad sample
                if template_sr_col is not None and not pd.isna(mm_df.at[i, template_sr_col]):
                    s = mm_df.at[i, template_sr_col]
                    parsed = _parse_gamma_placeholder(s)
                    if isinstance(parsed, dict):
                        new_sr[i] = float(_sample_with_index(parsed, i))
                    else:
                        if force_gamma and sr_gamma_def is not None:
                            new_sr[i] = float(_sample_with_index(sr_gamma_def, i))
                        elif force_gamma and isinstance(sr_def, dict):
                            try:
                                base_mu = float(sr_def.get('mean')) if sr_def.get('mean') is not None else (fallback_sr if fallback_sr is not None else 1e-6)
                                pct = var_pct if var_pct is not None else (sr_def.get('sigma')/sr_def.get('mean') if sr_def.get('mean') and sr_def.get('sigma') else 0.5)
                                sigma_i = max(abs(float(base_mu) * float(pct)), 1e-12)
                                tmp_def = {'dist': 'gamma', 'mean': float(base_mu), 'sigma': float(sigma_i)}
                                new_sr[i] = float(_sample_with_index(tmp_def, i))
                            except Exception:
                                new_sr[i] = float(fallback_sr) if fallback_sr is not None else 0.0
                        else:
                            try:
                                new_sr[i] = float(s)
                            except Exception:
                                if isinstance(sr_def, dict):
                                    new_sr[i] = float(_sample_with_index(sr_def, i))
                                else:
                                    new_sr[i] = float(fallback_sr) if fallback_sr is not None else 0.0

                else:
                    if force_gamma and sr_gamma_def is not None:
                        new_sr[i] = float(_sample_with_index(sr_gamma_def, i))
                    elif force_gamma and isinstance(sr_def, dict):
                        try:
                            base_mu = float(sr_def.get('mean')) if sr_def.get('mean') is not None else (fallback_sr if fallback_sr is not None else 1e-6)
                            pct = var_pct if var_pct is not None else (sr_def.get('sigma')/sr_def.get('mean') if sr_def.get('mean') and sr_def.get('sigma') else 0.5)
                            sigma_i = max(abs(float(base_mu) * float(pct)), 1e-12)
                            tmp_def = {'dist': 'gamma', 'mean': float(base_mu), 'sigma': float(sigma_i)}
                            new_sr[i] = float(_sample_with_index(tmp_def, i))
                        except Exception:
                            new_sr[i] = float(fallback_sr) if fallback_sr is not None else 0.0
                    elif isinstance(sr_def, dict):
                        new_sr[i] = float(_sample_with_index(sr_def, i))
                    else:
                        new_sr[i] = float(fallback_sr) if fallback_sr is not None else 0.0

                # Determine sigma_azi sample
                if template_sa_col is not None and not pd.isna(mm_df.at[i, template_sa_col]):
                    s = mm_df.at[i, template_sa_col]
                    parsed = _parse_gamma_placeholder(s)
                    if isinstance(parsed, dict):
                        new_sa[i] = float(_sample_with_index(parsed, i))
                    else:
                        if force_gamma and sa_gamma_def is not None:
                            new_sa[i] = float(_sample_with_index(sa_gamma_def, i))
                        elif force_gamma and isinstance(sa_def, dict):
                            try:
                                base_mu = float(sa_def.get('mean')) if sa_def.get('mean') is not None else (fallback_sa if fallback_sa is not None else 1e-6)
                                pct = var_pct if var_pct is not None else (sa_def.get('sigma')/sa_def.get('mean') if sa_def.get('mean') and sa_def.get('sigma') else 0.5)
                                sigma_i = max(abs(float(base_mu) * float(pct)), 1e-12)
                                tmp_def = {'dist': 'gamma', 'mean': float(base_mu), 'sigma': float(sigma_i)}
                                new_sa[i] = float(_sample_with_index(tmp_def, i))
                            except Exception:
                                new_sa[i] = float(fallback_sa) if fallback_sa is not None else 0.0
                        else:
                            try:
                                new_sa[i] = float(s)
                            except Exception:
                                if isinstance(sa_def, dict):
                                    new_sa[i] = float(_sample_with_index(sa_def, i))
                                else:
                                    new_sa[i] = float(fallback_sa) if fallback_sa is not None else 0.0
                else:
                    if force_gamma and sa_gamma_def is not None:
                        new_sa[i] = float(_sample_with_index(sa_gamma_def, i))
                    elif force_gamma and isinstance(sa_def, dict):
                        try:
                            base_mu = float(sa_def.get('mean')) if sa_def.get('mean') is not None else (fallback_sa if fallback_sa is not None else 1e-6)
                            pct = var_pct if var_pct is not None else (sa_def.get('sigma')/sa_def.get('mean') if sa_def.get('mean') and sa_def.get('sigma') else 0.5)
                            sigma_i = max(abs(float(base_mu) * float(pct)), 1e-12)
                            tmp_def = {'dist': 'gamma', 'mean': float(base_mu), 'sigma': float(sigma_i)}
                            new_sa[i] = float(_sample_with_index(tmp_def, i))
                        except Exception:
                            new_sa[i] = float(fallback_sa) if fallback_sa is not None else 0.0
                    elif isinstance(sa_def, dict):
                        new_sa[i] = float(_sample_with_index(sa_def, i))
                    else:
                        new_sa[i] = float(fallback_sa) if fallback_sa is not None else 0.0

                # detailed per-index logging removed

            new_sr = _np.where(new_sr <= 0, eps, new_sr)
            new_sa = _np.where(new_sa <= 0, eps, new_sa)
        else:
            new_sr = _np.zeros(n)
            new_sa = _np.zeros(n)
            for i in range(n):
                # sigma_rad: enforce per-index gamma sampling when requested
                if force_gamma and sr_gamma_def is not None:
                    new_sr[i] = float(_sample_with_index(sr_gamma_def, i))
                elif force_gamma and isinstance(sr_def, dict):
                    try:
                        base_mu = float(sr_def.get('mean')) if sr_def.get('mean') is not None else (fallback_sr if fallback_sr is not None else 1e-6)
                        pct = var_pct if var_pct is not None else (sr_def.get('sigma')/sr_def.get('mean') if sr_def.get('mean') and sr_def.get('sigma') else 0.5)
                        sigma_i = max(abs(float(base_mu) * float(pct)), 1e-12)
                        tmp_def = {'dist': 'gamma', 'mean': float(base_mu), 'sigma': float(sigma_i)}
                        new_sr[i] = float(_sample_with_index(tmp_def, i))
                    except Exception:
                        new_sr[i] = float(fallback_sr) if fallback_sr is not None else 0.0
                elif isinstance(sr_def, dict):
                    new_sr[i] = float(_sample_with_index(sr_def, i))
                else:
                    new_sr[i] = float(fallback_sr) if fallback_sr is not None else 0.0

                # sigma_azi: enforce per-index gamma sampling when requested
                if force_gamma and sa_gamma_def is not None:
                    new_sa[i] = float(_sample_with_index(sa_gamma_def, i))
                elif force_gamma and isinstance(sa_def, dict):
                    try:
                        base_mu = float(sa_def.get('mean')) if sa_def.get('mean') is not None else (fallback_sa if fallback_sa is not None else 1e-6)
                        pct = var_pct if var_pct is not None else (sa_def.get('sigma')/sa_def.get('mean') if sa_def.get('mean') and sa_def.get('sigma') else 0.5)
                        sigma_i = max(abs(float(base_mu) * float(pct)), 1e-12)
                        tmp_def = {'dist': 'gamma', 'mean': float(base_mu), 'sigma': float(sigma_i)}
                        new_sa[i] = float(_sample_with_index(tmp_def, i))
                    except Exception:
                        new_sa[i] = float(fallback_sa) if fallback_sa is not None else 0.0
                elif isinstance(sa_def, dict):
                    new_sa[i] = float(_sample_with_index(sa_def, i))
                else:
                    new_sa[i] = float(fallback_sa) if fallback_sa is not None else 0.0

                # detailed per-index logging removed
            new_sr = _np.where(new_sr <= 0, eps, new_sr)
            new_sa = _np.where(new_sa <= 0, eps, new_sa)

        # write back
        for i in range(n):
            mm_df.at[i, sr_col] = float(new_sr[i])
            mm_df.at[i, sa_col] = float(new_sa[i])

        # Sample alphas if definitions exist and alpha columns present
        if alpha_rad_cols and alpha_azi_cols:
            ar_col = alpha_rad_cols[0]
            aa_col = alpha_azi_cols[0]
            ar_def = entry.get('alpha_rad') if entry else None
            aa_def = entry.get('alpha_azi') if entry else None
            # if preset names include an explicit alpha (e.g. '(alpha 10%)'), use
            # a gamma constructed from the preset alpha mean when available;
            # otherwise fall back to a fixed fraction value as before.
            if parsed_alpha is not None:
                if ar_gamma_def is not None:
                    ar_def = ar_gamma_def
                else:
                    ar_def = {'dist': 'fixed', 'value': float(parsed_alpha)}
                if aa_gamma_def is not None:
                    aa_def = aa_gamma_def
                else:
                    aa_def = {'dist': 'fixed', 'value': float(parsed_alpha)}
            # Sample alphas per-index deterministically
            for i in range(n):
                try:
                    if ar_def is None:
                        ar_val = 0.5
                    else:
                        ar_val = float(_sample_with_index(ar_def, i))
                except Exception:
                    ar_val = 0.5
                try:
                    if aa_def is None:
                        aa_val = 0.5
                    else:
                        aa_val = float(_sample_with_index(aa_def, i))
                except Exception:
                    aa_val = 0.5
                ar_val = float(min(max(ar_val, 0.0), 1.0))
                aa_val = float(min(max(aa_val, 0.0), 1.0))
                mm_df.at[i, ar_col] = ar_val
                mm_df.at[i, aa_col] = aa_val

        # persist updated sheet by replacing it in workbook
        from openpyxl import load_workbook
        wb = load_workbook(fp)
        if 'MM_PSF' in wb.sheetnames:
            wb.remove(wb['MM_PSF'])
        from openpyxl.utils.dataframe import dataframe_to_rows
        ws = wb.create_sheet('MM_PSF')
        for r in dataframe_to_rows(mm_df, index=False, header=True):
            ws.append(r)
        # Log sampling details for reproducibility in a dedicated sheet
        try:
            import datetime as _dt
            samp_sheet = 'MM_PSF_SAMPLING'
            if samp_sheet in wb.sheetnames:
                ws2 = wb[samp_sheet]
            else:
                ws2 = wb.create_sheet(samp_sheet)
                ws2.append(['timestamp_utc', 'preset_name', 'seed', 'n', 'sr_mean', 'sr_std', 'sr_shape', 'sr_scale', 'sa_mean', 'sa_std', 'sa_shape', 'sa_scale'])
            sr_mean = float(new_sr.mean()) if hasattr(new_sr, 'mean') else float(sum(new_sr)/len(new_sr))
            sr_std = float(new_sr.std()) if hasattr(new_sr, 'std') else float(0.0)
            sa_mean = float(new_sa.mean()) if hasattr(new_sa, 'mean') else float(sum(new_sa)/len(new_sa))
            sa_std = float(new_sa.std()) if hasattr(new_sa, 'std') else float(0.0)
            # compute gamma params if applicable
            sr_shape = None
            sr_scale = None
            sa_shape = None
            sa_scale = None
            try:
                if isinstance(sr_def, dict) and (sr_def.get('dist') == 'gamma' or force_gamma):
                    mu = float(sr_def.get('mean', 0.0)); sigma = float(sr_def.get('sigma', 0.0))
                    if sigma > 0 and mu > 0:
                        sr_shape = (mu / sigma) ** 2
                        sr_scale = (sigma ** 2) / mu
                if isinstance(sa_def, dict) and (sa_def.get('dist') == 'gamma' or force_gamma):
                    mu = float(sa_def.get('mean', 0.0)); sigma = float(sa_def.get('sigma', 0.0))
                    if sigma > 0 and mu > 0:
                        sa_shape = (mu / sigma) ** 2
                        sa_scale = (sigma ** 2) / mu
            except Exception:
                pass
            ws2.append([_dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'), str(chosen_name), int(h), int(n), sr_mean, sr_std, sr_shape, sr_scale, sa_mean, sa_std, sa_shape, sa_scale])
        except Exception:
            pass
        wb.save(fp)

        # detailed per-index sampling CSV output removed

    def _mask_alpha_for_non_pseudo_voigt(workbook_path: Path, num_mm: int):
        """Set alpha_rad/alpha_azi to '-' for per-MM rows where the distribution
        column does not indicate a pseudo-voigt / voigt distribution.

        Only touches the first `num_mm` data rows and only writes into the
        alpha columns (if present). This keeps alpha cells explicit '-' for
        gaussian/uniform presets.
        """
        try:
            from openpyxl import load_workbook
            fp = Path(workbook_path)
            wb = load_workbook(fp)
            # find MM_PSF sheet (case-insensitive)
            sheet_key = None
            for s in wb.sheetnames:
                if s.lower().replace(' ', '_').startswith('mm_psf') or s.lower() == 'mm_psf':
                    sheet_key = s
                    break
            if sheet_key is None:
                wb.save(fp)
                return
            ws = wb[sheet_key]
            # Map header names to columns
            header = {}
            for c in range(1, ws.max_column + 1):
                hv = ws.cell(row=1, column=c).value
                if hv is None:
                    continue
                header[str(hv).strip().lower()] = c

            # Heuristic: find distribution column (name containing 'distrib' or 'distribution')
            distrib_col = None
            for k, v in header.items():
                if 'distrib' in k or 'distribution' in k:
                    distrib_col = v
                    break
            # find alpha columns by fragments
            ar_col = None
            aa_col = None
            for k, v in header.items():
                if 'alpha_rad' in k or 'alpha rad' in k:
                    ar_col = v
                if 'alpha_azi' in k or 'alpha azi' in k:
                    aa_col = v

            # If no alpha columns found, try tolerant matching on header tokens
            if ar_col is None or aa_col is None:
                for c in range(1, ws.max_column + 1):
                    hv = str(ws.cell(row=1, column=c).value or '').strip().lower()
                    if ar_col is None and 'alpha' in hv and 'rad' in hv:
                        ar_col = c
                    if aa_col is None and 'alpha' in hv and 'azi' in hv:
                        aa_col = c

            # Only operate on rows 2..1+num_mm
            for r in range(2, min(ws.max_row, 1 + int(num_mm)) + 1):
                # determine distribution text for this row
                is_pseudo = False
                if distrib_col is not None:
                    val = ws.cell(row=r, column=distrib_col).value
                    if val is not None:
                        s = str(val).strip().lower()
                        if 'pseudo' in s or 'voigt' in s or 'pv' in s:
                            is_pseudo = True
                # If not pseudo-voigt, set alpha cells to '-'
                if not is_pseudo:
                    if ar_col is not None:
                        try:
                            ws.cell(row=r, column=ar_col, value='-')
                        except Exception:
                            pass
                    if aa_col is not None:
                        try:
                            ws.cell(row=r, column=aa_col, value='-')
                        except Exception:
                            pass
            wb.save(fp)
        except Exception:
            return

    def _enforce_mmpsf_column_bounds(workbook_path: Path, baseline_path: Path, num_mm: int):
        """Ensure MM_PSF per-MM rows only differ in columns B..H.

        For rows 2..(1+num_mm), restore any columns beyond H (index 8) from the
        baseline workbook. Also set alpha_rad (G,7) and alpha_azi (H,8) to '-'
        for rows that are not pseudo-voigt distributions.
        """
        try:
            from openpyxl import load_workbook
            fp = Path(workbook_path)
            bp = Path(baseline_path)
            wb = load_workbook(fp)
            # load baseline evaluated values
            try:
                wb_base = load_workbook(bp, data_only=True)
            except Exception:
                wb_base = None

            sheet_key = None
            for s in wb.sheetnames:
                if s.lower().replace(' ', '_').startswith('mm_psf') or s.lower() == 'mm_psf':
                    sheet_key = s
                    break
            if sheet_key is None:
                wb.save(fp)
                return
            ws = wb[sheet_key]

            base_ws = None
            if wb_base is not None and sheet_key in wb_base.sheetnames:
                base_ws = wb_base[sheet_key]

            # ensure header exists and extend columns up to H (8) if needed
            max_need_col = 8
            if ws.max_column < max_need_col:
                # create header placeholders as needed
                for c in range(ws.max_column + 1, max_need_col + 1):
                    if ws.cell(row=1, column=c).value is None:
                        ws.cell(row=1, column=c, value=f'col{c}')

            # Identify distribution column if present
            distrib_col = None
            for c in range(1, ws.max_column + 1):
                hv = ws.cell(row=1, column=c).value
                if hv is None:
                    continue
                hvs = str(hv).strip().lower()
                if 'distrib' in hvs or 'distribution' in hvs:
                    distrib_col = c
                    break

            # For each per-MM row, enforce alpha masking and restore tail columns
            for r in range(2, min(ws.max_row, 1 + int(num_mm)) + 1):
                # check if this row is pseudo-voigt by distrib_col
                is_pseudo = False
                if distrib_col is not None:
                    v = ws.cell(row=r, column=distrib_col).value
                    if v is not None and isinstance(v, str):
                        s = v.strip().lower()
                        if 'pseudo' in s or 'voigt' in s or 'pv' in s:
                            is_pseudo = True
                # set alpha_rad (G=7) and alpha_azi (H=8)
                try:
                    if not is_pseudo:
                        ws.cell(row=r, column=7, value='-')
                        ws.cell(row=r, column=8, value='-')
                except Exception:
                    pass

                # restore baseline columns beyond H (9..end) if baseline exists
                if base_ws is not None:
                    for c in range(9, ws.max_column + 1):
                        try:
                            base_val = base_ws.cell(row=r, column=c).value if base_ws.max_column >= c and base_ws.max_row >= r else None
                            ws.cell(row=r, column=c, value=base_val)
                        except Exception:
                            continue

            wb.save(fp)
        except Exception:
            return


    def _apply_alignment_preset_to_workbook(workbook_path: Path, specs: dict, num_mm: int, preset_name: str | None):
        """Apply a standard Alignment preset by sampling per-MM values and writing into columns B..E.

        This is a fallback that does not require `generate_data_from_distributions` and
        uses a deterministic RNG seeded from the filename + preset_name for reproducibility.
        """
        try:
            import numpy as _np
            from openpyxl import load_workbook
            fp = Path(workbook_path)
            if specs is None:
                return

            # Build distribution defs for each param label
            if DATA_TYPES and isinstance(DATA_TYPES.get('Alignment', {}).get('params', []), list):
                param_labels = DATA_TYPES.get('Alignment', {}).get('params', [])
            else:
                # fallback labels (match GUI expectations)
                param_labels = [
                    'd_align_rad [µm]', 'd_align_azi [µm]', 'd_align_z [µm]', 'd_align_rotz [arcsec]'
                ]
            defs = {}
            # robust lookup for spec strings (full label, base name, underscore variant)
            def _lookup_spec(sdict, label):
                if not sdict:
                    return None
                if label in sdict:
                    return sdict.get(label)
                base = str(label).split(' ')[0]
                if base in sdict:
                    return sdict.get(base)
                alt = base.replace(' ', '_')
                if alt in sdict:
                    return sdict.get(alt)
                return None

            import re as _re
            # detect column-letter specifications (e.g. 'W' or 'W,X,Y,AB') and prefer direct column reads
            colref_found = False
            # also allow a single cell containing a comma-separated list of letters mapping
            # to the param_labels in order (e.g. 'W,X,Y,AB')
            group_mapping = None
            try:
                for v in (specs or {}).values():
                    if not isinstance(v, str):
                        continue
                    vs = v.strip()
                    if _re.fullmatch(r'[A-Za-z]{1,3}(?:,[A-Za-z]{1,3})+', vs):
                        parts = [p.strip() for p in vs.split(',')]
                        if len(parts) >= len(param_labels):
                            group_mapping = {param_labels[i]: parts[i] for i in range(len(param_labels))}
                            colref_found = True
                            break
            except Exception:
                group_mapping = None
            for p_label in param_labels:
                spec_str = _lookup_spec(specs, p_label)
                if spec_str is None and group_mapping is not None and p_label in group_mapping:
                    spec_str = group_mapping.get(p_label)
                if spec_str is None:
                    defs[p_label] = {'dist': 'fixed', 'value': 0.0}
                    continue
                s = str(spec_str).strip()
                # column-letter pattern (one or more letters optionally comma separated)
                if _re.fullmatch(r'[A-Za-z]{1,3}(?:,[A-Za-z]{1,3})*', s):
                    # store column-letter reference to be read from sheet later
                    defs[p_label] = {'col_ref': s}
                    colref_found = True
                    continue
                try:
                    kind, a, b = _parse_standard_dist_spec(spec_str)
                    if kind == 'fixed':
                        defs[p_label] = {'dist': 'fixed', 'value': float(a)}
                    elif kind == 'gaussian':
                        defs[p_label] = {'dist': 'gaussian', 'mean': float(a), 'sigma': float(b)}
                    else:
                        defs[p_label] = {'dist': 'uniform', 'min': float(min(a, b)), 'max': float(max(a, b))}
                except Exception:
                    defs[p_label] = {'dist': 'fixed', 'value': 0.0}

            # deterministic RNG
            import hashlib
            h = int(hashlib.sha256((str(fp.name) + str(preset_name)).encode('utf-8')).hexdigest()[:8], 16)
            rng = _np.random.default_rng(h)

            def _sample_one(defn, size):
                if defn is None:
                    return _np.zeros(size)
                if defn.get('dist') == 'fixed':
                    return _np.full(size, float(defn.get('value', 0.0)))
                if defn.get('dist') == 'gaussian':
                    mu = float(defn.get('mean', 0.0))
                    sigma = float(defn.get('sigma', 0.0))
                    if sigma <= 0:
                        return _np.full(size, mu)
                    return rng.normal(loc=mu, scale=sigma, size=size)
                if defn.get('dist') == 'uniform':
                    lo = float(defn.get('min', 0.0))
                    hi = float(defn.get('max', lo))
                    return rng.uniform(lo, hi, size=size)
                return _np.zeros(size)

            n = int(num_mm)
            samples = {p: _sample_one(defs.get(p), n) for p in param_labels}

            # If the preset specifies explicit Excel column letters, copy per-MM values
            # directly from those source columns (e.g. W,X,Y,AB) into the target
            # parameter columns in the same sheet. This is deterministic and avoids
            # doing sampling when an explicit mapping is present.
            if colref_found:
                try:
                    from openpyxl.utils import column_index_from_string
                    from openpyxl import load_workbook
                    # load workbook twice: one evaluated (data_only=True) to read numeric values,
                    # and one raw to overwrite cells (preserve formulas elsewhere).
                    wb_val = load_workbook(fp, data_only=True)
                    wb_raw = load_workbook(fp, data_only=False)
                    sheet_key = None
                    for sname in wb_raw.sheetnames:
                        if sname.lower() == 'alignment':
                            sheet_key = sname
                            break
                    if sheet_key is not None and sheet_key in wb_val.sheetnames:
                        ws_val = wb_val[sheet_key]
                        ws_raw = wb_raw[sheet_key]
                        # map target param labels to their workbook columns (header row 1)
                        target_col_map = {}
                        for col in range(1, ws_raw.max_column + 1):
                            hv = ws_raw.cell(row=1, column=col).value
                            if isinstance(hv, str) and hv.strip() in param_labels:
                                target_col_map[hv.strip()] = col
                        # ensure defaults B..G for missing targets, but do not assign beyond G
                        next_col = 2
                        for p_label in param_labels:
                            if p_label not in target_col_map:
                                if next_col <= 7:
                                    target_col_map[p_label] = next_col
                                    next_col += 1
                                else:
                                    # cannot assign beyond column G; leave absent
                                    pass

                        for p_label in param_labels:
                            crefs = defs.get(p_label, {}).get('col_ref')
                            if not crefs:
                                continue
                            # take first letter if comma-separated list
                            letter = str(crefs).split(',')[0].strip()
                            try:
                                src_idx = column_index_from_string(letter)
                            except Exception:
                                continue
                            tgt_idx = target_col_map.get(p_label)
                            if tgt_idx is None:
                                continue
                            for i in range(n):
                                try:
                                    # prefer evaluated numeric value when available
                                    v = ws_val.cell(row=2 + i, column=src_idx).value
                                    if v is None:
                                        # fallback to raw cell if evaluation not available
                                        v = ws_raw.cell(row=2 + i, column=src_idx).value
                                    # force numeric overwrite into the raw workbook, only in B..G
                                    if 2 <= tgt_idx <= 7:
                                        try:
                                            ws_raw.cell(row=2 + i, column=tgt_idx, value=float(v) if v is not None else 0.0)
                                        except Exception:
                                            ws_raw.cell(row=2 + i, column=tgt_idx, value=v)
                                except Exception:
                                    continue
                        wb_raw.save(fp)
                        # we've applied direct numeric copies; no further sampling needed
                        return
                except Exception:
                    pass

            # Use pandas to read/write the Alignment sheet for robust handling of headers/templates
            try:
                df_align = pd.read_excel(fp, sheet_name='Alignment', engine='openpyxl')
            except Exception:
                # build a minimal dataframe if sheet missing
                df_align = pd.DataFrame(columns=['Position #'] + param_labels)
                # ensure at least n rows with Position #
                for ii in range(n):
                    df_align.loc[ii, 'Position #'] = ii + 1

            # Ensure columns exist; if not, append them after Position #
            for p_label in param_labels:
                if p_label not in df_align.columns:
                    df_align[p_label] = 0.0
            # Ensure dataframe has at least `n` rows
            if df_align.shape[0] < n:
                for ii in range(df_align.shape[0], n):
                    df_align.loc[ii] = [None] * len(df_align.columns)
                    if 'Position #' in df_align.columns:
                        df_align.at[ii, 'Position #'] = ii + 1

            # Coerce alignment parameter columns to float to avoid dtype warnings and
            # ensure numeric dtype before assignment.
            for p_label in param_labels:
                if p_label not in df_align.columns:
                    df_align[p_label] = 0.0
                try:
                    df_align[p_label] = pd.to_numeric(df_align[p_label], errors='coerce').astype(float)
                except Exception:
                    # fallback: create a float column of zeros if coercion fails
                    df_align[p_label] = 0.0


            # Write sampled values into the alignment parameter columns for first n rows
            for i in range(n):
                for p_label in param_labels:
                    try:
                        df_align.at[i, p_label] = float(samples.get(p_label)[i])
                    except Exception:
                        df_align.at[i, p_label] = samples.get(p_label)[i]

            # Write numeric per-MM alignment values directly into the workbook using openpyxl
            try:
                from openpyxl import load_workbook
                wb = load_workbook(fp)
                # case-insensitive sheet lookup
                sheet_key = None
                for s in wb.sheetnames:
                    if s.lower() == 'alignment':
                        sheet_key = s
                        break
                if sheet_key is None:
                    # nothing to write
                    wb.save(fp)
                else:
                    ws = wb[sheet_key]
                    # map header names (row 1) to columns
                    col_map = {}
                    for col in range(1, ws.max_column + 1):
                        hv = ws.cell(row=1, column=col).value
                        if isinstance(hv, str):
                            hvn = hv.strip()
                            for p_label in param_labels:
                                if hvn == p_label:
                                    col_map[p_label] = col
                    # default to B..G for missing labels but do not assign beyond G
                    next_col = 2
                    for p_label in param_labels:
                        if p_label not in col_map:
                            if next_col <= 7:
                                col_map[p_label] = next_col
                                next_col += 1
                            else:
                                # cannot map further columns beyond G
                                pass

                    for i in range(n):
                        for p_label in param_labels:
                            col_idx = col_map.get(p_label, None)
                            if col_idx is None or not (2 <= col_idx <= 7):
                                continue
                            v = samples.get(p_label)[i]
                            try:
                                ws.cell(row=2 + i, column=col_idx, value=float(v))
                            except Exception:
                                ws.cell(row=2 + i, column=col_idx, value=v)
                    wb.save(fp)
            except Exception:
                # If openpyxl write fails, fall back to pandas write of the dataframe
                try:
                    write_sheet(fp, 'Alignment', df_align)
                except Exception:
                    pass
        except Exception:
            return

    def _apply_alignment_preset_to_sheets(sheets: dict, specs: dict, num_mm: int, preset_name: str | None):
        # minimal stub for in-memory preset application; real implementation
        # may sample or copy columns from baseline DataFrames if present.
        return

    def _apply_thermal_preset_to_workbook(workbook_path: Path, specs: dict, num_mm: int, preset_name: str | None):
        """Apply a standard Thermal preset by sampling per-MM values and writing into the `Thermal` sheet."""
        try:
            import numpy as _np
            fp = Path(workbook_path)
            if specs is None:
                return

            if DATA_TYPES and isinstance(DATA_TYPES.get('Thermal', {}).get('params', []), list):
                param_labels = DATA_TYPES.get('Thermal', {}).get('params', [])
            else:
                param_labels = ['d_therm_x [µm]', 'd_therm_y [µm]', 'd_therm_z [µm]', 'd_therm_rotz [arcsec]']

            defs = {}
            def _lookup_spec(sdict, label):
                if not sdict:
                    return None
                if label in sdict:
                    return sdict.get(label)
                base = str(label).split(' ')[0]
                if base in sdict:
                    return sdict.get(base)
                alt = base.replace(' ', '_')
                if alt in sdict:
                    return sdict.get(alt)
                return None

            import re as _re
            colref_found = False
            group_mapping = None
            try:
                for v in (specs or {}).values():
                    if not isinstance(v, str):
                        continue
                    vs = v.strip()
                    if _re.fullmatch(r'[A-Za-z]{1,3}(?:,[A-Za-z]{1,3})+', vs):
                        parts = [p.strip() for p in vs.split(',')]
                        if len(parts) >= len(param_labels):
                            group_mapping = {param_labels[i]: parts[i] for i in range(len(param_labels))}
                            colref_found = True
                            break
            except Exception:
                group_mapping = None
            for p_label in param_labels:
                spec_str = _lookup_spec(specs, p_label)
                if spec_str is None:
                    defs[p_label] = {'dist': 'fixed', 'value': 0.0}
                    continue
                s = str(spec_str).strip()
                if (group_mapping is not None) and (not s):
                    s = group_mapping.get(p_label, s)
                if _re.fullmatch(r'[A-Za-z]{1,3}(?:,[A-Za-z]{1,3})*', s):
                    defs[p_label] = {'col_ref': s}
                    colref_found = True
                    continue
                try:
                    kind, a, b = _parse_standard_dist_spec(spec_str)
                    if kind == 'fixed':
                        defs[p_label] = {'dist': 'fixed', 'value': float(a)}
                    elif kind == 'gaussian':
                        defs[p_label] = {'dist': 'gaussian', 'mean': float(a), 'sigma': float(b)}
                    else:
                        defs[p_label] = {'dist': 'uniform', 'min': float(min(a, b)), 'max': float(max(a, b))}
                except Exception:
                    defs[p_label] = {'dist': 'fixed', 'value': 0.0}

            import hashlib
            h = int(hashlib.sha256((str(fp.name) + str(preset_name)).encode('utf-8')).hexdigest()[:8], 16)
            rng = _np.random.default_rng(h)

            def _sample_one(defn, size):
                if defn is None:
                    return _np.zeros(size)
                if defn.get('dist') == 'fixed':
                    return _np.full(size, float(defn.get('value', 0.0)))
                if defn.get('dist') == 'gaussian':
                    mu = float(defn.get('mean', 0.0))
                    sigma = float(defn.get('sigma', 0.0))
                    if sigma <= 0:
                        return _np.full(size, mu)
                    return rng.normal(loc=mu, scale=sigma, size=size)
                if defn.get('dist') == 'uniform':
                    lo = float(defn.get('min', 0.0))
                    hi = float(defn.get('max', lo))
                    return rng.uniform(lo, hi, size=size)
                return _np.zeros(size)

            n = int(num_mm)
            samples = {p: _sample_one(defs.get(p), n) for p in param_labels}

            if colref_found:
                try:
                    from openpyxl.utils import column_index_from_string
                    from openpyxl import load_workbook
                    wb_val = load_workbook(fp, data_only=True)
                    wb_raw = load_workbook(fp, data_only=False)
                    sheet_key = None
                    for sname in wb_raw.sheetnames:
                        if sname.lower() == 'thermal':
                            sheet_key = sname
                            break
                    if sheet_key is not None and sheet_key in wb_val.sheetnames:
                        ws_val = wb_val[sheet_key]
                        ws_raw = wb_raw[sheet_key]
                        target_col_map = {}
                        for col in range(1, ws_raw.max_column + 1):
                            hv = ws_raw.cell(row=1, column=col).value
                            if isinstance(hv, str) and hv.strip() in param_labels:
                                target_col_map[hv.strip()] = col
                        next_col = 2
                        for p_label in param_labels:
                            if p_label not in target_col_map:
                                if next_col <= 7:
                                    target_col_map[p_label] = next_col
                                    next_col += 1
                                else:
                                    pass

                        for p_label in param_labels:
                            crefs = defs.get(p_label, {}).get('col_ref')
                            if not crefs:
                                continue
                            letter = str(crefs).split(',')[0].strip()
                            try:
                                src_idx = column_index_from_string(letter)
                            except Exception:
                                continue
                            tgt_idx = target_col_map.get(p_label)
                            if tgt_idx is None:
                                continue
                            for i in range(n):
                                try:
                                    v = ws_val.cell(row=2 + i, column=src_idx).value
                                    if v is None:
                                        v = ws_raw.cell(row=2 + i, column=src_idx).value
                                    # only write into allowed columns B..G
                                    if 2 <= tgt_idx <= 7:
                                        try:
                                            ws_raw.cell(row=2 + i, column=tgt_idx, value=float(v) if v is not None else 0.0)
                                        except Exception:
                                            ws_raw.cell(row=2 + i, column=tgt_idx, value=v)
                                except Exception:
                                    continue
                        wb_raw.save(fp)
                        return
                except Exception:
                    pass

            try:
                df_sheet = pd.read_excel(fp, sheet_name='Thermal', engine='openpyxl')
            except Exception:
                df_sheet = pd.DataFrame(columns=['Position #'] + param_labels)
                for ii in range(n):
                    df_sheet.loc[ii, 'Position #'] = ii + 1

            for p_label in param_labels:
                if p_label not in df_sheet.columns:
                    df_sheet[p_label] = 0.0
                try:
                    df_sheet[p_label] = pd.to_numeric(df_sheet[p_label], errors='coerce').astype(float)
                except Exception:
                    df_sheet[p_label] = 0.0

            if df_sheet.shape[0] < n:
                for ii in range(df_sheet.shape[0], n):
                    df_sheet.loc[ii] = [None] * len(df_sheet.columns)
                    if 'Position #' in df_sheet.columns:
                        df_sheet.at[ii, 'Position #'] = ii + 1

            for i in range(n):
                for p_label in param_labels:
                    try:
                        df_sheet.at[i, p_label] = float(samples.get(p_label)[i])
                    except Exception:
                        df_sheet.at[i, p_label] = samples.get(p_label)[i]

            try:
                from openpyxl import load_workbook
                wb = load_workbook(fp)
                sheet_key = None
                for s in wb.sheetnames:
                    if s.lower() == 'thermal':
                        sheet_key = s
                        break
                if sheet_key is None:
                    wb.save(fp)
                else:
                    ws = wb[sheet_key]
                    col_map = {}
                    for col in range(1, ws.max_column + 1):
                        hv = ws.cell(row=1, column=col).value
                        if isinstance(hv, str):
                            hvn = hv.strip()
                            for p_label in param_labels:
                                if hvn == p_label:
                                    col_map[p_label] = col
                    next_col = 2
                    for p_label in param_labels:
                        if p_label not in col_map:
                            if next_col <= 7:
                                col_map[p_label] = next_col
                                next_col += 1
                            else:
                                pass

                    for i in range(n):
                        for p_label in param_labels:
                            col_idx = col_map.get(p_label, None)
                            if col_idx is None or not (2 <= col_idx <= 7):
                                continue
                            v = samples.get(p_label)[i]
                            try:
                                ws.cell(row=2 + i, column=col_idx, value=float(v))
                            except Exception:
                                ws.cell(row=2 + i, column=col_idx, value=v)
                    wb.save(fp)
            except Exception:
                try:
                    write_sheet(fp, 'Thermal', df_sheet)
                except Exception:
                    pass
        except Exception:
            return

    def _apply_gravity_preset_to_workbook(workbook_path: Path, specs: dict, num_mm: int, preset_name: str | None):
        """Apply a standard Gravity offload preset by sampling per-MM values and writing into the `Gravity offload` sheet."""
        try:
            import numpy as _np
            fp = Path(workbook_path)
            if specs is None:
                return

            if DATA_TYPES and isinstance(DATA_TYPES.get('Gravity offload', {}).get('params', []), list):
                param_labels = DATA_TYPES.get('Gravity offload', {}).get('params', [])
            else:
                param_labels = ['d_grav_x [µm]', 'd_grav_y [µm]', 'd_grav_z [µm]', 'd_grav_rotz [arcsec]']

            defs = {}
            def _lookup_spec(sdict, label):
                if not sdict:
                    return None
                if label in sdict:
                    return sdict.get(label)
                base = str(label).split(' ')[0]
                if base in sdict:
                    return sdict.get(base)
                alt = base.replace(' ', '_')
                if alt in sdict:
                    return sdict.get(alt)
                return None

            import re as _re
            # detect group mapping (comma-separated column letters) in specs
            colref_found = False
            group_mapping = None
            try:
                for v in (specs or {}).values():
                    if not isinstance(v, str):
                        continue
                    vs = v.strip()
                    if _re.fullmatch(r'[A-Za-z]{1,3}(?:,[A-Za-z]{1,3})+', vs):
                        parts = [p.strip() for p in vs.split(',')]
                        if len(parts) >= len(param_labels):
                            group_mapping = {param_labels[i]: parts[i] for i in range(len(param_labels))}
                            colref_found = True
                            break
            except Exception:
                group_mapping = None

            for p_label in param_labels:
                spec_str = _lookup_spec(specs, p_label)
                if spec_str is None and group_mapping is not None and p_label in group_mapping:
                    spec_str = group_mapping.get(p_label)
                if spec_str is None:
                    defs[p_label] = {'dist': 'fixed', 'value': 0.0}
                    continue
                s = str(spec_str).strip()
                # if spec is a column-letter or comma-separated letters, treat as col_ref
                if _re.fullmatch(r'[A-Za-z]{1,3}(?:,[A-Za-z]{1,3})*', s):
                    defs[p_label] = {'col_ref': s}
                    colref_found = True
                    continue
                try:
                    kind, a, b = _parse_standard_dist_spec(spec_str)
                    if kind == 'fixed':
                        defs[p_label] = {'dist': 'fixed', 'value': float(a)}
                    elif kind == 'gaussian':
                        defs[p_label] = {'dist': 'gaussian', 'mean': float(a), 'sigma': float(b)}
                    else:
                        defs[p_label] = {'dist': 'uniform', 'min': float(min(a, b)), 'max': float(max(a, b))}
                except Exception:
                    defs[p_label] = {'dist': 'fixed', 'value': 0.0}

            import hashlib
            h = int(hashlib.sha256((str(fp.name) + str(preset_name)).encode('utf-8')).hexdigest()[:8], 16)
            rng = _np.random.default_rng(h)

            def _sample_one(defn, size):
                if defn is None:
                    return _np.zeros(size)
                if defn.get('dist') == 'fixed':
                    return _np.full(size, float(defn.get('value', 0.0)))
                if defn.get('dist') == 'gaussian':
                    mu = float(defn.get('mean', 0.0))
                    sigma = float(defn.get('sigma', 0.0))
                    if sigma <= 0:
                        return _np.full(size, mu)
                    return rng.normal(loc=mu, scale=sigma, size=size)
                if defn.get('dist') == 'uniform':
                    lo = float(defn.get('min', 0.0))
                    hi = float(defn.get('max', lo))
                    return rng.uniform(lo, hi, size=size)
                return _np.zeros(size)

            n = int(num_mm)
            samples = {p: _sample_one(defs.get(p), n) for p in param_labels}

            try:
                df_sheet = pd.read_excel(fp, sheet_name='Gravity offload', engine='openpyxl')
            except Exception:
                df_sheet = pd.DataFrame(columns=['Position #'] + param_labels)
                for ii in range(n):
                    df_sheet.loc[ii, 'Position #'] = ii + 1

            for p_label in param_labels:
                if p_label not in df_sheet.columns:
                    df_sheet[p_label] = 0.0
                try:
                    df_sheet[p_label] = pd.to_numeric(df_sheet[p_label], errors='coerce').astype(float)
                except Exception:
                    df_sheet[p_label] = 0.0

            if df_sheet.shape[0] < n:
                for ii in range(df_sheet.shape[0], n):
                    df_sheet.loc[ii] = [None] * len(df_sheet.columns)
                    if 'Position #' in df_sheet.columns:
                        df_sheet.at[ii, 'Position #'] = ii + 1

            for i in range(n):
                for p_label in param_labels:
                    try:
                        df_sheet.at[i, p_label] = float(samples.get(p_label)[i])
                    except Exception:
                        df_sheet.at[i, p_label] = samples.get(p_label)[i]
            # If explicit column references were provided, prefer direct numeric copy
            if colref_found:
                try:
                    from openpyxl.utils import column_index_from_string
                    from openpyxl import load_workbook
                    wb_val = load_workbook(fp, data_only=True)
                    wb_raw = load_workbook(fp, data_only=False)
                    sheet_key = None
                    for sname in wb_raw.sheetnames:
                        if sname.lower() == 'gravity offload':
                            sheet_key = sname
                            break
                    if sheet_key is not None and sheet_key in wb_val.sheetnames:
                        ws_val = wb_val[sheet_key]
                        ws_raw = wb_raw[sheet_key]
                        col_map = {}
                        for col in range(1, ws_raw.max_column + 1):
                            hv = ws_raw.cell(row=1, column=col).value
                            if isinstance(hv, str):
                                hvn = hv.strip()
                                for p_label in param_labels:
                                    if hvn == p_label:
                                        col_map[p_label] = col
                        next_col = 2
                        for p_label in param_labels:
                            if p_label not in col_map:
                                col_map[p_label] = next_col
                                next_col += 1

                        for i in range(n):
                            for p_label in param_labels:
                                crefs = defs.get(p_label, {}).get('col_ref')
                                if not crefs:
                                    continue
                                letter = str(crefs).split(',')[0].strip()
                                try:
                                    src_idx = column_index_from_string(letter)
                                except Exception:
                                    continue
                                tgt_idx = col_map.get(p_label)
                                if tgt_idx is None:
                                    continue
                                try:
                                    v = ws_val.cell(row=2 + i, column=src_idx).value
                                    if v is None:
                                        v = ws_raw.cell(row=2 + i, column=src_idx).value
                                    try:
                                        ws_raw.cell(row=2 + i, column=tgt_idx, value=float(v) if v is not None else 0.0)
                                    except Exception:
                                        ws_raw.cell(row=2 + i, column=tgt_idx, value=v)
                                except Exception:
                                    continue
                        wb_raw.save(fp)
                        return
                except Exception:
                    pass

            try:
                from openpyxl import load_workbook
                wb = load_workbook(fp)
                sheet_key = None
                for s in wb.sheetnames:
                    if s.lower() == 'gravity offload':
                        sheet_key = s
                        break
                if sheet_key is None:
                    wb.save(fp)
                else:
                    ws = wb[sheet_key]
                    col_map = {}
                    for col in range(1, ws.max_column + 1):
                        hv = ws.cell(row=1, column=col).value
                        if isinstance(hv, str):
                            hvn = hv.strip()
                            for p_label in param_labels:
                                if hvn == p_label:
                                    col_map[p_label] = col
                    next_col = 2
                    for p_label in param_labels:
                        if p_label not in col_map:
                            col_map[p_label] = next_col
                            next_col += 1

                    for i in range(n):
                        for p_label in param_labels:
                            col_idx = col_map.get(p_label, None)
                            if col_idx is None:
                                continue
                            v = samples.get(p_label)[i]
                            try:
                                ws.cell(row=2 + i, column=col_idx, value=float(v))
                            except Exception:
                                ws.cell(row=2 + i, column=col_idx, value=v)
                    wb.save(fp)
            except Exception:
                try:
                    write_sheet(fp, 'Gravity offload', df_sheet)
                except Exception:
                    pass
        except Exception:
            return

    # Generate input workbooks for each combo
    input_files = []
    # Allow forcing the timestamp for reproducible filenames (useful for
    # regenerating exactly-named pickles). The CLI flag --force-timestamp
    # provides a UTC timestamp string like '20260125T000945Z'. If not set,
    # fall back to current UTC time.
    if getattr(args, 'force_timestamp', None):
        ts = args.force_timestamp
    else:
        ts = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
    combo_id_map = {}
    for i, combo in enumerate(combos, start=1):
        print(f"Processing combo {i}/{len(combos)}: {combo}")
        parts = [f"{k}={v}" for k, v in combo.items()]
        name_suffix = '_'.join(_sanitize_filename(p) for p in parts)
        out_name = f"{ts}_{i}_{name_suffix}.xlsx"
        # Fast CSV-only path: produce a single MM_PSF CSV per combo and skip
        # expensive workbook copying / openpyxl operations to speed up runs.
        if getattr(args, 'csv_only', False):
            out_name = f"{ts}_{i}_{name_suffix}.csv"
            out_path = input_dir / out_name
            try:
                # If a standard MM_PSF preset is selected, try to match it to a
                # standard preset (supports a few label variants) and build a
                # numeric per-MM DataFrame and write directly to CSV.
                sd = None
                if 'MM_PSF' in combo:
                    mk = _find_std_mm_psf_key(combo.get('MM_PSF'), std_mm_psf)
                    if mk:
                        sd = std_mm_psf[mk]
                    else:
                        # no match; silently proceed
                        pass
                if sd is not None:
                    params = {}
                    params['m_rad [arcsec]'] = ('fixed', 0.0, 0.0)
                    params['m_azi [arcsec]'] = ('fixed', 0.0, 0.0)
                    sr = sd.get('sigma_rad')
                    if sr is None:
                        params['sigma_rad [arcsec]'] = ('fixed', 0.0, 0.0)
                    else:
                        if sr.get('dist') == 'fixed':
                            params['sigma_rad [arcsec]'] = ('fixed', sr.get('value', 0.0), 0.0)
                        elif sr.get('dist') == 'gaussian':
                            params['sigma_rad [arcsec]'] = ('gaussian', sr.get('mean', 0.0), sr.get('sigma', 0.0))
                        else:
                            params['sigma_rad [arcsec]'] = ('uniform', sr.get('min', 0.0), sr.get('max', 0.0))
                    sa = sd.get('sigma_azi')
                    if sa is None:
                        params['sigma_azi [arcsec]'] = ('fixed', 0.0, 0.0)
                    else:
                        if sa.get('dist') == 'fixed':
                            params['sigma_azi [arcsec]'] = ('fixed', sa.get('value', 0.0), 0.0)
                        elif sa.get('dist') == 'gaussian':
                            params['sigma_azi [arcsec]'] = ('gaussian', sa.get('mean', 0.0), sa.get('sigma', 0.0))
                        else:
                            params['sigma_azi [arcsec]'] = ('uniform', sa.get('min', 0.0), sa.get('max', 0.0))
                    ar = sd.get('alpha_rad')
                    if ar is None:
                        params['alpha_rad'] = ('fixed', 0.5, 0.0)
                    else:
                        if ar.get('dist') == 'fixed':
                            params['alpha_rad'] = ('fixed', ar.get('value', 0.5), 0.0)
                        else:
                            params['alpha_rad'] = ('fixed', ar.get('mean', 0.5), 0.0)
                    aa = sd.get('alpha_azi')
                    if aa is None:
                        params['alpha_azi'] = ('fixed', 0.5, 0.0)
                    else:
                        if aa.get('dist') == 'fixed':
                            params['alpha_azi'] = ('fixed', aa.get('value', 0.5), 0.0)
                        else:
                            params['alpha_azi'] = ('fixed', aa.get('mean', 0.5), 0.0)
                    if generate_data_from_distributions is not None:
                        df_gen = generate_data_from_distributions(params, num_mm, DATA_TYPES['MM_PSF'])
                        df_gen.insert(0, 'MM #', mm_list[:len(df_gen)])
                        
                    else:
                        df_gen = None
                        

                    # Deterministic per-MM sampling for sigma/alpha when preset
                    # defines a non-fixed distribution (use hash of timestamp+name+i)
                    try:
                        pass
                        import numpy as _np, hashlib as _hashlib
                        entry = sd
                        # find sigma/alpha cols (df_gen may be None when using fallback)
                        sr_col = None
                        sa_col = None
                        if df_gen is not None:
                            if 'sigma_rad [arcsec]' in df_gen.columns:
                                sr_col = 'sigma_rad [arcsec]'
                            else:
                                sr_col = next((c for c in df_gen.columns if 'sigma_rad' in c.lower()), None)
                            if 'sigma_azi [arcsec]' in df_gen.columns:
                                sa_col = 'sigma_azi [arcsec]'
                            else:
                                sa_col = next((c for c in df_gen.columns if 'sigma_azi' in c.lower()), None)

                        # If the GUI helper is unavailable, construct a numeric df_gen
                        # from the standard preset entry by deterministic sampling.
                        
                        if generate_data_from_distributions is None and isinstance(entry, dict):
                            # build a minimal dataframe matching expected CSV layout
                            import pandas as _pd
                            n = int(num_mm)
                            cols = [
                                'MM #', 'm_rad [arcsec]', 'm_azi [arcsec]',
                                'sigma_rad [arcsec]', 'sigma_azi [arcsec]', 'distribution',
                                'alpha_rad', 'alpha_azi'
                            ]
                            rows = []
                            # seed base for deterministic sampling
                            seed_base = int(_hashlib.sha256((out_name + ':' + str(mk or combo.get('MM_PSF'))).encode('utf-8')).hexdigest()[:16], 16)

                            def _sample(defn, idx, base_seed):
                                if defn is None:
                                    return 0.0
                                dist = defn.get('dist')
                                rng = _np.random.default_rng(base_seed + int(idx))
                                if dist == 'fixed':
                                    return float(defn.get('value', 0.0))
                                if dist == 'gaussian':
                                    mu = float(defn.get('mean', 0.0))
                                    sigma = float(defn.get('sigma', 0.0))
                                    if sigma <= 0 or mu <= 0:
                                        return float(mu)
                                    # truncated normal
                                    val = rng.normal(mu, sigma)
                                    attempts = 0
                                    while val <= 0 and attempts < 100:
                                        val = rng.normal(mu, sigma)
                                        attempts += 1
                                    return float(val if val > 0 else max(mu, 1e-6))
                                if dist == 'gamma':
                                    mu = float(defn.get('mean', 0.0))
                                    sigma = float(defn.get('sigma', 0.0))
                                    if sigma <= 0 or mu <= 0:
                                        return float(mu if mu > 0 else 1e-6)
                                    k = (mu / sigma) ** 2
                                    theta = (sigma ** 2) / mu
                                    return float(rng.gamma(k, theta))
                                if dist == 'uniform':
                                    lo = float(defn.get('min', 0.0))
                                    hi = float(defn.get('max', lo))
                                    return float(rng.uniform(lo, hi))
                                return 0.0

                            for i_idx in range(n):
                                v_sr = _sample(entry.get('sigma_rad'), i_idx, seed_base)
                                v_sa = _sample(entry.get('sigma_azi'), i_idx, seed_base + 0x1000)
                                # alpha values: fixed if provided, else default 0.5
                                ar = entry.get('alpha_rad')
                                aa = entry.get('alpha_azi')
                                if isinstance(ar, dict) and ar.get('dist') == 'fixed':
                                    v_ar = float(ar.get('value', 0.5))
                                elif isinstance(ar, dict) and ar.get('dist') in ('gaussian','gamma','uniform'):
                                    v_ar = _sample(ar, i_idx, seed_base + 0x2000)
                                else:
                                    v_ar = 0.5
                                if isinstance(aa, dict) and aa.get('dist') == 'fixed':
                                    v_aa = float(aa.get('value', 0.5))
                                elif isinstance(aa, dict) and aa.get('dist') in ('gaussian','gamma','uniform'):
                                    v_aa = _sample(aa, i_idx, seed_base + 0x3000)
                                else:
                                    v_aa = 0.5
                                # choose distribution label
                                lname = str(mk or combo.get('MM_PSF') or '').lower()
                                if 'voigt' in lname or 'pseudo' in lname:
                                    dlab = 'pseudo-voigt'
                                elif 'gauss' in lname or 'gaussian' in lname:
                                    dlab = 'gaussian'
                                else:
                                    dlab = 'gaussian'
                                rows.append([i_idx+1, 0.0, 0.0, float(v_sr), float(v_sa), dlab, float(v_ar), float(v_aa)])
                            df_gen = _pd.DataFrame(rows, columns=cols)
                            # debug: report constructed df_gen
                            try:
                                pass
                            except Exception:
                                pass
                            # ensure df_gen has expected column names used below
                            if 'MM #' in df_gen.columns and sr_col is None:
                                sr_col = 'sigma_rad [arcsec]'
                                sa_col = 'sigma_azi [arcsec]'
                        else:
                            # perform sampling row-wise using existing df_gen
                            if sr_col and sa_col:
                                for ridx in range(len(df_gen)):
                                    seed_src = f"{out_name}:{mk or combo.get('MM_PSF')}:{ridx}"
                                    h = int(_hashlib.sha256(seed_src.encode('utf-8')).hexdigest()[:16], 16)
                                    rng = _np.random.RandomState(h % (2**32))
                                    if entry and isinstance(entry.get('sigma_rad'), dict) and entry.get('sigma_rad').get('dist') != 'fixed':
                                        mu = float(entry['sigma_rad'].get('mean', 0.0))
                                        sigma = float(entry['sigma_rad'].get('sigma', max(1e-6, mu*0.1)))
                                        df_gen.at[ridx, sr_col] = float(max(1e-9, rng.normal(mu, sigma)))
                                    if entry and isinstance(entry.get('sigma_azi'), dict) and entry.get('sigma_azi').get('dist') != 'fixed':
                                        mu2 = float(entry['sigma_azi'].get('mean', 0.0))
                                        sigma2 = float(entry['sigma_azi'].get('sigma', max(1e-6, mu2*0.1)))
                                        df_gen.at[ridx, sa_col] = float(max(1e-9, rng.normal(mu2, sigma2)))
                    except Exception as e:
                        pass

                    try:
                        # Build a sheets dict that contains MM_PSF and any baseline sheets
                        sheets_out = {}
                        try:
                            # prefer existing sheets if created earlier
                            if 'sheets' in locals() and isinstance(sheets, dict):
                                sheets_out.update(sheets)
                        except Exception:
                            pass
                        sheets_out['MM_PSF'] = df_gen
                        # If out_path is a CSV, write as multi-sheet CSV; otherwise fallback to single-sheet CSV
                        if str(out_path).lower().endswith('.csv'):
                            try:
                                # Attempt to materialize a temporary workbook, apply
                                # per-combo Alignment/Thermal/Gravity presets using the
                                # workbook helpers, then read the modified sheets back
                                # into sheets_out so CSV contains the per-combo variants.
                                tmp_fp = None
                                try:
                                    import tempfile as _tempfile
                                    # create a temp copy of the baseline workbook we can modify
                                    tf = _tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                                    tf.close()
                                    tmp_fp = Path(tf.name)
                                    shutil.copy2(baseline, tmp_fp)
                                    # Apply A_eff per-row expansion if requested in combo
                                    try:
                                        if 'A_eff' in combo:
                                            import re as _re
                                            aeff_val = str(combo.get('A_eff'))
                                            mrow = _re.match(r'^(.+?)\s*\[row\s*(\d+)\]$', aeff_val, flags=_re.IGNORECASE)
                                            if mrow:
                                                preset_name = mrow.group(1).strip()
                                                target_row = int(mrow.group(2))
                                                try:
                                                    aeff_raw = pd.read_excel(baseline, sheet_name='A_eff', engine='openpyxl', header=0)
                                                except Exception:
                                                    aeff_raw = None
                                                if isinstance(aeff_raw, pd.DataFrame):
                                                    # find preset column in A_eff sheet
                                                    preset_col = None
                                                    for c in aeff_raw.columns:
                                                        try:
                                                            if preset_name.lower() in str(c).lower():
                                                                preset_col = c
                                                                break
                                                        except Exception:
                                                            continue
                                                    # build mm -> row mapping if mm_cfg is available
                                                    mm_to_row = {}
                                                    try:
                                                        if mm_cfg is not None:
                                                            mm_col = None
                                                            row_col = None
                                                            for c in mm_cfg.columns:
                                                                if isinstance(c, str) and 'mm' in c.lower():
                                                                    mm_col = c
                                                                    break
                                                            for c in mm_cfg.columns:
                                                                if isinstance(c, str) and 'row' in c.lower():
                                                                    row_col = c
                                                                    break
                                                            if mm_col and row_col:
                                                                for _, r in mm_cfg.iterrows():
                                                                    try:
                                                                        mmn = int(r.get(mm_col))
                                                                        rown = int(r.get(row_col))
                                                                        mm_to_row[mmn] = rown
                                                                    except Exception:
                                                                        continue
                                                    except Exception:
                                                        mm_to_row = {}
                                                    # construct new A_eff DF
                                                    rows = []
                                                    for _, r in aeff_raw.iterrows():
                                                        try:
                                                            mmn = int(r.get('MM #')) if 'MM #' in aeff_raw.columns else None
                                                        except Exception:
                                                            mmn = None
                                                        if mmn is None:
                                                            continue
                                                        rown = mm_to_row.get(mmn, None)
                                                        if rown == target_row and preset_col is not None:
                                                            try:
                                                                wt = float(r.get(preset_col, 0.0))
                                                            except Exception:
                                                                wt = 0.0
                                                        else:
                                                            try:
                                                                wt = float(r.get('A_eff', 0.0)) if 'A_eff' in aeff_raw.columns else 0.0
                                                            except Exception:
                                                                wt = 0.0
                                                        rows.append({'MM #': int(mmn), 'A_eff': float(wt)})
                                                    new_aeff_df = pd.DataFrame(rows)
                                                    try:
                                                        sheets_tmp = pd.read_excel(tmp_fp, sheet_name=None, engine='openpyxl')
                                                    except Exception:
                                                        sheets_tmp = {}
                                                    sheets_tmp['A_eff'] = new_aeff_df
                                                    try:
                                                        with pd.ExcelWriter(tmp_fp, engine='openpyxl') as writer:
                                                            for sname, sdf in (sheets_tmp or {}).items():
                                                                try:
                                                                    if isinstance(sdf, pd.DataFrame):
                                                                        sdf.to_excel(writer, sheet_name=sname, index=False)
                                                                except Exception:
                                                                    continue
                                                    except Exception:
                                                        pass
                                    except Exception:
                                        pass
                                    # Apply standard presets if present in combo
                                    try:
                                        local_std_alignment = load_standard_alignment_defs(baseline)
                                        if 'Alignment' in combo and local_std_alignment:
                                            try:
                                                import re as _re
                                                target = combo.get('Alignment')
                                                target_norm = _re.sub(r"[^a-z0-9 ]", "", str(target).lower().replace('_', ' ')).strip()
                                                mk_a = None
                                                for k in (local_std_alignment or {}).keys():
                                                    if _re.sub(r"[^a-z0-9 ]", "", str(k).lower().replace('_', ' ')).strip() == target_norm:
                                                        mk_a = k
                                                        break
                                            except Exception:
                                                mk_a = None
                                            if mk_a:
                                                specs = local_std_alignment.get(mk_a)
                                                _apply_alignment_preset_to_workbook(tmp_fp, specs, num_mm, mk_a)
                                    except Exception:
                                        pass
                                    try:
                                        local_std_thermal = load_standard_thermal_defs(baseline)
                                        if 'Thermal' in combo and local_std_thermal:
                                            try:
                                                import re as _re
                                                target = combo.get('Thermal')
                                                target_norm = _re.sub(r"[^a-z0-9 ]", "", str(target).lower().replace('_', ' ')).strip()
                                                mk_t = None
                                                for k in (local_std_thermal or {}).keys():
                                                    if _re.sub(r"[^a-z0-9 ]", "", str(k).lower().replace('_', ' ')).strip() == target_norm:
                                                        mk_t = k
                                                        break
                                            except Exception:
                                                mk_t = None
                                            if mk_t:
                                                specs = local_std_thermal.get(mk_t)
                                                _apply_thermal_preset_to_workbook(tmp_fp, specs, num_mm, mk_t)
                                    except Exception:
                                        pass
                                    try:
                                        local_std_gravity = load_standard_gravity_defs(baseline)
                                        if 'Gravity offload' in combo and local_std_gravity:
                                            try:
                                                import re as _re
                                                target = combo.get('Gravity offload')
                                                target_norm = _re.sub(r"[^a-z0-9 ]", "", str(target).lower().replace('_', ' ')).strip()
                                                mk_g = None
                                                for k in (local_std_gravity or {}).keys():
                                                    if _re.sub(r"[^a-z0-9 ]", "", str(k).lower().replace('_', ' ')).strip() == target_norm:
                                                        mk_g = k
                                                        break
                                            except Exception:
                                                mk_g = None
                                            if mk_g:
                                                specs = local_std_gravity.get(mk_g)
                                                _apply_gravity_preset_to_workbook(tmp_fp, specs, num_mm, mk_g)
                                    except Exception:
                                        pass

                                    # Read back modified sheets where available and merge into sheets_out
                                    try:
                                        for sname in ('A_eff', 'MM configuration', 'Alignment', 'Thermal', 'Gravity offload'):
                                            try:
                                                df_s = pd.read_excel(tmp_fp, sheet_name=sname, engine='openpyxl')
                                                # prefer per-combo modified sheet over existing
                                                sheets_out[sname] = df_s
                                            except Exception:
                                                # leave existing sheets_out entry if present
                                                pass
                                    except Exception:
                                        pass
                                except Exception:
                                    pass
                                finally:
                                    # cleanup temp file if created
                                    try:
                                        if tmp_fp is not None and tmp_fp.exists():
                                            tmp_fp.unlink()
                                    except Exception:
                                        pass

                                # Ensure A_eff and other baseline sheets are present for downstream consumers
                                if isinstance(sheets_out, dict):
                                    if 'A_eff' not in sheets_out:
                                        try:
                                            aeff_full = pd.read_excel(baseline, sheet_name='A_eff', engine='openpyxl')
                                            sheets_out['A_eff'] = aeff_full
                                        except Exception:
                                            pass
                                    for sname in ('MM configuration', 'Alignment', 'Thermal', 'Gravity offload'):
                                        if sname not in sheets_out:
                                            try:
                                                df_s = pd.read_excel(baseline, sheet_name=sname, engine='openpyxl')
                                                sheets_out[sname] = df_s
                                            except Exception:
                                                pass
                                write_multisheet_csv(sheets_out, Path(out_path))
                            except Exception:
                                pass
                        else:
                            # write MM_PSF only into CSV for backwards compatibility
                            df_gen.to_csv(out_path, index=False)
                    except Exception:
                        pass
                else:
                    # Fallback: try to read MM_PSF from baseline and write CSV
                    try:
                        base_df = pd.read_excel(baseline, sheet_name='MM_PSF', engine='openpyxl')
                        # write baseline sheets as a multi-sheet CSV if requested
                        try:
                            sheets_base = pd.read_excel(baseline, sheet_name=None, engine='openpyxl')
                            write_multisheet_csv(sheets_base, Path(out_path))
                        except Exception:
                            base_df.to_csv(out_path, index=False)
                    except Exception:
                        # if all fails, write an empty placeholder
                        pd.DataFrame().to_csv(out_path, index=False)
            except Exception:
                pass
            input_files.append((out_path, combo))
            combo_id_map[out_name] = i
            # cleanup input directory to cap number of stored workbooks
            try:
                _cleanup_input_dir(input_dir, max_files=100)
            except Exception:
                pass
            continue
        if args.no_excel:
            # create pickled sheets dict instead of Excel workbook
            try:
                sheets = pd.read_excel(baseline, sheet_name=None, engine='openpyxl')
            except Exception:
                sheets = {}
            # If MM_PSF requested and we can generate, build df_gen and replace
            # Resolve MM_PSF preset name via fuzzy matching to support label variants
            mk = _find_std_mm_psf_key(combo.get('MM_PSF'), std_mm_psf) if 'MM_PSF' in combo else None
            if mk is not None and generate_data_from_distributions is not None:
                try:
                    sd = std_mm_psf[mk]
                    params = {}
                    params['m_rad [arcsec]'] = ('fixed', 0.0, 0.0)
                    params['m_azi [arcsec]'] = ('fixed', 0.0, 0.0)
                    sr = sd.get('sigma_rad')
                    if sr is None:
                        params['sigma_rad [arcsec]'] = ('fixed', 0.0, 0.0)
                    else:
                        if sr['dist']=='fixed':
                            params['sigma_rad [arcsec]'] = ('fixed', sr['value'], 0.0)
                        elif sr['dist']=='gaussian':
                            params['sigma_rad [arcsec]'] = ('gaussian', sr['mean'], sr['sigma'])
                        else:
                            params['sigma_rad [arcsec]'] = ('uniform', sr.get('min',0.0), sr.get('max',0.0))
                    sa = sd.get('sigma_azi')
                    if sa is None:
                        params['sigma_azi [arcsec]'] = ('fixed', 0.0, 0.0)
                    else:
                        if sa['dist']=='fixed':
                            params['sigma_azi [arcsec]'] = ('fixed', sa['value'], 0.0)
                        elif sa['dist']=='gaussian':
                            params['sigma_azi [arcsec]'] = ('gaussian', sa['mean'], sa['sigma'])
                        else:
                            params['sigma_azi [arcsec]'] = ('uniform', sa.get('min',0.0), sa.get('max',0.0))
                    ar = sd.get('alpha_rad')
                    if ar is None:
                        params['alpha_rad'] = ('fixed', 0.5, 0.0)
                    else:
                        if ar['dist']=='fixed':
                            params['alpha_rad'] = ('fixed', ar['value'], 0.0)
                        else:
                            params['alpha_rad'] = ('fixed', ar.get('mean', 0.5), 0.0)
                    aa = sd.get('alpha_azi')
                    if aa is None:
                        params['alpha_azi'] = ('fixed', 0.5, 0.0)
                    else:
                        if aa['dist']=='fixed':
                            params['alpha_azi'] = ('fixed', aa['value'], 0.0)
                        else:
                            params['alpha_azi'] = ('fixed', aa.get('mean', 0.5), 0.0)
                    df_gen = generate_data_from_distributions(params, num_mm, DATA_TYPES['MM_PSF'])
                    df_gen.insert(0, 'MM #', mm_list[:len(df_gen)])
                    sheets['MM_PSF'] = df_gen
                    # Ensure per-MM numeric sigma columns are populated for in-memory mode
                    try:
                        # deterministic sampling similar to _sample_per_mm_sigmas_and_write
                        import numpy as _np
                        import hashlib as _hashlib
                        entry = std_mm_psf.get(mk)
                        if entry is not None:
                            mm_df = sheets.get('MM_PSF')
                            if isinstance(mm_df, pd.DataFrame):
                                sigma_rad_cols = [c for c in mm_df.columns if isinstance(c, str) and 'sigma_rad' in c.lower()]
                                sigma_azi_cols = [c for c in mm_df.columns if isinstance(c, str) and 'sigma_azi' in c.lower()]
                                if sigma_rad_cols and sigma_azi_cols:
                                    sr_col = sigma_rad_cols[0]
                                    sa_col = sigma_azi_cols[0]
                                    sr_vals = pd.to_numeric(mm_df.loc[:num_mm-1, sr_col], errors='coerce')
                                    sa_vals = pd.to_numeric(mm_df.loc[:num_mm-1, sa_col], errors='coerce')
                                    need_sampling = sr_vals.isnull().any() or sa_vals.isnull().any() or (sr_vals <= 0).any() or (sa_vals <= 0).any()
                                    if need_sampling:
                                        def _sample(defn, size):
                                            if defn is None:
                                                return _np.zeros(size)
                                            if defn.get('dist') == 'fixed':
                                                return _np.full(size, float(defn.get('value', 0.0)))
                                            if defn.get('dist') == 'gaussian':
                                                mu = float(defn.get('mean', 0.0))
                                                sigma = float(defn.get('sigma', 0.0))
                                                if sigma <= 0:
                                                    return _np.full(size, mu)
                                                return rng.normal(loc=mu, scale=sigma, size=size)
                                            if defn.get('dist') == 'uniform':
                                                lo = float(defn.get('min', 0.0))
                                                hi = float(defn.get('max', lo))
                                                return rng.uniform(lo, hi, size=size)
                                            return _np.zeros(size)

                                        # seed on out_name + preset for reproducibility
                                        h = int(_hashlib.sha256((out_name + str(mk)).encode('utf-8')).hexdigest()[:8], 16)
                                        rng = _np.random.default_rng(h)
                                        sr_def = entry.get('sigma_rad')
                                        sa_def = entry.get('sigma_azi')
                                        n = int(min(num_mm, mm_df.shape[0]))
                                        new_sr = _sample(sr_def, n)
                                        new_sa = _sample(sa_def, n)
                                        # fallback: if sampling produced all zeros (rare), use preset means
                                        try:
                                            if hasattr(new_sr, 'max') and float(new_sr.max()) == 0.0:
                                                if isinstance(sr_def, dict) and sr_def.get('dist') == 'gaussian':
                                                    new_sr = _np.full(n, float(sr_def.get('mean', 0.0)))
                                        except Exception:
                                            pass
                                        try:
                                            if hasattr(new_sa, 'max') and float(new_sa.max()) == 0.0:
                                                if isinstance(sa_def, dict) and sa_def.get('dist') == 'gaussian':
                                                    new_sa = _np.full(n, float(sa_def.get('mean', 0.0)))
                                        except Exception:
                                            pass
                                        eps = 1e-6
                                        new_sr = _np.where(new_sr <= 0, eps, new_sr)
                                        new_sa = _np.where(new_sa <= 0, eps, new_sa)
                                        for ii in range(n):
                                            mm_df.at[ii, sr_col] = float(new_sr[ii])
                                            mm_df.at[ii, sa_col] = float(new_sa[ii])
                                            # Ensure canonical column names (with units) are populated
                                            try:
                                                # copy from alternate column names if the unit-bearing ones remain zero
                                                def _copy_if_needed(src_names, tgt_name):
                                                    if tgt_name in mm_df.columns:
                                                        vals = pd.to_numeric(mm_df[tgt_name], errors='coerce')
                                                        if vals.isnull().all() or float(vals.sum()) == 0.0:
                                                            for s in src_names:
                                                                if s in mm_df.columns:
                                                                    alt = pd.to_numeric(mm_df[s], errors='coerce')
                                                                    if not alt.isnull().all() and float(alt.sum()) != 0.0:
                                                                        mm_df[tgt_name] = alt.values
                                                                        break
                                                _copy_if_needed(['sigma_rad_', 'sigma_rad'], 'sigma_rad [arcsec]')
                                                _copy_if_needed(['sigma_azi_', 'sigma_azi'], 'sigma_azi [arcsec]')
                                            except Exception:
                                                pass
                                        # update distribution column if template present
                                        try:
                                            dist_cols = [c for c in mm_df.columns if isinstance(c, str) and 'distrib' in c.lower()]
                                            if dist_cols:
                                                dcol = dist_cols[0]
                                                lname = str(mk).lower() if mk is not None else str(combo.get('MM_PSF')).lower()
                                                if 'voigt' in lname or 'pseudo' in lname:
                                                    mm_df[dcol] = 'pseudo-voigt'
                                                elif 'gauss' in lname or 'gaussian' in lname:
                                                    mm_df[dcol] = 'gaussian'
                                        except Exception:
                                            pass
                                        sheets['MM_PSF'] = mm_df
                    except Exception:
                        pass
                except Exception:
                    pass
            # conservative zeroing of B..E on sheets if requested
            def _zero_df_sheet(df):
                try:
                    # ensure at least 5 columns
                    for c in range(1,5):
                        if c < len(df.columns):
                            df.iloc[:, c] = 0
                except Exception:
                    pass
                return df

            try:
                if 'Alignment' in combo and _is_zero_choice(combo.get('Alignment')) and 'Alignment' in sheets:
                    sheets['Alignment'] = _zero_df_sheet(sheets['Alignment'])
                if 'Gravity offload' in combo and _is_zero_choice(combo.get('Gravity offload')) and 'Gravity offload' in sheets:
                    sheets['Gravity offload'] = _zero_df_sheet(sheets['Gravity offload'])
                if 'Thermal' in combo and _is_zero_choice(combo.get('Thermal')) and 'Thermal' in sheets:
                    sheets['Thermal'] = _zero_df_sheet(sheets['Thermal'])
            except Exception:
                pass

            # Materialize a temporary .xlsx, run the workbook-based MM_PSF expansion
            # & sampling helpers (to ensure canonical sigma columns are numeric),
            # then read back sheets and pickle them for in-memory runs.
            tmp_xlsx = input_dir / out_name
            try:
                # copy baseline workbook to a temporary .xlsx we can operate on
                shutil.copy2(baseline, tmp_xlsx)
                # If MM_PSF preset selected, attempt to preserve template and expand
                try:
                    # try to match preset name and expand/sample using canonical key
                    mk_tmp = _find_std_mm_psf_key(combo.get('MM_PSF'), std_mm_psf) if 'MM_PSF' in combo else None
                    if mk_tmp is not None:
                        try:
                            # attempt the preservative template write + expansion
                            write_mmpsf_preserve_template_and_expand(baseline, tmp_xlsx, df_gen if 'df_gen' in locals() else None, std_mm_psf, mk_tmp)
                        except Exception:
                            pass
                        try:
                            _sample_per_mm_sigmas_and_write(tmp_xlsx, std_mm_psf, mk_tmp, num_mm)
                        except Exception:
                            pass
                except Exception:
                    pass

                # Ensure an A_eff sheet exists with numeric weights for every MM (main.py expects this)
                try:
                    sheets_check = pd.read_excel(tmp_xlsx, sheet_name=None, engine='openpyxl')
                    daf = sheets_check.get('A_eff')
                    if daf is None or not isinstance(daf, pd.DataFrame):
                        rows = [['MM #', 'A_eff']]
                        for mmn in mm_list:
                            rows.append([int(mmn), float(0.0)])
                        sheets_check['A_eff'] = pd.DataFrame(rows)
                    else:
                        try:
                            if 'MM #' in daf.columns and 'A_eff' in daf.columns:
                                rows = [['MM #', 'A_eff']]
                                for _, r in daf.iterrows():
                                    try:
                                        mmn = int(r.get('MM #'))
                                    except Exception:
                                        mmn = None
                                    try:
                                        wt = float(r.get('A_eff', 0.0))
                                    except Exception:
                                        wt = 0.0
                                    rows.append([mmn if mmn is not None else '', wt])
                                sheets_check['A_eff'] = pd.DataFrame(rows)
                        except Exception:
                            pass
                    # finally, prefer the in-memory modified sheets_check (which may
                    # include a synthesized/normalized 'A_eff') when available; fall
                    # back to reading the workbook from disk otherwise.
                    try:
                        if 'sheets_check' in locals() and isinstance(sheets_check, dict):
                            sheets = sheets_check
                        else:
                            sheets = pd.read_excel(tmp_xlsx, sheet_name=None, engine='openpyxl')
                    except Exception:
                        # last resort: try reading workbook, or leave sheets as-is
                        try:
                            sheets = pd.read_excel(tmp_xlsx, sheet_name=None, engine='openpyxl')
                        except Exception:
                            pass
                except Exception:
                    # fallback: try to reuse previously-read sheets dict
                    pass

                # Persist modified sheets back into the temporary workbook (.xlsx)
                try:
                    with pd.ExcelWriter(tmp_xlsx, engine='openpyxl') as writer:
                        for sname, sdf in (sheets or {}).items():
                            try:
                                if isinstance(sdf, pd.DataFrame):
                                    sdf.to_excel(writer, sheet_name=sname, index=False)
                                else:
                                    # skip non-DataFrame sheets
                                    continue
                            except Exception:
                                continue
                    out_path = tmp_xlsx
                except Exception:
                    # fallback: keep the temporary xlsx as out_path
                    out_path = tmp_xlsx
            except Exception:
                # fallback to writing baseline copy if any of the above fails
                out_path = input_dir / out_name
                shutil.copy2(baseline, out_path)
        else:
            out_path = input_dir / out_name
            shutil.copy2(baseline, out_path)

        # If combo selects a standard Alignment preset, apply it immediately to the copied workbook.
        try:
            # Reload standard alignment defs from the baseline to ensure we use the
            # same workbook that was copied into the input folder.
            local_std_alignment = load_standard_alignment_defs(baseline)
            if 'Alignment' in combo and local_std_alignment:
                def _norm_name_quick(x):
                    if x is None:
                        return ''
                    s = str(x).lower()
                    s = s.replace('_', ' ')
                    s = re.sub(r"[^a-z0-9 ]", "", s)
                    s = re.sub(r"\s+", " ", s).strip()
                    return s

                combo_val = combo.get('Alignment')
                target_norm = _norm_name_quick(combo_val)
                match_key = None
                for k in (local_std_alignment or {}).keys():
                    kn = _norm_name_quick(k)
                    if kn == target_norm:
                        match_key = k
                        break
                if match_key:
                    try:
                        specs = local_std_alignment.get(match_key)
                        _apply_alignment_preset_to_workbook(out_path, specs, num_mm, match_key)
                        print(f"Applied standard Alignment preset '{match_key}' to {out_path.name}")
                    except Exception as _e:
                        print(f"Warning: failed to apply Alignment preset '{match_key}': {_e}")
        except Exception:
            pass

        # If combo selects a standard Thermal preset, apply it immediately to the copied workbook.
        try:
            local_std_thermal = load_standard_thermal_defs(baseline)
            if 'Thermal' in combo and local_std_thermal:
                combo_val = combo.get('Thermal')
                target_norm = _norm_name_quick(combo_val)
                match_key = None
                for k in (local_std_thermal or {}).keys():
                    kn = _norm_name_quick(k)
                    if kn == target_norm:
                        match_key = k
                        break
                if match_key:
                    try:
                        specs = local_std_thermal.get(match_key)
                        _apply_thermal_preset_to_workbook(out_path, specs, num_mm, match_key)
                        print(f"Applied standard Thermal preset '{match_key}' to {out_path.name}")
                    except Exception as _e:
                        print(f"Warning: failed to apply Thermal preset '{match_key}': {_e}")
        except Exception:
            pass

        # If combo selects a standard Gravity offload preset, apply it immediately to the copied workbook.
        try:
            local_std_gravity = load_standard_gravity_defs(baseline)
            if 'Gravity offload' in combo and local_std_gravity:
                combo_val = combo.get('Gravity offload')
                target_norm = _norm_name_quick(combo_val)
                match_key = None
                for k in (local_std_gravity or {}).keys():
                    kn = _norm_name_quick(k)
                    if kn == target_norm:
                        match_key = k
                        break
                if match_key:
                    try:
                        specs = local_std_gravity.get(match_key)
                        _apply_gravity_preset_to_workbook(out_path, specs, num_mm, match_key)
                        print(f"Applied standard Gravity preset '{match_key}' to {out_path.name}")
                    except Exception as _e:
                        print(f"Warning: failed to apply Gravity preset '{match_key}': {_e}")
        except Exception:
            pass
        except Exception:
            pass

        # If the Sensitivity spec set Alignment/Gravity offload/Thermal to zero,
        # force the corresponding parameter values in the workbook to 0.
        def _is_zero_choice(val) -> bool:
            try:
                if isinstance(val, (int, float)):
                    return float(val) == 0.0
                s = str(val).strip()
                if s == '0' or s == '0.0':
                    return True
                # allow numeric strings with whitespace
                try:
                    return float(s) == 0.0
                except Exception:
                    return False
            except Exception:
                return False

        def _zero_out_sheet_params(workbook_path: Path, sheet_name: str, param_cores: list[str]):
            """Zero only the input variable columns (B-E) on the given sheet.

            This is intentionally conservative: only touch columns B..G (2..7)
            to avoid altering template columns or other sheets such as `MM_PSF`.
            """
            try:
                from openpyxl import load_workbook
                wb = load_workbook(workbook_path)
                # case-insensitive sheet lookup
                sheet_key = None
                for s in wb.sheetnames:
                    if s.lower() == sheet_name.lower():
                        sheet_key = s
                        break
                if sheet_key is None:
                    wb.save(workbook_path)
                    return
                # Defensive: never touch MM_PSF sheet
                if sheet_key.lower().replace(' ', '_').startswith('mm_psf') or 'mm_psf' in sheet_key.lower():
                    wb.save(workbook_path)
                    return
                ws = wb[sheet_key]

                # Attempt to find header columns that match any of the requested
                # param core names (case-insensitive, tolerant matching). If any
                # are found, zero those columns only. Otherwise fall back to
                # zeroing the conservative default B..E range.
                header_map = {}
                try:
                    for col in range(1, ws.max_column + 1):
                        hv = ws.cell(row=1, column=col).value
                        if hv is None:
                            continue
                        hn = str(hv).strip().lower().replace(' ', '').replace('-', '').replace('.', '').replace('\n','')
                        header_map[hn] = col
                except Exception:
                    header_map = {}

                # normalize requested cores and attempt to match header columns
                cores_norm = [str(x).strip().lower().replace(' ', '').replace('-', '').replace('.', '') for x in (param_cores or [])]
                matched_cols = []
                for hn, col in header_map.items():
                    for pc in cores_norm:
                        if not pc:
                            continue
                        if pc in hn or hn in pc:
                            # Only consider matches that fall within allowed columns B..G (2..7)
                            if 2 <= col <= 7:
                                matched_cols.append(col)
                            break

                if matched_cols:
                    for col_idx in set(matched_cols):
                        for r in range(2, ws.max_row + 1):
                            ws.cell(row=r, column=col_idx, value=0)
                else:
                    # fallback: target columns B..G -> indices 2..7
                    for col_idx in range(2, min(8, ws.max_column + 1)):
                        for r in range(2, ws.max_row + 1):
                            ws.cell(row=r, column=col_idx, value=0)

                wb.save(workbook_path)
            except Exception:
                return

        def _set_specific_zero_columns(workbook_path: Path, sheet_name: str, target_names: list[str], num_rows: int):
            """Set columns matching any of `target_names` to 0 for rows 2..(1+num_rows).

            Matching is tolerant (case-insensitive, ignore spaces/dashes/dots). Only
            writes into columns B..G (2..7) to avoid touching templates.
            """
            try:
                from openpyxl import load_workbook
                wb = load_workbook(workbook_path)
                sheet_key = None
                for s in wb.sheetnames:
                    if s.lower() == sheet_name.lower():
                        sheet_key = s
                        break
                if sheet_key is None:
                    wb.save(workbook_path)
                    return
                ws = wb[sheet_key]

                # build normalized header map
                header_map = {}
                for col in range(1, ws.max_column + 1):
                    hv = ws.cell(row=1, column=col).value
                    if hv is None:
                        continue
                    hn = str(hv).strip().lower()
                    hn = hn.replace(' ', '').replace('-', '').replace('.', '')
                    header_map[hn] = col

                targets_norm = [str(x).strip().lower().replace(' ', '').replace('-', '').replace('.', '') for x in (target_names or [])]
                matched = []
                for hn, col in header_map.items():
                    for t in targets_norm:
                        if not t:
                            continue
                        if t in hn or hn in t:
                            if 2 <= col <= 7:
                                matched.append(col)
                            break

                if not matched:
                    wb.save(workbook_path)
                    return

                for col_idx in set(matched):
                    for r in range(2, min(ws.max_row, 1 + int(num_rows)) + 1):
                        ws.cell(row=r, column=col_idx, value=0)
                wb.save(workbook_path)
            except Exception:
                return

        # apply zeroing logic (constrained to B:E on the sheet)
        try:
            # Enforce explicit zeros for requested parameter pairs, constrained to B..G
            if 'Alignment' in combo and _is_zero_choice(combo.get('Alignment')):
                _set_specific_zero_columns(out_path, 'Alignment', ['d_align_rotazi', 'd_align_rotrad'], num_mm)
            if 'Thermal' in combo and _is_zero_choice(combo.get('Thermal')):
                _set_specific_zero_columns(out_path, 'Thermal', ['d_therm_rotx', 'd_therm_roty'], num_mm)
            if 'Gravity offload' in combo and _is_zero_choice(combo.get('Gravity offload')):
                _set_specific_zero_columns(out_path, 'Gravity offload', ['d_grav_rotx', 'd_grav_roty'], num_mm)
            # keep conservative fallback zeroing for any remaining related inputs
            try:
                if 'Alignment' in combo and _is_zero_choice(combo.get('Alignment')):
                    _zero_out_sheet_params(out_path, 'Alignment', ['d_align_rotazi', 'd_align_rotrad', 'd_align_rad', 'd_align_azi', 'd_align_z', 'd_align_rotz'])
                if 'Gravity offload' in combo and _is_zero_choice(combo.get('Gravity offload')):
                    _zero_out_sheet_params(out_path, 'Gravity offload', ['d_grav_rotx', 'd_grav_rott', 'd_grav_x', 'd_grav_y', 'd_grav_z', 'd_grav_rotz'])
                if 'Thermal' in combo and _is_zero_choice(combo.get('Thermal')):
                    _zero_out_sheet_params(out_path, 'Thermal', ['d_therm_rotx', 'd_therm_roty', 'd_therm_x', 'd_therm_y', 'd_therm_z', 'd_therm_rotz'])
            except Exception:
                pass
        except Exception:
            pass

        # If combo includes MM_PSF and it matches a standard preset, generate MM_PSF sheet
        mk = _find_std_mm_psf_key(combo.get('MM_PSF'), std_mm_psf) if 'MM_PSF' in combo else None
        if mk is not None and generate_data_from_distributions is not None:
                sd = std_mm_psf[combo['MM_PSF']]
                # build params dict for generate_data_from_distributions
                params = {}
                # default m_rad/m_azi fixed 0
                params['m_rad [arcsec]'] = ('fixed', 0.0, 0.0)
                params['m_azi [arcsec]'] = ('fixed', 0.0, 0.0)
                # sigma_rad
                sr = sd.get('sigma_rad')
                if sr is None:
                    params['sigma_rad [arcsec]'] = ('fixed', 0.0, 0.0)
                else:
                    if sr['dist'] == 'fixed':
                        params['sigma_rad [arcsec]'] = ('fixed', sr['value'], 0.0)
                    elif sr['dist'] == 'gaussian':
                        params['sigma_rad [arcsec]'] = ('gaussian', sr['mean'], sr['sigma'])
                    else:
                        params['sigma_rad [arcsec]'] = ('uniform', sr.get('min', 0.0), sr.get('max', 0.0))
                # sigma_azi
                sa = sd.get('sigma_azi')
                if sa is None:
                    params['sigma_azi [arcsec]'] = ('fixed', 0.0, 0.0)
                else:
                    if sa['dist'] == 'fixed':
                        params['sigma_azi [arcsec]'] = ('fixed', sa['value'], 0.0)
                    elif sa['dist'] == 'gaussian':
                        params['sigma_azi [arcsec]'] = ('gaussian', sa['mean'], sa['sigma'])
                    else:
                        params['sigma_azi [arcsec]'] = ('uniform', sa.get('min', 0.0), sa.get('max', 0.0))
                # alpha params
                ar = sd.get('alpha_rad')
                if ar is None:
                    params['alpha_rad'] = ('fixed', 0.5, 0.0)
                else:
                    if ar['dist'] == 'fixed':
                        params['alpha_rad'] = ('fixed', ar['value'], 0.0)
                    elif ar['dist'] == 'gaussian':
                        params['alpha_rad'] = ('gaussian', ar['mean'], ar['sigma'])
                    else:
                        params['alpha_rad'] = ('uniform', ar.get('min', 0.5), ar.get('max', 0.5))
                aa = sd.get('alpha_azi')
                if aa is None:
                    params['alpha_azi'] = ('fixed', 0.5, 0.0)
                else:
                    if aa['dist'] == 'fixed':
                        params['alpha_azi'] = ('fixed', aa['value'], 0.0)
                    elif aa['dist'] == 'gaussian':
                        params['alpha_azi'] = ('gaussian', aa['mean'], aa['sigma'])
                    else:
                        params['alpha_azi'] = ('uniform', aa.get('min', 0.5), aa.get('max', 0.5))

                try:
                    df_gen = generate_data_from_distributions(params, num_mm, DATA_TYPES['MM_PSF'])
                    # add MM # column
                    df_gen.insert(0, 'MM #', mm_list[:len(df_gen)])
                    # write generated MM_PSF preserving baseline template columns if present
                    try:
                        write_mmpsf_preserve_template_and_expand(baseline, out_path, df_gen, std_mm_psf, mk)
                    except Exception:
                        write_sheet(out_path, 'MM_PSF', df_gen)
                    print(f"Wrote generated MM_PSF to {out_path.name}")
                    # Ensure per-MM sigma columns are numeric for this input immediately
                    try:
                        _sample_per_mm_sigmas_and_write(out_path, std_mm_psf, mk, num_mm)
                    except Exception:
                        pass
                    try:
                        _mask_alpha_for_non_pseudo_voigt(out_path, num_mm)
                    except Exception:
                        pass
                    try:
                        _enforce_mmpsf_column_bounds(out_path, baseline, num_mm)
                    except Exception:
                        pass
                except Exception as e:
                    print(f"Failed to generate MM_PSF for combo {combo}: {e}")
        elif mk is not None:
                # generate_data_from_distributions not available: preserve baseline per-MM left table
                try:
                    df_base = pd.read_excel(baseline, sheet_name='MM_PSF', engine='openpyxl')
                    # Determine left-side per-MM column count (heuristic: start_col used in loader)
                    left_cols_count = 10
                    # Build df_gen from left-side columns and first num_mm rows
                    if df_base.shape[1] >= left_cols_count:
                        df_gen = df_base.iloc[:num_mm, :left_cols_count].reset_index(drop=True)
                    else:
                        df_gen = df_base.iloc[:num_mm].reset_index(drop=True)
                    # write preserving template and expanding chosen preset into numeric template cells
                    try:
                        write_mmpsf_preserve_template_and_expand(baseline, out_path, df_gen, std_mm_psf, mk)
                    except Exception:
                        write_sheet(out_path, 'MM_PSF', df_gen)
                    print(f"Wrote MM_PSF (expanded template) to {out_path.name}")
                    # Ensure per-MM sigma columns are numeric for this input immediately
                    try:
                        _sample_per_mm_sigmas_and_write(out_path, std_mm_psf, mk, num_mm)
                    except Exception:
                        pass
                    try:
                        _mask_alpha_for_non_pseudo_voigt(out_path, num_mm)
                    except Exception:
                        pass
                    try:
                        _enforce_mmpsf_column_bounds(out_path, baseline, num_mm)
                    except Exception:
                        pass
                except Exception as e:
                    print(f"Failed to write expanded MM_PSF for combo {combo}: {e}")

        # Handle A_eff per-row expansion: e.g. '1keV[row1]' applies preset '1keV' only to Row #1 positions
        if 'A_eff' in combo:
            aeff_val = str(combo['A_eff']).strip()
            m = re.match(r'^(.+?)\s*\[row\s*(\d+)\]$', aeff_val, flags=re.IGNORECASE)
            if m:
                preset_name = m.group(1).strip()
                target_row = int(m.group(2))
                try:
                    # Read baseline A_eff sheet to find preset column if present
                    aeff_raw = pd.read_excel(baseline, sheet_name='A_eff', engine='openpyxl', header=0)
                except Exception:
                    aeff_raw = None
                # Build mm -> row mapping from mm_cfg (if available). Be flexible with header names.
                mm_to_row = {}
                try:
                    if mm_cfg is not None:
                        # find MM column name
                        mm_col = None
                        row_col = None
                        for c in mm_cfg.columns:
                            if isinstance(c, str) and 'mm' in c.lower():
                                mm_col = c
                                break
                        for c in mm_cfg.columns:
                            if isinstance(c, str) and ('row' in c.lower()):
                                row_col = c
                                break
                        # fallback: if row_col not found, try 3rd column (index 2) as per TestDistribution.xlsx
                        if row_col is None and len(mm_cfg.columns) >= 3:
                            row_col = mm_cfg.columns[2]
                        if mm_col is None and len(mm_cfg.columns) >= 1:
                            mm_col = mm_cfg.columns[0]

                        for _, r in mm_cfg.iterrows():
                            try:
                                mmn = int(r[mm_col])
                            except Exception:
                                continue
                            try:
                                rown = int(r[row_col]) if row_col in r and pd.notna(r[row_col]) else None
                            except Exception:
                                rown = None
                            mm_to_row[mmn] = rown
                except Exception:
                    mm_to_row = {}

                # Determine preset weights mapping MM# -> value
                preset_map = {}
                valnum = None
                if aeff_raw is not None:
                    # find column matching preset_name (case-insensitive substring)
                    cols = [c for c in aeff_raw.columns if isinstance(c, str) and preset_name.lower() in c.lower()]
                    if cols:
                        col = cols[0]
                        # try to detect MM number column in A_eff sheet
                        try:
                            # locate a column that looks like MM # (numeric)
                            key_col = None
                            for kc in aeff_raw.columns[:3]:
                                if pd.to_numeric(aeff_raw[kc], errors='coerce').notna().any():
                                    key_col = kc
                                    break
                            if key_col is None:
                                # fallback to first column
                                key_col = aeff_raw.columns[0]
                            keys = pd.to_numeric(aeff_raw[key_col], errors='coerce')
                            vals = pd.to_numeric(aeff_raw[col], errors='coerce').fillna(0.0)
                            for k, v in zip(keys.tolist(), vals.tolist()):
                                try:
                                    mm_i = int(k)
                                except Exception:
                                    continue
                                preset_map[mm_i] = float(v)
                        except Exception:
                            preset_map = {}

                # Fallback: try parse numeric prefix (e.g., '1keV' -> 1.0)
                if not preset_map:
                    mnum = re.match(r'([0-9]+(?:\.[0-9]+)?)', preset_name)
                    if mnum:
                        valnum = float(mnum.group(1))
                        for mmn in mm_list:
                            preset_map[mmn] = valnum

                # Build A_eff dataframe: set preset value for MMs whose row equals target_row, else 0
                aeff_rows = []
                for mmn in mm_list:
                    rr = mm_to_row.get(mmn)
                    val = 0.0
                    if rr == target_row:
                        # prefer preset_map value for this MM, else use valnum if available
                        if mmn in preset_map:
                            val = preset_map.get(mmn, 0.0)
                        elif valnum is not None:
                            val = valnum
                    aeff_rows.append({'MM #': int(mmn), 'A_eff': float(val)})

                df_aeff = pd.DataFrame(aeff_rows)
                try:
                    write_sheet(out_path, 'A_eff', df_aeff)
                    print(f"Wrote A_eff (row {target_row} preset '{preset_name}') to {out_path.name}")
                except Exception as e:
                    print(f"Failed to write A_eff for combo {combo}: {e}")

        # If Alignment is set to a standard preset, populate per-MM d_align_* columns
        try:
            if 'Alignment' in combo and std_alignment:
                # Normalize preset names (ignore case, spaces/underscores, punctuation)
                def _norm_name(x):
                    if x is None:
                        return ''
                    s = str(x).lower()
                    s = s.replace('_', ' ')
                    s = re.sub(r"[^a-z0-9 ]", "", s)
                    s = re.sub(r"\s+", " ", s).strip()
                    return s

                combo_align_val = combo.get('Alignment')
                match_key = None
                target_norm = _norm_name(combo_align_val)
                for k in (std_alignment or {}).keys():
                    if _norm_name(k) == target_norm:
                        match_key = k
                        break
                if match_key and generate_data_from_distributions is not None:
                    specs = std_alignment.get(match_key)
                    # build params dict for generate_data_from_distributions
                    params_align = {}
                    for p_label in DATA_TYPES.get('Alignment', {}).get('params', []):
                        # lookup spec robustly: try full label, then base name
                        def _lookup_spec(sdict, label):
                            if not sdict:
                                return None
                            if label in sdict:
                                return sdict.get(label)
                            base = str(label).split(' ')[0]
                            if base in sdict:
                                return sdict.get(base)
                            # last resort: try underscore variant
                            alt = base.replace(' ', '_')
                            return sdict.get(alt)

                        spec_str = _lookup_spec(specs, p_label)
                        if spec_str is None:
                            # default to fixed 0
                            params_align[p_label] = ('fixed', 0.0, 0.0)
                            continue
                        try:
                            kind, a, b = _parse_standard_dist_spec(spec_str)
                            if kind == 'fixed':
                                params_align[p_label] = ('fixed', a, 0.0)
                            elif kind == 'gaussian':
                                params_align[p_label] = ('gaussian', a, b)
                            else:
                                params_align[p_label] = ('uniform', a, b)
                        except Exception:
                            params_align[p_label] = ('fixed', 0.0, 0.0)

                    try:
                        df_align_gen = generate_data_from_distributions(params_align, num_mm, DATA_TYPES['Alignment'])
                        # write values into columns B..E (rows 2..)
                        from openpyxl import load_workbook
                        wb = load_workbook(out_path)
                        # case-insensitive sheet match
                        sheet_key = None
                        for s in wb.sheetnames:
                            if s.lower() == 'alignment':
                                sheet_key = s
                                break
                        if sheet_key:
                            ws = wb[sheet_key]
                            for idx in range(min(len(df_align_gen), num_mm)):
                                for j, col_label in enumerate(DATA_TYPES['Alignment']['params'], start=2):
                                    try:
                                        val = float(df_align_gen.iloc[idx, j-2])
                                    except Exception:
                                        val = df_align_gen.iloc[idx, j-2]
                                    ws.cell(row=2+idx, column=j, value=val)
                            wb.save(out_path)
                    except Exception as e:
                        print(f"Failed to apply standard Alignment preset '{match_key}' to {out_path.name}: {e}")
                    # Ensure alignment values are applied even if generate_data_from_distributions failed
                    try:
                        _apply_alignment_preset_to_workbook(out_path, specs, num_mm, match_key)
                    except Exception as ex:
                        print(f"Warning: fallback alignment application failed for preset '{match_key}': {ex}")
        except Exception:
            pass
        # If requested, produce a CSV-only input for main.py instead of Excel/pickle.
        if getattr(args, 'csv_only', False):
            try:
                import tempfile as _tmp_tempfile
                # Prefer sheets['MM_PSF'] if available (created earlier), else df_gen
                df_to_write = None
                try:
                    if 'sheets' in locals() and isinstance(sheets, dict) and 'MM_PSF' in sheets:
                        df_to_write = sheets.get('MM_PSF')
                except Exception:
                    df_to_write = None
                if df_to_write is None:
                    try:
                        df_to_write = df_gen
                    except Exception:
                        df_to_write = None

                if df_to_write is None:
                    # last resort: try to read MM_PSF from baseline
                    try:
                        df_to_write = pd.read_excel(baseline, sheet_name='MM_PSF', engine='openpyxl')
                    except Exception:
                        df_to_write = None

                if df_to_write is None:
                    # no MM_PSF data available; fallback to copying baseline as CSV (best-effort)
                    csv_path = input_dir / out_name.replace('.xlsx', '.csv')
                    try:
                        baseline_df = pd.read_excel(baseline, sheet_name='MM_PSF', engine='openpyxl')
                        baseline_df.to_csv(csv_path, index=False)
                        out_path = csv_path
                    except Exception:
                        # leave out_path as-is if we cannot produce CSV
                        pass
                else:
                    # Create CSV file inside input_dir (ephemeral if non-persistent)
                    csv_name = out_name.replace('.xlsx', '.csv')
                    csv_path = input_dir / csv_name
                    try:
                        df_to_write.to_csv(csv_path, index=False)
                        out_path = csv_path
                    except Exception:
                        pass
            except Exception:
                pass

        input_files.append((out_path, combo))
        # Record generated filename for later lookups
        combo_id_map[out_name] = i
        # cleanup input directory to cap number of stored workbooks
        try:
            _cleanup_input_dir(input_dir, max_files=100)
        except Exception:
            pass

    if getattr(args, 'csv_only', False):
        print(f"Generated {len(input_files)} input CSVs in {input_dir}")
    else:
        print(f"Generated {len(input_files)} input workbooks in {input_dir}")
    if args.generate_only:
        print("generate-only: skipping job execution and post-processing")
        return

    # Enforce standard MM_PSF template and numeric per-MM sigmas for all inputs
    for out_path, combo in input_files:
        try:
            mk = _find_std_mm_psf_key(combo.get('MM_PSF'), std_mm_psf) if 'MM_PSF' in combo else None
            if mk is not None:
                # reconstruct a left-side per-MM df from baseline to preserve layout
                try:
                    df_base = pd.read_excel(baseline, sheet_name='MM_PSF', engine='openpyxl')
                    left_cols_count = 10
                    if df_base.shape[1] >= left_cols_count:
                        df_gen = df_base.iloc[:num_mm, :left_cols_count].reset_index(drop=True)
                    else:
                        df_gen = df_base.iloc[:num_mm].reset_index(drop=True)
                except Exception:
                    df_gen = None

                if df_gen is not None:
                    try:
                        write_mmpsf_preserve_template_and_expand(baseline, out_path, df_gen, std_mm_psf, mk)
                    except Exception:
                        try:
                            write_sheet(out_path, 'MM_PSF', df_gen)
                        except Exception:
                            pass

                # Ensure per-MM sigma columns are numeric (deterministic sampling when needed)
                    try:
                        _sample_per_mm_sigmas_and_write(out_path, std_mm_psf, mk, num_mm)
                    except Exception:
                        pass
                    try:
                        _mask_alpha_for_non_pseudo_voigt(out_path, num_mm)
                    except Exception:
                        pass
                    try:
                        _enforce_mmpsf_column_bounds(out_path, baseline, num_mm)
                    except Exception:
                        pass
        except Exception:
            pass

    # Run all input files through main.py in parallel
    results = []

    def run_one(path_combo):
        path, combo = path_combo
        # Diagnostic: verify MM_PSF sigma columns are numeric for the first num_mm rows.
        try:
            import numpy as _np
            fp = Path(path)
            # Prefer CSV reader for CSV inputs, otherwise read MM_PSF from workbook
            if str(path).lower().endswith('.csv'):
                try:
                    mm_df = pd.read_csv(path)
                except Exception:
                    mm_df = pd.read_excel(fp, sheet_name='MM_PSF', engine='openpyxl')
            else:
                mm_df = pd.read_excel(fp, sheet_name='MM_PSF', engine='openpyxl')
            # find sigma columns
            sigma_rad_cols = [c for c in mm_df.columns if isinstance(c, str) and 'sigma_rad' in c.lower()]
            sigma_azi_cols = [c for c in mm_df.columns if isinstance(c, str) and 'sigma_azi' in c.lower()]
            bad = False
            if sigma_rad_cols and sigma_azi_cols:
                sr = pd.to_numeric(mm_df.loc[:num_mm-1, sigma_rad_cols[0]], errors='coerce')
                sa = pd.to_numeric(mm_df.loc[:num_mm-1, sigma_azi_cols[0]], errors='coerce')
                if sr.isnull().any() or sa.isnull().any() or (sr <= 0).any() or (sa <= 0).any():
                    bad = True
            else:
                bad = True

            mk = _find_std_mm_psf_key(combo.get('MM_PSF'), std_mm_psf) if 'MM_PSF' in combo else None
            if bad and mk is not None:
                try:
                    # read left-side per-MM columns to build df_gen (heuristic 10 columns)
                    left_cols_count = 10
                    df_gen = mm_df.iloc[:num_mm, :left_cols_count].reset_index(drop=True)
                    write_mmpsf_preserve_template_and_expand(baseline, fp, df_gen, std_mm_psf, mk)
                    # reload mm_df after repair
                    mm_df = pd.read_excel(fp, sheet_name='MM_PSF', engine='openpyxl')
                    logging.info("Repaired MM_PSF template in %s", fp.name)
                except Exception as _e:
                    logging.exception("Failed to auto-repair MM_PSF in %s: %s", fp.name, _e)
        except Exception:
            # if anything goes wrong in diagnostic, continue to run and let main.py report errors
            pass
        # If CSV-only input was generated, call main.py with --input-csv so it reads CSVs
        if isinstance(path, (str,)) and str(path).lower().endswith('.csv'):
            cmd = ["python3", str(ROOT / 'main.py'), "-f", str(path), "--input-csv", str(path), "--placement", "elliptical", "--return_metrics_only"]
        else:
            # Default: pass the generated file path to main.py
            cmd = ["python3", str(ROOT / 'main.py'), "-f", str(path), "--placement", "elliptical", "--return_metrics_only"]
        try:
            proc = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, timeout=600)
            if proc.returncode != 0:
                return {'combo': combo, 'file': str(path), 'error': proc.stderr}
            # stdout should be JSON
            out = proc.stdout.strip()
            try:
                metrics = json.loads(out)
            except Exception:
                # try to extract last JSON block
                j = out[out.rfind('{'):]
                metrics = json.loads(j)
            return {'combo': combo, 'file': str(path), 'metrics': metrics}
        except Exception as e:
            return {'combo': combo, 'file': str(path), 'error': str(e)}

    # Force 64 workers as requested
    max_workers = 64
    batch_size = 64
    total_jobs = len(input_files)
    print(f"Running {total_jobs} jobs in batches of {batch_size} (workers={max_workers})...")
    # If user requested a single-file check-and-fix, perform it now and run that single job.
    if args.check_and_fix:
        # allow auto-detection of the latest placed 4.3 file
        target = None
        if args.check_and_fix.lower() == 'auto':
            from glob import glob
            pats = list((SENS_DIR / 'input').glob('*4.3*.xlsx'))[::-1]
            for p in pats:
                try:
                    raw = pd.read_excel(p, sheet_name='MM_PSF', engine='openpyxl', header=None)
                except Exception:
                    continue
                if raw.fillna('').astype(str).apply(lambda col: col.str.contains('10% Variable Sym Gaussian 4.3"', regex=False)).any().any():
                    target = p
                    break
        else:
            target = Path(args.check_and_fix)

        if target is None or not target.exists():
            print(f"check-and-fix target not found: {args.check_and_fix}")
        else:
            print(f"Running check-and-fix on: {target}")
            # locate combo entry for this path if present
            found_combo = None
            for p, c in input_files:
                if Path(p).resolve() == target.resolve():
                    found_combo = (p, c)
                    break
            # if not found, run diagnostic directly using a dummy combo with MM_PSF key guessed
            if not found_combo:
                # attempt to detect MM_PSF preset name in the template area
                try:
                    raw = pd.read_excel(target, sheet_name='MM_PSF', engine='openpyxl', header=None)
                    chosen = None
                    start_col = 10
                    for i in range(raw.shape[0]):
                        if raw.shape[1] > start_col:
                            v = raw.iloc[i, start_col]
                            if isinstance(v, str):
                                for k in std_mm_psf.keys():
                                    if k.lower() in v.lower():
                                        chosen = k
                                        break
                        if chosen:
                            break
                    combo_guess = {'MM_PSF': chosen} if chosen else {}
                    found_combo = (str(target), combo_guess)
                except Exception:
                    found_combo = (str(target), {})

            # run diagnostic+job
            res = run_one(found_combo)
            # print MM_PSF first 20 rows
            try:
                df_mm = pd.read_excel(target, sheet_name='MM_PSF', engine='openpyxl')
                print('MM_PSF (first 20 rows):')
                print(df_mm.head(20).to_string(index=False))
            except Exception:
                pass
            print('Run result:')
            print(json.dumps(res, indent=2, ensure_ascii=False))
            return
    # Process inputs in batches: generate up to `batch_size` workbooks, run them
    # in parallel, append partial results as each job finishes, then optionally
    # delete the batch input files and continue until all jobs are processed.
    def _write_partial_row(res):
        try:
            partial_dir = SENS_DIR / 'results'
            partial_dir.mkdir(parents=True, exist_ok=True)
            partial_path = partial_dir / 'sensitivity_run_partial.csv'
            def _flatten_result(r):
                row = {}
                combo = r.get('combo', {}) or {}
                try:
                    fid = r.get('file')
                    if fid:
                        bn = Path(str(fid)).name
                        row['combo_id'] = combo_id_map.get(bn)
                    else:
                        row['combo_id'] = None
                except Exception:
                    row['combo_id'] = None
                for k, v in combo.items():
                    row[k] = v
                try:
                    row['input_file'] = Path(str(r.get('file'))).name if r.get('file') else None
                except Exception:
                    row['input_file'] = r.get('file')
                if 'metrics' in r and isinstance(r.get('metrics'), dict):
                    m = r.get('metrics')
                    for key in ['hew_origin_arcsec', 'hew_best_arcsec', 'eef90_origin_arcsec', 'eef90_best_arcsec', 'hew_opt_arcsec', 'eef90_opt_arcsec']:
                        row[key] = m.get(key)
                else:
                    row['error'] = r.get('error')
                return row

            single_row = _flatten_result(res)
            import pandas as _pd
            df_row = _pd.DataFrame([single_row])
            if not partial_path.exists():
                df_row.to_csv(partial_path, index=False)
            else:
                df_row.to_csv(partial_path, mode='a', header=False, index=False)
        except Exception:
            pass

    # iterate batches
    for batch_start in range(0, total_jobs, batch_size):
        batch = input_files[batch_start: batch_start + batch_size]
        batch_idx = (batch_start // batch_size) + 1
        print(f"Running batch {batch_idx}: jobs {batch_start+1}-{batch_start+len(batch)} (count={len(batch)})")
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = {ex.submit(run_one, pc): pc for pc in batch}
            for fut in as_completed(futures):
                res = fut.result()
                results.append(res)
                _write_partial_row(res)

        # After batch completes, optionally delete generated input files to free space
        if not args.persist:
            for path_obj, _combo in batch:
                try:
                    p = Path(path_obj)
                    if p.exists():
                        p.unlink()
                except Exception:
                    continue

    # Consolidate results into DataFrame
    rows = []
    for r in results:
        row = {}
        combo = r.get('combo', {})
        # Determine combo id from file name if available
        try:
            fid = r.get('file')
            if fid:
                bn = Path(str(fid)).name
                row['combo_id'] = combo_id_map.get(bn)
            else:
                row['combo_id'] = None
        except Exception:
            row['combo_id'] = None
        for k, v in combo.items():
            row[k] = v
        # Describe MM_PSF preset if present
        try:
            chosen_name = combo.get('MM_PSF') if isinstance(combo, dict) else None
            mk = _find_std_mm_psf_key(chosen_name, std_mm_psf) if chosen_name else None
            use_name = mk if mk is not None else chosen_name
            def _describe_preset(name):
                import re
                if not name:
                    return {'distribution_type': 'custom', 'fixed_variable': '-', 'variability_level': '-', 'symmetry': '-'}
                s = str(name)
                low = s.lower()
                if 'voigt' in low or 'pseudo' in low or 'alpha' in low:
                    dtype = 'pseudo-voigt'
                elif 'gauss' in low:
                    dtype = 'gaussian'
                else:
                    dtype = 'custom'
                # fixed vs variable
                fixed_variable = 'fixed'
                variability_level = '-'
                if 'variable' in low or re.search(r'\d+\s*%', s):
                    fixed_variable = 'variable'
                    m = re.search(r'(\d+)\s*%', s)
                    if m:
                        variability_level = f"{m.group(1)}%"
                else:
                    # check std_mm_psf entry if available
                    try:
                        ent = std_mm_psf.get(s) or next((v for k, v in (std_mm_psf or {}).items() if k.lower() == low), None)
                        if ent:
                            for p in ('sigma_rad', 'sigma_azi', 'alpha_rad', 'alpha_azi'):
                                v = ent.get(p)
                                if isinstance(v, dict) and v.get('dist') != 'fixed':
                                    fixed_variable = 'variable'
                                    break
                    except Exception:
                        pass
                symmetry = '-'
                if 'sym' in low or 'symmetric' in low:
                    symmetry = 'Symmetric'
                if 'asym' in low or 'assym' in low or 'asymmetric' in low:
                    symmetry = 'Asymmetric'
                return {'distribution_type': dtype, 'fixed_variable': fixed_variable, 'variability_level': variability_level, 'symmetry': symmetry}
            desc = _describe_preset(chosen)
            row['distribution_type'] = desc.get('distribution_type')
            row['distribution_fixed_variable'] = desc.get('fixed_variable')
            row['distribution_variability_level'] = desc.get('variability_level')
            row['distribution_symmetry'] = desc.get('symmetry')
        except Exception:
            row['distribution_type'] = '-'
            row['distribution_fixed_variable'] = '-'
            row['distribution_variability_level'] = '-'
            row['distribution_symmetry'] = '-'
        # Normalize A_eff and extract expanded row if present
        try:
            aeff_val = combo.get('A_eff') if isinstance(combo, dict) else None
            if aeff_val is not None:
                aeff_s = str(aeff_val)
                m_a = re.match(r'^(.+?)\s*\[row\s*(\d+)\]$', aeff_s, flags=re.IGNORECASE)
                if m_a:
                    row['A_eff'] = m_a.group(1).strip()
                    row['A_eff_row'] = int(m_a.group(2))
                else:
                    row['A_eff'] = aeff_s
                    row['A_eff_row'] = '-'
            else:
                row['A_eff'] = '-'
                row['A_eff_row'] = '-'
        except Exception:
            row['A_eff'] = combo.get('A_eff', '-')
            row['A_eff_row'] = '-'
        try:
            fid = r.get('file')
            if fid:
                row['input_file'] = Path(str(fid)).name
            else:
                row['input_file'] = None
        except Exception:
            row['input_file'] = r.get('file')
        if 'metrics' in r:
            m = r['metrics']
            # pick hew_opt_arcsec, eef90_opt_arcsec, hew_best_arcsec, eef90_best_arcsec
            for key in ['hew_origin_arcsec', 'hew_best_arcsec', 'eef90_origin_arcsec', 'eef90_best_arcsec', 'hew_opt_arcsec', 'eef90_opt_arcsec']:
                row[key] = m.get(key)
        else:
            row['error'] = r.get('error')
        rows.append(row)

    out_df = pd.DataFrame(rows)
    # Ensure combo_id is numeric and sort results by combo_id so the first column
    # corresponds to the combo ordering used during generation.
    try:
        out_df['combo_id'] = pd.to_numeric(out_df.get('combo_id'), errors='coerce')
        out_df = out_df.sort_values(by='combo_id', na_position='last')
    except Exception:
        pass
    out_path = SENS_DIR / 'results' / 'sensitivity_run_results.xlsx'
    # Reorder columns to requested layout when possible, keep any extras after
    desired_order = [
        'combo_id', 'A_eff', 'A_eff_row', 'MM_PSF',
        'distribution_type', 'distribution_symmetry', 'distribution_fixed_variable', 'distribution_variability_level',
        'Alignment', 'Thermal', 'Gravity offload',
        'hew_origin_arcsec', 'hew_best_arcsec', 'hew_opt_arcsec',
        'eef90_origin_arcsec', 'eef90_best_arcsec', 'eef90_opt_arcsec',
        'input_file'
    ]
    cols_present = [c for c in desired_order if c in out_df.columns]
    remaining = [c for c in out_df.columns if c not in cols_present]
    out_df = out_df[cols_present + remaining]
    out_df.to_excel(out_path, index=False)
    print(f"Wrote results to: {out_path}")
    # Do not remove input workbooks; keep partial results for inspection


if __name__ == '__main__':
    main()
