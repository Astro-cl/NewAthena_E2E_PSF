from pathlib import Path
import sys
import hashlib
import pandas as pd


def load_standard_mm_psf_defs(path: Path) -> dict:
    import re
    try:
        df = pd.read_excel(path, sheet_name='MM_PSF', header=None, engine='openpyxl')
    except Exception:
        return {}
    std = {}
    start_row = 0
    start_col = 10
    if df.shape[0] <= start_row or df.shape[1] <= start_col:
        return {}
    row_idx = start_row + 1
    while row_idx < df.shape[0]:
        name = df.iloc[row_idx, start_col]
        if pd.isna(name) or str(name).strip() == '':
            break
        key_name = str(name).strip()
        dist_def = {'name': key_name}
        for i, param in enumerate(['sigma_rad', 'sigma_azi', 'alpha_rad', 'alpha_azi']):
            cell_value = None
            if start_col + 1 + i < df.shape[1]:
                cell_value = df.iloc[row_idx, start_col + 1 + i]
            if pd.isna(cell_value) or cell_value == '':
                dist_def[param] = None
                continue
            cell_str = str(cell_value).strip()
            # simple numeric parse or gaussian(...) pattern
            try:
                m = re.match(r'^(?:gaussian|normal)\s*\(\s*([0-9.+-Ee]+)\s*,\s*([0-9.+-Ee%*()\/]+)\s*\)', cell_str, re.IGNORECASE)
                if m:
                    mean = float(m.group(1))
                    # handle percent like '10%*1.82635' -> sigma = 0.1*1.82635 or if percent on itself -> percent*mean
                    rhs = m.group(2)
                    pct = None
                    if '%' in rhs:
                        pm = re.match(r'([0-9.]+)\%\s*\*\s*([0-9.+-Ee]+)', rhs)
                        if pm:
                            pct = float(pm.group(1)) / 100.0
                            sigma = float(pm.group(2)) * pct
                        else:
                            # fallback: percent of mean
                            pm2 = re.match(r'([0-9.]+)\%', rhs)
                            if pm2:
                                pct = float(pm2.group(1)) / 100.0
                                sigma = mean * pct
                            else:
                                sigma = float(rhs)
                    else:
                        sigma = float(rhs)
                    dist_def[param] = {'dist': 'gaussian', 'mean': mean, 'sigma': abs(float(sigma))}
                else:
                    # numeric
                    val = float(cell_str)
                    dist_def[param] = {'dist': 'fixed', 'value': val}
            except Exception:
                dist_def[param] = None
        std[key_name] = dist_def
        row_idx += 1
    return std


def sample_and_write(fp: Path, baseline: Path):
    import numpy as _np
    fp = Path(fp)
    std_defs = load_standard_mm_psf_defs(baseline)
    # try to detect chosen preset in template area
    raw = pd.read_excel(fp, sheet_name='MM_PSF', header=None, engine='openpyxl')
    chosen = None
    sc = 10
    for i in range(raw.shape[0]):
        for j in range(sc, raw.shape[1]):
            v = raw.iloc[i, j]
            if isinstance(v, str):
                for k in std_defs.keys():
                    if k.lower() in str(v).lower():
                        chosen = k
                        break
            if chosen:
                break
        if chosen:
            break
    if not chosen:
        print('Could not determine chosen preset from template; aborting')
        return False

    mm_df = pd.read_excel(fp, sheet_name='MM_PSF', engine='openpyxl')
    sigma_rad_cols = [c for c in mm_df.columns if isinstance(c, str) and 'sigma_rad' in c.lower()]
    sigma_azi_cols = [c for c in mm_df.columns if isinstance(c, str) and 'sigma_azi' in c.lower()]
    if not sigma_rad_cols or not sigma_azi_cols:
        print('sigma columns not found; aborting')
        return False
    sr_col = sigma_rad_cols[0]
    sa_col = sigma_azi_cols[0]

    entry = std_defs.get(chosen)
    if not entry:
        print('preset not found in standard defs; aborting')
        return False

    sr_def = entry.get('sigma_rad')
    sa_def = entry.get('sigma_azi')

    # check existing values
    sr_vals = pd.to_numeric(mm_df[sr_col].iloc[:], errors='coerce')
    sa_vals = pd.to_numeric(mm_df[sa_col].iloc[:], errors='coerce')
    force_sample = False
    try:
        if isinstance(sr_def, dict) and sr_def.get('dist') != 'fixed':
            force_sample = True
        if isinstance(sa_def, dict) and sa_def.get('dist') != 'fixed':
            force_sample = True
    except Exception:
        force_sample = False

    need_sampling = sr_vals.isnull().any() or sa_vals.isnull().any() or (sr_vals <= 0).any() or (sa_vals <= 0).any() or force_sample

    # deterministic seed
    h = int(hashlib.sha256((fp.name + str(chosen)).encode('utf-8')).hexdigest()[:8], 16)
    rng = _np.random.default_rng(h)

    def _sample(defn, size):
        if defn is None:
            return _np.zeros(size)
        if defn.get('dist') == 'fixed':
            return _np.full(size, float(defn.get('value', 0.0)))
        if defn.get('dist') == 'gaussian':
            mu = float(defn.get('mean', 0.0))
            sigma = float(defn.get('sigma', 0.0))
            if sigma <= 0:
                return _np.full(size, mu)
            return rng.normal(loc=mu, scale=sigma, size=size)
        if defn.get('dist') == 'uniform':
            lo = float(defn.get('min', 0.0))
            hi = float(defn.get('max', lo))
            return rng.uniform(lo, hi, size=size)
        return _np.zeros(size)

    n = int(min(mm_df.shape[0], len(mm_df)))
    # If the preset defines fixed values, always overwrite per-MM sigmas with that fixed value
    if isinstance(sr_def, dict) and sr_def.get('dist') == 'fixed':
        new_sr = _np.full(n, float(sr_def.get('value', 0.0)))
    else:
        new_sr = _sample(sr_def, n)
    if isinstance(sa_def, dict) and sa_def.get('dist') == 'fixed':
        new_sa = _np.full(n, float(sa_def.get('value', 0.0)))
    else:
        new_sa = _sample(sa_def, n)
    eps = 1e-6
    new_sr = _np.where(new_sr <= 0, eps, new_sr)
    new_sa = _np.where(new_sa <= 0, eps, new_sa)

    for i in range(n):
        mm_df.at[i, sr_col] = float(new_sr[i])
        mm_df.at[i, sa_col] = float(new_sa[i])

    # write back
    from openpyxl import load_workbook
    wb = load_workbook(fp)
    if 'MM_PSF' in wb.sheetnames:
        wb.remove(wb['MM_PSF'])
    from openpyxl.utils.dataframe import dataframe_to_rows
    ws = wb.create_sheet('MM_PSF')
    for r in dataframe_to_rows(mm_df, index=False, header=True):
        ws.append(r)
    # log
    samp_sheet = 'MM_PSF_SAMPLING'
    import datetime as _dt
    if samp_sheet in wb.sheetnames:
        ws2 = wb[samp_sheet]
    else:
        ws2 = wb.create_sheet(samp_sheet)
        ws2.append(['timestamp_utc', 'preset_name', 'seed', 'n', 'sr_mean', 'sr_std', 'sa_mean', 'sa_std'])
    ws2.append([_dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'), str(chosen), int(h), int(n), float(new_sr.mean()), float(new_sr.std()), float(new_sa.mean()), float(new_sa.std())])
    wb.save(fp)
    return True


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('usage: fix_place_set_sigmas.py <placed-workbook-path> [baseline-path]')
        raise SystemExit(1)
    target = Path(sys.argv[1])
    baseline = Path(sys.argv[2]) if len(sys.argv) > 2 else (Path(__file__).resolve().parents[1] / 'Distributions' / 'TestDistribution.xlsx')
    if not target.exists():
        print('Target not found:', target)
        raise SystemExit(1)
    if not baseline.exists():
        print('Baseline not found, continuing without std defs:', baseline)
    ok = sample_and_write(target, baseline)
    print('Fix applied:', ok)
