#!/usr/bin/env python3
from pathlib import Path
import sys
import pandas as pd
import numpy as np
import hashlib
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
baseline = ROOT / 'Distributions' / 'Test_Distribution.xlsx'

if len(sys.argv) > 1:
    target = Path(sys.argv[1])
else:
    target = Path('sensitivity/input/20260123T232615Z_3_A_eff1_keV_MM_PSF10_Variable_Sym_Gaussian_4.3_Alignment0.0_Gravity_offload0.0_Thermal0.0_placed.xlsx')

if not target.exists():
    print('Target missing:', target)
    sys.exit(1)

import sys
sys.path.insert(0, str(ROOT))
try:
    from sensitivity.sensitivity_run import load_standard_mm_psf_defs
    std_defs = load_standard_mm_psf_defs(baseline)
except Exception:
    # fallback: read template area directly
    std_defs = {}
    try:
        df = pd.read_excel(baseline, sheet_name='MM_PSF', header=None, engine='openpyxl')
        start_col = 10
        if df.shape[0] > 1 and df.shape[1] > start_col:
            row_idx = 1
            while row_idx < df.shape[0]:
                name = df.iloc[row_idx, start_col]
                if pd.isna(name) or str(name).strip() == '':
                    break
                key = str(name).strip()
                std_defs[key] = {'name': key}
                # parse next two columns into raw strings if present
                try:
                    std_defs[key]['sigma_rad'] = {'dist': 'raw', 'value': df.iloc[row_idx, start_col+1]}
                    std_defs[key]['sigma_azi'] = {'dist': 'raw', 'value': df.iloc[row_idx, start_col+2]}
                except Exception:
                    pass
                row_idx += 1
    except Exception:
        std_defs = {}

# detect chosen preset name from template area or filename
raw = pd.read_excel(target, sheet_name='MM_PSF', header=None, engine='openpyxl')
start_col = 10
chosen = None
# try filename normalized
import re
def _norm(s):
    if not s:
        return ''
    return re.sub(r'[^0-9a-z]', '', str(s).lower())
fname = _norm(target.name)
for k in std_defs.keys():
    if _norm(k) in fname:
        chosen = k
        break
if not chosen:
    for i in range(raw.shape[0]):
        if raw.shape[1] > start_col:
            v = raw.iloc[i, start_col]
            if isinstance(v, str):
                for k in std_defs.keys():
                    if k.lower() in v.lower():
                        chosen = k
                        break
        if chosen:
            break

if not chosen:
    print('Could not determine preset for', target)
    sys.exit(1)

entry = std_defs.get(chosen)
if not entry:
    print('No std def for', chosen)
    sys.exit(1)

# load mm df
mm = pd.read_excel(target, sheet_name='MM_PSF', engine='openpyxl')
# find per-mm sigma columns
sigma_rad_cols = [c for c in mm.columns if isinstance(c, str) and 'sigma_rad' in c.lower()]
sigma_azi_cols = [c for c in mm.columns if isinstance(c, str) and 'sigma_azi' in c.lower()]
if not sigma_rad_cols or not sigma_azi_cols:
    print('sigma columns missing, cols=', mm.columns.tolist())
    sys.exit(1)
sr_col = sigma_rad_cols[0]
sa_col = sigma_azi_cols[0]

# determine num_mm from baseline
try:
    mm_cfg = pd.read_excel(baseline, sheet_name='MM configuration', engine='openpyxl')
    if 'MM #' in mm_cfg.columns:
        num_mm = int(mm_cfg['MM #'].dropna().shape[0])
    else:
        num_mm = int(mm_cfg.shape[0])
except Exception:
    num_mm = min(50, mm.shape[0])

n = min(num_mm, mm.shape[0])

# sampling defs
sr_def = entry.get('sigma_rad')
sa_def = entry.get('sigma_azi')

# deterministic seed
h = int(hashlib.sha256((target.name + str(chosen)).encode('utf-8')).hexdigest()[:8], 16)
rng = np.random.default_rng(h)

def _sample(defn, size):
    if defn is None:
        return np.zeros(size)
    if defn.get('dist') == 'fixed':
        return np.full(size, float(defn.get('value', 0.0)))
    if defn.get('dist') == 'gaussian':
        mu = float(defn.get('mean', 0.0))
        sigma = float(defn.get('sigma', 0.0))
        if sigma <= 0:
            return np.full(size, mu)
        return rng.normal(loc=mu, scale=sigma, size=size)
    if defn.get('dist') == 'uniform':
        lo = float(defn.get('min', 0.0))
        hi = float(defn.get('max', lo))
        return rng.uniform(lo, hi, size=size)
    return np.zeros(size)

new_sr = _sample(sr_def, n)
new_sa = _sample(sa_def, n)
new_sr = np.where(new_sr <= 0, 1e-6, new_sr)
new_sa = np.where(new_sa <= 0, 1e-6, new_sa)

for i in range(n):
    mm.at[i, sr_col] = float(new_sr[i])
    mm.at[i, sa_col] = float(new_sa[i])

# write back replacing sheet
wb = load_workbook(target)
if 'MM_PSF' in wb.sheetnames:
    wb.remove(wb['MM_PSF'])
from openpyxl.utils.dataframe import dataframe_to_rows
ws = wb.create_sheet('MM_PSF')
for r in dataframe_to_rows(mm, index=False, header=True):
    ws.append(r)
# write sampling log
samp='MM_PSF_SAMPLING'
if samp in wb.sheetnames:
    ws2=wb[samp]
else:
    ws2=wb.create_sheet(samp)
    ws2.append(['timestamp_utc','preset_name','seed','n','sr_mean','sr_std','sa_mean','sa_std'])
ws2.append([__import__('datetime').datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'), chosen, int(h), int(n), float(new_sr.mean()), float(new_sr.std()), float(new_sa.mean()), float(new_sa.std())])
wb.save(target)
print('Wrote sampling log and updated', target)
