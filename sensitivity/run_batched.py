#!/usr/bin/env python3
"""
Batched sensitivity run with proper preset application.
"""

import os
import re
import shutil
import subprocess
import json
from pathlib import Path
from datetime import datetime
from itertools import product
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import numpy as np

ROOT = Path(__file__).resolve().parent.parent
SENS_DIR = ROOT / 'sensitivity'
INPUT_DIR = SENS_DIR / 'input'
RESULTS_DIR = SENS_DIR / 'results'
PARTIAL_CSV = RESULTS_DIR / 'sensitivity_run_partial.csv'
FINAL_XLSX = RESULTS_DIR / 'sensitivity_results.xlsx'

BATCH_SIZE = 64
NUM_WORKERS = 64

def get_timestamp():
    return datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')

def cleanup_input_dir():
    if INPUT_DIR.exists():
        for f in INPUT_DIR.iterdir():
            if f.is_file():
                try:
                    f.unlink()
                except Exception:
                    pass

def sanitize_filename(s):
    s = str(s)
    s = s.replace('"', '').replace("'", '')
    s = re.sub(r'\s+', '_', s)
    s = re.sub(r'[^A-Za-z0-9._-]', '', s)
    return s

def norm_name(s):
    if s is None:
        return ''
    return re.sub(r'[^a-z0-9]', '', str(s).lower().replace('_', ' ').replace('-', ''))

# ===== PRESET LOADING FUNCTIONS =====

def _safe_eval_numeric_expr(expr):
    import ast
    if expr is None:
        raise ValueError('Empty expression')
    s = str(expr).strip().replace('\r', '').replace('\n', ' ')
    s = s.strip()
    
    def _eval(node):
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        if isinstance(node, ast.Num):
            return float(node.n)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            v = _eval(node.operand)
            return +v if isinstance(node.op, ast.UAdd) else -v
        if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
            a = _eval(node.left)
            b = _eval(node.right)
            if isinstance(node.op, ast.Add):
                return a + b
            if isinstance(node.op, ast.Sub):
                return a - b
            if isinstance(node.op, ast.Mult):
                return a * b
            if isinstance(node.op, ast.Div):
                return a / b
        raise ValueError(f'Unsupported expression')
    
    node = ast.parse(s, mode='eval')
    return float(_eval(node.body))

def _parse_dist_spec(spec):
    if spec is None or (isinstance(spec, float) and pd.isna(spec)):
        raise ValueError('Empty spec')
    s = str(spec).strip()
    if not s:
        raise ValueError('Empty spec')
    
    compact = s.replace(' ', '')
    if re.fullmatch(r'[-+]?\d+(?:\.\d+)?', compact):
        return ('fixed', float(compact), 0.0)
    
    m = re.match(r'^\s*(gaussian|normal|uniform)\s*\(\s*(.+)\s*\)\s*$', s, re.IGNORECASE)
    if not m:
        raise ValueError(f'Unsupported distribution spec: {spec!r}')
    
    kind = m.group(1).lower()
    inner = m.group(2)
    parts = [p.strip() for p in inner.split(',') if p.strip()]
    if len(parts) < 1:
        raise ValueError(f'Expected parameters in spec: {spec!r}')
    
    left = parts[0]
    right = parts[1] if len(parts) > 1 else ''
    
    try:
        left_norm = re.sub(r'([+-]?\d+(?:\.\d+)?)\s*%', r'(\1/100)', left)
        a = _safe_eval_numeric_expr(left_norm)
    except Exception:
        a = _safe_eval_numeric_expr(left)
    
    try:
        rstr = re.sub(r'([+-]?\d+(?:\.\d+)?)\s*%', r'(\1/100)', right.strip())
        b = _safe_eval_numeric_expr(rstr)
    except Exception:
        raise ValueError(f'Unsupported numeric expression: {right!r}')
    
    if kind in ('gaussian', 'normal'):
        return ('gaussian', a, abs(b))
    if kind == 'uniform':
        return ('uniform', a, b)
    
    raise ValueError(f'Unsupported distribution spec: {spec!r}')

def load_standard_mm_psf_defs(path):
    try:
        df = pd.read_excel(path, sheet_name='MM_PSF', header=None, engine='openpyxl')
    except Exception:
        return {}
    
    std = {}
    start_row = 0
    start_col = 10  # Column K
    
    if df.shape[0] <= start_row or df.shape[1] <= start_col:
        return {}
    
    row_idx = start_row + 1
    
    while row_idx < df.shape[0]:
        name = df.iloc[row_idx, start_col]
        if pd.isna(name) or not str(name).strip():
            break
        
        key_name = str(name).strip()
        dist_def = {'name': key_name}
        
        for i, param in enumerate(['sigma_rad', 'sigma_azi', 'alpha_rad', 'alpha_azi']):
            cell_value = None
            if start_col + 1 + i < df.shape[1]:
                cell_value = df.iloc[row_idx, start_col + 1 + i]
            
            if pd.isna(cell_value) or not str(cell_value).strip():
                dist_def[param] = None
                continue
            
            cell_str = str(cell_value).strip()
            try:
                res = _parse_dist_spec(cell_str)
                kind = res[0]
                
                if kind == 'fixed':
                    dist_def[param] = {'dist': 'fixed', 'value': float(res[1])}
                elif kind == 'gaussian':
                    dist_def[param] = {'dist': 'gaussian', 'mean': float(res[1]), 'sigma': float(res[2])}
                elif kind == 'uniform':
                    dist_def[param] = {'dist': 'uniform', 'min': float(res[1]), 'max': float(res[2])}
                else:
                    dist_def[param] = None
            except Exception:
                dist_def[param] = None
        
        std[key_name] = dist_def
        row_idx += 1
    
    return std

def load_standard_alignment_defs(path):
    try:
        df = pd.read_excel(path, sheet_name='Alignment', header=None, engine='openpyxl')
    except Exception:
        return {}
    
    std = {}
    start_row = 0
    # Try to auto-detect name column and first variable column to be robust
    name_col = None
    first_var_col = None

    # detect first variable column by looking for header cells with 'd_' prefix
    for c in range(0, df.shape[1]):
        try:
            h = df.iloc[start_row, c]
            if isinstance(h, str) and 'd_' in h.lower():
                first_var_col = c
                break
        except Exception:
            continue

    # detect name column as the column (left of first_var_col) with mostly text values
    def _detect_name_col(df, start_row, maxcols=10):
        limit = min(df.shape[1], maxcols if maxcols else df.shape[1])
        for c in range(0, limit):
            count_strings = 0
            count_total = 0
            for r in range(start_row + 1, min(df.shape[0], start_row + 11)):
                try:
                    val = df.iloc[r, c]
                except Exception:
                    continue
                if pd.isna(val):
                    continue
                count_total += 1
                if isinstance(val, str) and val.strip() and not re.fullmatch(r'[-+]?\d+(?:\.\d+)?', val.strip()):
                    count_strings += 1
            if count_total > 0 and count_strings >= max(1, count_total // 2):
                return c
        return None

    if first_var_col is None:
        first_var_col = 7

    name_col = _detect_name_col(df, start_row, maxcols=df.shape[1])
    if name_col is None:
        name_col = 6

    if df.shape[0] <= start_row or df.shape[1] <= first_var_col:
        return {}

    headers = {}
    for c in range(first_var_col, min(df.shape[1], first_var_col + 16)):
        h = df.iloc[start_row, c]
        if pd.isna(h) or not str(h).strip():
            continue
        headers[c] = str(h).strip()
    
    row_idx = start_row + 1
    while row_idx < df.shape[0]:
        preset_name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
        if pd.isna(preset_name) or not str(preset_name).strip():
            break
        
        preset_name = str(preset_name).strip()
        preset_specs = {}
        
        for c, h in headers.items():
            if c >= df.shape[1]:
                continue
            raw_spec = df.iloc[row_idx, c]
            if pd.isna(raw_spec) or not str(raw_spec).strip():
                continue
            var = str(h).strip()
            if var.endswith('_'):
                var = var[:-1]
            preset_specs[var] = str(raw_spec).strip()
        
        if preset_specs:
            std[preset_name] = preset_specs
        
        row_idx += 1
    
    return std

def load_standard_thermal_defs(path):
    try:
        df = pd.read_excel(path, sheet_name='Thermal', header=None, engine='openpyxl')
    except Exception:
        return {}
    
    std = {}
    start_row = 0

    # auto-detect first variable column and name column similar to alignment
    first_var_col = None
    for c in range(0, df.shape[1]):
        try:
            h = df.iloc[start_row, c]
            if isinstance(h, str) and 'd_' in h.lower():
                first_var_col = c
                break
        except Exception:
            continue

    def _detect_name_col(df, start_row, maxcols=10):
        limit = min(df.shape[1], maxcols if maxcols else df.shape[1])
        for c in range(0, limit):
            count_strings = 0
            count_total = 0
            for r in range(start_row + 1, min(df.shape[0], start_row + 11)):
                try:
                    val = df.iloc[r, c]
                except Exception:
                    continue
                if pd.isna(val):
                    continue
                count_total += 1
                if isinstance(val, str) and val.strip() and not re.fullmatch(r'[-+]?\d+(?:\.\d+)?', val.strip()):
                    count_strings += 1
            if count_total > 0 and count_strings >= max(1, count_total // 2):
                return c
        return None

    if first_var_col is None:
        first_var_col = 7

    name_col = _detect_name_col(df, start_row, maxcols=df.shape[1])
    if name_col is None:
        name_col = 6

    if df.shape[0] <= start_row or df.shape[1] <= first_var_col:
        return {}

    headers = {}
    for c in range(first_var_col, min(df.shape[1], first_var_col + 16)):
        h = df.iloc[start_row, c]
        if pd.isna(h) or not str(h).strip():
            continue
        headers[c] = str(h).strip()
    
    row_idx = start_row + 1
    while row_idx < df.shape[0]:
        preset_name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
        if pd.isna(preset_name) or not str(preset_name).strip():
            break
        
        preset_name = str(preset_name).strip()
        preset_specs = {}
        
        for c, h in headers.items():
            if c >= df.shape[1]:
                continue
            raw_spec = df.iloc[row_idx, c]
            if pd.isna(raw_spec) or not str(raw_spec).strip():
                continue
            var = str(h).strip()
            if var.endswith('_'):
                var = var[:-1]
            preset_specs[var] = str(raw_spec).strip()
        
        if preset_specs:
            std[preset_name] = preset_specs
        
        row_idx += 1
    
    return std

def load_standard_gravity_defs(path):
    try:
        df = pd.read_excel(path, sheet_name='Gravity offload', header=None, engine='openpyxl')
    except Exception:
        return {}
    
    std = {}
    start_row = 0

    # auto-detect first variable column and name column similar to alignment/thermal
    first_var_col = None
    for c in range(0, df.shape[1]):
        try:
            h = df.iloc[start_row, c]
            if isinstance(h, str) and 'd_' in h.lower():
                first_var_col = c
                break
        except Exception:
            continue

    def _detect_name_col(df, start_row, maxcols=10):
        limit = min(df.shape[1], maxcols if maxcols else df.shape[1])
        for c in range(0, limit):
            count_strings = 0
            count_total = 0
            for r in range(start_row + 1, min(df.shape[0], start_row + 11)):
                try:
                    val = df.iloc[r, c]
                except Exception:
                    continue
                if pd.isna(val):
                    continue
                count_total += 1
                if isinstance(val, str) and val.strip() and not re.fullmatch(r'[-+]?\d+(?:\.\d+)?', val.strip()):
                    count_strings += 1
            if count_total > 0 and count_strings >= max(1, count_total // 2):
                return c
        return None

    if first_var_col is None:
        first_var_col = 7

    name_col = _detect_name_col(df, start_row, maxcols=df.shape[1])
    if name_col is None:
        name_col = 6

    if df.shape[0] <= start_row or df.shape[1] <= first_var_col:
        return {}

    headers = {}
    for c in range(first_var_col, min(df.shape[1], first_var_col + 16)):
        h = df.iloc[start_row, c]
        if pd.isna(h) or not str(h).strip():
            continue
        headers[c] = str(h).strip()
    
    row_idx = start_row + 1
    while row_idx < df.shape[0]:
        preset_name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
        if pd.isna(preset_name) or not str(preset_name).strip():
            break
        
        preset_name = str(preset_name).strip()
        preset_specs = {}
        
        for c, h in headers.items():
            if c >= df.shape[1]:
                continue
            raw_spec = df.iloc[row_idx, c]
            if pd.isna(raw_spec) or not str(raw_spec).strip():
                continue
            var = str(h).strip()
            if var.endswith('_'):
                var = var[:-1]
            preset_specs[var] = str(raw_spec).strip()
        
        if preset_specs:
            std[preset_name] = preset_specs
        
        row_idx += 1
    
    return std

def find_mm_psf_key(name, std_map):
    if not name or not std_map:
        return None
    
    target = norm_name(name)
    
    if name in std_map:
        return name
    
    for k in std_map.keys():
        if k.lower() == name.lower():
            return k
    
    for k in std_map.keys():
        nk = norm_name(k)
        if nk and (nk in target or target in nk):
            return k
    
    return None

# ===== PRESET APPLICATION FUNCTIONS =====

def apply_mm_psf_preset(workbook_path, specs, num_mm):
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(workbook_path)
        if 'MM_PSF' not in wb.sheetnames:
            wb.save(workbook_path)
            return
        
        ws = wb['MM_PSF']
        
        # Find header row
        header_row = 1
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, min(15, ws.max_column + 1))]
            if any('sigma' in v for v in row_vals):
                header_row = r
                break
        
        # Find columns
        sigma_rad_col = None
        sigma_azi_col = None
        alpha_rad_col = None
        alpha_azi_col = None
        
        for c in range(1, min(20, ws.max_column + 1)):
            h = str(ws.cell(row=header_row, column=c).value).lower()
            if 'sigma_rad' in h and sigma_rad_col is None:
                sigma_rad_col = c
            if 'sigma_azi' in h and sigma_azi_col is None:
                sigma_azi_col = c
            if 'alpha_rad' in h and alpha_rad_col is None:
                alpha_rad_col = c
            if 'alpha_azi' in h and alpha_azi_col is None:
                alpha_azi_col = c
        
        # Parse specs
        def get_val(param, default=0.0):
            s = specs.get(param)
            if s is None:
                return ('fixed', default, 0)
            if isinstance(s, dict):
                if s.get('dist') == 'fixed':
                    return ('fixed', s.get('value', 0), 0)
                if s.get('dist') == 'gaussian':
                    return ('gaussian', s.get('mean', 0), s.get('sigma', 0))
                if s.get('dist') == 'uniform':
                    return ('uniform', s.get('min', 0), s.get('max', 0))
            return ('fixed', default, 0)
        
        sr = get_val('sigma_rad')
        sa = get_val('sigma_azi')
        ar = get_val('alpha_rad', 0.5)
        aa = get_val('alpha_azi', 0.5)
        
        # Generate per-MM values
        rng = np.random.default_rng(hash(str(workbook_path)) % (2**32))
        
        def sample(spec, idx):
            kind, a, b = spec
            if kind == 'fixed':
                return a
            elif kind == 'gaussian':
                return rng.normal(a, b) if b > 0 else a
            elif kind == 'uniform':
                return rng.uniform(a, b)
            return a
        
        # Write values
        for i in range(num_mm):
            row = header_row + 1 + i
            if row > ws.max_row:
                break
            if sigma_rad_col:
                ws.cell(row=row, column=sigma_rad_col, value=sample(sr, i))
            if sigma_azi_col:
                ws.cell(row=row, column=sigma_azi_col, value=sample(sa, i))
            if alpha_rad_col:
                ws.cell(row=row, column=alpha_rad_col, value=sample(ar, i))
            if alpha_azi_col:
                ws.cell(row=row, column=alpha_azi_col, value=sample(aa, i))
        
        wb.save(workbook_path)
    except Exception as e:
        print(f"  Warning: MM_PSF preset failed: {e}")

def apply_alignment_preset(workbook_path, specs, num_mm):
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(workbook_path)
        if 'Alignment' not in wb.sheetnames:
            wb.save(workbook_path)
            return
        
        ws = wb['Alignment']
        
        # Find header row
        header_row = 1
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, min(15, ws.max_column + 1))]
            if any('d_align' in v for v in row_vals):
                header_row = r
                break
        
        # Find columns
        cols = {}
        for c in range(1, min(15, ws.max_column + 1)):
            h = str(ws.cell(row=header_row, column=c).value).lower()
            for param in ['d_align_rad', 'd_align_azi', 'd_align_z', 'd_align_rotz']:
                if param in h and param not in cols:
                    cols[param] = c
        
        # Parse specs
        def get_val(key):
            s = specs.get(key)
            if s is None:
                return ('fixed', 0, 0)
            try:
                return _parse_dist_spec(s)
            except:
                return ('fixed', 0, 0)
        
        specs_parsed = {k: get_val(k) for k in cols.keys()}
        
        # Generate and write values
        rng = np.random.default_rng(hash(str(workbook_path)) % (2**32))
        
        def sample(spec, idx):
            kind, a, b = spec
            if kind == 'fixed':
                return a
            elif kind == 'gaussian':
                return rng.normal(a, b) if b > 0 else a
            elif kind == 'uniform':
                return rng.uniform(a, b)
            return a
        
        for i in range(num_mm):
            row = header_row + 1 + i
            if row > ws.max_row:
                break
            for param, col in cols.items():
                ws.cell(row=row, column=col, value=sample(specs_parsed.get(param, ('fixed', 0, 0)), i))
        
        wb.save(workbook_path)
    except Exception as e:
        print(f"  Warning: Alignment preset failed: {e}")

def apply_thermal_preset(workbook_path, specs, num_mm):
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(workbook_path)
        if 'Thermal' not in wb.sheetnames:
            wb.save(workbook_path)
            return
        
        ws = wb['Thermal']
        
        header_row = 1
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, min(15, ws.max_column + 1))]
            if any('d_therm' in v for v in row_vals):
                header_row = r
                break
        
        cols = {}
        for c in range(1, min(15, ws.max_column + 1)):
            h = str(ws.cell(row=header_row, column=c).value).lower()
            for param in ['d_therm_x', 'd_therm_y', 'd_therm_z', 'd_therm_rotz']:
                if param in h and param not in cols:
                    cols[param] = c
        
        def get_val(key):
            s = specs.get(key)
            if s is None:
                return ('fixed', 0, 0)
            try:
                return _parse_dist_spec(s)
            except:
                return ('fixed', 0, 0)
        
        specs_parsed = {k: get_val(k) for k in cols.keys()}
        
        rng = np.random.default_rng(hash(str(workbook_path)) % (2**32))
        
        def sample(spec, idx):
            kind, a, b = spec
            if kind == 'fixed':
                return a
            elif kind == 'gaussian':
                return rng.normal(a, b) if b > 0 else a
            elif kind == 'uniform':
                return rng.uniform(a, b)
            return a
        
        for i in range(num_mm):
            row = header_row + 1 + i
            if row > ws.max_row:
                break
            for param, col in cols.items():
                ws.cell(row=row, column=col, value=sample(specs_parsed.get(param, ('fixed', 0, 0)), i))
        
        wb.save(workbook_path)
    except Exception as e:
        print(f"  Warning: Thermal preset failed: {e}")

def apply_gravity_preset(workbook_path, specs, num_mm):
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(workbook_path)
        if 'Gravity offload' not in wb.sheetnames:
            wb.save(workbook_path)
            return
        
        ws = wb['Gravity offload']
        
        header_row = 1
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, min(15, ws.max_column + 1))]
            if any('d_grav' in v for v in row_vals):
                header_row = r
                break
        
        cols = {}
        for c in range(1, min(15, ws.max_column + 1)):
            h = str(ws.cell(row=header_row, column=c).value).lower()
            for param in ['d_grav_x', 'd_grav_y', 'd_grav_z', 'd_grav_rotz']:
                if param in h and param not in cols:
                    cols[param] = c
        
        def get_val(key):
            s = specs.get(key)
            if s is None:
                return ('fixed', 0, 0)
            try:
                return _parse_dist_spec(s)
            except:
                return ('fixed', 0, 0)
        
        specs_parsed = {k: get_val(k) for k in cols.keys()}
        
        rng = np.random.default_rng(hash(str(workbook_path)) % (2**32))
        
        def sample(spec, idx):
            kind, a, b = spec
            if kind == 'fixed':
                return a
            elif kind == 'gaussian':
                return rng.normal(a, b) if b > 0 else a
            elif kind == 'uniform':
                return rng.uniform(a, b)
            return a
        
        for i in range(num_mm):
            row = header_row + 1 + i
            if row > ws.max_row:
                break
            for param, col in cols.items():
                ws.cell(row=row, column=col, value=sample(specs_parsed.get(param, ('fixed', 0, 0)), i))
        
        wb.save(workbook_path)
    except Exception as e:
        print(f"  Warning: Gravity preset failed: {e}")

def apply_aeff_preset(workbook_path, aeff_value, num_mm):
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(workbook_path)
        
        m = re.match(r'^(.+?)\s*\[row\s*(\d+)\]$', str(aeff_value), flags=re.IGNORECASE)
        
        if m:
            preset_name = m.group(1).strip()
            target_row = int(m.group(2))
            
            # Load MM config
            try:
                mm_cfg = pd.read_excel(workbook_path, sheet_name='MM configuration', engine='openpyxl')
                mm_to_row = {}
                for _, r in mm_cfg.iterrows():
                    try:
                        mmn = int(r.get('MM #'))
                        rown = int(r.get('Row #'))
                        mm_to_row[mmn] = rown
                    except:
                        pass
            except:
                mm_to_row = {}
            
            # Create A_eff sheet
            rows = []
            for mmn in range(1, num_mm + 1):
                rown = mm_to_row.get(mmn)
                weight = 1.0 if rown == target_row else 0.0
                rows.append({'MM #': mmn, 'A_eff': weight})
            
            df_aeff = pd.DataFrame(rows)
            
            if 'A_eff' in wb.sheetnames:
                del wb['A_eff']
            ws = wb.create_sheet('A_eff')
            ws.append(['MM #', 'A_eff'])
            for _, row in df_aeff.iterrows():
                ws.append([int(row['MM #']), float(row['A_eff'])])
        else:
            # Simple numeric value
            try:
                val = float(aeff_value)
                if 'A_eff' in wb.sheetnames:
                    del wb['A_eff']
                ws = wb.create_sheet('A_eff')
                ws.append(['MM #', 'A_eff'])
                for i in range(1, num_mm + 1):
                    ws.append([i, val])
            except:
                pass
        
        wb.save(workbook_path)
    except Exception as e:
        print(f"  Warning: A_eff preset failed: {e}")

# ===== MAIN COMBO PROCESSING =====

def apply_combo_to_file(baseline, out_path, combo, std_mm_psf, std_alignment, std_thermal, std_gravity, num_mm):
    """Apply combo parameters to create a unique input file."""
    # Copy baseline
    shutil.copy2(baseline, out_path)
    # Prefer the richer sensitivity_run helpers when available to ensure
    # generated workbooks follow the same rules as `sensitivity/sensitivity_run.py`.
    try:
        import sensitivity.sensitivity_run as sr
    except Exception:
        sr = None

    # Apply MM_PSF preset (prefer sensitivity_run logic)
    if combo.get('MM_PSF') and std_mm_psf:
        mk = None
        try:
            if sr and hasattr(sr, '_find_std_mm_psf_key'):
                mk = sr._find_std_mm_psf_key(combo['MM_PSF'], std_mm_psf)
            else:
                mk = find_mm_psf_key(combo['MM_PSF'], std_mm_psf)
        except Exception:
            mk = find_mm_psf_key(combo['MM_PSF'], std_mm_psf)

        if mk and mk in std_mm_psf:
            print(f"    Applying MM_PSF preset: {mk}")
            # If sensitivity_run provides the preservative writer and sampling helpers,
            # use them to preserve template columns and ensure numeric per-MM sigmas.
            try:
                if sr and hasattr(sr, 'write_mmpsf_preserve_template_and_expand'):
                    # construct a minimal left-side df from baseline to pass through
                    try:
                        import pandas as _pd
                        base_df = _pd.read_excel(baseline, sheet_name='MM_PSF', engine='openpyxl')
                        left_cols_count = 10
                        if base_df.shape[1] >= left_cols_count:
                            df_gen = base_df.iloc[:num_mm, :left_cols_count].reset_index(drop=True)
                        else:
                            df_gen = base_df.iloc[:num_mm].reset_index(drop=True)
                    except Exception:
                        df_gen = None
                    try:
                        sr.write_mmpsf_preserve_template_and_expand(baseline, out_path, df_gen, std_mm_psf, mk)
                    except Exception:
                        # fallback to simple applicator
                        apply_mm_psf_preset(out_path, std_mm_psf[mk], num_mm)
                    # ensure numeric sampling/masking/enforcement
                    try:
                        if hasattr(sr, '_sample_per_mm_sigmas_and_write'):
                            sr._sample_per_mm_sigmas_and_write(out_path, std_mm_psf, mk, num_mm)
                    except Exception:
                        pass
                    try:
                        if hasattr(sr, '_mask_alpha_for_non_pseudo_voigt'):
                            sr._mask_alpha_for_non_pseudo_voigt(out_path, num_mm)
                    except Exception:
                        pass
                    try:
                        if hasattr(sr, '_enforce_mmpsf_column_bounds'):
                            sr._enforce_mmpsf_column_bounds(out_path, baseline, num_mm)
                    except Exception:
                        pass
                else:
                    apply_mm_psf_preset(out_path, std_mm_psf[mk], num_mm)
            except Exception:
                apply_mm_psf_preset(out_path, std_mm_psf[mk], num_mm)

    # Apply Alignment preset (prefer sensitivity_run logic)
    if combo.get('Alignment') and std_alignment:
        match_key = None
        try:
            if sr and hasattr(sr, '_norm_name'):
                # use internal normalizer if available
                target_norm = sr._norm_name(combo['Alignment'])
                for k in std_alignment:
                    if sr._norm_name(k) == target_norm:
                        match_key = k
                        break
            else:
                target_norm = norm_name(combo['Alignment'])
                for k in std_alignment:
                    if norm_name(k) == target_norm:
                        match_key = k
                        break
        except Exception:
            target_norm = norm_name(combo['Alignment'])
            for k in std_alignment:
                if norm_name(k) == target_norm:
                    match_key = k
                    break

        if match_key:
            print(f"    Applying Alignment preset: {match_key}")
            try:
                if sr and hasattr(sr, '_apply_alignment_preset_to_workbook'):
                    sr._apply_alignment_preset_to_workbook(out_path, std_alignment[match_key], num_mm, match_key)
                else:
                    apply_alignment_preset(out_path, std_alignment[match_key], num_mm)
            except Exception:
                try:
                    apply_alignment_preset(out_path, std_alignment[match_key], num_mm)
                except Exception:
                    pass

    # Apply Thermal preset
    if combo.get('Thermal') and std_thermal:
        match_key = None
        try:
            if sr and hasattr(sr, '_norm_name'):
                target_norm = sr._norm_name(combo['Thermal'])
                for k in std_thermal:
                    if sr._norm_name(k) == target_norm:
                        match_key = k
                        break
            else:
                target_norm = norm_name(combo['Thermal'])
                for k in std_thermal:
                    if norm_name(k) == target_norm:
                        match_key = k
                        break
        except Exception:
            target_norm = norm_name(combo['Thermal'])
            for k in std_thermal:
                if norm_name(k) == target_norm:
                    match_key = k
                    break

        if match_key:
            print(f"    Applying Thermal preset: {match_key}")
            try:
                if sr and hasattr(sr, '_apply_thermal_preset_to_workbook'):
                    sr._apply_thermal_preset_to_workbook(out_path, std_thermal[match_key], num_mm, match_key)
                else:
                    apply_thermal_preset(out_path, std_thermal[match_key], num_mm)
            except Exception:
                try:
                    apply_thermal_preset(out_path, std_thermal[match_key], num_mm)
                except Exception:
                    pass

    # Apply Gravity preset
    if combo.get('Gravity offload') and std_gravity:
        match_key = None
        try:
            if sr and hasattr(sr, '_norm_name'):
                target_norm = sr._norm_name(combo['Gravity offload'])
                for k in std_gravity:
                    if sr._norm_name(k) == target_norm:
                        match_key = k
                        break
            else:
                target_norm = norm_name(combo['Gravity offload'])
                for k in std_gravity:
                    if norm_name(k) == target_norm:
                        match_key = k
                        break
        except Exception:
            target_norm = norm_name(combo['Gravity offload'])
            for k in std_gravity:
                if norm_name(k) == target_norm:
                    match_key = k
                    break

        if match_key:
            print(f"    Applying Gravity preset: {match_key}")
            try:
                if sr and hasattr(sr, '_apply_gravity_preset_to_workbook'):
                    sr._apply_gravity_preset_to_workbook(out_path, std_gravity[match_key], num_mm, match_key)
                else:
                    apply_gravity_preset(out_path, std_gravity[match_key], num_mm)
            except Exception:
                try:
                    apply_gravity_preset(out_path, std_gravity[match_key], num_mm)
                except Exception:
                    pass

    # Apply A_eff preset
    if combo.get('A_eff'):
        print(f"    Applying A_eff preset: {combo['A_eff']}")
        apply_aeff_preset(out_path, combo['A_eff'], num_mm)

def run_combo(args):
    """Run main.py for a single combo."""
    combo_id, combo, baseline, std_mm_psf, std_alignment, std_thermal, std_gravity, num_mm = args
    
    # Generate unique filename
    ts = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
    parts = [f"{k}={v}" for k, v in combo.items()]
    name_suffix = '_'.join(sanitize_filename(p) for p in parts)
    filename = f"{ts}_{combo_id}_{name_suffix}.xlsx"
    input_path = INPUT_DIR / filename
    
    # Apply combo parameters
    apply_combo_to_file(baseline, input_path, combo, std_mm_psf, std_alignment, std_thermal, std_gravity, num_mm)
    
    # Run main.py
    cmd = [
        "python3", str(ROOT / 'main.py'),
        "-f", str(input_path),
        "--placement", "elliptical",
        "--return_metrics_only"
    ]
    
    try:
        proc = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, timeout=600)
        
        if proc.returncode != 0:
            return {
                'combo_id': combo_id,
                'combo': combo,
                'file': str(input_path),
                'error': proc.stderr
            }
        
        # Parse JSON from stdout
        output = proc.stdout.strip()
        start = output.rfind('{')
        if start == -1:
            return {
                'combo_id': combo_id,
                'combo': combo,
                'file': str(input_path),
                'error': 'No JSON in output'
            }
        
        try:
            metrics = json.loads(output[start:])
        except Exception:
            return {
                'combo_id': combo_id,
                'combo': combo,
                'file': str(input_path),
                'error': 'JSON parse error'
            }
        
        return {
            'combo_id': combo_id,
            'combo': combo,
            'file': str(input_path),
            'metrics': metrics,
            'error': None
        }
            
    except subprocess.TimeoutExpired:
        return {
            'combo_id': combo_id,
            'combo': combo,
            'file': str(input_path),
            'error': 'Timeout'
        }
    except Exception as e:
        return {
            'combo_id': combo_id,
            'combo': combo,
            'file': str(input_path),
            'error': str(e)
        }
    finally:
        # Cleanup input file
        try:
            if input_path.exists():
                input_path.unlink()
        except:
            pass

def save_results(results):
    """Save only valid results."""
    valid = [r for r in results if r.get('metrics') and not r.get('error')]
    
    if not valid:
        print("No valid results to save!")
        return
    
    for r in valid:
        combo_id = r['combo_id']
        combo = r['combo']
        
        row = {
            'combo_id': combo_id,
            'A_eff': combo.get('A_eff', '-'),
            'MM_PSF': combo.get('MM_PSF', '-'),
            'Alignment': combo.get('Alignment', '-'),
            'Thermal': combo.get('Thermal', '-'),
            'Gravity offload': combo.get('Gravity offload', '-'),
            'input_file': Path(r['file']).name,
            'hew_origin_arcsec': r['metrics'].get('hew_origin_arcsec'),
            'hew_best_arcsec': r['metrics'].get('hew_best_arcsec'),
            'hew_opt_arcsec': r['metrics'].get('hew_opt_arcsec'),
            'eef90_origin_arcsec': r['metrics'].get('eef90_origin_arcsec'),
            'eef90_best_arcsec': r['metrics'].get('eef90_best_arcsec'),
            'eef90_opt_arcsec': r['metrics'].get('eef90_opt_arcsec'),
        }
        
        mode = 'a' if PARTIAL_CSV.exists() else 'w'
        header = mode == 'w'
        
        df_row = pd.DataFrame([row])
        df_row.to_csv(PARTIAL_CSV, mode=mode, header=header, index=False)
    
    print(f"Saved {len(valid)} results")

def finalize():
    """Merge partial CSV to final XLSX."""
    if not PARTIAL_CSV.exists():
        return
    
    df = pd.read_csv(PARTIAL_CSV)
    if df.empty:
        return
    
    df = df.drop_duplicates(subset=['combo_id'], keep='last')
    df = df.sort_values('combo_id')
    df.to_excel(FINAL_XLSX, index=False)
    print(f"Final: {len(df)} results in {FINAL_XLSX}")

def main():
    print("=" * 60)
    print("Batched Sensitivity Run")
    print("=" * 60)
    
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    
    baseline = ROOT / 'Distributions' / 'TestDistribution.xlsx'
    if not baseline.exists():
        print(f"Baseline not found: {baseline}")
        return
    
    # Load presets
    print("Loading presets...")
    std_mm_psf = load_standard_mm_psf_defs(baseline)
    std_alignment = load_standard_alignment_defs(baseline)
    std_thermal = load_standard_thermal_defs(baseline)
    std_gravity = load_standard_gravity_defs(baseline)
    print(f"  {len(std_mm_psf)} MM_PSF, {len(std_alignment)} Alignment, {len(std_thermal)} Thermal, {len(std_gravity)} Gravity")
    
    # Get num_mm
    try:
        mm_cfg = pd.read_excel(baseline, sheet_name='MM configuration', engine='openpyxl')
        num_mm = len(mm_cfg['MM #'].dropna()) if 'MM #' in mm_cfg.columns else 8
    except:
        num_mm = 8
    
    # Load combos
    sens_df = pd.read_excel(SENS_DIR / 'sensitivity_input.xlsx', engine='openpyxl')
    choices = {}
    for col in sens_df.columns:
        vals = []
        for v in sens_df[col].dropna():
            if isinstance(v, str) and (',' in v or '\n' in v):
                vals.extend([p.strip() for p in re.split('[,\n]', v) if p.strip()])
            else:
                vals.append(v)
        choices[col] = list(dict.fromkeys(vals))
    
    combos = [dict(zip(choices.keys(), prod)) for prod in product(*(choices[c] for c in choices))]
    print(f"Total combos: {len(combos)}")
    
    # Process batches
    for batch_start in range(0, len(combos), BATCH_SIZE):
        batch = combos[batch_start:batch_start + BATCH_SIZE]
        batch_num = batch_start // BATCH_SIZE + 1
        
        print(f"\n{'=' * 60}")
        print(f"Batch {batch_num}: combos {batch_start + 1} to {min(batch_start + BATCH_SIZE, len(combos))}")
        print(f"{'=' * 60}")
        
        cleanup_input_dir()
        
        # Prepare args for parallel execution
        args_list = [(batch_start + i + 1, combo, baseline, std_mm_psf, std_alignment, std_thermal, std_gravity, num_mm) for i, combo in enumerate(batch)]
        
        # Run in parallel
        max_workers = min(NUM_WORKERS, len(args_list))
        print(f"Running {len(args_list)} combos with {max_workers} workers...")
        
        results = []
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(run_combo, args): args for args in args_list}
            
            for future in as_completed(futures):
                result = future.result()
                results.append(result)
                
                combo_id = result['combo_id']
                if result.get('error'):
                    print(f"  ERROR {combo_id}: {str(result.get('error', ''))[:50]}")
                else:
                    hew = result.get('metrics', {}).get('hew_origin_arcsec', 'N/A')
                    print(f"  OK {combo_id}: HEW={hew}")
        
        # Save results
        save_results(results)
        
        # Cleanup
        cleanup_input_dir()
        
        # Progress
        success = sum(1 for r in results if r.get('metrics') and not r.get('error'))
        errors = len(results) - success
        print(f"\nBatch {batch_num}: {success} OK, {errors} errors")
    
    # Finalize
    print(f"\n{'=' * 60}")
    print("Finalizing results...")
    print(f"{'=' * 60}")
    finalize()
    print(f"\nDone! Results in {FINAL_XLSX}")

if __name__ == '__main__':
    main()

