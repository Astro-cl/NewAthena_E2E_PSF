from openpyxl import load_workbook
import os,glob
pkg_dir='Exports/Export_reallistic_20260416_112020/2__Ang__20__E__7__Def__0_reallistic_20260416_112145'
print('Package dir exists:', os.path.isdir(pkg_dir))
# find fitparams
fit_paths=glob.glob(os.path.join(pkg_dir,'fitparams_aeffloss.xlsx'))
if not fit_paths:
    fit_paths=glob.glob(os.path.join(pkg_dir,'EEF_fittingparams*.xlsx'))
print('Found fitparams:', fit_paths)
for fp in fit_paths:
    print('\n== Fitparams:',fp)
    try:
        wb=load_workbook(fp,data_only=True)
        if 'Fit_parameters' in wb.sheetnames:
            ws=wb['Fit_parameters']
            for r in range(1, ws.max_row+1):
                k=ws.cell(row=r,column=1).value
                v=ws.cell(row=r,column=2).value
                print(k,'=>',v)
        else:
            print('No Fit_parameters sheet, sheets:', wb.sheetnames)
    except Exception as e:
        print('Failed to read fitparams:',e)

# find exported workbook (input file copied into package)
xls = None
for cand in os.listdir(pkg_dir):
    if cand.lower().endswith('.xlsx') and not cand.lower().startswith('eef_fittingparams') and 'fitparams' not in cand.lower():
        xls=os.path.join(pkg_dir,cand)
        break
print('\nExported workbook:', xls)
if xls:
    try:
        wb2=load_workbook(xls,data_only=True)
        if 'A_eff' in wb2.sheetnames:
            ws2=wb2['A_eff']
            sumB=0.0
            sumC=0.0
            count=0
            for r in range(2, ws2.max_row+1):
                mm=ws2.cell(row=r,column=1).value
                if mm is None:
                    continue
                try:
                    float(mm)
                except Exception:
                    continue
                vB=ws2.cell(row=r,column=2).value
                vC=ws2.cell(row=r,column=3).value
                try:
                    sumB += float(vB) if vB is not None else 0.0
                except Exception:
                    pass
                try:
                    sumC += float(vC) if vC is not None else 0.0
                except Exception:
                    pass
                count+=1
            print('\nA_eff rows:',count,'sumB',sumB,'sumC',sumC)
        else:
            print('No A_eff sheet in exported workbook')
    except Exception as e:
        print('Failed to read exported workbook:',e)
