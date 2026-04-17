"""
gui_distributions.py
--------------------
Tkinter-based interactive GUI for generating and exporting per-MM
distribution parameters used by the PSF analysis pipeline.

Key responsibilities:
- Load Excel workbooks containing MM configuration, A_eff, MM_PSF presets
    and perturbation tables (Alignment, Thermal, Gravity offload).
- Provide a user-friendly preset selection and per-MM application flow.
- Export generated parameter tables back to Excel while preserving other
    workbook sheets and formatting where possible.

Notes:
- The A_eff tab includes an "Apply vignetting factors when exporting"
    checkbox which, when set and used with a standard A_eff preset, will copy
    the selected energy column from the `Vignetting rotazi` / `Vignetting rotrad`
    sheets into column B of those sheets prior to saving the workbook.
"""
import matplotlib
matplotlib.use('TkAgg')  # Use Tkinter backend (works on most systems)
# matplotlib.use('Agg')  # Use this for non-GUI/headless environments
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import numpy as np
import pandas as pd
import sys
import re
import traceback
import ast
from pathlib import Path
import logging
# Candidate vignette sheet names (support both legacy and new workbook layouts)
VIG_ROT_AZI_CANDIDATES = ('MM vignetting rotazi', 'Vignetting rotazi')
VIG_ROT_RAD_CANDIDATES = ('MM vignetting rotrad', 'Vignetting rotrad')

def _pick_sheet_in_wb(wb, candidates):
    try:
        for s in candidates:
            if s in wb.sheetnames:
                return s
    except Exception:
        pass
    return None

# Minimal DATA_TYPES structure required by the GUI module. This mirrors the
# structure expected by the rest of the code and provides reasonable defaults
# for non-interactive tests. The full definitions are normally loaded from
# the Excel sheets at runtime.
DATA_TYPES = {
    'Alignment': {
        'sheet_name': 'Alignment',
        'tab_label': 'Alignment',
        # Keep translational parameters first; move rotation-related
        # parameters to the end so they appear last in the GUI.
        'params': ['d_align_rad [µm]', 'd_align_azi [µm]', 'd_align_z [µm]', 'd_align_rotazi [arcsec]', 'd_align_rotrad [arcsec]', 'd_align_rotz [arcsec]'],
        'defaults': {
            'd_align_rad [µm]': (0, 0.1),
            'd_align_azi [µm]': (0, 0.1),
            'd_align_z [µm]': (0, 0.1),
            'd_align_rotazi [arcsec]': (0, 0.01),
            'd_align_rotrad [arcsec]': (0, 0.01),
            'd_align_rotz [arcsec]': (0, 0.01)
        }
    },
    'Thermal': {
        'sheet_name': 'Thermal',
        'tab_label': 'Thermal',
        'params': ['d_therm_x [µm]', 'd_therm_y [µm]', 'd_therm_z [µm]', 'd_therm_rotx [arcsec]', 'd_therm_roty [arcsec]', 'd_therm_rotz [arcsec]'],
        'defaults': {
            'd_therm_x [µm]': (0, 0.1),
            'd_therm_y [µm]': (0, 0.1),
            'd_therm_z [µm]': (0, 0.1),
            'd_therm_rotx [arcsec]': (0, 0.01),
            'd_therm_roty [arcsec]': (0, 0.01),
            'd_therm_rotz [arcsec]': (0, 0.01)
        }
    },
    'Gravity offload': {
        'sheet_name': 'Gravity offload',
        'tab_label': 'Gravity offload',
        'params': ['d_grav_x [µm]', 'd_grav_y [µm]', 'd_grav_z [µm]', 'd_grav_rotx [arcsec]', 'd_grav_roty [arcsec]', 'd_grav_rotz [arcsec]'],
        'defaults': {
            'd_grav_x [µm]': (0, 0.1),
            'd_grav_y [µm]': (0, 0.1),
            'd_grav_z [µm]': (0, 0.1),
            'd_grav_rotx [arcsec]': (0, 0.01),
            'd_grav_roty [arcsec]': (0, 0.01),
            'd_grav_rotz [arcsec]': (0, 0.01)
        }
    },
    'MM_PSF': {
        'sheet_name': 'MM_PSF',
        'tab_label': 'PSF',
        'params': ['sigma_rad [arcsec]', 'sigma_azi [arcsec]'],
        'alpha_params': ['alpha_rad', 'alpha_azi'],
        'defaults': {
            'sigma_rad [arcsec]': (3.0, 1.0),
            'sigma_azi [arcsec]': (3.0, 1.0),
            'alpha_rad': (0.5, 0.1),
            'alpha_azi': (0.5, 0.1)
        },
        'has_distribution': True
    }
}


def apply_macos_input_fixes(root: tk.Tk) -> None:
    """Improve ttk widget click reliability on macOS trackpads.

    On Aqua/Tk, a slight pointer drift between press/release can prevent ttk buttons
    and some readonly comboboxes from activating, which feels like clicks are ignored.
    This applies conservative, macOS-only class bindings.
    """
    try:
        is_aqua = (root.tk.call('tk', 'windowingsystem') == 'aqua')
    except Exception:
        is_aqua = (sys.platform == 'darwin')

    if not is_aqua:
        return

    def _ensure_focus() -> None:
        # Ensure these are always defined for later summary logic
        successes = []
        failures = []
        failure_reasons = []

        try:
            # Make sure the window is the active/focused window.
            # On macOS the first click can be consumed just to activate the app.
            root.update_idletasks()
            root.deiconify()
            root.lift()
            try:
                root.attributes('-topmost', True)
                root.update_idletasks()
                root.attributes('-topmost', False)
            except Exception:
                pass

            root.focus_force()
        except Exception:
            pass

    # Give the WM time to map the window before forcing focus.
    root.after(200, _ensure_focus)

    def _press_invokeable(event):
        w = event.widget
        try:
            w.focus_set()
        except Exception:
            pass
        try:
            w.state(['pressed'])
        except Exception:
            pass
        setattr(w, '_mac_pressed', True)
        return 'break'

    def _release_invokeable(event):
        w = event.widget
        pressed = getattr(w, '_mac_pressed', False)
        setattr(w, '_mac_pressed', False)
        try:
            w.state(['!pressed'])
        except Exception:
            pass

        if pressed:
            try:
                w.invoke()
            except Exception:
                pass
        return 'break'

    # Replace the default mouse bindings for ttk buttons (macOS only).
    root.bind_class('TButton', '<ButtonPress-1>', _press_invokeable, add=False)
    root.bind_class('TButton', '<ButtonRelease-1>', _release_invokeable, add=False)
    root.bind_class('TCheckbutton', '<ButtonPress-1>', _press_invokeable, add=False)
    root.bind_class('TCheckbutton', '<ButtonRelease-1>', _release_invokeable, add=False)
    root.bind_class('TRadiobutton', '<ButtonPress-1>', _press_invokeable, add=False)
    root.bind_class('TRadiobutton', '<ButtonRelease-1>', _release_invokeable, add=False)

    def _tcombobox_click(event):
        w = event.widget
        try:
            w.focus_set()
        except Exception:
            pass
        try:
            state = str(w.cget('state'))
        except Exception:
            state = ''

        # For readonly comboboxes, make the whole widget clickable to open the list.
        if state == 'readonly':
            try:
                w.focus_set()
                w.tk.call('ttk::combobox::Post', w)
                # Don't return 'break' to allow default selection behavior
            except Exception:
                pass
        return None

    # Use ButtonRelease to avoid fighting the entry caret placement.
    root.bind_class('TCombobox', '<ButtonRelease-1>', _tcombobox_click, add=True)


def generate_values(dist, a, b, count):
    """Sample `count` values from a distribution spec.

    Parameters
    - dist: distribution type string ('fixed', 'gaussian', 'gamma', 'uniform')
    - a, b: distribution parameters (mean/scale or bounds depending on type)
    - count: number of samples to draw

    Returns a NumPy array of length `count`.
    """
    count = int(count)
    if dist == 'fixed':
        # Return array of fixed value (a parameter)
        return np.full(count, float(a))
    elif dist == 'gaussian':
        # Default gaussian sampling; callers may apply one-sided logic for
        # sigmas/alphas. Keep raw normal here for generic uses.
        rng = np.random.default_rng()
        return rng.normal(float(a), float(b), count)
    elif dist == 'gamma':
        # a = mean, b = std -> derive shape k and scale theta
        mu = float(a)
        sigma = float(b)
        if sigma <= 0 or mu <= 0:
            return np.full(count, mu if mu > 0 else 1e-6)
        k = (mu / sigma) ** 2
        theta = (sigma ** 2) / mu
        rng = np.random.default_rng()
        return rng.gamma(shape=k, scale=theta, size=count)
    elif dist == 'uniform':
        rng = np.random.default_rng()
        return rng.uniform(float(a), float(b), count)
    else:
        raise ValueError('Unknown distribution: ' + dist)


def generate_data_from_distributions(params, num_mm, data_type_config):
    """Generate data parameters for each MM based on distributions."""
    samples = {}
    for key, (dist, a, b) in params.items():
        samples[key] = generate_values(dist, a, b, num_mm)
    
    df_dict = {}
    # Add regular parameters
    for param in data_type_config['params']:
        # Use the full param name (with units) as the column name
        if param in samples:
            # For PSF sigmas and alpha mixing parameters we require positive
            # one-sided behaviour. Implement truncated-at-zero sampling by
            # resampling negative draws where necessary. For alpha parameters
            # we also clip to [0,1].
            vals = samples[param]
            if 'sigma' in param.lower() or 'alpha' in param.lower():
                # vectorized rejection sampling: replace non-positive entries
                # by resampling up to a small iteration cap.
                rng = np.random.default_rng()
                mask = vals <= 0
                attempts = 0
                # use the distribution spec for proper resampling
                dist_spec = params.get(param, ('gaussian', a, b))[0]
                while mask.any() and attempts < 100:
                    if dist_spec == 'gaussian':
                        vals[mask] = rng.normal(loc=float(a), scale=float(b), size=mask.sum())
                    elif dist_spec == 'gamma':
                        mu = float(a); sigma = float(b)
                        if sigma <= 0 or mu <= 0:
                            vals[mask] = mu if mu > 0 else 1e-6
                        else:
                            k = (mu / sigma) ** 2
                            theta = (sigma ** 2) / mu
                            vals[mask] = rng.gamma(shape=k, scale=theta, size=mask.sum())
                    elif dist_spec == 'uniform':
                        vals[mask] = rng.uniform(low=float(a), high=float(b), size=mask.sum())
                    else:
                        vals[mask] = float(a)
                    mask = vals <= 0
                    attempts += 1
                vals[vals <= 0] = 1e-6
                if 'alpha' in param.lower():
                    vals = np.clip(vals, 0.0, 1.0)
                df_dict[param] = vals
            else:
                df_dict[param] = samples[param]
    
    # Add alpha parameters if present
    if 'alpha_params' in data_type_config:
        for param in data_type_config['alpha_params']:
            if param in samples:
                df_dict[param] = samples[param]
    
    df = pd.DataFrame(df_dict)
    
    # Validate positive sigmas for PSF
    sigma_cols = [col for col in df.columns if 'sigma' in col.lower()]
    if sigma_cols:
        for col in sigma_cols:
            if (df[col] <= 0).any():
                raise ValueError('Generated sigmas must be positive. Please adjust the distribution parameters.')
    
    return df


class ExtendedGUI:
    """Main GUI application class for interactive generation and export.

    Responsibilities:
    - Load Excel workbooks and extract MM configuration and standard presets
      (A_eff, MM_PSF, Alignment, Thermal, Gravity offload).
    - Provide per-data-type distribution editors and sampling controls.
    - Preview generated tables and export updated workbooks or plots.
    """
    def __init__(self, root):
        self.root = root
        root.title('MM Configuration Generator GUI')
        root.geometry('1100x850')

        self.excel_path = None
        self.mm_config_df = None
        self.selected_mm_numbers = []
        self.aeff_weights = {}
        self.aeff_raw_df = None
        self.aeff_standard_presets = {}  # preset_name -> values_expr
        self.aeff_mode_var = tk.StringVar(value='standard')  # 'standard' or 'fixed'
        self.aeff_fixed_var = tk.StringVar(value='')
        self.aeff_selected_preset_var = tk.StringVar(value='')
        self.aeff_pending_export = False
        # Vignetting is applied automatically based on selected A_eff preset
        # or selected energy when using a fixed A_eff.
        # Free-mode vignetting energy selector (user-entered or chosen)
        self.aeff_free_energy_var = tk.StringVar(value='')
        # When True, do not show modal 'Applied preset' dialogs (useful during UI setup)
        self.suppress_standard_apply_modals = False
        self.export_mode_var = tk.StringVar(value='current')
        self.export_path_var = tk.StringVar(value='')
        
        # Store data for each type
        self.data_dfs = {key: None for key in DATA_TYPES.keys()}
        
        # Store which data types are enabled
        self.enabled_data_types = []
        self.data_type_checkboxes = {}
        self.aeff_checkbox_var = tk.BooleanVar(value=False)
        
        # Create main notebook
        self.main_notebook = ttk.Notebook(root)
        self.main_notebook.pack(fill='both', expand=True)

        # Load tab (common)
        self.tab_load = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.tab_load, text='Load File')
        
        # MM Configuration tab (common)
        self.tab_config = ttk.Frame(self.main_notebook)

        # A_eff tab (common)
        self.tab_aeff = ttk.Frame(self.main_notebook)
        
        # Preview/Export tab (common)
        self.tab_preview = ttk.Frame(self.main_notebook)
        
        # Data type tabs will be created dynamically
        self.data_type_tabs = {}
        self.tab_status_labels = {}  # data_type_key -> ttk.Label for pending export status
        self.dist_entries_by_type = {}
        self.param_labels_by_type = {}
        self.distribution_widgets = {}  # Store distribution type and alpha widgets
        self.alpha_entries_by_type = {}  # Store alpha parameter widgets separately
        self.standard_distributions = {}  # Store standard distribution definitions from Excel
        self.psf_mode_var = tk.StringVar(value='standard')  # 'standard' or 'free' - default to standard

        # Standard Alignment presets (loaded from Alignment sheet columns starting at G1)
        self.alignment_standard_presets = {}  # preset_name -> {param_label: spec_str}
        self.align_mode_var = tk.StringVar(value='standard')  # 'standard' or 'free' - default to standard
        # Thermal and Gravity standard presets
        self.thermal_standard_presets = {}
        self.thermal_mode_var = tk.StringVar(value='standard')
        self.gravity_standard_presets = {}
        self.gravity_mode_var = tk.StringVar(value='standard')

        # Custom PSF-from-file support (MM_PSF)
        self.CUSTOM_PSF_OPTION = 'Custom PSF (select file...)'
        self.custom_psf_path_var = tk.StringVar(value='')
        self.custom_psf_stem_var = tk.StringVar(value='')

        # Build initial tabs
        self.build_load_tab()
        self.build_config_tab()
        self.build_aeff_tab()
        self.build_preview_tab()

    def build_load_tab(self):
        """Build the initial 'Load File' tab.

        The Load tab provides a button to open an Excel workbook, and when a
        workbook is selected the function loads relevant sheets (MM configuration,
        A_eff, Alignment, Thermal, Gravity offload, MM_PSF) into memory and
        prepares standard preset lists for the UI.
        """
        frame = self.tab_load
        ttk.Label(frame, text='Load Excel File with MM Configuration', font=('Arial', 14)).pack(pady=10)
        ttk.Button(frame, text='Load Excel File', command=self.load_excel).pack(pady=5)
        self.load_status = ttk.Label(frame, text='', foreground='blue')
        self.load_status.pack(pady=10)
        
        # Data type selection frame (initially hidden)
        self.selection_frame = ttk.LabelFrame(frame, text='Select Data Types to Modify', padding=10)
        # Don't pack it yet - will be shown after successful load
        
        ttk.Label(self.selection_frame, text='Choose which data types you want to generate/modify:', 
                  font=('Arial', 10)).pack(anchor='w', pady=5)

        # A_eff is a special case: it uses its own tab (not a DATA_TYPES generate tab)
        ttk.Checkbutton(
            self.selection_frame,
            text='A_eff',
            variable=self.aeff_checkbox_var,
        ).pack(anchor='w', padx=20, pady=2)
        
        # Create checkboxes for each data type.
        # Requirement: PSF should appear second after A_eff, so add MM_PSF explicitly,
        # then add the remaining data types in the original order.
        # Add PSF (MM_PSF) as the second checkbox when available in DATA_TYPES.
        if 'MM_PSF' in DATA_TYPES:
            psf_cfg = DATA_TYPES['MM_PSF']
            var = tk.BooleanVar(value=False)
            self.data_type_checkboxes['MM_PSF'] = var
            ttk.Checkbutton(self.selection_frame, text=psf_cfg['tab_label'], variable=var).pack(anchor='w', padx=20, pady=2)

        # Add the rest of the data types (skip MM_PSF since already added)
        for data_type_key, config in DATA_TYPES.items():
            if data_type_key == 'MM_PSF':
                continue
            var = tk.BooleanVar(value=False)
            self.data_type_checkboxes[data_type_key] = var
            ttk.Checkbutton(self.selection_frame, text=config['tab_label'], variable=var).pack(anchor='w', padx=20, pady=2)
        
        ttk.Button(self.selection_frame, text='Apply Selection', command=self.apply_data_type_selection).pack(pady=10)
        
        # Feedback label for when selection is applied
        self.selection_feedback = ttk.Label(frame, text='', foreground='green', font=('Arial', 10, 'bold'))
        self.selection_feedback.pack(pady=5)

    def load_excel(self):
        """Open an Excel file and load known configuration sheets.

        Loads the `MM configuration` sheet into `self.mm_config_df` and any
        known data-type sheets (as defined in `DATA_TYPES`) into
        `self.data_dfs`. Also attempts to load standard presets from the
        `MM_PSF`, `Alignment`, `Thermal`, `Gravity offload` and `A_eff`
        sheets when present. Any recoverable errors are printed rather than
        raised to keep the GUI responsive.
        """
        path = filedialog.askopenfilename(initialdir="./Distributions", filetypes=[('Excel files', '*.xlsx *.xls')])
        if not path:
            return
        try:
            self.excel_path = path
            self.mm_config_df = pd.read_excel(path, sheet_name='MM configuration', engine='openpyxl')
            
            # Load existing data for each type if available
            with pd.ExcelFile(path, engine='openpyxl') as xls:
                for data_type_key, config in DATA_TYPES.items():
                    sheet_name = config['sheet_name']
                    if sheet_name in xls.sheet_names:
                        self.data_dfs[data_type_key] = pd.read_excel(path, sheet_name=sheet_name, engine='openpyxl')
            
            # Load standard MM_PSF distributions table if available
            try:
                std_dist_df = pd.read_excel(path, sheet_name='MM_PSF', engine='openpyxl', header=None)
                self.load_standard_distributions(std_dist_df)
                try:
                    self.refresh_standard_distribution_controls()
                except Exception:
                    pass
            except Exception as e:
                logging.debug("Could not load standard distributions: %s", e)
                self.standard_distributions = {}

            # Load standard Alignment presets table if available
            try:
                align_raw = pd.read_excel(path, sheet_name='Alignment', engine='openpyxl', header=None)
                self.load_standard_alignment_presets(align_raw)
                try:
                    self.refresh_standard_distribution_controls()
                except Exception:
                    pass
            except Exception as e:
                logging.debug("Could not load standard alignment presets: %s", e)
                self.alignment_standard_presets = {}
            # Load standard Thermal presets table if available
            try:
                therm_raw = pd.read_excel(path, sheet_name='Thermal', engine='openpyxl', header=None)
                self.load_standard_thermal_presets(therm_raw)
                try:
                    self.refresh_standard_distribution_controls()
                except Exception:
                    pass
            except Exception as e:
                logging.debug("Could not load standard thermal presets: %s", e)
                self.thermal_standard_presets = {}

            # Load standard Gravity offload presets table if available
            try:
                grav_raw = pd.read_excel(path, sheet_name='Gravity offload', engine='openpyxl', header=None)
                self.load_standard_gravity_presets(grav_raw)
                try:
                    self.refresh_standard_distribution_controls()
                except Exception:
                    pass
            except Exception as e:
                logging.debug("Could not load standard gravity presets: %s", e)
                self.gravity_standard_presets = {}
            
            # Load A_eff weights + standard preset table (headerless)
            self.aeff_raw_df = None
            self.aeff_weights = {}
            self.aeff_standard_presets = {}
            try:
                # Load A_eff sheet using openpyxl directly to get cached formula values
                import openpyxl as _openpyxl
                wb = _openpyxl.load_workbook(path, data_only=True)
                ws = wb['A_eff']
                
                # Convert worksheet to DataFrame
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                aeff_raw = pd.DataFrame(data)
                self.aeff_raw_df = aeff_raw

                # Expect MM list in col A and weights in col B.
                # Build map MM# -> weight from rows with numeric MM#.
                for r in range(1, aeff_raw.shape[0]):
                    mm = aeff_raw.iloc[r, 0] if aeff_raw.shape[1] > 0 else None
                    w = aeff_raw.iloc[r, 1] if aeff_raw.shape[1] > 1 else None
                    if pd.isna(mm) or mm == '':
                        continue
                    try:
                        mm_i = int(float(mm))
                    except Exception:
                        continue
                    try:
                        w_f = float(w)
                    except Exception:
                        continue
                    self.aeff_weights[mm_i] = w_f

                # Load standard presets from columns starting at D (if any)
                self.load_standard_aeff_presets(aeff_raw)

                # Refresh A_eff UI controls with any newly loaded presets
                try:
                    self.refresh_aeff_preset_controls()
                except Exception:
                    pass

            except Exception as e:
                logging.debug("Could not load A_eff sheet: %s", e)
                self.aeff_raw_df = None
                self.aeff_weights = {}
                self.aeff_standard_presets = {}
            
            # Validate Row # column
            if 'Row #' in self.mm_config_df.columns:
                for idx, val in enumerate(self.mm_config_df['Row #']):
                    try:
                        float(val)
                    except (ValueError, TypeError):
                        messagebox.showerror('Error', f'Row # at index {idx} contains non-numeric value: "{val}".')
                        return

            # Validate Position # column (if present)
            if 'Position #' in self.mm_config_df.columns:
                for idx, val in enumerate(self.mm_config_df['Position #']):
                    try:
                        float(val)
                    except (ValueError, TypeError):
                        messagebox.showerror('Error', f'Position # at index {idx} contains non-numeric value: "{val}".')
                        return
            
            self.load_status.configure(text=f'Loaded successfully! Found {len(self.mm_config_df)} MMs.')
            self.export_path_var.set(self.excel_path)
            self.update_config_display()

            # Default A_eff preset selection in the UI (if available)
            if self.aeff_standard_presets:
                preset_names = list(self.aeff_standard_presets.keys())
                if preset_names and not self.aeff_selected_preset_var.get().strip():
                    self.aeff_selected_preset_var.set(preset_names[0])
                try:
                    self.on_aeff_standard_selected()
                except Exception:
                    pass
            
            # Show the data type selection frame
            self.selection_frame.pack(pady=20, padx=20, fill='x')
            self.selection_feedback.configure(text='')
            
            messagebox.showinfo('Success', f'File loaded with {len(self.mm_config_df)} MM configurations!\n\nNow select which data types you want to modify.')
            self.update_preview()
        except Exception as e:
            messagebox.showerror('Error', str(e))
    
    def load_standard_distributions(self, df):
        """Load standard MM_PSF distribution definitions from Excel table starting at M1."""
        self.standard_distributions = {}
        
        try:
            # Find the table starting at M1 (column 12, row 0 in 0-indexed)
            # Columns I-J hold degraded sigma_rad/sigma_azi; K-L are reserved.
            # Expected format: Name | sigma_rad | sigma_azi | alpha_rad | alpha_azi
            start_row = 0  # M1 is row 1, 0-indexed = 0
            start_col = 12  # M is column 13, 0-indexed = 12
            
            # Check if there's a header row
            if df.shape[0] <= start_row or df.shape[1] <= start_col:
                return
            
            # Read headers
            headers = [df.iloc[start_row, start_col + i] for i in range(5)]
            
            # Read data rows
            row_idx = start_row + 1
            while row_idx < df.shape[0]:
                name = df.iloc[row_idx, start_col]
                if pd.isna(name) or name == '':
                    break

                key_name = str(name).strip()
                dist_def = {'name': key_name}
                
                # Determine distribution type from name
                if 'pseudo-voigt' in str(name).lower() or 'voigt' in str(name).lower():
                    dist_def['type'] = 'pseudo-voigt'
                else:
                    dist_def['type'] = 'gaussian'
                
                # Parse parameters
                for i, param in enumerate(['sigma_rad', 'sigma_azi', 'alpha_rad', 'alpha_azi']):
                    cell_value = df.iloc[row_idx, start_col + 1 + i]
                    
                    if pd.isna(cell_value) or cell_value == '':
                        dist_def[param] = None
                        continue
                    
                    cell_str = str(cell_value).strip()

                    # Support numeric expressions and percentages inside gaussian(...)
                    # e.g. gaussian(3.397, 10%*3.397)
                    try:
                        kind, a, b = self._parse_standard_dist_spec(cell_str)
                        if kind == 'fixed':
                            dist_def[param] = {'dist': 'fixed', 'value': float(a)}
                        elif kind == 'gaussian':
                            dist_def[param] = {'dist': 'gaussian', 'mean': float(a), 'sigma': float(abs(b))}
                        elif kind == 'uniform':
                            lo = float(min(a, b))
                            hi = float(max(a, b))
                            dist_def[param] = {'dist': 'uniform', 'min': lo, 'max': hi}
                        else:
                            dist_def[param] = None
                    except Exception:
                        dist_def[param] = None
                
                # Always key by the stripped string name to avoid invisible whitespace mismatches.
                self.standard_distributions[key_name] = dist_def
                row_idx += 1
            
            logging.debug("Loaded %d standard distributions: %s", len(self.standard_distributions), list(self.standard_distributions.keys()))
        except Exception as e:
            logging.exception("Error loading standard distributions: %s", e)
            self.standard_distributions = {}
        # If none loaded from workbook, populate with fallback list provided by user
        if not self.standard_distributions:
            fallback = [
                'Standard medialario',
                '10% worse medialario',
                '30% worse medialario',
                '50% worse medialario',
                '100% worse medialario',
            ]
            for name in fallback:
                # minimal placeholder definition
                self.standard_distributions[name] = {'name': name, 'type': 'gaussian', 'sigma_rad': None, 'sigma_azi': None, 'alpha_rad': None, 'alpha_azi': None}
            logging.debug("Populated fallback standard distributions: %s", fallback)

    def refresh_standard_distribution_controls(self) -> None:
        """Refresh standard preset dropdowns after (re)loading an Excel file.

        This is important when the user loads a different workbook without restarting the GUI.
        """
        # MM_PSF standard distributions
        dt = 'MM_PSF'
        if dt in self.distribution_widgets and 'std_dist_combo' in self.distribution_widgets[dt]:
            combo = self.distribution_widgets[dt]['std_dist_combo']
            values = list(self.standard_distributions.keys()) + [self.CUSTOM_PSF_OPTION]
            try:
                combo.config(values=values)
            except Exception:
                try:
                    combo['values'] = values
                except Exception:
                    pass
            # Keep selection if still valid; otherwise pick first.
            cur = str(combo.get()).strip()
            if cur not in values:
                if self.standard_distributions:
                    combo.set(list(self.standard_distributions.keys())[0])
                elif values:
                    combo.set(values[0])

        # Alignment standard presets
        dt = 'Alignment'
        if dt in self.distribution_widgets and 'align_std_combo' in self.distribution_widgets[dt]:
            combo = self.distribution_widgets[dt]['align_std_combo']
            values = list(self.alignment_standard_presets.keys())
            try:
                combo.config(values=values)
            except Exception:
                try:
                    combo['values'] = values
                except Exception:
                    pass
            cur = str(combo.get()).strip()
            if cur not in values and values:
                combo.set(values[0])

        # Thermal standard presets
        dt = 'Thermal'
        if dt in self.distribution_widgets and 'therm_std_combo' in self.distribution_widgets[dt]:
            combo = self.distribution_widgets[dt]['therm_std_combo']
            values = list(self.thermal_standard_presets.keys())
            try:
                combo.config(values=values)
            except Exception:
                try:
                    combo['values'] = values
                except Exception:
                    pass
            cur = str(combo.get()).strip()
            if cur not in values and values:
                if '0 deg FMS tilt' in values:
                    combo.set('0 deg FMS tilt')
                else:
                    combo.set(values[0])
            # Enforce UI state for thermal mode
            try:
                self.toggle_thermal_mode('Thermal')
            except Exception:
                pass

        # Gravity standard presets
        dt = 'Gravity offload'
        if dt in self.distribution_widgets and 'grav_std_combo' in self.distribution_widgets[dt]:
            combo = self.distribution_widgets[dt]['grav_std_combo']
            values = list(self.gravity_standard_presets.keys())
            try:
                combo.config(values=values)
            except Exception:
                try:
                    combo['values'] = values
                except Exception:
                    pass
            cur = str(combo.get()).strip()
            if cur not in values and values:
                if 'GZ' in values:
                    combo.set('GZ')
                else:
                    combo.set(values[0])
            # Enforce UI state for gravity mode
            try:
                self.toggle_gravity_mode('Gravity offload')
            except Exception:
                pass

    def _safe_eval_numeric_expr(self, expr: str) -> float:
        """Safely evaluate simple numeric expressions like '12/3' or '1.5*110%'."""
        if expr is None:
            raise ValueError('Empty expression')
        s = str(expr).strip()
        if not s:
            raise ValueError('Empty expression')

        # Convert percentages: '110%' -> '(110/100)'
        s = re.sub(r'(\d+(?:\.\d+)?)\s*%', r'(\1/100)', s)

        # Only allow safe characters
        if re.search(r'[^0-9\s\+\-\*\/\(\)\.]', s):
            raise ValueError(f'Unsupported characters in expression: {expr!r}')

        node = ast.parse(s, mode='eval')

        def _eval(n):
            if isinstance(n, ast.Expression):
                return _eval(n.body)
            if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
                return float(n.value)
            if isinstance(n, ast.UnaryOp) and isinstance(n.op, (ast.UAdd, ast.USub)):
                v = _eval(n.operand)
                return +v if isinstance(n.op, ast.UAdd) else -v
            if isinstance(n, ast.BinOp) and isinstance(n.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
                a = _eval(n.left)
                b = _eval(n.right)
                if isinstance(n.op, ast.Add):
                    return a + b
                if isinstance(n.op, ast.Sub):
                    return a - b
                if isinstance(n.op, ast.Mult):
                    return a * b
                if isinstance(n.op, ast.Div):
                    return a / b
            raise ValueError(f'Unsupported expression: {expr!r}')

        return float(_eval(node))

    def _parse_standard_dist_spec(self, spec: str) -> tuple[str, float, float]:
        """Parse specs like 'gaussian(0,12/3)' or 'uniform(1.5,9)' into (dist,a,b)."""
        if spec is None or (isinstance(spec, float) and pd.isna(spec)):
            raise ValueError('Empty spec')
        s = str(spec).strip()
        if not s:
            raise ValueError('Empty spec')

        # Accept plain number as fixed
        compact = s.replace(' ', '')
        if re.fullmatch(r'[-+]?\d+(?:\.\d+)?', compact):
            return ('fixed', float(compact), 0.0)

        m = re.match(r'^\s*(gaussian|normal|uniform|gamma)\s*\(\s*(.+)\s*\)\s*$', s, re.IGNORECASE)
        if not m:
            raise ValueError(f'Unsupported distribution spec: {spec!r}')

        kind = m.group(1).lower()
        inner = m.group(2)
        if ',' not in inner:
            raise ValueError(f'Expected two parameters in spec: {spec!r}')
        left, right = inner.split(',', 1)

        a = self._safe_eval_numeric_expr(left)
        b = self._safe_eval_numeric_expr(right)

        # Note: gaussian specs are sampled as one-sided (truncated at zero)
        # when used for sigma/alpha parameters in MM_PSF presets. The GUI
        # continues to accept the familiar 'gaussian(mean,std)' syntax.
        if kind in {'gaussian', 'normal', 'gamma'}:
            return ('gaussian', a, abs(b))
        if kind == 'uniform':
            return ('uniform', a, b)

        raise ValueError(f'Unsupported distribution spec: {spec!r}')

    def _is_column_letter_spec(self, spec: str) -> bool:
        """Return True if spec looks like comma-separated Excel column letters."""
        if spec is None:
            return False
        s = str(spec).strip()
        if not s:
            return False
        # allow groups like 'W' or 'W,X,Y,AB' (letters only, commas, optional spaces)
        return bool(re.fullmatch(r"[A-Za-z]+(\s*,\s*[A-Za-z]+)*", s))

    def load_standard_alignment_presets(self, df: pd.DataFrame) -> None:
        """Load Alignment standard presets from the table starting at G1.

        Expected layout (0-indexed):
        - Column G (index 6): preset names (row 1..)
        - Columns H-K (indexes 7..10): variable specs (headers at row 0 like 'd_align_rad_')
        """
        self.alignment_standard_presets = {}

        try:
            # Determine header row by scanning the top few rows for expected variable names
            expected_bases = [p.split(' ')[0].strip() for p in DATA_TYPES['Alignment']['params']]
            header_row = None
            for r in range(min(5, df.shape[0])):
                row_vals = [str(x).strip() for x in df.iloc[r].tolist()]
                found = False
                for v in row_vals:
                    if not v:
                        continue
                    for base in expected_bases:
                        if base in v or v.startswith(base):
                            found = True
                            break
                    if found:
                        break
                if found:
                    header_row = r
                    break
            if header_row is None:
                header_row = 0

            # Historically presets lived starting at column G (index 6). New workbooks
            # may place preset names in M/N (indexes 12/13). Try common candidates
            # and fall back to 6 when none are populated. Look for non-empty name
            # cells below the detected header row.
            # Prefer column I (index 8) for standard preset names, then historical G (6), then M/N (12/13)
            candidate_name_cols = [8, 6, 12, 13]
            name_col = None
            for c in candidate_name_cols:
                if c < df.shape[1]:
                    col_vals = df.iloc[header_row + 1 : min(header_row + 21, df.shape[0]), c]
                    # Prefer columns with at least one non-empty, letter-containing value
                    if col_vals.dropna().astype(str).str.contains(r'[A-Za-z]').any():
                        name_col = c
                        break
            if name_col is None:
                # fallback: scan for any column with non-numeric entries below header
                for c in range(min(df.shape[1], 30)):
                    col_vals = df.iloc[header_row + 1 : min(header_row + 21, df.shape[0]), c]
                    if col_vals.dropna().astype(str).str.strip().apply(lambda s: not re.fullmatch(r"\s*\d+(?:\.\d+)?\s*", s)).any():
                        name_col = c
                        break
            if name_col is None:
                name_col = 6

            # variable specs historically start to the right of the name column
            first_var_col = name_col + 1

            if df.shape[0] <= header_row or df.shape[1] <= first_var_col:
                return

            headers: dict[int, str] = {}
            for c in range(first_var_col, min(df.shape[1], first_var_col + 8)):
                h = df.iloc[header_row, c]
                if pd.isna(h) or str(h).strip() == '':
                    continue
                headers[c] = str(h).strip()

            # Map base variable names to the GUI parameter labels
            param_map: dict[str, str] = {}
            for p in DATA_TYPES['Alignment']['params']:
                base = str(p).split(' ')[0].strip()
                param_map[base] = p

            row_idx = header_row + 1
            while row_idx < df.shape[0]:
                preset_name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
                if pd.isna(preset_name) or str(preset_name).strip() == '':
                    break
                # Avoid treating plain numeric values (e.g., 0) as preset names
                if isinstance(preset_name, (int, float)) and float(preset_name) == 0:
                    row_idx += 1
                    continue
                preset_name = str(preset_name).strip()

                preset_specs: dict[str, str] = {}
                for c, h in headers.items():
                    if c >= df.shape[1]:
                        continue
                    raw_spec = df.iloc[row_idx, c]
                    if pd.isna(raw_spec) or str(raw_spec).strip() == '':
                        continue
                    var = str(h).strip()
                    if var.endswith('_'):
                        var = var[:-1]
                    if var in param_map:
                        preset_specs[param_map[var]] = str(raw_spec).strip()

                if preset_specs:
                    self.alignment_standard_presets[preset_name] = preset_specs
                row_idx += 1

            if self.alignment_standard_presets:
                self.align_mode_var.set('standard')
                logging.debug("Loaded %d standard alignment presets: %s", len(self.alignment_standard_presets), list(self.alignment_standard_presets.keys()))
        except Exception as e:
            logging.exception("Error loading standard alignment presets: %s", e)
            self.alignment_standard_presets = {}
    
    def load_standard_thermal_presets(self, df: pd.DataFrame) -> None:
        """Load Thermal standard presets from the table starting at G1.

        Expected layout (0-indexed):
        - Column G (index 6): preset names
        - Subsequent columns: variable specs with headers at row 0 (e.g., 'd_therm_x_')
        """
        self.thermal_standard_presets = {}
        try:
            start_row = 0
            # Try historically common positions first, then scan broadly for a
            # column that contains non-numeric, non-empty preset names in rows 1..20.
            preferred = [8, 6, 12, 13, 9]
            name_col = None
            for c in preferred:
                if c < df.shape[1]:
                    col_vals = df.iloc[start_row + 1 : min(start_row + 21, df.shape[0]), c]
                    # Prefer columns with at least one non-empty cell containing letters
                    if col_vals.dropna().astype(str).str.contains(r'[A-Za-z]').any():
                        name_col = c
                        break
            if name_col is None:
                # Fallback: scan all columns for a likely name column
                for c in range(min(df.shape[1], 30)):
                    col_vals = df.iloc[start_row + 1 : min(start_row + 21, df.shape[0]), c]
                    if col_vals.dropna().astype(str).str.strip().apply(lambda s: not re.fullmatch(r"\s*\d+(?:\.\d+)?\s*", s)).any():
                        name_col = c
                        break
            if name_col is None:
                name_col = 6
            first_var_col = name_col + 1
            # If the header is embedded earlier/later, try to find it by looking
            # for expected variable base names in the top few rows.
            header_row = None
            expected_bases = [p.split(' ')[0].strip() for p in DATA_TYPES['Thermal']['params']]
            for r in range(min(5, df.shape[0])):
                row_vals = [str(x).strip() for x in df.iloc[r].tolist()]
                found = False
                for v in row_vals:
                    if not v:
                        continue
                    for base in expected_bases:
                        if base in v or v.startswith(base):
                            found = True
                            break
                    if found:
                        break
                if found:
                    header_row = r
                    break

            if header_row is None:
                header_row = start_row

            if df.shape[0] <= header_row:
                return

            headers = {}
            for c in range(first_var_col, min(df.shape[1], first_var_col + 16)):
                h = df.iloc[header_row, c]
                if pd.isna(h) or str(h).strip() == '':
                    continue
                headers[c] = str(h).strip()

            param_map = {}
            for p in DATA_TYPES['Thermal']['params']:
                base = str(p).split(' ')[0].strip()
                param_map[base] = p
            # Build normalized map to tolerate header typos (remove non-alnum, lowercase)
            def _norm(s: str) -> str:
                return re.sub(r'[^a-z0-9]', '', str(s).lower())
            param_map_norm = { _norm(k): v for k, v in param_map.items() }

            row_idx = header_row + 1
            while row_idx < df.shape[0]:
                preset_name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
                if pd.isna(preset_name) or str(preset_name).strip() == '':
                    break
                preset_name = str(preset_name).strip()

                preset_specs = {}
                for c, h in headers.items():
                    if c >= df.shape[1]:
                        continue
                    raw_spec = df.iloc[row_idx, c]
                    if pd.isna(raw_spec) or str(raw_spec).strip() == '':
                        continue
                    var = str(h).strip()
                    if var.endswith('_'):
                        var = var[:-1]
                    # normalize and attempt fuzzy match to param_map (tolerate typos)
                    var_norm = _norm(var)
                    matched = None
                    if var_norm in param_map_norm:
                        matched = param_map_norm[var_norm]
                    else:
                        # try substring matches (header may be abbreviated/typo)
                        for pk, pv in param_map_norm.items():
                            if pk in var_norm or var_norm in pk:
                                matched = pv
                                break
                    if matched:
                        preset_specs[matched] = str(raw_spec).strip()

                if preset_specs:
                    self.thermal_standard_presets[preset_name] = preset_specs
                row_idx += 1

            if self.thermal_standard_presets:
                self.thermal_mode_var.set('standard')
                logging.debug("Loaded %d standard thermal presets: %s", len(self.thermal_standard_presets), list(self.thermal_standard_presets.keys()))
        except Exception as e:
            logging.exception("Error loading standard thermal presets: %s", e)
            self.thermal_standard_presets = {}

    def load_standard_gravity_presets(self, df: pd.DataFrame) -> None:
        """Load Gravity offload standard presets from the table starting at G1.

        Expected layout (0-indexed):
        - Column G (index 6): preset names
        - Subsequent columns: variable specs with headers at row 0 (e.g., 'd_grav_x_')
        """
        self.gravity_standard_presets = {}
        try:
            start_row = 0
            preferred = [8, 6, 12, 13, 9]
            name_col = None
            for c in preferred:
                if c < df.shape[1]:
                    col_vals = df.iloc[start_row + 1 : min(start_row + 21, df.shape[0]), c]
                    # Prefer columns with at least one non-empty cell containing letters
                    if col_vals.dropna().astype(str).str.contains(r'[A-Za-z]').any():
                        name_col = c
                        break
            if name_col is None:
                for c in range(min(df.shape[1], 30)):
                    col_vals = df.iloc[start_row + 1 : min(start_row + 21, df.shape[0]), c]
                    if col_vals.dropna().astype(str).str.contains(r'[A-Za-z]').any():
                        name_col = c
                        break
            if name_col is None:
                name_col = 6

            first_var_col = name_col + 1

            # Detect header row containing variable headers (look for expected base names)
            header_row = None
            expected_bases = [p.split(' ')[0].strip() for p in DATA_TYPES['Gravity offload']['params']]
            for r in range(min(5, df.shape[0])):
                row_vals = [str(x).strip() for x in df.iloc[r].tolist()]
                found = False
                for v in row_vals:
                    if not v:
                        continue
                    for base in expected_bases:
                        if base in v or v.startswith(base):
                            found = True
                            break
                    if found:
                        break
                if found:
                    header_row = r
                    break

            if header_row is None:
                header_row = start_row

            if df.shape[0] <= header_row or df.shape[1] <= first_var_col:
                return

            headers = {}
            for c in range(first_var_col, min(df.shape[1], first_var_col + 16)):
                h = df.iloc[header_row, c]
                if pd.isna(h) or str(h).strip() == '':
                    continue
                headers[c] = str(h).strip()

            param_map = {}
            for p in DATA_TYPES['Gravity offload']['params']:
                base = str(p).split(' ')[0].strip()
                param_map[base] = p

            row_idx = header_row + 1
            while row_idx < df.shape[0]:
                preset_name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
                if pd.isna(preset_name) or str(preset_name).strip() == '':
                    break
                preset_name = str(preset_name).strip()

                preset_specs = {}
                for c, h in headers.items():
                    if c >= df.shape[1]:
                        continue
                    raw_spec = df.iloc[row_idx, c]
                    if pd.isna(raw_spec) or str(raw_spec).strip() == '':
                        continue
                    var = str(h).strip()
                    if var.endswith('_'):
                        var = var[:-1]
                    if var in param_map:
                        preset_specs[param_map[var]] = str(raw_spec).strip()

                if preset_specs:
                    self.gravity_standard_presets[preset_name] = preset_specs
                row_idx += 1

            if self.gravity_standard_presets:
                self.gravity_mode_var.set('standard')
                logging.debug("Loaded %d standard gravity presets: %s", len(self.gravity_standard_presets), list(self.gravity_standard_presets.keys()))
        except Exception as e:
            logging.exception("Error loading standard gravity presets: %s", e)
            self.gravity_standard_presets = {}

    def load_standard_aeff_presets(self, df: pd.DataFrame) -> None:
        """Load A_eff standard presets from the table starting at D1.

        Expected layout (0-indexed):
        - Column D (index 3): preset name (row 1..)
        - Column E (index 4): values expression (row 1..)
          Examples: 'J', 'L+gaussian(0,20%L)', 'J+gaussian(0,5%*J)'
        """
        self.aeff_standard_presets = {}
        try:
            # Try to robustly locate the 'Standard' / 'Values' columns instead of assuming fixed D/E.
            # Search the top few rows for header-like cells containing 'standard' and 'values'.
            start_row = 0
            name_col = None
            values_col = None

            # Search first 3 rows and all columns for header hints
            for r in range(min(3, df.shape[0])):
                for c in range(df.shape[1]):
                    cell = df.iloc[r, c]
                    try:
                        s = str(cell).strip().lower()
                    except Exception:
                        s = ''
                    if not s:
                        continue
                    if 'standard' in s and name_col is None:
                        name_col = c
                    if 'value' in s and values_col is None:
                        values_col = c

            # Fallback to original D/E if detection failed
            if name_col is None:
                name_col = 3
            if values_col is None:
                values_col = 4

            if df.shape[0] <= start_row + 1 or df.shape[1] <= values_col:
                return

            row_idx = start_row + 1
            while row_idx < df.shape[0]:
                name = df.iloc[row_idx, name_col] if name_col < df.shape[1] else None
                if pd.isna(name) or str(name).strip() == '':
                    row_idx += 1
                    continue
                expr = df.iloc[row_idx, values_col] if values_col < df.shape[1] else None
                if pd.isna(expr) or str(expr).strip() == '':
                    row_idx += 1
                    continue
                # Synthesize gaussian expressions for "Variable" presets that
                # name a percent but have a simple base column in the Values
                # cell (e.g. preset name "Variable 10% 1 keV" with Values "L").
                # Turn that into an additive gaussian around the column base
                # with sigma = pct * base, e.g. "gaussian(L,10%L)" so the
                # evaluator later interprets percent-relative sigma.
                preset_name = str(name).strip()
                preset_expr = str(expr).strip()
                try:
                    lname = preset_name.lower()
                    if (('variable' in lname or '%' in preset_name) and re.fullmatch(r"[A-Za-z]+", preset_expr.strip())):
                        m_pct = re.search(r"(\d+(?:\.\d+)?)\s*%", preset_name)
                        if m_pct:
                            pct = m_pct.group(1)
                            # produce gaussian(base, pct%base)
                            synthesized = f"gaussian({preset_expr},{pct}%{preset_expr})"
                            self.aeff_standard_presets[preset_name] = synthesized
                        else:
                            # no explicit percent found — store original expr
                            self.aeff_standard_presets[preset_name] = preset_expr
                    else:
                        self.aeff_standard_presets[preset_name] = preset_expr
                except Exception:
                    # Fallback: store raw string
                    self.aeff_standard_presets[str(name).strip()] = str(expr).strip()
                row_idx += 1

            if self.aeff_standard_presets:
                self.aeff_mode_var.set('standard')
                logging.debug("Loaded %d standard A_eff presets: %s", len(self.aeff_standard_presets), list(self.aeff_standard_presets.keys()))
        except Exception as e:
            logging.exception("Error loading standard A_eff presets: %s", e)
            self.aeff_standard_presets = {}
    
    def apply_data_type_selection(self):
        """Apply the data type selection and provide feedback."""
        # Check if at least one data type is selected
        selected = [key for key, var in self.data_type_checkboxes.items() if var.get()]
        aeff_selected = bool(self.aeff_checkbox_var.get())
        if not selected and not aeff_selected:
            messagebox.showwarning('No Selection', 'Please select at least one data type to modify.')
            return
        
        # Update tabs; suppress automatic 'Applied preset' modals during setup
        try:
            self.suppress_standard_apply_modals = True
            self.update_data_type_tabs()
        finally:
            self.suppress_standard_apply_modals = False
        
        # Provide feedback
        selected_names = [DATA_TYPES[key]['tab_label'] for key in selected]
        if aeff_selected:
            selected_names.append('A_eff')
        feedback_msg = f"✓ Selection applied: {', '.join(selected_names)}"
        self.selection_feedback.configure(text=feedback_msg)
        
        selected_list = ', '.join(selected_names)
        messagebox.showinfo('Success', f'Tabs created for: {selected_list}\n\nYou can now proceed to configure each data type.')
    
    def update_data_type_tabs(self):
        """Dynamically add or remove tabs based on selected data types."""
        # Get newly selected data types
        new_enabled = [key for key, var in self.data_type_checkboxes.items() if var.get()]
        aeff_enabled = bool(self.aeff_checkbox_var.get())
        
        # Remove tabs for deselected data types
        for data_type_key in self.enabled_data_types:
            if data_type_key not in new_enabled and data_type_key in self.data_type_tabs:
                # Find and remove the tab
                notebook = self.data_type_tabs[data_type_key]['notebook']
                for i in range(self.main_notebook.index('end')):
                    if self.main_notebook.tab(i, 'text') == DATA_TYPES[data_type_key]['tab_label']:
                        self.main_notebook.forget(i)
                        break
        
        # Add tabs for newly selected data types
        for data_type_key in new_enabled:
            if data_type_key not in self.enabled_data_types:
                # Create the tab if it doesn't exist
                if data_type_key not in self.data_type_tabs:
                    config = DATA_TYPES[data_type_key]
                    
                    # Create a notebook for this data type with 2 sub-tabs
                    type_notebook = ttk.Notebook(self.main_notebook)
                    
                    # Selection tab
                    selection_frame = ttk.Frame(type_notebook)
                    type_notebook.add(selection_frame, text='Select MMs')
                    
                    # Generation tab
                    generation_frame = ttk.Frame(type_notebook)
                    type_notebook.add(generation_frame, text='Generate')
                    
                    self.data_type_tabs[data_type_key] = {
                        'notebook': type_notebook,
                        'selection': selection_frame,
                        'generation': generation_frame
                    }
                    
                    # Build the tabs
                    self.build_selection_tab(data_type_key)
                    self.build_generation_tab(data_type_key)
                
                # Insert tab before the Preview/Export tab
                preview_index = None
                for i in range(self.main_notebook.index('end')):
                    if self.main_notebook.tab(i, 'text') == 'Preview / Export':
                        preview_index = i
                        break
                
                if preview_index is not None:
                    self.main_notebook.insert(preview_index, self.data_type_tabs[data_type_key]['notebook'], 
                                              text=DATA_TYPES[data_type_key]['tab_label'])
                else:
                    self.main_notebook.add(self.data_type_tabs[data_type_key]['notebook'], 
                                           text=DATA_TYPES[data_type_key]['tab_label'])
        
        # Ensure MM Configuration tab is present if any data type is selected
        if new_enabled or aeff_enabled:
            # Check if MM Configuration tab is already in the notebook
            config_tab_exists = False
            for i in range(self.main_notebook.index('end')):
                try:
                    if self.main_notebook.tab(i, 'text') == 'MM Configuration':
                        config_tab_exists = True
                        break
                except:
                    pass
            
            # Add MM Configuration tab after Load tab if not present
            if not config_tab_exists:
                # ttk.Notebook.insert(pos, ...) inserts *before* an existing index.
                # If only the Load tab exists, index 1 is out of bounds; in that case, just add.
                if self.main_notebook.index('end') >= 2:
                    self.main_notebook.insert(1, self.tab_config, text='MM Configuration')
                else:
                    self.main_notebook.add(self.tab_config, text='MM Configuration')

            # Insert/remove A_eff tab immediately after MM Configuration depending on selection
            aeff_tab_exists = False
            config_index = None
            for i in range(self.main_notebook.index('end')):
                try:
                    t = self.main_notebook.tab(i, 'text')
                    if t == 'MM Configuration':
                        config_index = i
                    if t == 'A_eff':
                        aeff_tab_exists = True
                except Exception:
                    pass
            if aeff_enabled:
                if not aeff_tab_exists and config_index is not None:
                    # Insert after MM Configuration when possible; otherwise add at end.
                    end_idx = self.main_notebook.index('end')
                    target_idx = config_index + 1
                    if 0 <= target_idx < end_idx:
                        self.main_notebook.insert(target_idx, self.tab_aeff, text='A_eff')
                    else:
                        self.main_notebook.add(self.tab_aeff, text='A_eff')
            else:
                if aeff_tab_exists:
                    for i in range(self.main_notebook.index('end')):
                        try:
                            if self.main_notebook.tab(i, 'text') == 'A_eff':
                                self.main_notebook.forget(i)
                                break
                        except Exception:
                            pass
        else:
            # If everything is deselected, remove MM Configuration and A_eff.
            for i in range(self.main_notebook.index('end')):
                try:
                    if self.main_notebook.tab(i, 'text') == 'MM Configuration':
                        self.main_notebook.forget(i)
                        break
                except Exception:
                    pass
            for i in range(self.main_notebook.index('end')):
                try:
                    if self.main_notebook.tab(i, 'text') == 'A_eff':
                        self.main_notebook.forget(i)
                        break
                except Exception:
                    pass
        
        # Ensure Preview/Export tab is always at the end
        try:
            for i in range(self.main_notebook.index('end')):
                if self.main_notebook.tab(i, 'text') == 'Preview / Export':
                    self.main_notebook.forget(i)
                    break
        except:
            pass
        self.main_notebook.add(self.tab_preview, text='Preview / Export')
        
        self.enabled_data_types = new_enabled

    def build_config_tab(self):
        """Build the MM Configuration tab UI.

        This tab displays the per-MM table (selection, filtering, sorting)
        and provides convenience buttons to select/deselect/toggle rows.
        """
        frame = self.tab_config
        ttk.Label(frame, text='MM Configuration - Select MMs', font=('Arial', 14)).pack(pady=10)
        
        control_frame = ttk.Frame(frame)
        control_frame.pack(fill='x', padx=5, pady=5)
        
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(side='left', padx=2)
        
        ttk.Button(button_frame, text='Select All', command=self.select_all_mms).pack(side='left', padx=2)
        ttk.Button(button_frame, text='Deselect All', command=self.deselect_all_mms).pack(side='left', padx=2)
        ttk.Button(button_frame, text='Toggle Selected', command=self.toggle_selected_rows).pack(side='left', padx=2)
        
        filter_frame = ttk.LabelFrame(control_frame, text='Filter by Row #:', padding=5)
        filter_frame.pack(side='left', padx=10)
        
        self.row_filter_var = tk.StringVar()
        row_filter_entry = ttk.Entry(filter_frame, textvariable=self.row_filter_var, width=20)
        row_filter_entry.pack(side='left', padx=5)
        row_filter_entry.bind('<KeyRelease>', lambda e: self.validate_and_apply_filter())
        
        self.row_filter_error = ttk.Label(filter_frame, text='', foreground='red', font=('Arial', 9))
        self.row_filter_error.pack(side='left', padx=5)
        
        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.tree = ttk.Treeview(tree_frame, height=25)
        self.tree['columns'] = ('select', 'pos', 'row', 'mm_num', 'x_mm', 'y_mm', 'z_mm', 'r_mm')
        
        self.tree.column('#0', width=0, stretch=tk.NO)
        self.tree.column('select', anchor='center', width=50)
        self.tree.column('pos', anchor='center', width=80)
        self.tree.column('row', anchor='center', width=60)
        self.tree.column('mm_num', anchor='center', width=70)
        self.tree.column('x_mm', anchor='center', width=100)
        self.tree.column('y_mm', anchor='center', width=100)
        self.tree.column('z_mm', anchor='center', width=100)
        self.tree.column('r_mm', anchor='center', width=100)
        
        self.tree.heading('#0', text='')
        self.tree.heading('select', text='Select', command=lambda: self.sort_tree('select'))
        self.tree.heading('pos', text='Position #', command=lambda: self.sort_tree('pos'))
        self.tree.heading('row', text='Row #', command=lambda: self.sort_tree('row'))
        self.tree.heading('mm_num', text='MM #', command=lambda: self.sort_tree('mm_num'))
        self.tree.heading('x_mm', text='x_MM [m]', command=lambda: self.sort_tree('x_mm'))
        self.tree.heading('y_mm', text='y_MM [m]', command=lambda: self.sort_tree('y_mm'))
        self.tree.heading('z_mm', text='z_MM [m]', command=lambda: self.sort_tree('z_mm'))
        self.tree.heading('r_mm', text='r_MM [m]', command=lambda: self.sort_tree('r_mm'))
        
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Trackpad-friendly bindings (macOS): some settings trigger events on release.
        # Allow toggling by clicking on the checkbox column or anywhere in a row.
        self.tree.bind('<Button-1>', self.on_tree_click)
        self.tree.bind('<space>', self.on_tree_space)
        
        self.mm_data = {}
        self.sort_column = None
        self.sort_reverse = False

    def on_tree_click(self, event):
        # Ignore clicks on headers/separators.
        region = self.tree.identify_region(event.x, event.y)
        if region not in ('cell', 'tree'):
            return

        item = self.tree.identify('item', event.x, event.y)
        if not item or item not in self.mm_data:
            return

        # Toggle selection for the clicked row (works when clicking anywhere in the row or on the checkbox)
        current_val = self.mm_data[item]['selected']
        self.mm_data[item]['selected'] = not current_val
        self.update_tree_display()
        self.update_selected_mms()

        # Prevent default Treeview selection behavior to avoid interference
        return 'break'

    def on_tree_space(self, event):
        # Spacebar toggles all highlighted rows (keyboard-friendly, trackpad-independent).
        selected_items = self.tree.selection()
        if not selected_items:
            return

        for item in selected_items:
            if item in self.mm_data:
                self.mm_data[item]['selected'] = not self.mm_data[item]['selected']

        self.update_tree_display()
        self.update_selected_mms()
        return

    def select_all_mms(self):
        for mm_num in self.mm_data:
            self.mm_data[mm_num]['selected'] = True
        self.update_tree_display()
        self.update_selected_mms()

    def deselect_all_mms(self):
        for mm_num in self.mm_data:
            self.mm_data[mm_num]['selected'] = False
        self.update_tree_display()
        self.update_selected_mms()

    def toggle_selected_rows(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo('No Selection', 'Please select one or more rows in the table first.')
            return
        
        for item in selected_items:
            if item in self.mm_data:
                self.mm_data[item]['selected'] = not self.mm_data[item]['selected']
        
        self.update_tree_display()
        self.update_selected_mms()

    def sort_tree(self, col):
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False
        self.update_tree_display()

    def validate_and_apply_filter(self):
        row_filter = self.row_filter_var.get().strip()
        
        if not row_filter:
            self.row_filter_error.configure(text='')
            self.apply_filter()
            return
        
        try:
            float(row_filter)
            self.row_filter_error.configure(text='')
            self.apply_filter()
        except ValueError:
            self.row_filter_error.configure(text='Error: Numbers only')
    
    def apply_filter(self):
        self.update_tree_display()

    def update_tree_display(self):
        """Refresh the visible Treeview contents from `self.mm_data`.

        Applies optional row filtering and sorting state then inserts rows
        with formatted numeric columns for display. Selection markers are
        rendered as simple unicode checkboxes for a compact UI.
        """
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        row_filter = self.row_filter_var.get().strip()
        
        items_list = []
        for mm_id, data in self.mm_data.items():
            if row_filter:
                try:
                    filter_value = int(float(row_filter))
                    if data['row'] != filter_value:
                        continue
                except ValueError:
                    pass
            items_list.append((mm_id, data))
        
        if self.sort_column:
            reverse = self.sort_reverse
            if self.sort_column == 'select':
                items_list.sort(key=lambda x: x[1]['selected'], reverse=reverse)
            elif self.sort_column == 'pos':
                items_list.sort(key=lambda x: int(x[1].get('pos', 0)), reverse=reverse)
            elif self.sort_column == 'row':
                items_list.sort(key=lambda x: int(x[1]['row']), reverse=reverse)
            elif self.sort_column == 'mm_num':
                items_list.sort(key=lambda x: int(x[1]['mm_num']), reverse=reverse)
            else:
                items_list.sort(key=lambda x: float(x[1].get(self.sort_column, 0)), reverse=reverse)
        
        for mm_id, data in items_list:
            select_icon = '✓' if data['selected'] else '☐'
            self.tree.insert('', 'end', mm_id, values=(
                select_icon,
                int(data.get('pos', 0)),
                data['row'],
                data['mm_num'],
                f"{float(data.get('x_mm', 0)):.6f}",
                f"{float(data.get('y_mm', 0)):.6f}",
                f"{float(data.get('z_mm', 0)):.6f}",
                f"{float(data.get('r_mm', 0)):.6f}"
            ))

    def update_config_display(self):
        """Populate the treeview data store from the loaded MM configuration.

        Converts workbook values to the minimal set of fields expected by the
        tree display (selected, pos, row, mm_num, x_mm, y_mm, z_mm, r_mm).
        Called after loading a workbook or when selections change.
        """
        self.mm_data.clear()
        self.selected_mm_numbers = []
        
        if self.mm_config_df is not None:
            for idx, row in self.mm_config_df.iterrows():
                mm_num_raw = row.get('MM #', idx + 1)
                try:
                    mm_num = int(float(mm_num_raw))
                except Exception:
                    mm_num = idx + 1
                mm_id = f"mm_{mm_num}"
                pos_val = row.get('Position #', idx + 1)
                self.mm_data[mm_id] = {
                    'selected': True,
                    'pos': int(float(pos_val)) if not pd.isna(pos_val) else (idx + 1),
                    'row': int(float(row.get('Row #', idx + 1))),
                    'mm_num': mm_num,
                    'x_mm': row.get('x_MM [m]', 0),
                    'y_mm': row.get('y_MM [m]', 0),
                    'z_mm': row.get('z_MM [m]', 0),
                    'r_mm': row.get('r_MM [m]', 0)
                }
            
            self.sort_column = None
            self.sort_reverse = False
            self.row_filter_var.set('')
            
            self.update_tree_display()
            self.update_selected_mms()

    def update_selected_mms(self):
        """Update `self.selected_mm_numbers` from the current tree selection.

        Iterates `self.mm_data` and builds a sorted list of MM numbers that are
        currently marked selected. This list is used by generation and export
        operations to decide which MMs to include.
        """
        selected = []
        for data in self.mm_data.values():
            if not data.get('selected'):
                continue
            try:
                selected.append(int(float(data.get('mm_num'))))
            except Exception:
                continue
        self.selected_mm_numbers = sorted(selected)

    def build_aeff_tab(self):
        """Build the A_eff tab controls.

        The tab allows selecting a standard A_eff preset or a fixed value and
        applying the chosen weights to selected MMs. When a standard preset is
        selected the UI exposes an optional checkbox `Apply vignetting factors
        when exporting` which, if enabled, triggers the export-time behavior
        that copies the chosen energy column from Vignetting sheets into column
        B before saving.
        """
        frame = self.tab_aeff
        ttk.Label(frame, text='A_eff - Apply weights to selected MMs', font=('Arial', 14)).pack(pady=10)

        container = ttk.Frame(frame)
        container.pack(fill='x', padx=10, pady=5)

        ttk.Label(container, text='Mode:', font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        mode_frame = ttk.Frame(container)
        mode_frame.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        ttk.Radiobutton(
            mode_frame,
            text='Standard Distribution',
            variable=self.aeff_mode_var,
            value='standard',
            command=self.toggle_aeff_mode,
        ).pack(side='left', padx=5)
        ttk.Radiobutton(
            mode_frame,
            text='Fixed Value',
            variable=self.aeff_mode_var,
            value='fixed',
            command=self.toggle_aeff_mode,
        ).pack(side='left', padx=5)

        # Standard preset controls
        self.aeff_std_row = ttk.Frame(container)
        self.aeff_std_row.grid(row=1, column=0, columnspan=3, sticky='w', padx=5, pady=5)
        ttk.Label(self.aeff_std_row, text='Standard:', font=('Arial', 10)).grid(row=0, column=0, sticky='w', padx=0)
        self.aeff_std_combo = ttk.Combobox(
            self.aeff_std_row,
            textvariable=self.aeff_selected_preset_var,
            values=[],
            width=35,
            state='readonly',
        )
        self.aeff_std_combo.grid(row=0, column=1, sticky='w', padx=5)
        self.aeff_std_combo.bind('<<ComboboxSelected>>', lambda e: self.on_aeff_standard_selected())
        self.aeff_expr_label = ttk.Label(self.aeff_std_row, text='', font=('Arial', 9), foreground='gray')
        self.aeff_expr_label.grid(row=1, column=1, sticky='w', padx=5, pady=(2, 0))

        # Fixed-value controls (shown when user selects Fixed mode)
        self.aeff_fixed_row = ttk.Frame(container)
        self.aeff_fixed_row.grid(row=2, column=0, columnspan=3, sticky='w', padx=5, pady=5)
        ttk.Label(self.aeff_fixed_row, text='Fixed A_eff:', font=('Arial', 10)).grid(row=0, column=0, sticky='w')
        self.aeff_fixed_entry = ttk.Entry(self.aeff_fixed_row, textvariable=self.aeff_fixed_var, width=20)
        self.aeff_fixed_entry.grid(row=0, column=1, sticky='w', padx=5)

        # Free-mode vignetting energy selector
        self.aeff_free_row = ttk.Frame(container)
        self.aeff_free_row.grid(row=3, column=0, columnspan=3, sticky='w', padx=5, pady=5)
        ttk.Label(self.aeff_free_row, text='Vignetting energy (keV):', font=('Arial', 10)).grid(row=0, column=0, sticky='w')
        self.aeff_free_energy_combo = ttk.Combobox(self.aeff_free_row, textvariable=self.aeff_free_energy_var, values=[], width=8, state='readonly')
        self.aeff_free_energy_combo.grid(row=0, column=1, sticky='w', padx=5)

        # Informational note: vignetting factors are applied automatically
        note_row = ttk.Frame(container)
        note_row.grid(row=4, column=0, columnspan=3, sticky='w', padx=5, pady=5)
        ttk.Label(note_row, text='Vignetting factors are applied automatically from A_eff selection', font=('Arial', 9), foreground='gray').pack(side='left')

        # Apply button for A_eff actions
        self.aeff_apply_row = ttk.Frame(container)
        self.aeff_apply_row.grid(row=5, column=0, columnspan=3, sticky='w', padx=5, pady=10)
        ttk.Button(self.aeff_apply_row, text='Apply to Selected MMs', command=self.apply_aeff_to_selected).pack()

        # Populate preset list if already loaded
        self.refresh_aeff_preset_controls()
        self.toggle_aeff_mode()

    def refresh_aeff_preset_controls(self):
        """Refresh the A_eff preset combobox values and select a sensible default.

        Prefers a preset named '1 keV' when present, otherwise selects the
        first available preset. Triggers an update to the displayed expression
        summary via `on_aeff_standard_selected`.
        """
        presets = list(self.aeff_standard_presets.keys()) if self.aeff_standard_presets else []
        try:
            self.aeff_std_combo['values'] = presets
        except Exception:
            pass
        # Default to the '1 keV' standard preset when available, otherwise the first one.
        if presets and not self.aeff_selected_preset_var.get():
            if '1 keV' in presets:
                self.aeff_selected_preset_var.set('1 keV')
            else:
                self.aeff_selected_preset_var.set(presets[0])
        try:
            self.on_aeff_standard_selected()
        except Exception:
            pass
        # Populate free-energy combobox with numeric energies extracted from standard presets
        try:
            energies_list = []
            # First try to read distinct numeric energy tokens from the vignetting sheets (column J)
            try:
                if self.excel_path:
                    import pandas as _pd
                    try:
                        xls = _pd.ExcelFile(self.excel_path, engine='openpyxl')
                        # collect any present vignette sheets from candidate lists
                        sheets_to_check = []
                        for name in list(VIG_ROT_AZI_CANDIDATES) + list(VIG_ROT_RAD_CANDIDATES):
                            if name in xls.sheet_names and name not in sheets_to_check:
                                sheets_to_check.append(name)
                        energy_set = set()
                        for s in sheets_to_check:
                            try:
                                vdf = _pd.read_excel(self.excel_path, sheet_name=s, engine='openpyxl', header=None)
                                if vdf is None or vdf.empty:
                                    continue
                                if vdf.shape[1] > 9:
                                    col = vdf.iloc[:, 9].dropna()
                                    for v in col.tolist():
                                        try:
                                            fv = float(v)
                                            energy_set.add(fv)
                                        except Exception:
                                            continue
                            except Exception:
                                continue
                        if energy_set:
                            energies_list = sorted(list(energy_set))
                            # format: drop .0 for integer energies
                            def _fmt(f):
                                try:
                                    if abs(f - int(f)) < 1e-9:
                                        return str(int(f))
                                except Exception:
                                    pass
                                return str(float(f))
                            self.aeff_free_energy_combo['values'] = [_fmt(e) for e in energies_list]
                        else:
                            raise RuntimeError('no energies in vignetting sheets')
                    finally:
                        try:
                            del xls
                        except Exception:
                            pass
            except Exception:
                # Fallback: parse preset names for numeric tokens
                energies = []
                for name in presets:
                    m = re.search(r"(\d+(?:\.\d*)?)\s*(?:keV)?", str(name), flags=re.IGNORECASE)
                    if m:
                        energies.append(m.group(1))
                # remove duplicates while preserving order
                seen = set()
                uniq = [e for e in energies if not (e in seen or seen.add(e))]
                self.aeff_free_energy_combo['values'] = uniq
        except Exception:
            pass

    def toggle_aeff_mode(self):
        mode = self.aeff_mode_var.get()
        if mode == 'fixed':
            self.aeff_std_row.grid_remove()
            self.aeff_fixed_row.grid()
            try:
                self.aeff_fixed_entry.state(['!disabled'])
            except Exception:
                pass
            # In fixed mode show the vignetting controls and the free-energy selector
            try:
                self.aeff_vig_row.grid()
            except Exception:
                pass
            try:
                self.aeff_free_row.grid()
            except Exception:
                pass
        else:
            self.aeff_fixed_row.grid_remove()
            self.aeff_std_row.grid()
            try:
                self.aeff_fixed_entry.state(['disabled'])
            except Exception:
                pass
            try:
                self.aeff_vig_row.grid()
            except Exception:
                pass
            try:
                self.aeff_free_row.grid_remove()
            except Exception:
                pass

    def on_aeff_standard_selected(self):
        """Update the small expression summary when a standard preset is chosen.

        Reads the chosen preset name from `self.aeff_selected_preset_var` and
        writes a human-friendly 'Values: ...' string to the UI label.
        """
        name = self.aeff_selected_preset_var.get().strip()
        expr = self.aeff_standard_presets.get(name, '')
        self.aeff_expr_label.configure(text=f'Values: {expr}')
        # Parse numeric energy from preset name (e.g. 'Variable 10% 1 keV')
        # and set the free-energy var so export will write this into vignetting
        # sheets' C2. This makes the selected-energy explicit when a preset
        # includes the energy token.
        try:
            # Prefer an explicit numeric token immediately followed by 'keV'.
            m = re.search(r"(\d+(?:\.\d*)?)\s*(?:keV)", name, flags=re.IGNORECASE)
            if not m:
                # Fallback: take the last numeric token in the name (handles "Variable 10% 1 keV")
                all_nums = re.findall(r"(\d+(?:\.\d*)?)", name)
                if all_nums:
                    sval = all_nums[-1]
                else:
                    sval = None
            else:
                sval = m.group(1)
                try:
                    fval = float(sval)
                    # Format as integer when appropriate
                    if abs(fval - int(fval)) < 1e-9:
                        display = str(int(fval))
                    else:
                        display = str(fval)
                except Exception:
                    display = sval
                try:
                    self.aeff_free_energy_var.set(display)
                except Exception:
                    pass
                # Store parsed numeric energy for downstream logic if needed
                try:
                    self._aeff_selected_energy = float(sval)
                except Exception:
                    self._aeff_selected_energy = None
        except Exception:
            pass

    def _safe_eval_expr_with_vars(self, expr: str, variables: dict[str, float]) -> float:
        """Safely evaluate a simple arithmetic expression with optional single-letter variables.

        Supports: + - * / ( ) numbers, percentages, and variable names present in `variables`.
        Also supports '20%L' shorthand (treated as 0.2*L).
        """
        if expr is None:
            raise ValueError('Empty expression')
        s = str(expr).strip()
        if not s:
            raise ValueError('Empty expression')

        # Normalize "20%L" -> "(20/100)*L" (and similar)
        s = re.sub(r'(\d+(?:\.\d+)?)\s*%\s*([A-Za-z])', r'(\1/100)*\2', s)
        # Convert remaining percentages: '110%' -> '(110/100)'
        s = re.sub(r'(\d+(?:\.\d+)?)\s*%', r'(\1/100)', s)

        # Allow only safe characters (letters for variables)
        if re.search(r'[^0-9A-Za-z\s\+\-\*\/\(\)\.]', s):
            raise ValueError(f'Unsupported characters in expression: {expr!r}')

        node = ast.parse(s, mode='eval')

        def _eval(n):
            if isinstance(n, ast.Expression):
                return _eval(n.body)
            if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
                return float(n.value)
            if isinstance(n, ast.Name):
                if n.id not in variables:
                    raise ValueError(f'Unknown variable: {n.id!r}')
                return float(variables[n.id])
            if isinstance(n, ast.UnaryOp) and isinstance(n.op, (ast.UAdd, ast.USub)):
                v = _eval(n.operand)
                return +v if isinstance(n.op, ast.UAdd) else -v
            if isinstance(n, ast.BinOp) and isinstance(n.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
                a = _eval(n.left)
                b = _eval(n.right)
                if isinstance(n.op, ast.Add):
                    return a + b
                if isinstance(n.op, ast.Sub):
                    return a - b
                if isinstance(n.op, ast.Mult):
                    return a * b
                if isinstance(n.op, ast.Div):
                    return a / b
            raise ValueError(f'Unsupported expression: {expr!r}')

        return float(_eval(node))

    def _get_aeff_row_for_mm(self, mm: int) -> int | None:
        """Return the row index in self.aeff_raw_df corresponding to MM#, or None."""
        if self.aeff_raw_df is None:
            return None
        # MM values in the sheet are 1-based (MM # = 1..N). Prefer a fast
        # path that checks the first column at index (mm-1). If that doesn't
        # match, fall back to scanning the column for the MM value.
        try:
            mm_int = int(float(mm))
        except Exception:
            return None

        try:
            nrows = self.aeff_raw_df.shape[0]
            if 1 <= mm_int <= nrows:
                v = self.aeff_raw_df.iloc[mm_int - 1, 0]
                if not pd.isna(v) and int(float(v)) == mm_int:
                    return mm_int - 1
        except Exception:
            pass

        # Fallback: search the first column for a matching MM value. Be robust
        # against header text or other non-numeric cells by catching parse
        # errors per-cell instead of letting a ValueError abort the whole
        # operation.
        try:
            col = self.aeff_raw_df.iloc[:, 0]

            def _is_mm_match(x):
                try:
                    if pd.isna(x):
                        return False
                    s = str(x).strip()
                    if s == '':
                        return False
                    return int(float(s)) == mm_int
                except Exception:
                    return False

            matches = col[col.apply(_is_mm_match)]
            if len(matches) == 0:
                return None
            return int(matches.index[0])
        except Exception:
            return None

    def _value_from_column_letter(self, row_idx: int, col_letter: str) -> float:
        """Return the numeric value from `self.aeff_raw_df` at the given row and Excel column letter.

        row_idx is a zero-based index into the dataframe. `col_letter` may be
        a single Excel column letter (e.g. 'A', 'J'). Raises ValueError when the
        sheet is not loaded or the requested cell is non-numeric or out of
        range.
        """
        if self.aeff_raw_df is None:
            raise ValueError('A_eff sheet not loaded')
        col_letter = str(col_letter).strip().upper()
        try:
            from openpyxl.utils import column_index_from_string

            col_idx = column_index_from_string(col_letter) - 1
        except Exception:
            raise ValueError(f'Invalid column letter: {col_letter!r}')

        if col_idx < 0 or col_idx >= self.aeff_raw_df.shape[1]:
            raise ValueError(f'Column {col_letter} not present in A_eff sheet')
        v = self.aeff_raw_df.iloc[row_idx, col_idx]
        try:
            return float(v)
        except Exception:
            # Check if the value is None/NaN and provide helpful error message
            if pd.isna(v) or v is None:
                raise ValueError(
                    f'Column {col_letter} contains no numeric data at row {row_idx + 1}. '
                    f'This column may contain formulas that have not been calculated. '
                    f'Please open the Excel file in Microsoft Excel, press Ctrl+Shift+F9 to recalculate all formulas, '
                    f'then save the file to cache the calculated values.'
                )
            else:
                raise ValueError(f'Non-numeric value at A_eff row {row_idx + 1}, col {col_letter}: {v!r}')

    def _evaluate_aeff_preset_for_mm(self, mm: int, values_expr: str) -> float:
        """Evaluate a `Values` expression from an A_eff preset for MM `mm`.

        Supported expressions include simple column-letter references like
        'J', additive forms like 'J+gaussian(0,20%J)', and gaussian(...) forms
        referencing a column. The function reads the base column value for the
        requested MM and evaluates the expression returning a float.
        """
        s = str(values_expr).strip()
        if not s:
            raise ValueError('Empty Values expression')

        row_idx = self._get_aeff_row_for_mm(mm)
        if row_idx is None:
            raise ValueError(f'MM #{mm} not found in A_eff sheet')

        # Simple: 'J'
        m = re.fullmatch(r'([A-Za-z]+)', s.replace(' ', ''))
        if m:
            base_col = m.group(1).upper()
            return self._value_from_column_letter(row_idx, base_col)

        # Direct gaussian around a column value: support both
        # 'gaussian(J, mean_expr, sigma_expr)' and the shorthand
        # 'gaussian(J, sigma_expr)' where mean=0.
        m = re.match(r'^\s*gaussian\s*\(\s*([A-Za-z]+)\s*,\s*(.+?)\s*,\s*(.+?)\s*\)\s*$', s, re.IGNORECASE)
        if m:
            base_col = m.group(1).upper()
            mean_expr = m.group(2)
            sigma_expr = m.group(3)
            base = self._value_from_column_letter(row_idx, base_col)
            vars_map = {base_col: base}
            mean = self._safe_eval_expr_with_vars(mean_expr, vars_map)
            sigma = abs(self._safe_eval_expr_with_vars(sigma_expr, vars_map))
            return float(base + np.random.normal(loc=mean, scale=sigma))

        m = re.match(r'^\s*gaussian\s*\(\s*([A-Za-z]+)\s*,\s*(.+?)\s*\)\s*$', s, re.IGNORECASE)
        if m:
            base_col = m.group(1).upper()
            sigma_expr = m.group(2)
            base = self._value_from_column_letter(row_idx, base_col)
            vars_map = {base_col: base}
            mean = 0.0
            sigma = abs(self._safe_eval_expr_with_vars(sigma_expr, vars_map))
            return float(base + np.random.normal(loc=mean, scale=sigma))

        # Additive Gaussian: 'J+gaussian(0,20%*J)' or 'L+gaussian(0,20%L)'
        m = re.match(r'^\s*([A-Za-z]+)\s*\+\s*gaussian\s*\(\s*(.+?)\s*,\s*(.+?)\s*\)\s*$', s, re.IGNORECASE)
        if m:
            base_col = m.group(1).upper()
            mean_expr = m.group(2)
            sigma_expr = m.group(3)

            base = self._value_from_column_letter(row_idx, base_col)
            vars_map = {base_col: base}

            mean = self._safe_eval_expr_with_vars(mean_expr, vars_map)
            sigma = abs(self._safe_eval_expr_with_vars(sigma_expr, vars_map))
            return float(base + np.random.normal(loc=mean, scale=sigma))

        raise ValueError(f'Unsupported Values expression: {values_expr!r}')

    def apply_aeff_to_selected(self):
        """Apply the currently selected A_eff mode (fixed or standard preset).

        For `fixed` mode the user-provided numeric value is applied to all
        selected MMs. For `standard` mode the preset's `Values` expression is
        evaluated per-MM using the A_eff sheet; failures are reported to the
        user while successful updates mark `self.aeff_pending_export` so the
        export operation can commit changes to disk.
        """
        if not self.selected_mm_numbers:
            messagebox.showwarning('No MMs selected', 'Select MMs in the "MM Configuration" tab first.')
            return

        mode = self.aeff_mode_var.get()
        updated = 0
        # Ensure result collectors are always defined so later summary
        # logic can reference them regardless of mode (fixed vs standard).
        successes = []
        failures = []
        failure_reasons = []

        try:
            if mode == 'fixed':
                raw = self.aeff_fixed_var.get().strip()
                if not raw:
                    raise ValueError('Please enter a fixed A_eff value')
                fixed = float(raw)
                for mm in self.selected_mm_numbers:
                    self.aeff_weights[int(mm)] = float(fixed)
                    updated += 1
            else:
                if not self.aeff_standard_presets:
                    raise ValueError('No standard A_eff presets found in the loaded file')
                preset = self.aeff_selected_preset_var.get().strip()
                if not preset:
                    raise ValueError('Please select a standard preset')
                expr = self.aeff_standard_presets.get(preset)
                if not expr:
                    raise ValueError('Selected preset has no Values expression')

                # Ask for confirmation before applying to selected MMs
                msg = f"Apply standard A_eff preset '{preset}' to {len(self.selected_mm_numbers)} selected MMs?"
                if not messagebox.askyesno('Confirm A_eff Apply', msg):
                    return

                successes = []
                failures = []
                failure_reasons = []
                for mm in self.selected_mm_numbers:
                    try:
                        v = self._evaluate_aeff_preset_for_mm(int(mm), expr)
                        self.aeff_weights[int(mm)] = float(v)
                        updated += 1
                        successes.append(int(mm))
                    except Exception as e:
                        failures.append(mm)
                        failure_reasons.append(f"MM {mm}: {e}")

            self.aeff_pending_export = True
            # Provide a modal confirmation summary; do NOT show persistent in-window blue status text
            if updated > 0:
                msg = f'Updated A_eff for {updated} MMs (pending export).'
                if failures:
                    msg += f"\n{len(failures)} MMs could not be updated (missing A_eff row or parse error)."
                messagebox.showinfo('A_eff Applied', msg)
            else:
                if failures:
                    # Provide more detailed feedback about why updates failed.
                    sample = '\n'.join(failure_reasons[:6])
                    detail_msg = f'No MMs could be updated: missing A_eff rows or invalid preset.\n\nExamples:\n{sample}'
                    messagebox.showwarning('A_eff Not Applied', detail_msg)
                else:
                    messagebox.showinfo('A_eff', 'No MMs were selected.')
            # Clear the in-window A_eff status label if present
            try:
                self.aeff_status_label.configure(text='')
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror('A_eff error', str(e))
            return

    def build_selection_tab(self, data_type_key):
        """Build the MM selection tab for a specific data type."""
        frame = self.data_type_tabs[data_type_key]['selection']
        config = DATA_TYPES[data_type_key]
        
        ttk.Label(frame, text=f'Select MMs for {config["tab_label"]}', font=('Arial', 12)).pack(pady=10)
        ttk.Label(frame, text='Use the "MM Configuration" tab to select which MMs to modify.', 
                  font=('Arial', 10)).pack(pady=5)
        ttk.Label(frame, text=f'Then use the "Generate" tab to create {config["tab_label"]} data.', 
                  font=('Arial', 10)).pack(pady=5)

    def build_generation_tab(self, data_type_key):
        """Build the generation tab for a specific data type."""
        frame = self.data_type_tabs[data_type_key]['generation']
        config = DATA_TYPES[data_type_key]
        
        ttk.Label(frame, text=f'Generate {config["tab_label"]} Data', font=('Arial', 14)).grid(row=0, column=0, columnspan=7, pady=10)
        
        dist_entries = {}
        param_labels = {}
        row_start = 2

        # Add standard/free mode selection for Alignment if presets are available
        if data_type_key == 'Alignment' and self.alignment_standard_presets:
            ttk.Label(frame, text='Mode:', font=('Arial', 10, 'bold')).grid(row=1, column=0, padx=5, pady=10, sticky='w')

            mode_frame = ttk.Frame(frame)
            mode_frame.grid(row=1, column=1, columnspan=3, padx=5, pady=10, sticky='w')

            ttk.Radiobutton(
                mode_frame,
                text='Standard Distribution',
                variable=self.align_mode_var,
                value='standard',
                command=lambda: self.toggle_alignment_mode(data_type_key),
            ).pack(side='left', padx=5)
            ttk.Radiobutton(
                mode_frame,
                text='Free (Custom)',
                variable=self.align_mode_var,
                value='free',
                command=lambda: self.toggle_alignment_mode(data_type_key),
            ).pack(side='left', padx=5)

            ttk.Label(frame, text='Standard:', font=('Arial', 10)).grid(row=1, column=4, padx=5, pady=10, sticky='w')
            std_values = list(self.alignment_standard_presets.keys())
            std_combo = ttk.Combobox(frame, values=std_values, width=25, state='readonly')
            if std_values:
                std_combo.set(std_values[0])
            std_combo.grid(row=1, column=5, columnspan=2, padx=5, pady=10, sticky='w')
            std_combo.bind('<<ComboboxSelected>>', lambda e: (setattr(self, 'suppress_standard_apply_modals', True), self.on_alignment_standard_selected(), setattr(self, 'suppress_standard_apply_modals', False)))

            if data_type_key not in self.distribution_widgets:
                self.distribution_widgets[data_type_key] = {}
            self.distribution_widgets[data_type_key]['align_std_combo'] = std_combo

            row_start = 2
        
        # Add standard/free mode selection for Thermal if presets are available
        if data_type_key == 'Thermal' and self.thermal_standard_presets:
            ttk.Label(frame, text='Mode:', font=('Arial', 10, 'bold')).grid(row=1, column=0, padx=5, pady=10, sticky='w')
            mode_frame = ttk.Frame(frame)
            mode_frame.grid(row=1, column=1, columnspan=3, padx=5, pady=10, sticky='w')
            ttk.Radiobutton(mode_frame, text='Standard Distribution', variable=self.thermal_mode_var, value='standard', command=lambda: self.toggle_thermal_mode(data_type_key)).pack(side='left', padx=5)
            ttk.Radiobutton(mode_frame, text='Free (Custom)', variable=self.thermal_mode_var, value='free', command=lambda: self.toggle_thermal_mode(data_type_key)).pack(side='left', padx=5)
            ttk.Label(frame, text='Standard:', font=('Arial', 10)).grid(row=1, column=4, padx=5, pady=10, sticky='w')
            std_values = list(self.thermal_standard_presets.keys())
            std_combo = ttk.Combobox(frame, values=std_values, width=25, state='readonly')
            # Default to '0 deg FMS tilt' when available
            if std_values:
                if '0 deg FMS tilt' in std_values:
                    std_combo.set('0 deg FMS tilt')
                else:
                    std_combo.set(std_values[0])
            std_combo.grid(row=1, column=5, columnspan=2, padx=5, pady=10, sticky='w')
            std_combo.bind('<<ComboboxSelected>>', lambda e: (setattr(self, 'suppress_standard_apply_modals', True), self.on_thermal_standard_selected(), setattr(self, 'suppress_standard_apply_modals', False)))
            # Ensure UI reflects current mode immediately (disable lower controls if standard)
            try:
                self.toggle_thermal_mode(data_type_key)
            except Exception:
                pass
            if data_type_key not in self.distribution_widgets:
                self.distribution_widgets[data_type_key] = {}
            self.distribution_widgets[data_type_key]['therm_std_combo'] = std_combo
            row_start = 2

        # Add standard/free mode selection for Gravity offload if presets are available
        if data_type_key == 'Gravity offload' and self.gravity_standard_presets:
            ttk.Label(frame, text='Mode:', font=('Arial', 10, 'bold')).grid(row=1, column=0, padx=5, pady=10, sticky='w')
            mode_frame = ttk.Frame(frame)
            mode_frame.grid(row=1, column=1, columnspan=3, padx=5, pady=10, sticky='w')
            ttk.Radiobutton(mode_frame, text='Standard Distribution', variable=self.gravity_mode_var, value='standard', command=lambda: self.toggle_gravity_mode(data_type_key)).pack(side='left', padx=5)
            ttk.Radiobutton(mode_frame, text='Free (Custom)', variable=self.gravity_mode_var, value='free', command=lambda: self.toggle_gravity_mode(data_type_key)).pack(side='left', padx=5)
            ttk.Label(frame, text='Standard:', font=('Arial', 10)).grid(row=1, column=4, padx=5, pady=10, sticky='w')
            std_values = list(self.gravity_standard_presets.keys())
            std_combo = ttk.Combobox(frame, values=std_values, width=25, state='readonly')
            if std_values:
                # Default to 'GZ' preset when available
                if 'GZ' in std_values:
                    std_combo.set('GZ')
                else:
                    std_combo.set(std_values[0])
            std_combo.grid(row=1, column=5, columnspan=2, padx=5, pady=10, sticky='w')
            std_combo.bind('<<ComboboxSelected>>', lambda e: (setattr(self, 'suppress_standard_apply_modals', True), self.on_gravity_standard_selected(), setattr(self, 'suppress_standard_apply_modals', False)))
            # Ensure UI reflects current mode immediately (disable lower controls if standard)
            try:
                self.toggle_gravity_mode(data_type_key)
            except Exception:
                pass
            if data_type_key not in self.distribution_widgets:
                self.distribution_widgets[data_type_key] = {}
            self.distribution_widgets[data_type_key]['grav_std_combo'] = std_combo
            row_start = 2
        
        # Add standard/free mode selection for MM_PSF if standard distributions are available
        if config.get('has_distribution', False) and data_type_key == 'MM_PSF' and self.standard_distributions:
            ttk.Label(frame, text='Mode:', font=('Arial', 10, 'bold')).grid(row=1, column=0, padx=5, pady=10, sticky='w')
            
            mode_frame = ttk.Frame(frame)
            mode_frame.grid(row=1, column=1, columnspan=3, padx=5, pady=10, sticky='w')
            
            ttk.Radiobutton(mode_frame, text='Standard Distribution', variable=self.psf_mode_var, 
                           value='standard', command=lambda: self.toggle_psf_mode(data_type_key)).pack(side='left', padx=5)
            ttk.Radiobutton(mode_frame, text='Free (Custom)', variable=self.psf_mode_var, 
                           value='free', command=lambda: self.toggle_psf_mode(data_type_key)).pack(side='left', padx=5)
            
            # Standard distribution dropdown
            ttk.Label(frame, text='Standard:', font=('Arial', 10)).grid(row=1, column=4, padx=5, pady=10, sticky='w')
            std_values = list(self.standard_distributions.keys()) + [self.CUSTOM_PSF_OPTION]
            std_dist_combo = ttk.Combobox(frame, values=std_values, 
                                         width=25, state='readonly')
            if self.standard_distributions:
                std_dist_combo.set(list(self.standard_distributions.keys())[0])
            std_dist_combo.grid(row=1, column=5, columnspan=2, padx=5, pady=10, sticky='w')
            std_dist_combo.bind('<<ComboboxSelected>>', lambda e: (setattr(self, 'suppress_standard_apply_modals', True), self.on_mm_psf_standard_selected(data_type_key), setattr(self, 'suppress_standard_apply_modals', False)))

            btn_pick = ttk.Button(frame, text='Choose PSF file...', command=self.choose_custom_psf_file)
            btn_pick.grid(row=1, column=7, padx=5, pady=10, sticky='w')
            btn_pick.grid_remove()  # Only visible for Custom PSF selection

            lbl_pick = ttk.Label(frame, textvariable=self.custom_psf_path_var, font=('Arial', 8), foreground='gray')
            lbl_pick.grid(row=2, column=5, columnspan=3, padx=5, pady=0, sticky='w')
            lbl_pick.grid_remove()  # Only visible for Custom PSF selection
            
            if data_type_key not in self.distribution_widgets:
                self.distribution_widgets[data_type_key] = {}
            self.distribution_widgets[data_type_key]['std_dist_combo'] = std_dist_combo
            self.distribution_widgets[data_type_key]['btn_pick_psf'] = btn_pick
            self.distribution_widgets[data_type_key]['lbl_pick_psf'] = lbl_pick
            
            row_start = 2
        
        # Add distribution type selection for MM_PSF
        if config.get('has_distribution', False):
            ttk.Label(frame, text='Distribution Type:', font=('Arial', 10, 'bold')).grid(row=row_start, column=0, padx=5, pady=10, sticky='w')
            dist_type_combo = ttk.Combobox(frame, values=['gaussian', 'pseudo-voigt'], width=15, state='readonly')
            dist_type_combo.set('gaussian')
            dist_type_combo.grid(row=row_start, column=1, columnspan=2, padx=5, pady=10, sticky='w')
            
            # Bind event to enable/disable alpha controls
            dist_type_combo.bind('<<ComboboxSelected>>', lambda e: self.toggle_eta_entry(data_type_key))
            
            row_start = row_start + 1
        
        # Add regular parameters first
        for i, param in enumerate(config['params']):
            row = row_start + i
            ttk.Label(frame, text=param).grid(row=row, column=0, padx=5, pady=5)
            
            dist_box = ttk.Combobox(frame, values=['fixed', 'gaussian', 'gamma', 'uniform'], width=12, state='readonly')
            dist_box.set('fixed')
            dist_box.grid(row=row, column=1)
            dist_box.bind('<<ComboboxSelected>>', lambda e, dt=data_type_key, p=param: self.update_param_labels(dt, p))
            
            label_a = ttk.Label(frame, text='Value:')
            label_a.grid(row=row, column=2, padx=5, pady=5)
            a_entry = ttk.Entry(frame, width=12)
            a_entry.grid(row=row, column=3, padx=5)
            
            label_b = ttk.Label(frame, text='')
            label_b.grid(row=row, column=4, padx=5, pady=5)
            label_b.grid_remove()  # Hide initially since 'fixed' is default
            b_entry = ttk.Entry(frame, width=12, state='disabled')
            b_entry.grid(row=row, column=5, padx=5)
            b_entry.grid_remove()  # Hide initially since 'fixed' is default
            
            # Set default values
            default_mean, default_sigma = config['defaults'].get(param, (0, 1))
            a_entry.insert(0, str(default_mean))
            b_entry.insert(0, str(default_sigma))
            
            dist_entries[param] = (dist_box, a_entry, b_entry)
            param_labels[param] = (label_a, label_b)
        
        # Add alpha parameters for pseudo-voigt (initially hidden)
        alpha_entries = {}
        alpha_labels = {}
        if config.get('has_distribution', False) and 'alpha_params' in config:
            alpha_row_start = row_start + len(config['params'])
            for i, param in enumerate(config['alpha_params']):
                row = alpha_row_start + i
                param_label = ttk.Label(frame, text=f'{param} [0-1]')
                param_label.grid(row=row, column=0, padx=5, pady=5)
                
                dist_box = ttk.Combobox(frame, values=['fixed', 'gaussian', 'gamma', 'uniform'], width=12, state='readonly')
                dist_box.set('fixed')
                dist_box.grid(row=row, column=1)
                dist_box.bind('<<ComboboxSelected>>', lambda e, dt=data_type_key, p=param: self.update_alpha_param_labels(dt, p))
                
                label_a = ttk.Label(frame, text='Value:')
                label_a.grid(row=row, column=2, padx=5, pady=5)
                a_entry = ttk.Entry(frame, width=12)
                a_entry.grid(row=row, column=3, padx=5)
                
                label_b = ttk.Label(frame, text='')
                label_b.grid(row=row, column=4, padx=5, pady=5)
                label_b.grid_remove()
                b_entry = ttk.Entry(frame, width=12, state='disabled')
                b_entry.grid(row=row, column=5, padx=5)
                b_entry.grid_remove()
                
                description_label = ttk.Label(frame, text='(0=Gaussian, 1=Lorentzian)', font=('Arial', 8), foreground='gray')
                description_label.grid(row=row, column=6, padx=5, sticky='w')
                
                # Set default values
                default_mean, default_sigma = config['defaults'].get(param, (0.5, 0.1))
                a_entry.insert(0, str(default_mean))
                b_entry.insert(0, str(default_sigma))
                
                alpha_entries[param] = (param_label, dist_box, a_entry, b_entry, description_label)
                alpha_labels[param] = (label_a, label_b)
                
                # Hide alpha parameters initially (will be shown when pseudo-voigt is selected)
                param_label.grid_remove()
                dist_box.grid_remove()
                label_a.grid_remove()
                a_entry.grid_remove()
                label_b.grid_remove()
                b_entry.grid_remove()
                description_label.grid_remove()
            
            self.alpha_entries_by_type[data_type_key] = alpha_entries
            if data_type_key not in self.param_labels_by_type:
                self.param_labels_by_type[data_type_key] = {}
            self.param_labels_by_type[data_type_key].update(alpha_labels)
            
            # Store reference to distribution type combo for toggling alpha parameters
            if data_type_key not in self.distribution_widgets:
                self.distribution_widgets[data_type_key] = {}
            self.distribution_widgets[data_type_key]['dist_type'] = dist_type_combo
        
        self.dist_entries_by_type[data_type_key] = dist_entries
        if data_type_key not in self.param_labels_by_type:
            self.param_labels_by_type[data_type_key] = {}
        self.param_labels_by_type[data_type_key].update(param_labels)
        # For Thermal and Gravity, enforce mode UI after parameter widgets exist
        # Force the mode variables to 'standard' so manual controls remain
        # disabled until the user explicitly switches to 'free'. Call the
        # toggle functions to apply the UI state immediately.
        try:
            if data_type_key == 'Thermal':
                try:
                    self.thermal_mode_var.set('standard')
                except Exception:
                    pass
                try:
                    self.toggle_thermal_mode('Thermal')
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if data_type_key == 'Gravity offload':
                try:
                    self.gravity_mode_var.set('standard')
                except Exception:
                    pass
                try:
                    self.toggle_gravity_mode('Gravity offload')
                except Exception:
                    pass
        except Exception:
            pass
        
        # Initialize mode (enable/disable controls based on default mode)
        if config.get('has_distribution', False) and data_type_key == 'MM_PSF' and self.standard_distributions:
            self.toggle_psf_mode(data_type_key)

        if data_type_key == 'Alignment' and self.alignment_standard_presets:
            self.toggle_alignment_mode(data_type_key)
        
        # Position button after alpha rows if they exist, otherwise after last parameter
        button_row = (row_start + len(config['params']) + 
                     len(config.get('alpha_params', [])) + 1 
                     if config.get('has_distribution', False) and 'alpha_params' in config 
                     else row + 1)
        gen_btn = ttk.Button(frame, text=f'Generate {config["tab_label"]} Data', 
                   command=lambda dt=data_type_key: self.generate_data(dt))
        gen_btn.grid(row=button_row, column=0, columnspan=5, pady=20, sticky='w')
        # Status label to show pending export after applying/generating
        status_label = ttk.Label(frame, text='', foreground='blue')
        status_label.grid(row=button_row, column=5, columnspan=2, pady=20, sticky='w')
        try:
            self.tab_status_labels[data_type_key] = status_label
        except Exception:
            pass

    def _set_entry_text(self, entry: ttk.Entry, value: str) -> None:
        """Set entry text even if disabled."""
        try:
            prior_state = str(entry.cget('state'))
        except Exception:
            prior_state = 'normal'
        try:
            entry.config(state='normal')
        except Exception:
            pass
        try:
            entry.delete(0, tk.END)
            entry.insert(0, value)
        finally:
            try:
                entry.config(state=prior_state)
            except Exception:
                pass

    def on_alignment_standard_selected(self) -> None:
        """Apply selected Alignment standard preset to the per-parameter distribution controls."""
        data_type_key = 'Alignment'
        if data_type_key not in self.distribution_widgets or 'align_std_combo' not in self.distribution_widgets[data_type_key]:
            return
        preset_name = self.distribution_widgets[data_type_key]['align_std_combo'].get()
        preset = self.alignment_standard_presets.get(preset_name)
        if not preset:
            return

        # Temporarily enable free controls so we can populate them
        self._set_alignment_free_controls_state(state='normal')
        is_variable_preset = bool(re.search(r"\bVariable\b", preset_name, re.IGNORECASE))

        # If preset name starts with 'Fixed', force controls to fixed/disabled
        is_fixed_preset = bool(re.match(r'^\s*fixed', str(preset_name), re.IGNORECASE))

        for param_label, spec in preset.items():
            if data_type_key not in self.dist_entries_by_type or param_label not in self.dist_entries_by_type[data_type_key]:
                continue
            dist_box, a_entry, b_entry = self.dist_entries_by_type[data_type_key][param_label]
            try:
                dist, a, b = self._parse_standard_dist_spec(spec)
            except Exception:
                continue

            try:
                # If this is a Variable preset but the spec is gaussian, present the mean/sigma
                # as-is (gaussian) so the GUI shows Mean/Sigma from the preset table.
                if is_variable_preset and dist in ('gaussian', 'normal'):
                    dist_box.set('gaussian')
                else:
                    dist_box.set(dist)
            except Exception:
                pass
            # Insert numeric text into the entries (preserve text even when disabled)
            self._set_entry_text(a_entry, f"{a}")
            self._set_entry_text(b_entry, f"{b}")

            try:
                self.update_param_labels(data_type_key, param_label)
            except Exception:
                pass

            # If we're in standard mode, disable the controls for this parameter
            try:
                if self.align_mode_var.get() == 'standard' or is_fixed_preset:
                    try:
                        dist_box.config(state='disabled')
                    except Exception:
                        pass
                    try:
                        a_entry.config(state='disabled')
                    except Exception:
                        pass
                    try:
                        if dist in ('gaussian', 'normal', 'gamma'):
                            b_entry.config(state='disabled')
                    except Exception:
                        pass
            except Exception:
                pass

        # If we're in standard mode, disable again after applying
        if self.align_mode_var.get() == 'standard':
            self._set_alignment_free_controls_state(state='disabled')
            # Inform user via modal unless we're suppressing standard-apply modals
            try:
                if not getattr(self, 'suppress_standard_apply_modals', False):
                    messagebox.showinfo('Alignment Applied', f"Applied preset '{preset_name}' for Alignment (pending export).")
            except Exception:
                pass
            try:
                if 'Alignment' in self.tab_status_labels:
                    self.tab_status_labels['Alignment'].configure(text='')
            except Exception:
                pass

    def on_thermal_standard_selected(self) -> None:
        """Apply selected Thermal standard preset to per-parameter controls."""
        data_type_key = 'Thermal'
        if data_type_key not in self.distribution_widgets or 'therm_std_combo' not in self.distribution_widgets[data_type_key]:
            return
        preset_name = self.distribution_widgets[data_type_key]['therm_std_combo'].get()
        preset = self.thermal_standard_presets.get(preset_name)
        if not preset:
            return

        self._set_thermal_free_controls_state(state='normal')

        for param_label, spec in preset.items():
            if data_type_key not in self.dist_entries_by_type or param_label not in self.dist_entries_by_type[data_type_key]:
                continue
            dist_box, a_entry, b_entry = self.dist_entries_by_type[data_type_key][param_label]
            # If spec is a column-letter mapping (e.g. 'W' or 'W,X,Y,AB'), preserve it
            if self._is_column_letter_spec(spec):
                try:
                    dist_box.set('fixed')
                except Exception:
                    pass
                # Show the raw spec so the user can see the mapping
                self._set_entry_text(a_entry, str(spec).strip())
                self._set_entry_text(b_entry, '0')
                # Preserve mapping for export/processing
                self.distribution_widgets.setdefault(data_type_key, {}).setdefault('col_mappings', {})[param_label] = str(spec).strip()
                try:
                    self.update_param_labels(data_type_key, param_label)
                except Exception:
                    pass
                continue

            try:
                dist, a, b = self._parse_standard_dist_spec(spec)
            except Exception:
                continue

            try:
                dist_box.set(dist)
            except Exception:
                pass
            self._set_entry_text(a_entry, f"{a}")
            self._set_entry_text(b_entry, f"{b}")

            try:
                self.update_param_labels(data_type_key, param_label)
            except Exception:
                pass

        # If preset name starts with 'Fixed', treat all controls as fixed and disabled
        is_fixed_preset = bool(re.match(r'^\s*fixed', str(preset_name), re.IGNORECASE))

        if self.thermal_mode_var.get() == 'standard' or is_fixed_preset:
            self._set_thermal_free_controls_state(state='disabled')
        try:
            if not getattr(self, 'suppress_standard_apply_modals', False):
                messagebox.showinfo('Thermal Applied', f"Applied preset '{preset_name}' for Thermal (pending export).")
        except Exception:
            pass
        try:
            if 'Thermal' in self.tab_status_labels:
                self.tab_status_labels['Thermal'].configure(text='')
        except Exception:
            pass

    def _set_thermal_free_controls_state(self, state: str) -> None:
        data_type_key = 'Thermal'
        if data_type_key not in self.dist_entries_by_type:
            return
        for _, widgets in self.dist_entries_by_type[data_type_key].items():
            dist_box, a_entry, b_entry = widgets
            try:
                dist_box.config(state=state)
            except Exception:
                pass
            try:
                a_entry.config(state=state)
            except Exception:
                pass
            try:
                b_entry.config(state=state)
            except Exception:
                pass

    def toggle_thermal_mode(self, data_type_key: str) -> None:
        if data_type_key != 'Thermal':
            return
        mode = self.thermal_mode_var.get()
        std_combo = None
        if data_type_key in self.distribution_widgets:
            std_combo = self.distribution_widgets[data_type_key].get('therm_std_combo')
        if std_combo is not None:
            try:
                std_combo.config(state='readonly' if mode == 'standard' else 'disabled')
            except Exception:
                pass
        if mode == 'standard':
            try:
                self.suppress_standard_apply_modals = True
                self.on_thermal_standard_selected()
            finally:
                self.suppress_standard_apply_modals = False
            self._set_thermal_free_controls_state(state='disabled')
        else:
            self._set_thermal_free_controls_state(state='normal')

    def on_gravity_standard_selected(self) -> None:
        data_type_key = 'Gravity offload'
        if data_type_key not in self.distribution_widgets or 'grav_std_combo' not in self.distribution_widgets[data_type_key]:
            return
        preset_name = self.distribution_widgets[data_type_key]['grav_std_combo'].get()
        preset = self.gravity_standard_presets.get(preset_name)
        if not preset:
            return

        self._set_gravity_free_controls_state(state='normal')

        for param_label, spec in preset.items():
            if data_type_key not in self.dist_entries_by_type or param_label not in self.dist_entries_by_type[data_type_key]:
                continue
            dist_box, a_entry, b_entry = self.dist_entries_by_type[data_type_key][param_label]
            # If spec looks like an Excel column-letter mapping, preserve it and show it
            if self._is_column_letter_spec(spec):
                try:
                    dist_box.set('fixed')
                except Exception:
                    pass
                self._set_entry_text(a_entry, str(spec).strip())
                self._set_entry_text(b_entry, '0')
                self.distribution_widgets.setdefault(data_type_key, {}).setdefault('col_mappings', {})[param_label] = str(spec).strip()
                try:
                    self.update_param_labels(data_type_key, param_label)
                except Exception:
                    pass
                continue

            try:
                dist, a, b = self._parse_standard_dist_spec(spec)
            except Exception:
                continue

            try:
                dist_box.set(dist)
            except Exception:
                pass
            self._set_entry_text(a_entry, f"{a}")
            self._set_entry_text(b_entry, f"{b}")

            try:
                self.update_param_labels(data_type_key, param_label)
            except Exception:
                pass

        # If preset name starts with 'Fixed', treat all controls as fixed and disabled
        is_fixed_preset = bool(re.match(r'^\s*fixed', str(preset_name), re.IGNORECASE))

        if self.gravity_mode_var.get() == 'standard' or is_fixed_preset:
            self._set_gravity_free_controls_state(state='disabled')
        try:
            if not getattr(self, 'suppress_standard_apply_modals', False):
                messagebox.showinfo('Gravity Applied', f"Applied preset '{preset_name}' for Gravity (pending export).")
        except Exception:
            pass
        try:
            if 'Gravity offload' in self.tab_status_labels:
                self.tab_status_labels['Gravity offload'].configure(text='')
        except Exception:
            pass

    def _set_gravity_free_controls_state(self, state: str) -> None:
        data_type_key = 'Gravity offload'
        if data_type_key not in self.dist_entries_by_type:
            return
        for _, widgets in self.dist_entries_by_type[data_type_key].items():
            dist_box, a_entry, b_entry = widgets
            try:
                dist_box.config(state=state)
            except Exception:
                pass
            try:
                a_entry.config(state=state)
            except Exception:
                pass
            try:
                b_entry.config(state=state)
            except Exception:
                pass

    def toggle_gravity_mode(self, data_type_key: str) -> None:
        if data_type_key != 'Gravity offload':
            return
        mode = self.gravity_mode_var.get()
        std_combo = None
        if data_type_key in self.distribution_widgets:
            std_combo = self.distribution_widgets[data_type_key].get('grav_std_combo')
        if std_combo is not None:
            try:
                std_combo.config(state='readonly' if mode == 'standard' else 'disabled')
            except Exception:
                pass
        if mode == 'standard':
            try:
                self.suppress_standard_apply_modals = True
                self.on_gravity_standard_selected()
            finally:
                self.suppress_standard_apply_modals = False
            self._set_gravity_free_controls_state(state='disabled')
        else:
            self._set_gravity_free_controls_state(state='normal')

    def _set_alignment_free_controls_state(self, state: str) -> None:
        """Enable/disable Alignment free-mode controls (dist type + value entries)."""
        data_type_key = 'Alignment'
        if data_type_key not in self.dist_entries_by_type:
            return
        for _, widgets in self.dist_entries_by_type[data_type_key].items():
            dist_box, a_entry, b_entry = widgets
            try:
                dist_box.config(state=state)
            except Exception:
                pass
            try:
                a_entry.config(state=state)
            except Exception:
                pass
            try:
                b_entry.config(state=state)
            except Exception:
                pass

    def toggle_alignment_mode(self, data_type_key: str) -> None:
        """Toggle between standard and free mode for Alignment."""
        if data_type_key != 'Alignment':
            return

        mode = self.align_mode_var.get()

        std_combo = None
        if data_type_key in self.distribution_widgets:
            std_combo = self.distribution_widgets[data_type_key].get('align_std_combo')

        if std_combo is not None:
            try:
                std_combo.config(state='readonly' if mode == 'standard' else 'disabled')
            except Exception:
                pass

        if mode == 'standard':
            # Apply preset then lock free controls (do not show modal on toggle)
            try:
                self.suppress_standard_apply_modals = True
                self.on_alignment_standard_selected()
            finally:
                self.suppress_standard_apply_modals = False
            self._set_alignment_free_controls_state(state='disabled')
        else:
            # Unlock free controls
            self._set_alignment_free_controls_state(state='normal')

    def update_param_labels(self, data_type_key, param):
        """Update parameter labels based on distribution type."""
        dist = self.dist_entries_by_type[data_type_key][param][0].get()
        label_a, label_b = self.param_labels_by_type[data_type_key][param]
        entry_a, entry_b = self.dist_entries_by_type[data_type_key][param][1:]
        
        if dist == 'fixed':
            label_a.config(text='Value:')
            # Hide second field and label for fixed
            label_b.grid_remove()
            entry_b.grid_remove()
        elif dist == 'gaussian' or dist == 'gamma':
            label_a.config(text='Mean:')
            label_b.config(text='Sigma:')
            # Show second field and label
            label_b.grid()
            entry_b.config(state='normal')
            entry_b.grid()
        elif dist == 'uniform':
            label_a.config(text='Min:')
            label_b.config(text='Max:')
            # Show second field and label
            label_b.grid()
            entry_b.config(state='normal')
            entry_b.grid()
    
    def toggle_psf_mode(self, data_type_key):
        """Toggle between standard and free mode for MM_PSF."""
        mode = self.psf_mode_var.get()
        # Enable/disable standard distribution dropdown
        if data_type_key in self.distribution_widgets and 'std_dist_combo' in self.distribution_widgets[data_type_key]:
            std_combo = self.distribution_widgets[data_type_key]['std_dist_combo']
            if mode == 'standard':
                std_combo.config(state='readonly')
                # Apply preset without showing modal (toggle should not show the modal)
                try:
                    self.suppress_standard_apply_modals = True
                    logging.debug("toggle_psf_mode invoking on_mm_psf_standard_selected (mode=standard) for %s", data_type_key)
                    self.on_mm_psf_standard_selected(data_type_key)
                finally:
                    self.suppress_standard_apply_modals = False
                logging.debug("toggle_psf_mode finished on_mm_psf_standard_selected, calling enforce_psf_alpha_ui")
                try:
                    self.enforce_psf_alpha_ui(data_type_key)
                except Exception as e:
                    logging.debug("enforce_psf_alpha_ui raised in toggle_psf_mode: %s", e)
            else:
                std_combo.config(state='disabled')

        # Show/hide custom PSF file picker controls
        if data_type_key in self.distribution_widgets:
            std_name = ''
            if 'std_dist_combo' in self.distribution_widgets[data_type_key]:
                std_name = self.distribution_widgets[data_type_key]['std_dist_combo'].get()
            show_custom = (mode == 'standard' and std_name == self.CUSTOM_PSF_OPTION)

            btn = self.distribution_widgets[data_type_key].get('btn_pick_psf')
            lbl = self.distribution_widgets[data_type_key].get('lbl_pick_psf')
            if btn is not None:
                if show_custom:
                    btn.grid()
                else:
                    btn.grid_remove()
            if lbl is not None:
                if show_custom:
                    lbl.grid()
                else:
                    lbl.grid_remove()

        # When switching to Free mode, ensure all distribution controls are enabled
        # so the user can edit per-parameter comboboxes and text entries.
        try:
            if mode == 'free' and data_type_key in self.distribution_widgets:
                # distribution type combobox
                try:
                    if 'dist_type' in self.distribution_widgets[data_type_key]:
                        self.distribution_widgets[data_type_key]['dist_type'].config(state='readonly')
                except Exception:
                    pass

                # enable per-parameter distribution boxes and entries
                if data_type_key in self.dist_entries_by_type:
                    for p, widgets in self.dist_entries_by_type[data_type_key].items():
                        try:
                            dist_box, entry_a, entry_b = widgets
                            dist_box.config(state='readonly')
                            entry_a.config(state='normal')
                            # If the second field is visible, enable it; otherwise leave hidden
                            try:
                                if entry_b.winfo_ismapped() or entry_b.winfo_viewable():
                                    entry_b.config(state='normal')
                            except Exception:
                                entry_b.config(state='normal')
                            # Refresh labels/states to reflect selection
                            try:
                                self.update_param_labels(data_type_key, p)
                            except Exception:
                                pass
                        except Exception:
                            pass

                # enable alpha parameter controls when present
                if data_type_key in self.alpha_entries_by_type:
                    for ui_param, widgets in self.alpha_entries_by_type[data_type_key].items():
                        try:
                            _, dist_box, entry_a, entry_b, _ = widgets
                            dist_box.config(state='readonly')
                            entry_a.config(state='normal')
                            try:
                                if entry_b.winfo_ismapped() or entry_b.winfo_viewable():
                                    entry_b.config(state='normal')
                            except Exception:
                                entry_b.config(state='normal')
                            try:
                                self.update_alpha_param_labels(data_type_key, ui_param)
                            except Exception:
                                pass
                        except Exception:
                            pass
        except Exception:
            pass


    def on_mm_psf_standard_selected(self, data_type_key: str) -> None:
        """Handle selection of a standard MM_PSF preset or the custom PSF file option."""
        try:
            if data_type_key not in self.distribution_widgets or 'std_dist_combo' not in self.distribution_widgets[data_type_key]:
                return

            std_name = self.distribution_widgets[data_type_key]['std_dist_combo'].get()

            logging.debug("on_mm_psf_standard_selected called for '%s', std_name='%s'", data_type_key, std_name)

            if std_name == self.CUSTOM_PSF_OPTION:
                # Keep all other MM_PSF entries disabled (standard mode already does this).
                # Ensure picker is visible.
                if 'btn_pick_psf' in self.distribution_widgets[data_type_key]:
                    self.distribution_widgets[data_type_key]['btn_pick_psf'].grid()
                if 'lbl_pick_psf' in self.distribution_widgets[data_type_key]:
                    self.distribution_widgets[data_type_key]['lbl_pick_psf'].grid()

                # Extra safety: disable dist type + m_rad/m_azi dist dropdowns while using a preset.
                if 'dist_type' in self.distribution_widgets[data_type_key]:
                    try:
                        self.distribution_widgets[data_type_key]['dist_type'].config(state='disabled')
                    except Exception:
                        pass
                for p in ['m_rad [arcsec]', 'm_azi [arcsec]']:
                    if data_type_key in self.dist_entries_by_type and p in self.dist_entries_by_type[data_type_key]:
                        try:
                            dist_box, entry_a, entry_b = self.dist_entries_by_type[data_type_key][p]
                            dist_box.config(state='disabled')
                            entry_a.config(state='disabled')
                            entry_b.config(state='disabled')
                        except Exception:
                            pass

                if not self.custom_psf_path_var.get().strip():
                    self.choose_custom_psf_file()
                    return

                return

            # Normal preset
            self.custom_psf_path_var.set('')
            self.custom_psf_stem_var.set('')

            # Hide picker when not on Custom
            if 'btn_pick_psf' in self.distribution_widgets[data_type_key]:
                self.distribution_widgets[data_type_key]['btn_pick_psf'].grid_remove()
            if 'lbl_pick_psf' in self.distribution_widgets[data_type_key]:
                self.distribution_widgets[data_type_key]['lbl_pick_psf'].grid_remove()

            # Extra safety: disable dist type + m_rad/m_azi dist dropdowns while using a preset.
            if 'dist_type' in self.distribution_widgets[data_type_key]:
                try:
                    self.distribution_widgets[data_type_key]['dist_type'].config(state='disabled')
                except Exception:
                    pass
            for p in ['m_rad [arcsec]', 'm_azi [arcsec]']:
                if data_type_key in self.dist_entries_by_type and p in self.dist_entries_by_type[data_type_key]:
                    try:
                        dist_box, entry_a, entry_b = self.dist_entries_by_type[data_type_key][p]
                        dist_box.config(state='disabled')
                        entry_a.config(state='disabled')
                        entry_b.config(state='disabled')
                    except Exception:
                        pass

            logging.debug("calling apply_standard_distribution for '%s'", std_name)
            self.apply_standard_distribution(data_type_key)
            logging.debug("apply_standard_distribution returned for '%s', now calling enforce_psf_alpha_ui", std_name)
            try:
                self.enforce_psf_alpha_ui(data_type_key)
            except Exception as e:
                logging.exception("enforce_psf_alpha_ui raised: %s", e)
        except Exception as e:
            logging.exception("Error in on_mm_psf_standard_selected: %s", e)
            try:
                messagebox.showerror('PSF Apply Error', f"Error applying PSF preset: {e}")
            except Exception:
                pass


    def choose_custom_psf_file(self) -> None:
        """Prompt the user to choose an external PSF matrix file and copy it into `CustomPSFs/`.

        The chosen file is moved into the repository `CustomPSFs` directory when
        possible (or copied if it lives outside the repo). Updates
        `self.custom_psf_path_var` and `self.custom_psf_stem_var` with the
        destination path and stem respectively.
        """
        path = filedialog.askopenfilename(initialdir="./Distributions", filetypes=[('Excel files', '*.xlsx *.xls')])
        if not path:
            return
        src = Path(path)

        # Move/copy into CustomPSFs so projects stay self-contained.
        repo_root = Path(__file__).resolve().parent
        dest_dir = repo_root / 'CustomPSFs'
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / src.name

        try:
            if src.resolve().parent != dest_dir.resolve():
                # If the file is inside this repo, move it; otherwise copy it.
                import shutil
                try:
                    src.resolve().relative_to(repo_root.resolve())
                    shutil.move(str(src), str(dest))
                except Exception:
                    shutil.copy2(str(src), str(dest))
        except Exception:
            # Fall back to using the original path if anything goes wrong.
            dest = src

        self.custom_psf_path_var.set(str(dest))
        self.custom_psf_stem_var.set(dest.stem)
    
    
    def apply_standard_distribution(self, data_type_key):
        """Apply the selected standard distribution parameters to the UI."""
        def clamp01(x):
            try:
                return min(max(float(x), 0.0), 1.0)
            except Exception:
                return 0.0

        if data_type_key not in self.distribution_widgets or 'std_dist_combo' not in self.distribution_widgets[data_type_key]:
            logging.debug("No std_dist_combo found for %s", data_type_key)
            return
        
        # Determine current mode for this data type (standard/free)
        mode = 'standard'
        try:
            if data_type_key == 'MM_PSF':
                mode = self.psf_mode_var.get()
            elif data_type_key == 'Alignment':
                mode = self.align_mode_var.get()
            elif data_type_key == 'Thermal':
                mode = self.thermal_mode_var.get()
            elif data_type_key == 'Gravity offload':
                mode = self.gravity_mode_var.get()
        except Exception:
            mode = 'standard'

        # Enable/disable per-parameter controls and allowed distributions
        if data_type_key in self.dist_entries_by_type:
            for param, widgets in self.dist_entries_by_type[data_type_key].items():
                dist_box, entry_a, entry_b = widgets
                try:
                    if mode == 'free':
                        # Free mode: allow fixed, gaussian, uniform (no gamma)
                        try:
                            dist_box.config(values=['fixed', 'gaussian', 'uniform'])
                        except Exception:
                            pass
                        try:
                            dist_box.config(state='normal')
                            entry_a.config(state='normal')
                            entry_b.config(state='normal')
                        except Exception:
                            pass
                    else:
                        # Standard mode: allow gamma as representation of Variable presets
                        try:
                            dist_box.config(values=['fixed', 'gaussian', 'gamma', 'uniform'])
                        except Exception:
                            pass
                        # Most fields will be disabled after applying preset
                        try:
                            dist_box.config(state='disabled')
                            entry_a.config(state='disabled')
                            entry_b.config(state='disabled')
                        except Exception:
                            pass
                except Exception:
                    pass

        # Alpha parameter controls: enable in free mode, otherwise leave handled by apply_standard_distribution
        if data_type_key in self.alpha_entries_by_type:
            for ui_param, widgets in self.alpha_entries_by_type[data_type_key].items():
                _, dist_box, entry_a, entry_b, _ = widgets
                try:
                    if mode == 'free':
                        dist_box.config(values=['fixed', 'gaussian', 'uniform'])
                        dist_box.config(state='normal')
                        entry_a.config(state='normal')
                        entry_b.config(state='normal')
                    else:
                        dist_box.config(values=['fixed', 'gaussian', 'gamma', 'uniform'])
                        dist_box.config(state='disabled')
                        entry_a.config(state='disabled')
                        entry_b.config(state='disabled')
                except Exception:
                    pass
        std_name = self.distribution_widgets[data_type_key]['std_dist_combo'].get()
        if std_name == self.CUSTOM_PSF_OPTION:
            return
        if not std_name or std_name not in self.standard_distributions:
            logging.debug("std_name '%s' not in standard_distributions: %s", std_name, list(self.standard_distributions.keys()))
            return
        
        std_def = self.standard_distributions[std_name]
        logging.debug("Applying standard distribution '%s': %s", std_name, std_def)
        
        # Set distribution type
        if 'dist_type' in self.distribution_widgets[data_type_key]:
            # For fixed pseudo-voigt, ensure type is set to 'pseudo-voigt' (not 'gaussian')
            if std_def.get('type') == 'pseudo-voigt':
                self.distribution_widgets[data_type_key]['dist_type'].set('pseudo-voigt')
            else:
                self.distribution_widgets[data_type_key]['dist_type'].set(std_def['type'])
            self.toggle_eta_entry(data_type_key)

        

        # Standard presets assume zero mean offsets unless explicitly controlled elsewhere.
        # Ensure m_rad/m_azi fields are consistent when selecting different presets.
        if data_type_key in self.dist_entries_by_type:
            for p in ('m_rad [arcsec]', 'm_azi [arcsec]'):
                if p in self.dist_entries_by_type[data_type_key]:
                    dist_box, entry_a, entry_b = self.dist_entries_by_type[data_type_key][p]
                    label_a, label_b = self.param_labels_by_type[data_type_key][p]
                    try:
                        dist_box.config(state='normal')
                        entry_a.config(state='normal')
                        entry_b.config(state='normal')
                    except Exception:
                        pass
                    dist_box.set('fixed')
                    entry_a.delete(0, tk.END)
                    entry_a.insert(0, '0')
                    label_a.config(text='Value:')
                    label_b.grid_remove()
                    entry_b.grid_remove()
                    if self.psf_mode_var.get() == 'standard':
                        try:
                            dist_box.config(state='disabled')
                            entry_a.config(state='disabled')
                            entry_b.config(state='disabled')
                        except Exception:
                            pass
        
        # Apply sigma parameters.
        sigma_ui = {
            'sigma_rad [arcsec]': 'sigma_rad',
            'sigma_azi [arcsec]': 'sigma_azi',
        }

        sigma_defs: dict[str, dict | None] = {}
        preset_name = str(std_def.get('name', '')).strip()

        # If the preset name starts with 'Fixed', disable all downstream
        # distribution controls (distribution type comboboxes and text boxes).
        is_fixed_preset = bool(re.match(r'^\s*fixed', str(preset_name), re.IGNORECASE))
        if is_fixed_preset:
            # Disable main dist_type selector if present
            try:
                if 'dist_type' in self.distribution_widgets[data_type_key]:
                    # For fixed pseudo-voigt, set to 'pseudo-voigt', else fallback
                    if std_def.get('type') == 'pseudo-voigt':
                        try:
                            self.distribution_widgets[data_type_key]['dist_type'].set('pseudo-voigt')
                        except Exception:
                            pass
                    else:
                        try:
                            self.distribution_widgets[data_type_key]['dist_type'].set('gaussian')
                        except Exception:
                            pass
                    self.distribution_widgets[data_type_key]['dist_type'].config(state='disabled')
            except Exception:
                pass

            # Disable per-parameter distribution widgets and set to fixed where possible
            if data_type_key in self.dist_entries_by_type:
                for param, widgets in self.dist_entries_by_type[data_type_key].items():
                    try:
                        dist_box, entry_a, entry_b = widgets
                        try:
                            dist_box.set('fixed')
                        except Exception:
                            pass
                        try:
                            dist_box.config(state='disabled')
                        except Exception:
                            pass
                        try:
                            entry_a.config(state='disabled')
                        except Exception:
                            pass
                        try:
                            entry_b.config(state='disabled')
                        except Exception:
                            pass
                    except Exception:
                        pass

            # Disable alpha parameter widgets as well
            if data_type_key in self.alpha_entries_by_type:
                for ui_param, widgets in self.alpha_entries_by_type[data_type_key].items():
                    try:
                        _, dist_box, entry_a, entry_b, _ = widgets
                        try:
                            dist_box.set('fixed')
                        except Exception:
                            pass
                        try:
                            dist_box.config(state='disabled')
                        except Exception:
                            pass
                        try:
                            entry_a.config(state='disabled')
                        except Exception:
                            pass
                        try:
                            entry_b.config(state='disabled')
                        except Exception:
                            pass
                    except Exception:
                        pass

        # Start from the parsed std_def
        for ui_param, std_param in sigma_ui.items():
            if ui_param not in self.dist_entries_by_type[data_type_key] or std_param not in std_def:
                continue
            sigma_defs[std_param] = std_def.get(std_param)

        # Fallback: derive gamma(mean,sigma) when the table cell was empty and the preset indicates % variability.
        # For symmetric Gaussians, HEW_diameter = 2*sqrt(2 ln2)*sigma.
        for std_param in ('sigma_rad', 'sigma_azi'):
            if std_param not in sigma_defs or sigma_defs.get(std_param):
                continue
            if std_def.get('type') not in ('gaussian', 'pseudo-voigt'):
                continue
            m_pct = re.search(r"(\d+)\%\s*Variable", preset_name)
            m_hew = re.search(r"(\d+(?:\.\d+)?)\"", preset_name)
            if not (m_pct and m_hew):
                continue
            pct = float(m_pct.group(1)) / 100.0
            hew_val = float(m_hew.group(1))
            # Symmetric Gaussian mapping: HEW_diameter = 2*sqrt(2 ln2)*sigma
            import math
            mean_val = hew_val / (2.0 * math.sqrt(2.0 * math.log(2.0)))
            # Use gamma distribution to describe per-MM variability from presets labelled 'Variable'
            sigma_defs[std_param] = {'dist': 'gamma', 'mean': mean_val, 'sigma': abs(pct * mean_val)}

        # Push sigma defs into the UI
        for ui_param, std_param in sigma_ui.items():
            if ui_param not in self.dist_entries_by_type[data_type_key]:
                continue
            param_def = sigma_defs.get(std_param)
            if not param_def:
                continue

            dist_box, entry_a, entry_b = self.dist_entries_by_type[data_type_key][ui_param]
            label_a, label_b = self.param_labels_by_type[data_type_key][ui_param]

            logging.debug("Setting %s to %s", ui_param, param_def)
            is_variable_preset = bool(re.search(r"\bVariable\b", preset_name, re.IGNORECASE))

            # Temporarily enable everything to update
            dist_box.config(state='normal')
            entry_a.config(state='normal')
            entry_b.config(state='normal')

            if param_def['dist'] == 'fixed':
                dist_box.set('fixed')
                entry_a.delete(0, tk.END)
                entry_a.insert(0, str(param_def['value']))
                label_a.config(text='Value:')
                label_b.grid_remove()
                entry_b.grid_remove()
            elif param_def['dist'] == 'gaussian' or param_def['dist'] == 'gamma':
                try:
                    # If the preset is a Variable preset, present it as a gamma distribution
                    if is_variable_preset:
                        dist_box.set('gamma')
                    else:
                        dist_box.set(param_def['dist'])
                except Exception:
                    pass
                entry_a.delete(0, tk.END)
                entry_a.insert(0, str(param_def.get('mean')))
                entry_b.delete(0, tk.END)
                entry_b.insert(0, str(param_def.get('sigma')))
                label_a.config(text='Mean:')
                label_b.config(text='Sigma:')
                label_b.grid()
                entry_b.grid()
            elif param_def['dist'] == 'uniform':
                dist_box.set('uniform')
                entry_a.delete(0, tk.END)
                entry_a.insert(0, str(param_def['min']))
                entry_b.delete(0, tk.END)
                entry_b.insert(0, str(param_def['max']))
                label_a.config(text='Min:')
                label_b.config(text='Max:')
                label_b.grid()
                entry_b.grid()

            try:
                logging.debug("dist_box=%s, entry_a=%s, entry_b visible=%s", dist_box.get(), entry_a.get(), entry_b.winfo_viewable())
            except Exception:
                logging.debug("dist box state logging failed for %s", ui_param)

            # Ensure labels and states are consistent with current distribution selection
            try:
                self.update_param_labels(data_type_key, ui_param)
            except Exception:
                pass

            # Re-enforce disabled state for Fixed presets as well (ensure GUI helpers
            # that temporarily enabled fields do not leave them writable).
            try:
                if is_fixed_preset:
                    try:
                        dist_box.config(state='disabled')
                    except Exception:
                        pass
                    try:
                        entry_a.config(state='disabled')
                    except Exception:
                        pass
                    try:
                        entry_b.config(state='disabled')
                    except Exception:
                        pass
            except Exception:
                pass

            # Re-apply disabled state for Variable presets while in standard mode.
            # `update_param_labels` may re-enable the sigma entry, so enforce disable
            # after the label/state update.
            try:
                if self.psf_mode_var.get() == 'standard' and is_variable_preset:
                    try:
                        dist_box.config(state='disabled')
                    except Exception:
                        pass
                    try:
                        entry_a.config(state='disabled')
                    except Exception:
                        pass
                    try:
                        # For Variable presets we want sigma fields disabled as well
                        if param_def.get('dist') != 'fixed':
                            entry_b.config(state='disabled')
                    except Exception:
                        pass
            except Exception:
                pass
        
        # Apply alpha parameters if pseudo-voigt
        if std_def['type'] == 'pseudo-voigt' and data_type_key in self.alpha_entries_by_type:
            is_variable_preset = bool(re.search(r"\bVariable\b", preset_name, re.IGNORECASE))
            alpha_mapping = {
                'alpha_rad': 'alpha_rad',
                'alpha_azi': 'alpha_azi'
            }

            # try to extract alpha-specific percent (e.g., 'alpha (10%)') or overall '% Variable'
            m_alpha = re.search(r"alpha\s*\(?\s*(\d+(?:\.\d+)?)\s*%\s*\)?", preset_name, re.IGNORECASE)
            alpha_pct = float(m_alpha.group(1)) / 100.0 if m_alpha else None
            m_overall_pct = re.search(r"(\d+(?:\.\d+)?)\%\s*Variable", preset_name, re.IGNORECASE)
            overall_pct = float(m_overall_pct.group(1)) / 100.0 if m_overall_pct else None

            for ui_param, std_param in alpha_mapping.items():
                # prefer explicit spec coming from the spreadsheet
                explicit = std_def.get(std_param)
                if explicit and explicit:
                    param_def = explicit
                    # If preset explicitly contains a fixed alpha but the preset name
                    # indicates a Variable pseudo-voigt with an alpha(...) percent,
                    # prefer showing it as a gamma(mean,sigma) in the UI (disabled).
                    if (param_def.get('dist') == 'fixed') and is_variable_preset and (alpha_pct is not None or overall_pct is not None):
                        # derive baseline if possible or use defaults
                        import math
                        hew_match = re.search(r"(\d+(?:\.\d+)?)\"", preset_name)
                        hew_str = hew_match.group(1) if hew_match else None
                        baseline_alpha = {'alpha_rad': None, 'alpha_azi': None}
                        if hew_str and hasattr(self, 'standard_distributions'):
                            for preset_name2, preset2 in self.standard_distributions.items():
                                try:
                                    if preset2.get('type') == 'pseudo-voigt' and 'fixed' in str(preset_name2).lower() and hew_str in str(preset_name2):
                                        ar = preset2.get('alpha_rad')
                                        aa = preset2.get('alpha_azi')
                                        if ar and aa and ar.get('dist') == 'fixed' and aa.get('dist') == 'fixed':
                                            baseline_alpha['alpha_rad'] = float(ar['value'])
                                            baseline_alpha['alpha_azi'] = float(aa['value'])
                                            break
                                except Exception:
                                    pass
                        if baseline_alpha['alpha_rad'] is None:
                            baseline_alpha['alpha_rad'] = 0.77
                        if baseline_alpha['alpha_azi'] is None:
                            baseline_alpha['alpha_azi'] = 0.29
                        mean_val = baseline_alpha[ui_param]
                        use_pct = alpha_pct if alpha_pct is not None else overall_pct
                        sigma_val = abs(use_pct * mean_val) if use_pct is not None else 0.0
                        param_def = {'dist': 'gamma', 'mean': clamp01(mean_val), 'sigma': sigma_val}
                else:
                    # If this is a Variable preset and no explicit alpha spec exists,
                    # derive a gamma(mean,sigma) using either matching fixed presets
                    # or sensible defaults (0.77 / 0.29) and the percent from the name.
                    if is_variable_preset and (alpha_pct is not None or overall_pct is not None):
                        # attempt to find baseline alpha means from matching fixed preset (same HEW)
                        import math
                        hew_match = re.search(r"(\d+(?:\.\d+)?)\"", preset_name)
                        hew_str = hew_match.group(1) if hew_match else None
                        baseline_alpha = {'alpha_rad': None, 'alpha_azi': None}
                        if hew_str and hasattr(self, 'standard_distributions'):
                            for preset_name2, preset2 in self.standard_distributions.items():
                                try:
                                    if preset2.get('type') == 'pseudo-voigt' and 'fixed' in str(preset_name2).lower() and hew_str in str(preset_name2):
                                        ar = preset2.get('alpha_rad')
                                        aa = preset2.get('alpha_azi')
                                        if ar and aa and ar.get('dist') == 'fixed' and aa.get('dist') == 'fixed':
                                            baseline_alpha['alpha_rad'] = float(ar['value'])
                                            baseline_alpha['alpha_azi'] = float(aa['value'])
                                            break
                                except Exception:
                                    pass
                        # defaults when baseline not found
                        if baseline_alpha['alpha_rad'] is None:
                            baseline_alpha['alpha_rad'] = 0.77
                        if baseline_alpha['alpha_azi'] is None:
                            baseline_alpha['alpha_azi'] = 0.29

                        mean_val = baseline_alpha[ui_param]
                        # prefer alpha-specific percent, otherwise fall back to overall Variable percent
                        use_pct = alpha_pct if alpha_pct is not None else overall_pct
                        sigma_val = abs(use_pct * mean_val) if use_pct is not None else 0.0
                        param_def = {'dist': 'gamma', 'mean': clamp01(mean_val), 'sigma': sigma_val}
                    else:
                        # no explicit spec and not a variable preset with percent: skip
                        continue

                # At this point param_def is defined (either explicit or derived)
                _, dist_box, entry_a, entry_b, _ = self.alpha_entries_by_type[data_type_key][ui_param]
                label_a, label_b = self.param_labels_by_type[data_type_key][ui_param]

                # Clamp alpha means into [0,1]
                if param_def.get('dist') == 'fixed':
                    param_def = {**param_def, 'value': clamp01(param_def['value'])}
                elif param_def.get('dist') in ('gaussian', 'gamma'):
                    param_def = {**param_def, 'mean': clamp01(param_def.get('mean'))}

                logging.debug("Setting alpha %s to %s", ui_param, param_def)

                # Temporarily enable everything to update
                try:
                    dist_box.config(state='normal')
                    entry_a.config(state='normal')
                    entry_b.config(state='normal')
                except Exception:
                    pass

                if param_def.get('dist') == 'fixed':
                    try:
                        dist_box.set('fixed')
                    except Exception:
                        pass
                    entry_a.delete(0, tk.END)
                    entry_a.insert(0, str(param_def['value']))
                    label_a.config(text='Value:')
                    label_b.grid_remove()
                    entry_b.grid_remove()
                    # Ensure fields remain disabled for Fixed presets
                    try:
                        if is_fixed_preset or self.psf_mode_var.get() == 'standard':
                            dist_box.config(state='disabled')
                            entry_a.config(state='disabled')
                    except Exception:
                        pass
                elif param_def.get('dist') in ('gaussian', 'gamma'):
                    try:
                        # For Variable presets always present as gamma
                        if is_variable_preset:
                            dist_box.set('gamma')
                        else:
                            dist_box.set(param_def.get('dist'))
                    except Exception:
                        pass
                    entry_a.delete(0, tk.END)
                    entry_a.insert(0, str(param_def.get('mean')))
                    entry_b.delete(0, tk.END)
                    entry_b.insert(0, str(param_def.get('sigma')))
                    label_a.config(text='Mean:')
                    label_b.config(text='Sigma:')
                    label_b.grid()
                    entry_b.grid()

                try:
                    logging.debug("alpha dist_box=%s, entry_a=%s", dist_box.get(), entry_a.get())
                except Exception:
                    logging.debug("alpha dist box state logging failed for %s", ui_param)

                # Update labels and possibly re-disable fields
                try:
                    self.update_alpha_param_labels(data_type_key, ui_param)
                except Exception:
                    pass

                # Re-enforce disabled state for alpha sigma fields if Variable preset
                try:
                    if self.psf_mode_var.get() == 'standard' and is_variable_preset:
                        try:
                            dist_box.config(state='disabled')
                        except Exception:
                            pass
                        try:
                            entry_a.config(state='disabled')
                        except Exception:
                            pass
                        try:
                            if param_def.get('dist') != 'fixed':
                                entry_b.config(state='disabled')
                        except Exception:
                            pass
                except Exception:
                    pass

                # Re-enforce disabled state for Fixed presets as well
                try:
                    if is_fixed_preset:
                        try:
                            dist_box.config(state='disabled')
                        except Exception:
                            pass
                        try:
                            entry_a.config(state='disabled')
                        except Exception:
                            pass
                        try:
                            entry_b.config(state='disabled')
                        except Exception:
                            pass
                except Exception:
                    pass
        
        # Fallback for alpha params when missing in pseudo-voigt presets (e.g., "(alpha 10%)")
        if std_def['type'] == 'pseudo-voigt' and data_type_key in self.alpha_entries_by_type:
            import math
            name = std_def.get('name', '')
            m_alpha = re.search(r"alpha\s*(\d+)\%", name, re.IGNORECASE)
            if m_alpha:
                pct = float(m_alpha.group(1)) / 100.0
                # Try to infer baseline alpha means from a matching fixed pseudo-voigt preset (same HEW)
                hew_match = re.search(r"(\d+(?:\.\d+)?)\"", name)
                hew_str = hew_match.group(1) if hew_match else None
                baseline_alpha = {'alpha_rad': None, 'alpha_azi': None}
                if hew_str and hasattr(self, 'standard_distributions'):
                    for preset_name, preset in self.standard_distributions.items():
                        try:
                            if preset.get('type') == 'pseudo-voigt' and 'fixed' in str(preset_name).lower() and hew_str in str(preset_name):
                                ar = preset.get('alpha_rad')
                                aa = preset.get('alpha_azi')
                                if ar and aa and ar.get('dist') == 'fixed' and aa.get('dist') == 'fixed':
                                    baseline_alpha['alpha_rad'] = float(ar['value'])
                                    baseline_alpha['alpha_azi'] = float(aa['value'])
                                    break
                        except Exception:
                            pass
                # If not found in presets, fall back to requested constants 0.77/0.29
                if baseline_alpha['alpha_rad'] is None:
                    baseline_alpha['alpha_rad'] = 0.77
                if baseline_alpha['alpha_azi'] is None:
                    baseline_alpha['alpha_azi'] = 0.29
                for ui_param in ('alpha_rad', 'alpha_azi'):
                    if ui_param in self.alpha_entries_by_type[data_type_key]:
                        _, dist_box, entry_a, entry_b, _ = self.alpha_entries_by_type[data_type_key][ui_param]
                        label_a, label_b = self.param_labels_by_type[data_type_key][ui_param]
                        # determine mean: from matching fixed preset or requested defaults
                        mean_val = baseline_alpha[ui_param]
                        sigma_val = abs(pct * mean_val)
                        # For Variable presets, model alpha variability with a gamma distribution
                        param_def = {'dist': 'gamma', 'mean': clamp01(mean_val), 'sigma': sigma_val}
                        logging.debug("Derived %s from name '%s' using baseline %s: sigma=%s", ui_param, name, param_def.get('mean'), sigma_val)
                        # Temporarily enable to update
                        dist_box.config(state='normal')
                        entry_a.config(state='normal')
                        entry_b.config(state='normal')
                        dist_box.set(param_def['dist'])
                        entry_a.delete(0, tk.END)
                        entry_a.insert(0, str(param_def['mean']))
                        entry_b.delete(0, tk.END)
                        entry_b.insert(0, str(param_def['sigma']))
                        label_a.config(text='Mean:')
                        label_b.config(text='Sigma:')
                        label_b.grid()
                        entry_b.grid()
                        if self.psf_mode_var.get() == 'standard' and is_variable_preset:
                            # If standard Variable preset, keep the gamma shown but disabled
                            try:
                                dist_box.config(state='disabled')
                            except Exception:
                                pass
                            try:
                                entry_a.config(state='disabled')
                            except Exception:
                                pass
                            try:
                                entry_b.config(state='disabled')
                            except Exception:
                                pass

        # Ensure that when in standard PSF mode for pseudo-voigt presets, alpha fields are shown as disabled
        # and hold the mean value (not editable). This covers 'Variable' presets where the alpha
        # distribution is not intended for per-MM editing.
        try:
            if self.psf_mode_var.get() == 'standard' and std_def.get('type') == 'pseudo-voigt':
                # Determine if this preset is a 'Variable' preset
                is_variable_preset = bool(re.search(r"\bVariable\b", preset_name, re.IGNORECASE))
                for ui_param in ('alpha_rad', 'alpha_azi'):
                    if data_type_key in self.alpha_entries_by_type and ui_param in self.alpha_entries_by_type[data_type_key]:
                        _, dist_box, entry_a, entry_b, _ = self.alpha_entries_by_type[data_type_key][ui_param]
                        label_a, label_b = self.param_labels_by_type[data_type_key][ui_param]
                        # prefer explicit spec, otherwise try derived baseline
                        pdef = std_def.get(ui_param)
                        if pdef and pdef.get('dist') == 'fixed':
                            mean_val = float(pdef.get('value'))
                            sigma_val = 0.0
                        elif pdef and pdef.get('dist') in ('gaussian', 'gamma'):
                            mean_val = float(pdef.get('mean'))
                            sigma_val = float(pdef.get('sigma', 0.0))
                        else:
                            # If no explicit alpha spec, try to reuse derived values from earlier fallback
                            mean_val = None
                            sigma_val = None
                        if mean_val is None:
                            # try to read whatever is currently in the entry_a/b
                            try:
                                mean_val = float(entry_a.get())
                            except Exception:
                                mean_val = 0.0
                            try:
                                sigma_val = float(entry_b.get())
                            except Exception:
                                sigma_val = 0.0

                        # If this is a Variable preset, show gamma(mean,sigma) and disable editing
                        try:
                            if is_variable_preset:
                                try:
                                    dist_box.set('gamma')
                                except Exception:
                                    pass
                                self._set_entry_text(entry_a, f"{mean_val}")
                                self._set_entry_text(entry_b, f"{sigma_val}")
                                label_a.config(text='Mean:')
                                label_b.config(text='Sigma:')
                                label_b.grid()
                                entry_b.grid()
                                try:
                                    dist_box.config(state='disabled')
                                    entry_a.config(state='disabled')
                                    entry_b.config(state='disabled')
                                except Exception:
                                    pass
                            else:
                                # Non-variable: present fixed mean and hide sigma
                                try:
                                    dist_box.set('fixed')
                                except Exception:
                                    pass
                                self._set_entry_text(entry_a, f"{mean_val}")
                                self._set_entry_text(entry_b, '0')
                                label_a.config(text='Value:')
                                label_b.grid_remove()
                                entry_b.grid_remove()
                                try:
                                    dist_box.config(state='disabled')
                                    entry_a.config(state='disabled')
                                except Exception:
                                    pass
                        except Exception:
                            pass
        except Exception:
            pass

        # Final enforcement: ensure Variable pseudo-voigt alpha fields are shown as gamma(mean,sigma)
        # This guarantees UI consistency even if other helpers toggled widgets.
        try:
            if std_def.get('type') == 'pseudo-voigt' and is_variable_preset and data_type_key in self.alpha_entries_by_type:
                # determine percent to use for sigma (alpha-specific or overall)
                use_pct = alpha_pct if alpha_pct is not None else overall_pct
                for ui_param in ('alpha_rad', 'alpha_azi'):
                    try:
                        _, dist_box, entry_a, entry_b, _ = self.alpha_entries_by_type[data_type_key][ui_param]
                        label_a, label_b = self.param_labels_by_type[data_type_key][ui_param]
                        # pick mean from explicit spec if present, otherwise fallback
                        pdef = std_def.get(ui_param)
                        if pdef and pdef.get('dist') in ('gaussian', 'gamma'):
                            mean_val = clamp01(pdef.get('mean'))
                            sigma_val = float(pdef.get('sigma', 0.0))
                        else:
                            # try to find baseline fixed preset or defaults
                            baseline = None
                            hew_match = re.search(r"(\d+(?:\.\d+)?)\"", preset_name)
                            hew_str = hew_match.group(1) if hew_match else None
                            if hew_str and hasattr(self, 'standard_distributions'):
                                for pname, preset in self.standard_distributions.items():
                                    try:
                                        if preset.get('type') == 'pseudo-voigt' and 'fixed' in str(pname).lower() and hew_str in str(pname):
                                            b = preset.get(ui_param)
                                            if b and b.get('dist') == 'fixed':
                                                baseline = float(b.get('value'))
                                                break
                                    except Exception:
                                        pass
                            if baseline is None:
                                baseline = 0.77 if ui_param == 'alpha_rad' else 0.29
                            mean_val = clamp01(baseline)
                            sigma_val = abs(use_pct * mean_val) if use_pct is not None else 0.0

                        # set UI to gamma mean/sigma and disable
                        try:
                            dist_box.set('gamma')
                        except Exception:
                            pass
                        self._set_entry_text(entry_a, f"{mean_val}")
                        self._set_entry_text(entry_b, f"{sigma_val}")
                        label_a.config(text='Mean:')
                        label_b.config(text='Sigma:')
                        label_b.grid()
                        entry_b.grid()
                        try:
                            dist_box.config(state='disabled')
                            entry_a.config(state='disabled')
                            entry_b.config(state='disabled')
                        except Exception:
                            pass
                    except Exception:
                        pass
        except Exception:
            pass

        # Notify user via modal and clear any persistent in-window status
        try:
            if not getattr(self, 'suppress_standard_apply_modals', False):
                messagebox.showinfo(f"{DATA_TYPES[data_type_key]['tab_label']} Applied",
                                    f"Applied standard preset '{preset_name}' for {DATA_TYPES[data_type_key]['tab_label']} (pending export).")
        except Exception:
            pass
        try:
            if data_type_key in self.tab_status_labels:
                self.tab_status_labels[data_type_key].configure(text='')
        except Exception:
            pass

    def update_alpha_param_labels(self, data_type_key, param):
        """Update alpha parameter labels based on distribution type."""
        if data_type_key not in self.alpha_entries_by_type:
            return
        if param not in self.alpha_entries_by_type[data_type_key]:
            return
        
        _, dist_box, entry_a, entry_b, _ = self.alpha_entries_by_type[data_type_key][param]
        label_a, label_b = self.param_labels_by_type[data_type_key][param]
        
        dist = dist_box.get()
        
        if dist == 'fixed':
            label_a.config(text='Value:')
            label_b.grid_remove()
            entry_b.grid_remove()
        elif dist == 'gaussian' or dist == 'gamma':
            label_a.config(text='Mean:')
            label_b.config(text='Sigma:')
            label_b.grid()
            entry_b.config(state='normal')
            entry_b.grid()
        elif dist == 'uniform':
            label_a.config(text='Min:')
            label_b.config(text='Max:')
            label_b.grid()
            entry_b.config(state='normal')
            entry_b.grid()

    def enforce_psf_alpha_ui(self, data_type_key):
        """Runtime enforcement: ensure alpha fields show gamma(mean,sigma) and are disabled
        for 'Variable' pseudo-voigt standard presets. Call from live handlers to guarantee
        UI consistency when users toggle modes or select presets interactively."""
        try:
            def _clamp01(x):
                try:
                    return min(max(float(x), 0.0), 1.0)
                except Exception:
                    return 0.0
            if data_type_key not in self.distribution_widgets or 'std_dist_combo' not in self.distribution_widgets[data_type_key]:
                return
            std_name = self.distribution_widgets[data_type_key]['std_dist_combo'].get()
            if not std_name or std_name == self.CUSTOM_PSF_OPTION:
                return
            std_def = self.standard_distributions.get(std_name)
            if not std_def or std_def.get('type') != 'pseudo-voigt':
                return
            is_variable_preset = bool(re.search(r"\bVariable\b", std_name, re.IGNORECASE))
            if not is_variable_preset:
                return

            # prefer alpha-specific percent in name, else overall '% Variable'
            m_alpha = re.search(r"alpha\s*\(?\s*(\d+(?:\.\d+)?)\s*%\s*\)?", std_name, re.IGNORECASE)
            alpha_pct = float(m_alpha.group(1)) / 100.0 if m_alpha else None
            m_overall = re.search(r"(\d+(?:\.\d+)?)\%\s*Variable", std_name, re.IGNORECASE)
            overall_pct = float(m_overall.group(1)) / 100.0 if m_overall else None
            use_pct = alpha_pct if alpha_pct is not None else overall_pct

            for ui_param in ('alpha_rad', 'alpha_azi'):
                if data_type_key not in self.alpha_entries_by_type or ui_param not in self.alpha_entries_by_type[data_type_key]:
                    continue
                try:
                    _, dist_box, entry_a, entry_b, _ = self.alpha_entries_by_type[data_type_key][ui_param]
                    label_a, label_b = self.param_labels_by_type[data_type_key][ui_param]

                    # Determine mean: prefer explicit spec, else derive baseline or default
                    pdef = std_def.get(ui_param)
                    if pdef and pdef.get('dist') in ('gaussian', 'gamma'):
                        mean_val = _clamp01(pdef.get('mean'))
                    else:
                        # try to find matching fixed preset baseline
                        baseline = None
                        hew_match = re.search(r"(\d+(?:\.\d+)?)\"", std_name)
                        hew_str = hew_match.group(1) if hew_match else None
                        if hew_str and hasattr(self, 'standard_distributions'):
                            for pname, preset in self.standard_distributions.items():
                                try:
                                    if preset.get('type') == 'pseudo-voigt' and 'fixed' in str(pname).lower() and hew_str in str(pname):
                                        b = preset.get(ui_param)
                                        if b and b.get('dist') == 'fixed':
                                            baseline = float(b.get('value'))
                                            break
                                except Exception:
                                    pass
                        if baseline is None:
                            baseline = 0.77 if ui_param == 'alpha_rad' else 0.29
                        mean_val = _clamp01(baseline)

                    sigma_val = abs(use_pct * mean_val) if use_pct is not None else 0.0

                    logging.debug("ENFORCE: setting %s -> gamma(mean=%s, sigma=%s) for preset '%s'", ui_param, mean_val, sigma_val, std_name)
                    try:
                        dist_box.set('gamma')
                    except Exception:
                        pass
                    self._set_entry_text(entry_a, f"{mean_val}")
                    self._set_entry_text(entry_b, f"{sigma_val}")
                    label_a.config(text='Mean:')
                    label_b.config(text='Sigma:')
                    label_b.grid()
                    entry_b.grid()
                    try:
                        dist_box.config(state='disabled')
                        entry_a.config(state='disabled')
                        entry_b.config(state='disabled')
                    except Exception:
                        pass
                except Exception:
                    pass
        except Exception as e:
            logging.exception("ENFORCE ERROR: %s", e)
    
    def toggle_eta_entry(self, data_type_key):
        """Show/hide alpha controls based on distribution type."""
        if data_type_key in self.distribution_widgets:
            dist_type = self.distribution_widgets[data_type_key]['dist_type'].get()
            
            # Show/hide alpha parameters for pseudo-voigt
            if data_type_key in self.alpha_entries_by_type:
                for param, widgets in self.alpha_entries_by_type[data_type_key].items():
                    param_label, dist_box, entry_a, entry_b, description_label = widgets
                    label_a, label_b = self.param_labels_by_type[data_type_key][param]
                    
                    if dist_type == 'pseudo-voigt':
                        # Show alpha parameters
                        param_label.grid()
                        dist_box.grid()
                        label_a.grid()
                        entry_a.grid()
                        description_label.grid()
                        # Check if second field should be shown based on distribution
                        dist = dist_box.get()
                        if dist == 'fixed':
                            label_b.grid_remove()
                            entry_b.grid_remove()
                        else:
                            label_b.grid()
                            entry_b.config(state='normal')
                            entry_b.grid()
                    else:
                        # Hide alpha parameters for gaussian
                        param_label.grid_remove()
                        dist_box.grid_remove()
                        label_a.grid_remove()
                        entry_a.grid_remove()
                        label_b.grid_remove()
                        entry_b.grid_remove()
                        description_label.grid_remove()
    
    def generate_data(self, data_type_key):
        """Generate data for a specific data type."""
        if self.mm_config_df is None:
            messagebox.showwarning('No Data', 'Please load an Excel file first!')
            return
        
        if not self.selected_mm_numbers:
            messagebox.showwarning('No Selection', 'Please select at least one MM in the MM Configuration tab!')
            return
        
        try:
            config = DATA_TYPES[data_type_key]
            num_mm = len(self.selected_mm_numbers)

            # Special case: MM_PSF from a custom PSF matrix file
            if (
                data_type_key == 'MM_PSF'
                and self.psf_mode_var.get() == 'standard'
                and data_type_key in self.distribution_widgets
                and 'std_dist_combo' in self.distribution_widgets[data_type_key]
                and self.distribution_widgets[data_type_key]['std_dist_combo'].get() == self.CUSTOM_PSF_OPTION
            ):
                stem = self.custom_psf_stem_var.get().strip()
                path = self.custom_psf_path_var.get().strip()
                if not stem or not path:
                    messagebox.showwarning('Missing PSF file', 'Please choose a PSF matrix Excel file first.')
            


            # (removed accidental debug path that re-launched the GUI when
            # this module was executed as __main__; keep normal in-app flow)

            params = {}
            col_specs: dict[str, str] = {}

            # Collect regular parameters
            for param, widgets in self.dist_entries_by_type[data_type_key].items():
                dist, a, b = widgets
                dist_type = dist.get()

                a_val = a.get().strip() if hasattr(a, 'get') else str(a)
                # If user entered an Excel column-letter mapping (e.g. 'W' or 'W,X,Y,AB'), record it
                try:
                    if self._is_column_letter_spec(a_val):
                        col_specs[param] = a_val
                        # mark as a special 'colref' distribution so downstream code can recognize it
                        params[param] = ('colref', a_val, 0.0)
                        continue
                except Exception:
                    pass

                # For 'fixed' distribution, only use first parameter (second is disabled)
                if dist_type == 'fixed':
                    params[param] = (dist_type, float(a_val), 0.0)
                else:
                    params[param] = (dist_type, float(a_val), float(b.get()))
            
            # Collect alpha parameters if pseudo-voigt is selected
            if (config.get('has_distribution', False) and 
                data_type_key in self.distribution_widgets and
                data_type_key in self.alpha_entries_by_type):
                dist_type = self.distribution_widgets[data_type_key]['dist_type'].get()
                
                if dist_type == 'pseudo-voigt':
                    for param, widgets in self.alpha_entries_by_type[data_type_key].items():
                        _, dist_box, entry_a, entry_b, _ = widgets
                        alpha_dist = dist_box.get()
                        
                        if alpha_dist == 'fixed':
                            params[param] = (alpha_dist, float(entry_a.get()), 0.0)
                        else:
                            params[param] = (alpha_dist, float(entry_a.get()), float(entry_b.get()))
            
            # Generate data
            # If any parameters reference Excel columns, read those numeric values directly
            if col_specs and self.excel_path:
                try:
                    from openpyxl import load_workbook
                    from openpyxl.utils import column_index_from_string
                    wb = load_workbook(self.excel_path, data_only=True)
                    sheet_name = config.get('sheet_name')
                    sheet_key = None
                    for s in wb.sheetnames:
                        if s.lower() == sheet_name.lower():
                            sheet_key = s
                            break
                    if sheet_key is not None:
                        ws = wb[sheet_key]
                        # Determine Position # mapping for selected MMs
                        mm_cfg = self.mm_config_df.copy()
                        if 'Position #' in mm_cfg.columns:
                            mm_cfg['Position #'] = pd.to_numeric(mm_cfg['Position #'], errors='coerce')
                            mm_cfg = mm_cfg[mm_cfg['Position #'].notna()]
                            mm_to_pos = dict(zip(mm_cfg['MM #'].astype(int), mm_cfg['Position #'].astype(int)))
                        else:
                            mm_to_pos = {int(r['MM #']): int(i + 1) for i, (_, r) in enumerate(mm_cfg.iterrows()) if not pd.isna(r.get('MM #'))}

                        selected_positions = [mm_to_pos.get(int(mm)) for mm in sorted(self.selected_mm_numbers)]
                        # Build dataframe with Position # and filled params
                        data_rows = []
                        for pos in selected_positions:
                            row = {'Position #': int(pos) if pos is not None else None}
                            for p_label in config['params']:
                                if p_label in col_specs:
                                    # use first column letter if comma-separated
                                    letter = col_specs[p_label].split(',')[0].strip()
                                    try:
                                        src_idx = column_index_from_string(letter)
                                        # workbook rows start at 1; assume data rows align with Position #
                                        v = ws.cell(row=1 + int(pos), column=src_idx).value
                                    except Exception:
                                        v = None
                                    row[p_label] = v
                                else:
                                    # placeholder, will be filled from generated samples
                                    row[p_label] = None
                            data_rows.append(row)
                        # Build a dataframe of the Excel-sourced values for overlay
                        excel_df = pd.DataFrame(data_rows)
                        # Prepare params for sampling by excluding colref entries
                        sample_params = {k: v for k, v in params.items() if v[0] != 'colref'}
                        sampled_df = generate_data_from_distributions(sample_params, num_mm, config) if sample_params else pd.DataFrame()
                        # Ensure sampled_df has all expected param columns
                        for p_label in config['params']:
                            if p_label not in sampled_df.columns:
                                sampled_df[p_label] = [None] * num_mm
                        # Overlay Excel values for col_specs into the sampled dataframe
                        data_df = sampled_df.reset_index(drop=True)
                        for i, row in excel_df.iterrows():
                            for p_label in config['params']:
                                if p_label in col_specs:
                                    data_df.at[i, p_label] = row.get(p_label)
                    else:
                        # sheet not found; fall back to sampling
                        data_df = generate_data_from_distributions(params, num_mm, config)
                except Exception:
                    data_df = generate_data_from_distributions(params, num_mm, config)
            else:
                data_df = generate_data_from_distributions(params, num_mm, config)
            
            # Ensure parameter columns are in the canonical order defined by config['params']
            try:
                param_order = [p for p in config.get('params', []) if p in data_df.columns]
                other_cols = [c for c in data_df.columns if c not in param_order]
                data_df = data_df.loc[:, param_order + other_cols]
            except Exception:
                pass

            # Add distribution column for MM_PSF BEFORE alpha parameters
            if config.get('has_distribution', False) and data_type_key in self.distribution_widgets:
                dist_type = self.distribution_widgets[data_type_key]['dist_type'].get()
                
                # Insert distribution column after regular params, before alpha params
                # Get position of last regular param
                regular_params = config['params']
                insert_pos = len(regular_params)

                # For fixed pseudo-voigt, ensure distribution column is set to 'pseudo-voigt'
                # (not 'gaussian')
                if dist_type == 'gaussian' and hasattr(self, 'standard_distributions'):
                    # Try to detect if the selected preset is a fixed pseudo-voigt
                    std_combo = self.distribution_widgets[data_type_key].get('std_dist_combo')
                    if std_combo:
                        std_name = std_combo.get()
                        std_def = self.standard_distributions.get(std_name)
                        if std_def and std_def.get('type') == 'pseudo-voigt':
                            dist_type = 'pseudo-voigt'

                # Insert distribution column at the correct position
                data_df.insert(insert_pos, 'distribution', dist_type)

                # If the user selects a Gaussian distribution, alpha parameters are not used.
                # During export we still want to overwrite any existing alpha columns in Excel,
                # so write them explicitly as '-' (instead of leaving old numeric values).
                if dist_type == 'gaussian' and 'alpha_params' in config:
                    after_dist = data_df.columns.get_loc('distribution') + 1
                    # Keep a stable column order: distribution, alpha_rad, alpha_azi
                    if 'alpha_rad' not in data_df.columns:
                        data_df.insert(after_dist, 'alpha_rad', '-')
                        after_dist += 1
                    else:
                        data_df['alpha_rad'] = '-'
                    if 'alpha_azi' not in data_df.columns:
                        data_df.insert(after_dist, 'alpha_azi', '-')
                    else:
                        data_df['alpha_azi'] = '-'

                # For pseudo-voigt, clamp alpha values to [0, 1] range
                if dist_type == 'pseudo-voigt':
                    if 'alpha_rad' in data_df.columns:
                        data_df['alpha_rad'] = np.clip(data_df['alpha_rad'], 0.0, 1.0)
                    if 'alpha_azi' in data_df.columns:
                        data_df['alpha_azi'] = np.clip(data_df['alpha_azi'], 0.0, 1.0)
            
            # Debug: Print column names to verify they have units
            logging.debug("Generated %s data with columns: %s", config['tab_label'], data_df.columns.tolist())
            
            # Add key column.
            # MM_PSF is keyed by MM #. Alignment/Gravity/Thermal are keyed by Position #.
            if data_type_key == 'MM_PSF':
                data_df.insert(0, 'MM #', sorted(self.selected_mm_numbers))
            else:
                if self.mm_config_df is None or 'MM #' not in self.mm_config_df.columns:
                    raise ValueError('MM configuration must include "MM #" to generate position-based sheets.')

                mm_cfg = self.mm_config_df.copy()
                if 'Position #' in mm_cfg.columns:
                    mm_cfg['Position #'] = pd.to_numeric(mm_cfg['Position #'], errors='coerce')
                    mm_cfg = mm_cfg[mm_cfg['Position #'].notna()]
                    mm_to_pos = dict(zip(mm_cfg['MM #'].astype(int), mm_cfg['Position #'].astype(int)))
                else:
                    # Backward-compatible fallback: infer Position # from row order
                    mm_to_pos = {int(r['MM #']): int(i + 1) for i, (_, r) in enumerate(mm_cfg.iterrows()) if not pd.isna(r.get('MM #'))}

                selected_positions = []
                for mm in sorted(self.selected_mm_numbers):
                    pos = mm_to_pos.get(int(mm))
                    if pos is None:
                        raise ValueError(f'Could not determine Position # for MM # {mm}.')
                    selected_positions.append(int(pos))

                data_df.insert(0, 'Position #', selected_positions)
            
            self.data_dfs[data_type_key] = data_df
            # Inform user that data was generated and is pending export; do not use persistent in-window label
            try:
                messagebox.showinfo('Success', f"{config['tab_label']} data generated for {num_mm} selected MMs!\n\n(Pending export)")
            except Exception:
                pass
            try:
                if data_type_key in self.tab_status_labels:
                    self.tab_status_labels[data_type_key].configure(text='')
            except Exception:
                pass
            self.update_preview()
        except ValueError as e:
            messagebox.showerror('Error', str(e))

    def build_preview_tab(self):
        """Build the Preview & Export tab UI.

        The tab displays a textual preview of the generated per-MM tables and
        exposes simple export controls. Export supports overwriting the loaded
        workbook or saving to a new file while preserving existing sheets and
        formulae where possible.
        """
        frame = self.tab_preview
        ttk.Label(frame, text='Preview & Export', font=('Arial', 14)).pack(pady=10)
        
        preview_frame = ttk.Frame(frame)
        preview_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.preview_text = tk.Text(preview_frame, height=25, width=130)
        self.preview_text.pack()
        
        export_frame = ttk.Frame(frame)
        export_frame.pack(fill='x', padx=5, pady=5)

        ttk.Label(export_frame, text='Export destination:').grid(row=0, column=0, sticky='w', padx=2)

        ttk.Radiobutton(export_frame, text='Overwrite loaded file', variable=self.export_mode_var, 
                        value='current', command=self.update_export_controls).grid(row=1, column=0, sticky='w', padx=2, pady=2)
        ttk.Radiobutton(export_frame, text='Save as new file', variable=self.export_mode_var, 
                        value='new', command=self.update_export_controls).grid(row=2, column=0, sticky='w', padx=2, pady=2)

        self.export_path_entry = ttk.Entry(export_frame, textvariable=self.export_path_var, width=70)
        self.export_path_entry.grid(row=2, column=1, padx=5, pady=2, sticky='we')
        ttk.Button(export_frame, text='Browse...', command=self.browse_export_path).grid(row=2, column=2, padx=2, pady=2)

        export_frame.columnconfigure(1, weight=1)

        ttk.Button(frame, text='Export to Excel', command=self.export_to_excel).pack(pady=10)
        self.update_export_controls()

    def update_export_controls(self):
        """Enable or disable the export path entry depending on mode.

        When `export_mode_var` is `'current'` the path entry is disabled since
        the loaded file will be updated. When `'new'` the entry is enabled so
        the user can choose a destination path.
        """
        mode = self.export_mode_var.get()
        if mode == 'current':
            self.export_path_entry.state(['disabled'])
        else:
            self.export_path_entry.state(['!disabled'])

    def browse_export_path(self):
        """Ask the user for a save-path when exporting to a new file.

        Updates `self.export_path_var` when the user selects a path.
        """
        path = filedialog.asksaveasfilename(initialdir="./Distributions", defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx *.xls')])
        if path:
            self.export_path_var.set(path)

    def update_preview(self):
        """Write a brief textual preview of selected MM configuration and generated data.

        Shows the first few rows of each generated dataframe to help the user
        verify content before exporting.
        """
        self.preview_text.delete('1.0', tk.END)

        if self.mm_config_df is not None:
            config_preview = 'MM Configuration (Selected):\n'
            selected_config = self.mm_config_df[self.mm_config_df['MM #'].isin(self.selected_mm_numbers)]
            config_preview += str(selected_config.head(10))
            self.preview_text.insert(tk.END, config_preview + '\n\n')

        # Only show preview for enabled data types
        for data_type_key, config in DATA_TYPES.items():
            if data_type_key in self.enabled_data_types and self.data_dfs[data_type_key] is not None:
                self.preview_text.insert(tk.END, f"{config['tab_label']} Preview:\n")
                self.preview_text.insert(tk.END, str(self.data_dfs[data_type_key].head(20)) + '\n\n')

    def export_to_excel(self):
        """Export generated data and/or pending A_eff edits to an Excel workbook.

        Writes generated dataframes to their corresponding sheets, updates the
        A_eff sheet when pending edits exist, and (optionally) copies the
        selected vignetting column into column B of vignetting sheets.
        """

        # Check if any data has been generated for enabled types OR A_eff has pending edits
        has_data = any(self.data_dfs[dt] is not None for dt in self.enabled_data_types)
        if not has_data and not self.aeff_pending_export:
            messagebox.showwarning('No Data', 'Please generate data for at least one enabled type first (or apply A_eff edits).')
            return

        # Determine target path
        target_path = None
        if self.export_mode_var.get() == 'current':
            target_path = self.excel_path
            if target_path is None:
                messagebox.showwarning('No File', 'Please load an Excel file first or choose "Save as new file".')
                return
        else:
            target_path = self.export_path_var.get().strip()
            if not target_path:
                target_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx *.xls')])
                if not target_path:
                    return
                self.export_path_var.set(target_path)
        
        try:
            import time as _time
            logging.debug("export_to_excel: start export at %s", _time.time())
            # Load existing workbook to preserve formatting and formulas
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Load existing workbook or create new one
            if self.excel_path:
                logging.debug("export_to_excel: loading workbook %s at %s", self.excel_path, _time.time())
                # Always load existing workbook to preserve formatting, even when saving to new file
                wb = load_workbook(self.excel_path)
                logging.debug("export_to_excel: loaded workbook, sheets=%s at %s", wb.sheetnames[:6], _time.time())
            else:
                # Create new workbook only if no file was loaded
                from openpyxl import Workbook
                wb = Workbook()
                # Remove default sheet if present
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # Update each data type (only for enabled types with generated data)
            selected_set = set(self.selected_mm_numbers)
            
            for data_type_key, config in DATA_TYPES.items():
                if data_type_key not in self.enabled_data_types:
                    continue
                
                generated = self.data_dfs[data_type_key]
                if generated is None:
                    continue

                # Per-MM export cleanup for MM_PSF:
                # If a particular MM row uses a Gaussian PSF, alpha parameters are not used.
                # Write '-' for alpha_rad/alpha_azi for those rows so any previous numeric
                # values in the Excel sheet get overwritten.
                if data_type_key == 'MM_PSF' and 'distribution' in generated.columns:
                    g = generated.copy()
                    dist = g['distribution'].astype(str).str.lower().fillna('')
                    gaussian_mask = (dist == 'gaussian')

                    # Custom PSF: distribution holds the PSF file stem.
                    # Set all MM_PSF parameter/shift columns to '-' except 'distribution'.
                    custom_mask = ~dist.isin(['gaussian', 'pseudo-voigt', 'voigt']) & dist.ne('')
                    if custom_mask.any():
                        for col in ['m_rad [arcsec]', 'm_azi [arcsec]', 'sigma_rad [arcsec]', 'sigma_azi [arcsec]', 'alpha_rad', 'alpha_azi']:
                            if col in g.columns:
                                g.loc[custom_mask, col] = '-'

                    if 'alpha_rad' not in g.columns:
                        try:
                            insert_at = g.columns.get_loc('distribution') + 1
                            g.insert(insert_at, 'alpha_rad', None)
                        except Exception:
                            g['alpha_rad'] = None
                    if 'alpha_azi' not in g.columns:
                        try:
                            insert_at = g.columns.get_loc('distribution') + 1
                            # If alpha_rad was inserted, alpha_azi goes after it
                            if 'alpha_rad' in g.columns:
                                insert_at = g.columns.get_loc('alpha_rad') + 1
                            g.insert(insert_at, 'alpha_azi', None)
                        except Exception:
                            g['alpha_azi'] = None

                    # Use NaN for numeric alpha columns to avoid dtype conflicts
                    import numpy as _np
                    if 'alpha_rad' in g.columns:
                        # preserve object columns (custom PSF) but set numeric entries to NaN
                        try:
                            g['alpha_rad'] = pd.to_numeric(g['alpha_rad'], errors='coerce')
                            g.loc[gaussian_mask, 'alpha_rad'] = _np.nan
                        except Exception:
                            g.loc[gaussian_mask, 'alpha_rad'] = None
                    if 'alpha_azi' in g.columns:
                        try:
                            g['alpha_azi'] = pd.to_numeric(g['alpha_azi'], errors='coerce')
                            g.loc[gaussian_mask, 'alpha_azi'] = _np.nan
                        except Exception:
                            g.loc[gaussian_mask, 'alpha_azi'] = None
                    generated = g

                # If a standard preset with variable sigmas was selected, sample
                # per-MM numeric sigma values now (deterministic per target file
                # and preset) and write them into the canonical sigma columns so
                # columns B/E in the workbook are numeric for each of the MMs.
                try:
                    if data_type_key == 'MM_PSF' and self.psf_mode_var.get() == 'standard' and 'std_dist_combo' in self.distribution_widgets.get(data_type_key, {}):
                        import numpy as _np, hashlib
                        from pathlib import Path as _Path
                        target_name = _Path(target_path).name if target_path else (_Path(self.excel_path).name if self.excel_path else 'export.xlsx')
                        # Iterate rows and sample per-row values from the preset definition
                        for idx, row in generated.iterrows():
                            preset_name = str(row.get('distribution') or '').strip()
                            if not preset_name:
                                continue
                            std_def = self.standard_distributions.get(preset_name)
                            if not isinstance(std_def, dict):
                                continue
                            # Detect Variable presets and parse percent/alpha if present
                            force_gamma = False
                            var_pct = None
                            parsed_alpha = None
                            try:
                                lname = preset_name.lower()
                                if 'variable' in lname or '%' in preset_name:
                                    force_gamma = True
                                    import re as _re
                                    m = _re.search(r'(\d+)\s*%\s*variable', preset_name, flags=_re.IGNORECASE)
                                    if not m:
                                        m = _re.search(r'^(\d+)\s*%\b', preset_name)
                                    if m:
                                        var_pct = float(m.group(1)) / 100.0
                                    m2 = _re.search(r'alpha\s*[:(]?\s*(\d+)\s*%?', preset_name, flags=_re.IGNORECASE)
                                    if m2:
                                        parsed_alpha = float(m2.group(1)) / 100.0
                            except Exception:
                                force_gamma = False

                            # Build deterministic RNG per (file, preset)
                            h = int(hashlib.sha256((target_name + str(preset_name)).encode('utf-8')).hexdigest()[:8], 16)
                            rng = _np.random.default_rng(h + int(idx))

                            for ui_param, std_param in (('sigma_rad [arcsec]', 'sigma_rad'), ('sigma_azi [arcsec]', 'sigma_azi')):
                                try:
                                    pdef = std_def.get(std_param)
                                    if not isinstance(pdef, dict):
                                        continue
                                    mu = float(pdef.get('mean', 0.0)) if pdef.get('mean') is not None else 0.0
                                    sigma = float(pdef.get('sigma', 0.0)) if pdef.get('sigma') is not None else 0.0
                                    # If preset indicates Variable(...) but the table cell omitted
                                    # an explicit sigma, derive sigma from the parsed percent
                                    if force_gamma and (sigma <= 0) and mu > 0 and var_pct is not None:
                                        sigma = abs(mu * float(var_pct))
                                    # If distribution describes variability (gaussian/gamma),
                                    # convert mean/std to a gamma shape/scale and sample.
                                    if pdef.get('dist') in ('gaussian', 'gamma') and mu > 0 and sigma > 0:
                                        k = (mu / sigma) ** 2
                                        theta = (sigma ** 2) / mu
                                        val = float(rng.gamma(shape=k, scale=theta, size=1)[0])
                                    else:
                                        # fallback to fixed mean
                                        val = mu if mu > 0 else 1e-6
                                    # Enforce positive floor
                                    if val <= 0:
                                        val = 1e-6
                                    generated.at[idx, ui_param] = float(val)
                                except Exception:
                                    continue
                            # Optional: override alpha params when preset encodes a parsed alpha
                            if parsed_alpha is not None:
                                for alpha_ui, alpha_param in (('alpha_rad', 'alpha_rad'), ('alpha_azi', 'alpha_azi')):
                                    try:
                                        pdef = std_def.get(alpha_param)
                                        if not isinstance(pdef, dict):
                                            continue
                                        mu_a = float(pdef.get('mean', 0.0)) if pdef.get('mean') is not None else None
                                        # If an explicit mean exists, construct gamma from parsed_alpha
                                        if mu_a is not None and mu_a > 0:
                                            sigma_a = max(abs(mu_a * float(parsed_alpha)), 1e-12)
                                            k = (mu_a / sigma_a) ** 2
                                            theta = (sigma_a ** 2) / mu_a
                                            aval = float(rng.gamma(shape=k, scale=theta, size=1)[0])
                                            aval = min(max(aval, 0.0), 1.0)
                                            generated.at[idx, alpha_ui] = float(aval)
                                    except Exception:
                                        continue
                except Exception:
                    pass
                
                sheet_name = config['sheet_name']
                
                # Get or create worksheet
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
                
                # Read header row to find column indices
                if ws.max_row > 0:
                    header = [cell.value for cell in ws[1]]
                else:
                    header = []
                
                # Build column mapping from generated dataframe
                gen_columns = list(generated.columns)

                # Determine key column (MM # for PSF, Position # for the perturbation sheets)
                key_col = 'Position #' if 'Position #' in gen_columns else 'MM #'
                
                # If worksheet is empty or missing the key column, write all data from scratch
                if not header or key_col not in header:
                    # Clear sheet
                    ws.delete_rows(1, ws.max_row)
                    # Write all data including headers
                    for r_idx, row in enumerate(dataframe_to_rows(generated, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    continue
                
                # Find key column index
                key_col_idx = header.index(key_col) + 1
                
                # Find existing columns and add any missing ones
                existing_cols = {col: idx + 1 for idx, col in enumerate(header) if col}
                
                # Add new columns if they don't exist
                next_col = len(header) + 1
                for col in gen_columns:
                    if col not in existing_cols:
                        existing_cols[col] = next_col
                        ws.cell(row=1, column=next_col, value=col)
                        next_col += 1
                
                # Build a map of MM# to row index in the worksheet
                key_to_row = {}
                for row_idx in range(2, ws.max_row + 1):
                    k = ws.cell(row=row_idx, column=key_col_idx).value
                    if k is not None:
                        key_to_row[k] = row_idx
                
                # Update existing rows or add new rows
                rows_to_add = []
                for _, gen_row in generated.iterrows():
                    k = gen_row[key_col]
                    
                    if k in key_to_row:
                        # Update existing row
                        row_idx = key_to_row[k]
                        for col_name, value in gen_row.items():
                            col_idx = existing_cols[col_name]
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    else:
                        # Collect rows to add
                        rows_to_add.append(gen_row)
                
                # Add new rows at the end
                next_row = ws.max_row + 1
                for gen_row in rows_to_add:
                    for col_name, value in gen_row.items():
                        col_idx = existing_cols[col_name]
                        ws.cell(row=next_row, column=col_idx, value=value)
                    next_row += 1

            # Update A_eff sheet column B for selected MMs (if modified via A_eff tab
            # or if there are in-memory A_eff weights available). Previously the
            # GUI only updated the sheet when `aeff_pending_export` was True; in
            # some flows the pending flag can be cleared or missed which left the
            # A_eff column unwritten. Treat any in-memory weights as intent to
            # export so GUI exports reliably populate column B.
            if self.aeff_pending_export or (hasattr(self, 'aeff_weights') and bool(self.aeff_weights)):
                if 'A_eff' in wb.sheetnames:
                    ws = wb['A_eff']
                else:
                    ws = wb.create_sheet('A_eff')
                    ws.cell(row=1, column=1, value='MM #')
                    ws.cell(row=1, column=2, value='A_eff')

                # Build MM# -> row mapping from column A
                mm_to_row = {}
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(row=r, column=1).value
                    if v is None:
                        continue
                    try:
                        mm_to_row[int(float(v))] = r
                    except Exception:
                        continue

                mode = self.aeff_mode_var.get()
                if mode == 'fixed':
                    # When fixed mode is used, set entire column B to the fixed value
                    raw = self.aeff_fixed_var.get().strip()
                    if not raw:
                        raise ValueError('Please enter a fixed A_eff value before exporting')
                    fixed_val = float(raw)
                    # Ensure there is at least one MM row; if none, add selected MMs
                    if ws.max_row < 2 and self.selected_mm_numbers:
                        for mm in self.selected_mm_numbers:
                            r = ws.max_row + 1
                            ws.cell(row=r, column=1, value=int(mm))
                    # Overwrite A_eff values for all rows (from row 2..max_row)
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=2, value=float(fixed_val))
                    # Ensure GUI export does not populate column C — clear any values
                    for r in range(2, ws.max_row + 1):
                        try:
                            ws.cell(row=r, column=3, value=None)
                        except Exception:
                            pass
                    # Also update in-memory weights map for completeness
                    for mm in list(mm_to_row.keys()):
                        try:
                            self.aeff_weights[int(mm)] = float(fixed_val)
                        except Exception:
                            continue
                else:
                    # Standard preset: write the preset expression into column B for
                    # the selected MMs so GUI exports record the chosen A_eff
                    # definition (e.g. "gaussian(L,10%*L)") rather than numeric
                    # per-MM values. This keeps CLI/main behavior (which writes
                    # adjusted numeric A_eff into column C) unchanged.
                    preset_name = str(self.aeff_selected_preset_var.get()).strip()
                    expr_text = None
                    if preset_name:
                        expr_text = self.aeff_standard_presets.get(preset_name)
                    # Fallback to the displayed preset name if no Values expression
                    if not expr_text:
                        expr_text = preset_name or ''

                    # Evaluate the preset per-MM (numeric) and write numeric A_eff
                    for mm in self.selected_mm_numbers:
                        mm_i = int(mm)
                        if mm_i not in mm_to_row:
                            # Append missing MM row
                            r = ws.max_row + 1
                            ws.cell(row=r, column=1, value=mm_i)
                            mm_to_row[mm_i] = r
                        try:
                            # Use the same evaluator used by the Apply action so
                            # randomness and sampling are consistent.
                            val = float(self._evaluate_aeff_preset_for_mm(mm_i, expr_text))
                        except Exception:
                            # If evaluation fails, fall back to any numeric weight
                            # already present or leave blank.
                            val = None
                            if mm_i in self.aeff_weights:
                                try:
                                    val = float(self.aeff_weights[mm_i])
                                except Exception:
                                    val = None
                        # If evaluator returned NaN or None, attempt a percent-based
                        # fallback: parse percent from preset name and use the
                        # canonical A_eff_base (column B) as the mean if present.
                        try:
                            import math
                            if val is None or (isinstance(val, float) and math.isnan(val)):
                                # try parse percent from preset_name
                                pct = None
                                try:
                                    m_pct = re.search(r"(\d+(?:\.\d+)?)\s*%", preset_name)
                                    if m_pct:
                                        pct = float(m_pct.group(1)) / 100.0
                                except Exception:
                                    pct = None
                                # try to read A_eff_base from aeff_raw_df col index 1
                                base_val = None
                                try:
                                    row_idx = self._get_aeff_row_for_mm(mm_i)
                                    if row_idx is not None and self.aeff_raw_df is not None and self.aeff_raw_df.shape[1] > 1:
                                        bv = self.aeff_raw_df.iloc[row_idx, 1]
                                        base_val = float(bv) if bv is not None and not pd.isna(bv) else None
                                except Exception:
                                    base_val = None
                                if base_val is not None and pct is not None:
                                    import numpy as _np
                                    sigma = abs(float(base_val) * float(pct))
                                    val = float(base_val + _np.random.normal(loc=0.0, scale=sigma))
                                elif base_val is not None and val is None:
                                    val = float(base_val)
                        except Exception:
                            pass
                        if val is not None:
                            ws.cell(row=mm_to_row[mm_i], column=2, value=float(val))
                            # Keep in-memory weights in sync
                            try:
                                self.aeff_weights[int(mm_i)] = float(val)
                            except Exception:
                                pass

                    # Clear column C for all rows so GUI export never writes adjusted A_eff
                    for r in range(2, ws.max_row + 1):
                        try:
                            ws.cell(row=r, column=3, value=None)
                        except Exception:
                            pass

                    # Ensure column B contains the final numeric A_eff values for every MM row.
                    # Preference: `self.aeff_weights` (user-applied GUI changes) then raw sheet
                    # second-column values when available.
                    try:
                        # Build a mapping of MM -> numeric A_eff to write. Preference:
                        # 1) explicit GUI mapping `self.aeff_weights` (already applied)
                        # 2) If none, evaluate current UI selection (fixed or standard)
                        # 3) fallback to the raw A_eff sheet second-column values
                        mapping = {}
                        try:
                            if hasattr(self, 'aeff_weights') and self.aeff_weights:
                                mapping.update({int(k): float(v) for k, v in self.aeff_weights.items()})
                        except Exception:
                            pass

                        # If mapping is empty, evaluate current UI selection for selected MMs.
                        # If no MMs are selected, evaluate for all MMs present in the loaded MM configuration.
                        if not mapping:
                            try:
                                # Determine the MM list to evaluate: prefer explicit selection,
                                # otherwise fall back to all MMs from the loaded MM configuration.
                                mm_list = list(self.selected_mm_numbers) if self.selected_mm_numbers else []
                                if not mm_list and getattr(self, 'mm_config_df', None) is not None:
                                    try:
                                        mm_list = [int(float(x)) for x in self.mm_config_df['MM #'].dropna().tolist()]
                                    except Exception:
                                        mm_list = []

                                mode = self.aeff_mode_var.get()
                                if mode == 'fixed':
                                    raw = self.aeff_fixed_var.get().strip()
                                    if raw:
                                        fixed = float(raw)
                                        for mm in mm_list:
                                            mapping[int(mm)] = float(fixed)
                                else:
                                    preset = str(self.aeff_selected_preset_var.get()).strip()
                                    expr = self.aeff_standard_presets.get(preset)
                                    if preset and expr and mm_list:
                                        for mm in mm_list:
                                            try:
                                                v = self._evaluate_aeff_preset_for_mm(int(mm), expr)
                                                mapping[int(mm)] = float(v)
                                            except Exception:
                                                # leave missing MM out of mapping so fallback can apply
                                                continue
                            except Exception:
                                pass

                        # Determine selected energy preset name (if any) for special-case wide-table A_eff sheets
                        sel_preset_name = None
                        try:
                            # Prefer explicit free-energy control when present, otherwise preset name
                            fe = str(self.aeff_free_energy_var.get()).strip() if hasattr(self, 'aeff_free_energy_var') else ''
                            if fe:
                                sel_preset_name = fe
                            else:
                                pn = str(self.aeff_selected_preset_var.get()).strip() if hasattr(self, 'aeff_selected_preset_var') else ''
                                if pn:
                                    sel_preset_name = pn
                        except Exception:
                            sel_preset_name = None

                        # Attempt to locate energy column in the workbook sheet when sel_preset_name looks like an energy token
                        energy_col_index = None
                        header_row_idx = None
                        try:
                            if sel_preset_name and ws is not None:
                                # Find a sensible header row (search first 1..8 rows for a cell containing 'MM' or 'MM #')
                                for hr in range(1, min(9, ws.max_row) + 1):
                                    for c in range(1, min(40, ws.max_column) + 1):
                                        try:
                                            v = ws.cell(row=hr, column=c).value
                                            if v is None:
                                                continue
                                            s = str(v).strip()
                                            if re.search(r"^MM\s*#?$", s, flags=re.IGNORECASE) or re.search(r"mm\s*#", s, flags=re.IGNORECASE):
                                                header_row_idx = hr
                                                break
                                        except Exception:
                                            continue
                                    if header_row_idx is not None:
                                        break
                                if header_row_idx is None:
                                    header_row_idx = 1

                                # Find column whose header matches the energy token (exact or contains)
                                for c in range(1, min(200, ws.max_column) + 1):
                                    try:
                                        hv = ws.cell(row=header_row_idx, column=c).value
                                        if hv is None:
                                            continue
                                        hs = str(hv).strip()
                                        if hs == sel_preset_name or sel_preset_name in hs or hs in sel_preset_name:
                                            energy_col_index = c
                                            break
                                    except Exception:
                                        continue
                        except Exception:
                            energy_col_index = None

                        # Now write mapping where present, otherwise fall back to raw sheet values
                        for r in range(2, ws.max_row + 1):
                            try:
                                mm_val = ws.cell(row=r, column=1).value
                                if mm_val is None:
                                    continue
                                mm_i = int(float(mm_val))
                            except Exception:
                                continue

                            if mm_i in mapping:
                                try:
                                    ws.cell(row=r, column=2, value=float(mapping[mm_i]))
                                except Exception:
                                    pass
                                continue

                            # Special-case: if we have a selected energy column, try to copy the per-row
                            # energy value when the row's 'Standard distributions' cell matches the preset.
                            try:
                                # detect the 'Standard distributions' column index from header row
                                std_col_idx = None
                                try:
                                    # prefer header label match in header_row_idx
                                    if header_row_idx is None:
                                        # attempt to find header row similarly as above
                                        for hr in range(1, min(9, ws.max_row) + 1):
                                            for c in range(1, min(40, ws.max_column) + 1):
                                                try:
                                                    v = ws.cell(row=hr, column=c).value
                                                    if v is None:
                                                        continue
                                                    s = str(v).strip().lower()
                                                    if 'standard' in s and 'a_eff' in s.replace(' ', '') or 'standard' in s:
                                                        std_col_idx = c
                                                        break
                                                except Exception:
                                                    continue
                                            if std_col_idx is not None:
                                                break
                                except Exception:
                                    std_col_idx = None

                                # fallback column index D
                                if std_col_idx is None:
                                    std_col_idx = 4

                                got_from_wide = False
                                if energy_col_index is not None:
                                    std_val = ws.cell(row=r, column=std_col_idx).value
                                    if std_val is not None and str(std_val).strip() and sel_preset_name and str(std_val).strip() == str(sel_preset_name).strip():
                                        try:
                                            ev = ws.cell(row=r, column=energy_col_index).value
                                            if ev is not None:
                                                ws.cell(row=r, column=2, value=float(ev))
                                                got_from_wide = True
                                        except Exception:
                                            got_from_wide = False
                                if got_from_wide:
                                    continue
                                # Try to locate the energy column in the loaded raw dataframe (preferred, contains evaluated numbers).
                                # Look for a column whose header contains the energy token (e.g. '@10 keV') and has numeric values.
                                try:
                                    if self.aeff_raw_df is not None and sel_preset_name:
                                        # Scan headers (row 0) to find the energy column
                                        preset_col = None
                                        for j in range(min(self.aeff_raw_df.shape[1], 200)):
                                            try:
                                                header_val = str(self.aeff_raw_df.iloc[0, j]).strip()
                                                # Match: header contains the preset energy (e.g. '10 keV' in 'A_eff @10 keV')
                                                if str(sel_preset_name).strip() in header_val:
                                                    preset_col = j
                                                    break
                                            except Exception:
                                                continue
                                        if preset_col is not None:
                                            # find the row index in the raw df for this MM
                                            row_idx = self._get_aeff_row_for_mm(mm_i)
                                            if row_idx is not None and preset_col < self.aeff_raw_df.shape[1]:
                                                try:
                                                    val = self.aeff_raw_df.iloc[row_idx, preset_col]
                                                    if not pd.isna(val):
                                                        ws.cell(row=r, column=2, value=float(val))
                                                        continue
                                                except Exception:
                                                    pass
                                except Exception:
                                    pass
                            except Exception:
                                pass

                            # fallback: copy existing second-column from loaded raw df if available
                            try:
                                if self.aeff_raw_df is not None:
                                    row_idx = self._get_aeff_row_for_mm(mm_i)
                                    if row_idx is not None and self.aeff_raw_df.shape[1] > 1:
                                        bv = self.aeff_raw_df.iloc[row_idx, 1]
                                        if not pd.isna(bv):
                                            try:
                                                ws.cell(row=r, column=2, value=float(bv))
                                            except Exception:
                                                pass
                            except Exception:
                                pass
                    except Exception:
                        pass

            # Ensure A_eff sheet C2 stores the selected/preset energy when exporting.
            try:
                sel_energy_for_aeff = None
                if self.aeff_mode_var.get() == 'fixed':
                    # take numeric value from free-energy dropdown when fixed
                    free_e = str(self.aeff_free_energy_var.get()).strip()
                    if free_e:
                        m = re.search(r"(\d+(?:\.\d*)?)", free_e)
                        if m:
                            try:
                                sel_energy_for_aeff = float(m.group(1))
                            except Exception:
                                sel_energy_for_aeff = None
                else:
                    preset_name = str(self.aeff_selected_preset_var.get()).strip()
                    if preset_name:
                        m = re.search(r"(\d+(?:\.\d*)?)\s*(?:keV)?", preset_name, flags=re.IGNORECASE)
                        if m:
                            try:
                                sel_energy_for_aeff = float(m.group(1))
                            except Exception:
                                sel_energy_for_aeff = None
                # Do NOT write selected energy into the A_eff sheet C2.
                # The vignetting sheets' C2 values are updated below instead.
            except Exception:
                pass

            # Update vignetting sheet selected-energy cell C2 based on A_eff choice.
            try:
                sel_energy = None
                # fixed mode: read selected energy from free-energy combobox
                if self.aeff_mode_var.get() == 'fixed':
                    fe = str(self.aeff_free_energy_var.get()).strip()
                    if fe:
                        m = re.search(r"(\d+(?:\.\d*)?)", fe)
                        if m:
                            try:
                                sel_energy = float(m.group(1))
                            except Exception:
                                sel_energy = None
                else:
                    # standard mode: prefer explicit free-energy control if set,
                    # otherwise parse the preset name preferring a number followed
                    # by 'keV' and falling back to the last numeric token.
                    fe = str(self.aeff_free_energy_var.get()).strip()
                    if fe:
                        m = re.search(r"(\d+(?:\.\d*)?)", fe)
                        if m:
                            try:
                                sel_energy = float(m.group(1))
                            except Exception:
                                sel_energy = None
                    if sel_energy is None:
                        preset_name = str(self.aeff_selected_preset_var.get()).strip()
                        if preset_name:
                            # prefer numeric token immediately followed by 'keV'
                            m = re.search(r"(\d+(?:\.\d*)?)\s*(?:keV)", preset_name, flags=re.IGNORECASE)
                            if not m:
                                all_nums = re.findall(r"(\d+(?:\.\d*)?)", preset_name)
                                sval = all_nums[-1] if all_nums else None
                            else:
                                sval = m.group(1)
                            if sval is not None:
                                try:
                                    sel_energy = float(sval)
                                except Exception:
                                    sel_energy = None

                # Fallback: if still None, try to read from existing vignetting C2
                if sel_energy is None:
                    try:
                        name = _pick_sheet_in_wb(wb, VIG_ROT_AZI_CANDIDATES)
                        if name:
                            vws = wb[name]
                            vval = vws.cell(row=2, column=3).value
                            if vval is not None:
                                sel_energy = float(vval)
                    except Exception:
                        sel_energy = None
                if sel_energy is None:
                    try:
                        name = _pick_sheet_in_wb(wb, VIG_ROT_RAD_CANDIDATES)
                        if name:
                            vws = wb[name]
                            vval = vws.cell(row=2, column=3).value
                            if vval is not None:
                                sel_energy = float(vval)
                    except Exception:
                        sel_energy = None

                # Write selected energy into C2 of both vignetting sheets if present
                if sel_energy is not None:
                    try:
                        name = _pick_sheet_in_wb(wb, VIG_ROT_AZI_CANDIDATES)
                        if name:
                            vws = wb[name]
                            vws.cell(row=2, column=3, value=float(sel_energy))
                    except Exception:
                        pass
                    try:
                        name = _pick_sheet_in_wb(wb, VIG_ROT_RAD_CANDIDATES)
                        if name:
                            vws = wb[name]
                            vws.cell(row=2, column=3, value=float(sel_energy))
                    except Exception:
                        pass
            except Exception:
                # Non-fatal: do not block export on vignetting metadata update
                pass
            
            # Save workbook in a background thread to avoid blocking the GUI
            import threading as _threading
            import os as _os
            import tempfile as _tempfile
            from pathlib import Path as _Path
            import time as _time

            try:
                progress_win = tk.Toplevel(self.root)
                progress_win.title('Saving...')
                ttk.Label(progress_win, text='Saving file, please wait...').pack(padx=20, pady=20)
                progress_win.transient(self.root)
                try:
                    progress_win.grab_set()
                except Exception:
                    pass
                progress_win.update()
            except Exception:
                progress_win = None

            save_result = {'ok': False, 'error': None}

            def _save_worker():
                try:
                    dirp = _Path(target_path).parent if target_path else None
                    if dirp is None:
                        wb.save(target_path)
                    else:
                        tf = _tempfile.NamedTemporaryFile(prefix='.tmp_export_', suffix='.xlsx', dir=str(dirp), delete=False)
                        tf.close()
                        tmpname = tf.name
                        wb.save(tmpname)
                        try:
                            _os.replace(tmpname, target_path)
                        except Exception:
                            wb.save(target_path)
                    save_result['ok'] = True
                except Exception as _e:
                    save_result['error'] = str(_e)
                finally:
                    def _finish():
                        try:
                            if progress_win is not None:
                                try:
                                    progress_win.grab_release()
                                except Exception:
                                    pass
                                try:
                                    progress_win.destroy()
                                except Exception:
                                    pass
                            if save_result['ok']:
                                self.aeff_pending_export = False
                                try:
                                    messagebox.showinfo('Success', f'File saved to:\n{target_path}\nwith data for {len(self.selected_mm_numbers)} MMs across multiple sheets!')
                                except Exception:
                                    pass
                            else:
                                try:
                                    messagebox.showerror('Error', save_result['error'])
                                except Exception:
                                    pass
                        finally:
                            return
                    try:
                        self.root.after(50, _finish)
                    except Exception:
                        _finish()

            t = _threading.Thread(target=_save_worker, daemon=True)
            t.start()
            # Return now; UI will be updated when save completes in background
            return
        except Exception as e:
            messagebox.showerror('Error', str(e))


def apply_vignetting_to_workbook(target_path, preset=None, verbose=False):
    """Apply per-position vignette factors to `target_path` workbook.

    Writes vignette factors into column B of the vignetting sheets and writes
    adjusted A_eff into column C of the A_eff sheet. Returns a summary dict.
    """
    import pandas as _pd
    import numpy as _np
    import re as _re
    from openpyxl import load_workbook
    from pathlib import Path as _Path
    import tempfile as _tempfile

    wb = load_workbook(target_path)

    tf = _tempfile.NamedTemporaryFile(prefix='gui_vig_', suffix='.xlsx', delete=False)
    tf.close()
    tmp_read_path = tf.name
    try:
        wb.save(tmp_read_path)
    except Exception:
        tmp_read_path = target_path

    mm_to_pos = {}
    pos_to_cfg_row = {}
    mm_config_map = {}
    try:
        mmc_df = _pd.read_excel(tmp_read_path, sheet_name='MM configuration', engine='openpyxl')
        if 'MM #' in mmc_df.columns:
            for order_i, (_, row) in enumerate(mmc_df.iterrows()):
                mm_num = row.get('MM #')
                if _pd.isna(mm_num):
                    continue
                mm_num_i = int(mm_num)
                if 'Position #' in mmc_df.columns:
                    pos_val = row.get('Position #')
                    if not _pd.isna(pos_val):
                        try:
                            mm_to_pos[mm_num_i] = int(float(pos_val))
                        except Exception:
                            pass
                    if mm_num_i not in mm_to_pos:
                        mm_to_pos[mm_num_i] = int(order_i) + 1
                else:
                    mm_to_pos[mm_num_i] = int(order_i) + 1
                cfg_row_number = int(order_i) + 1
                pos_for_row = mm_to_pos.get(mm_num_i, cfg_row_number)
                pos_to_cfg_row[pos_for_row] = cfg_row_number
                mm_config_map[mm_num_i] = {
                    'x_MM': row.get('x_MM [m]', 0),
                    'y_MM': row.get('y_MM [m]', 0),
                    'z_MM': row.get('z_MM [m]', 0),
                    'r_MM': row.get('r_MM [m]', 0),
                }
    except Exception:
        pass

    if verbose:
        logging.debug('apply_vignetting_to_workbook: mm_to_pos=%d pos_to_cfg_row=%d', len(mm_to_pos), len(pos_to_cfg_row))

    def _read_pos_map(sheet_name, rot_keys):
        out = {}
        try:
            dfp = _pd.read_excel(tmp_read_path, sheet_name=sheet_name, engine='openpyxl')
            if 'Position #' in dfp.columns:
                for _, r in dfp.iterrows():
                    pos = r.get('Position #')
                    if _pd.isna(pos):
                        continue
                    pos_i = int(pos)
                    out[pos_i] = {k: r.get(k, 0.0) for k in rot_keys}
            elif 'MM #' in dfp.columns and mm_to_pos:
                for _, r in dfp.iterrows():
                    mmn = r.get('MM #')
                    if _pd.isna(mmn):
                        continue
                    pos = mm_to_pos.get(int(mmn))
                    if pos is None:
                        continue
                    out[pos] = {k: r.get(k, 0.0) for k in rot_keys}
        except Exception:
            pass
        return out

    alignment_by_pos = _read_pos_map('Alignment', ['d_align_rotrad', 'd_align_rotazi'])
    gravity_by_pos = _read_pos_map('Gravity offload', ['d_grav_rotx', 'd_grav_roty', 'd_grav_rotrad', 'd_grav_rotazi'])
    thermal_by_pos = _read_pos_map('Thermal', ['d_therm_rotx', 'd_therm_roty', 'd_therm_rotrad', 'd_therm_rotazi'])

    try:
        import sys as _sys
        _sys.path.insert(0, str(_Path(__file__).resolve().parents[0]))
        import main as _main
        _, _, rot_rad_map, rot_azi_map = _main.compute_total_rot_polar(mm_to_pos, mm_config_map, alignment_by_pos, gravity_by_pos, thermal_by_pos)
    except Exception:
        rot_rad_map = {}
        rot_azi_map = {}

    if verbose:
        logging.debug('apply_vignetting_to_workbook: rot maps sizes -> rot_rad=%d rot_azi=%d', len(rot_rad_map), len(rot_azi_map))

    def _parse_vig(sheet_name):
        ys_by_pos = {}
        try:
            vdf = _pd.read_excel(tmp_read_path, sheet_name=sheet_name, engine='openpyxl', header=None)
            if vdf is None or vdf.empty or vdf.shape[1] < 2:
                return ys_by_pos
            if vdf.shape[1] >= 11:
                col_H = vdf.iloc[:, 7]
                if col_H.notna().any():
                    for _, r in vdf.iterrows():
                        try:
                            cfg_row = r.iloc[7]
                            if _pd.isna(cfg_row):
                                continue
                            cfg_row = int(float(cfg_row))
                        except Exception:
                            continue
                        energy_marker = r.iloc[9] if vdf.shape[1] > 9 else None
                        try:
                            # column I holds rot delta in arcmin; convert to arcsec
                            xval = float(r.iloc[8]) * 60.0 if not _pd.isna(r.iloc[8]) else None
                        except Exception:
                            xval = None
                        try:
                            yval = float(r.iloc[10]) if vdf.shape[1] > 10 and not _pd.isna(r.iloc[10]) else None
                        except Exception:
                            yval = None
                        if xval is None or yval is None:
                            continue
                        key_str = (cfg_row, str(energy_marker).strip())
                        key_num = None
                        try:
                            key_num = (cfg_row, float(energy_marker))
                        except Exception:
                            key_num = None
                        if key_str not in ys_by_pos:
                            ys_by_pos[key_str] = {'xs': [], 'ys': []}
                        ys_by_pos[key_str]['xs'].append(xval)
                        ys_by_pos[key_str]['ys'].append(yval)
                        if key_num is not None:
                            if key_num not in ys_by_pos:
                                ys_by_pos[key_num] = {'xs': [], 'ys': []}
                            ys_by_pos[key_num]['xs'].append(xval)
                            ys_by_pos[key_num]['ys'].append(yval)
                    for k, v in list(ys_by_pos.items()):
                        order = _np.argsort(v['xs'])
                        xs_sorted = _np.array(v['xs'], dtype=float)[order]
                        ys_sorted = _np.array(v['ys'], dtype=float)[order]
                        ys_by_pos[k] = (xs_sorted, ys_sorted)
        except Exception:
            pass
        return ys_by_pos

    ys_by_pos_azi = _parse_vig('Vignetting rotazi')
    ys_by_pos_rad = _parse_vig('Vignetting rotrad')

    if verbose:
        logging.debug('apply_vignetting_to_workbook: parsed vignetting -> azi keys=%d rad keys=%d', len(ys_by_pos_azi), len(ys_by_pos_rad))

    sel_energy = None
    if preset:
        m = _re.search(r"(\d+(?:\.\d*)?)\s*(?:keV)?", str(preset), flags=_re.IGNORECASE)
        if m:
            try:
                sel_energy = float(m.group(1))
            except Exception:
                sel_energy = None
    # If no preset energy was supplied, attempt to read energy from vignetting
    # sheet cell C2 (row=2,col=3). Prefer 'Vignetting rotazi' then 'Vignetting rotrad'.
    if sel_energy is None:
        try:
            if 'Vignetting rotazi' in wb.sheetnames:
                vws = wb['Vignetting rotazi']
                vval = vws.cell(row=2, column=3).value
                if vval is not None:
                    sel_energy = float(vval)
        except Exception:
            sel_energy = None
    if sel_energy is None:
        try:
            if 'Vignetting rotrad' in wb.sheetnames:
                vws = wb['Vignetting rotrad']
                vval = vws.cell(row=2, column=3).value
                if vval is not None:
                    sel_energy = float(vval)
        except Exception:
            sel_energy = None

    vig_vals_azi = {}
    vig_vals_rad = {}
    for mm, pos in mm_to_pos.items():
        cfg_row = pos_to_cfg_row.get(pos)
        f_azi = 1.0
        try:
            used = False
            if cfg_row is not None and sel_energy is not None and (cfg_row, float(sel_energy)) in ys_by_pos_azi:
                xs, ys = ys_by_pos_azi[(cfg_row, float(sel_energy))]
                f_azi = float(_np.interp(abs(float(rot_azi_map.get(pos, 0.0))), xs, ys))
                used = True
            if not used and cfg_row is not None and (cfg_row, str(preset).strip()) in ys_by_pos_azi:
                xs, ys = ys_by_pos_azi[(cfg_row, str(preset).strip())]
                f_azi = float(_np.interp(abs(float(rot_azi_map.get(pos, 0.0))), xs, ys))
                used = True
            if not used and cfg_row is not None:
                matches = [k for k in ys_by_pos_azi.keys() if k[0] == cfg_row]
                if matches:
                    xs, ys = ys_by_pos_azi[matches[0]]
                    f_azi = float(_np.interp(abs(float(rot_azi_map.get(pos, 0.0))), xs, ys))
        except Exception:
            f_azi = 1.0

        f_rad = 1.0
        try:
            used = False
            if cfg_row is not None and sel_energy is not None and (cfg_row, float(sel_energy)) in ys_by_pos_rad:
                xs, ys = ys_by_pos_rad[(cfg_row, float(sel_energy))]
                f_rad = float(_np.interp(abs(float(rot_rad_map.get(pos, 0.0))), xs, ys))
                used = True
            if not used and cfg_row is not None and (cfg_row, str(preset).strip()) in ys_by_pos_rad:
                xs, ys = ys_by_pos_rad[(cfg_row, str(preset).strip())]
                f_rad = float(_np.interp(abs(float(rot_rad_map.get(pos, 0.0))), xs, ys))
                used = True
            if not used and cfg_row is not None:
                matches = [k for k in ys_by_pos_rad.keys() if k[0] == cfg_row]
                if matches:
                    xs, ys = ys_by_pos_rad[matches[0]]
                    f_rad = float(_np.interp(abs(float(rot_rad_map.get(pos, 0.0))), xs, ys))
        except Exception:
            f_rad = 1.0

        vig_vals_azi[pos] = float(f_azi)
        vig_vals_rad[pos] = float(f_rad)

    if verbose:
        logging.debug('apply_vignetting_to_workbook: computed vig_vals counts -> azi=%d rad=%d', len(vig_vals_azi), len(vig_vals_rad))

    def _write_vig_sheet_in_wb(wb_obj, sheet_name, vig_map):
        if sheet_name not in wb_obj.sheetnames:
            return 0
        ws = wb_obj[sheet_name]
        pos_to_row = {}
        try:
            max_col = min(40, ws.max_column or 40)
        except Exception:
            max_col = 40
        # Use fast row iteration to avoid accessing ws._cells repeatedly
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True), start=1):
            for cell in row:
                try:
                    if cell is None:
                        continue
                    if isinstance(cell, (int, float)) and float(cell).is_integer():
                        v = int(cell)
                    else:
                        s = str(cell).strip()
                        if s.isdigit():
                            v = int(s)
                        else:
                            continue
                    if v not in pos_to_row:
                        pos_to_row[v] = row_idx
                except Exception:
                    continue
        count = 0
        for pos_k, val in (vig_map or {}).items():
            if pos_k in pos_to_row:
                r = pos_to_row[pos_k]
                ws.cell(row=r, column=2, value=float(val))
                count += 1
        return count

    c1 = _write_vig_sheet_in_wb(wb, 'Vignetting rotazi', vig_vals_azi)
    c2 = _write_vig_sheet_in_wb(wb, 'Vignetting rotrad', vig_vals_rad)

    c3 = 0
    if 'A_eff' in wb.sheetnames:
        ws_a = wb['A_eff']
        mm_to_row = {}
        for r in range(2, ws_a.max_row+1):
            try:
                mm = ws_a.cell(row=r, column=1).value
                if mm is None:
                    continue
                mm_to_row[int(float(mm))] = r
            except Exception:
                continue
        for mm, pos in mm_to_pos.items():
            r = mm_to_row.get(mm)
            if r is None:
                continue
            base = ws_a.cell(row=r, column=2).value
            try:
                base_f = float(base)
            except Exception:
                base_f = 1.0
            f_azi = vig_vals_azi.get(pos, 1.0)
            f_rad = vig_vals_rad.get(pos, 1.0)
            adj = base_f * float(f_azi) * float(f_rad)
            # Write the (possibly vignetted) A_eff into the canonical A_eff
            # column (column B) so exported A_eff values appear in column B
            # as expected by downstream tools/users. Previously this wrote
            # adjusted values into column C which caused confusion.
            ws_a.cell(row=r, column=2, value=float(adj))
            c3 += 1

        # Ensure any GUI-driven A_eff choice (aeff_weights) is reflected for
        # all MM rows in column B. Preference order when deciding what to
        # write for an MM row:
        # 1. `self.aeff_weights` mapping (if provided by caller via closure)
        # 2. existing numeric value in the raw A_eff sheet second column
        # 3. leave cell as-is
        try:
            # aeff_weights may not be in this scope when called as a helper;
            # attempt to fetch from outer closure (gui export) otherwise skip
            aeff_weights_local = locals().get('self', None)
        except Exception:
            aeff_weights_local = None

    try:
        if verbose:
            logging.debug('apply_vignetting_to_workbook: saving workbook to %s counts: %d %d %d', target_path, c1, c2, c3)
        wb.save(target_path)
    except Exception:
        pass

    return {'rotazi_written': c1, 'rotrad_written': c2, 'aeff_adjusted_written': c3}


def sync_aeff_column_b_in_workbook(wb, aeff_weights: dict | None = None, aeff_raw_df: 'pd.DataFrame | None' = None):
    """Ensure column B of the `A_eff` sheet contains the canonical numeric
    A_eff for each MM row. Preference order:
    1. `aeff_weights` mapping (MM -> numeric)
    2. `aeff_raw_df` second column value for that MM (if present)

    This is offered as a standalone helper so unit tests can validate GUI
    export semantics without instantiating the full UI.
    """
    try:
        if 'A_eff' not in wb.sheetnames:
            return 0
        ws = wb['A_eff']
        count = 0
        for r in range(2, ws.max_row + 1):
            try:
                mmv = ws.cell(row=r, column=1).value
                if mmv is None:
                    continue
                mm_i = int(float(mmv))
            except Exception:
                continue
            written = False
            try:
                if aeff_weights and mm_i in aeff_weights:
                    ws.cell(row=r, column=2, value=float(aeff_weights[mm_i]))
                    written = True
                    count += 1
            except Exception:
                written = False
            if not written and aeff_raw_df is not None:
                try:
                    # Attempt to find row index in raw df; assume first column lists MM
                    # Use the same logic as ExtendedGUI._get_aeff_row_for_mm
                    row_idx = None
                    try:
                        if 1 <= mm_i <= aeff_raw_df.shape[0]:
                            v = aeff_raw_df.iloc[mm_i - 1, 0]
                            if not pd.isna(v) and int(float(v)) == mm_i:
                                row_idx = mm_i - 1
                    except Exception:
                        row_idx = None
                    if row_idx is None:
                        # fallback search
                        col = aeff_raw_df.iloc[:, 0]
                        for idx, x in enumerate(col):
                            try:
                                if pd.isna(x):
                                    continue
                                if int(float(x)) == mm_i:
                                    row_idx = idx
                                    break
                            except Exception:
                                continue
                    if row_idx is not None and aeff_raw_df.shape[1] > 1:
                        bv = aeff_raw_df.iloc[row_idx, 1]
                        if not pd.isna(bv):
                            ws.cell(row=r, column=2, value=float(bv))
                            count += 1
                except Exception:
                    pass
        return count
    except Exception:
        return 0


if __name__ == '__main__':
    root = tk.Tk()
    apply_macos_input_fixes(root)
    app = ExtendedGUI(root)
    root.mainloop()