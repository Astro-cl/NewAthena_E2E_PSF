#!/usr/bin/env python3
"""Debug lookup table reading."""

import openpyxl

wb = openpyxl.load_workbook('Distributions/test10kev.xlsx')
aeff = wb['A_eff']

print("Source data (Rows 4-20, Columns T-AA):")
print("\nRow  | S | T | U | V | W | X | Y | Z | AA")
print("-" * 80)

for row_num in range(4, 21):
    s_val = aeff.cell(row_num, 19).value  # Column S (Row label)
    
    row_vals = [str(s_val)[:6] if s_val is not None else "-"]
    for col_idx in range(20, 28):  # T to AA (20 to 27)
        cell_val = aeff.cell(row_num, col_idx).value
        if isinstance(cell_val, float):
            row_vals.append(f"{cell_val:.3f}")
        elif cell_val is None:
            row_vals.append("-")
        else:
            row_vals.append(str(cell_val)[:6])
    
    print(f"{row_num:3}  | " + " | ".join(row_vals))
