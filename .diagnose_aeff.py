import glob, os, importlib.util
from openpyxl import load_workbook

# locate a Distributions file with _E__7_
files = glob.glob('Distributions/*_E__7_*')
if not files:
    print('NO_7KEV_FILES')
    raise SystemExit(0)
fp = files[0]
print('FILE:', fp)

# load main.py evaluator without triggering CLI
spec = importlib.util.spec_from_file_location('mymain','main.py')
m = importlib.util.module_from_spec(spec)
spec.loader.exec_module(m)
eval_fn = getattr(m, '_evaluate_vlookup_xlookup', None)
print('Evaluator found:', bool(eval_fn))

# open workbook normal and data_only
wb = load_workbook(fp, data_only=False)
wb_vals = load_workbook(fp, data_only=True)
if 'A_eff' not in wb.sheetnames:
    print('NO_AEFF_SHEET')
    raise SystemExit(0)
ws = wb['A_eff']
ws_vals = wb_vals['A_eff'] if 'A_eff' in wb_vals.sheetnames else None

# print mapping rows D/E up to row 40
print('\nMapping rows (row, D, E):')
for r in range(1, min(ws.max_row, 40)+1):
    d = ws.cell(row=r, column=4).value
    e = ws.cell(row=r, column=5).value
    if d is not None and e is not None:
        print(r, repr(d), '->', repr(e))

# find mapping for energy ~7
import re
chosen = None
for r in range(1, min(ws.max_row, 40)+1):
    d = ws.cell(row=r, column=4).value
    e = ws.cell(row=r, column=5).value
    if d is None or e is None:
        continue
    m = re.search(r"(\d+(?:\.\d*)?)", str(d))
    if not m:
        continue
    if abs(float(m.group(1)) - 7.0) < 1e-6:
        chosen = (r,e)
        break
print('Chosen mapping row:', chosen)

# resolve chosen to column index if possible
from openpyxl.utils import column_index_from_string
if chosen:
    r,e = chosen
    src_idx = None
    if isinstance(e,(int,float)):
        src_idx = int(e)
    else:
        s = str(e).strip()
        if s.isalpha():
            try:
                src_idx = column_index_from_string(s.upper())
            except Exception:
                src_idx = None
        else:
            m2 = re.search(r"(\d+)", s)
            if m2:
                src_idx = int(m2.group(1))
    print('Resolved src_idx:', src_idx)
    # print a few sample rows in that column
    if src_idx is not None:
        print('\nSample cells from chosen column (row, formula/value, cached):')
        for rr in range(2, min(ws.max_row, 12)):
            cell = ws.cell(row=rr, column=src_idx)
            val = cell.value
            cached = None
            if ws_vals is not None:
                try:
                    cached = ws_vals.cell(row=rr, column=src_idx).value
                except Exception:
                    cached = None
            print(rr, repr(val)[:200], 'CACHED->', repr(cached)[:200])
            if isinstance(val, str) and val.startswith('=') and eval_fn is not None:
                try:
                    ev = eval_fn(val, wb, ws, rr)
                except Exception as e:
                    ev = f'EXC:{e}'
                print(' EVAL->', ev)

        # Attempt to write evaluated values into column B of a copy and save
        outp = fp.replace('.xlsx', '_diag_modified.xlsx')
        print('\nWriting evaluated values into', outp)
        wb_copy = load_workbook(fp, data_only=False)
        ws_copy = wb_copy['A_eff']
        for rr in range(2, ws.max_row+1):
            mmcell = ws.cell(row=rr, column=1).value
            try:
                if mmcell is None:
                    continue
                _ = int(float(mmcell))
            except Exception:
                continue
            cell = ws.cell(row=rr, column=src_idx)
            val = cell.value
            evaluated = None
            if isinstance(val, str) and val.startswith('=') and eval_fn is not None:
                try:
                    evaluated = eval_fn(val, wb, ws, rr)
                except Exception:
                    evaluated = None
            elif isinstance(val, (int, float)):
                evaluated = float(val)
            # fallback to adjusted col 3
            if evaluated is None:
                try:
                    alt = ws.cell(row=rr, column=3).value
                    if alt is not None:
                        evaluated = float(alt)
                except Exception:
                    evaluated = None
            if evaluated is not None:
                try:
                    ws_copy.cell(row=rr, column=2).value = float(evaluated)
                except Exception:
                    pass
        wb_copy.save(outp)
        # reopen saved file and show a few B samples
        wb_new = load_workbook(outp, data_only=True)
        ws_new = wb_new['A_eff']
        print('\nModified file sample column B:')
        for rr in range(2, min(ws_new.max_row, 12)):
            try:
                print(rr, ws_new.cell(row=rr, column=2).value)
            except Exception:
                print(rr, 'ERR')
