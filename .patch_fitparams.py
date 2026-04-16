from openpyxl import load_workbook
import os,glob
pkg_dir='Exports/Export_reallistic_20260416_112020/2__Ang__20__E__7__Def__0_reallistic_20260416_112145'
fit_fp=os.path.join(pkg_dir,'fitparams_aeffloss.xlsx')
wb_fp=None
if os.path.exists(fit_fp):
    wb_fp=fit_fp
else:
    cand=os.path.join(pkg_dir,'EEF_fittingparams.xlsx')
    if os.path.exists(cand):
        wb_fp=cand
if not wb_fp:
    print('No fitparams workbook found to patch')
    raise SystemExit(1)
# compute sums from workbook inside package
xls=None
for cand in os.listdir(pkg_dir):
    if cand.lower().endswith('.xlsx') and not cand.lower().startswith('eef_fittingparams') and 'fitparams' not in cand.lower():
        xls=os.path.join(pkg_dir,cand)
        break
if not xls:
    print('No exported workbook found in package')
    raise SystemExit(1)
wb_mod=load_workbook(xls,data_only=True)
if 'A_eff' not in wb_mod.sheetnames:
    print('No A_eff sheet in exported workbook')
    raise SystemExit(1)
ws=wb_mod['A_eff']
sumB=0.0
sumC=0.0
for r in range(2, ws.max_row+1):
    mm=ws.cell(row=r,column=1).value
    if mm is None: continue
    try:
        float(mm)
    except Exception:
        continue
    vB=ws.cell(row=r,column=2).value
    vC=ws.cell(row=r,column=3).value
    try: sumB += float(vB) if vB is not None else 0.0
    except: pass
    try: sumC += float(vC) if vC is not None else 0.0
    except: pass
# patch fitparams workbook (rewrite values)
wb2=load_workbook(wb_fp)
if 'Fit_parameters' not in wb2.sheetnames:
    print('Fit_parameters sheet not present in',wb_fp)
    raise SystemExit(1)
ws2=wb2['Fit_parameters']
# find rows
name_to_row={}
for r in range(1, ws2.max_row+1):
    k=ws2.cell(row=r,column=1).value
    if k is None: continue
    name_to_row[str(k)]=r
# write or append
def write_param(name, val):
    if name in name_to_row:
        r=name_to_row[name]
        ws2.cell(row=r,column=2).value = float(val) if val is not None else None
    else:
        nr=ws2.max_row+1
        ws2.cell(row=nr,column=1).value = name
        ws2.cell(row=nr,column=2).value = float(val) if val is not None else None

write_param('Aeff_sum_orig', sumB)
write_param('Aeff_sum_mod', sumC)
if sumB is not None and sumB!=0:
    aeff_loss = 1.0 - (sumC / sumB)
else:
    aeff_loss = None
write_param('Aeff_loss', aeff_loss)
wb2.save(wb_fp)
print('Patched', wb_fp, 'with sumB=',sumB,'sumC=',sumC,'loss=',aeff_loss)
