import argparse
import math
import random
import os
import shutil
from copy import deepcopy
from itertools import permutations
from multiprocessing import Pool, cpu_count
import time
from functools import partial
from pathlib import Path

import numpy as np
import pandas as pd
from distributions_rotated import gaussian_2d_rotated, pseudo_voigt_2d_rotated


def _theta_degrees_from_xy(x: float, y: float) -> float:
    """Match compute_theta_maps() logic but for a single slot."""
    theta_ccw = float(np.degrees(np.arctan2(x, y)))
    theta_cw_from_y = theta_ccw if theta_ccw >= 0 else 180.0 + theta_ccw
    return -(theta_cw_from_y - 90.0)


def _mux_muy_for_slot(m_rad: float, m_azi: float, x_mm: float, y_mm: float, r_mm: float) -> tuple[float, float]:
    r = float(r_mm) if float(r_mm) != 0.0 else 1e-9
    u_rad_x = float(x_mm) / r
    u_rad_y = float(y_mm) / r
    u_azi_x = -float(y_mm) / r
    u_azi_y = float(x_mm) / r
    mux = u_rad_x * float(m_rad) + u_azi_x * float(m_azi)
    muy = u_rad_y * float(m_rad) + u_azi_y * float(m_azi)
    return mux, muy


def _radius_for_fraction(frac: np.ndarray, r: np.ndarray, target: float = 0.5) -> float:
    """Estimate radius where cumulative fraction == target using local cubic fit.

    Falls back to linear interpolation when the local window is too small or
    a valid cubic root cannot be found.
    """
    if frac.size == 0:
        return 0.0
    # Ensure arrays are 1D and aligned
    frac = np.asarray(frac).ravel()
    r = np.asarray(r).ravel()
    # Quick linear fallback when shapes mismatch
    if frac.size != r.size or frac.size < 2:
        return float(np.interp(target, frac, r))

    idx = int(np.searchsorted(frac, target))
    # Choose up to 4 points around the crossing for a local cubic fit
    start = max(0, idx - 2)
    end = min(frac.size, idx + 2)
    # If we don't have enough points, use linear interp
    if end - start < 2:
        return float(np.interp(target, frac, r))

    r_win = r[start:end]
    f_win = frac[start:end]

    # If fewer than 4 points, prefer linear interpolation to avoid poor fits
    if f_win.size < 4:
        return float(np.interp(target, frac, r))

    try:
        coeffs = np.polyfit(r_win, f_win, 3)
        coeffs[-1] -= float(target)
        roots = np.roots(coeffs)
        real_roots = roots[np.isreal(roots)].real
        rmin, rmax = float(r_win.min()), float(r_win.max())
        for rt in real_roots:
            if rmin - 1e-12 <= float(rt) <= rmax + 1e-12:
                return float(rt)
    except Exception:
        # Fall through to linear interp on any numerical issue
        pass

    return float(np.interp(target, frac, r))


class IncrementalHEWApprox:
    """Incremental HEW objective on a fixed polar grid.

    This is intentionally approximate: we keep the polar sampling grid fixed
    (center + r_max) based on the starting configuration. This makes swap
    evaluation incremental (O(1) distributions instead of O(N)).
    """

    def __init__(
        self,
        base_params: pd.DataFrame,
        mm_config: pd.DataFrame,
        alignment_by_pos: dict[int, dict] | None = None,
        gravity_by_pos: dict[int, dict] | None = None,
        thermal_by_pos: dict[int, dict] | None = None,
        n_r: int = 60,
        n_theta: int = 40,
    ):
        self.normalize = True
        self.n_r = int(n_r)
        self.n_theta = int(n_theta)

        self.alignment_by_pos = alignment_by_pos or {}
        self.gravity_by_pos = gravity_by_pos or {}
        self.thermal_by_pos = thermal_by_pos or {}

        self.slot_indices: list[int] = mm_config.index.tolist()
        self.idx_to_pos: dict[int, int] = {idx: pos for pos, idx in enumerate(self.slot_indices)}

        # Slot (physical location) data
        x = mm_config["x_MM [m]"].astype(float).to_numpy()
        y = mm_config["y_MM [m]"].astype(float).to_numpy()
        z = mm_config["z_MM [m]"].astype(float).to_numpy() if "z_MM [m]" in mm_config.columns else np.zeros(len(mm_config), dtype=float)
        r = mm_config["r_MM [m]"].astype(float).to_numpy()
        self.slot_x = x
        self.slot_y = y
        self.slot_z = z
        self.slot_r = r
        self.slot_theta_deg = np.array([_theta_degrees_from_xy(float(xx), float(yy)) for xx, yy in zip(x, y)], dtype=float)

        # Slot Position # values (authoritative join key for Alignment/Gravity/Thermal).
        if "Position #" in mm_config.columns:
            pos_series = pd.to_numeric(mm_config["Position #"], errors="coerce")
            fallback = pd.Series(range(1, len(mm_config) + 1), index=mm_config.index, dtype=float)
            pos_series = pos_series.fillna(fallback)
            self.slot_posnum = pos_series.astype(int).to_numpy()
        else:
            self.slot_posnum = np.arange(1, len(mm_config) + 1, dtype=int)

        # MM assignment per slot position (same ordering as mm_config rows)
        self.assignment = mm_config["MM #"].astype(int).to_numpy().copy()

        # Fast lookup of MM params by MM#
        needed_cols = [
            "MM #",
            "m_rad",
            "m_azi",
            "sigma_rad",
            "sigma_azi",
            "weight",
            "distribution",
            "alpha_azi",
            "alpha_rad",
        ]
        self.params = base_params[needed_cols].copy()
        self.params["MM #"] = self.params["MM #"].astype(int)
        self.param_by_mm = {int(row["MM #"]): row for _, row in self.params.iterrows()}

        # Fix the polar grid based on the starting configuration.
        df0 = rebuild_df(base_params, mm_config)
        total_weight = float(df0["weight"].sum())
        if total_weight <= 0:
            self.cx0 = 0.0
            self.cy0 = 0.0
        else:
            self.cx0 = float((df0["mux"] * df0["weight"]).sum() / total_weight)
            self.cy0 = float((df0["muy"] * df0["weight"]).sum() / total_weight)

        max_sigma = float(max(df0["sigmax"].max(), df0["sigmay"].max()))
        max_center = float(np.sqrt((df0["mux"] - self.cx0) ** 2 + (df0["muy"] - self.cy0) ** 2).max())
        self.r_max = max(max_center + 3.0 * max_sigma, 1e-6)

        theta = np.linspace(0.0, 2.0 * np.pi, self.n_theta, endpoint=False)
        r_grid = np.linspace(0.0, self.r_max, self.n_r)
        self.theta = theta
        self.r = r_grid
        self.dtheta = float(theta[1] - theta[0]) if self.n_theta > 1 else 2.0 * math.pi
        self.dr = float(r_grid[1] - r_grid[0]) if self.n_r > 1 else float(self.r_max)
        self.R, self.TH = np.meshgrid(r_grid, theta)
        self.Xp = self.cx0 + self.R * np.cos(self.TH)
        self.Yp = self.cy0 + self.R * np.sin(self.TH)

        # Build initial summed PSF on the fixed grid.
        self.Zp = np.zeros_like(self.Xp)
        for pos in range(len(self.assignment)):
            mm_num = int(self.assignment[pos])
            self.Zp += self._contribution(mm_num, pos)

    def _contribution(self, mm_num: int, slot_pos: int) -> np.ndarray:
        p = self.param_by_mm.get(int(mm_num))
        if p is None:
            return np.zeros_like(self.Xp)

        # Position number comes from MM configuration (Position #) when available.
        pos_num = int(self.slot_posnum[int(slot_pos)])

        m_rad = float(p["m_rad"])
        m_azi = float(p["m_azi"])

        # Per-position alignment deltas in polar components
        if pos_num in self.alignment_by_pos:
            m_rad += float(self.alignment_by_pos[pos_num].get("d_align_rad", 0.0))
            m_azi += float(self.alignment_by_pos[pos_num].get("d_align_azi", 0.0))

        # rotz coupling (sum of contributions)
        d_rotz_arcsec = 0.0
        if pos_num in self.alignment_by_pos:
            d_rotz_arcsec += float(self.alignment_by_pos[pos_num].get("d_align_rotz", 0.0))
        if pos_num in self.gravity_by_pos:
            d_rotz_arcsec += float(self.gravity_by_pos[pos_num].get("d_grav_rotz", 0.0))
        if pos_num in self.thermal_by_pos:
            d_rotz_arcsec += float(self.thermal_by_pos[pos_num].get("d_therm_rotz", 0.0))
        if d_rotz_arcsec != 0.0:
            d_rotz_rad = np.radians(d_rotz_arcsec / 3600.0)
            m_azi += float(self.slot_r[slot_pos]) * d_rotz_rad

        mux, muy = _mux_muy_for_slot(
            m_rad,
            m_azi,
            float(self.slot_x[slot_pos]),
            float(self.slot_y[slot_pos]),
            float(self.slot_r[slot_pos]),
        )
        sigmax = float(p["sigma_rad"])
        sigmay = float(p["sigma_azi"])
        theta_deg = float(self.slot_theta_deg[slot_pos])
        weight = float(p["weight"])
        dist_type = str(p.get("distribution", "gaussian"))

        # Apply per-position xy offsets and dz projection after polar conversion.
        mux2 = float(mux)
        muy2 = float(muy)

        if pos_num in self.gravity_by_pos:
            mux2 += float(self.gravity_by_pos[pos_num].get("d_grav_x", 0.0))
            muy2 += float(self.gravity_by_pos[pos_num].get("d_grav_y", 0.0))
        if pos_num in self.thermal_by_pos:
            mux2 += float(self.thermal_by_pos[pos_num].get("d_therm_x", 0.0))
            muy2 += float(self.thermal_by_pos[pos_num].get("d_therm_y", 0.0))

        dz = 0.0
        if pos_num in self.alignment_by_pos:
            dz += float(self.alignment_by_pos[pos_num].get("d_align_z", 0.0))
        if pos_num in self.gravity_by_pos:
            dz += float(self.gravity_by_pos[pos_num].get("d_grav_z", 0.0))
        if pos_num in self.thermal_by_pos:
            dz += float(self.thermal_by_pos[pos_num].get("d_therm_z", 0.0))
        if dz != 0.0:
            denom = 12.0 - float(self.slot_z[slot_pos])
            if denom != 0.0:
                mux2 += dz * float(self.slot_x[slot_pos]) / denom
                muy2 += dz * float(self.slot_y[slot_pos]) / denom

        if dist_type in ["pseudo-voigt", "voigt"]:
            return pseudo_voigt_2d_rotated(
                self.Xp,
                self.Yp,
                muazi=mux2,
                murad=muy2,
                sigmaazi=sigmax,
                sigmarad=sigmay,
                theta=theta_deg,
                alphaazi=float(p.get("alpha_azi", 0.5)),
                alpharad=float(p.get("alpha_rad", 0.5)),
                amplitude=weight,
                normalize=self.normalize,
                degrees=True,
            )

        return gaussian_2d_rotated(
            self.Xp,
            self.Yp,
            mux=mux2,
            muy=muy2,
            sigmax=sigmax,
            sigmay=sigmay,
            theta=theta_deg,
            amplitude=weight,
            normalize=self.normalize,
            degrees=True,
        )

    def hew(self) -> float:
        radial_energy = np.sum(self.Zp * self.R, axis=0) * self.dtheta
        cumulative = np.cumsum(radial_energy * self.dr)
        total_energy = float(cumulative[-1]) if cumulative.size else 1.0
        frac = cumulative / total_energy if total_energy > 0 else cumulative
        # Use local cubic interpolation for the 50% radius, then return HEW diameter
        radius50 = _radius_for_fraction(frac, self.r, target=0.5)
        return float(2.0 * radius50)

    def swap_slots(self, pos_a: int, pos_b: int) -> None:
        """Apply a swap between two slot positions and update Zp incrementally."""
        if pos_a == pos_b:
            return
        a_mm = int(self.assignment[pos_a])
        b_mm = int(self.assignment[pos_b])

        # Remove old contributions
        self.Zp -= self._contribution(a_mm, pos_a)
        self.Zp -= self._contribution(b_mm, pos_b)

        # Swap assignment
        self.assignment[pos_a], self.assignment[pos_b] = self.assignment[pos_b], self.assignment[pos_a]

        # Add new contributions
        self.Zp += self._contribution(b_mm, pos_a)
        self.Zp += self._contribution(a_mm, pos_b)


def _azimuthal_place_mm_config(
    mm_config: pd.DataFrame,
    base_params: pd.DataFrame,
    alignment_by_pos: dict[int, dict] | None = None,
    gravity_by_pos: dict[int, dict] | None = None,
    thermal_by_pos: dict[int, dict] | None = None,
    seed: int = 42,
) -> pd.DataFrame:
    """In-memory azimuthal placement (no Excel IO)."""
    rng = random.Random(seed)
    placed = mm_config.copy()

    mm_numbers = sorted(placed["MM #"].dropna().unique().astype(int).tolist())
    mm_hew_pairs: list[tuple[int, float]] = []

    # Local cache for this run
    hew_cache: dict[int, float] = {}
    for mm_num in mm_numbers:
        if mm_num not in hew_cache:
            hew_cache[mm_num] = compute_mm_hew_at_origin(
                base_params,
                mm_num,
                mm_config_df=mm_config,
                alignment_by_pos=alignment_by_pos,
                gravity_by_pos=gravity_by_pos,
                thermal_by_pos=thermal_by_pos,
            )
        mm_hew_pairs.append((mm_num, hew_cache[mm_num]))

    mm_hew_pairs.sort(key=lambda x: x[1])
    sorted_mms = [mm for mm, _ in mm_hew_pairs]

    positions = []
    for idx, row in placed.iterrows():
        x = float(row.get("x_MM [m]", 0.0))
        y = float(row.get("y_MM [m]", 0.0))
        theta = float(np.degrees(np.arctan2(y, x)))
        if theta < 0:
            theta += 360.0
        positions.append({"index": idx, "theta": theta})

    positions.sort(key=lambda p: p["theta"])
    assigned = [False] * len(positions)
    new_assignment: dict[int, int] = {}

    def find_closest_free_position(target_theta: float) -> int | None:
        best_idx = None
        best_diff = float("inf")
        for i, pos in enumerate(positions):
            if assigned[i]:
                continue
            diff = abs(float(pos["theta"]) - float(target_theta))
            diff = min(diff, 360.0 - diff)
            if diff < best_diff:
                best_diff = diff
                best_idx = i
        return best_idx

    start_theta_offset = rng.uniform(0.0, 360.0)

    for group_start in range(0, len(sorted_mms), 4):
        group = sorted_mms[group_start : group_start + 4]

        if len(group) >= 1:
            idx1 = find_closest_free_position(start_theta_offset % 360.0)
            if idx1 is None:
                break
            assigned[idx1] = True
            new_assignment[positions[idx1]["index"]] = int(group[0])
            ref_theta_1 = float(positions[idx1]["theta"])
            start_theta_offset = (ref_theta_1 + 15.0) % 360.0

        if len(group) >= 2:
            idx2 = find_closest_free_position((ref_theta_1 + 180.0) % 360.0)
            if idx2 is not None:
                assigned[idx2] = True
                new_assignment[positions[idx2]["index"]] = int(group[1])

        if len(group) >= 3:
            idx3 = find_closest_free_position((ref_theta_1 + 90.0) % 360.0)
            if idx3 is None:
                continue
            assigned[idx3] = True
            new_assignment[positions[idx3]["index"]] = int(group[2])
            ref_theta_3 = float(positions[idx3]["theta"])

        if len(group) >= 4:
            idx4 = find_closest_free_position((ref_theta_3 + 180.0) % 360.0)
            if idx4 is not None:
                assigned[idx4] = True
                new_assignment[positions[idx4]["index"]] = int(group[3])

    for pos_idx, mm_num in new_assignment.items():
        placed.at[pos_idx, "MM #"] = int(mm_num)

    return placed


def _place_mm_config_within_rows(
    mm_config: pd.DataFrame,
    base_params: pd.DataFrame,
    placer,
    alignment_by_pos: dict[int, dict] | None = None,
    gravity_by_pos: dict[int, dict] | None = None,
    thermal_by_pos: dict[int, dict] | None = None,
    seed: int = 42,
) -> pd.DataFrame:
    """Apply a placement strategy but *only* by permuting within each Row #.

    If 'Row #' is missing, falls back to applying the placer to the whole table.
    """
    placed = mm_config.copy()
    if "Row #" not in placed.columns:
        return placer(
            placed,
            base_params,
            alignment_by_pos=alignment_by_pos,
            gravity_by_pos=gravity_by_pos,
            thermal_by_pos=thermal_by_pos,
            seed=seed,
        )

    # Deterministic row order
    row_values = [v for v in placed["Row #"].dropna().unique().tolist()]
    try:
        row_values = sorted(row_values)
    except Exception:
        pass

    for rv in row_values:
        idxs = placed.index[placed["Row #"] == rv].tolist()
        if len(idxs) < 2:
            continue
        sub = placed.loc[idxs].copy()
        sub2 = placer(
            sub,
            base_params,
            alignment_by_pos=alignment_by_pos,
            gravity_by_pos=gravity_by_pos,
            thermal_by_pos=thermal_by_pos,
            seed=seed,
        )
        # Only patch MM assignment back into the full table.
        if "MM #" in sub2.columns:
            placed.loc[idxs, "MM #"] = sub2.loc[idxs, "MM #"].astype(sub["MM #"].dtype if "MM #" in sub.columns else sub2["MM #"].dtype)

    return placed


def _xaxis_place_mm_config(
    mm_config: pd.DataFrame,
    base_params: pd.DataFrame,
    alignment_by_pos: dict[int, dict] | None = None,
    gravity_by_pos: dict[int, dict] | None = None,
    thermal_by_pos: dict[int, dict] | None = None,
    seed: int = 42,
) -> pd.DataFrame:
    """In-memory placement biased toward the +/-x axis.

    Strategy (deterministic, geometry-driven):
    1) Sort MMs by individual HEW (best to worst)
    2) Place best MM as close as possible to +x axis (theta≈0°)
    3) Place 2nd best as close as possible to -x axis (theta≈180°)
     4) Place remaining MMs alternating between +x and -x sides
     5) Within each side, alternate above and below the x-axis (y>0 then y<0)
         while staying as close as possible (in azimuth) to that side's axis.

    Notes:
    - Angles are computed from slot coordinates as theta = atan2(y, x) in degrees, mapped to [0, 360).
    """
    rng = random.Random(seed)
    placed = mm_config.copy()

    mm_numbers = sorted(placed["MM #"].dropna().unique().astype(int).tolist())
    if not mm_numbers:
        return placed

    hew_cache: dict[int, float] = {}
    mm_hew_pairs: list[tuple[int, float]] = []
    for mm_num in mm_numbers:
        if mm_num not in hew_cache:
            hew_cache[mm_num] = compute_mm_hew_at_origin(
                base_params,
                mm_num,
                mm_config_df=mm_config,
                alignment_by_pos=alignment_by_pos,
                gravity_by_pos=gravity_by_pos,
                thermal_by_pos=thermal_by_pos,
            )
        mm_hew_pairs.append((mm_num, hew_cache[mm_num]))
    mm_hew_pairs.sort(key=lambda x: x[1])
    sorted_mms = [mm for mm, _ in mm_hew_pairs]

    positions: list[dict] = []
    for idx, row in placed.iterrows():
        x = float(row.get("x_MM [m]", 0.0))
        y = float(row.get("y_MM [m]", 0.0))
        theta = float(np.degrees(np.arctan2(y, x)))
        if theta < 0:
            theta += 360.0
        positions.append({"index": idx, "theta": theta, "y": y})

    positions.sort(key=lambda p: p["theta"])
    npos = len(positions)
    if npos == 0:
        return placed

    assigned = [False] * npos
    new_assignment: dict[int, int] = {}

    def ang_diff(theta_a: float, theta_b: float) -> float:
        d = abs(float(theta_a) - float(theta_b))
        return min(d, 360.0 - d)

    def find_closest_free_position(target_theta: float) -> int | None:
        best_idx = None
        best_diff = float("inf")
        # Random tie-break for identical diffs (rare, but keeps behavior stable across equal-angle slots).
        tie_break = rng.random()
        for i, pos in enumerate(positions):
            if assigned[i]:
                continue
            diff = ang_diff(float(pos["theta"]), float(target_theta))
            if diff < best_diff:
                best_diff = diff
                best_idx = i
                tie_break = rng.random()
            elif diff == best_diff and best_idx is not None:
                # pseudo-random tie-break
                if rng.random() < tie_break:
                    best_idx = i
                    tie_break = rng.random()
        return best_idx

    # Rank 1 -> +x axis
    idx_pos = find_closest_free_position(0.0)
    if idx_pos is None:
        return placed
    assigned[idx_pos] = True
    new_assignment[positions[idx_pos]["index"]] = int(sorted_mms[0])

    # Rank 2 -> -x axis (if available)
    idx_neg = None
    if len(sorted_mms) >= 2:
        idx_neg = find_closest_free_position(180.0)
        if idx_neg is not None:
            assigned[idx_neg] = True
            new_assignment[positions[idx_neg]["index"]] = int(sorted_mms[1])

    # Build remaining candidate lists for each axis side, split by above/below x-axis.
    # Above x-axis => y>0 (theta in (0, 180)); Below => y<0 (theta in (180, 360)).
    # For +x side we prefer smaller angular difference to 0°.
    # For -x side we prefer smaller angular difference to 180°.
    remaining = [i for i in range(npos) if not assigned[i]]
    rng.shuffle(remaining)  # stable-ish tie-breaking before sorting

    pos_above = [
        i for i in remaining
        if float(positions[i].get("y", 0.0)) > 0.0
    ]
    pos_below = [
        i for i in remaining
        if float(positions[i].get("y", 0.0)) < 0.0
    ]
    # y==0 can go anywhere; treat it as above first.
    pos_axis = [
        i for i in remaining
        if float(positions[i].get("y", 0.0)) == 0.0
    ]
    pos_above = pos_axis + pos_above

    pos_above.sort(key=lambda i: ang_diff(float(positions[i]["theta"]), 0.0))
    pos_below.sort(key=lambda i: ang_diff(float(positions[i]["theta"]), 0.0))

    neg_above = [
        i for i in remaining
        if float(positions[i].get("y", 0.0)) > 0.0
    ]
    neg_below = [
        i for i in remaining
        if float(positions[i].get("y", 0.0)) < 0.0
    ]
    neg_axis = [
        i for i in remaining
        if float(positions[i].get("y", 0.0)) == 0.0
    ]
    neg_above = neg_axis + neg_above

    neg_above.sort(key=lambda i: ang_diff(float(positions[i]["theta"]), 180.0))
    neg_below.sort(key=lambda i: ang_diff(float(positions[i]["theta"]), 180.0))

    def pop_next(candidates: list[int], fallback: list[int]) -> int | None:
        while candidates:
            i = candidates.pop(0)
            if not assigned[i]:
                return i
        while fallback:
            i = fallback.pop(0)
            if not assigned[i]:
                return i
        return None

    toggle_pos_above = True
    toggle_neg_above = True

    # Remaining ranks: alternate between +x side (odd rank index => 3rd, 5th, ...) and -x side.
    for rank in range(2, len(sorted_mms)):
        mm_num = int(sorted_mms[rank])

        if rank % 2 == 0 or idx_neg is None:
            # +x side: 3rd, 5th, 7th, ... (or all remaining if no -x anchor)
            if toggle_pos_above:
                nxt = pop_next(pos_above, pos_below)
            else:
                nxt = pop_next(pos_below, pos_above)
            toggle_pos_above = not toggle_pos_above
        else:
            # -x side: 4th, 6th, 8th, ...
            if toggle_neg_above:
                nxt = pop_next(neg_above, neg_below)
            else:
                nxt = pop_next(neg_below, neg_above)
            toggle_neg_above = not toggle_neg_above

        if nxt is None:
            break
        assigned[nxt] = True
        new_assignment[positions[nxt]["index"]] = mm_num

    for pos_idx, mm_num in new_assignment.items():
        placed.at[pos_idx, "MM #"] = int(mm_num)

    return placed


def _elliptical_place_mm_config(
    mm_config: pd.DataFrame,
    base_params: pd.DataFrame,
    alignment_by_pos: dict[int, dict] | None = None,
    gravity_by_pos: dict[int, dict] | None = None,
    thermal_by_pos: dict[int, dict] | None = None,
    seed: int = 42,
) -> pd.DataFrame:
    """Row-wise placement: best MMs toward x-axis, worst toward y-axis.

    This strategy *does not* move MMs across rows. Within each Row #, it permutes the
    MM assignments so that:
    - lowest-HEW (best) MMs go to slots closest to the x axis (theta≈0° or 180°)
    - highest-HEW (worst) MMs go to slots closest to the y axis (theta≈90° or 270°)

    Implementation detail:
    - For each slot, compute x_score = min(|Δθ to 0|, |Δθ to 180|)
      and y_score = min(|Δθ to 90|, |Δθ to 270|).
    - Sort slots by (x_score - y_score): x-like first (negative), y-like last (positive).
    - Assign row's MMs sorted by individual HEW (best->worst) to that slot order.
    """
    rng = random.Random(seed)
    placed = mm_config.copy()

    def ang_diff(theta_a: float, theta_b: float) -> float:
        d = abs(float(theta_a) - float(theta_b))
        return min(d, 360.0 - d)

    def slot_theta(row: pd.Series) -> float:
        x = float(row.get("x_MM [m]", 0.0))
        y = float(row.get("y_MM [m]", 0.0))
        t = float(np.degrees(np.arctan2(y, x)))
        if t < 0:
            t += 360.0
        return t

    # Cache per-MM intrinsic HEW
    hew_cache: dict[int, float] = {}

    if "Row #" in placed.columns:
        row_values = [v for v in placed["Row #"].dropna().unique().tolist()]
    else:
        row_values = [None]

    for rv in row_values:
        if rv is None:
            idxs = placed.index.tolist()
        else:
            idxs = placed.index[placed["Row #"] == rv].tolist()
        if len(idxs) < 2:
            continue

        row_mm = placed.loc[idxs, "MM #"].dropna().astype(int).tolist()
        if len(row_mm) < 2:
            continue

        mm_hew_pairs: list[tuple[int, float]] = []
        for mm_num in row_mm:
            if mm_num not in hew_cache:
                hew_cache[mm_num] = compute_mm_hew_at_origin(
                    base_params,
                    mm_num,
                    mm_config_df=mm_config,
                    alignment_by_pos=alignment_by_pos,
                    gravity_by_pos=gravity_by_pos,
                    thermal_by_pos=thermal_by_pos,
                )
            mm_hew_pairs.append((mm_num, hew_cache[mm_num]))
        mm_hew_pairs.sort(key=lambda x: x[1])
        sorted_mms = [mm for mm, _ in mm_hew_pairs]

        slots: list[dict] = []
        for idx in idxs:
            t = slot_theta(placed.loc[idx])
            x_score = min(ang_diff(t, 0.0), ang_diff(t, 180.0))
            y_score = min(ang_diff(t, 90.0), ang_diff(t, 270.0))
            slots.append(
                {
                    "index": idx,
                    "theta": t,
                    "x_score": float(x_score),
                    "y_score": float(y_score),
                }
            )

        rng.shuffle(slots)
        slots.sort(key=lambda s: (float(s["x_score"]) - float(s["y_score"]), float(s["x_score"])))

        # If some entries are NaN in MM #, keep them at the end (do not assign new MMs into empty slots).
        filled_slots = [s for s in slots if not pd.isna(placed.at[s["index"], "MM #"])]
        if len(filled_slots) != len(sorted_mms):
            # Defensive: align lengths to avoid accidental row growth.
            n = min(len(filled_slots), len(sorted_mms))
            filled_slots = filled_slots[:n]
            sorted_mms = sorted_mms[:n]

        for s, mm_num in zip(filled_slots, sorted_mms):
            placed.at[s["index"], "MM #"] = int(mm_num)

    return placed


def load_all_sheets(path: str) -> dict:
    # Support in-memory sentinel and CSV-only inputs.
    try:
        if isinstance(path, str) and path == '__INMEM__':
            # pd.read_excel has been overridden by main.py to serve in-memory dicts
            try:
                return pd.read_excel('__INMEM__', sheet_name=None)
            except Exception:
                return {}
        if isinstance(path, str) and str(path).lower().endswith('.csv'):
            try:
                df = pd.read_csv(path)
                return {'MM_PSF': df}
            except Exception:
                return {}
    except Exception:
        pass

    xls = pd.ExcelFile(path, engine="openpyxl")
    sheets = {}
    for name in xls.sheet_names:
        sheets[name] = pd.read_excel(xls, sheet_name=name, engine="openpyxl")
    return sheets


def _write_optimised_workbook_preserving_formatting(
    input_path: str,
    output_path: str,
    updated_mm_config: pd.DataFrame,
) -> None:
    """Write an output workbook while preserving original formatting/formulas/images.

    Implementation:
    - Copy input workbook bytes to a temp file
    - Patch only the 'MM #' cells in the 'MM configuration' sheet
      (matched by 'Position #' when available; else by row order)
    - Atomic replace into output_path
    """
    from openpyxl import load_workbook
    import tempfile

    # If input_path is the in-memory sentinel, materialize a temporary
    # workbook from the in-memory sheets (formatting will be lost).
    created_tmp_input = None
    if isinstance(input_path, str) and input_path == '__INMEM__':
        try:
            sheets = load_all_sheets(input_path)
            tf = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            tmp_input_path = tf.name
            tf.close()
            with pd.ExcelWriter(tmp_input_path, engine='openpyxl') as writer:
                for name, df in sheets.items():
                    if df is None:
                        continue
                    try:
                        df.to_excel(writer, sheet_name=name, index=False)
                    except Exception:
                        try:
                            pd.DataFrame(df).to_excel(writer, sheet_name=name, index=False)
                        except Exception:
                            pass
            created_tmp_input = tmp_input_path
            input_path = tmp_input_path
        except Exception:
            created_tmp_input = None

    base_out, ext_out = os.path.splitext(output_path)
    tmp_output_path = f"{base_out}.tmp.{os.getpid()}{ext_out}"

    # Start from a byte-for-byte copy so all workbook artifacts are retained.
    shutil.copy2(input_path, tmp_output_path)

    keep_vba = ext_out.lower() == ".xlsm"
    wb = load_workbook(tmp_output_path, keep_vba=keep_vba, data_only=False, keep_links=True)

    # Locate MM configuration sheet (case-insensitive fallback).
    sheet_name = None
    if "MM configuration" in wb.sheetnames:
        sheet_name = "MM configuration"
    else:
        for n in wb.sheetnames:
            if str(n).strip().lower() == "mm configuration":
                sheet_name = n
                break
    if sheet_name is None:
        raise ValueError("MM configuration sheet missing in workbook")
    ws = wb[sheet_name]

    # Find header row + column indices.
    header_row = None
    mm_col = None
    pos_col = None
    max_header_scan_rows = min(20, ws.max_row)
    for r in range(1, max_header_scan_rows + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if "MM #" in row_vals:
            header_row = r
            mm_col = row_vals.index("MM #") + 1
            if "Position #" in row_vals:
                pos_col = row_vals.index("Position #") + 1
            break
    if header_row is None or mm_col is None:
        raise ValueError("Could not locate 'MM #' header in MM configuration sheet")

    data_start_row = header_row + 1

    # Prefer updating by Position # when present in both sheet and dataframe.
    use_pos = (pos_col is not None) and ("Position #" in updated_mm_config.columns)
    if use_pos:
        ws_pos_to_row: dict[int, int] = {}
        for r in range(data_start_row, ws.max_row + 1):
            v = ws.cell(row=r, column=pos_col).value
            if v is None:
                continue
            try:
                pos = int(float(v))
            except Exception:
                continue
            ws_pos_to_row[pos] = r

        missing: list[int] = []
        for _, row in updated_mm_config.iterrows():
            pos_v = row.get("Position #")
            if pd.isna(pos_v):
                continue
            pos = int(float(pos_v))
            target_row = ws_pos_to_row.get(pos)
            if target_row is None:
                missing.append(pos)
                continue
            new_mm = int(row.get("MM #"))
            old_mm = ws.cell(row=target_row, column=mm_col).value
            try:
                old_mm_i = int(float(old_mm)) if old_mm is not None else None
            except Exception:
                old_mm_i = None
            if old_mm_i != new_mm:
                ws.cell(row=target_row, column=mm_col, value=new_mm)

        if missing:
            raise ValueError(f"MM configuration Position # values not found in sheet: {sorted(set(missing))[:20]}")
    else:
        # Fallback: update by row order (first data row corresponds to first df row).
        for i, (_, row) in enumerate(updated_mm_config.iterrows()):
            target_row = data_start_row + i
            if target_row > ws.max_row:
                break
            new_mm = int(row.get("MM #"))
            old_mm = ws.cell(row=target_row, column=mm_col).value
            try:
                old_mm_i = int(float(old_mm)) if old_mm is not None else None
            except Exception:
                old_mm_i = None
            if old_mm_i != new_mm:
                ws.cell(row=target_row, column=mm_col, value=new_mm)

    wb.save(tmp_output_path)
    os.replace(tmp_output_path, output_path)

    # Clean up any temporary input we created for __INMEM__ materialization
    try:
        if created_tmp_input is not None:
            os.remove(created_tmp_input)
    except Exception:
        pass


def compute_theta_maps(mm_config_df: pd.DataFrame) -> tuple[dict, dict]:
    """Compute theta_position and theta_degrees maps from MM configuration."""
    theta_map = {}
    theta_pos_map = {}
    x = mm_config_df.get("x_MM [m]", pd.Series(dtype=float))
    y = mm_config_df.get("y_MM [m]", pd.Series(dtype=float))
    mm_nums = mm_config_df["MM #"].astype(int)
    if "y_MM [m]" in mm_config_df.columns:
        theta_ccw = np.degrees(np.arctan2(x, y))
        theta_cw_from_y = np.where(theta_ccw >= 0, theta_ccw, 180 + theta_ccw)
        theta_vals = -(theta_cw_from_y - 90)
    else:
        r = mm_config_df.get("r_MM [m]", pd.Series(dtype=float))
        y_comp = np.sqrt(np.clip(r**2 - x**2, a_min=0, a_max=None))
        theta_ccw = np.degrees(np.arctan2(x, y_comp))
        theta_cw_from_y = np.where(theta_ccw >= 0, theta_ccw, 180 + theta_ccw)
        theta_vals = -(theta_cw_from_y - 90)
    for mm, t in zip(mm_nums, theta_vals):
        theta_map[int(mm)] = t
        theta_pos_map[int(mm)] = t
    return theta_map, theta_pos_map


def convert_polar_to_cartesian(row, mm_config_map):
    # Ensure we lookup mm_config_map using an int key: some DataFrames have
    # MM # as float (e.g. 1.0) which won't match int keys in mm_config_map.
    try:
        mm_num = int(row["MM #"])
    except Exception:
        mm_num = row["MM #"]
    cfg = mm_config_map.get(mm_num)
    if not cfg:
        return row["mux"], row["muy"]
    x_mm = cfg.get("x_MM", 0.0)
    y_mm = cfg.get("y_MM", 0.0)
    r_mm = cfg.get("r_MM", 1e-9)
    u_rad_x = x_mm / r_mm
    u_rad_y = y_mm / r_mm
    u_azi_x = -y_mm / r_mm
    u_azi_y = x_mm / r_mm
    m_rad = row["m_rad"]
    m_azi = row["m_azi"]
    mux = u_rad_x * m_rad + u_azi_x * m_azi
    muy = u_rad_y * m_rad + u_azi_y * m_azi
    return mux, muy


def rebuild_df(params_df: pd.DataFrame, mm_config_df: pd.DataFrame) -> pd.DataFrame:
    theta_map, theta_pos_map = compute_theta_maps(mm_config_df)
    mm_config_map = {
        int(row["MM #"]): {
            "x_MM": row.get("x_MM [m]", 0.0),
            "y_MM": row.get("y_MM [m]", 0.0),
            "z_MM": row.get("z_MM [m]", 0.0),
            "r_MM": row.get("r_MM [m]", 0.0),
        }
        for _, row in mm_config_df.iterrows()
    }
    df = params_df.copy()
    # Compute theta maps but preserve any runtime-provided theta values
    try:
        comp_theta = df["MM #"].map(theta_map)
        comp_theta_pos = df["MM #"].map(theta_pos_map)
        if "theta_degrees" in df.columns:
            df["theta_degrees"] = df["theta_degrees"].where(~df["theta_degrees"].isna(), comp_theta)
        else:
            df["theta_degrees"] = comp_theta.fillna(0.0)
        if "theta_position" in df.columns:
            df["theta_position"] = df["theta_position"].where(~df["theta_position"].isna(), comp_theta_pos)
        else:
            df["theta_position"] = comp_theta_pos.fillna(0.0)
    except Exception:
        df["theta_degrees"] = df.get("theta_degrees", 0.0)
        df["theta_position"] = df.get("theta_position", 0.0)

    # Compute canonical mux/muy from polar params but preserve any runtime
    # values present in params_df. We only fill missing/NaN entries.
    try:
        comp = df.apply(lambda r: pd.Series(convert_polar_to_cartesian(r, mm_config_map)), axis=1)
        comp.columns = ["_comp_mux", "_comp_muy"]
        comp.index = df.index
        if "mux" in df.columns:
            df["mux"] = df["mux"].where(~df["mux"].isna(), comp["_comp_mux"])
        else:
            df["mux"] = comp["_comp_mux"]
        if "muy" in df.columns:
            df["muy"] = df["muy"].where(~df["muy"].isna(), comp["_comp_muy"])
        else:
            df["muy"] = comp["_comp_muy"]
    except Exception:
        # Fallback: ensure mux/muy exist even if computation fails
        if "mux" not in df.columns:
            df["mux"] = 0.0
        if "muy" not in df.columns:
            df["muy"] = 0.0

    # Preserve existing sigmax/sigmay from params_df when present, otherwise
    # derive from sigma_rad/sigma_azi. Only fill NaNs.
    if "sigmax" in df.columns:
        try:
            df["sigmax"] = df["sigmax"].where(~df["sigmax"].isna(), df.get("sigma_rad", pd.Series(dtype=float)))
        except Exception:
            df["sigmax"] = df.get("sigma_rad", pd.Series(dtype=float))
    else:
        df["sigmax"] = df.get("sigma_rad", pd.Series(dtype=float))

    if "sigmay" in df.columns:
        try:
            df["sigmay"] = df["sigmay"].where(~df["sigmay"].isna(), df.get("sigma_azi", pd.Series(dtype=float)))
        except Exception:
            df["sigmay"] = df.get("sigma_azi", pd.Series(dtype=float))
    else:
        df["sigmay"] = df.get("sigma_azi", pd.Series(dtype=float))
    # Debug dump: save input params and rebuilt df for inspection
    try:
        root = Path(__file__).resolve().parent
        outdir = root.parent / 'Figures'
        outdir.mkdir(parents=True, exist_ok=True)
        ts = int(time.time())
        params_df.to_csv(outdir / f'rebuild_params_{ts}.csv', index=False)
        df.to_csv(outdir / f'rebuild_result_{ts}.csv', index=False)
        # Per-row compact diff between params_df and rebuilt df
        try:
            eps = 1e-12
            rows = []
            # merge on 'MM #' when available, else align by index
            if 'MM #' in params_df.columns and 'MM #' in df.columns:
                left = params_df.set_index('MM #')
                right = df.set_index('MM #')
                keys = sorted(set(left.index.tolist()) | set(right.index.tolist()))
                for k in keys:
                    b = left.loc[k] if k in left.index else None
                    a = right.loc[k] if k in right.index else None
                    def _val(obj, col):
                        try:
                            return float(obj[col]) if (obj is not None and col in obj.index and not pd.isna(obj[col])) else float('nan')
                        except Exception:
                            return float('nan')
                    b_sigx = _val(b, 'sigmax') if b is not None else float('nan')
                    a_sigx = _val(a, 'sigmax') if a is not None else float('nan')
                    b_sigy = _val(b, 'sigmay') if b is not None else float('nan')
                    a_sigy = _val(a, 'sigmay') if a is not None else float('nan')
                    b_mux = _val(b, 'mux') if b is not None else float('nan')
                    a_mux = _val(a, 'mux') if a is not None else float('nan')
                    b_muy = _val(b, 'muy') if b is not None else float('nan')
                    a_muy = _val(a, 'muy') if a is not None else float('nan')
                    changed = False
                    for x, y in ((b_sigx, a_sigx), (b_sigy, a_sigy), (b_mux, a_mux), (b_muy, a_muy)):
                        try:
                            if (pd.isna(x) and not pd.isna(y)) or (pd.isna(y) and not pd.isna(x)):
                                changed = True
                                break
                            if pd.isna(x) and pd.isna(y):
                                continue
                            if abs(float(x) - float(y)) > eps:
                                changed = True
                                break
                        except Exception:
                            changed = True
                            break
                    rows.append({
                        'MM': k,
                        'before_sigmax': b_sigx,
                        'after_sigmax': a_sigx,
                        'before_sigmay': b_sigy,
                        'after_sigmay': a_sigy,
                        'before_mux': b_mux,
                        'after_mux': a_mux,
                        'before_muy': b_muy,
                        'after_muy': a_muy,
                        'changed': changed,
                    })
            else:
                for idx in params_df.index:
                    b = params_df.loc[idx]
                    a = df.loc[idx] if idx in df.index else None
                    def _val2(obj, col):
                        try:
                            return float(obj[col]) if (obj is not None and col in obj and not pd.isna(obj[col])) else float('nan')
                        except Exception:
                            return float('nan')
                    b_sigx = _val2(b, 'sigmax')
                    a_sigx = _val2(a, 'sigmax') if a is not None else float('nan')
                    b_sigy = _val2(b, 'sigmay')
                    a_sigy = _val2(a, 'sigmay') if a is not None else float('nan')
                    b_mux = _val2(b, 'mux') if 'mux' in b.index else float('nan')
                    a_mux = _val2(a, 'mux') if (a is not None and 'mux' in a) else float('nan')
                    b_muy = _val2(b, 'muy') if 'muy' in b.index else float('nan')
                    a_muy = _val2(a, 'muy') if (a is not None and 'muy' in a) else float('nan')
                    changed = False
                    for x, y in ((b_sigx, a_sigx), (b_sigy, a_sigy), (b_mux, a_mux), (b_muy, a_muy)):
                        try:
                            if (pd.isna(x) and not pd.isna(y)) or (pd.isna(y) and not pd.isna(x)):
                                changed = True
                                break
                            if pd.isna(x) and pd.isna(y):
                                continue
                            if abs(float(x) - float(y)) > eps:
                                changed = True
                                break
                        except Exception:
                            changed = True
                            break
                    rows.append({
                        'index': idx,
                        'before_sigmax': b_sigx,
                        'after_sigmax': a_sigx,
                        'before_sigmay': b_sigy,
                        'after_sigmay': a_sigy,
                        'before_mux': b_mux,
                        'after_mux': a_mux,
                        'before_muy': b_muy,
                        'after_muy': a_muy,
                        'changed': changed,
                    })
            try:
                pd.DataFrame(rows).to_csv(outdir / f'rebuild_perrow_{ts}.csv', index=False)
            except Exception:
                pass

            # Post-write assertions: detect unexpected zeroing of mux/muy
            try:
                for r in rows:
                    b_mux = r.get('before_mux')
                    a_mux = r.get('after_mux')
                    b_muy = r.get('before_muy')
                    a_muy = r.get('after_muy')
                    if not pd.isna(b_mux) and not pd.isna(a_mux):
                        if abs(float(b_mux)) > 1e-15 and float(a_mux) == 0.0:
                            raise AssertionError(f'Unexpected mux zeroing during rebuild_df for MM/index {r.get("index") or r.get("MM")}: before={b_mux} after={a_mux}')
                    if not pd.isna(b_muy) and not pd.isna(a_muy):
                        if abs(float(b_muy)) > 1e-15 and float(a_muy) == 0.0:
                            raise AssertionError(f'Unexpected muy zeroing during rebuild_df for MM/index {r.get("index") or r.get("MM")}: before={b_muy} after={a_muy}')
            except AssertionError:
                raise
        except Exception:
            pass
    except Exception:
        pass
    # As a final step, if the original params_df carried explicit theta values
    # (e.g. copied from runtime), prefer those authoritative values by mapping
    # them into the rebuilt frame using `MM #` as the join key.
    try:
        if 'MM #' in params_df.columns and 'MM #' in df.columns:
            if 'theta_degrees' in params_df.columns:
                try:
                    src = params_df.set_index('MM #')['theta_degrees']
                    df['theta_degrees'] = df['MM #'].astype(int).map(src).fillna(df.get('theta_degrees'))
                except Exception:
                    pass
            if 'theta_position' in params_df.columns:
                try:
                    src2 = params_df.set_index('MM #')['theta_position']
                    df['theta_position'] = df['MM #'].astype(int).map(src2).fillna(df.get('theta_position'))
                except Exception:
                    pass
    except Exception:
        pass
    return df



def _load_position_deltas(path: str) -> tuple[dict[int, dict], dict[int, dict], dict[int, dict]]:
    """Load per-position deltas from Alignment/Gravity offload/Thermal.

    Returns (alignment_by_pos, gravity_by_pos, thermal_by_pos)
    where shifts are in meters and rotz in arcsec.
    For sheets with duplicate rows per position, values are summed.
    """
    alignment_by_pos: dict[int, dict] = {}
    gravity_by_pos: dict[int, dict] = {}
    thermal_by_pos: dict[int, dict] = {}

    try:
        align_df = pd.read_excel(path, sheet_name="Alignment", engine="openpyxl")
        if "Position #" in align_df.columns:
            tmp = align_df.copy()
            tmp["Position #"] = pd.to_numeric(tmp["Position #"], errors="coerce")
            tmp = tmp[tmp["Position #"].notna()]
            for _, row in tmp.iterrows():
                pos = int(row["Position #"])
                alignment_by_pos[pos] = {
                    "d_align_rad": float(row.get("d_align_rad [µm]", 0.0)) * 1e-6,
                    "d_align_azi": float(row.get("d_align_azi [µm]", 0.0)) * 1e-6,
                    "d_align_z": float(row.get("d_align_z [µm]", 0.0)) * 1e-6,
                    "d_align_rotz": float(row.get("d_align_rotz [arcsec]", 0.0)),
                }
    except Exception:
        pass

    def _sum_by_position(df: pd.DataFrame, pos_col: str, cols: list[str]) -> pd.DataFrame:
        tmp = df.copy()
        tmp[pos_col] = pd.to_numeric(tmp[pos_col], errors="coerce")
        tmp = tmp[tmp[pos_col].notna()]
        for c in cols:
            if c in tmp.columns:
                tmp[c] = pd.to_numeric(tmp[c], errors="coerce").fillna(0.0)
        return tmp.groupby(pos_col, as_index=False).sum(numeric_only=True)

    try:
        grav_df = pd.read_excel(path, sheet_name="Gravity offload", engine="openpyxl")
        if "Position #" in grav_df.columns:
            grp = _sum_by_position(
                grav_df,
                "Position #",
                ["d_grav_x [µm]", "d_grav_y [µm]", "d_grav_z [µm]", "d_grav_rotz [arcsec]"],
            )
            for _, row in grp.iterrows():
                pos = int(row["Position #"])
                gravity_by_pos[pos] = {
                    "d_grav_x": float(row.get("d_grav_x [µm]", 0.0)) * 1e-6,
                    "d_grav_y": float(row.get("d_grav_y [µm]", 0.0)) * 1e-6,
                    "d_grav_z": float(row.get("d_grav_z [µm]", 0.0)) * 1e-6,
                    "d_grav_rotz": float(row.get("d_grav_rotz [arcsec]", 0.0)),
                }
    except Exception:
        pass

    try:
        therm_df = pd.read_excel(path, sheet_name="Thermal", engine="openpyxl")
        if "Position #" in therm_df.columns:
            grp = _sum_by_position(
                therm_df,
                "Position #",
                ["d_therm_x [µm]", "d_therm_y [µm]", "d_therm_z [µm]", "d_therm_rotz [arcsec]"],
            )
            for _, row in grp.iterrows():
                pos = int(row["Position #"])
                thermal_by_pos[pos] = {
                    "d_therm_x": float(row.get("d_therm_x [µm]", 0.0)) * 1e-6,
                    "d_therm_y": float(row.get("d_therm_y [µm]", 0.0)) * 1e-6,
                    "d_therm_z": float(row.get("d_therm_z [µm]", 0.0)) * 1e-6,
                    "d_therm_rotz": float(row.get("d_therm_rotz [arcsec]", 0.0)),
                }
    except Exception:
        pass

    return alignment_by_pos, gravity_by_pos, thermal_by_pos


def _apply_position_deltas_to_df(
    df: pd.DataFrame,
    mm_config_df: pd.DataFrame,
    alignment_by_pos: dict[int, dict],
    gravity_by_pos: dict[int, dict],
    thermal_by_pos: dict[int, dict],
) -> pd.DataFrame:
    """Apply per-position deltas to a rebuilt PSF parameter DataFrame.

    Uses explicit Position # column from mm_config_df when present,
    else falls back to Position # = 1..N by row order.
    """
    out = df.copy()

    # Lightweight debug snapshot: save input before applying deltas
    try:
        from pathlib import Path
        import time
        root = Path(__file__).resolve().parent
        outdir = root.parent / 'Figures'
        outdir.mkdir(parents=True, exist_ok=True)
        ts = int(time.time())
        out.to_csv(outdir / f'apply_posdeltas_before_{ts}.csv', index=False)
    except Exception:
        pass

    # Map MM# -> row index in mm_config_df (each row is a fixed slot/position)
    mm_to_idx = {int(row["MM #"]): idx for idx, row in mm_config_df.iterrows()}

    # Map row index -> position number
    idx_list = list(mm_config_df.index)
    if "Position #" in mm_config_df.columns:
        pos_series = pd.to_numeric(mm_config_df["Position #"], errors="coerce")
        idx_to_pos = {}
        for i, idx in enumerate(idx_list):
            v = pos_series.loc[idx]
            if pd.isna(v):
                idx_to_pos[idx] = i + 1
            else:
                idx_to_pos[idx] = int(float(v))
    else:
        idx_to_pos = {idx: i + 1 for i, idx in enumerate(idx_list)}

    # Geometry by row index
    geom_by_idx = {
        idx: {
            "x": float(row.get("x_MM [m]", 0.0)),
            "y": float(row.get("y_MM [m]", 0.0)),
            "z": float(row.get("z_MM [m]", 0.0)),
            "r": float(row.get("r_MM [m]", 0.0)),
        }
        for idx, row in mm_config_df.iterrows()
    }

    # Apply to polar offsets first (m_rad/m_azi + rotz coupling)
    for i, row in out.iterrows():
        mm_num = int(row["MM #"])
        idx = mm_to_idx.get(mm_num)
        if idx is None:
            continue
        pos = idx_to_pos[idx]

        m_rad = float(row["m_rad"])
        m_azi = float(row["m_azi"])

        if pos in alignment_by_pos:
            m_rad += float(alignment_by_pos[pos].get("d_align_rad", 0.0))
            m_azi += float(alignment_by_pos[pos].get("d_align_azi", 0.0))

        d_rotz_arcsec = 0.0
        if pos in alignment_by_pos:
            d_rotz_arcsec += float(alignment_by_pos[pos].get("d_align_rotz", 0.0))
        if pos in gravity_by_pos:
            d_rotz_arcsec += float(gravity_by_pos[pos].get("d_grav_rotz", 0.0))
        if pos in thermal_by_pos:
            d_rotz_arcsec += float(thermal_by_pos[pos].get("d_therm_rotz", 0.0))

        if d_rotz_arcsec != 0.0:
            r_mm = float(geom_by_idx[idx]["r"])
            d_rotz_rad = np.radians(d_rotz_arcsec / 3600.0)
            m_azi += r_mm * d_rotz_rad

        out.at[i, "m_rad"] = m_rad
        out.at[i, "m_azi"] = m_azi

    # Recompute mux/muy based on modified m_rad/m_azi
    # (reuse rebuild_df conversion logic by rebuilding mm_config_map)
    mm_config_map = {
        int(r["MM #"]): {
            "x_MM": float(r.get("x_MM [m]", 0.0)),
            "y_MM": float(r.get("y_MM [m]", 0.0)),
            "z_MM": float(r.get("z_MM [m]", 0.0)),
            "r_MM": float(r.get("r_MM [m]", 0.0)),
        }
        for _, r in mm_config_df.iterrows()
    }
    out[["mux", "muy"]] = out.apply(lambda r: pd.Series(convert_polar_to_cartesian(r, mm_config_map)), axis=1)
    out["sigmax"] = out["sigma_rad"]
    out["sigmay"] = out["sigma_azi"]

    # Apply xy offsets from gravity/thermal and z-projection from combined dz.
    for i, row in out.iterrows():
        mm_num = int(row["MM #"])
        idx = mm_to_idx.get(mm_num)
        if idx is None:
            continue
        pos = idx_to_pos[idx]

        mux = float(row["mux"])
        muy = float(row["muy"])

        d_align_z = float(alignment_by_pos.get(pos, {}).get("d_align_z", 0.0))
        d_grav_z = float(gravity_by_pos.get(pos, {}).get("d_grav_z", 0.0))
        d_therm_z = float(thermal_by_pos.get(pos, {}).get("d_therm_z", 0.0))
        dz = d_align_z + d_grav_z + d_therm_z

        if pos in gravity_by_pos:
            mux += float(gravity_by_pos[pos].get("d_grav_x", 0.0))
            muy += float(gravity_by_pos[pos].get("d_grav_y", 0.0))
        if pos in thermal_by_pos:
            mux += float(thermal_by_pos[pos].get("d_therm_x", 0.0))
            muy += float(thermal_by_pos[pos].get("d_therm_y", 0.0))

        if dz != 0.0:
            g = geom_by_idx[idx]
            denom = 12.0 - float(g["z"])
            if denom != 0.0:
                mux += dz * float(g["x"]) / denom
                muy += dz * float(g["y"]) / denom

        out.at[i, "mux"] = mux
        out.at[i, "muy"] = muy

    # Save output snapshot for debugging
    try:
        out.to_csv(outdir / f'apply_posdeltas_after_{ts}.csv', index=False)
        # Compact per-row diff for apply_posdeltas (before -> after)
        try:
            eps = 1e-12
            before = pd.read_csv(outdir / f'apply_posdeltas_before_{ts}.csv') if (outdir / f'apply_posdeltas_before_{ts}.csv').exists() else None
            after = out.copy()
            rows = []
            if before is not None:
                # align by MM # when present
                if 'MM #' in before.columns and 'MM #' in after.columns:
                    left = before.set_index('MM #')
                    right = after.set_index('MM #')
                    keys = sorted(set(left.index.tolist()) | set(right.index.tolist()))
                    for k in keys:
                        b = left.loc[k] if k in left.index else None
                        a = right.loc[k] if k in right.index else None
                        def _v(obj, col):
                            try:
                                return float(obj[col]) if (obj is not None and col in obj.index and not pd.isna(obj[col])) else float('nan')
                            except Exception:
                                return float('nan')
                        b_sigx = _v(b, 'sigmax')
                        a_sigx = _v(a, 'sigmax')
                        b_sigy = _v(b, 'sigmay')
                        a_sigy = _v(a, 'sigmay')
                        b_mux = _v(b, 'mux')
                        a_mux = _v(a, 'mux')
                        b_muy = _v(b, 'muy')
                        a_muy = _v(a, 'muy')
                        changed = False
                        for x, y in ((b_sigx, a_sigx), (b_sigy, a_sigy), (b_mux, a_mux), (b_muy, a_muy)):
                            try:
                                if (pd.isna(x) and not pd.isna(y)) or (pd.isna(y) and not pd.isna(x)):
                                    changed = True
                                    break
                                if pd.isna(x) and pd.isna(y):
                                    continue
                                if abs(float(x) - float(y)) > eps:
                                    changed = True
                                    break
                            except Exception:
                                changed = True
                                break
                        rows.append({'MM': k, 'before_sigmax': b_sigx, 'after_sigmax': a_sigx, 'before_sigmay': b_sigy, 'after_sigmay': a_sigy, 'before_mux': b_mux, 'after_mux': a_mux, 'before_muy': b_muy, 'after_muy': a_muy, 'changed': changed})
            if rows:
                try:
                    pd.DataFrame(rows).to_csv(outdir / f'apply_posdeltas_perrow_{ts}.csv', index=False)
                except Exception:
                    pass

                # Post-write assertions: detect unexpected zeroing of mux/muy
                try:
                    for r in rows:
                        b_mux = r.get('before_mux')
                        a_mux = r.get('after_mux')
                        b_muy = r.get('before_muy')
                        a_muy = r.get('after_muy')
                        if not pd.isna(b_mux) and not pd.isna(a_mux):
                            if abs(float(b_mux)) > 1e-15 and float(a_mux) == 0.0:
                                raise AssertionError(f'Unexpected mux zeroing during apply_posdeltas for MM/index {r.get("index") or r.get("MM")}: before={b_mux} after={a_mux}')
                        if not pd.isna(b_muy) and not pd.isna(a_muy):
                            if abs(float(b_muy)) > 1e-15 and float(a_muy) == 0.0:
                                raise AssertionError(f'Unexpected muy zeroing during apply_posdeltas for MM/index {r.get("index") or r.get("MM")}: before={b_muy} after={a_muy}')
                except AssertionError:
                    raise
        except Exception:
            pass
    except Exception:
        pass
    return out


def hew_at_best_focus(df: pd.DataFrame, fast: bool = True, timeout: float = 60.0) -> float:
    """
    Calculates the Half Energy Width (HEW) at the best focus point.

    This function computes the HEW of a point spread function described by a DataFrame
    of Gaussian or Pseudo-Voigt distributions. It includes an optimization step to find
    the center of the distribution that minimizes the HEW. The accuracy of the calculation
    can be controlled with the `fast` parameter, which adjusts the grid size for the
    radial profile calculation. A timeout is included to prevent excessively long
    calculations.

    Args:
        df (pd.DataFrame): DataFrame containing the parameters of the distributions
                           (e.g., position, sigma, weight).
        fast (bool): If True, uses a coarser grid for faster but less accurate
                     calculation. Defaults to True.
        timeout (float): The maximum time in seconds allowed for the HEW calculation.
                         Defaults to 60.0.

    Returns:
        float: The calculated HEW value in meters.
    """
    start_time = time.time()
    normalize = True

    def radial_profile(cx, cy, n_r=320, n_theta=180, r_margin_factor=5.0):
        if time.time() - start_time > timeout:
            raise TimeoutError("HEW calculation timed out")
        max_sigma = max(df["sigmax"].max(), df["sigmay"].max())
        max_center = np.sqrt((df["mux"] - cx) ** 2 + (df["muy"] - cy) ** 2).max()
        r_max = max(max_center + r_margin_factor * max_sigma, 1e-6)
        theta = np.linspace(0.0, 2.0 * np.pi, n_theta, endpoint=False)
        r = np.linspace(0.0, r_max, n_r)
        dtheta = theta[1] - theta[0]
        dr = r[1] - r[0] if n_r > 1 else r_max
        R, TH = np.meshgrid(r, theta)
        Xp = cx + R * np.cos(TH)
        Yp = cy + R * np.sin(TH)
        Zp = np.zeros_like(Xp)
        for _, row in df.iterrows():
            dist_type = row.get("distribution", "gaussian")
            if dist_type in ["pseudo-voigt", "voigt"]:
                Zp += pseudo_voigt_2d_rotated(
                    Xp,
                    Yp,
                    muazi=row["mux"],
                    murad=row["muy"],
                    sigmaazi=row["sigmax"],
                    sigmarad=row["sigmay"],
                    theta=row["theta_degrees"],
                    alphaazi=row.get("alpha_azi", 0.5),
                    alpharad=row.get("alpha_rad", 0.5),
                    amplitude=row["weight"],
                    normalize=normalize,
                    degrees=True,
                )
            else:
                Zp += gaussian_2d_rotated(
                    Xp,
                    Yp,
                    mux=row["mux"],
                    muy=row["muy"],
                    sigmax=row["sigmax"],
                    sigmay=row["sigmay"],
                    theta=row["theta_degrees"],
                    amplitude=row["weight"],
                    normalize=normalize,
                    degrees=True,
                )
        radial_energy = np.sum(Zp * R, axis=0) * dtheta
        cumulative = np.cumsum(radial_energy * dr)
        total_energy = cumulative[-1] if cumulative.size else 1.0
        return r, cumulative, total_energy

    def hew_at_center(cx, cy):
        r, cumulative, total_energy = radial_profile(
            cx, cy, n_r=120 if fast else 200, n_theta=60 if fast else 120
        )
        frac = cumulative / total_energy if total_energy > 0 else cumulative
        return float(np.interp(0.5, frac, r))

    total_weight = float(df["weight"].sum())
    if not np.isfinite(total_weight) or total_weight <= 0:
        # Fallback to unweighted centroid when weights are zero/invalid
        cx0 = float(df["mux"].mean()) if "mux" in df.columns else 0.0
        cy0 = float(df["muy"].mean()) if "muy" in df.columns else 0.0
    else:
        cx0 = float((df["mux"] * df["weight"]).sum() / total_weight)
        cy0 = float((df["muy"] * df["weight"]).sum() / total_weight)

    # Time-bounded coordinate-descent search (no SciPy dependency).
    # This is deterministic, stable, and respects the timeout.
    best_cx, best_cy = cx0, cy0
    best_hew = hew_at_center(best_cx, best_cy)

    # A couple of alternative starting points can help when the weighted center
    # is not close to best focus.
    starts = [(cx0, cy0), (0.0, 0.0)]
    step = 2e-6 if fast else 1e-6
    min_step = 2e-9 if fast else 1e-9

    for sx, sy in starts:
        if time.time() - start_time > timeout:
            break
        cx, cy = float(sx), float(sy)
        try:
            cur = hew_at_center(cx, cy)
        except TimeoutError:
            break

        local_step = step
        while local_step >= min_step and (time.time() - start_time) <= timeout:
            improved = False
            for tx, ty in (
                (cx + local_step, cy),
                (cx - local_step, cy),
                (cx, cy + local_step),
                (cx, cy - local_step),
            ):
                try:
                    val = hew_at_center(tx, ty)
                except TimeoutError:
                    val = None
                if val is not None and val < cur:
                    cx, cy, cur = tx, ty, val
                    improved = True
                    break
            if not improved:
                local_step *= 0.5

        if cur < best_hew:
            best_cx, best_cy, best_hew = cx, cy, cur

    r_final, cumulative_final, total_final = radial_profile(
        best_cx, best_cy, n_r=200 if fast else 360, n_theta=100 if fast else 200
    )
    frac_final = cumulative_final / total_final if total_final > 0 else cumulative_final
    return float(np.interp(0.5, frac_final, r_final))


def hew_fast_approximate(df: pd.DataFrame) -> float:
    """Fast approximate HEW using very coarse grid for permutation testing."""
    normalize = True
    total_weight = float(df["weight"].sum())
    if not np.isfinite(total_weight) or total_weight <= 0:
        # Fallback to unweighted centroid when weights are zero/invalid
        cx = float(df["mux"].mean()) if "mux" in df.columns else 0.0
        cy = float(df["muy"].mean()) if "muy" in df.columns else 0.0
    else:
        cx = float((df["mux"] * df["weight"]).sum() / total_weight)
        cy = float((df["muy"] * df["weight"]).sum() / total_weight)
    
    # Enforce a small sigma floor to avoid zero-width Gaussians which raise
    # in the Gaussian implementation and break approximate HEW computations.
    MIN_SIG_M = 1e-9
    if 'sigmax' in df.columns:
        df['sigmax'] = np.maximum(df['sigmax'].to_numpy(dtype=float, copy=False), MIN_SIG_M)
    if 'sigmay' in df.columns:
        df['sigmay'] = np.maximum(df['sigmay'].to_numpy(dtype=float, copy=False), MIN_SIG_M)

    max_sigma = max(df["sigmax"].max(), df["sigmay"].max())
    max_center = np.sqrt((df["mux"] - cx) ** 2 + (df["muy"] - cy) ** 2).max()
    # Reduce radius margin and sampling to make placement/approx HEW fast.
    # Placement uses this approximate HEW many times; prioritize speed over
    # perfect accuracy. Use conservative caps to avoid runaway cost for
    # pseudo-Voigt heavy tails.
    r_max = max(max_center + 4.0 * max_sigma, 1e-6)

    # Moderate-resolution grid suitable for ranking/placement decisions.
    # These values keep runtime low while producing stable ordering of MMs.
    n_r, n_theta = 600, 180
    theta = np.linspace(0.0, 2.0 * np.pi, n_theta, endpoint=False)
    r = np.linspace(0.0, r_max, n_r)
    dtheta = theta[1] - theta[0]
    dr = r[1] - r[0] if n_r > 1 else r_max
    R, TH = np.meshgrid(r, theta)
    Xp = cx + R * np.cos(TH)
    Yp = cy + R * np.sin(TH)
    Zp = np.zeros_like(Xp)
    
    for _, row in df.iterrows():
        dist_type = row.get("distribution", "gaussian")
        if dist_type in ["pseudo-voigt", "voigt"]:
            Zp += pseudo_voigt_2d_rotated(
                Xp, Yp,
                muazi=row["mux"], murad=row["muy"],
                sigmaazi=row["sigmax"], sigmarad=row["sigmay"],
                theta=row["theta_degrees"],
                alphaazi=row.get("alpha_azi", 0.5),
                alpharad=row.get("alpha_rad", 0.5),
                amplitude=row["weight"],
                normalize=normalize,
                degrees=True,
            )
        else:
            Zp += gaussian_2d_rotated(
                Xp, Yp,
                mux=row["mux"], muy=row["muy"],
                sigmax=row["sigmax"], sigmay=row["sigmay"],
                theta=row["theta_degrees"],
                amplitude=row["weight"],
                normalize=normalize,
                degrees=True,
            )
    
    radial_energy = np.sum(Zp * R, axis=0) * dtheta
    cumulative = np.cumsum(radial_energy * dr)
    total_energy = cumulative[-1] if cumulative.size else 1.0
    frac = cumulative / total_energy if total_energy > 0 else cumulative
    # Use local cubic interpolation for the 50% radius, then return HEW diameter
    radius50 = _radius_for_fraction(frac, r, target=0.5)
    return float(2.0 * radius50)


def hew_fast_approximate_center(df: pd.DataFrame, cx: float, cy: float) -> float:
    """Approximate HEW computed with a fixed center (cx, cy) instead of using
    the dataset centroid. Useful for ranking single-MM PSFs with respect to the
    instrument origin.
    """
    normalize = True
    # use the same sigma floor as hew_fast_approximate
    MIN_SIG_M = 1e-9
    if 'sigmax' in df.columns:
        df['sigmax'] = np.maximum(df['sigmax'].to_numpy(dtype=float, copy=False), MIN_SIG_M)
    if 'sigmay' in df.columns:
        df['sigmay'] = np.maximum(df['sigmay'].to_numpy(dtype=float, copy=False), MIN_SIG_M)

    max_sigma = max(df['sigmax'].max(), df['sigmay'].max())
    max_center = np.sqrt((df['mux'] - cx) ** 2 + (df['muy'] - cy) ** 2).max()
    r_max = max(max_center + 4.0 * max_sigma, 1e-6)

    n_r, n_theta = 600, 180
    theta = np.linspace(0.0, 2.0 * np.pi, n_theta, endpoint=False)
    r = np.linspace(0.0, r_max, n_r)
    dtheta = theta[1] - theta[0]
    dr = r[1] - r[0] if n_r > 1 else r_max
    R, TH = np.meshgrid(r, theta)
    Xp = cx + R * np.cos(TH)
    Yp = cy + R * np.sin(TH)
    Zp = np.zeros_like(Xp)

    for _, row in df.iterrows():
        dist_type = row.get('distribution', 'gaussian')
        if dist_type in ['pseudo-voigt', 'voigt']:
            Zp += pseudo_voigt_2d_rotated(
                Xp, Yp,
                muazi=row['mux'], murad=row['muy'],
                sigmaazi=row['sigmax'], sigmarad=row['sigmay'],
                theta=row['theta_degrees'],
                alphaazi=row.get('alpha_azi', 0.5),
                alpharad=row.get('alpha_rad', 0.5),
                amplitude=row['weight'],
                normalize=normalize,
                degrees=True,
            )
        else:
            Zp += gaussian_2d_rotated(
                Xp, Yp,
                mux=row['mux'], muy=row['muy'],
                sigmax=row['sigmax'], sigmay=row['sigmay'],
                theta=row['theta_degrees'],
                amplitude=row['weight'],
                normalize=normalize,
                degrees=True,
            )

    radial_energy = np.sum(Zp * R, axis=0) * dtheta
    cumulative = np.cumsum(radial_energy * dr)
    total_energy = cumulative[-1] if cumulative.size else 1.0
    frac = cumulative / total_energy if total_energy > 0 else cumulative
    radius50 = _radius_for_fraction(frac, r, target=0.5)
    return float(2.0 * radius50)


def compute_mm_hew_at_origin(
    params_df: pd.DataFrame,
    mm_num: int,
    mm_config_df: pd.DataFrame | None = None,
    alignment_by_pos: dict[int, dict] | None = None,
    gravity_by_pos: dict[int, dict] | None = None,
    thermal_by_pos: dict[int, dict] | None = None,
) -> float:
    """Compute the HEW for a single MM measured about the origin (0,0).

    This constructs a single-row DataFrame for the MM with mux/muy set to the
    MM's position including alignment/gravity/thermal deltas (if mm_config_df
    is provided) and returns the approximate HEW computed about (0,0).
    """
    mm_params = params_df[params_df['MM #'] == mm_num]
    if mm_params.empty:
        return float('inf')

    # Build a single-row DF matching hew_fast_approximate expectations
    row = mm_params.iloc[0].copy()
    # Default mux/muy from intrinsic polar offsets
    mux = 0.0
    muy = 0.0
    theta_deg = 0.0

    if mm_config_df is not None:
        try:
            slot_row = mm_config_df[mm_config_df['MM #'].astype(int) == int(mm_num)].iloc[0]
        except Exception:
            slot_row = None
        if slot_row is not None:
            x_mm = float(slot_row.get('x_MM [m]', 0.0))
            y_mm = float(slot_row.get('y_MM [m]', 0.0))
            r_mm = float(slot_row.get('r_MM [m]', 0.0))
            if r_mm == 0.0:
                r_mm = float(np.hypot(x_mm, y_mm)) if (x_mm or y_mm) else 1e-9
            # compute polar offsets including per-position deltas
            m_rad = float(row.get('m_rad', 0.0))
            m_azi = float(row.get('m_azi', 0.0))
            pos_num = None
            if 'Position #' in mm_config_df.columns:
                try:
                    pos_num = int(float(slot_row.get('Position #')))
                except Exception:
                    pos_num = None

            # Apply alignment deltas to polar offsets
            if pos_num is not None and alignment_by_pos and pos_num in alignment_by_pos:
                m_rad += float(alignment_by_pos[pos_num].get('d_align_rad', 0.0))
                m_azi += float(alignment_by_pos[pos_num].get('d_align_azi', 0.0))

            # rotz coupling
            d_rotz_arcsec = 0.0
            if alignment_by_pos and pos_num in alignment_by_pos:
                d_rotz_arcsec += float(alignment_by_pos[pos_num].get('d_align_rotz', 0.0))
            if gravity_by_pos and pos_num in gravity_by_pos:
                d_rotz_arcsec += float(gravity_by_pos[pos_num].get('d_grav_rotz', 0.0))
            if thermal_by_pos and pos_num in thermal_by_pos:
                d_rotz_arcsec += float(thermal_by_pos[pos_num].get('d_therm_rotz', 0.0))
            if d_rotz_arcsec != 0.0:
                d_rotz_rad = np.radians(d_rotz_arcsec / 3600.0)
                m_azi += r_mm * d_rotz_rad

            # Convert polar->cartesian using same convention
            u_rad_x = x_mm / r_mm
            u_rad_y = y_mm / r_mm
            u_azi_x = -y_mm / r_mm
            u_azi_y = x_mm / r_mm
            mux = u_rad_x * m_rad + u_azi_x * m_azi
            muy = u_rad_y * m_rad + u_azi_y * m_azi

            # Apply gravity/thermal xy
            if gravity_by_pos and pos_num in gravity_by_pos:
                mux += float(gravity_by_pos[pos_num].get('d_grav_x', 0.0))
                muy += float(gravity_by_pos[pos_num].get('d_grav_y', 0.0))
            if thermal_by_pos and pos_num in thermal_by_pos:
                mux += float(thermal_by_pos[pos_num].get('d_therm_x', 0.0))
                muy += float(thermal_by_pos[pos_num].get('d_therm_y', 0.0))

            # z projection
            dz = 0.0
            if alignment_by_pos and pos_num in alignment_by_pos:
                dz += float(alignment_by_pos[pos_num].get('d_align_z', 0.0))
            if gravity_by_pos and pos_num in gravity_by_pos:
                dz += float(gravity_by_pos[pos_num].get('d_grav_z', 0.0))
            if thermal_by_pos and pos_num in thermal_by_pos:
                dz += float(thermal_by_pos[pos_num].get('d_therm_z', 0.0))
            if dz != 0.0:
                denom = 12.0 - float(slot_row.get('z_MM', 0.0))
                if denom != 0.0:
                    mux += dz * float(slot_row.get('x_MM [m]', 0.0)) / denom
                    muy += dz * float(slot_row.get('y_MM [m]', 0.0)) / denom

            # Theta for the slot
            theta_deg = float(np.degrees(np.arctan2(float(y_mm), float(x_mm)))) if (x_mm or y_mm) else 0.0
            if theta_deg < 0:
                theta_deg += 360.0

    # Build single-row DataFrame
    single = pd.DataFrame([
        {
            'MM #': int(row.get('MM #')),
            'mux': float(mux),
            'muy': float(muy),
            'sigmax': float(row.get('sigma_rad', row.get('sigmax', 0.0))),
            'sigmay': float(row.get('sigma_azi', row.get('sigmay', 0.0))),
            'theta_degrees': float(row.get('theta_degrees', theta_deg)),
            'weight': float(row.get('weight', 1.0)),
            'distribution': row.get('distribution', 'gaussian'),
            'alpha_azi': row.get('alpha_azi', 0.5),
            'alpha_rad': row.get('alpha_rad', 0.5),
        }
    ])

    return hew_fast_approximate_center(single, 0.0, 0.0)


def _load_base_params_from_workbook(input_path: str) -> pd.DataFrame:
    """Load intrinsic MM PSF params from MM_PSF (no position deltas applied)."""
    arcsec_to_m = 12 * np.pi / 180 / 3600
    psf = pd.read_excel(input_path, sheet_name="MM_PSF", engine="openpyxl")
    base_params = pd.DataFrame({
        "MM #": pd.to_numeric(psf["MM #"], errors="coerce").astype(int),
        "m_rad": pd.to_numeric(psf["m_rad [arcsec]"], errors="coerce").fillna(0.0) * arcsec_to_m,
        "m_azi": pd.to_numeric(psf["m_azi [arcsec]"], errors="coerce").fillna(0.0) * arcsec_to_m,
        "sigma_rad": pd.to_numeric(psf["sigma_rad [arcsec]"], errors="coerce").fillna(0.0) * arcsec_to_m,
        "sigma_azi": pd.to_numeric(psf["sigma_azi [arcsec]"], errors="coerce").fillna(0.0) * arcsec_to_m,
    })
    if "distribution" in psf.columns:
        base_params["distribution"] = psf["distribution"].astype(str).str.lower().fillna("gaussian")
    else:
        base_params["distribution"] = "gaussian"
    base_params["alpha_azi"] = pd.to_numeric(psf.get("alpha_azi", 0.5), errors="coerce").fillna(0.5)
    base_params["alpha_rad"] = pd.to_numeric(psf.get("alpha_rad", 0.5), errors="coerce").fillna(0.5)

    from main import load_aeff_weight_map
    aeff_map = load_aeff_weight_map(input_path)
    base_params["weight"] = base_params["MM #"].map(aeff_map)
    missing = base_params["weight"].isna()
    if missing.any():
        missing_mm = sorted(set(base_params.loc[missing, "MM #"].astype(int).tolist()))
        raise ValueError(
            "Missing A_eff weights for some MMs. "
            "A_eff column B must contain a numeric weight for every MM used. "
            f"Missing examples: {missing_mm[:20]}"
        )
    return base_params


def compute_individual_mm_hew(
    params_df: pd.DataFrame,
    mm_num: int,
    mm_config_df: pd.DataFrame | None = None,
    alignment_by_pos: dict[int, dict] | None = None,
    gravity_by_pos: dict[int, dict] | None = None,
    thermal_by_pos: dict[int, dict] | None = None,
) -> float:
    """Compute a per-MM ranking score that includes PSF shifts.

    Why this exists:
    - A "single-MM HEW" centered on its own centroid is mostly insensitive to
      centroid shifts, but for system placement those shifts matter.
    - We therefore rank by: intrinsic_HEW + |centroid_shift|, where centroid_shift
      includes m_rad/m_azi and all per-position deltas (alignment/gravity/thermal).

    If mm_config_df and delta dicts are not provided, falls back to intrinsic HEW only.
    """
    mm_params = params_df[params_df["MM #"] == mm_num]
    if mm_params.empty:
        return float("inf")

    # Intrinsic spot-size term (independent of centroid shift)
    single_mm = mm_params.copy()
    single_mm["mux"] = 0.0
    single_mm["muy"] = 0.0
    single_mm["sigmax"] = single_mm["sigma_rad"]
    single_mm["sigmay"] = single_mm["sigma_azi"]
    single_mm["theta_degrees"] = 0.0
    intrinsic = float(hew_fast_approximate(single_mm))

    if mm_config_df is None:
        return intrinsic

    try:
        slot_row = mm_config_df[mm_config_df["MM #"].astype(int) == int(mm_num)].iloc[0]
    except Exception:
        return intrinsic

    # Determine position number (needed to look up deltas)
    if "Position #" in mm_config_df.columns:
        pos_val = pd.to_numeric(slot_row.get("Position #", np.nan), errors="coerce")
        if pd.isna(pos_val):
            # fall back to row order in the config sheet
            pos_num = int(list(mm_config_df.index).index(slot_row.name)) + 1
        else:
            pos_num = int(float(pos_val))
    else:
        pos_num = int(list(mm_config_df.index).index(slot_row.name)) + 1

    alignment_by_pos = alignment_by_pos or {}
    gravity_by_pos = gravity_by_pos or {}
    thermal_by_pos = thermal_by_pos or {}

    x_mm = float(slot_row.get("x_MM [m]", 0.0))
    y_mm = float(slot_row.get("y_MM [m]", 0.0))
    z_mm = float(slot_row.get("z_MM [m]", 0.0))
    r_mm = float(slot_row.get("r_MM [m]", 0.0))
    if r_mm == 0.0:
        r_mm = float(np.hypot(x_mm, y_mm))
    if r_mm == 0.0:
        r_mm = 1e-9

    m_rad = float(mm_params.iloc[0]["m_rad"])
    m_azi = float(mm_params.iloc[0]["m_azi"])

    # Alignment rad/azi deltas
    if pos_num in alignment_by_pos:
        m_rad += float(alignment_by_pos[pos_num].get("d_align_rad", 0.0))
        m_azi += float(alignment_by_pos[pos_num].get("d_align_azi", 0.0))

    # rotz coupling affects azimuthal shift
    d_rotz_arcsec = 0.0
    if pos_num in alignment_by_pos:
        d_rotz_arcsec += float(alignment_by_pos[pos_num].get("d_align_rotz", 0.0))
    if pos_num in gravity_by_pos:
        d_rotz_arcsec += float(gravity_by_pos[pos_num].get("d_grav_rotz", 0.0))
    if pos_num in thermal_by_pos:
        d_rotz_arcsec += float(thermal_by_pos[pos_num].get("d_therm_rotz", 0.0))
    if d_rotz_arcsec != 0.0:
        d_rotz_rad = np.radians(d_rotz_arcsec / 3600.0)
        m_azi += r_mm * d_rotz_rad

    # Convert polar -> cartesian (same convention as rebuild_df)
    u_rad_x = x_mm / r_mm
    u_rad_y = y_mm / r_mm
    u_azi_x = -y_mm / r_mm
    u_azi_y = x_mm / r_mm
    mux = u_rad_x * m_rad + u_azi_x * m_azi
    muy = u_rad_y * m_rad + u_azi_y * m_azi

    # Apply gravity/thermal xy
    if pos_num in gravity_by_pos:
        mux += float(gravity_by_pos[pos_num].get("d_grav_x", 0.0))
        muy += float(gravity_by_pos[pos_num].get("d_grav_y", 0.0))
    if pos_num in thermal_by_pos:
        mux += float(thermal_by_pos[pos_num].get("d_therm_x", 0.0))
        muy += float(thermal_by_pos[pos_num].get("d_therm_y", 0.0))

    # z projection from combined dz
    dz = (
        float(alignment_by_pos.get(pos_num, {}).get("d_align_z", 0.0))
        + float(gravity_by_pos.get(pos_num, {}).get("d_grav_z", 0.0))
        + float(thermal_by_pos.get(pos_num, {}).get("d_therm_z", 0.0))
    )
    if dz != 0.0:
        denom = 12.0 - z_mm
        if denom != 0.0:
            mux += dz * x_mm / denom
            muy += dz * y_mm / denom

    shift_mag = float(np.hypot(mux, muy))
    return intrinsic + shift_mag


def azimuthal_placement(
    input_path: str,
    output_path: str,
    seed: int = 42,
    write_output: bool = True,
) -> float:
    """
    Place MMs using an azimuthal pattern based on individual MM HEW quality.
    
    Strategy:
    1. Sort MMs by individual HEW (best to worst)
    2. Place best MM at a random free position
    3. Place 2nd best diametrically opposite
    4. Place 3rd best ~90° from 1st
    5. Place 4th best diametrically opposite to 3rd
    6. Repeat pattern, rotating start position clockwise each cycle
    
    Returns:
        Final system HEW after placement
    """
    sheets = load_all_sheets(input_path)
    
    if "MM configuration" not in sheets:
        raise ValueError("MM configuration sheet missing")
    if "MM_PSF" not in sheets:
        raise ValueError("MM_PSF sheet missing")
    
    mm_config = sheets["MM configuration"].copy()
    
    base_params = _load_base_params_from_workbook(input_path)
    alignment_by_pos, gravity_by_pos, thermal_by_pos = _load_position_deltas(input_path)

    # Placement is constrained to permute only within each Row #.
    mm_config = _place_mm_config_within_rows(
        mm_config,
        base_params,
        placer=_azimuthal_place_mm_config,
        alignment_by_pos=alignment_by_pos,
        gravity_by_pos=gravity_by_pos,
        thermal_by_pos=thermal_by_pos,
        seed=seed,
    )
    
    # Compute final HEW
    final_df = rebuild_df(base_params, mm_config)
    final_hew = hew_fast_approximate(final_df)

    # Write output Excel while preserving formatting/formulas/images (optional).
    if write_output:
        _write_optimised_workbook_preserving_formatting(input_path, output_path, mm_config)
    
    return final_hew


def x_axis_placement(
    input_path: str,
    output_path: str,
    seed: int = 42,
    write_output: bool = True,
) -> float:
    """Place MMs biased toward the +/-x axis, alternating above/below the x-axis.

    This uses the same individual-MM HEW ranking as `azimuthal_placement`, but a different
    slot assignment order (see `_xaxis_place_mm_config`).

    Returns:
        Final system HEW after placement
    """
    sheets = load_all_sheets(input_path)

    if "MM configuration" not in sheets:
        raise ValueError("MM configuration sheet missing")
    if "MM_PSF" not in sheets:
        raise ValueError("MM_PSF sheet missing")

    mm_config = sheets["MM configuration"].copy()

    base_params = _load_base_params_from_workbook(input_path)
    alignment_by_pos, gravity_by_pos, thermal_by_pos = _load_position_deltas(input_path)

    # Placement is constrained to permute only within each Row #.
    mm_config = _place_mm_config_within_rows(
        mm_config,
        base_params,
        placer=_xaxis_place_mm_config,
        alignment_by_pos=alignment_by_pos,
        gravity_by_pos=gravity_by_pos,
        thermal_by_pos=thermal_by_pos,
        seed=seed,
    )

    # Compute final HEW
    final_df = rebuild_df(base_params, mm_config)
    final_hew = hew_fast_approximate(final_df)

    # Write output Excel while preserving formatting/formulas/images (optional).
    if write_output:
        _write_optimised_workbook_preserving_formatting(input_path, output_path, mm_config)

    return final_hew


def elliptical_placement(
    input_path: str,
    output_path: str,
    seed: int = 42,
    write_output: bool = True,
) -> float:
    """Row-wise placement: best MMs near x-axis, worst near y-axis.

    This permutes MM assignments only within each row of the MM configuration.
    """
    sheets = load_all_sheets(input_path)

    if "MM configuration" not in sheets:
        raise ValueError("MM configuration sheet missing")
    if "MM_PSF" not in sheets:
        raise ValueError("MM_PSF sheet missing")

    mm_config = sheets["MM configuration"].copy()

    base_params = _load_base_params_from_workbook(input_path)
    alignment_by_pos, gravity_by_pos, thermal_by_pos = _load_position_deltas(input_path)

    mm_config = _elliptical_place_mm_config(
        mm_config,
        base_params,
        alignment_by_pos=alignment_by_pos,
        gravity_by_pos=gravity_by_pos,
        thermal_by_pos=thermal_by_pos,
        seed=seed,
    )

    final_df = rebuild_df(base_params, mm_config)
    final_hew = hew_fast_approximate(final_df)

    if write_output:
        _write_optimised_workbook_preserving_formatting(input_path, output_path, mm_config)

    return final_hew


# Public/CLI-friendly aliases
cross_placement = azimuthal_placement
xaxis_placement = x_axis_placement


def permute_row(
    row_value,
    mm_config_df: pd.DataFrame,
    seed: int,
    swaps: int,
    params_df: pd.DataFrame,
    fast: bool,
    deadline: float | None = None,
):
    """Permute MM # assignments within a row while keeping locations (x,y,z,r) fixed.
    
    This finds the best assignment of MM numbers to physical locations in a row,
    using MM_PSF parameters to evaluate which assignment minimizes HEW.
    
    Note: Alignment, Gravity offload, and Thermal sheets are indexed by MM#, so they
    automatically follow the MM when it's assigned to a new location.
    """
    rng = random.Random(seed)
    if deadline is not None and time.time() >= deadline:
        return mm_config_df, None, None
    row_mask = mm_config_df["Row #"] == row_value
    indices = mm_config_df.index[row_mask].tolist()
    if len(indices) < 2:
        return mm_config_df, None, None
    
    # Extract MM numbers and locations for this row
    mm_numbers = mm_config_df.loc[indices, "MM #"].values.copy()
    locations = mm_config_df.loc[indices, ["x_MM [m]", "y_MM [m]", "z_MM [m]", "r_MM [m]"]].copy()
    
    working = mm_config_df.copy()
    best = working.copy()
    base_params = params_df.copy()
    best_df = rebuild_df(base_params, best)
    best_hew = hew_fast_approximate(best_df)
    
    # Simulated annealing over MM# permutations within this row.
    # We keep a "current" state that can accept worse moves early on, and a
    # separate "best" state that is always the best found so far.
    current_mm_order = mm_numbers.copy()
    current_hew = best_hew
    start_time = time.time()

    # Temperature schedule (in HEW units, meters). Start relatively hot to
    # encourage exploration, then cool down to become greedy.
    start_temp = max(1e-12, 0.25 * float(current_hew))
    end_temp = max(1e-12, start_temp * 1e-4)

    # Ensure working row matches current order (usually already true).
    for pos, idx in enumerate(indices):
        working.at[idx, "MM #"] = current_mm_order[pos]

    def time_fraction() -> float:
        if deadline is None:
            return 0.0
        total = max(1e-9, float(deadline) - float(start_time))
        return min(1.0, max(0.0, (time.time() - start_time) / total))

    for step in range(int(swaps)):
        if deadline is not None and time.time() >= deadline:
            break

        # Use wall-clock fraction when a deadline exists; otherwise use step-based.
        if deadline is None:
            frac = step / max(1.0, float(swaps) - 1.0)
        else:
            frac = time_fraction()

        # Exponential cooling.
        temp = start_temp * ((end_temp / start_temp) ** frac)
        temp = max(1e-12, float(temp))

        # Propose a simple move: swap two positions within the row.
        i1, i2 = rng.sample(range(len(indices)), 2)
        current_mm_order[i1], current_mm_order[i2] = current_mm_order[i2], current_mm_order[i1]
        idx1, idx2 = indices[i1], indices[i2]
        working.at[idx1, "MM #"] = current_mm_order[i1]
        working.at[idx2, "MM #"] = current_mm_order[i2]

        trial_df = rebuild_df(base_params, working)
        trial_hew = hew_fast_approximate(trial_df)

        delta = float(trial_hew) - float(current_hew)
        if delta <= 0.0:
            accept = True
        else:
            # Accept worse moves with Boltzmann probability.
            accept = (rng.random() < math.exp(-delta / temp))

        if accept:
            current_hew = float(trial_hew)
            if trial_hew < best_hew:
                best_hew = trial_hew
                best = working.copy()
                best_df = trial_df
        else:
            # Revert the move.
            current_mm_order[i1], current_mm_order[i2] = current_mm_order[i2], current_mm_order[i1]
            working.at[idx1, "MM #"] = current_mm_order[i1]
            working.at[idx2, "MM #"] = current_mm_order[i2]

        # Optional occasional "kick" to escape deep local minima.
        # Kept rare to preserve time bounds.
        if (step + 1) % 200 == 0 and (deadline is None or time.time() < deadline):
            j1, j2 = rng.sample(range(len(indices)), 2)
            current_mm_order[j1], current_mm_order[j2] = current_mm_order[j2], current_mm_order[j1]
            id1, id2 = indices[j1], indices[j2]
            working.at[id1, "MM #"] = current_mm_order[j1]
            working.at[id2, "MM #"] = current_mm_order[j2]
            kick_df = rebuild_df(base_params, working)
            kick_hew = hew_fast_approximate(kick_df)
            delta_k = float(kick_hew) - float(current_hew)
            if delta_k <= 0.0 or (rng.random() < math.exp(-delta_k / temp)):
                current_hew = float(kick_hew)
                if kick_hew < best_hew:
                    best_hew = kick_hew
                    best = working.copy()
                    best_df = kick_df
            else:
                current_mm_order[j1], current_mm_order[j2] = current_mm_order[j2], current_mm_order[j1]
                working.at[id1, "MM #"] = current_mm_order[j1]
                working.at[id2, "MM #"] = current_mm_order[j2]
    
    return best, best_hew, best_df


def optimize_rows(
    input_path: str,
    output_path: str,
    mode: str,
    optimize: bool = True,
    time_budget_s: float = 55.0,
    start_placement: str = "cross",
    write_output: bool = True,
):
    # Keep wall-time bounded.
    # Note: HEW evaluation is the expensive part, so we cap iterations.
    # Upper bound: work will stop early once the deadline hits.
    if mode not in {"coarse", "fine", "extra-fine"}:
        raise ValueError(f"Unknown mode: {mode}")

    is_coarse = (mode == "coarse")
    is_extra_fine = (mode == "extra-fine")

    if is_coarse:
        swaps_per_row = 160
    elif is_extra_fine:
        swaps_per_row = 1200
    else:
        swaps_per_row = 220
    seed = 42  # Fixed seed for reproducibility

    # Hard wall-clock budget for optimization work.
    # (Caller should leave time for Excel IO / plotting as needed.)
    opt_deadline = time.time() + float(time_budget_s)
    
    sheets = load_all_sheets(input_path)
    if "MM configuration" not in sheets:
        raise ValueError("MM configuration sheet missing")
    mm_config = sheets["MM configuration"].copy()
    if "MM_PSF" not in sheets:
        raise ValueError("MM_PSF sheet missing")
    psf_df = sheets["MM_PSF"].copy()

    # Load base MM PSF parameters (intrinsic to MM) without applying position deltas.
    # Position deltas must stay with the slot during swapping, so we apply them
    # dynamically during evaluation.
    base_params = _load_base_params_from_workbook(input_path)

    # Per-position deltas
    alignment_by_pos, gravity_by_pos, thermal_by_pos = _load_position_deltas(input_path)

    best_mm_config = mm_config.copy()
    final_hew = None
    
    if optimize:
        # Iterated local search starting from azimuthal placement.
        # We keep search scoped to a subset of rows (mode-dependent) for speed,
        # but the starting point is a full placement.
        row_counts = mm_config["Row #"].value_counts(dropna=True)
        row_values = [int(v) for v in row_counts.index.tolist()]
        if is_coarse:
            selected_rows = row_values[:2]
        elif is_extra_fine:
            selected_rows = row_values
        else:
            selected_rows = row_values[:6]

        # Step 0: start from a full placement (user-selectable strategy).
        if time.time() < opt_deadline:
            if start_placement in {"cross", "azimuthal90"}:
                current_mm_config = _place_mm_config_within_rows(
                    mm_config,
                    base_params,
                    placer=_azimuthal_place_mm_config,
                    alignment_by_pos=alignment_by_pos,
                    gravity_by_pos=gravity_by_pos,
                    thermal_by_pos=thermal_by_pos,
                    seed=seed,
                )
            elif start_placement in {"x_axis", "xaxis"}:
                current_mm_config = _place_mm_config_within_rows(
                    mm_config,
                    base_params,
                    placer=_xaxis_place_mm_config,
                    alignment_by_pos=alignment_by_pos,
                    gravity_by_pos=gravity_by_pos,
                    thermal_by_pos=thermal_by_pos,
                    seed=seed,
                )
            elif start_placement == "elliptical":
                current_mm_config = _elliptical_place_mm_config(
                    mm_config,
                    base_params,
                    alignment_by_pos=alignment_by_pos,
                    gravity_by_pos=gravity_by_pos,
                    thermal_by_pos=thermal_by_pos,
                    seed=seed,
                )
            else:
                raise ValueError(f"Unknown start_placement: {start_placement}")
        else:
            current_mm_config = mm_config.copy()

        # Incremental objective evaluator (fixed polar grid).
        if time.time() >= opt_deadline:
            best_mm_config = current_mm_config
            best_df = rebuild_df(base_params, best_mm_config)
            final_hew = hew_fast_approximate(best_df)
        else:
            evaluator = IncrementalHEWApprox(
                base_params=base_params,
                mm_config=current_mm_config,
                alignment_by_pos=alignment_by_pos,
                gravity_by_pos=gravity_by_pos,
                thermal_by_pos=thermal_by_pos,
                n_r=60,
                n_theta=40,
            )

            # Restrict moves to selected rows, but build indices in evaluator space.
            idx_to_pos = evaluator.idx_to_pos
            row_to_positions: dict[int, list[int]] = {}
            for rv in selected_rows:
                mask = current_mm_config["Row #"] == rv
                idxs = current_mm_config.index[mask].tolist()
                positions = [idx_to_pos[i] for i in idxs if i in idx_to_pos]
                if len(positions) >= 2:
                    row_to_positions[rv] = positions

            movable_rows = list(row_to_positions.keys())
            if not movable_rows:
                best_mm_config = current_mm_config
                best_df = rebuild_df(base_params, best_mm_config)
                final_hew = hew_fast_approximate(best_df)
            else:
                rng = random.Random(seed)

                current_score = evaluator.hew()
                best_score = float(current_score)
                best_assignment = evaluator.assignment.copy()

                # SA temperature schedule for acceptance between local minima.
                start_time = time.time()
                start_temp = max(1e-12, 0.10 * float(current_score))
                end_temp = max(1e-12, start_temp * 1e-4)

                def temperature() -> float:
                    frac = min(1.0, max(0.0, (time.time() - start_time) / max(1e-9, opt_deadline - start_time)))
                    t = start_temp * ((end_temp / start_temp) ** frac)
                    return max(1e-12, float(t))

                # ILS parameters (mode-dependent)
                if is_coarse:
                    stagnation_limit = 120
                    perturb_k = 6
                elif is_extra_fine:
                    stagnation_limit = 800
                    perturb_k = 16
                else:
                    stagnation_limit = 300
                    perturb_k = 10

                stagnation = 0
                steps = 0

                while time.time() < opt_deadline:
                    steps += 1

                    rv = rng.choice(movable_rows)
                    positions = row_to_positions[rv]
                    pos_a, pos_b = rng.sample(positions, 2)

                    # Propose swap (incremental update)
                    evaluator.swap_slots(pos_a, pos_b)
                    new_score = evaluator.hew()
                    delta = float(new_score) - float(current_score)
                    t = temperature()
                    accept = (delta <= 0.0) or (rng.random() < math.exp(-delta / t))

                    if accept:
                        current_score = float(new_score)
                        if current_score < best_score:
                            best_score = float(current_score)
                            best_assignment = evaluator.assignment.copy()
                            stagnation = 0
                        else:
                            stagnation += 1
                    else:
                        # Revert swap
                        evaluator.swap_slots(pos_a, pos_b)
                        stagnation += 1

                    # Perturbation when stuck (iterated local search)
                    if stagnation >= stagnation_limit and time.time() < opt_deadline:
                        for _ in range(int(perturb_k)):
                            rv2 = rng.choice(movable_rows)
                            pos_list = row_to_positions[rv2]
                            a2, b2 = rng.sample(pos_list, 2)
                            evaluator.swap_slots(a2, b2)
                        current_score = float(evaluator.hew())
                        stagnation = 0

                # Materialize best assignment into output config
                best_mm_config = current_mm_config.copy()
                for pos, idx in enumerate(evaluator.slot_indices):
                    best_mm_config.at[idx, "MM #"] = int(best_assignment[pos])

                best_df = rebuild_df(base_params, best_mm_config)
                best_df = _apply_position_deltas_to_df(
                    best_df,
                    best_mm_config,
                    alignment_by_pos,
                    gravity_by_pos,
                    thermal_by_pos,
                )
                remaining = max(1.0, opt_deadline - time.time())
                try:
                    final_hew = hew_at_best_focus(best_df, fast=is_coarse, timeout=remaining)
                except TimeoutError:
                    final_hew = hew_fast_approximate(best_df)

        if best_mm_config["MM #"].astype(int).equals(mm_config["MM #"].astype(int)):
            print("Note: optimizer did not find an improving MM# permutation within the time budget; MM configuration is unchanged.")


    # Write output Excel while preserving formatting/formulas/images (optional).
    # Only the MM configuration's MM# cells are patched.
    if write_output:
        _write_optimised_workbook_preserving_formatting(input_path, output_path, best_mm_config)
    
    return final_hew


if __name__ == "__main__":
    # This block is essential for multiprocessing on Windows and macOS
    parser = argparse.ArgumentParser(description="Copy MM configuration to new file, optionally optimizing MM positions within each row to minimize HEW.")
    parser.add_argument("-f", "--file", required=True, help="Input Excel file with MM configuration and MM_PSF sheets")
    parser.add_argument("-o", "--output", default=None, help="Output Excel file path (default: input_optimised.xlsx)")
    parser.add_argument(
        "--mode",
        type=str,
        choices=["coarse", "fine", "extra-fine"],
        default="coarse",
        help="Runtime mode: coarse, fine, or extra-fine. Controls optimization speed/accuracy.",
    )
    parser.add_argument(
        "--optimize",
        action="store_true",
        default=False,
        help="Enable MM position optimization (uses --mode for speed/accuracy).",
    )
    args = parser.parse_args()
    
    # Auto-generate output filename if not provided
    output_path = args.output
    if output_path is None:
        import os
        base, ext = os.path.splitext(args.file)
        output_path = f"{base}_optimised{ext}"
    
    if args.optimize:
        # Keep this aligned with main.py so users get consistent behavior.
        if args.mode == "coarse":
            time_budget_s = 18.0
        elif args.mode == "extra-fine":
            time_budget_s = 240.0
        else:
            time_budget_s = 45.0
        best_hew = optimize_rows(
            input_path=args.file,
            output_path=output_path,
            mode=args.mode,
            optimize=True,
            time_budget_s=time_budget_s,
        )
        print(f"Optimized HEW: {best_hew:.6e} m")
    else:
        # If not optimizing, just copy the file
        import shutil
        shutil.copy(args.file, output_path)
        print("No optimization performed, file copied.")

    print(f"Output saved to: {output_path}")
