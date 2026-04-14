import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Distributions/test10kev.xlsx')
ws = wb['A_eff']

print(f"Sheet: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Check columns 9-16 (which are the energy preset columns)
print("\n=== Energy preset columns (9-16) ===")
for col_idx in range(9, 17):
    cell_header = ws.cell(row=1, column=col_idx)
    cell_data = ws.cell(row=2, column=col_idx)  # MM 1 data
    print(f"Col {col_idx}: header='{cell_header.value}', MM1_cell={cell_data}")
    print(f"  Cell type: {cell_data.data_type}, Value: {cell_data.value}, Formula: {cell_data.value if cell_data.data_type == 'f' else 'N/A'}")

# Also check the left columns
print("\n=== Left columns (1-5) ===")
for col_idx in range(1, 6):
    cell_header = ws.cell(row=1, column=col_idx)
    cell_data = ws.cell(row=2, column=col_idx)  # MM 1 data
    print(f"Col {col_idx}: header='{cell_header.value}', MM1_value='{cell_data.value}' (type: {cell_data.data_type})")
