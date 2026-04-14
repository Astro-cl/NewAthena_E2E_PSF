#!/usr/bin/env python3
"""Test the corrected compute logic."""

import openpyxl

# Use a different test file or copy the original
import shutil

# Make a fresh copy
print("Making a copy to test with...")
try:
    shutil.copy('Distributions/test10kev.xlsx', 'Distributions/test10kev_fresh.xlsx')
    print("Created test10kev_fresh.xlsx")
except Exception as e:
    print(f"Could not copy: {e}")
    # Try listing what files we have
    import os
    files = os.listdir('Distributions')
    print(f"Files in Distributions: {[f for f in files if 'test' in f.lower() or 'ref' in f.lower()]}")
    exit(1)

# Now process the fresh copy
wb = openpyxl.load_workbook('Distributions/test10kev_fresh.xlsx')
aeff = wb['A_eff']

# Check a sample formula before processing
print("\nBefore processing:")
print(f"  J2 data_type: {aeff['J2'].data_type}, value: {aeff['J2'].value}")

# Now reverse the previous change by reloading with data_only=True to get cached values
# Actually, let me just manually verify the logic would work

mm_config = wb['MM configuration']
mm_to_row = {}
for row in mm_config.iter_rows(min_row=2, max_row=610):
    row_cell = row[2] if len(row) > 2 else None  # Column C (Row #)
    mm_cell = row[3] if len(row) > 3 else None    # Column D (MM #)
    
    if mm_cell is None or row_cell is None:
        break
    if mm_cell.value is not None and row_cell.value is not None:
        try:
            mm = int(float(mm_cell.value))
            row_num = int(float(row_cell.value))
            mm_to_row[mm] = row_num
        except Exception:
            pass

# Get system margins
system_margins = {}
for col_idx in range(20, 28):
    val = aeff.cell(5, col_idx).value
    if val is not None:
        try:
            system_margins[col_idx] = float(val)
        except Exception:
            system_margins[col_idx] = 0

# Compute lookup data for just a few rows
lookup_data = {}
for source_row_num in range(6, 9):  # Just rows 6-8 for testing
    lookup_row = source_row_num - 5
    focal_len = aeff.cell(source_row_num, 19).value
    
    if focal_len is None:
        break
    
    try:
        focal_len = float(focal_len)
    except Exception:
        continue
    
    for col_idx in range(20, 23):  # Just a few columns
        base_val = aeff.cell(source_row_num, col_idx).value
        if base_val is None:
            continue
        
        try:
            base_val = float(base_val)
            margin = system_margins.get(col_idx, 0)
            
            if focal_len != 0:
                adjusted = base_val * (1.0 - margin) / focal_len
            else:
                adjusted = 0
            
            lookup_data[(lookup_row, col_idx)] = adjusted
        except Exception:
            pass

print("\nComputed lookup data sample:")
for key, val in sorted(lookup_data.items()):
    print(f"  Lookup row {key[0]}, Col {openpyxl.utils.get_column_letter(key[1])}: {val:.6f}")

# Now show what would be filled for MM 1 (which has Row# 1)
print("\n For MM 1 (Row# 1):")
mm_row = mm_to_row.get(1)
print(f"  MM 1 -> Row# {mm_row}")
for col_idx in range(10, 13):  # J, K, L
    energy_col = col_idx + 10  # J(10) -> 20(T), etc.
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    if (mm_row, energy_col) in lookup_data:
        val = lookup_data[(mm_row, energy_col)]
        print(f"  Column {col_letter}: {val:.6f}")
    else:
        print(f"  Column {col_letter}: NOT FOUND")
