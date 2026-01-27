from pathlib import Path
import pandas as pd, shutil, tempfile, json, subprocess, sys
from main import load_gaussians_from_excel
from tools.run_sensitivity import apply_mm_psf_choice_to_df, build_aeff_mapping

BASE=Path('Distributions/Test_Distribution.xlsx')
df=load_gaussians_from_excel(str(BASE), sheet='MM_PSF')
aeff_map=build_aeff_mapping()
from numpy.random import default_rng
rng=default_rng(12345)
choice='Fixed Sym Gaussian (sigma_rad=0.05 micron, sigma_azi=0.05 micron)'
df2=apply_mm_psf_choice_to_df(df.copy(), choice, aeff_map, rng)
# write temp workbook
tmpdir=tempfile.mkdtemp(prefix='debug_')
tmp_in=str(Path(tmpdir)/BASE.name)
shutil.copy2(str(BASE), tmp_in)
from openpyxl import load_workbook
wb = load_workbook(tmp_in)
if 'MM_PSF' in wb.sheetnames:
    std = wb['MM_PSF']
    wb.remove(std)
wb.save(tmp_in)
with pd.ExcelWriter(tmp_in, engine='openpyxl', mode='a') as w:
    df2.to_excel(w, sheet_name='MM_PSF', index=False)
cmd=[sys.executable, 'main.py', '-f', tmp_in, '--return_metrics_only', '--metrics-nr-final','50','--metrics-ntheta-final','12','--metrics-r-margin','6.0']
print('Running:', ' '.join(cmd))
proc=subprocess.run(cmd, capture_output=True, text=True, timeout=120)
print('RC:', proc.returncode)
print('STDOUT (first 2000 chars):')
print(proc.stdout[:2000])
print('STDERR (first 2000 chars):')
print(proc.stderr[:2000])
try:
    j=json.loads(proc.stdout)
    print('Top-level keys:', list(j.keys()))
    if 'post_raw' in j:
        print('post_raw.hew_opt_arcsec=', j['post_raw'].get('hew_opt_arcsec'))
    if 'hew_opt_arcsec' in j:
        print('hew_opt_arcsec top-level=', j.get('hew_opt_arcsec'))
except Exception as e:
    print('JSON load error', e)
finally:
    shutil.rmtree(tmpdir)
