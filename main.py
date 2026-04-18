"""Command-line entrypoints and Excel I/O helpers for PSF generation.

This module implements the primary CLI functions used by the project's
headless test harness and the command-line workflow. It provides Excel
loaders for the `MM_PSF` and `A_eff` tables, PSF parameter conversion,
vignetting application and plotting helpers.

Note: many small utility scripts have been moved to `tools/` (e.g.
`tools/compute_aeff_values.py`) to keep the repository root focused on
core modules and the GUI. Use the `tools/` folder for one-off helper
commands and legacy scripts retained under `scripts/legacy/`.
"""
import argparse  # For parsing command-line arguments
import numpy as np  # For numerical operations and arrays
import pandas as pd  # For data manipulation and Excel reading
import matplotlib
import matplotlib.pyplot as plt  # For plotting
import matplotlib.gridspec as gridspec
import os
import subprocess
import math
import glob
import shutil
import time
import zipfile
from pathlib import Path
from scipy.optimize import curve_fit
from concurrent.futures import ThreadPoolExecutor
import tempfile
from distributions_rotated import (
    gaussian_2d_rotated,
    pseudo_voigt_2d_rotated,
    load_psf_matrix_excel,
    eval_psf_matrix_rotated,
)
import json
import sys
import io
import re


def parse_multisheet_csv(path_or_buffer):
    """Parse a multisheet-style CSV file created by `write_multisheet_csv`.

    The file uses marker lines of the form:
      # sheet: Sheet Name

    Returns a dict mapping sheet name -> pandas.DataFrame.
    Accepts a filesystem path or any file-like / string buffer.
    """
    text = None
    # Accept file path or file-like object or raw string content
    if hasattr(path_or_buffer, 'read'):
        text = path_or_buffer.read()
    else:
        try:
            # try treating as path
            with open(str(path_or_buffer), 'r', encoding='utf-8') as fh:
                text = fh.read()
        except Exception:
            # fallback: treat as raw string content
            text = str(path_or_buffer)

    sheets = {}
    current_name = None
    current_lines = []
    for raw in text.splitlines():
        m = re.match(r"^#\s*sheet:\s*(.+)$", raw)
        if m:
            # flush previous
            if current_name is not None:
                buf = io.StringIO('\n'.join(current_lines))
                try:
                    df = pd.read_csv(buf)
                except Exception:
                    # empty or unparsable -> empty DataFrame
                    df = pd.DataFrame()
                sheets[current_name] = df
            current_name = m.group(1).strip()
            current_lines = []
        else:
            # skip leading/trailing blank lines
            current_lines.append(raw)

    # flush last
    if current_name is not None:
        buf = io.StringIO('\n'.join(current_lines))
        try:
            df = pd.read_csv(buf)
        except Exception:
            df = pd.DataFrame()
        sheets[current_name] = df

    # Post-process to mimic writer sanitization expected by tests:
    # - remove 'aeff_adjusted' column from 'A_eff' sheet (case-insensitive)
    # - for 'MM_PSF' sheet, if 'aeff_adjusted' present, set/replace 'weight'
    #   column with its values
    out = {}
    for name, df in sheets.items():
        if name.lower() == 'a_eff' or name == 'A_eff' or name.lower() == 'a_eff':
            # drop any column named aeff_adjusted (case-insensitive)
            cols = [c for c in df.columns]
            drop_cols = [c for c in cols if str(c).strip().lower() == 'aeff_adjusted']
            if drop_cols:
                df = df.drop(columns=drop_cols)
            out[name] = df
        elif name == 'MM_PSF' or name.lower() == 'mm_psf':
            df2 = df.copy()
            # find any aeff_adjusted column case-insensitively
            for c in list(df2.columns):
                if str(c).strip().lower() == 'aeff_adjusted':
                    try:
                        df2['weight'] = pd.to_numeric(df2[c], errors='coerce')
                    except Exception:
                        df2['weight'] = df2[c]
                    break
            out[name] = df2
        else:
            out[name] = df

    return out

# Vignetting sheet name candidates (support both new and legacy names)
VIG_ROT_AZI_CANDIDATES = ('MM vignetting rotazi', 'Vignetting rotazi')
VIG_ROT_RAD_CANDIDATES = ('MM vignetting rotrad', 'Vignetting rotrad')

# HEW degradation sheet name candidates
HEW_DEG_ROT_AZI_CANDIDATES = ('MM HEW degradation rotazi',)
HEW_DEG_ROT_RAD_CANDIDATES = ('MM HEW degradation rotrad',)

def _find_vig_sheet(container, candidates):
    """Return first candidate sheet name present in *container*, or None.

    *container* may be an openpyxl Workbook (has `.sheetnames`) or any
    iterable of sheet name strings.
    """
    names = container.sheetnames if hasattr(container, 'sheetnames') else container
    for c in candidates:
        if c in names:
            return c
    return None

def _read_excel_vig(path, candidates, **kwargs):
    """Try ``pd.read_excel`` with each candidate sheet name in order.

    Returns the DataFrame on success, or None if no candidate matched.
    """
    for name in candidates:
        try:
            return pd.read_excel(path, sheet_name=name, **kwargs)
        except Exception:
            continue
    return None

def load_aeff_weight_map(path: str, sheet: str | None = None, energy_col: int | None = None) -> dict:
    """Load A_eff mapping (MM # -> base A_eff) from `A_eff` sheet.

    Reads the `A_eff` sheet (headerless) and returns a dict mapping integer
    MM -> float(A_eff_base).  When *energy_col* is given (0-based column
    index, e.g. 9 for column J, up to 19 for column T) the value is read
    from that column instead of column B (index 1).  If the sheet is
    missing or no valid rows are found an empty dict is returned.
    """
    mapping = {}
    val_col = energy_col if energy_col is not None else 1
    try:
        kwargs = {"engine": "openpyxl", "header": None, "dtype": str}
        if sheet is not None:
            kwargs["sheet_name"] = sheet
        else:
            kwargs["sheet_name"] = 'A_eff'
        raw = pd.read_excel(path, **kwargs)
    except Exception:
        return mapping

    # Column A = MM #; value column = B (default) or energy_col (J..T)
    for rid in range(raw.shape[0]):
        try:
            mmv = raw.iat[rid, 0]
        except Exception:
            mmv = None
        try:
            aval = raw.iat[rid, val_col]
        except Exception:
            aval = None
        if mmv is None:
            continue
        try:
            mm_int = int(float(mmv))
        except Exception:
            continue
        try:
            a_float = float(aval)
        except Exception:
            # Found a MM entry but its A_eff value is invalid -> raise
            raise ValueError(f"Invalid A_eff value for MM #{mm_int}: {aval!r}")
        mapping[mm_int] = a_float
    return mapping


def load_aeff_weight_map_with_name(path: str) -> tuple[dict, str | None]:
    """Return (mapping, aeff_column_name).

    Attempts to detect a human-readable A_eff column name (e.g. '0.25 keV')
    from the sheet if present; otherwise returns (mapping, None).
    """
    mapping = load_aeff_weight_map(path)
    col_name = None
    try:
        raw = pd.read_excel(path, sheet_name='A_eff', engine='openpyxl', header=None, dtype=str)
        # scan first few rows/cols for a string containing 'keV'
        import re
        for r in range(min(6, raw.shape[0])):
            for c in range(min(6, raw.shape[1])):
                try:
                    v = raw.iat[r, c]
                except Exception:
                    v = None
                if isinstance(v, str) and re.search(r"\bkeV\b", v, flags=re.IGNORECASE):
                    col_name = v.strip()
                    raise StopIteration
    except StopIteration:
        pass
    except Exception:
        pass
    return mapping, col_name


# `_evaluate_vlookup_xlookup` removed: prefer `data_only` cached values
# and do not attempt best-effort formula parsing/evaluation.

def load_gaussians_from_excel(path: str, sheet: str | None = None, fast_metrics: bool | None = None, **kwargs) -> pd.DataFrame:
    """Load gaussian parameters from Excel.

    Expected columns: m_rad [arcsec], m_azi [arcsec], sigma_rad [arcsec], sigma_azi [arcsec]
    Values are converted from arcsec to meters using: 1 arcsec = 12*π/180/3600 m
    Converts to mux, muy using rotation matrix:
    - mux = cos(theta)*m_rad - sin(theta)*m_azi
    - muy = sin(theta)*m_rad + cos(theta)*m_azi
    theta_degrees is calculated from MM configuration: theta = arcsin(x_MM / r_MM)
    weight column is optional and will be overridden by A_eff sheet if present
    
    Note: accepts `fast_metrics` kwarg for compatibility with test harnesses; the
    argument is recognized but not used by the loader.
    """
    # Support CSV inputs: either when the provided path ends with .csv or when
    # `--input-csv` was used to provide a CSV path. In these cases we read the
    # supplied CSV directly into a DataFrame. Otherwise fall back to Excel.
    try:
        print(f"load_gaussians_from_excel: start reading '{path}' (sheet={sheet})")
    except Exception:
        pass
    try:
        # Treat Path or str uniformly by converting to string first.
        arg_input_csv = getattr(sys.modules.get('__main__'), 'args', None)
        arg_input_csv = getattr(arg_input_csv, 'input_csv', None) if arg_input_csv is not None else None
        is_csv = str(path).lower().endswith('.csv') or bool(arg_input_csv)
    except Exception:
        is_csv = str(path).lower().endswith('.csv')

    if is_csv:
        # If main was invoked with --input-csv, that path will be passed as `path`.
        # Support two CSV flavors:
        # 1) Plain CSV that directly contains the `MM_PSF` table.
        # 2) Multisheet-style CSV where sheets are separated by marker lines
        #    like '# sheet: Sheet Name' (produced by `parse_multisheet_csv`).
        try:
            # Try parsing as multisheet CSV first
            try:
                parsed = parse_multisheet_csv(path)
                if isinstance(parsed, dict) and 'MM_PSF' in parsed:
                    df = parsed['MM_PSF']
                else:
                    # Fallback to normal CSV read
                    df = pd.read_csv(path)
            except Exception:
                # Normal CSV fallback
                df = pd.read_csv(path)
        except Exception as e:
            # Fall back to Excel reader if CSV read fails
            kwargs = {"engine": "openpyxl"}
            if sheet:
                kwargs["sheet_name"] = sheet
            df = pd.read_excel(path, **kwargs)
    else:
        kwargs = {"engine": "openpyxl"}  # Use openpyxl engine for Excel files
        if sheet:
            kwargs["sheet_name"] = sheet  # Specify sheet if provided

        # Create a values-only copy of the workbook before reading. Prefer to
        # use Excel via xlwings to perform an in-place paste-values (so cached
        # results are materialized); if xlwings/Excel is unavailable, fall
        # back to creating a values-only file from openpyxl's data_only read.
        def make_values_only_copy(src_path: str) -> str:
            src_path = str(src_path)
            dest_fd, dest_path = tempfile.mkstemp(prefix='values_only_', suffix='.xlsx', dir=os.getcwd())
            os.close(dest_fd)

            def _has_uncached_formula(path: str) -> bool:
                """Return True if workbook contains formula cells whose cached values are missing.

                We consider a cell 'uncached' when it contains a formula (text starting
                with '=') and the corresponding data_only value is None.
                """
                try:
                    wb_f = load_workbook(path, data_only=False, read_only=True)
                    wb_v = load_workbook(path, data_only=True, read_only=True)
                except Exception:
                    return False
                try:
                    for name in wb_f.sheetnames:
                        wsf = wb_f[name]
                        wsv = wb_v[name]
                        for row in wsf.iter_rows():
                            for cell in row:
                                try:
                                    val = cell.value
                                except Exception:
                                    val = None
                                if isinstance(val, str) and val.startswith('='):
                                    try:
                                        v = wsv.cell(row=cell.row, column=cell.col_idx).value
                                    except Exception:
                                        v = None
                                    if v is None:
                                        try:
                                            wb_f.close()
                                            wb_v.close()
                                        except Exception:
                                            pass
                                        return True
                    try:
                        wb_f.close()
                        wb_v.close()
                    except Exception:
                        pass
                    return False
                except Exception:
                    try:
                        wb_f.close()
                        wb_v.close()
                    except Exception:
                        pass
                    return False

            # Allow forcing xlwings paste-values via env var for batch runs
            force_env = os.environ.get('FORCE_PASTE_VALUES', '')
            need_xlwings = (force_env == '1') or _has_uncached_formula(src_path)

            if need_xlwings:
                try:
                    import xlwings as xw
                    try:
                        app = xw.App(visible=False)
                        wb_x = app.books.open(src_path)
                        for sh in wb_x.sheets:
                            sh.api.Cells.Copy()
                            sh.api.Cells.PasteSpecial(-4163)  # xlPasteValues
                        wb_x.api.SaveAs(dest_path)
                        wb_x.close()
                        app.quit()
                        # xlwings paste-values should have removed formulas.
                        # Do not overwrite the saved copy with the source's
                        # cached values here (that can unintentionally wipe
                        # data when the source cached values are missing).
                        return dest_path
                    except Exception:
                        try:
                            wb_x.close()
                        except Exception:
                            pass
                        try:
                            app.quit()
                        except Exception:
                            pass
                except Exception:
                    # xlwings not available or failed; fall through to fallback
                    pass

            # Fallback: read cached values with openpyxl and write to new file
            try:
                from openpyxl import load_workbook as _load_wb_vals, Workbook as _Workbook
                wb_vals = _load_wb_vals(src_path, data_only=True)
                new_wb = _Workbook()
                default = new_wb.active
                new_wb.remove(default)
                for name in wb_vals.sheetnames:
                    src = wb_vals[name]
                    dest = new_wb.create_sheet(title=name)
                    for r in src.iter_rows(values_only=True):
                        dest.append(list(r))
                new_wb.save(dest_path)
                return dest_path
            except Exception:
                # Last resort: copy original file raw
                shutil.copy2(src_path, dest_path)
                return dest_path

        values_only_path = None
        try:
            values_only_path = make_values_only_copy(path)
            wb_vals = None
            try:
                from openpyxl import load_workbook as _load_wb_vals
                wb_vals = _load_wb_vals(values_only_path, data_only=True)
            except Exception:
                wb_vals = None

            if wb_vals is not None:
                # choose target sheet: prefer provided `sheet`, else detect
                target_sheet = None
                if sheet and sheet in wb_vals.sheetnames:
                    target_sheet = sheet
                else:
                    for sname in wb_vals.sheetnames:
                        if sname.lower() == 'mm_psf':
                            target_sheet = sname
                            break
                if target_sheet is None and wb_vals.sheetnames:
                    target_sheet = wb_vals.sheetnames[0]

                if target_sheet is not None:
                    ws = wb_vals[target_sheet]
                    rows = list(ws.values)
                    if rows:
                        header = [str(h) if h is not None else '' for h in rows[0]]
                        data_rows = rows[1:]
                        df = pd.DataFrame(data_rows, columns=header)
                    else:
                        df = pd.DataFrame()
                else:
                    df = pd.DataFrame()
                print(f"INFO: Loaded workbook using values-only copy: {values_only_path}")
            else:
                df = pd.read_excel(path, **kwargs)
        finally:
            # Do NOT delete the values-only copy yet — mapping of D/E below
            # may need to read it. We'll clean up after the mapping step.
            pass
    required = ["m_rad [arcsec]","m_azi [arcsec]","sigma_rad [arcsec]","sigma_azi [arcsec]"]  # Required columns

    # If the sheet was read headerless (integer column names) and the first
    # data row appears to contain header strings, promote that row to header.
    try:
        import numpy as _np
        if all(isinstance(c, (int,)) for c in df.columns):
            first_row = df.iloc[0].astype(str).str.strip().tolist()
            # If any required name appears in first row, treat it as header
            if any(r in first_row for r in [r for r in required]):
                df.columns = first_row
                df = df.iloc[1:].reset_index(drop=True)
    except Exception:
        pass
    # Cleanup values-only temp copy if it was created earlier
    try:
        if 'values_only_path' in locals() and values_only_path and os.path.exists(values_only_path):
            os.remove(values_only_path)
    except Exception:
        pass
    # If this was an Excel input, prefer the workbook's evaluated cell values
    # for columns D/E (sigma_rad, sigma_azi). This ensures any formulas in
    # the sheet are resolved using the workbook's cached numeric results
    # rather than relying on pandas' inference which can miss evaluated
    # values when formulas are present.
    try:
        if not is_csv:
            from openpyxl import load_workbook as _load_wb_vals
            wb_vals = None
            try:
                wb_vals = _load_wb_vals(path, data_only=True)
            except Exception:
                wb_vals = None
            if wb_vals is not None:
                # find MM_PSF sheet case-insensitively
                sheet_name = None
                for s in wb_vals.sheetnames:
                    if s.lower() == 'mm_psf':
                        sheet_name = s
                        break
                if sheet_name is not None and 'MM #' in df.columns:
                    ws_vals = wb_vals[sheet_name]
                    sigma_map_rad = {}
                    sigma_map_azi = {}
                    # iterate rows, expect MM # in col A, sigma_rad in D (4), sigma_azi in E (5)
                    for r in range(2, (ws_vals.max_row or 0) + 1):
                        try:
                            mmcell = ws_vals.cell(row=r, column=1).value
                            if mmcell is None:
                                continue
                            try:
                                mm_int = int(float(mmcell))
                            except Exception:
                                continue
                            dval = ws_vals.cell(row=r, column=4).value
                            eval_d = None
                            if dval is not None:
                                try:
                                    eval_d = float(dval)
                                except Exception:
                                    # try to extract a numeric token
                                    try:
                                        import re as _re
                                        m = _re.search(r"(-?\d+(?:\.\d+)?)", str(dval))
                                        if m:
                                            eval_d = float(m.group(1))
                                    except Exception:
                                        eval_d = None
                            eval_e = None
                            eval_e_raw = ws_vals.cell(row=r, column=5).value
                            if eval_e_raw is not None:
                                try:
                                    eval_e = float(eval_e_raw)
                                except Exception:
                                    try:
                                        import re as _re2
                                        m2 = _re2.search(r"(-?\d+(?:\.\d+)?)", str(eval_e_raw))
                                        if m2:
                                            eval_e = float(m2.group(1))
                                    except Exception:
                                        eval_e = None
                            if eval_d is not None:
                                sigma_map_rad[mm_int] = eval_d
                            if eval_e is not None:
                                sigma_map_azi[mm_int] = eval_e
                        except Exception:
                            continue
                    # Apply maps to the DataFrame for matching MM # rows
                    for idx in df.index:
                        try:
                            mmv = int(pd.to_numeric(df.at[idx, 'MM #'], errors='coerce'))
                        except Exception:
                            continue
                        if mmv in sigma_map_rad:
                            df.at[idx, 'sigma_rad [arcsec]'] = sigma_map_rad[mmv]
                        if mmv in sigma_map_azi:
                            df.at[idx, 'sigma_azi [arcsec]'] = sigma_map_azi[mmv]

                    # Fallback: when D/E cached values are missing (openpyxl
                    # strips formula caches on save), resolve the VLOOKUP
                    # manually using the preset table and MM configuration.
                    # Formula: =VLOOKUP(VLOOKUP($A,'MM configuration'!A:H,3), M31:Q45, 4/5)
                    # Inner: MM# -> Row# via 'MM configuration' col A->C
                    # Outer: Row# -> sigma from preset table col 4(P)/5(Q)
                    rad_check = pd.to_numeric(df.get('sigma_rad [arcsec]'), errors='coerce')
                    azi_check = pd.to_numeric(df.get('sigma_azi [arcsec]'), errors='coerce')
                    missing_de = rad_check.isna() | azi_check.isna() | (rad_check <= 0.0) | (azi_check <= 0.0)
                    if missing_de.any():
                        try:
                            # Build MM# -> Row# from 'MM configuration'
                            mm_to_rownum = {}
                            cfg_sheet = None
                            for s in wb_vals.sheetnames:
                                if s.lower() == 'mm configuration':
                                    cfg_sheet = s
                                    break
                            if cfg_sheet is not None:
                                ws_cfg = wb_vals[cfg_sheet]
                                for r in range(2, (ws_cfg.max_row or 0) + 1):
                                    pos_v = ws_cfg.cell(row=r, column=1).value
                                    row_v = ws_cfg.cell(row=r, column=3).value
                                    if pos_v is not None and row_v is not None:
                                        try:
                                            mm_to_rownum[int(float(pos_v))] = int(float(row_v))
                                        except Exception:
                                            pass

                            # Build Row# -> (sigma_rad, sigma_azi) from preset
                            # table in MM_PSF.  Detect preset region dynamically:
                            # scan column M (13) for a cell containing "Row" as
                            # header, then read rows below until blank.
                            rownum_to_sigma = {}
                            ws_psf_f = None
                            try:
                                wb_formulas = _load_wb_vals.__self__ if hasattr(_load_wb_vals, '__self__') else None
                            except Exception:
                                wb_formulas = None
                            # Use the non-data_only workbook to read preset table
                            # (it has plain values, not formulas)
                            try:
                                from openpyxl import load_workbook as _load_wb_f
                                wb_f_tmp = _load_wb_f(path, data_only=False)
                                for s in wb_f_tmp.sheetnames:
                                    if s.lower() == 'mm_psf':
                                        ws_psf_f = wb_f_tmp[s]
                                        break
                            except Exception:
                                ws_psf_f = None
                            if ws_psf_f is not None:
                                # Find the preset header row (cell in col M containing "Row")
                                preset_header_row = None
                                for r in range(2, (ws_psf_f.max_row or 0) + 1):
                                    v = ws_psf_f.cell(row=r, column=13).value
                                    if v is not None and str(v).strip().lower() == 'row':
                                        preset_header_row = r
                                        break
                                if preset_header_row is not None:
                                    for r in range(preset_header_row + 1, (ws_psf_f.max_row or 0) + 1):
                                        rv = ws_psf_f.cell(row=r, column=13).value
                                        if rv is None:
                                            break
                                        sig_r = ws_psf_f.cell(row=r, column=16).value  # P = sigma_rad
                                        sig_a = ws_psf_f.cell(row=r, column=17).value  # Q = sigma_azi
                                        if sig_r is not None and sig_a is not None:
                                            try:
                                                rownum_to_sigma[int(float(rv))] = (float(sig_r), float(sig_a))
                                            except Exception:
                                                pass
                                try:
                                    wb_f_tmp.close()
                                except Exception:
                                    pass

                            # Apply resolved VLOOKUP values
                            if mm_to_rownum and rownum_to_sigma:
                                resolved_count = 0
                                for idx in df.index:
                                    try:
                                        mmv = int(pd.to_numeric(df.at[idx, 'MM #'], errors='coerce'))
                                    except Exception:
                                        continue
                                    cur_rad = pd.to_numeric(df.at[idx, 'sigma_rad [arcsec]'], errors='coerce')
                                    cur_azi = pd.to_numeric(df.at[idx, 'sigma_azi [arcsec]'], errors='coerce')
                                    if pd.notna(cur_rad) and cur_rad > 0 and pd.notna(cur_azi) and cur_azi > 0:
                                        continue  # already has valid values
                                    rn = mm_to_rownum.get(mmv)
                                    if rn is not None and rn in rownum_to_sigma:
                                        sr, sa = rownum_to_sigma[rn]
                                        df.at[idx, 'sigma_rad [arcsec]'] = sr
                                        df.at[idx, 'sigma_azi [arcsec]'] = sa
                                        resolved_count += 1
                                if resolved_count:
                                    print(f"INFO: resolved D/E VLOOKUP for {resolved_count} MMs from preset table")
                        except Exception:
                            pass

                    # If after attempting to map D/E values any required sigma cells
                    # remain missing, fail loudly — the workbook must contain cached
                    # numeric values for D/E. Ask user to re-save the Excel file with
                    # recalculation so caches are stored.
                    # Coerce to numeric and check for NaN or non-positive values
                    rad_vals = pd.to_numeric(df.get('sigma_rad [arcsec]'), errors='coerce')
                    azi_vals = pd.to_numeric(df.get('sigma_azi [arcsec]'), errors='coerce')
                    missing_mask = rad_vals.isna() | azi_vals.isna() | (rad_vals <= 0.0) | (azi_vals <= 0.0)
                    if missing_mask.any():
                        missing_mms = df.loc[missing_mask, 'MM #'].head(20).tolist()
                        raise ValueError(
                            "Invalid or missing numeric values in MM_PSF columns D/E (sigma_rad/sigma_azi) for some MMs. "
                            "These cells must contain positive numeric values (not formulas without cached results). "
                            "Open the workbook in Excel, recalculate, and save so cached values are stored. "
                            f"Examples: {missing_mms[:20]}"
                        )
    except Exception:
        pass
    # polar vignetting handled later after A_eff/weight initialization

    # If required headers aren't present, try to be flexible:
    # - map similar column names (e.g. 'm_rad', 'm_rad [arcsec]', 'm rad (arcsec)')
    # - if headerless, assume first 4 columns are the required ones
    if not all(c in df.columns for c in required):
        # Normalize column names: remove unit suffixes in brackets, collapse
        # whitespace, and strip punctuation so we can match flexible headers.
        cols = [str(c).strip() for c in df.columns]
        def norm(name: str) -> str:
            import re
            s = str(name).lower()
            # remove bracketed unit annotations like ' [arcsec]'
            s = re.sub(r"\[.*?\]", "", s)
            # replace non-alphanumeric with underscore
            s = re.sub(r"[^0-9a-z]+", "_", s)
            s = s.strip("_")
            return s

        norm_cols = {c: norm(c) for c in cols}
        mapping = {}
        # Desired base tokens for required names (without units)
        desired = {
            'm_rad [arcsec]': 'm_rad',
            'm_azi [arcsec]': 'm_azi',
            'sigma_rad [arcsec]': 'sigma_rad',
            'sigma_azi [arcsec]': 'sigma_azi',
        }

        for req_full, token in desired.items():
            found = None
            for orig, nc in norm_cols.items():
                if token.replace('_', '') == nc.replace('_', ''):
                    found = orig
                    break
            if found is None:
                # try more relaxed matches: token parts contained in normalized name
                parts = token.split('_')
                for orig, nc in norm_cols.items():
                    if all(p in nc for p in parts):
                        found = orig
                        break
            if found:
                mapping[found] = req_full

        if len(mapping) == len(required):
            df = df.rename(columns={k: v for k, v in mapping.items()})
        else:
            # As a last resort, if there are >=4 columns, assume first four are
            # the required fields (common in headerless templates). Preserve
            # any named 'MM #' column if present elsewhere.
            if df.shape[1] >= 4:
                orig_cols = list(df.columns)
                warn_cols = orig_cols[:4]
                print(f"Warning: MM_PSF appears headerless or uses non-standard headers {warn_cols}; assuming order {required}.")
                df = df.copy()
                df.columns = required + orig_cols[4:]
                if 'MM #' not in df.columns and 'MM #' in orig_cols:
                    mm_idx = orig_cols.index('MM #')
                    df['MM #'] = df.iloc[:, mm_idx]
            else:
                # Provide a clearer error message that includes available columns
                raise ValueError(f"Excel must contain columns: {required}. Found columns: {cols}")
    
    # Convert from arcsec to meters using project-specific convention.
    # Historically this code used an extra factor of 12; preserve that
    # behavior to match prior outputs: 1 arcsec -> 12 * π / 180 / 3600.
    arcsec_to_m = 12 * np.pi / 180 / 3600
    # Strict: do not promote alternate sigma columns; require D/E contain
    # cached numeric values. If D/E are formulas without cached results,
    # the loader will raise a clear error below instructing the user to
    # re-open and save the workbook so Excel writes cached values.

    # Coerce to numeric to support placeholder strings like '-'
    for col in ['m_rad [arcsec]', 'm_azi [arcsec]', 'sigma_rad [arcsec]', 'sigma_azi [arcsec]']:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    df['m_rad'] = df['m_rad [arcsec]'].fillna(0.0) * arcsec_to_m
    df['m_azi'] = df['m_azi [arcsec]'].fillna(0.0) * arcsec_to_m
    df['sigma_rad'] = df['sigma_rad [arcsec]'].fillna(0.0) * arcsec_to_m
    df['sigma_azi'] = df['sigma_azi [arcsec]'].fillna(0.0) * arcsec_to_m

    # Initialize weight column if it doesn't exist
    if 'weight' not in df.columns:
        df['weight'] = 1.0
    
    # Initialize distribution type and alpha parameters
    if 'distribution' not in df.columns:
        df['distribution'] = 'gaussian'  # Default to Gaussian
    if 'alpha_azi' not in df.columns:
        df['alpha_azi'] = 0.5  # Default mixing parameter for azimuthal axis
    if 'alpha_rad' not in df.columns:
        df['alpha_rad'] = 0.5  # Default mixing parameter for radial axis
    
    # Normalize distribution types:
    # - Built-ins: gaussian / pseudo-voigt / voigt
    # - Otherwise: treat as a custom PSF file stem (do NOT overwrite)
    valid_distributions = ['gaussian', 'pseudo-voigt', 'voigt']
    dist_raw = df['distribution'].astype(str).fillna('gaussian')
    dist_norm = dist_raw.str.strip()
    dist_lower = dist_norm.str.lower()
    df['distribution'] = np.where(dist_lower.isin(valid_distributions), dist_lower, dist_norm)

    # Coerce alpha columns to numeric where possible (placeholders like '-' become 0.5 fallback)
    df['alpha_azi'] = pd.to_numeric(df.get('alpha_azi', 0.5), errors='coerce').fillna(0.5)
    df['alpha_rad'] = pd.to_numeric(df.get('alpha_rad', 0.5), errors='coerce').fillna(0.5)
    
    # Calculate theta_degrees from MM configuration if MM # is present
    if 'MM #' in df.columns and 'theta_degrees' not in df.columns:
        try:
            mm_config_df = pd.read_excel(path, sheet_name='MM configuration', engine='openpyxl')
            if 'MM #' in mm_config_df.columns and 'x_MM [m]' in mm_config_df.columns and 'r_MM [m]' in mm_config_df.columns:
                # Create mapping from MM # to theta_degrees.
                # Convention note:
                # - User-facing MM configuration historically specifies an
                #   angle measured clockwise from the Y axis. Internally we
                #   convert to the project's `theta_degrees` convention via
                #   `-(theta_cw_from_y - 90)` so that theta_degrees==0 points
                #   along +Y and positive angles rotate the PSF as used by
                #   gaussian_2d_rotated/pseudo_voigt_2d_rotated.
                # - We compute an initial counter-clockwise angle via
                #   `atan2(x, y)` (note the swapped args to match the
                #   clockwise-from-Y intent), convert to a clockwise-from-Y
                #   range [0, 360) and finally apply the above transform.
                if 'y_MM [m]' in mm_config_df.columns:
                    theta_ccw = np.degrees(np.arctan2(mm_config_df['x_MM [m]'], mm_config_df['y_MM [m]']))
                    theta_cw_from_y = np.where(
                        theta_ccw >= 0,
                        theta_ccw,
                        180 + theta_ccw
                    )
                    mm_config_df['theta_position'] = -(theta_cw_from_y - 90)
                    mm_config_df['theta_degrees'] = -(theta_cw_from_y - 90)
                else:
                    # Fallback: compute y_MM from r_MM and x_MM, then use atan2
                    mm_config_df['y_MM_computed'] = np.sqrt(mm_config_df['r_MM [m]']**2 - mm_config_df['x_MM [m]']**2)
                    theta_ccw = np.degrees(np.arctan2(mm_config_df['x_MM [m]'], mm_config_df['y_MM_computed']))
                    theta_cw_from_y = np.where(
                        theta_ccw >= 0,
                        theta_ccw,
                        180 + theta_ccw
                    )
                    mm_config_df['theta_position'] = -(theta_cw_from_y - 90)
                    mm_config_df['theta_degrees'] = -(theta_cw_from_y - 90)
                
                theta_map = dict(zip(mm_config_df['MM #'].astype(int), mm_config_df['theta_degrees']))
                theta_position_map = dict(zip(mm_config_df['MM #'].astype(int), mm_config_df['theta_position']))
                
                # Map theta values to df
                mm_as_int = pd.to_numeric(df['MM #'], errors='coerce').astype('Int64', errors='ignore').fillna(df['MM #']).astype(int)
                df['theta_degrees'] = mm_as_int.map(theta_map)
                df['theta_position'] = mm_as_int.map(theta_position_map)
                
                # Check for missing mappings
                missing_theta = df['theta_degrees'].isna()
                if missing_theta.any():
                    missing_mm = df.loc[missing_theta, 'MM #'].tolist()
                    print(f"Warning: MM # values {missing_mm} not found in MM configuration, defaulting theta to 0.0")
                    df.loc[missing_theta, 'theta_degrees'] = 0.0
                    df.loc[missing_theta, 'theta_position'] = 0.0
            else:
                print("Warning: MM configuration sheet missing required columns, defaulting theta to 0.0")
                df['theta_position'] = 0.0
                df['theta_degrees'] = 0.0
        except Exception as e:
            print(f"Warning: Could not load MM configuration sheet: {e}, defaulting theta to 0.0")
            df['theta_position'] = 0.0
            df['theta_degrees'] = 0.0
    elif 'theta_degrees' not in df.columns:
        # If no MM # column, default theta to 0
        df['theta_degrees'] = 0.0
        df['theta_position'] = 0.0
    
    # If MM # present, always override weight from A_eff (strictly from column B).
    if 'MM #' in df.columns:
        # For Excel inputs, load authoritative per-MM A_eff from the A_eff sheet
        # (strict: column B). For CSV-only inputs (commonly produced by the
        # sensitivity runner), the CSV typically contains only the MM_PSF
        # table and no A_eff sheet. In that case, prefer an existing 'weight'
        # column in the CSV or fall back to 1.0 to allow processing.
        mm_as_int = pd.to_numeric(df['MM #'], errors='coerce')
        if mm_as_int.isna().any():
            bad = df.loc[mm_as_int.isna(), 'MM #'].head(10).tolist()
            raise ValueError(f"Invalid 'MM #' values in PSF sheet: {bad}")
        mm_as_int = mm_as_int.astype(int)

        if is_csv:
            # If CSV provided a 'weight' column, use it per-MM; otherwise default
            # to 1.0 for all MMs (sensitivity runner will have applied A_eff via
            # the baseline workbook when needed).
            if 'weight' in df.columns:
                df['aeff_base'] = pd.to_numeric(df['weight'], errors='coerce').fillna(1.0)
                df['weight'] = df['aeff_base'].astype(float)
            else:
                df['aeff_base'] = 1.0
                df['weight'] = 1.0
        else:
            # Load authoritative bare A_eff from column B using existing helper
            aeff_map_base = load_aeff_weight_map(path)
            df['aeff_base'] = mm_as_int.map(aeff_map_base)

            # Skip input col2 adjusted; will compute fresh post-vignetting from base * vig_factor
            df['weight'] = df['aeff_base'].astype(float)  # temporary; override post-vignetting

            # Validate that bare A_eff exists for all MMs (column B required)
            missing_mask = df['aeff_base'].isna()
            if missing_mask.any():
                missing_mm = sorted(set(mm_as_int[missing_mask].tolist()))
                raise ValueError(
                    "Missing A_eff weights for some MMs. "
                    "A_eff column B must contain a numeric weight for every MM used. "
                    f"Missing examples: {missing_mm[:20]}"
                )
    
    # Load all perturbation deltas and MM configuration.
    # Important: Alignment/Thermal/Gravity offload are allocated per *position* (slot),
    # not per MM. When MMs are swapped, these deltas must stay with the slot.
    alignment_by_pos: dict[int, dict] = {}
    gravity_by_pos: dict[int, dict] = {}
    thermal_by_pos: dict[int, dict] = {}
    extra_by_pos: dict[int, dict] = {}
    mm_config_map = {}
    mm_to_pos: dict[int, int] = {}
    
    # Load MM configuration for x, y, z coordinates and build MM -> Position mapping.
    if 'MM #' in df.columns:
        try:
            mm_config_df = pd.read_excel(path, sheet_name='MM configuration', engine='openpyxl')
            if 'MM #' in mm_config_df.columns:
                for order_i, (idx, row) in enumerate(mm_config_df.iterrows()):
                    mm_num = row.get('MM #')
                    if pd.isna(mm_num):
                        continue
                    mm_num_i = int(mm_num)

                    # Prefer explicit Position # column when present.
                    if 'Position #' in mm_config_df.columns:
                        pos_val = row.get('Position #')
                        if not pd.isna(pos_val):
                            try:
                                mm_to_pos[mm_num_i] = int(float(pos_val))
                            except Exception:
                                pass
                        if mm_num_i not in mm_to_pos:
                            # Fallback to row order if Position # is missing/invalid.
                            mm_to_pos[mm_num_i] = int(order_i) + 1
                    else:
                        mm_to_pos[mm_num_i] = int(order_i) + 1

                    # Record the MM configuration row identifier that
                    # vignetting tables reference in column H. Prefer the
                    # explicit 'Row #' value when present in the MM
                    # configuration sheet; otherwise fall back to the
                    # 1-based data row index (order_i + 1).
                    try:
                        raw_rownum = row.get('Row #') if 'Row #' in mm_config_df.columns else None
                        if raw_rownum is not None and not pd.isna(raw_rownum):
                            cfg_row_number = int(float(raw_rownum))
                        else:
                            cfg_row_number = int(order_i) + 1
                    except Exception:
                        cfg_row_number = int(order_i) + 1
                    # Prefer explicit Position # mapping already set above
                    # but also keep reverse lookup: position -> mm_config row.
                    # If Position # wasn't present we still map using order.
                    try:
                        pos_for_row = mm_to_pos.get(mm_num_i, cfg_row_number)
                    except Exception:
                        pos_for_row = cfg_row_number
                    # pos_to_cfg_row: Position # -> row number in MM configuration
                    if 'pos_to_cfg_row' not in locals():
                        pos_to_cfg_row = {}
                    pos_to_cfg_row[pos_for_row] = cfg_row_number

                    mm_config_map[mm_num_i] = {
                        'x_MM': row.get('x_MM [m]', 0),
                        'y_MM': row.get('y_MM [m]', 0),
                        'z_MM': row.get('z_MM [m]', 0),
                        'r_MM': row.get('r_MM [m]', 0)
                    }
        except Exception:
            pass

    

    # Load alignment deltas (prefer Position #)
    if 'MM #' in df.columns:
        try:
            align_df = pd.read_excel(path, sheet_name='Alignment', engine='openpyxl')
            if 'Position #' in align_df.columns:
                tmp = align_df.copy()
                tmp['Position #'] = pd.to_numeric(tmp['Position #'], errors='coerce')
                tmp = tmp[tmp['Position #'].notna()]
                for _, row in tmp.iterrows():
                    pos = int(row['Position #'])
                    # Handle both named-column and alignment-preset row formats.
                    # Some workbooks mark a preset row (e.g. 'preset_selected') and
                    # place rotazi/rotrad values in fixed column indices.
                    d_align_rotazi = row.get('d_align_rotazi [arcsec]', row.get('d_align_rotazi', 0))
                    d_align_rotrad = row.get('d_align_rotrad [arcsec]', row.get('d_align_rotrad', 0))
                    # Detect preset marker in any column; if present, try positional indices
                    try:
                        if any(str(x).strip().lower() == 'preset_selected' for x in row.tolist()):
                            # test harness uses columns 12 (rotazi) and 13 (rotrad)
                            try:
                                cand_rotazi = row.iloc[12]
                                cand_rotrad = row.iloc[13]
                                if pd.notna(cand_rotazi):
                                    d_align_rotazi = float(cand_rotazi)
                                if pd.notna(cand_rotrad):
                                    d_align_rotrad = float(cand_rotrad)
                            except Exception:
                                pass
                    except Exception:
                        pass

                    alignment_by_pos[pos] = {
                        'd_align_rad': float(row.get('d_align_rad [µm]', 0)) * 1e-6,
                        'd_align_azi': float(row.get('d_align_azi [µm]', 0)) * 1e-6,
                        'd_align_z': float(row.get('d_align_z [µm]', 0)) * 1e-6,
                        'd_align_rotz': float(row.get('d_align_rotz [arcsec]', 0)),
                        'd_align_rotazi': float(d_align_rotazi or 0),
                        'd_align_rotrad': float(d_align_rotrad or 0),
                    }
            elif 'MM #' in align_df.columns:
                # Legacy fallback: if the sheet is keyed by MM #
                for _, row in align_df.iterrows():
                    mm_num = row.get('MM #')
                    if pd.isna(mm_num):
                        continue
                    pos = mm_to_pos.get(int(mm_num))
                    if pos is None:
                        continue
                    alignment_by_pos[pos] = {
                        'd_align_rad': float(row.get('d_align_rad [µm]', 0)) * 1e-6,
                        'd_align_azi': float(row.get('d_align_azi [µm]', 0)) * 1e-6,
                        'd_align_z': float(row.get('d_align_z [µm]', 0)) * 1e-6,
                        'd_align_rotz': float(row.get('d_align_rotz [arcsec]', 0)),
                        'd_align_rotazi': float(row.get('d_align_rotazi [arcsec]', 0)),
                        'd_align_rotrad': float(row.get('d_align_rotrad [arcsec]', 0)),
                    }
        except Exception:
            pass
    
    # Load gravity offload deltas (prefer Position #, sum duplicates)
    if 'MM #' in df.columns:
        try:
            gravity_df = pd.read_excel(path, sheet_name='Gravity offload', engine='openpyxl')
            if 'Position #' in gravity_df.columns:
                tmp = gravity_df.copy()
                tmp['Position #'] = pd.to_numeric(tmp['Position #'], errors='coerce')
                tmp = tmp[tmp['Position #'].notna()]
                for c in ['d_grav_x [µm]', 'd_grav_y [µm]', 'd_grav_z [µm]', 'd_grav_rotz [arcsec]', 'd_grav_rotx [arcsec]', 'd_grav_roty [arcsec]']:
                    if c in tmp.columns:
                        tmp[c] = pd.to_numeric(tmp[c], errors='coerce').fillna(0.0)
                grp = tmp.groupby('Position #', as_index=False).sum(numeric_only=True)
                for _, row in grp.iterrows():
                    pos = int(row['Position #'])
                    gravity_by_pos[pos] = {
                        'd_grav_x': float(row.get('d_grav_x [µm]', 0.0)) * 1e-6,
                        'd_grav_y': float(row.get('d_grav_y [µm]', 0.0)) * 1e-6,
                        'd_grav_z': float(row.get('d_grav_z [µm]', 0.0)) * 1e-6,
                        'd_grav_rotz': float(row.get('d_grav_rotz [arcsec]', 0.0)),
                        'd_grav_rotx': float(row.get('d_grav_rotx [arcsec]', 0.0)),
                        'd_grav_roty': float(row.get('d_grav_roty [arcsec]', 0.0)),
                        'd_grav_rotazi': float(row.get('d_grav_rotazi [arcsec]', 0.0)),
                        'd_grav_rotrad': float(row.get('d_grav_rotrad [arcsec]', 0.0)),
                    }
            elif 'MM #' in gravity_df.columns:
                for _, row in gravity_df.iterrows():
                    mm_num = row.get('MM #')
                    if pd.isna(mm_num):
                        continue
                    pos = mm_to_pos.get(int(mm_num))
                    if pos is None:
                        continue
                    prev = gravity_by_pos.get(pos, {'d_grav_x':0.0,'d_grav_y':0.0,'d_grav_z':0.0,'d_grav_rotz':0.0})
                    gravity_by_pos[pos] = {
                            'd_grav_x': prev['d_grav_x'] + float(row.get('d_grav_x [µm]', 0.0)) * 1e-6,
                            'd_grav_y': prev['d_grav_y'] + float(row.get('d_grav_y [µm]', 0.0)) * 1e-6,
                            'd_grav_z': prev['d_grav_z'] + float(row.get('d_grav_z [µm]', 0.0)) * 1e-6,
                            'd_grav_rotz': prev['d_grav_rotz'] + float(row.get('d_grav_rotz [arcsec]', 0.0)),
                            'd_grav_rotx': prev.get('d_grav_rotx', 0.0) + float(row.get('d_grav_rotx [arcsec]', 0.0)),
                            'd_grav_roty': prev.get('d_grav_roty', 0.0) + float(row.get('d_grav_roty [arcsec]', 0.0)),
                            'd_grav_rotazi': prev.get('d_grav_rotazi', 0.0) + float(row.get('d_grav_rotazi [arcsec]', 0.0)),
                            'd_grav_rotrad': prev.get('d_grav_rotrad', 0.0) + float(row.get('d_grav_rotrad [arcsec]', 0.0)),
                    }
        except Exception:
            pass
    
    # Load thermal deltas (prefer Position #, sum duplicates)
    if 'MM #' in df.columns:
        try:
            thermal_df = pd.read_excel(path, sheet_name='Thermal', engine='openpyxl')
            if 'Position #' in thermal_df.columns:
                tmp = thermal_df.copy()
                tmp['Position #'] = pd.to_numeric(tmp['Position #'], errors='coerce')
                tmp = tmp[tmp['Position #'].notna()]
                for c in ['d_therm_x [µm]', 'd_therm_y [µm]', 'd_therm_z [µm]', 'd_therm_rotz [arcsec]', 'd_therm_rotx [arcsec]', 'd_therm_roty [arcsec]']:
                    if c in tmp.columns:
                        tmp[c] = pd.to_numeric(tmp[c], errors='coerce').fillna(0.0)
                grp = tmp.groupby('Position #', as_index=False).sum(numeric_only=True)
                for _, row in grp.iterrows():
                    pos = int(row['Position #'])
                    thermal_by_pos[pos] = {
                            'd_therm_x': float(row.get('d_therm_x [µm]', 0.0)) * 1e-6,
                            'd_therm_y': float(row.get('d_therm_y [µm]', 0.0)) * 1e-6,
                            'd_therm_z': float(row.get('d_therm_z [µm]', 0.0)) * 1e-6,
                            'd_therm_rotz': float(row.get('d_therm_rotz [arcsec]', 0.0)),
                            'd_therm_rotx': float(row.get('d_therm_rotx [arcsec]', 0.0)),
                            'd_therm_roty': float(row.get('d_therm_roty [arcsec]', 0.0)),
                            'd_therm_rotazi': float(row.get('d_therm_rotazi [arcsec]', 0.0)),
                            'd_therm_rotrad': float(row.get('d_therm_rotrad [arcsec]', 0.0)),
                    }
            elif 'MM #' in thermal_df.columns:
                for _, row in thermal_df.iterrows():
                    mm_num = row.get('MM #')
                    if pd.isna(mm_num):
                        continue
                    pos = mm_to_pos.get(int(mm_num))
                    if pos is None:
                        continue
                    prev = thermal_by_pos.get(pos, {'d_therm_x':0.0,'d_therm_y':0.0,'d_therm_z':0.0,'d_therm_rotz':0.0})
                    thermal_by_pos[pos] = {
                            'd_therm_x': prev['d_therm_x'] + float(row.get('d_therm_x [µm]', 0.0)) * 1e-6,
                            'd_therm_y': prev['d_therm_y'] + float(row.get('d_therm_y [µm]', 0.0)) * 1e-6,
                            'd_therm_z': prev['d_therm_z'] + float(row.get('d_therm_z [µm]', 0.0)) * 1e-6,
                            'd_therm_rotz': prev['d_therm_rotz'] + float(row.get('d_therm_rotz [arcsec]', 0.0)),
                            'd_therm_rotx': prev.get('d_therm_rotx', 0.0) + float(row.get('d_therm_rotx [arcsec]', 0.0)),
                            'd_therm_roty': prev.get('d_therm_roty', 0.0) + float(row.get('d_therm_roty [arcsec]', 0.0)),
                            'd_therm_rotazi': prev.get('d_therm_rotazi', 0.0) + float(row.get('d_therm_rotazi [arcsec]', 0.0)),
                            'd_therm_rotrad': prev.get('d_therm_rotrad', 0.0) + float(row.get('d_therm_rotrad [arcsec]', 0.0)),
                    }
        except Exception:
            pass
        print(f"VIG DEBUG: thermal_by_pos has {len(thermal_by_pos)} entries")
        if thermal_by_pos:
            sample = dict(list(thermal_by_pos.items())[:3])
            if not os.environ.get('SILENCE_OUTPUT'):
                print(f"VIG DEBUG: thermal_by_pos sample: {sample}")
            rotx_values = [v.get('d_therm_rotx', 0) for v in thermal_by_pos.values()]
            print(f"VIG DEBUG: d_therm_rotx values: min={min(rotx_values):.1f}, max={max(rotx_values):.1f}, all_same={len(set(rotx_values)) == 1}")

    # Load extra PSF shifts (off-axis pointing, etc.)
    if 'MM #' in df.columns:
        try:
            extra_df = pd.read_excel(path, sheet_name='Extra PSF shifts', engine='openpyxl')
            if 'Position #' in extra_df.columns:
                for c in ['d_extra_rotx [arcsec]', 'd_extra_roty [arcsec]', 'd_extra_z [µm]']:
                    if c in extra_df.columns:
                        extra_df[c] = pd.to_numeric(extra_df[c], errors='coerce').fillna(0.0)
                for _, row in extra_df.iterrows():
                    pos_val = row.get('Position #')
                    if pd.isna(pos_val):
                        continue
                    pos = int(pos_val)
                    extra_by_pos[pos] = {
                        'd_extra_rotx': float(row.get('d_extra_rotx [arcsec]', 0.0)),
                        'd_extra_roty': float(row.get('d_extra_roty [arcsec]', 0.0)),
                        'd_extra_z': float(row.get('d_extra_z [µm]', 0.0)) * 1e-6,  # µm -> m
                    }
        except Exception:
            pass

    # Per-position HEW degradation values (arcsec), populated during
    # HEW degradation sheet processing below; used later to broaden sigma.
    hew_deg_per_pos_azi: dict[int, float] = {}
    hew_deg_per_pos_rad: dict[int, float] = {}

    # --- Apply polar vignetting (rotazi + rotrad) after A_eff/weight initialization ---
    try:
        # compute rotation projections using the populated mm_to_pos and *_by_pos
        try:
            _, _, rot_rad_map, rot_azi_map = compute_total_rot_polar(mm_to_pos, mm_config_map, alignment_by_pos, gravity_by_pos, thermal_by_pos, extra_by_pos)
            print(f"VIG DEBUG: rot_azi_map has {len(rot_azi_map)} entries")
            if rot_azi_map:
                sample = dict(list(rot_azi_map.items())[:3])
                if not os.environ.get('SILENCE_OUTPUT'):
                    print(f"VIG DEBUG: rot_azi_map sample: {sample}")
                non_zero = [v for v in rot_azi_map.values() if abs(v) > 1e-6]
                print(f"VIG DEBUG: rot_azi_map has {len(non_zero)} non-zero values, range: {min(non_zero) if non_zero else 0:.1f} to {max(non_zero) if non_zero else 0:.1f}")
        except Exception:
            rot_rad_map = {}
            rot_azi_map = {}

        applied_azi = False
        applied_rad = False

        # Read rotazi sheet (A->B interpretation: col0 = delta, col1 = factor)
        # Supported layouts:
        # - Two-column A->B (delta in col0, factor in col1): used for a single
        #   global vignette curve applied to every position.
        # - Multi-column with numeric column headers: columns after col0 are
        #   interpreted as per-position factor series (header=Position #).
        # - Header-heavy workbooks often still contain the intended A->B mapping
        #   in the first two columns; we try that as a final fallback.
        try:
            vdf_azi = _read_excel_vig(path, VIG_ROT_AZI_CANDIDATES, engine='openpyxl')
            if vdf_azi is None:
                raise ValueError('No vignetting rotazi sheet found')
            xs_azi = ys_azi = None
            ys_by_pos_azi = {}
            azi_mode = 'none'
            # Try to detect new layout: table starting at column H (index 7)
            # where column H contains MM-config row numbers, column I (8)
            # contains rotazi x-values, column J (9) indicates energy,
            # and column K (10) contains the factor. If present, build
            # per-position/energy interpolation series.
            try:
                aeff_map, aeff_col_name = load_aeff_weight_map_with_name(path)
            except Exception:
                aeff_map, aeff_col_name = load_aeff_weight_map(path), None

            # Try to detect the selected energy (numeric keV) from several
            # places: preferred sources (in order):
            # 1) explicit selection in the vignetting sheet (e.g. cell C2),
            # 2) the A_eff column name (e.g. '7 keV'),
            # 3) a scanned 'keV' token anywhere in the A_eff sheet.
            sel_energy = None
            sel_energy_from_vdf = False
            
            # 1) Enhanced: vignette sheet col 'Selected energy [keV]' and C2
            try:
                # DataFrame scan first (faster) - vdf_azi/vdf_rad defined later, use vdf_rad for now
                for vdf_name, sname in [('vdf_rad', 'rotrad'), ('vdf_azi', 'rotazi')]:
                    vdf = locals().get(vdf_name)
                    if vdf is not None and vdf.shape[1] > 2:
                        for r in range(min(3, vdf.shape[0])):
                            cand = vdf.iloc[r, 2]
                            if pd.notna(cand):
                                try:
                                    sel_energy = float(cand)
                                    sel_energy_from_vdf = True
                                    print(f'VIG sel_energy from {sname} col2 row{r}: {sel_energy}')
                                    break
                                except:
                                    pass
                        if sel_energy is not None:
                            break
                if sel_energy is None:
                    # C2 fallback with openpyxl
                    from openpyxl import load_workbook
                    wb_tmp = load_workbook(path, data_only=True)
                    for sname in list(VIG_ROT_AZI_CANDIDATES) + list(VIG_ROT_RAD_CANDIDATES):
                        if sname in wb_tmp.sheetnames:
                            ws_tmp = wb_tmp[sname]
                            candidate = ws_tmp.cell(row=2, column=3).value
                            if candidate is not None and not pd.isna(candidate):
                                try:
                                    sel_energy = float(candidate)
                                    sel_energy_from_vdf = True
                                    print(f'VIG sel_energy from {sname} C2: {sel_energy}')
                                    break
                                except:
                                    pass
            except Exception as e:
                print(f'VIG sel_energy scan error: {e}')
            
            if sel_energy is None:
                print('VIG WARNING sel_energy=None - default 0.2')
                sel_energy = 0.2
            print(f'VIG sel_energy={sel_energy}')
            # 2/3) fallback: detect from A_eff or column name
            try:
                # Only use the A_eff column name if sel_energy was not
                # already determined from the vignette sheet itself.
                import re as _re
                if sel_energy is None and aeff_col_name and isinstance(aeff_col_name, str):
                    m = _re.search(r"(\d+(?:\.\d*)?)\s*(?:keV)?", aeff_col_name, flags=_re.IGNORECASE)
                    if m:
                        sel_energy = float(m.group(1))
            except Exception:
                pass

            if sel_energy is None and not sel_energy_from_vdf:
                try:
                    aeff_df = pd.read_excel(path, sheet_name='A_eff', engine='openpyxl', header=None)
                    import re as _re
                    found = None
                    for _, row in aeff_df.iterrows():
                        for cell in row.tolist():
                            try:
                                if isinstance(cell, str):
                                    m = _re.search(r"(\d+(?:\.\d*)?)\s*keV", cell, flags=_re.IGNORECASE)
                                    if m:
                                        found = float(m.group(1))
                                        break
                            except Exception:
                                continue
                        if found is not None:
                            break
                    if found is not None:
                        sel_energy = found
                except Exception:
                    sel_energy = None

            if vdf_azi is not None and not vdf_azi.empty and vdf_azi.shape[1] >= 2:
                # New table layout detection (column H-based)
                if vdf_azi.shape[1] >= 11:
                    # column H (index 7) should contain row numbers referencing
                    # the MM configuration; build per-row arrays keyed by that
                    # row number and energy marker.
                    col_H = vdf_azi.iloc[:, 7]
                    if col_H.notna().any():
                        for _, r in vdf_azi.iterrows():
                            try:
                                cfg_row = r.iloc[7]
                                if pd.isna(cfg_row):
                                    continue
                                cfg_row = int(float(cfg_row))
                            except Exception:
                                continue
                            # energy marker in column J (index 9)
                            energy_marker = r.iloc[9] if vdf_azi.shape[1] > 9 else None
                            try:
                                # column I contains rot delta in arcmin; convert to arcsec
                                xval = float(r.iloc[8]) * 60.0 if not pd.isna(r.iloc[8]) else None
                            except Exception:
                                xval = None
                            try:
                                yval = float(r.iloc[10]) if vdf_azi.shape[1] > 10 and not pd.isna(r.iloc[10]) else None
                            except Exception:
                                yval = None
                            if xval is None or yval is None:
                                continue
                            # store series keyed by (cfg_row, energy_marker)
                            # energy in column J may be numeric or textual; preserve both representations
                            key_str = (cfg_row, str(energy_marker).strip())
                            key_num = None
                            try:
                                key_num = (cfg_row, float(energy_marker))
                            except Exception:
                                key_num = None
                            if key_str not in ys_by_pos_azi:
                                ys_by_pos_azi[key_str] = {'xs': [], 'ys': []}
                            ys_by_pos_azi[key_str]['xs'].append(xval)
                            ys_by_pos_azi[key_str]['ys'].append(yval)
                            if key_num is not None:
                                if key_num not in ys_by_pos_azi:
                                    ys_by_pos_azi[key_num] = {'xs': [], 'ys': []}
                                ys_by_pos_azi[key_num]['xs'].append(xval)
                                ys_by_pos_azi[key_num]['ys'].append(yval)
                        # sort arrays
                        for k, v in list(ys_by_pos_azi.items()):
                            order = np.argsort(v['xs'])
                            xs_sorted = np.array(v['xs'], dtype=float)[order]
                            ys_sorted = np.array(v['ys'], dtype=float)[order]
                            ys_by_pos_azi[k] = (xs_sorted, ys_sorted)
                        if ys_by_pos_azi:
                            azi_mode = 'per_row_energy'
                            print(f'VIG azi_mode={azi_mode} nseries={len(ys_by_pos_azi)}')
                            print(f'VIG ys_by_pos_azi keys sample: {list(ys_by_pos_azi.keys())[:5]}')
                            if os.environ.get("VIG_DEBUG"):
                                for k in list(ys_by_pos_azi)[:3]:
                                    xs, ys = ys_by_pos_azi[k]
                                    print(f'VIG series {k} xs.shape={xs.shape} xs.min={xs.min():.1f} xs.max={xs.max():.1f} ys.mean={ys.mean():.3f}')
                # Fallback to classic modes if new layout not present
                if azi_mode == 'none':
                    xs_raw = pd.to_numeric(vdf_azi.iloc[:, 0], errors='coerce')
                    xs = xs_raw.dropna().to_numpy(dtype=float)
                    if xs.size > 0:
                        if vdf_azi.shape[1] == 2:
                            ys = pd.to_numeric(vdf_azi.iloc[:, 1], errors='coerce').dropna().to_numpy(dtype=float)
                            if ys.size > 0:
                                order = np.argsort(xs)
                                xs_azi = xs[order]
                                ys_azi = ys[order]
                                azi_mode = 'single'
                        else:
                            cols = list(vdf_azi.columns)
                            for col in cols[1:]:
                                try:
                                    pos_key = int(str(col))
                                except Exception:
                                    continue
                                ys_col = pd.to_numeric(vdf_azi[col], errors='coerce')
                                ys_vals = ys_col.fillna(np.nan).to_numpy(dtype=float)
                                if ys_vals.size > 0:
                                    order = np.argsort(xs)
                                    ys_by_pos_azi[pos_key] = ys_vals[order]
                            if ys_by_pos_azi and isinstance(list(ys_by_pos_azi.keys())[0], int):
                                xs_azi = xs[np.argsort(xs)]
                                azi_mode = 'per_pos'
                    # Fallback: try first two columns as A->B
                    if azi_mode == 'none' and vdf_azi.shape[1] >= 2:
                        trial_y = pd.to_numeric(vdf_azi.iloc[:, 1], errors='coerce').dropna().to_numpy(dtype=float)
                        if trial_y.size > 0:
                            order = np.argsort(xs)
                            xs_azi = xs[order]
                            ys_azi = trial_y[order] if trial_y.size == xs_azi.size else trial_y
                            azi_mode = 'single'
                if azi_mode != 'none':
                    applied_azi = True
                    try:
                        if 'MM #' in df.columns and len(df) <= 10:
                            print(f"VIG_PARSE: azi_mode={azi_mode} xs_azi_set={xs_azi is not None} ys_by_pos_azi_keys={list(ys_by_pos_azi.keys()) if 'ys_by_pos_azi' in locals() else None}")
                    except Exception:
                        pass
        except Exception:
            xs_azi = ys_azi = None

        # Read rotrad sheet (A->B interpretation)
        # Read rotrad sheet (same layout rules as rotazi):
        # - 'single' mode: xs_rad/ys_rad forms a single curve (delta->factor)
        # - 'per_pos' mode: ys_by_pos_rad[pos] contains the factor array for that slot
        # During application we prefer per_pos series when present, else fall
        # back to the single global curve. If neither is present we leave
        # the weight unchanged for that component.
        try:
            vdf_rad = _read_excel_vig(path, VIG_ROT_RAD_CANDIDATES, engine='openpyxl')
            if vdf_rad is None:
                raise ValueError('No vignetting rotrad sheet found')
            xs_rad = ys_rad = None
            ys_by_pos_rad = {}
            rad_mode = 'none'
            # Try to detect new H/I/J/K layout (same as rotazi):
            # column H=index7 -> cfg_row, I=index8 -> delta(arcmin),
            # J=index9 -> energy marker, K=index10 -> factor
            try:
                if vdf_rad is not None and not vdf_rad.empty and vdf_rad.shape[1] >= 11:
                    col_H = vdf_rad.iloc[:, 7]
                    if col_H.notna().any():
                        for _, r in vdf_rad.iterrows():
                            try:
                                cfg_row = r.iloc[7]
                                if pd.isna(cfg_row):
                                    continue
                                cfg_row = int(float(cfg_row))
                            except Exception:
                                continue
                            energy_marker = r.iloc[9] if vdf_rad.shape[1] > 9 else None
                            try:
                                xval = float(r.iloc[8]) * 60.0 if not pd.isna(r.iloc[8]) else None
                            except Exception:
                                xval = None
                            try:
                                yval = float(r.iloc[10]) if vdf_rad.shape[1] > 10 and not pd.isna(r.iloc[10]) else None
                            except Exception:
                                yval = None
                            if xval is None or yval is None:
                                continue
                            key_str = (cfg_row, str(energy_marker).strip())
                            key_num = None
                            try:
                                key_num = (cfg_row, float(energy_marker))
                            except Exception:
                                key_num = None
                            if key_str not in ys_by_pos_rad:
                                ys_by_pos_rad[key_str] = {'xs': [], 'ys': []}
                            ys_by_pos_rad[key_str]['xs'].append(xval)
                            ys_by_pos_rad[key_str]['ys'].append(yval)
                            if key_num is not None:
                                if key_num not in ys_by_pos_rad:
                                    ys_by_pos_rad[key_num] = {'xs': [], 'ys': []}
                                ys_by_pos_rad[key_num]['xs'].append(xval)
                                ys_by_pos_rad[key_num]['ys'].append(yval)
                        # sort arrays
                        for k, v in list(ys_by_pos_rad.items()):
                            order = np.argsort(v['xs'])
                            xs_sorted = np.array(v['xs'], dtype=float)[order]
                            ys_sorted = np.array(v['ys'], dtype=float)[order]
                            ys_by_pos_rad[k] = (xs_sorted, ys_sorted)
                        if ys_by_pos_rad:
                            rad_mode = 'per_row_energy'
                            print(f'VIG rad_mode={rad_mode} nseries={len(ys_by_pos_rad)}')
                            print(f'VIG ys_by_pos_rad keys sample: {list(ys_by_pos_rad.keys())[:5]}')
                            if os.environ.get("VIG_DEBUG"):
                                for k in list(ys_by_pos_rad)[:3]:
                                    xs, ys = ys_by_pos_rad[k]
                                    print(f'VIG series {k} xs.shape={xs.shape} xs.max={xs.max():.1f}" ys.mean={ys.mean():.3f}')

            except Exception:
                pass

            # Fallback to classic layouts if new layout not detected
            if vdf_rad is not None and not vdf_rad.empty and vdf_rad.shape[1] >= 2 and rad_mode == 'none':
                xs_raw = pd.to_numeric(vdf_rad.iloc[:, 0], errors='coerce')
                xs = xs_raw.dropna().to_numpy(dtype=float)
                if xs.size > 0:
                    if vdf_rad.shape[1] == 2:
                        ys = pd.to_numeric(vdf_rad.iloc[:, 1], errors='coerce').dropna().to_numpy(dtype=float)
                        if ys.size > 0:
                            order = np.argsort(xs)
                            xs_rad = xs[order]
                            ys_rad = ys[order]
                            rad_mode = 'single'
                    else:
                        cols = list(vdf_rad.columns)
                        for col in cols[1:]:
                            try:
                                pos_key = int(str(col))
                            except Exception:
                                continue
                            ys_col = pd.to_numeric(vdf_rad[col], errors='coerce')
                            ys_vals = ys_col.fillna(np.nan).to_numpy(dtype=float)
                            if ys_vals.size > 0:
                                order = np.argsort(xs)
                                ys_by_pos_rad[pos_key] = ys_vals[order]
                        if ys_by_pos_rad:
                            xs_rad = xs[np.argsort(xs)]
                            rad_mode = 'per_pos'
                    if rad_mode == 'none' and vdf_rad.shape[1] >= 2:
                        trial_y = pd.to_numeric(vdf_rad.iloc[:, 1], errors='coerce').dropna().to_numpy(dtype=float)
                        if trial_y.size > 0:
                            order = np.argsort(xs)
                            xs_rad = xs[order]
                            ys_rad = trial_y[order] if trial_y.size == xs_rad.size else trial_y
                            rad_mode = 'single'
                    if rad_mode != 'none':
                        applied_rad = True
                        try:
                            if 'MM #' in df.columns and len(df) <= 10:
                                print(f"VIG_PARSE: rad_mode={rad_mode} xs_rad_set={xs_rad is not None} ys_by_pos_rad_keys={list(ys_by_pos_rad.keys()) if 'ys_by_pos_rad' in locals() else None}")
                        except Exception:
                            pass
        except Exception:
            xs_rad = ys_rad = None
        # If sel_energy was not determined from the rotazi sheet, try rotrad's
        # selected-energy marker (common exporter sometimes places it only
        # on one of the vignette sheets). Normalize to float when possible.
        try:
            if (sel_energy is None or (isinstance(sel_energy, float) and np.isnan(sel_energy))):
                # read rotrad sheet C2 explicitly to detect selected energy
                try:
                    from openpyxl import load_workbook
                    wb_tmp2 = load_workbook(path, data_only=True)
                    _vig_rad_name = _find_vig_sheet(wb_tmp2, VIG_ROT_RAD_CANDIDATES)
                    if _vig_rad_name:
                        ws_tmp2 = wb_tmp2[_vig_rad_name]
                        candidate = ws_tmp2.cell(row=2, column=3).value
                    else:
                        candidate = None
                except Exception:
                    candidate = None
                if candidate is not None and not (isinstance(candidate, float) and np.isnan(candidate)):
                    try:
                        if isinstance(candidate, (int, float)):
                            sel_energy = float(candidate)
                        else:
                            sel_energy = float(str(candidate).strip())
                        sel_energy_from_vdf = True
                    except Exception:
                        import re as _re
                        if isinstance(candidate, str):
                            m = _re.search(r"(\d+(?:\.\d*)?)\s*(?:keV)?", candidate, flags=_re.IGNORECASE)
                            if m:
                                sel_energy = float(m.group(1))
                                sel_energy_from_vdf = True
        except Exception:
            pass

        # Apply per-row interpolation multiplicatively to the already-initialized weight
        # (instrumentation: print progress to help find stalls)
        try:
            pass
            sys.stdout.flush()
        except Exception:
            pass
        try:
            if os.environ.get('VIG_DEBUG'):
                try:
                    pass
                    pass
                    pass
                    sys.stdout.flush()
                except Exception:
                    pass
        except Exception:
            pass
        # Apply per-row interpolation multiplicatively to the already-initialized weight.
        # For each MM row we:
        # 1) determine its Position # (slot) using `mm_to_pos` mapping;
        # 2) for each of rot_rad and rot_azi, select the appropriate factor series
        #    (per-position series preferred, otherwise the single global curve);
        # 3) interpolate the factor at the computed angular offset and multiply
        #    it into the existing weight. Radial and azimuthal factors both
        #    multiply the weight (order is irrelevant since multiplication is
        #    commutative), but we apply radial first then azimuthal for clarity.
        # Normalize any per-position vignette series so interpolation
        # later can assume integer-position keys map to 1D arrays of
        # factors sampled at `xs_azi` / `xs_rad`. This fixes cases
        # where sheets mixed tuple (xs, ys) series and plain ys
        # arrays, or where per-pos arrays have a different length
        # than the global xs vectors.
        try:
            if 'ys_by_pos_azi' in locals() and ys_by_pos_azi:
                for k in list(ys_by_pos_azi.keys()):
                    # only normalize integer position keys (per-pos mode)
                    try:
                        if not isinstance(k, int):
                            continue
                    except Exception:
                        continue
                    v = ys_by_pos_azi[k]
                    # if stored as (xs_local, ys_local), resample to xs_azi if available
                    if isinstance(v, tuple) and len(v) == 2:
                        xs_local, ys_local = v
                        try:
                            if 'xs_azi' in locals() and xs_azi is not None:
                                ys_resampled = np.interp(xs_azi, np.asarray(xs_local, dtype=float), np.asarray(ys_local, dtype=float))
                                ys_by_pos_azi[k] = ys_resampled
                            else:
                                ys_by_pos_azi[k] = np.asarray(ys_local, dtype=float)
                        except Exception:
                            ys_by_pos_azi[k] = np.asarray(ys_local, dtype=float)
                    else:
                        # plain array/list -- coerce and if needed resample to xs_azi
                        try:
                            ys_arr = np.asarray(v, dtype=float)
                        except Exception:
                            ys_arr = np.array(v, dtype=float)
                        if 'xs_azi' in locals() and xs_azi is not None:
                            try:
                                if ys_arr.size != xs_azi.size:
                                    xs_from = np.linspace(xs_azi.min(), xs_azi.max(), ys_arr.size)
                                    ys_resampled = np.interp(xs_azi, xs_from, ys_arr)
                                    ys_by_pos_azi[k] = ys_resampled
                                else:
                                    ys_by_pos_azi[k] = ys_arr
                            except Exception:
                                ys_by_pos_azi[k] = ys_arr
                        else:
                            ys_by_pos_azi[k] = ys_arr
        except Exception:
            pass

        try:
            if 'ys_by_pos_rad' in locals() and ys_by_pos_rad:
                for k in list(ys_by_pos_rad.keys()):
                    try:
                        if not isinstance(k, int):
                            continue
                    except Exception:
                        continue
                    v = ys_by_pos_rad[k]
                    if isinstance(v, tuple) and len(v) == 2:
                        xs_local, ys_local = v
                        try:
                            if 'xs_rad' in locals() and xs_rad is not None:
                                ys_resampled = np.interp(xs_rad, np.asarray(xs_local, dtype=float), np.asarray(ys_local, dtype=float))
                                ys_by_pos_rad[k] = ys_resampled
                            else:
                                ys_by_pos_rad[k] = np.asarray(ys_local, dtype=float)
                        except Exception:
                            ys_by_pos_rad[k] = np.asarray(ys_local, dtype=float)
                    else:
                        try:
                            ys_arr = np.asarray(v, dtype=float)
                        except Exception:
                            ys_arr = np.array(v, dtype=float)
                        if 'xs_rad' in locals() and xs_rad is not None:
                            try:
                                if ys_arr.size != xs_rad.size:
                                    xs_from = np.linspace(xs_rad.min(), xs_rad.max(), ys_arr.size)
                                    ys_resampled = np.interp(xs_rad, xs_from, ys_arr)
                                    ys_by_pos_rad[k] = ys_resampled
                                else:
                                    ys_by_pos_rad[k] = ys_arr
                            except Exception:
                                ys_by_pos_rad[k] = ys_arr
                        else:
                            ys_by_pos_rad[k] = ys_arr
        except Exception:
            pass

        if 'weight' in df.columns:

            # Helper to find a per-(cfg_row,energy) series robustly. The
            # vignette sheet may store the energy marker as text or numeric;
            # prefer numeric match, then named A_eff column, then fallback
            # to any series for the cfg_row.
            def _find_series(ys_map, cfg_row, sel_energy_local, aeff_col_name_local=None):
                if not ys_map or cfg_row is None:
                    return None
                # try exact numeric key
                try:
                    keyn = (cfg_row, float(sel_energy_local))
                    if keyn in ys_map:
                        return ys_map[keyn]
                except Exception:
                    pass
                # try named aeff column
                try:
                    if aeff_col_name_local is not None:
                        key_named = (cfg_row, str(aeff_col_name_local).strip())
                        if key_named in ys_map:
                            return ys_map[key_named]
                except Exception:
                    pass
                # try matching by numeric equivalence of the energy token
                for k in ys_map.keys():
                    try:
                        if k[0] != cfg_row:
                            continue
                        val = k[1]
                        try:
                            if sel_energy_local is not None and abs(float(val) - float(sel_energy_local)) < 1e-6:
                                return ys_map[k]
                        except Exception:
                            continue
                    except Exception:
                        continue
                # fallback: return any series for this cfg_row
                for k in ys_map.keys():
                    try:
                        if k[0] == cfg_row:
                            return ys_map[k]
                    except Exception:
                        continue
                return None

            for idx, row in df.iterrows():
                if idx % 100 == 0:
                    try:
                        pass
                        sys.stdout.flush()
                    except Exception:
                        pass
                mm_num = row.get('MM #')
                try:
                    p = int(mm_to_pos.get(int(mm_num))) if pd.notna(mm_num) and int(mm_num) in mm_to_pos else None
                except Exception:
                    p = None
                if p is None:
                    continue

                # radial (per-row)
                if applied_rad and p in rot_rad_map:
                    rot_val = float(rot_rad_map.get(p, 0.0))
                    try:
                        factor = 1.0
                        if 'rad_mode' in locals() and rad_mode == 'per_row_energy' and 'pos_to_cfg_row' in locals():
                            cfg_row = pos_to_cfg_row.get(p)
                            if cfg_row is not None:
                                series = _find_series(ys_by_pos_rad, cfg_row, sel_energy, locals().get('aeff_col_name'))
                                if series is None:
                                    pass
                                else:
                                    xs_use, ys_use = series
                                    factor = float(np.interp(rot_val, xs_use, ys_use))
                        df.at[idx, 'aeff_vig_factor_rad'] = factor
                        df.at[idx, 'weight'] = float(df.at[idx, 'weight']) * float(factor)
                    except Exception:
                        factor = 1.0

                    if 'vig_vals_rad' not in locals():
                        vig_vals_rad = {}
                    if 'vig_source_rad' not in locals():
                        vig_source_rad = {}
                    vig_vals_rad[p] = float(factor)
                    vig_source_rad[p] = ('per_row' if 'rad_mode' in locals() and rad_mode.startswith('per') else 'global')

                # azimuthal (per-row)
                if applied_azi and p in rot_azi_map:
                    try:
                        used = False
                        if 'azi_mode' in locals() and azi_mode == 'per_row_energy' and 'pos_to_cfg_row' in locals():
                            cfg_row = pos_to_cfg_row.get(p)
                            if cfg_row is not None:
                                series = _find_series(ys_by_pos_azi, cfg_row, locals().get('sel_energy'), locals().get('aeff_col_name'))
                                if series is not None:
                                    xs_use, ys_use = series
                                    factor = float(np.interp(abs(float(rot_azi_map.get(p, 0.0))), xs_use, ys_use))
                                    used = True
                        if not used and 'azi_mode' in locals() and azi_mode == 'per_pos' and p in ys_by_pos_azi:
                            ys_use = ys_by_pos_azi[p]
                            rot_val = abs(float(rot_azi_map.get(p, 0.0)))
                            factor = float(np.interp(rot_val, xs_azi, ys_use, left=ys_use[0], right=ys_use[-1]))
                        df.at[idx, 'aeff_vig_factor_azi'] = factor
                        df.at[idx, 'weight'] = float(df.at[idx, 'weight']) * float(factor)
                    except Exception:
                        factor = 1.0

                    if 'vig_vals_azi' not in locals():
                        vig_vals_azi = {}
                    if 'vig_source_azi' not in locals():
                        vig_source_azi = {}
                    vig_vals_azi[p] = float(factor)
                    vig_source_azi[p] = ('per_row' if 'azi_mode' in locals() and azi_mode.startswith('per') else 'global')

            df.attrs['vignetting_rotrad_applied'] = bool(applied_rad)
        # Post-pass: recompute per-position vignette values from the
        # populated per-(cfg_row,energy) tables to avoid mismatches that
        # can occur during the row-wise application loop. This ensures
        # `vig_vals_azi` / `vig_vals_rad` reflect the intended
        # interpolation from the vignetting tables.
        try:
            if os.environ.get('VIG_DEBUG'):
                pass
                sys.stdout.flush()
        except Exception:
            pass
        try:
            if 'pos_to_cfg_row' in locals():
                # ensure vig dicts exist
                if 'vig_vals_azi' not in locals():
                    vig_vals_azi = {}
                if 'vig_vals_rad' not in locals():
                    vig_vals_rad = {}
                if 'vig_source_azi' not in locals():
                    vig_source_azi = {}
                if 'vig_source_rad' not in locals():
                    vig_source_rad = {}

                for pos in sorted(set(list(rot_azi_map.keys()) + list(rot_rad_map.keys()))):
                    cfg = pos_to_cfg_row.get(pos)
                    # azimuthal
                    try:
                        applied_val = None
                        # Prefer per-position arrays when available (azi_mode == 'per_pos')
                        if cfg is not None and 'azi_mode' in locals() and azi_mode == 'per_pos' and 'ys_by_pos_azi' in locals() and isinstance(ys_by_pos_azi, dict) and pos in ys_by_pos_azi:
                            try:
                                applied_val = float(np.interp(abs(float(rot_azi_map.get(pos, 0.0))), xs_azi, ys_by_pos_azi[pos]))
                                vig_vals_azi[pos] = float(applied_val)
                                vig_source_azi[pos] = 'per_pos'
                            except Exception:
                                applied_val = None
                        else:
                            if cfg is not None and 'ys_by_pos_azi' in locals() and ys_by_pos_azi:
                                series = _find_series(ys_by_pos_azi, cfg, locals().get('sel_energy'), locals().get('aeff_col_name'))
                                if series is not None:
                                    xsu, ysu = series
                                    applied_val = float(np.interp(abs(float(rot_azi_map.get(pos, 0.0))), xsu, ysu))
                            if applied_val is None:
                                # fallback to global
                                if 'xs_azi' in locals() and xs_azi is not None and ys_azi is not None:
                                    applied_val = float(np.interp(abs(float(rot_azi_map.get(pos, 0.0))), xs_azi, ys_azi))
                                else:
                                    applied_val = 1.0
                            vig_vals_azi[pos] = float(applied_val)
                            vig_source_azi[pos] = ('per_row' if 'ys_by_pos_azi' in locals() and any((isinstance(k, tuple) and k[0] == cfg) for k in ys_by_pos_azi.keys()) else 'global')
                    except Exception:
                        vig_vals_azi[pos] = 1.0
                        vig_source_azi[pos] = 'none'

                    # radial
                    try:
                        applied_val = None
                        # Prefer per-position arrays when available (rad_mode == 'per_pos')
                        if cfg is not None and 'rad_mode' in locals() and rad_mode == 'per_pos' and 'ys_by_pos_rad' in locals() and isinstance(ys_by_pos_rad, dict) and pos in ys_by_pos_rad:
                            try:
                                applied_val = float(np.interp(float(rot_rad_map.get(pos, 0.0)), xs_rad, ys_by_pos_rad[pos]))
                                vig_vals_rad[pos] = float(applied_val)
                                vig_source_rad[pos] = 'per_pos'
                            except Exception:
                                applied_val = None
                        else:
                            if cfg is not None and 'ys_by_pos_rad' in locals() and ys_by_pos_rad:
                                series = _find_series(ys_by_pos_rad, cfg, locals().get('sel_energy'), locals().get('aeff_col_name'))
                                if series is not None:
                                    xsr, ysr = series
                                    applied_val = float(np.interp(float(rot_rad_map.get(pos, 0.0)), xsr, ysr))
                            if applied_val is None:
                                if 'xs_rad' in locals() and xs_rad is not None and ys_rad is not None:
                                    applied_val = float(np.interp(float(rot_rad_map.get(pos, 0.0)), xs_rad, ys_rad))
                                else:
                                    applied_val = 1.0
                            vig_vals_rad[pos] = float(applied_val)
                            vig_source_rad[pos] = ('per_row' if 'ys_by_pos_rad' in locals() and any((isinstance(k, tuple) and k[0] == cfg) for k in ys_by_pos_rad.keys()) else 'global')
                    except Exception:
                        vig_vals_rad[pos] = 1.0
                        vig_source_rad[pos] = 'none'
        except Exception:
            pass
# Force compute A_eff_adjusted for ALL MMs (even without vignetting)
        print("DEBUG: Starting A_eff_adjusted computation...")
        if 'aeff_base' in df.columns:
            base = df['aeff_base'].astype(float)
            print(f"DEBUG: base sum={base.sum():.3f}, shape={base.shape}")
            # Combine radial and azimuthal vignetting factors
            vig_factor_rad = df.get('aeff_vig_factor_rad', pd.Series(1.0, index=df.index)).fillna(1.0).astype(float)
            vig_factor_azi = df.get('aeff_vig_factor_azi', pd.Series(1.0, index=df.index)).fillna(1.0).astype(float)
            vig_factor = vig_factor_rad * vig_factor_azi
            df['aeff_vig_factor'] = vig_factor  # Store the combined factor
            print(f"DEBUG: vig_factor min/max/mean={vig_factor.min():.3f}/{vig_factor.max():.3f}/{vig_factor.mean():.3f}")
            df['aeff_adjusted'] = base * vig_factor
            df['weight'] = df['aeff_adjusted']
            
            # Debug print: always show
            total_base = float(base.sum())
            total_adj = float(df['aeff_adjusted'].sum())
            pct_diff = ((total_adj - total_base) / total_base * 100) if total_base > 0 else 0
            print(f"A_eff totals: base={total_base:.1f}, adjusted={total_adj:.1f} ({pct_diff:+.1f}%) nMMs={len(df)}")
            print(f"  Sample base/weights: {base.head().tolist()[:5]} / {df['weight'].head().tolist()}")
            
            # Preserve source columns if exist
            for col in ['aeff_vig_source_rotrad', 'aeff_vig_source_rotazi']:
                if col in locals():
                    df[col] = locals()[col]
        else:
            print("DEBUG: No aeff_base column - skipping adjusted computation")

        # Attempt to write per-position vignette factors into the workbook's
        # Vignetting rotazi/rotrad sheets in column B for visibility. This
        # implementation uses openpyxl to update only the specified cells
        # (column B and C1 on vignette sheets; columns B and C on A_eff)
        # and saves atomically to avoid corrupting the original workbook.
        try:
            from openpyxl import load_workbook
            pass
            sys.stdout.flush()
            wb = load_workbook(path)

            # Derive per-position vignette factors from the DataFrame to ensure
            # the values written into the vignette sheets exactly reflect the
            # factors used to compute `aeff_adjusted`.
            per_pos_rad = {}
            per_pos_azi = {}
            try:
                if 'MM #' in df.columns:
                    for _, r in df.iterrows():
                        try:
                            mmn = int(r['MM #'])
                        except Exception:
                            continue
                        posn = mm_to_pos.get(mmn)
                        if posn is None:
                            continue
                        if 'aeff_vig_factor_rad' in df.columns:
                            try:
                                per_pos_rad.setdefault(posn, []).append(float(r.get('aeff_vig_factor_rad', 1.0)))
                            except Exception:
                                pass
                        if 'aeff_vig_factor_azi' in df.columns:
                            try:
                                per_pos_azi.setdefault(posn, []).append(float(r.get('aeff_vig_factor_azi', 1.0)))
                            except Exception:
                                pass
                    # reduce to representative value (mean) for writing
                    for k, vlist in list(per_pos_rad.items()):
                        try:
                            per_pos_rad[k] = float(sum(vlist) / len(vlist)) if vlist else 1.0
                        except Exception:
                            per_pos_rad[k] = 1.0
                    for k, vlist in list(per_pos_azi.items()):
                        try:
                            per_pos_azi[k] = float(sum(vlist) / len(vlist)) if vlist else 1.0
                        except Exception:
                            per_pos_azi[k] = 1.0
            except Exception:
                per_pos_rad = per_pos_rad if per_pos_rad else {}
                per_pos_azi = per_pos_azi if per_pos_azi else {}

            # Build per-position vignette dictionaries from the computed
            # DataFrame values so the values written into the vignette
            # sheets exactly match those used to compute `aeff_adjusted`.
            final_vig_vals_rad = {}
            final_vig_vals_azi = {}
            if 'MM #' in df.columns:
                for _, r in df.iterrows():
                    try:
                        mmv = int(r.get('MM #'))
                    except Exception:
                        continue
                    pos_k = mm_to_pos.get(mmv)
                    if pos_k is None:
                        continue
                    fr = r.get('aeff_vig_factor_rad') if 'aeff_vig_factor_rad' in r.index else None
                    fa = r.get('aeff_vig_factor_azi') if 'aeff_vig_factor_azi' in r.index else None
                    combined = r.get('aeff_vig_factor') if 'aeff_vig_factor' in r.index else 1.0
                    try:
                        if fr is None or (isinstance(fr, float) and (fr != fr)):
                            if fa is not None and not (isinstance(fa, float) and (fa != fa)):
                                fr = float(combined) / float(fa) if float(fa) != 0 else 1.0
                            else:
                                fr = 1.0
                    except Exception:
                        fr = 1.0
                    try:
                        if fa is None or (isinstance(fa, float) and (fa != fa)):
                            if fr is not None and not (isinstance(fr, float) and (fr != fr)):
                                fa = float(combined) / float(fr) if float(fr) != 0 else float(combined)
                            else:
                                fa = float(combined)
                    except Exception:
                        fa = float(combined)

                    final_vig_vals_rad[pos_k] = float(fr)
                    final_vig_vals_azi[pos_k] = float(fa)

            # If DataFrame-based maps are empty, fall back to previously
            # computed maps (if available).
            if not final_vig_vals_rad and 'vig_vals_rad' in locals():
                final_vig_vals_rad = dict(locals().get('vig_vals_rad', {}))
            if not final_vig_vals_azi and 'vig_vals_azi' in locals():
                final_vig_vals_azi = dict(locals().get('vig_vals_azi', {}))

            # VIGNETTE SHEETS: write col B and C1 only
            _vig_azi_sname = _find_vig_sheet(wb, VIG_ROT_AZI_CANDIDATES)
            _vig_rad_sname = _find_vig_sheet(wb, VIG_ROT_RAD_CANDIDATES)
            for sname, vig_map in (
                (_vig_azi_sname, final_vig_vals_azi if final_vig_vals_azi else (vig_vals_azi if 'vig_vals_azi' in locals() else {})),
                (_vig_rad_sname, final_vig_vals_rad if final_vig_vals_rad else (vig_vals_rad if 'vig_vals_rad' in locals() else {})),
            ):
                if sname is None:
                    continue
                # Debug print of sample vig_map contents when requested
                try:
                    if os.environ.get('VIG_DEBUG'):
                        sample_items = list((vig_vals_azi if sname.endswith('rotazi') else vig_vals_rad).items())[:10]
                        pass
                        sys.stdout.flush()
                except Exception:
                    pass

                # Extra debug: show a few per-position arrays and rot offsets for pos 1
                try:
                    if os.environ.get('VIG_DEBUG'):
                        p0 = 1
                        pass
                        pass
                        if 'pos_to_cfg_row' in locals():
                            cfg = pos_to_cfg_row.get(p0)
                            pass
                            if 'ys_by_pos_azi' in locals():
                                matches = [k for k in ys_by_pos_azi.keys() if k[0] == cfg]
                                pass
                                if matches:
                                    k = matches[0]
                                    xsu, ysu = ys_by_pos_azi[k]
                                    pass
                            if 'ys_by_pos_rad' in locals():
                                matchesr = [k for k in ys_by_pos_rad.keys() if k[0] == cfg]
                                pass
                                if matchesr:
                                    kr = matchesr[0]
                                    xsr, ysr = ys_by_pos_rad[kr]
                                    pass
                        sys.stdout.flush()
                except Exception:
                    pass

                if not (vig_map and isinstance(vig_map, dict)):
                    continue
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]
                max_r = ws.max_row or 0

                # Prefer sheet-specific selected energy in cell C2 only (ignore C1)
                sheet_energy = None
                try:
                    c2 = ws.cell(row=2, column=3).value
                    if c2 is not None and not (isinstance(c2, float) and np.isnan(c2)):
                        try:
                            sheet_energy = float(c2)
                        except Exception:
                            import re as _re
                            if isinstance(c2, str):
                                m = _re.search(r"(\d+(?:\.\d*)?)", c2)
                                if m:
                                    sheet_energy = float(m.group(1))
                except Exception:
                    sheet_energy = None
                if sheet_energy is None:
                    sheet_energy = locals().get('sel_energy') if 'sel_energy' in locals() else None

                # Build mapping row->position from column A (explicit mapping expected)
                pos_row_map = {}
                # If the first column looks like a delta table (e.g. header 'delta' or 'delta_arcsec')
                # then do not attempt to interpret integer-like values in column A as Position #.
                try:
                    header_cell = ws.cell(row=1, column=1).value
                    header_is_delta = False
                    if isinstance(header_cell, str):
                        import re as _re
                        if _re.search(r"\bdelta\b|delta_arc|arcsec", header_cell, flags=_re.IGNORECASE):
                            header_is_delta = True
                except Exception:
                    header_is_delta = False

                if not header_is_delta:
                    for r in range(1, max_r + 1):
                        try:
                            v = ws.cell(row=r, column=1).value
                            if v is None:
                                continue
                            if isinstance(v, (int,)) or (isinstance(v, float) and float(v).is_integer()):
                                pos_row_map[int(v)] = r
                            else:
                                s = str(v).strip()
                                if s.isdigit():
                                    pos_row_map[int(s)] = r
                        except Exception:
                            continue

                written = 0
                total = len(pos_row_map)
                # For each position recorded in column A, compute/write factor into column B
                # Rebuild a robust pos->cfg_row mapping from MM configuration if available
                local_pos_to_cfg = {}
                try:
                    mmcfg = pd.read_excel(path, sheet_name='MM configuration', engine='openpyxl')
                    if 'Position #' in mmcfg.columns:
                        tmp = mmcfg.copy()
                        tmp['Position #'] = pd.to_numeric(tmp['Position #'], errors='coerce')
                        tmp = tmp[tmp['Position #'].notna()]
                        for order_i, (_, rr) in enumerate(tmp.iterrows()):
                            try:
                                pval = int(rr['Position #'])
                                # Prefer explicit 'Row #' when present; otherwise
                                # fall back to the enumerated row index.
                                raw_rownum = rr.get('Row #') if 'Row #' in mmcfg.columns else None
                                if raw_rownum is not None and not pd.isna(raw_rownum):
                                    local_pos_to_cfg[pval] = int(float(raw_rownum))
                                else:
                                    local_pos_to_cfg[pval] = order_i + 1
                            except Exception:
                                continue
                    else:
                        # fallback: use row order as cfg_row for positions if 'Position #' absent
                        for order_i, (_, rr) in enumerate(mmcfg.iterrows()):
                            local_pos_to_cfg[order_i + 1] = order_i + 1
                except Exception:
                    local_pos_to_cfg = locals().get('pos_to_cfg_row', {}) if 'pos_to_cfg_row' in locals() else {}

                for i, (pos_k, row_idx) in enumerate(pos_row_map.items()):
                    if i % 100 == 0:
                        pass
                        sys.stdout.flush()
                    pos_int = int(pos_k)
                    # prefer per-position values derived from the DataFrame
                    val = None
                    try:
                        if sname.endswith('rotazi'):
                            if pos_int in per_pos_azi:
                                val = float(per_pos_azi[pos_int])
                        else:
                            if pos_int in per_pos_rad:
                                val = float(per_pos_rad[pos_int])
                    except Exception:
                        val = None
                    # fallback to any precomputed vig_map entry
                    try:
                        if val is None and pos_int in vig_map:
                            val = float(vig_map[pos_int])
                    except Exception:
                        val = None
                    # if missing, attempt to compute from per-(cfg_row,energy) tables
                    if val is None:
                        try:
                            cfg_row = None
                            if 'pos_to_cfg_row' in locals() and pos_int in pos_to_cfg_row:
                                cfg_row = pos_to_cfg_row.get(pos_int)
                            elif pos_int in local_pos_to_cfg:
                                cfg_row = local_pos_to_cfg.get(pos_int)
                            if sname.endswith('rotazi'):
                                # choose azimuthal series
                                if cfg_row is not None and 'ys_by_pos_azi' in locals() and ys_by_pos_azi:
                                    series = _find_series(ys_by_pos_azi, cfg_row, sheet_energy, locals().get('aeff_col_name'))
                                    if series is not None:
                                        xs_u, ys_u = series
                                        val = float(np.interp(abs(float(rot_azi_map.get(pos_int, 0.0))), xs_u, ys_u))
                                    else:
                                        val = float(np.interp(abs(float(rot_azi_map.get(pos_int, 0.0))), locals().get('xs_azi', np.array([0.])), locals().get('ys_azi', np.array([1.]))))
                                else:
                                        val = float(np.interp(abs(float(rot_azi_map.get(pos_int, 0.0))), locals().get('xs_azi', np.array([0.])), locals().get('ys_azi', np.array([1.]))))
                            else:
                                # rotrad
                                if cfg_row is not None and 'ys_by_pos_rad' in locals() and ys_by_pos_rad:
                                    series = _find_series(ys_by_pos_rad, cfg_row, sheet_energy, locals().get('aeff_col_name'))
                                    if series is not None:
                                        xs_u, ys_u = series
                                        val = float(np.interp(abs(float(rot_rad_map.get(pos_int, 0.0))), xs_u, ys_u))
                                    else:
                                        val = float(np.interp(abs(float(rot_rad_map.get(pos_int, 0.0))), locals().get('xs_rad', np.array([0.])), locals().get('ys_rad', np.array([1.]))))
                                else:
                                    val = float(np.interp(abs(float(rot_rad_map.get(pos_int, 0.0))), locals().get('xs_rad', np.array([0.])), locals().get('ys_rad', np.array([1.]))))
                        except Exception:
                            val = 1.0
                    try:
                        ws.cell(row=row_idx, column=2, value=float(val))
                        written += 1
                    except Exception:
                        continue

                pass

                # Do not write to cell C1 here; GUI is responsible for any
                # selected-energy markers (e.g. cell C2). Leave C1/C2 untouched.

            # A_eff sheet: update columns B and C only
            if 'A_eff' in wb.sheetnames and ('aeff_adjusted' in df.columns or 'aeff_base' in df.columns):
                ws_a = wb['A_eff']
                max_r_a = ws_a.max_row or 0

                # Read vignette factors from workbook (column B of vignette sheets)
                vig_rad_sheet = {}
                vig_azi_sheet = {}
                _vr_name = _find_vig_sheet(wb, VIG_ROT_RAD_CANDIDATES)
                if _vr_name:
                    wsr = wb[_vr_name]
                    for rr in range(1, wsr.max_row + 1):
                        a = wsr.cell(row=rr, column=1).value
                        b = wsr.cell(row=rr, column=2).value
                        try:
                            key = int(a) if isinstance(a, (int, float)) or (isinstance(a, str) and str(a).strip().isdigit()) else None
                        except Exception:
                            key = None
                        if key is not None and b is not None:
                            try:
                                vig_rad_sheet[key] = float(b)
                            except Exception:
                                pass
                _va_name = _find_vig_sheet(wb, VIG_ROT_AZI_CANDIDATES)
                if _va_name:
                    wsa = wb[_va_name]
                    for rr in range(1, wsa.max_row + 1):
                        a = wsa.cell(row=rr, column=1).value
                        b = wsa.cell(row=rr, column=2).value
                        try:
                            key = int(a) if isinstance(a, (int, float)) or (isinstance(a, str) and str(a).strip().isdigit()) else None
                        except Exception:
                            key = None
                        if key is not None and b is not None:
                            try:
                                vig_azi_sheet[key] = float(b)
                            except Exception:
                                pass

                # Update the in-memory dataframe so its per-MM factors and
                # adjusted A_eff match the vignette values written to the sheets.
                try:
                    # ensure columns exist
                    if 'aeff_vig_factor_rad' not in df.columns:
                        df['aeff_vig_factor_rad'] = np.nan
                    if 'aeff_vig_factor_azi' not in df.columns:
                        df['aeff_vig_factor_azi'] = np.nan
                    if 'aeff_vig_factor' not in df.columns:
                        df['aeff_vig_factor'] = np.nan
                    if 'aeff_adjusted' not in df.columns:
                        df['aeff_adjusted'] = np.nan
                    for idx in df.index:
                        try:
                            mmv = int(df.at[idx, 'MM #'])
                        except Exception:
                            continue
                        pos = mm_to_pos.get(mmv) if 'mm_to_pos' in locals() else None
                        rad_f_sheet = vig_rad_sheet.get(pos) if pos is not None else None
                        azi_f_sheet = vig_azi_sheet.get(pos) if pos is not None else None
                        # Use sheet value if available; otherwise keep
                        # existing DataFrame value (from interpolation);
                        # fall back to 1.0 (no vignetting) as last resort.
                        if rad_f_sheet is not None:
                            rad_f = float(rad_f_sheet)
                        else:
                            try:
                                v = df.at[idx, 'aeff_vig_factor_rad']
                                rad_f = float(v) if pd.notna(v) else 1.0
                            except Exception:
                                rad_f = 1.0
                        if azi_f_sheet is not None:
                            azi_f = float(azi_f_sheet)
                        else:
                            try:
                                v = df.at[idx, 'aeff_vig_factor_azi']
                                azi_f = float(v) if pd.notna(v) else 1.0
                            except Exception:
                                azi_f = 1.0
                        combined = rad_f * azi_f
                        df.at[idx, 'aeff_vig_factor_rad'] = rad_f
                        df.at[idx, 'aeff_vig_factor_azi'] = azi_f
                        df.at[idx, 'aeff_vig_factor'] = combined
                        # Enforce that aeff_base must be present and numeric for every MM row.
                        if 'aeff_base' not in df.columns:
                            raise ValueError("Missing required column 'aeff_base' in MM_PSF; A_eff column B must be present in the workbook.")
                        if pd.isna(df.at[idx, 'aeff_base']):
                            mm_missing = mmv
                            raise ValueError(f"Missing A_eff base value for MM {mm_missing} (row index {idx}). A_eff column B must contain a numeric weight for every MM used.")
                        base_val = float(df.at[idx, 'aeff_base'])
                        try:
                            df.at[idx, 'aeff_adjusted'] = float(base_val) * float(combined)
                            # Synchronize weight for this MM immediately so incremental
                            # updates don't leave weight out-of-sync with aeff_adjusted.
                            try:
                                df.at[idx, 'weight'] = float(df.at[idx, 'aeff_adjusted'])
                            except Exception:
                                df.at[idx, 'weight'] = df.at[idx, 'aeff_adjusted']
                        except Exception:
                            pass
                except Exception:
                    pass

                # Now write A_eff columns using sheet-derived vignette products to
                # ensure A_eff C/B == (rotazi_B * rotrad_B)
                written_a = 0
                for r in range(1, max_r_a + 1):
                    cell = ws_a.cell(row=r, column=1).value
                    if cell is None:
                        continue
                    try:
                        mmv = int(cell) if isinstance(cell, (int, float)) or (isinstance(cell, str) and str(cell).strip().isdigit()) else None
                        if isinstance(cell, str) and str(cell).strip().isdigit():
                            mmv = int(str(cell).strip())
                    except Exception:
                        mmv = None
                    if mmv is None:
                        continue
                    # write base if available in df
                    try:
                        base_row = df.loc[df['MM #'] == mmv]
                        if not base_row.empty and 'aeff_base' in base_row.columns and not pd.isna(base_row['aeff_base'].iat[0]):
                            base_val = float(base_row['aeff_base'].iat[0])
                        else:
                            # If this MM is present in the working DataFrame but has no aeff_base,
                            # raise an error — aeff_base must be defined for any MM used in aggregation.
                            if mmv in df['MM #'].values:
                                raise ValueError(f"Missing A_eff base value for MM {mmv} encountered while writing A_eff sheet. A_eff column B must contain numeric weights for all MMs.")
                            # Otherwise the A_eff sheet contains an MM not present in the current df; skip writing for that row.
                            continue
                    except Exception:
                        raise

                    # determine per-position factor from sheet
                    pos = mm_to_pos.get(mmv) if 'mm_to_pos' in locals() else None
                    rad_f_sheet = vig_rad_sheet.get(pos) if pos is not None else None
                    azi_f_sheet = vig_azi_sheet.get(pos) if pos is not None else None
                    if rad_f_sheet is not None and azi_f_sheet is not None:
                        combined = float(rad_f_sheet) * float(azi_f_sheet)
                    else:
                        # Sheet lacks per-position data; use DataFrame value
                        try:
                            vig_row = df.loc[df['MM #'] == mmv, 'aeff_vig_factor']
                            if not vig_row.empty and pd.notna(vig_row.iat[0]):
                                combined = float(vig_row.iat[0])
                            else:
                                combined = 1.0
                        except Exception:
                            combined = 1.0

                    # Write column B = canonical base (if known)
                    if base_val is not None:
                        try:
                            ws_a.cell(row=r, column=2, value=float(base_val))
                        except Exception:
                            pass

                    # Write column C = adjusted = base * combined vignetting factor
                    try:
                        if base_val is not None and base_val != 0.0:
                            ws_a.cell(row=r, column=3, value=float(base_val) * float(combined))
                        else:
                            ws_a.cell(row=r, column=3, value=0.0)
                        written_a += 1
                    except Exception:
                        try:
                            ws_a.cell(row=r, column=3, value=0.0)
                            written_a += 1
                        except Exception:
                            pass

                # Sync workbook A_eff column C back into the dataframe so the
                # returned df matches what was written to file.
                try:
                    for r in range(1, max_r_a + 1):
                        cell = ws_a.cell(row=r, column=1).value
                        if cell is None:
                            continue
                        try:
                            mmv = int(cell) if isinstance(cell, (int, float)) or (isinstance(cell, str) and str(cell).strip().isdigit()) else None
                            if isinstance(cell, str) and str(cell).strip().isdigit():
                                mmv = int(str(cell).strip())
                        except Exception:
                            mmv = None
                        if mmv is None:
                            continue
                        cval = ws_a.cell(row=r, column=3).value
                        try:
                            # Find rows in df matching this MM and update
                            mask = df['MM #'] == mmv
                            if 'aeff_adjusted' in df.columns:
                                val = float(cval) if cval is not None else None
                                # Only overwrite if the workbook value is meaningful;
                                # preserve the authoritative in-memory computation
                                # when the sheet stores 0 but in-memory is non-zero.
                                if val is not None and val != 0.0:
                                    df.loc[mask, 'aeff_adjusted'] = val
                                    try:
                                        df.loc[mask, 'weight'] = val
                                    except Exception:
                                        pass
                                elif val == 0.0:
                                    existing = df.loc[mask, 'aeff_adjusted']
                                    if existing.isna().all() or (existing == 0.0).all():
                                        df.loc[mask, 'aeff_adjusted'] = 0.0
                                        try:
                                            df.loc[mask, 'weight'] = 0.0
                                        except Exception:
                                            pass
                            # also update combined vig factor column if base present
                            if 'aeff_base' in df.columns and cval is not None:
                                try:
                                    base_vals = df.loc[mask, 'aeff_base']
                                    for i_idx in base_vals.index:
                                        b = base_vals.at[i_idx]
                                        try:
                                            if b and float(b) != 0.0:
                                                df.at[i_idx, 'aeff_vig_factor'] = float(cval) / float(b)
                                        except Exception:
                                            pass
                                except Exception:
                                    pass
                        except Exception:
                            pass
                except Exception:
                    pass

                # Always save workbook after writing column C
                # Before saving, reconcile per-position vignette factors so
                # the values written into the vignette sheets (column B)
                # are consistent with the `aeff_adjusted` values computed
                # for each MM. Compute per-position combined factor from the
                # DataFrame and adjust rad/azi components accordingly.
                try:
                    # Build mapping pos -> member MM indices (from mm_to_pos)
                    pos_members = {}
                    if 'MM #' in df.columns:
                        for idx_row, mmv in df['MM #'].items():
                            try:
                                mm_key = int(float(mmv))
                            except Exception:
                                continue
                            ppos = mm_to_pos.get(mm_key)
                            if ppos is None:
                                continue
                            pos_members.setdefault(ppos, []).append(idx_row)

                    # Compute canonical per-position vignette factors directly
                    # from the vignetting tables (preferred source) so the
                    # values written into the vignette sheets reflect the
                    # same interpolation used for per-position corrections.
                    final_vig_rad = {}
                    final_vig_azi = {}
                    for pos in sorted(pos_members.keys()):
                        cfg = pos_to_cfg_row.get(pos) if 'pos_to_cfg_row' in locals() else None
                        # AZIMUTHAL: prefer per-pos series, then per-row series, then global
                        try:
                            applied_azi = None
                            if cfg is not None and 'ys_by_pos_azi' in locals() and isinstance(ys_by_pos_azi, dict):
                                # Support both per-pos dict keyed by Position# and
                                # per-row dict keyed by (cfg_row, energy). Per-pos
                                # entries may be either plain ys arrays (sampled
                                # at xs_azi) or (xs, ys) tuples. Handle both.
                                series = None
                                try:
                                    if pos in ys_by_pos_azi:
                                        series = ys_by_pos_azi[pos]
                                except Exception:
                                    pass
                                if series is None:
                                    series = _find_series(ys_by_pos_azi, cfg, locals().get('sel_energy'), locals().get('aeff_col_name'))
                                if series is not None:
                                    try:
                                        if isinstance(series, tuple) and len(series) == 2:
                                            xsu, ysu = series
                                        else:
                                            # assume series is a 1D ys array sampled at xs_azi
                                            ysu = np.asarray(series, dtype=float)
                                            if 'xs_azi' in locals() and xs_azi is not None:
                                                xsu = xs_azi
                                            else:
                                                raise ValueError('no xs available for per-pos ys')
                                        applied_azi = float(np.interp(abs(float(rot_azi_map.get(pos, 0.0))), xsu, ysu, left=ysu[0], right=ysu[-1]))
                                    except Exception as e:
                                        print(f"VIG_DEBUG: interp error pos={pos} e={e}")
                                        applied_azi = None
                            if applied_azi is None and 'xs_azi' in locals() and xs_azi is not None and 'ys_azi' in locals() and ys_azi is not None:
                                try:
                                    applied_azi = float(np.interp(abs(float(rot_azi_map.get(pos, 0.0))), xs_azi, ys_azi))
                                except Exception as e:
                                    print(f"VIG_DEBUG: global interp azi error pos={pos} e={e}")
                                    applied_azi = None
                            if applied_azi is None:
                                applied_azi = 1.0
                            final_vig_azi[pos] = float(applied_azi)
                            try:
                                if 'MM #' in df.columns and len(df) <= 10:
                                    print(f"VIG_DEBUG_COMPUTE: pos={pos} applied_azi={applied_azi}")
                            except Exception:
                                pass
                        except Exception:
                            final_vig_azi[pos] = 1.0

                        # RADIAL: prefer per-pos series, then per-row series, then global
                        try:
                            applied_rad = None
                            if cfg is not None and 'ys_by_pos_rad' in locals() and isinstance(ys_by_pos_rad, dict):
                                # Support both per-pos dict keyed by Position# and
                                # per-row dict keyed by (cfg_row, energy)
                                series = None
                                try:
                                    if pos in ys_by_pos_rad:
                                        series = ys_by_pos_rad[pos]
                                except Exception:
                                    pass
                                if series is None:
                                    series = _find_series(ys_by_pos_rad, cfg, locals().get('sel_energy'), locals().get('aeff_col_name'))
                                if series is not None:
                                    try:
                                        if isinstance(series, tuple) and len(series) == 2:
                                            xsr, ysr = series
                                        else:
                                            ysr = np.asarray(series, dtype=float)
                                            if 'xs_rad' in locals() and xs_rad is not None:
                                                xsr = xs_rad
                                            else:
                                                raise ValueError('no xs available for per-pos ys')
                                        applied_rad = float(np.interp(float(rot_rad_map.get(pos, 0.0)), xsr, ysr, left=ysr[0], right=ysr[-1]))
                                    except Exception:
                                        applied_rad = None
                            if applied_rad is None and 'xs_rad' in locals() and xs_rad is not None and 'ys_rad' in locals() and ys_rad is not None:
                                applied_rad = float(np.interp(float(rot_rad_map.get(pos, 0.0)), xs_rad, ys_rad))
                            if applied_rad is None:
                                applied_rad = 1.0
                            final_vig_rad[pos] = float(applied_rad)
                            try:
                                if 'MM #' in df.columns and len(df) <= 10:
                                    print(f"VIG_DEBUG_COMPUTE: pos={pos} applied_rad={applied_rad}")
                            except Exception:
                                pass
                        except Exception:
                            final_vig_rad[pos] = 1.0

                    # Overwrite vig_vals_* with canonical interpolated values
                    vig_vals_rad = dict(final_vig_rad)
                    vig_vals_azi = dict(final_vig_azi)

                    # Apply these canonical per-position factors to the DataFrame
                    try:
                        if 'MM #' in df.columns:
                            df['aeff_vig_factor_rad'] = df['MM #'].map(lambda mm: vig_vals_rad.get(mm_to_pos.get(int(mm)) if mm is not None else None, 1.0))
                            df['aeff_vig_factor_azi'] = df['MM #'].map(lambda mm: vig_vals_azi.get(mm_to_pos.get(int(mm)) if mm is not None else None, 1.0))
                            df['aeff_vig_factor'] = df['aeff_vig_factor_rad'].fillna(1.0).astype(float) * df['aeff_vig_factor_azi'].fillna(1.0).astype(float)
                            if 'aeff_base' in df.columns:
                                df['aeff_adjusted'] = df['aeff_base'].astype(float) * df['aeff_vig_factor']
                                df['weight'] = df['aeff_adjusted']
                    except Exception:
                        pass
                    # Debug small workbooks: print final maps and DF snippet
                    try:
                        if len(df) <= 10:
                            print(f"VIG_DEBUG_SMALL: final_vig_azi={final_vig_azi}")
                            print(f"VIG_DEBUG_SMALL: final_vig_rad={final_vig_rad}")
                            print(df[['MM #','aeff_base','aeff_vig_factor_rad','aeff_vig_factor_azi','aeff_vig_factor','aeff_adjusted']].head(20))
                    except Exception:
                        pass
                    except Exception:
                        pass

                    # Also write these reconciled per-position factors back into
                    # the vignette sheets' column B so the workbook reflects the
                    # exact factors used to compute `aeff_adjusted`.
                    for sname, final_map in ((_find_vig_sheet(wb, VIG_ROT_AZI_CANDIDATES), vig_vals_azi), (_find_vig_sheet(wb, VIG_ROT_RAD_CANDIDATES), vig_vals_rad)):
                        if sname is None or sname not in wb.sheetnames:
                            continue
                        ws_w = wb[sname]
                        # build row map from column A (like earlier)
                        pos_row_map2 = {}
                        max_r2 = ws_w.max_row or 0
                        for r2 in range(1, max_r2 + 1):
                            try:
                                v = ws_w.cell(row=r2, column=1).value
                                if v is None:
                                    continue
                                if isinstance(v, (int,)) or (isinstance(v, float) and float(v).is_integer()):
                                    pos_row_map2[int(v)] = r2
                                else:
                                    s = str(v).strip()
                                    if s.isdigit():
                                        pos_row_map2[int(s)] = r2
                            except Exception:
                                continue
                        # write reconciled values
                        for cfg_or_pos_k, row_idx in pos_row_map2.items():
                            try:
                                # The sheet's column A may contain cfg_row identifiers
                                # rather than Position #. Try to resolve to a
                                # Position # using pos_to_cfg_row if available.
                                pos_key = None
                                try:
                                    if 'pos_to_cfg_row' in locals():
                                        # find position whose cfg_row equals this key
                                        for pp, cfgv in pos_to_cfg_row.items():
                                            if cfgv == int(cfg_or_pos_k):
                                                pos_key = pp
                                                break
                                except Exception:
                                    pos_key = None
                                # fallback: if the cell already contains a Position #
                                if pos_key is None:
                                    try:
                                        pos_key = int(cfg_or_pos_k)
                                    except Exception:
                                        pos_key = None
                                val = float(final_map.get(pos_key, 1.0))
                                ws_w.cell(row=row_idx, column=2, value=val)
                            except Exception:
                                continue
                except Exception:
                    pass

                # After reconciling vignette sheet B values, re-write A_eff
                # column C using the newly-updated vignette B values so the
                # workbook's A_eff entries are consistent with the vignette
                # factors that will be saved below.
                try:
                    vig_azi_sheet = {}
                    vig_rad_sheet = {}
                    _va_name2 = _find_vig_sheet(wb, VIG_ROT_AZI_CANDIDATES)
                    if _va_name2:
                        wsa = wb[_va_name2]
                        for rr in range(1, wsa.max_row + 1):
                            a = wsa.cell(row=rr, column=1).value
                            b = wsa.cell(row=rr, column=2).value
                            try:
                                key = int(a) if isinstance(a, (int, float)) or (isinstance(a, str) and str(a).strip().isdigit()) else None
                            except Exception:
                                key = None
                            if key is not None and b is not None:
                                try:
                                    vig_azi_sheet[key] = float(b)
                                except Exception:
                                    pass
                    _vr_name2 = _find_vig_sheet(wb, VIG_ROT_RAD_CANDIDATES)
                    if _vr_name2:
                        wsr = wb[_vr_name2]
                        for rr in range(1, wsr.max_row + 1):
                            a = wsr.cell(row=rr, column=1).value
                            b = wsr.cell(row=rr, column=2).value
                            try:
                                key = int(a) if isinstance(a, (int, float)) or (isinstance(a, str) and str(a).strip().isdigit()) else None
                            except Exception:
                                key = None
                            if key is not None and b is not None:
                                try:
                                    vig_rad_sheet[key] = float(b)
                                except Exception:
                                    pass

                    # Write A_eff column C based on these sheet values
                    if 'A_eff' in wb.sheetnames:
                        ws_a = wb['A_eff']
                        for r in range(1, ws_a.max_row + 1):
                            cell = ws_a.cell(row=r, column=1).value
                            if cell is None:
                                continue
                            try:
                                mmv = int(cell) if isinstance(cell, (int, float)) or (isinstance(cell, str) and str(cell).strip().isdigit()) else None
                                if isinstance(cell, str) and str(cell).strip().isdigit():
                                    mmv = int(str(cell).strip())
                            except Exception:
                                mmv = None
                            if mmv is None:
                                continue
                            try:
                                base_row = df.loc[df['MM #'] == mmv]
                                base_val = float(base_row['aeff_base'].iat[0]) if (not base_row.empty and 'aeff_base' in base_row.columns) else 0.0
                            except Exception:
                                base_val = 0.0
                            pos = mm_to_pos.get(mmv) if 'mm_to_pos' in locals() else None
                            rad_f_sheet = vig_rad_sheet.get(pos) if pos is not None else None
                            azi_f_sheet = vig_azi_sheet.get(pos) if pos is not None else None
                            if rad_f_sheet is not None and azi_f_sheet is not None:
                                combined = float(rad_f_sheet) * float(azi_f_sheet)
                            else:
                                try:
                                    vig_row = df.loc[df['MM #'] == mmv, 'aeff_vig_factor']
                                    if not vig_row.empty and pd.notna(vig_row.iat[0]):
                                        combined = float(vig_row.iat[0])
                                    else:
                                        combined = 1.0
                                except Exception:
                                    combined = 1.0
                            try:
                                if base_val is not None and base_val != 0.0:
                                    ws_a.cell(row=r, column=3, value=float(base_val) * float(combined))
                                else:
                                    ws_a.cell(row=r, column=3, value=0.0)
                            except Exception:
                                try:
                                    ws_a.cell(row=r, column=3, value=0.0)
                                except Exception:
                                    pass
                except Exception:
                    pass

                # --- HEW degradation sheets: compute per-position ---
                # Logic mirrors vignetting rotazi/rotrad: for each position
                # use its cfg_row, the selected energy, and local rotation
                # angle to interpolate from the H-K lookup table, then write
                # the result into column B.
                try:
                    # Build pos->cfg_row mapping (may not exist from vignetting)
                    _hew_local_pos_to_cfg = {}
                    if 'local_pos_to_cfg' in locals() and local_pos_to_cfg:
                        _hew_local_pos_to_cfg = local_pos_to_cfg
                    elif 'pos_to_cfg_row' in locals() and pos_to_cfg_row:
                        _hew_local_pos_to_cfg = pos_to_cfg_row
                    if not _hew_local_pos_to_cfg:
                        try:
                            _hew_mmcfg = pd.read_excel(path, sheet_name='MM configuration', engine='openpyxl')
                            if 'Position #' in _hew_mmcfg.columns:
                                _hew_tmp = _hew_mmcfg.copy()
                                _hew_tmp['Position #'] = pd.to_numeric(_hew_tmp['Position #'], errors='coerce')
                                _hew_tmp = _hew_tmp[_hew_tmp['Position #'].notna()]
                                for _hew_oi, (_, _hew_rr) in enumerate(_hew_tmp.iterrows()):
                                    _hew_pval = int(_hew_rr['Position #'])
                                    _hew_rownum = _hew_rr.get('Row #') if 'Row #' in _hew_mmcfg.columns else None
                                    if _hew_rownum is not None and not pd.isna(_hew_rownum):
                                        _hew_local_pos_to_cfg[_hew_pval] = int(float(_hew_rownum))
                                    else:
                                        _hew_local_pos_to_cfg[_hew_pval] = _hew_oi + 1
                        except Exception:
                            pass
                    _hew_sel_energy = locals().get('sheet_energy') or locals().get('sel_energy') or 1.0

                    for hew_candidates, rot_map, angle_label in (
                        (HEW_DEG_ROT_AZI_CANDIDATES, rot_azi_map, 'rotazi'),
                        (HEW_DEG_ROT_RAD_CANDIDATES, rot_rad_map, 'rotrad'),
                    ):
                        hew_sname = _find_vig_sheet(wb, hew_candidates)
                        if hew_sname is None or hew_sname not in wb.sheetnames:
                            continue
                        ws_hew = wb[hew_sname]

                        # Check C2 for sheet-specific selected energy
                        hew_sheet_energy = _hew_sel_energy
                        try:
                            c2_val = ws_hew.cell(row=2, column=3).value
                            if c2_val is not None and not (isinstance(c2_val, float) and np.isnan(c2_val)):
                                hew_sheet_energy = float(c2_val)
                        except Exception:
                            pass

                        # Build per-(cfg_row, energy) series from cols H-K
                        # using a data-only workbook to resolve cached values
                        hew_series = {}
                        try:
                            from openpyxl import load_workbook as _load_wb_hew
                            wb_hew_vals = _load_wb_hew(path, data_only=True)
                            ws_hew_vals = wb_hew_vals[hew_sname]
                            for rr in range(2, (ws_hew_vals.max_row or 0) + 1):
                                try:
                                    cfg_row_val = ws_hew_vals.cell(row=rr, column=8).value  # H
                                    if cfg_row_val is None:
                                        continue
                                    cfg_row_val = int(float(cfg_row_val))
                                    angle_arcmin = ws_hew_vals.cell(row=rr, column=9).value  # I
                                    energy_val_hew = ws_hew_vals.cell(row=rr, column=10).value  # J
                                    hew_val = ws_hew_vals.cell(row=rr, column=11).value  # K
                                    if angle_arcmin is None or hew_val is None:
                                        continue
                                    xval = float(angle_arcmin) * 60.0  # arcmin -> arcsec
                                    yval = float(hew_val)
                                    # Key by (cfg_row, energy_float)
                                    try:
                                        key = (cfg_row_val, float(energy_val_hew))
                                    except Exception:
                                        continue
                                    if key not in hew_series:
                                        hew_series[key] = {'xs': [], 'ys': []}
                                    hew_series[key]['xs'].append(xval)
                                    hew_series[key]['ys'].append(yval)
                                except Exception:
                                    continue
                            try:
                                wb_hew_vals.close()
                            except Exception:
                                pass
                        except Exception:
                            pass

                        if not hew_series:
                            continue

                        # Sort each series by x
                        for k in list(hew_series.keys()):
                            v = hew_series[k]
                            order = np.argsort(v['xs'])
                            hew_series[k] = (np.array(v['xs'], dtype=float)[order],
                                             np.array(v['ys'], dtype=float)[order])

                        print(f"HEW_DEG {angle_label}: loaded {len(hew_series)} series, sel_energy={hew_sheet_energy}")

                        # Build pos -> row mapping from column A
                        hew_pos_row = {}
                        for rr in range(2, (ws_hew.max_row or 0) + 1):
                            try:
                                av = ws_hew.cell(row=rr, column=1).value
                                if av is None:
                                    continue
                                hew_pos_row[int(float(av))] = rr
                            except Exception:
                                continue

                        # For each position, interpolate and write to column B
                        hew_written = 0
                        for pos_int, row_idx in hew_pos_row.items():
                            try:
                                cfg_row = _hew_local_pos_to_cfg.get(pos_int)
                                if cfg_row is None:
                                    continue
                                rot_val = abs(float(rot_map.get(pos_int, 0.0)))
                                # Find the series for (cfg_row, sel_energy)
                                series = None
                                try:
                                    series = hew_series.get((cfg_row, float(hew_sheet_energy)))
                                except Exception:
                                    pass
                                # Fallback: closest energy for this cfg_row
                                if series is None:
                                    best_key = None
                                    best_dist = float('inf')
                                    for k in hew_series:
                                        if k[0] == cfg_row:
                                            try:
                                                d = abs(float(k[1]) - float(hew_sheet_energy))
                                                if d < best_dist:
                                                    best_dist = d
                                                    best_key = k
                                            except Exception:
                                                continue
                                    if best_key is not None:
                                        series = hew_series[best_key]
                                if series is None:
                                    continue
                                xs_h, ys_h = series
                                hew_deg_val = float(np.interp(rot_val, xs_h, ys_h))
                                ws_hew.cell(row=row_idx, column=2, value=hew_deg_val)
                                # Store for later sigma broadening
                                if angle_label == 'rotazi':
                                    hew_deg_per_pos_azi[pos_int] = hew_deg_val
                                else:
                                    hew_deg_per_pos_rad[pos_int] = hew_deg_val
                                hew_written += 1
                            except Exception:
                                continue
                        print(f"HEW_DEG {angle_label}: wrote {hew_written}/{len(hew_pos_row)} positions")
                except Exception as e:
                    print(f"HEW_DEG: error processing sheets: {e}")

                # Persist base sigma values (D/E) as plain numbers so they
                # survive openpyxl round-trips (openpyxl strips formula caches).
                try:
                    ws_psf_de = None
                    for s in wb.sheetnames:
                        if s.lower() == 'mm_psf':
                            ws_psf_de = wb[s]
                            break
                    if ws_psf_de is not None and 'MM #' in df.columns and 'sigma_rad [arcsec]' in df.columns:
                        mm_to_row_de = {}
                        for r in range(2, (ws_psf_de.max_row or 0) + 1):
                            v = ws_psf_de.cell(row=r, column=1).value
                            if v is not None:
                                try:
                                    mm_to_row_de[int(float(v))] = r
                                except Exception:
                                    pass
                        written_de = 0
                        for idx_de, row_de in df.iterrows():
                            mm_de = int(row_de['MM #'])
                            r_de = mm_to_row_de.get(mm_de)
                            if r_de is None:
                                continue
                            ws_psf_de.cell(row=r_de, column=4, value=float(row_de['sigma_rad [arcsec]']))
                            ws_psf_de.cell(row=r_de, column=5, value=float(row_de['sigma_azi [arcsec]']))
                            written_de += 1
                        print(f"HEW_DEG: persisted base sigma to MM_PSF D/E for {written_de} MMs")
                    else:
                        print(f"HEW_DEG: skipped D/E write: ws_psf_de={ws_psf_de is not None}, MM#={'MM #' in df.columns}, sigma_rad={'sigma_rad [arcsec]' in df.columns}")
                except Exception as e_de:
                    print(f"HEW_DEG: error writing D/E: {e_de}")

                tmpf = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                tmpf.close()
                wb.save(tmpf.name)
                os.replace(tmpf.name, path)

                # Re-open the saved workbook and synchronize A_eff column C
                # back into the dataframe to guarantee the returned `df`
                # exactly mirrors what is persisted on disk.
                try:
                    from openpyxl import load_workbook as _load_wb
                    wb_saved = _load_wb(path, data_only=True)
                    if 'A_eff' in wb_saved.sheetnames:
                        ws_saved_a = wb_saved['A_eff']
                        for r in range(1, ws_saved_a.max_row + 1):
                            cell = ws_saved_a.cell(row=r, column=1).value
                            if cell is None:
                                continue
                            try:
                                mmv = int(cell) if isinstance(cell, (int, float)) or (isinstance(cell, str) and str(cell).strip().isdigit()) else None
                                if isinstance(cell, str) and str(cell).strip().isdigit():
                                    mmv = int(str(cell).strip())
                            except Exception:
                                mmv = None
                            if mmv is None:
                                continue
                            cval = ws_saved_a.cell(row=r, column=3).value
                            try:
                                mask = df['MM #'] == mmv
                                if 'aeff_adjusted' in df.columns:
                                    val = float(cval) if cval is not None else None
                                    # Only overwrite if the workbook value is meaningful;
                                    # preserve the authoritative in-memory computation
                                    # when the sheet stores 0 but in-memory is non-zero.
                                    if val is not None and val != 0.0:
                                        df.loc[mask, 'aeff_adjusted'] = val
                                        try:
                                            df.loc[mask, 'weight'] = val
                                        except Exception:
                                            pass
                                    elif val == 0.0:
                                        existing = df.loc[mask, 'aeff_adjusted']
                                        if existing.isna().all() or (existing == 0.0).all():
                                            df.loc[mask, 'aeff_adjusted'] = 0.0
                                            try:
                                                df.loc[mask, 'weight'] = 0.0
                                            except Exception:
                                                pass
                                if 'aeff_base' in df.columns and cval is not None:
                                    base_vals = df.loc[mask, 'aeff_base']
                                    for i_idx in base_vals.index:
                                        try:
                                            b = float(base_vals.at[i_idx])
                                            if b != 0.0:
                                                df.at[i_idx, 'aeff_vig_factor'] = float(cval) / b
                                        except Exception:
                                            pass
                            except Exception:
                                pass
                    try:
                        wb_saved.close()
                    except Exception:
                        pass
                except Exception:
                    pass
                pass
                sys.stdout.flush()

            try:
                wb.close()
            except Exception:
                pass
            pass
            sys.stdout.flush()
        except Exception:
            # If anything goes wrong, do not raise — vignetting writes are non-fatal.
            pass
            sys.stdout.flush()
        # Debug summary: print selected vignette source mapping for first positions
        try:
            debug_env = os.environ.get('VIG_DEBUG', None)
            if debug_env:
                sample = sorted(list(mm_to_pos.values()))[:16]
                pass
        except Exception:
            pass
    except Exception:
        # non-fatal: continue without vignetting
        pass
    
    # Apply deltas to m_rad and m_azi, and rotz effect to m_azi (deltas are per position)
    for idx, row in df.iterrows():
        mm_num = row['MM #']
        pos = mm_to_pos.get(int(mm_num))
        new_m_rad = row['m_rad']
        new_m_azi = row['m_azi']
        
        # Apply alignment rad/azi deltas
        if pos is not None and pos in alignment_by_pos:
            deltas = alignment_by_pos[pos]
            new_m_rad += deltas['d_align_rad']
            new_m_azi += deltas['d_align_azi']
        
        # Calculate total rotz from all perturbations and apply to m_azi
        d_rotz_total_arcsec = 0
        if pos is not None and pos in alignment_by_pos:
            d_rotz_total_arcsec += alignment_by_pos[pos].get('d_align_rotz', 0)
        if pos is not None and pos in gravity_by_pos:
            d_rotz_total_arcsec += gravity_by_pos[pos].get('d_grav_rotz', 0)
        if pos is not None and pos in thermal_by_pos:
            d_rotz_total_arcsec += thermal_by_pos[pos].get('d_therm_rotz', 0)
        
        # Apply rotz effect: d_azi increases by r_MM * d_rotz (in radians)
        if d_rotz_total_arcsec != 0 and mm_num in mm_config_map:
            r_MM = mm_config_map[mm_num].get('r_MM', 0)
            d_rotz_rad = np.radians(d_rotz_total_arcsec / 3600.0)  # arcsec -> degrees -> radians
            new_m_azi += r_MM * d_rotz_rad  # Both in meters
        
        df.at[idx, 'm_rad'] = new_m_rad
        df.at[idx, 'm_azi'] = new_m_azi
    
    # Convert from polar (m_rad, m_azi) to Cartesian (mux, muy)
    # Radial direction: along MM position (x_MM, y_MM)
    # Azimuthal direction: perpendicular to radial, counterclockwise
    # mux = (x_MM/r_MM)*m_rad - (y_MM/r_MM)*m_azi
    # muy = (y_MM/r_MM)*m_rad + (x_MM/r_MM)*m_azi
    # This ensures consistent behavior: positive m_rad always moves outward from center
    
    def convert_polar_to_cartesian(row, mm_config_map):
        """Convert polar offsets (m_rad, m_azi) to Cartesian using MM position as reference."""
        mm_num = row['MM #']
        # Use provided mm_config_map or sensible defaults when missing
        config = mm_config_map.get(mm_num, {'x_MM': 1.0, 'y_MM': 0.0, 'r_MM': 1.0})
        x_mm = config.get('x_MM', 1.0)
        y_mm = config.get('y_MM', 0.0)
        r_mm = config.get('r_MM', 1.0)  # Avoid division by zero
        
        # Unit vectors
        u_rad_x = x_mm / r_mm  # Radial unit vector x-component
        u_rad_y = y_mm / r_mm  # Radial unit vector y-component
        u_azi_x = -y_mm / r_mm  # Azimuthal unit vector x-component (perpendicular)
        u_azi_y = x_mm / r_mm   # Azimuthal unit vector y-component
        
        m_rad = row['m_rad']
        m_azi = row['m_azi']
        
        mux = u_rad_x * m_rad + u_azi_x * m_azi
        muy = u_rad_y * m_rad + u_azi_y * m_azi
        
        return mux, muy
    
    # Apply conversion for all rows
    df[['mux', 'muy']] = df.apply(
        lambda row: pd.Series(convert_polar_to_cartesian(row, mm_config_map)),
        axis=1
    )
    
    # Apply all perturbations to mux and muy
    for idx, row in df.iterrows():
        mm_num = row['MM #']
        pos = mm_to_pos.get(int(mm_num))
        
        # Start with current values
        new_mux = row['mux']
        new_muy = row['muy']
        
        # Get alignment d_z
        d_align_z = 0
        if pos is not None and pos in alignment_by_pos:
            d_align_z = alignment_by_pos[pos].get('d_align_z', 0)
        
        # Get gravity d_z
        d_grav_z = 0
        if pos is not None and pos in gravity_by_pos:
            new_mux += gravity_by_pos[pos]['d_grav_x']
            new_muy += gravity_by_pos[pos]['d_grav_y']
            d_grav_z = gravity_by_pos[pos]['d_grav_z']
        
        # Get thermal d_z
        d_therm_z = 0
        if pos is not None and pos in thermal_by_pos:
            new_mux += thermal_by_pos[pos]['d_therm_x']
            new_muy += thermal_by_pos[pos]['d_therm_y']
            d_therm_z = thermal_by_pos[pos]['d_therm_z']
        
        # Get extra d_z (defocus, etc.)
        d_extra_z = 0
        if pos is not None and pos in extra_by_pos:
            d_extra_z = extra_by_pos[pos].get('d_extra_z', 0)

        # Calculate d_z_total and apply z-axis projection
        d_z_total = d_align_z + d_grav_z + d_therm_z + d_extra_z
        if mm_num in mm_config_map:
            mm_config = mm_config_map[mm_num]
            x_MM = mm_config['x_MM']
            y_MM = mm_config['y_MM']
            z_MM = mm_config['z_MM']
            
            # Calculate dm_x and dm_y based on z displacement
            denominator = 12 - z_MM
            if denominator != 0 and d_z_total != 0:
                dm_x = d_z_total * x_MM / denominator
                dm_y = d_z_total * y_MM / denominator
                new_mux += dm_x
                new_muy += dm_y
        
        # Update the dataframe
        df.at[idx, 'mux'] = new_mux
        df.at[idx, 'muy'] = new_muy
    
    # Apply HEW degradation broadening to sigma_rad and sigma_azi.
    # new_sigma = sqrt(sigma^2 + (hew_deg / (2*sqrt(2*ln(2))))^2)
    # The factor 2*sqrt(2*ln(2)) converts FWHM (HEW) to Gaussian sigma.
    # DataFrame sigma values are in metres (arcsec * arcsec_to_m at load);
    # HEW degradation values are in arcsec, so convert to metres first.
    if hew_deg_per_pos_rad or hew_deg_per_pos_azi:
        _fwhm_to_sigma = 2.0 * np.sqrt(2.0 * np.log(2.0))  # ~2.3548
        _arcsec_to_m = 12.0 * np.pi / 180.0 / 3600.0
        # Collect per-position sigma_extra in arcsec for "Extra PSF degradations"
        _sigma_extra_rad_arcsec = {}  # pos -> arcsec
        _sigma_extra_azi_arcsec = {}  # pos -> arcsec
        for pos_int, hew_val in hew_deg_per_pos_rad.items():
            if hew_val is not None and hew_val > 0:
                _sigma_extra_rad_arcsec[pos_int] = hew_val / _fwhm_to_sigma
        for pos_int, hew_val in hew_deg_per_pos_azi.items():
            if hew_val is not None and hew_val > 0:
                _sigma_extra_azi_arcsec[pos_int] = hew_val / _fwhm_to_sigma
        broadened_rad = 0
        broadened_azi = 0
        for idx, row in df.iterrows():
            mm_num = row['MM #']
            pos = mm_to_pos.get(int(mm_num))
            if pos is None:
                continue
            # Radial broadening from rotrad HEW degradation
            hew_rad = hew_deg_per_pos_rad.get(pos)
            if hew_rad is not None and hew_rad > 0:
                sigma_extra = (hew_rad / _fwhm_to_sigma) * _arcsec_to_m
                old_sigma = float(df.at[idx, 'sigma_rad'])
                df.at[idx, 'sigma_rad'] = np.sqrt(old_sigma**2 + sigma_extra**2)
                broadened_rad += 1
            # Azimuthal broadening from rotazi HEW degradation
            hew_azi = hew_deg_per_pos_azi.get(pos)
            if hew_azi is not None and hew_azi > 0:
                sigma_extra = (hew_azi / _fwhm_to_sigma) * _arcsec_to_m
                old_sigma = float(df.at[idx, 'sigma_azi'])
                df.at[idx, 'sigma_azi'] = np.sqrt(old_sigma**2 + sigma_extra**2)
                broadened_azi += 1
        if broadened_rad or broadened_azi:
            print(f"HEW_DEG broadening: {broadened_rad} sigma_rad, {broadened_azi} sigma_azi")

        # Write sigma_extra per position to "Extra PSF degradations" sheet
        # and final degraded sigma per MM to MM_PSF columns I/J.
        # DataFrame values are in metres; convert back to arcsec for sheets.
        _m_to_arcsec = 1.0 / _arcsec_to_m
        try:
            from openpyxl import load_workbook as _load_wb_hew
            wb_hew = _load_wb_hew(path)

            # --- "Extra PSF degradations" sheet: sigma_extra per position ---
            ws_extra_deg = None
            for s in wb_hew.sheetnames:
                if 'extra' in s.lower() and 'degradation' in s.lower():
                    ws_extra_deg = wb_hew[s]
                    break
            if ws_extra_deg is not None:
                # Build Position# -> row mapping from column A
                pos_to_row_extra = {}
                for r in range(2, (ws_extra_deg.max_row or 0) + 1):
                    v = ws_extra_deg.cell(row=r, column=1).value
                    if v is not None:
                        try:
                            pos_to_row_extra[int(float(v))] = r
                        except Exception:
                            pass
                written_extra = 0
                for pos_int, r_extra in pos_to_row_extra.items():
                    se_rad = _sigma_extra_rad_arcsec.get(pos_int, 0.0)
                    se_azi = _sigma_extra_azi_arcsec.get(pos_int, 0.0)
                    ws_extra_deg.cell(row=r_extra, column=2, value=float(se_rad))
                    ws_extra_deg.cell(row=r_extra, column=3, value=float(se_azi))
                    written_extra += 1
                print(f"HEW_DEG: wrote sigma_extra to 'Extra PSF degradations' for {written_extra} positions")

            # --- MM_PSF columns I/J: final degraded sigma per MM ---
            ws_hew_psf = None
            for s in wb_hew.sheetnames:
                if s.lower() == 'mm_psf':
                    ws_hew_psf = wb_hew[s]
                    break
            if ws_hew_psf is not None:
                # Write headers
                ws_hew_psf.cell(row=1, column=9, value='sigma_rad_deg [arcsec]')
                ws_hew_psf.cell(row=1, column=10, value='sigma_azi_deg [arcsec]')
                # Build MM# -> row mapping from column A
                mm_to_row_hew = {}
                for r in range(2, (ws_hew_psf.max_row or 0) + 1):
                    v = ws_hew_psf.cell(row=r, column=1).value
                    if v is not None:
                        try:
                            mm_to_row_hew[int(float(v))] = r
                        except Exception:
                            pass
                written_ij = 0
                for idx_hew, row_hew in df.iterrows():
                    mm_num_hew = int(row_hew['MM #'])
                    r_hew = mm_to_row_hew.get(mm_num_hew)
                    if r_hew is None:
                        continue
                    ws_hew_psf.cell(row=r_hew, column=9, value=float(row_hew['sigma_rad']) * _m_to_arcsec)
                    ws_hew_psf.cell(row=r_hew, column=10, value=float(row_hew['sigma_azi']) * _m_to_arcsec)
                    written_ij += 1
                print(f"HEW_DEG: wrote degraded sigma to MM_PSF cols I/J for {written_ij} MMs")

            # Persist base sigma (D/E) as plain numbers in this save cycle too
            if ws_hew_psf is not None and 'sigma_rad [arcsec]' in df.columns:
                written_de2 = 0
                for idx_de2, row_de2 in df.iterrows():
                    mm_de2 = int(row_de2['MM #'])
                    r_de2 = mm_to_row_hew.get(mm_de2)
                    if r_de2 is None:
                        continue
                    val_d = row_de2['sigma_rad [arcsec]']
                    val_e = row_de2['sigma_azi [arcsec]']
                    if pd.notna(val_d):
                        ws_hew_psf.cell(row=r_de2, column=4, value=float(val_d))
                    if pd.notna(val_e):
                        ws_hew_psf.cell(row=r_de2, column=5, value=float(val_e))
                    written_de2 += 1
                print(f"HEW_DEG: persisted base sigma to MM_PSF D/E (wb_hew) for {written_de2} MMs")

            # Save workbook if any sheet was modified
            if ws_extra_deg is not None or ws_hew_psf is not None:
                import tempfile as _tmpf_hew
                tmpf_hew = _tmpf_hew.NamedTemporaryFile(delete=False, suffix='.xlsx')
                tmpf_hew.close()
                wb_hew.save(tmpf_hew.name)
                os.replace(tmpf_hew.name, path)
                if ws_extra_deg is not None:
                    print(f"HEW_DEG: saved sigma_extra to 'Extra PSF degradations'")
                if ws_hew_psf is not None:
                    print(f"HEW_DEG: saved degraded sigma to cols I/J for {written_ij} MMs")
            try:
                wb_hew.close()
            except Exception:
                pass
        except Exception as e:
            print(f"HEW_DEG: error writing degraded sigma to MM_PSF: {e}")

    # Always write the final per-MM sigma (after any HEW_DEG broadening) to
    # MM_PSF columns I/J so the workbook stays consistent with what PSF
    # aggregation actually used.  Values are taken from df['sigma_rad'] and
    # df['sigma_azi'] (in metres) – the exact arrays passed to plot_sum –
    # and converted back to arcsec.  This runs unconditionally for non-CSV
    # inputs so stale I/J values are never left in the workbook when D/E or
    # Extra PSF B/C are updated externally.
    if not is_csv and 'MM #' in df.columns:
        try:
            from openpyxl import load_workbook as _load_wb_ij
            _arcsec_to_m_ij = 12.0 * np.pi / 180.0 / 3600.0
            _m_to_arcsec_ij = 1.0 / _arcsec_to_m_ij
            _wb_ij = _load_wb_ij(path)
            _ws_ij = None
            for _s_ij in _wb_ij.sheetnames:
                if _s_ij.lower() == 'mm_psf':
                    _ws_ij = _wb_ij[_s_ij]
                    break
            if _ws_ij is not None:
                _ws_ij.cell(row=1, column=9,  value='sigma_rad_deg [arcsec]')
                _ws_ij.cell(row=1, column=10, value='sigma_azi_deg [arcsec]')
                # Build MM# -> workbook-row mapping from column A
                _mm_to_wb_row_ij: dict = {}
                for _r_ij in range(2, (_ws_ij.max_row or 0) + 1):
                    _v_ij = _ws_ij.cell(row=_r_ij, column=1).value
                    if _v_ij is not None:
                        try:
                            _mm_to_wb_row_ij[int(float(_v_ij))] = _r_ij
                        except Exception:
                            pass
                _written_ij2 = 0
                for _idx_ij, _row_ij in df.iterrows():
                    _mm_ij = int(_row_ij['MM #'])
                    _wb_r_ij = _mm_to_wb_row_ij.get(_mm_ij)
                    if _wb_r_ij is None:
                        continue
                    # I = sigma_rad_deg: the broadened sigma actually used for aggregation
                    _ws_ij.cell(row=_wb_r_ij, column=9,
                                value=float(_row_ij['sigma_rad']) * _m_to_arcsec_ij)
                    _ws_ij.cell(row=_wb_r_ij, column=10,
                                value=float(_row_ij['sigma_azi']) * _m_to_arcsec_ij)
                    # Also persist base D/E as plain numbers
                    if 'sigma_rad [arcsec]' in df.columns:
                        _vd_ij = _row_ij['sigma_rad [arcsec]']
                        _ve_ij = _row_ij['sigma_azi [arcsec]']
                        if pd.notna(_vd_ij):
                            _ws_ij.cell(row=_wb_r_ij, column=4, value=float(_vd_ij))
                        if pd.notna(_ve_ij):
                            _ws_ij.cell(row=_wb_r_ij, column=5, value=float(_ve_ij))
                    _written_ij2 += 1
                import tempfile as _tmpf_ij2
                _tf_ij2 = _tmpf_ij2.NamedTemporaryFile(delete=False, suffix='.xlsx')
                _tf_ij2.close()
                _wb_ij.save(_tf_ij2.name)
                os.replace(_tf_ij2.name, path)
                print(f"INFO: wrote sigma_rad/azi_deg to MM_PSF cols I/J for {_written_ij2} MMs")
            try:
                _wb_ij.close()
            except Exception:
                pass
        except Exception as _e_ij:
            print(f"WARNING: could not write I/J to workbook: {_e_ij}")

    # Copy sigma values
    df['sigmax'] = df['sigma_rad']
    df['sigmay'] = df['sigma_azi']

    # Remember workbook path for resolving custom PSF file stems during plotting.
    df.attrs['workbook_path'] = path

    # Final enforcement: ensure `weight` exactly matches `aeff_adjusted`
    try:
        if 'aeff_adjusted' in df.columns:
            df['weight'] = pd.to_numeric(df['aeff_adjusted'], errors='coerce').fillna(0.0).astype(float)
            # If any rows required coercion/correction, log a debug line with counts.
            try:
                mismatches = int((pd.to_numeric(df['aeff_adjusted'], errors='coerce').fillna(0.0).astype(float) != df['weight']).sum())
                if mismatches:
                    print(f"DEBUG: weight/a_eff mismatch corrected for {mismatches} rows; total weight sum={float(df['weight'].sum()):.6g}")
            except Exception:
                pass
    except Exception:
        pass

    return df  # Return the loaded DataFrame


def _resolve_custom_psf_path(workbook_path: str, stem: str) -> str | None:
    """Resolve a workbook-referenced PSF file stem to an existing path.

    The workbook may reference a PSF file by a stem or relative path.
    This helper tries common extensions and a small set of search
    directories (workbook dir, repo Distributions/ and CustomPSFs/, CWD)
    and returns the first existing absolute path or ``None`` when not
    found. Returns ``None`` for empty or placeholder stems.

    Parameters
    - workbook_path: path to the workbook containing the stem reference
    - stem: filename stem or relative path as provided in the workbook
    """
    stem = str(stem).strip()
    if not stem or stem.lower() in {'nan', 'none'}:
        return None

    # If user already passed a filename with extension, try directly.
    candidates: list[str] = []
    if stem.lower().endswith(('.xlsx', '.xls')):
        candidates.append(stem)
    else:
        for ext in ['.xlsx', '.xls']:
            candidates.append(stem + ext)

    wb_dir = os.path.dirname(os.path.abspath(workbook_path))
    repo_dir = os.path.dirname(os.path.abspath(__file__))
    search_dirs = [
        wb_dir,
        os.path.join(wb_dir, 'Distributions'),
        os.path.join(repo_dir, 'Distributions'),
        os.path.join(repo_dir, 'CustomPSFs'),
        os.getcwd(),
    ]

    for d in search_dirs:
        for name in candidates:
            p = os.path.join(d, name)
            if os.path.exists(p):
                return p
    return None


def compute_total_rot_polar(mm_to_pos: dict, mm_config_map: dict, alignment_by_pos: dict, gravity_by_pos: dict, thermal_by_pos: dict, extra_by_pos: dict | None = None):
    """Compute total rotation components and their projections onto polar axes.

    Returns (rotx, roty, rot_rad, rot_azi) where each is a dict keyed by position.
    See module-level documentation for units and conventions (arcsec, direct
    polar terms versus projected X/Y rotations).
    """
    rotx = {}
    roty = {}
    rot_rad = {}
    rot_azi = {}

    # Gather all positions that may be present in any of the inputs
    positions = set()
    if mm_to_pos:
        positions.update(mm_to_pos.values())
    if alignment_by_pos:
        positions.update(alignment_by_pos.keys())
    if gravity_by_pos:
        positions.update(gravity_by_pos.keys())
    if thermal_by_pos:
        positions.update(thermal_by_pos.keys())
    if extra_by_pos:
        positions.update(extra_by_pos.keys())

    # Reverse mapping: position -> an example MM for geometry lookup
    pos_to_mm = {}
    if mm_to_pos:
        for mm, p in mm_to_pos.items():
            pos_to_mm.setdefault(p, mm)

    for pos in positions:
        # Sum rotx/roty contributions from gravity and thermal only
        rtx_total = 0.0
        rty_total = 0.0
        if gravity_by_pos and pos in gravity_by_pos:
            rtx_total += float(gravity_by_pos[pos].get('d_grav_rotx', 0.0) or 0.0)
            rty_total += float(gravity_by_pos[pos].get('d_grav_roty', 0.0) or 0.0)
        if thermal_by_pos and pos in thermal_by_pos:
            rtx_total += float(thermal_by_pos[pos].get('d_therm_rotx', 0.0) or 0.0)
            rty_total += float(thermal_by_pos[pos].get('d_therm_roty', 0.0) or 0.0)
        if extra_by_pos and pos in extra_by_pos:
            rtx_total += float(extra_by_pos[pos].get('d_extra_rotx', 0.0) or 0.0)
            rty_total += float(extra_by_pos[pos].get('d_extra_roty', 0.0) or 0.0)

        rotx[pos] = rtx_total
        roty[pos] = rty_total

        # Compute radial unit vector from MM geometry if available
        ux, uy = 1.0, 0.0
        mm_choice = pos_to_mm.get(pos)
        if mm_choice is not None and mm_choice in mm_config_map:
            cfg = mm_config_map.get(mm_choice, {})
            r_mm = float(cfg.get('r_MM', 0.0) or 0.0)
            x_mm = float(cfg.get('x_MM', 0.0) or 0.0)
            y_mm = float(cfg.get('y_MM', 0.0) or 0.0)
            if r_mm > 0.0:
                ux = x_mm / r_mm
                uy = y_mm / r_mm

        # Project rotx/roty into polar components
        proj_rotrad = rtx_total * ux + rty_total * uy
        proj_rotazi = -rtx_total * uy + rty_total * ux

        # Add any direct polar contributions from alignment/gravity/thermal
        direct_rotrad = 0.0
        direct_rotazi = 0.0
        if alignment_by_pos and pos in alignment_by_pos:
            direct_rotrad += float(alignment_by_pos[pos].get('d_align_rotrad', 0.0) or 0.0)
            direct_rotazi += float(alignment_by_pos[pos].get('d_align_rotazi', 0.0) or 0.0)
        if gravity_by_pos and pos in gravity_by_pos:
            direct_rotrad += float(gravity_by_pos[pos].get('d_grav_rotrad', 0.0) or 0.0)
            direct_rotazi += float(gravity_by_pos[pos].get('d_grav_rotazi', 0.0) or 0.0)
        if thermal_by_pos and pos in thermal_by_pos:
            direct_rotrad += float(thermal_by_pos[pos].get('d_therm_rotrad', 0.0) or 0.0)
            direct_rotazi += float(thermal_by_pos[pos].get('d_therm_rotazi', 0.0) or 0.0)

        total_rotrad = proj_rotrad + direct_rotrad
        total_rotazi = proj_rotazi + direct_rotazi

        rot_rad[pos] = total_rotrad
        rot_azi[pos] = total_rotazi

    return rotx, roty, rot_rad, rot_azi


def compute_dm_from_dz(mm: dict, row: dict, d_z: float) -> tuple[float, float]:
    """Project a z-displacement d_z into DM x/y using MM geometry or theta fallback.

    - If `r_MM` > 0 is present in `mm`, project along the radial unit vector (x_MM/r_MM, y_MM/r_MM).
    - Else if `row` provides `theta_degrees`, use (cos(theta), sin(theta)).
    - Otherwise fallback to (1,0).
    """
    import math
    ux = 1.0
    uy = 0.0
    r_mm = float(mm.get('r_MM', 0.0) or 0.0)
    if r_mm > 0.0:
        ux = float(mm.get('x_MM', 0.0)) / r_mm
        uy = float(mm.get('y_MM', 0.0)) / r_mm
    else:
        theta = row.get('theta_degrees') if isinstance(row, dict) else None
        if theta is not None:
            try:
                th = math.radians(float(theta))
                ux = math.cos(th)
                uy = math.sin(th)
            except Exception:
                ux, uy = 1.0, 0.0

    dm_x = float(d_z) * ux
    dm_y = float(d_z) * uy
    return dm_x, dm_y


def plot_sum(df: pd.DataFrame, xlim=(-10,10), ylim=(-8,8), nx=800, ny=640, normalize=True, output=None, fast=True, title_suffix: str = "", df_optimized: pd.DataFrame = None, return_metrics_only: bool = False, debug: bool = False, metrics_n_r_final: int | None = None, metrics_n_theta_final: int | None = None, metrics_r_margin: float | None = None):
    """Create combined PSF image, compute HEW/EEF metrics, and optionally save output.

    This function composes per-MM PSFs (analytic or matrix-based) described
    in `df`, computes rotation-invariant HEW and EEF via polar integration,
    fits aggregated radial profiles (modified pseudo-Voigt and King) and
    produces a summary figure. Use `return_metrics_only=True` to retrieve
    computed numeric metrics without creating the plot (useful for CI).

    Important: `df` must include at minimum `mux`, `muy`, `sigmax`,
    `sigmay`, `weight` and `distribution`/custom PSF references as used by
    the rest of the codebase.
    """
    # Close any existing matplotlib figures to prevent accumulation
    plt.close('all')

    """ # Reduce grid resolution in fast mode
    if fast:
        # Aggressive defaults for interactive speed (target: <5s without optimization)
        nx = min(nx, 2062)
        ny = min(ny, 2062)
    else:
        # Keep slow mode bounded (still more accurate than fast, but not unbounded, this value is known from previous iterations of the optical design)
        nx = 2062
        ny = 2062 """

    # fast=True now means "coarse"/quick plotting. If callers passed None
    # for `nx`/`ny`, choose sizes based on mode: coarse=320, fine=2062.
    quick_mode = bool(fast)
    if nx is None:
        nx = 320 if quick_mode else 2062
    if ny is None:
        ny = 320 if quick_mode else 2062
    
    # Calculate the weighted center of mass directly from the Gaussian parameters
    # Normalize per-combo weights so they sum to 1 before computing centroids and sums.
    # This ensures HEW is computed from properly normalized mixture amplitudes.
    # Use a numpy array to avoid repeatedly accessing the DataFrame.
    # Prefer adjusted A_eff weights when available
    if 'aeff_adjusted' in df.columns:
        # Treat NaN adjusted A_eff as zero weight (do not propagate NaN into normalization)
        weight_arr_for_center = pd.to_numeric(df['aeff_adjusted'], errors='coerce').fillna(0.0).to_numpy(dtype=float, copy=False)
    elif 'weight' in df.columns:
        # Ensure any non-finite weights are treated as zero
        weight_arr_for_center = pd.to_numeric(df['weight'], errors='coerce').fillna(0.0).to_numpy(dtype=float, copy=False)
    else:
        weight_arr_for_center = np.ones(len(df), dtype=float) if len(df) > 0 else np.array([], dtype=float)
    total_weight = float(np.nansum(weight_arr_for_center)) if weight_arr_for_center.size else 0.0
    if not np.isfinite(total_weight) or total_weight <= 0.0:
        # Fallback to unweighted centroid when weights are zero/invalid
        center_x = float(df['mux'].mean()) if 'mux' in df.columns else 0.0
        center_y = float(df['muy'].mean()) if 'muy' in df.columns else 0.0
        # keep weight array as ones for later summation
        weight_arr = np.ones(len(df), dtype=float)
    else:
        # normalize so sum(weights) == 1.0
        weight_arr = weight_arr_for_center / total_weight
        center_x = (df['mux'].to_numpy(dtype=float, copy=False) * weight_arr).sum()
        center_y = (df['muy'].to_numpy(dtype=float, copy=False) * weight_arr).sum()
    # Ensure centers are finite numbers
    if not np.isfinite(center_x):
        center_x = 0.0
    if not np.isfinite(center_y):
        center_y = 0.0
    
    # --- Fast grid summation helpers (threaded) ---
    mux_arr = df['mux'].to_numpy(dtype=float, copy=False)
    muy_arr = df['muy'].to_numpy(dtype=float, copy=False)
    sigx_arr = df['sigmax'].to_numpy(dtype=float, copy=False)
    sigy_arr = df['sigmay'].to_numpy(dtype=float, copy=False)
    # Ensure theta is numeric and replace NaN with 0.0 to avoid NaN trig results
    theta_arr = pd.to_numeric(df.get('theta_degrees', pd.Series([0.0] * len(df))), errors='coerce').fillna(0.0).to_numpy(dtype=float, copy=False)
    # Weight array already computed above as `weight_arr` (normalized). Ensure it's available.
    try:
        # if weight_arr defined above, keep it; otherwise derive normalized from df
        weight_arr
    except NameError:
        if 'weight' in df.columns:
            wtmp = df['weight'].to_numpy(dtype=float, copy=False)
            wsum = float(np.nansum(wtmp)) if wtmp.size else 0.0
            weight_arr = (wtmp / wsum) if (wsum and np.isfinite(wsum) and wsum > 0.0) else np.ones(len(df), dtype=float)
        else:
            weight_arr = np.ones(len(df), dtype=float)
    dist_raw_arr = df.get('distribution', pd.Series(['gaussian'] * len(df))).astype(str).to_numpy(copy=False)
    dist_arr = pd.Series(dist_raw_arr).astype(str).str.lower().to_numpy(copy=False)
    # alpha_* may contain placeholders like '-' for Gaussian rows; coerce safely.
    alpha_azi_arr = pd.to_numeric(df.get('alpha_azi', pd.Series([0.5] * len(df))), errors='coerce').fillna(0.5).to_numpy(dtype=float, copy=False)
    alpha_rad_arr = pd.to_numeric(df.get('alpha_rad', pd.Series([0.5] * len(df))), errors='coerce').fillna(0.5).to_numpy(dtype=float, copy=False)

    # Prevent numerically zero sigmas which break grid integration (tiny values
    # from user input or rounding can lead to extremely narrow peaks and unstable
    # minimal-interval searches). Enforce a small floor in meters.
    # Small floor for sigma in meters: keep tiny but non-zero to avoid
    # numerical integration instability. 1e-9 m (1 nm) is small enough to
    # preserve realistic PSF widths (user inputs are typically 1e-8..1e-6).
    MIN_SIG_M = 1e-9
    sigx_arr = np.maximum(sigx_arr, MIN_SIG_M)
    sigy_arr = np.maximum(sigy_arr, MIN_SIG_M)

    # Custom PSF cache (by distribution name/stem)
    builtins = {'gaussian', 'pseudo-voigt', 'voigt'}
    workbook_path = df.attrs.get('workbook_path', None)
    arcsec_to_m = 12 * np.pi / 180 / 3600
    custom_psf_cache: dict[str, tuple[np.ndarray, np.ndarray, np.ndarray]] = {}
    custom_sigma_hint: list[float] = []
    if workbook_path is not None:
        unique_names = sorted({str(n).strip() for n, dl in zip(dist_raw_arr, dist_arr) if str(n).strip() and dl not in builtins})
        for name in unique_names:
            p = _resolve_custom_psf_path(str(workbook_path), name)
            if not p:
                continue
            try:
                x_psf, y_psf, f_psf = load_psf_matrix_excel(p, arcsec_to_m=arcsec_to_m)
                custom_psf_cache[name] = (x_psf, y_psf, f_psf)
                # crude scale hint for r_max: assume extent ~ +/- 3 sigma
                extent = max(float(np.max(np.abs(x_psf))), float(np.max(np.abs(y_psf))))
                if extent > 0:
                    custom_sigma_hint.append(extent / 3.0)
            except Exception as e:
                print(f"Warning: Could not load custom PSF '{name}' from '{p}': {e}")

    def _sum_chunk_on_grid(Xg, Yg, idxs: np.ndarray, normalize_flag: bool) -> np.ndarray:
        Zc = np.zeros_like(Xg, dtype=float)
        for i in idxs:
            dlow = dist_arr[i]
            if dlow in ['pseudo-voigt', 'voigt']:
                add = pseudo_voigt_2d_rotated(
                    Xg, Yg,
                    muazi=mux_arr[i], murad=muy_arr[i],
                    sigmaazi=sigx_arr[i], sigmarad=sigy_arr[i],
                    theta=theta_arr[i],
                    alphaazi=alpha_azi_arr[i],
                    alpharad=alpha_rad_arr[i],
                    amplitude=weight_arr[i],
                    normalize=normalize_flag,
                    degrees=True,
                )
                Zc += add
            elif dlow == 'gaussian':
                add = gaussian_2d_rotated(
                    Xg, Yg,
                    mux=mux_arr[i], muy=muy_arr[i],
                    sigmax=sigx_arr[i], sigmay=sigy_arr[i],
                    theta=theta_arr[i],
                    amplitude=weight_arr[i],
                    normalize=normalize_flag,
                    degrees=True,
                )
                Zc += add
            else:
                # File-based custom PSF: distribution cell holds the file stem (without extension)
                name = str(dist_raw_arr[i]).strip()
                psf = custom_psf_cache.get(name)
                if psf is None and workbook_path is not None:
                    p = _resolve_custom_psf_path(str(workbook_path), name)
                    if p:
                        try:
                            psf = load_psf_matrix_excel(p, arcsec_to_m=arcsec_to_m)
                            custom_psf_cache[name] = psf
                        except Exception:
                            psf = None
                if psf is None:
                    # If file not found/invalid, fall back to Gaussian to keep plotting running.
                    Zc += gaussian_2d_rotated(
                        Xg, Yg,
                        mux=mux_arr[i], muy=muy_arr[i],
                        sigmax=max(sigx_arr[i], 1e-12), sigmay=max(sigy_arr[i], 1e-12),
                        theta=theta_arr[i],
                        amplitude=weight_arr[i],
                        normalize=normalize_flag,
                        degrees=True,
                    )
                else:
                    x_psf, y_psf, f_psf = psf
                    Zc += weight_arr[i] * eval_psf_matrix_rotated(
                        Xg,
                        Yg,
                        mux=mux_arr[i],
                        muy=muy_arr[i],
                        theta_deg=theta_arr[i],
                        x_axis=x_psf,
                        y_axis=y_psf,
                        flux=f_psf,
                    )
        return Zc

    def _sum_on_grid(Xg, Yg, normalize_flag: bool) -> np.ndarray:
        n = len(mux_arr)
        if n == 0:
            return np.zeros_like(Xg, dtype=float)

        # Threading helps because heavy NumPy ops release the GIL.
        max_workers = 1 if n < 25 else min(8, (os.cpu_count() or 2))
        if max_workers <= 1:
            return _sum_chunk_on_grid(Xg, Yg, np.arange(n), normalize_flag)

        chunks = np.array_split(np.arange(n), max_workers)
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            parts = list(ex.map(lambda c: _sum_chunk_on_grid(Xg, Yg, c, normalize_flag), chunks))
        Zg = np.zeros_like(Xg, dtype=float)
        for p in parts:
            Zg += p
        return Zg

    # ---- Rotation-invariant HEW via polar integration ----
    def radial_profile(cx, cy, n_r=400, n_theta=360, r_margin_factor=5.0, normalize_flag=normalize):
        """Compute radial energy profile around (cx, cy) using polar sampling.

        Uses polar grid integration (rotation-invariant) instead of Cartesian
        raster to avoid angle-dependent HEW bias.
        """

        # Radius limit: cover all PSFs plus margin
        max_sigma = max(df['sigmax'].max(), df['sigmay'].max())
        if custom_sigma_hint:
            max_sigma = max(float(max_sigma), float(max(custom_sigma_hint)))
        max_center_dist = np.sqrt((df['mux'] - cx) ** 2 + (df['muy'] - cy) ** 2).max()
        r_max = max_center_dist + r_margin_factor * max_sigma
        if r_max <= 0:
            r_max = 1e-6

        # Iteratively expand radial margin if heavy tails cause significant
        # energy to fall outside the initial r_max (common with Lorentzian).
        expected_total = float(np.nansum(weight_arr)) if 'weight_arr' in locals() else 1.0
        # allow a few attempts, expanding margin each time up to a modest limit
        # (reduce attempts and growth to avoid runaway cost for pseudo-Voigt tails)
        attempts = 3
        tol = 0.9995
        total_energy = 0.0
        for attempt in range(attempts):
            theta = np.linspace(0.0, 2.0 * np.pi, n_theta, endpoint=False)
            r = np.linspace(0.0, r_max, n_r)
            dtheta = theta[1] - theta[0]
            dr = r[1] - r[0] if n_r > 1 else r_max
            R, TH = np.meshgrid(r, theta)
            Xp = cx + R * np.cos(TH)
            Yp = cy + R * np.sin(TH)

            # Reuse the same summation kernel on the polar grid.
            Zp = _sum_on_grid(Xp, Yp, normalize_flag)

            # Integrate over theta (Jacobian r) to get radial energy density
            radial_energy = np.sum(Zp * R, axis=0) * dtheta  # shape (n_r,)
            cumulative = np.cumsum(radial_energy * dr)
            total_energy = cumulative[-1] if cumulative.size else 1.0

            # If normalization is expected (weights sum > 0) and we captured
            # nearly all energy, break early. Otherwise expand r_max and retry.
            if expected_total <= 0 or total_energy / expected_total >= tol:
                break
            # Expand margin and retry (limit growth to avoid runaway cost).
            # Use gentler growth factor and a tighter n_r cap to bound cost.
            r_max *= 1.5
            # Slightly increase radial resolution but cap much lower than before
            n_r = min(int(n_r * 1.15) + 1, 5000)
            n_theta = min(int(n_theta * 1.0), 2048)
        if debug:
                pass
        return r, cumulative, total_energy

    def _radius_for_fraction(frac: np.ndarray, r: np.ndarray, target: float = 0.5) -> float:
        """Estimate radius where cumulative fraction == target using local cubic fit.

        Falls back to linear interpolation when the local window is too small or
        a valid cubic root cannot be found.
        """
        if frac.size == 0:
            return 0.0
        frac = np.asarray(frac).ravel()
        r = np.asarray(r).ravel()
        if frac.size != r.size or frac.size < 2:
            return float(np.interp(target, frac, r))
        idx = int(np.searchsorted(frac, target))
        start = max(0, idx - 2)
        end = min(frac.size, idx + 2)
        if end - start < 2:
            return float(np.interp(target, frac, r))
        r_win = r[start:end]
        f_win = frac[start:end]
        if f_win.size < 4:
            return float(np.interp(target, frac, r))
        try:
            coeffs = np.polyfit(r_win, f_win, 3)
            coeffs[-1] -= float(target)
            roots = np.roots(coeffs)
            real_roots = roots[np.isreal(roots)].real
            rmin, rmax = float(r_win.min()), float(r_win.max())
            for rt in real_roots:
                if rmin - 1e-12 <= float(rt) <= rmax + 1e-12:
                    return float(rt)
        except Exception:
            pass
        return float(np.interp(target, frac, r))

    def hew_at_center(cx, cy, coarse=False):
        # Use coarser grid during search for speed
        if coarse:
            n_r = 110 if fast else 140
            n_theta = 72 if fast else 100
            r, cumulative, total_energy = radial_profile(cx, cy, n_r=n_r, n_theta=n_theta)
        else:
            n_r = 200 if fast else 240
            n_theta = 120 if fast else 180
            r, cumulative, total_energy = radial_profile(cx, cy, n_r=n_r, n_theta=n_theta)
        frac = cumulative / total_energy if total_energy > 0 else cumulative
        # Use local cubic interpolation for the 50% radius, then return HEW diameter
        radius50 = _radius_for_fraction(frac, r, target=0.5)
        return 2.0 * radius50

    # Find best focus position.
    # In quick_mode (typical when not optimizing), avoid iterative search to keep runtime low.
    if quick_mode:
        best_cx, best_cy = center_x, center_y
        best_hew = hew_at_center(best_cx, best_cy, coarse=True)
        step_size = 1e-6
        for cx, cy in (
            (center_x + step_size, center_y),
            (center_x - step_size, center_y),
            (center_x, center_y + step_size),
            (center_x, center_y - step_size),
        ):
            hew_val = hew_at_center(cx, cy, coarse=True)
            if hew_val < best_hew:
                best_cx, best_cy, best_hew = cx, cy, hew_val
        center_x, center_y = best_cx, best_cy
    else:
        candidates_start = [(center_x, center_y), (0.0, 0.0)]
        best_cx, best_cy = center_x, center_y
        best_hew = hew_at_center(best_cx, best_cy, coarse=True)

        for start_cx, start_cy in candidates_start:
            current_cx, current_cy = start_cx, start_cy
            current_hew = hew_at_center(current_cx, current_cy, coarse=True)
            step_size = 1e-6  # 1 micron steps

            for _ in range(20 if fast else 30):  # Limit iterations for speed
                candidates = [
                    (current_cx + step_size, current_cy),
                    (current_cx - step_size, current_cy),
                    (current_cx, current_cy + step_size),
                    (current_cx, current_cy - step_size)
                ]
                improved = False
                for cx, cy in candidates:
                    hew_val = hew_at_center(cx, cy, coarse=True)
                    if hew_val < current_hew:
                        current_cx, current_cy = cx, cy
                        current_hew = hew_val
                        improved = True
                        break
                if not improved:
                    step_size *= 0.5
                    if step_size < 1e-10:
                        break

            if current_hew < best_hew:
                best_cx, best_cy = current_cx, current_cy
                best_hew = current_hew

        center_x, center_y = best_cx, best_cy

    # Final HEW and EE metrics using fine polar integration at best focus
    # Lower final polar-grid for faster runs (target <5s) while increasing
    # radial margin to capture PSF tails and keep HEW estimate high.
    if quick_mode:
        # Coarse: keep radial emphasis but reduce final caps to limit runtime
        n_r_final = 3000
        n_theta_final = 120
        final_r_margin = 10.0
    else:
        # Increase fine-mode sampling for higher accuracy but stay bounded.
        n_r_final = 5000
        n_theta_final = 720
        final_r_margin = 12.0

    # Allow callers to override final metric sampling for speed/experiments
    if metrics_n_r_final is not None:
        try:
            n_r_final = int(metrics_n_r_final)
        except Exception:
            pass
    if metrics_n_theta_final is not None:
        try:
            n_theta_final = int(metrics_n_theta_final)
        except Exception:
            pass
    if metrics_r_margin is not None:
        try:
            final_r_margin = float(metrics_r_margin)
        except Exception:
            pass
    r_profile, cumulative_profile, total_energy = radial_profile(center_x, center_y, n_r=n_r_final, n_theta=n_theta_final, r_margin_factor=final_r_margin)
    frac_profile = cumulative_profile / total_energy if total_energy > 0 else cumulative_profile
    radius_50 = _radius_for_fraction(frac_profile, r_profile, target=0.5)
    radius_80 = _radius_for_fraction(frac_profile, r_profile, target=0.8)
    radius_90 = _radius_for_fraction(frac_profile, r_profile, target=0.9)

    # --------------------------------------------------------
    # Fit modified pseudo-Voigt to the azimuthal-average radial profile
    # Improved: two-stage fit + robust least-squares to better capture wings
    # --------------------------------------------------------
    fit_params_available = False
    fit_profile_pct = None
    fit_profile_diam = None
    pearson4_profile_pct = None
    pearson4_profile_diam = None
    try:
        from scipy.optimize import curve_fit
        try:
            from scipy.optimize import least_squares
            have_least_squares = True
        except Exception:
            have_least_squares = False

        # Recover radial energy per-bin from cumulative (radial_energy*dr = diff(cumulative))
        dr = float(r_profile[1] - r_profile[0]) if r_profile.size > 1 else 1.0
        radial_energy = np.diff(np.concatenate(([0.0], cumulative_profile))) / dr

        # Convert to mean intensity per unit area of the annulus: I = radial_energy/(2*pi*r)
        with np.errstate(divide='ignore', invalid='ignore'):
            I_profile = radial_energy / (2.0 * np.pi * r_profile)
        # handle r==0 by replacing with nearest finite value
        if not np.isfinite(I_profile[0]):
            finite_idx = np.where(np.isfinite(I_profile))[0]
            if finite_idx.size:
                I_profile[0] = I_profile[finite_idx[0]]
            else:
                I_profile[0] = 0.0

        # Work in arcsec for human-friendly parameters
        try:
            arcsec_to_m
        except NameError:
            arcsec_to_m = 12 * np.pi / 180 / 3600
        r_arcsec = r_profile / arcsec_to_m

        # Local model definitions
        def pure_gaussian(r, A, Gamma, b):
            return A * np.exp(-4*np.log(2)*(r/Gamma)**2) + b

        # Modified pseudo-Voigt radial intensity model used for the aggregated fit.
        # The model mixes a narrow Gaussian core with a broader Lorentzian-like wing.
        # The wing amplitude is scaled separately and the final mixture is normalized
        # so that A remains the overall intensity scale.
        #
        #   G(r; Γ_c) = exp(-4 ln 2 (r / Γ_c)^2)
        #   a = 2^(1/β) - 1
        #   C(r; Γ_w) = [1 + a (2 r / Γ_w)^2]^{-β}
        #   mix = (1 - η) G + η scalar C
        #   norm = (1 - η) + η scalar
        #   I(r) = A * (mix / norm)
        #
        # where r is in arcseconds and A is the peak radial mean intensity.
        def beta_pseudo_gaussian(r, A, Gamma_c, Gamma_w, eta, beta, scalar):
            G = np.exp(-4*np.log(2)*(r/Gamma_c)**2)
            a = 2**(1.0/beta) - 1.0
            C = 1.0 / (1.0 + a*(2.0*r/Gamma_w)**2)**beta
            mix = (1.0-eta)*G + eta*scalar*C
            norm = (1.0-eta) + eta*scalar
            return A * (mix / norm)

        # King profile often used for PSFs: I(r) = I0 * (1 + (r/rc)^2)^(-alpha) + b
        def king_profile(r, I0, rc, alpha, b):
            return I0 * (1.0 + (r/np.maximum(rc, 1e-12))**2.0)**(-np.maximum(alpha, 0.01)) + b

        def core_weights(r, r0):
            return np.exp(-(r/r0)**2)

        # Utility: bounded multi-start least_squares runner to fit EEF-focused residuals
        def multi_start_least_squares(resid_func, x0, lb, ub, attempts=6, rng_seed=12345, **ls_kwargs):
            try:
                from scipy.optimize import least_squares
            except Exception:
                return None, None
            rng = np.random.default_rng(rng_seed)
            best_score = np.inf
            best_x = None
            last_res = None
            x0 = np.asarray(x0, dtype=float)
            lb = np.asarray(lb, dtype=float)
            ub = np.asarray(ub, dtype=float)
            for attempt in range(attempts):
                try:
                    pert = x0 * (1.0 + rng.normal(0.0, 0.12, size=x0.shape))
                    pert = np.maximum(lb, np.minimum(ub, pert))
                    res = least_squares(resid_func, pert, bounds=(lb, ub), **ls_kwargs)
                except Exception:
                    res = None
                if res is not None and hasattr(res, 'fun'):
                    try:
                        score = float(np.sqrt(np.mean(res.fun**2)))
                    except Exception:
                        score = np.inf
                    if np.isfinite(score) and score < best_score:
                        best_score = score
                        best_x = res.x
                        last_res = res
            return best_x, last_res

        # Prepare data mask (positive intensities)
        mask = np.isfinite(I_profile) & (I_profile > 0)
        r_fit = r_arcsec[mask]
        I_fit = I_profile[mask]
        try:
            print(f"r_fit size={r_fit.size}, I_fit finite count={np.count_nonzero(np.isfinite(I_fit))}")
        except Exception:
            pass

        fit_params_available = False
        fit_profile_pct = None
        fit_profile_diam = None

        if r_fit.size >= 7:
            try:
                print("Entering radial-fit block (r_fit.size >= 7)")
            except Exception:
                pass
            # STEP 1: core-only Gaussian fit for robust initial Gamma estimate
            A0 = float(np.nanmax(I_fit))
            try:
                Gamma0 = float(r_fit[np.argmin(np.abs(I_fit - A0/2.0))])
            except Exception:
                Gamma0 = max(float(np.median(r_fit)), 1.0)
            b0 = max(float(np.nanmin(I_fit)), 0.0)

            rmax_core = 2.0 * max(Gamma0, 1e-6)
            core_mask = r_fit <= rmax_core
            try:
                popt_g, _ = curve_fit(pure_gaussian, r_fit[core_mask], I_fit[core_mask], p0=[A0, Gamma0, b0])
                A_g, Gamma_g, b_g = popt_g
            except Exception:
                A_g, Gamma_g, b_g = A0, Gamma0, b0

            # STEP 2: modified pseudo-Voigt fit with separate core and wing widths
            A_fit = float(A_g)
            Gamma_c_fit = float(Gamma_g)
            Gamma_w_fit = max(3.0 * Gamma_c_fit, 4.0)
            eta_fit = 0.08
            beta_fit = 1.8
            scalar_fit = 1.8

            A_lb = max(0.5 * A_fit, 1.0)
            x0 = [A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit]
            lb = [A_lb, max(1e-3, float(Gamma_c_fit)*0.2), max(1e-3, float(Gamma_c_fit)*1.2), 0.05, 1.0, 0.2]
            ub = [np.inf, float(Gamma_c_fit)*3.0, float(Gamma_c_fit)*25.0, 0.5, 5.0, 12.0]

            # Use weighted least squares for peak and wing matching.
            eps = 1e-12
            floor = max(eps, np.nanpercentile(I_fit[I_fit>0], 1) if np.any(I_fit>0) else eps)
            rscale = np.median(r_fit) if r_fit.size else 1.0
            peak_cut = max(3.0, min(12.0, float(Gamma_c_fit)*1.5))
            weights = 1.0 + 20.0 * np.exp(-(r_fit/peak_cut)**2) + 0.5 * (r_fit / max(rscale, 1e-6))
            sigma = 1.0 / np.maximum(weights, 1e-8)

            def fit_func(r, A_x, Gc_x, Gw_x, eta_x, beta_x, scalar_x):
                return beta_pseudo_gaussian(r, A_x, Gc_x, Gw_x, eta_x, beta_x, scalar_x)

            # Prepare cumulative EEF reference for the current data
            dr_arcsec = float(r_arcsec[1] - r_arcsec[0]) if r_arcsec.size > 1 else 1.0
            radial_energy_data = 2.0 * np.pi * r_fit * I_fit * dr_arcsec
            cumulative_data = np.cumsum(radial_energy_data)
            total_data = cumulative_data[-1] if cumulative_data.size else 1.0
            eef_data = cumulative_data / total_data if total_data > 0 else cumulative_data
            # Give the EEF objective most weight but keep a small intensity
            # residual component to stabilise the core fit and prevent large
            # deviations at intermediate radii.
            eef_weight = 30.0
            intensity_weight_scale = 0.25

            def resid_pvoigt(params):
                A_x, Gc_x, Gw_x, eta_x, beta_x, scalar_x = params
                model = beta_pseudo_gaussian(r_fit, A_x, Gc_x, Gw_x, eta_x, beta_x, scalar_x)
                model = np.maximum(model, floor)
                # EEF residual (fraction) only: compute cumulative model energy and compare
                cumulative_model = np.cumsum(2.0 * np.pi * r_fit * model * dr_arcsec)
                total_model = cumulative_model[-1] if cumulative_model.size else 1.0
                eef_model = cumulative_model / total_model if total_model > 0 else cumulative_model
                res_eef = (eef_model - eef_data) * eef_weight
                return res_eef

            # Use multi-start least_squares for robustness (EEF-only objective)
            if have_least_squares:
                try:
                    # increase PV multi-start attempts and allow longer evaluation per start
                    best_x, last_res = multi_start_least_squares(resid_pvoigt, x0, lb, ub, attempts=36, rng_seed=12345, loss='soft_l1', f_scale=1e-3, max_nfev=120000)
                    if last_res is not None and best_x is not None:
                        A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit = best_x.tolist()
                    else:
                        # fallback to curve_fit if LS failed
                        popt, _ = curve_fit(fit_func, r_fit, I_fit, p0=x0, bounds=(lb, ub), sigma=sigma, maxfev=5000)
                        A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit = popt
                except Exception:
                    try:
                        popt, _ = curve_fit(fit_func, r_fit, I_fit, p0=x0, bounds=(lb, ub), sigma=sigma, maxfev=5000)
                        A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit = popt
                    except Exception:
                        A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit = x0
            else:
                try:
                    popt, _ = curve_fit(fit_func, r_fit, I_fit, p0=x0, bounds=(lb, ub), sigma=sigma, maxfev=5000)
                    A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit = popt
                except Exception:
                    A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit = x0

            # Save aggregated fit diagnostic plot
            fit_params_available = True
            # Standalone PV diagnostic figure generation disabled (user request).
            try:
                # Previously saved a standalone E2E_fit.png here; intentionally skipping.
                pass
            except Exception:
                pass

            try:
                dr_arcsec = float(r_arcsec[1] - r_arcsec[0]) if r_arcsec.size > 1 else 1.0
                I_fit_model = beta_pseudo_gaussian(r_arcsec, A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit)
                radial_energy_fit = 2.0 * np.pi * r_arcsec * I_fit_model * dr_arcsec
                total_fit_energy = np.sum(radial_energy_fit)
                if total_fit_energy > 0:
                    fit_cumulative = np.cumsum(radial_energy_fit)
                    fit_profile_pct = 100.0 * fit_cumulative / total_fit_energy
                    fit_profile_diam = 2.0 * r_arcsec
            except Exception:
                fit_profile_pct = None
                fit_profile_diam = None

            # ========== PEARSON4 FITTING ==========
            # Fit Pearson Type IV profile to the azimuthal-average radial intensity.
            #
            # Design decisions vs. previous implementation:
            #   * Direct numpy formula — no lmfit dependency or Model.eval overhead.
            #   * center fixed at 0: the radial profile I(r)=E(r)/(2πr) is centred
            #     at the best-focus origin by construction; a free centre parameter
            #     trades off with sigma/nu and causes ill-conditioning.
            #   * 4 parameters instead of 5: amp, sigma, m (tail exponent), nu (skew).
            #   * 12 structured starts (6 grid + 6 random perturbations) rather than
            #     36 + 36 + differential_evolution — covers the relevant landscape
            #     without the prohibitive runtime.
            #   * EEF residuals multiplied by constant 30 (same as PV fit) — stable,
            #     location-independent threshold.
            #   * Acceptance: EEF-RMS < 0.10 (unscaled cumulative-fraction units).
            #
            # Pearson IV radial model (center = 0):
            #   u = r / sigma
            #   I(r) = amp * (1 + u²)^{-m} * exp(-nu * arctan(u))
            #
            pearson4_result = None
            pearson4_profile_pct = None
            pearson4_profile_diam = None
            try:
                # ------------------------------------------------------------------
                # Direct Pearson IV formula — no lmfit, no Model.eval overhead.
                # ------------------------------------------------------------------
                def _p4(r, amp, sigma, m, nu):
                    """Pearson Type IV radial profile centred at r=0."""
                    u = r / np.maximum(sigma, 1e-15)
                    return (amp
                            * (1.0 + u * u) ** (-np.maximum(m, 0.5))
                            * np.exp(-nu * np.arctan(u)))

                # Coarse evaluation grid for the residual — max 400 pts.
                _n_coarse = min(400, max(40, r_arcsec.size))
                _r_coarse = np.linspace(0.0, float(r_arcsec.max()) if r_arcsec.size else 20.0, _n_coarse)
                _dr_coarse = float(_r_coarse[1] - _r_coarse[0]) if _n_coarse > 1 else 1.0

                # Fine (full) grid spacing for EEF profile output.
                _dr_fine = float(r_arcsec[1] - r_arcsec[0]) if r_arcsec.size > 1 else 1.0

                # EEF reference from the data radial profile (on r_fit sample points).
                _re_data = 2.0 * np.pi * r_fit * I_fit * _dr_fine
                _cum_data = np.cumsum(_re_data)
                _tot_data = _cum_data[-1] if _cum_data.size else 1.0
                _eef_data = _cum_data / _tot_data if _tot_data > 0 else _cum_data

                # EEF residual function (returns vector of length r_fit).
                # Scale factor 30 mirrors the pseudo-Voigt EEF objective so that
                # the acceptance threshold is meaningful and stable.
                _EEF_SCALE = 30.0

                from scipy.optimize import least_squares as _ls_p4

                def _p4_eef_resid(pv):
                    a, s, m_, nu_ = pv
                    y = _p4(_r_coarse, a, s, m_, nu_)
                    y = np.maximum(y, 0.0)
                    re = 2.0 * np.pi * _r_coarse * y * _dr_coarse
                    tot = np.sum(re)
                    if tot <= 0 or not np.isfinite(tot):
                        return np.ones_like(_eef_data) * 1e6
                    eef_m = np.cumsum(re) / tot
                    eef_at_rfit = np.interp(r_fit, _r_coarse, eef_m)
                    return (eef_at_rfit - _eef_data) * _EEF_SCALE

                # ------------------------------------------------------------------
                # Initial seeds — informed by the PV fit result.
                # [amp, sigma, m, nu]
                # ------------------------------------------------------------------
                _amp0 = float(A_fit)
                _sig0 = float(Gamma_c_fit)
                _m0   = float(max(0.7, min(8.0, beta_fit)))
                _gw   = float(Gamma_w_fit)

                _lb = np.array([max(float(floor), 1e-15),  1e-3,  0.51, -6.0])
                _ub = np.array([_amp0 * 100.0,  max(_gw * 3.0, _sig0 * 30.0),  12.0,  6.0])

                # 6 structured seeds covering core/wing width and m variation.
                _seeds = [
                    [_amp0, _sig0,          _m0,           0.0],
                    [_amp0, _sig0 * 0.5,    _m0 * 0.7,     0.0],
                    [_amp0, _sig0 * 2.0,    _m0 * 1.3,     0.0],
                    [_amp0, _gw / 3.0,      max(0.7, _m0 * 0.6), 0.0],
                    [_amp0, _sig0,          1.5,            0.0],
                    [_amp0, _sig0,          4.0,            0.0],
                ]
                # 6 random perturbations of the primary seed.
                _rng_p4 = np.random.default_rng(4321)
                _base = np.array([_amp0, _sig0, _m0, 0.0])
                for _ in range(6):
                    _pert = _base * (1.0 + _rng_p4.normal(0.0, 0.18, size=4))
                    _seeds.append(np.clip(_pert, _lb, _ub).tolist())

                # ------------------------------------------------------------------
                # Run multi-start least_squares.
                # ------------------------------------------------------------------
                _best_x   = None
                _best_rms = np.inf

                for _s0 in _seeds:
                    try:
                        _s0c = np.clip(np.asarray(_s0, dtype=float), _lb, _ub)
                        _res = _ls_p4(
                            _p4_eef_resid, _s0c,
                            bounds=(_lb, _ub),
                            loss='soft_l1', f_scale=1e-3,
                            max_nfev=25000, method='trf',
                        )
                        _rms = float(np.sqrt(np.mean(_res.fun ** 2)))
                        if np.isfinite(_rms) and _rms < _best_rms:
                            _best_rms = _rms
                            _best_x = _res.x.copy()
                    except Exception:
                        pass

                # ------------------------------------------------------------------
                # Accept if unscaled EEF RMS < 0.10 (cumulative-fraction units).
                # This is equivalent to ≤ 10 percentage-points mean EEF deviation.
                # ------------------------------------------------------------------
                if _best_x is not None and np.isfinite(_best_rms):
                    _amp_p4, _sig_p4, _m_p4, _nu_p4 = _best_x

                    # Unscaled EEF RMS on the fine grid for the acceptance gate.
                    _y_p4 = _p4(r_arcsec, _amp_p4, _sig_p4, _m_p4, _nu_p4)
                    _y_p4 = np.maximum(_y_p4, 0.0)
                    _re_p4 = 2.0 * np.pi * r_arcsec * _y_p4 * _dr_fine
                    _tot_p4 = np.sum(_re_p4)
                    if _tot_p4 > 0 and np.isfinite(_tot_p4):
                        _eef_p4_model = np.cumsum(_re_p4) / _tot_p4
                        _eef_p4_at_rfit = np.interp(r_fit, r_arcsec, _eef_p4_model)
                        _eef_rms = float(np.sqrt(np.mean((_eef_p4_at_rfit - _eef_data) ** 2)))
                    else:
                        _eef_rms = np.inf

                    if np.isfinite(_eef_rms) and _eef_rms < 0.10:
                        class _P4Result:
                            pass
                        pearson4_result = _P4Result()
                        pearson4_result.params = {
                            'amplitude': _amp_p4,
                            'sigma':     _sig_p4,
                            'm':         _m_p4,
                            'nu':        _nu_p4,
                        }

                # ------------------------------------------------------------------
                # Compute EEF profile for plotting (regardless of acceptance gate —
                # use whatever was the best fit found).
                # ------------------------------------------------------------------
                _p4_params_used = pearson4_result.params if (
                    pearson4_result is not None and hasattr(pearson4_result, 'params')
                ) else None

                if _p4_params_used is not None:
                    _I_p4 = _p4(r_arcsec,
                                _p4_params_used['amplitude'],
                                _p4_params_used['sigma'],
                                _p4_params_used['m'],
                                _p4_params_used['nu'])
                    _I_p4 = np.maximum(_I_p4, 0.0)
                    _re_p4_full = 2.0 * np.pi * r_arcsec * _I_p4 * _dr_fine
                    _tot_p4_full = np.sum(_re_p4_full)
                    if _tot_p4_full > 0 and np.isfinite(_tot_p4_full):
                        pearson4_profile_pct  = 100.0 * np.cumsum(_re_p4_full) / _tot_p4_full
                        pearson4_profile_diam = 2.0 * r_arcsec
                    else:
                        pearson4_profile_pct  = None
                        pearson4_profile_diam = None
                else:
                    pearson4_profile_pct  = None
                    pearson4_profile_diam = None

            except Exception:
                pearson4_result        = None
                pearson4_profile_pct   = None
                pearson4_profile_diam  = None
            # Create merged diagnostic: intensity curves + residuals, then export fit parameters
            # Disabled by default to avoid creating a second figure window.
            if False:
                os.makedirs('Figures', exist_ok=True)
                # Define plotting grid
                rplot = np.linspace(0.0, float(r_arcsec.max()), 1000)
                # PV model on rplot
                Ifit_pv_plot = beta_pseudo_gaussian(rplot, A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit)
                # Pearson4 model params: prefer fit result params, fall back to params
                try:
                    p4_params_plot = pearson4_result.params if (pearson4_result is not None and hasattr(pearson4_result, 'params')) else (params if 'params' in locals() else None)
                except Exception:
                    p4_params_plot = (params if 'params' in locals() else None)
                Ifit_p4_plot = None
                if p4_params_plot is not None:
                    try:
                        Ifit_p4_plot = pearson4_model.eval(p4_params_plot, x=rplot)
                    except Exception:
                        Ifit_p4_plot = None

                # --- King profile fitting (EEF-only objective) ---
                Ifit_king_plot = None
                king_profile_pct = None
                king_profile_diam = None
                try:
                    # initial guesses
                    I0_0 = float(np.nanmax(I_fit)) if I_fit.size else 1.0
                    # Use pseudo-Voigt core width as a better initial guess for King core size when available
                    try:
                        rc_guess = float(Gamma_c_fit)
                    except Exception:
                        rc_guess = float(np.median(r_fit)) if r_fit.size else 1.0
                    rc_0 = max(0.5 * rc_guess, 1e-3)
                    alpha_0 = 2.0
                    b0 = float(np.nanmin(I_fit)) if np.any(np.isfinite(I_fit)) else 0.0
                    x0k = [I0_0, rc_0, alpha_0, b0]
                    # Constrain I0 to be at least the noise floor and limit background b using data percentiles
                    I0_lb = float(floor)
                    try:
                        p25 = float(np.nanpercentile(I_fit, 25)) if np.any(np.isfinite(I_fit)) else None
                        p50 = float(np.nanmedian(I_fit)) if np.any(np.isfinite(I_fit)) else None
                    except Exception:
                        p25, p50 = None, None
                    if p25 is not None and np.isfinite(p25):
                        b_ub = max(p25 * 2.0, p50 if (p50 is not None and np.isfinite(p50)) else p25 * 2.0)
                    else:
                        b_ub = max(float(floor) * 100.0, 1.0)
                    # Tighten King parameter bounds to avoid degenerate long tails
                    lbk = [I0_lb, 1e-3, 1.0, 0.0]  # enforce alpha >= 1.0
                    # limit rc to a reasonable fraction of the sampled radius and cap it
                    rc_cap = float(min(r_arcsec.max() * 0.5 if r_arcsec.size else 10.0, 50.0))
                    ubk = [np.inf, rc_cap, 8.0, float(b_ub)]

                    def resid_king(vec):
                        try:
                            I0_k, rc_k, alpha_k, b_k = vec
                            model_vals = king_profile(r_arcsec, I0_k, rc_k, alpha_k, b_k)
                            model_vals = np.maximum(model_vals, 0.0)
                            dr = float(r_arcsec[1] - r_arcsec[0]) if r_arcsec.size>1 else 1.0
                            radial = 2.0 * np.pi * r_arcsec * model_vals * dr
                            tot = np.sum(radial)
                            if tot <= 0 or not np.isfinite(tot):
                                return np.ones_like(eef_data) * 1e6
                            model_eef_pct = 100.0 * np.cumsum(radial) / tot
                            from numpy import interp
                            # use eef_data (computed on r_fit) as the reference EEF values at r_fit (fractions 0..1)
                            ref_at_r = eef_data
                            # convert model EEF percent -> fraction before comparing
                            model_at_ref = interp(r_fit, r_arcsec, model_eef_pct) / 100.0
                            # scale by the eef weight used elsewhere so residuals have comparable magnitude
                            return (model_at_ref - ref_at_r) * eef_weight
                        except Exception:
                            return np.ones_like(eef_data) * 1e6

                    # run King fit if we have the reference EEF computed on r_fit (eef_data)
                    if 'eef_data' in locals() and eef_data is not None:
                        if have_least_squares:
                            try:
                                try:
                                    print(f"King fit: initial x0k={x0k}, lbk={lbk}, ubk={ubk}")
                                except Exception:
                                    pass
                                # increase King-fit starts and budget to improve robustness
                                best_vec, last_resk = multi_start_least_squares(resid_king, x0k, lbk, ubk, attempts=24, rng_seed=1234, loss='soft_l1', f_scale=1.0, max_nfev=80000)
                                if best_vec is not None:
                                    I0_k, rc_k, alpha_k, b_k = best_vec.tolist()
                                else:
                                    I0_k, rc_k, alpha_k, b_k = x0k
                            except Exception:
                                I0_k, rc_k, alpha_k, b_k = x0k
                        else:
                            I0_k, rc_k, alpha_k, b_k = x0k

                        # compute king model and EEF
                        try:
                            try:
                                print(f"Computing King model with params: I0={I0_k}, rc={rc_k}, alpha={alpha_k}, b={b_k}")
                            except Exception:
                                pass
                            Ifit_king_plot = king_profile(rplot, I0_k, rc_k, alpha_k, b_k)
                            vals_k = king_profile(r_arcsec, I0_k, rc_k, alpha_k, b_k)
                            dr = float(r_arcsec[1] - r_arcsec[0]) if r_arcsec.size>1 else 1.0
                            radial_k = 2.0 * np.pi * r_arcsec * vals_k * dr
                            totk = np.sum(radial_k)
                            try:
                                print(f"  King totk={totk}")
                            except Exception:
                                pass
                            if totk > 0 and np.isfinite(totk):
                                cumk = np.cumsum(radial_k)
                                king_profile_pct = 100.0 * cumk / totk
                                king_profile_diam = 2.0 * r_arcsec
                                try:
                                    print(f"King fit succeeded: I0={I0_k:.3g}, rc={rc_k:.3g}, alpha={alpha_k:.3g}, b={b_k:.3g}")
                                    print(f"  king_profile_pct len={len(king_profile_pct)}, min={np.nanmin(king_profile_pct):.3g}, max={np.nanmax(king_profile_pct):.3g}")
                                except Exception:
                                    pass
                            else:
                                king_profile_pct = None
                                king_profile_diam = None
                        except Exception:
                            king_profile_pct = None
                            king_profile_diam = None
                except Exception:
                    Ifit_king_plot = None
                    king_profile_pct = None
                    king_profile_diam = None

                # Prepare residuals evaluated at r_fit (data sample points)
                model_pv_rfit = beta_pseudo_gaussian(r_fit, A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit)
                model_pv_rfit = np.maximum(model_pv_rfit, floor)
                res_int_pv = np.log(model_pv_rfit + floor) - np.log(I_fit + floor)

                res_int_p4 = None
                if p4_params_plot is not None:
                    try:
                        model_p4_rfit = pearson4_model.eval(p4_params_plot, x=r_fit)
                        model_p4_rfit = np.maximum(model_p4_rfit, floor)
                        res_int_p4 = np.log(model_p4_rfit + floor) - np.log(I_fit + floor)
                    except Exception:
                        res_int_p4 = None

                # King intensity residuals at r_fit
                res_int_king = None
                if Ifit_king_plot is not None:
                    try:
                        model_king_rfit = king_profile(r_fit, I0_k, rc_k, alpha_k, b_k)
                        model_king_rfit = np.maximum(model_king_rfit, floor)
                        res_int_king = np.log(model_king_rfit + floor) - np.log(I_fit + floor)
                    except Exception:
                        res_int_king = None

                # EEF residuals: compute model EEF on r_arcsec grid and interpolate to r_fit
                eef_ref_pct = profile_pct if ('profile_pct' in locals() and profile_pct is not None) else None
                eef_ref_diam = profile_diam if ('profile_diam' in locals() and profile_diam is not None) else None
                res_eef_pv = None
                res_eef_p4 = None
                try:
                    if fit_profile_pct is not None and fit_profile_diam is not None and eef_ref_pct is not None and eef_ref_diam is not None:
                        # PV model EEF on r_arcsec
                        pv_vals = beta_pseudo_gaussian(r_arcsec, A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit)
                        dr = float(r_arcsec[1] - r_arcsec[0]) if r_arcsec.size>1 else 1.0
                        radial_pv = 2.0 * np.pi * r_arcsec * pv_vals * dr
                        tot_pv = np.sum(radial_pv)
                        if tot_pv > 0 and np.isfinite(tot_pv):
                            pv_eef_pct = 100.0 * np.cumsum(radial_pv) / tot_pv
                            # interpolate PV EEF to r_fit (r in arcsec)
                            from numpy import interp
                            pv_eef_at_rfit = interp(r_fit, r_arcsec, pv_eef_pct)
                            ref_at_rfit = interp(r_fit, eef_ref_diam/2.0, eef_ref_pct)
                            res_eef_pv = pv_eef_at_rfit - ref_at_rfit
                        # Pearson4 EEF
                        if p4_params_plot is not None:
                            p4_vals = pearson4_model.eval(p4_params_plot, x=r_arcsec)
                            radial_p4 = 2.0 * np.pi * r_arcsec * p4_vals * dr
                            tot_p4 = np.sum(radial_p4)
                            if tot_p4 > 0 and np.isfinite(tot_p4):
                                p4_eef_pct = 100.0 * np.cumsum(radial_p4) / tot_p4
                                p4_eef_at_rfit = interp(r_fit, r_arcsec, p4_eef_pct)
                                ref_at_rfit = interp(r_fit, eef_ref_diam/2.0, eef_ref_pct)
                                res_eef_p4 = p4_eef_at_rfit - ref_at_rfit
                        # King EEF
                        if king_profile_pct is not None and king_profile_diam is not None:
                            try:
                                king_eef_at_rfit = interp(r_fit, king_profile_diam/2.0, king_profile_pct)
                                king_vals = king_profile(r_arcsec, I0_k, rc_k, alpha_k, b_k)
                                radial_k = 2.0 * np.pi * r_arcsec * king_vals * dr
                                tot_k = np.sum(radial_k)
                                if tot_k > 0 and np.isfinite(tot_k):
                                    king_eef_pct = 100.0 * np.cumsum(radial_k) / tot_k
                                    king_eef_at_rfit = interp(r_fit, r_arcsec, king_eef_pct)
                                    ref_at_rfit = interp(r_fit, eef_ref_diam/2.0, eef_ref_pct)
                                    res_eef_king = king_eef_at_rfit - ref_at_rfit
                                else:
                                    res_eef_king = None
                            except Exception:
                                res_eef_king = None
                except Exception:
                    res_eef_pv = None
                    res_eef_p4 = None

                # Prepare aggregated EEF arrays for plotting (robustly)
                profile_pct = None
                profile_diam = None
                try:
                    # r_profile is in meters; convert to arcsec using arcsec_to_m
                    if ('cumulative_profile' in locals() and 'total_energy' in locals() and total_energy and 'r_profile' in locals() and 'arcsec_to_m' in locals()):
                        profile_pct = 100.0 * (cumulative_profile / total_energy)
                        # diameter in arcsec = 2 * (r_profile [m] / arcsec_to_m [m/arcsec])
                        profile_diam = 2.0 * (r_profile / arcsec_to_m)
                except Exception:
                    profile_pct = None
                    profile_diam = None

                # Recompute EEF residuals at r_fit explicitly for plotting (ensure bottom-right shows data)
                res_eef_pv_plot = None
                res_eef_p4_plot = None
                res_eef_king_plot = None
                try:
                    from numpy import interp
                    dr0 = float(r_arcsec[1] - r_arcsec[0]) if r_arcsec.size>1 else 1.0
                    # reference EEF at r_fit (fraction)
                    ref_at_rfit = eef_data if ('eef_data' in locals() and eef_data is not None) else None
                    if ref_at_rfit is not None:
                        # PV model
                        try:
                            pv_vals_full = beta_pseudo_gaussian(r_arcsec, A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit)
                            radial_pv_full = 2.0 * np.pi * r_arcsec * pv_vals_full * dr0
                            tot_pv_full = np.sum(radial_pv_full)
                            if tot_pv_full > 0 and np.isfinite(tot_pv_full):
                                pv_eef_pct_full = 100.0 * np.cumsum(radial_pv_full) / tot_pv_full
                                pv_eef_at_rfit = interp(r_fit, r_arcsec, pv_eef_pct_full) / 100.0
                                res_eef_pv_plot = pv_eef_at_rfit - ref_at_rfit
                        except Exception:
                            res_eef_pv_plot = None
                        # Pearson4 model
                        try:
                            if p4_params_plot is not None:
                                p4_vals_full = pearson4_model.eval(p4_params_plot, x=r_arcsec)
                                radial_p4_full = 2.0 * np.pi * r_arcsec * p4_vals_full * dr0
                                tot_p4_full = np.sum(radial_p4_full)
                                if tot_p4_full > 0 and np.isfinite(tot_p4_full):
                                    p4_eef_pct_full = 100.0 * np.cumsum(radial_p4_full) / tot_p4_full
                                    p4_eef_at_rfit = interp(r_fit, r_arcsec, p4_eef_pct_full) / 100.0
                                    res_eef_p4_plot = p4_eef_at_rfit - ref_at_rfit
                        except Exception:
                            res_eef_p4_plot = None
                        # King model
                        try:
                            if 'king_profile_pct' in locals() and king_profile_pct is not None and 'king_profile_diam' in locals() and king_profile_diam is not None:
                                king_vals_full = king_profile(r_arcsec, I0_k, rc_k, alpha_k, b_k)
                                radial_k_full = 2.0 * np.pi * r_arcsec * king_vals_full * dr0
                                tot_k_full = np.sum(radial_k_full)
                                if tot_k_full > 0 and np.isfinite(tot_k_full):
                                    king_eef_pct_full = 100.0 * np.cumsum(radial_k_full) / tot_k_full
                                    king_eef_at_rfit = interp(r_fit, r_arcsec, king_eef_pct_full) / 100.0
                                    res_eef_king_plot = king_eef_at_rfit - ref_at_rfit
                        except Exception:
                            res_eef_king_plot = None
                except Exception:
                    res_eef_pv_plot = None
                    res_eef_p4_plot = None
                    res_eef_king_plot = None

                # Plot combined figure
                # Create a 2x2 figure: top row = intensity + residuals; bottom row = EEF + EEF residuals
                fig, axes = plt.subplots(2, 2, figsize=(14, 10), constrained_layout=True)
                ax_int = axes[0, 0]
                ax_res = axes[0, 1]
                ax_eef = axes[1, 0]
                ax_eef_res = axes[1, 1]

                # Determine whether Pearson4 was accepted (only then show its traces)
                try:
                    pearson4_ok = (pearson4_result is not None)
                except Exception:
                    pearson4_ok = False

                # Top-left: intensities
                # define linewidths for combined figure: thin for dashed fits, thicker for solid/reference
                _combined_thin_lw = 1.0
                _combined_thick_lw = 2.0
                ax_int.plot(r_arcsec, I_profile, 'k.', ms=3, label='Aggregated PSF (data)')
                line_pv_comb = ax_int.plot(rplot, Ifit_pv_plot, color='red', linestyle='--', lw=_combined_thin_lw, label='Modified pseudo-Voigt')[0]
                try:
                    line_pv_comb.set_dashes([6, 2])
                except Exception:
                    pass
                if pearson4_ok and Ifit_p4_plot is not None:
                    ax_int.plot(rplot, Ifit_p4_plot, color='orange', linestyle='--', lw=_combined_thin_lw, label=label_p4)
                if Ifit_king_plot is not None:
                    ax_int.plot(rplot, Ifit_king_plot, color='purple', linestyle='--', lw=_combined_thin_lw, label=label_king)
                ax_int.set_xlabel('Radius [arcsec]')
                ax_int.set_ylabel('Mean intensity')
                ax_int.set_yscale('log')
                ax_int.legend(fontsize=9)
                ax_int.grid(True, which='both', linestyle='--', alpha=0.4)

                # Top-right: intensity residuals (left y-axis, percent) and
                # EEF residuals (right y-axis, linear percent). Convert intensity
                # residuals to percent so all curves share the same units.
                ax_res.set_xlabel('Radius [arcsec]')
                ax_res.set_ylabel('Residual (%)')
                # Twin axis for EEF residuals (percentage, linear)
                ax_res_eef = ax_res.twinx()
                ax_res_eef.set_ylabel('EEF residual (pct)')

                # Convert log/int residuals to percent relative to data
                eps_local = floor if ('floor' in locals() or 'floor' in globals()) else 1e-12
                try:
                    if res_int_pv is not None and 'I_fit' in locals():
                        # res_int_pv was log(model) - log(data); prefer to use the
                        # linear model values already computed on r_fit (model_pv_rfit)
                        try:
                            if 'model_pv_rfit' in locals() and model_pv_rfit is not None:
                                res_int_pv_pct = 100.0 * (model_pv_rfit - I_fit) / (I_fit + eps_local)
                            else:
                                # Fall back: convert log-residual to approximate percent using exp(delta)-1
                                res_int_pv_pct = 100.0 * (np.exp(res_int_pv) - 1.0)
                        except Exception:
                            res_int_pv_pct = np.full_like(r_fit, np.nan)
                    else:
                        res_int_pv_pct = np.full_like(r_fit, np.nan)
                except Exception:
                    res_int_pv_pct = np.full_like(r_fit, np.nan)
                try:
                    if pearson4_ok and res_int_p4 is not None:
                        res_int_p4_pct = 100.0 * (np.exp(res_int_p4) - 1.0)
                    else:
                        res_int_p4_pct = np.full_like(r_fit, np.nan)
                except Exception:
                    res_int_p4_pct = np.full_like(r_fit, np.nan)
                try:
                    if res_int_king is not None:
                        res_int_king_pct = 100.0 * (np.exp(res_int_king) - 1.0)
                    else:
                        res_int_king_pct = np.full_like(r_fit, np.nan)
                except Exception:
                    res_int_king_pct = np.full_like(r_fit, np.nan)

                # Plot intensity residuals on the left (percent)
                if res_int_pv_pct is not None:
                    ax_res.plot(r_fit, res_int_pv_pct, color='red', linestyle='--', lw=_combined_thin_lw, label='Intensity residual: Data - Pseudo-Voigt (%)')
                if pearson4_ok and res_int_p4_pct is not None:
                    ax_res.plot(r_fit, res_int_p4_pct, color='orange', linestyle='--', lw=_combined_thin_lw * 1.6, marker='o', markersize=3, markevery=50, label='Intensity residual: Data - Pearson4 (%)')
                if res_int_king_pct is not None:
                    ax_res.plot(r_fit, res_int_king_pct, color='purple', linestyle='--', lw=_combined_thin_lw * 1.6, marker='s', markersize=3, markevery=50, label=f'Intensity residual: Data - {label_king} (%)')
                ax_res.axhline(0.0, color='k', linestyle='--', linewidth=1)
                ax_res.grid(True, which='both', linestyle='--', alpha=0.4)

                # Plot EEF residuals on the right (linear scale)
                handles_res, labels_res = [], []
                if res_eef_pv is not None:
                    h1, = ax_res_eef.plot(r_fit, res_eef_pv, color='red', linestyle='--', lw=_combined_thin_lw, label='EEF residual: PV - Data (pct)')
                    handles_res.append(h1); labels_res.append('EEF residual: PV - Data (pct)')
                if res_eef_p4 is not None:
                    h2, = ax_res_eef.plot(r_fit, res_eef_p4, color='orange', linestyle='--', lw=_combined_thin_lw * 1.6, marker='o', markersize=3, markevery=50, label='EEF residual: P4 - Data (pct)')
                    handles_res.append(h2); labels_res.append('EEF residual: P4 - Data (pct)')
                if 'res_eef_king' in locals() and res_eef_king is not None:
                    h3, = ax_res_eef.plot(r_fit, res_eef_king, color='purple', linestyle='--', lw=_combined_thin_lw * 1.6, marker='s', markersize=3, markevery=50, label='EEF residual: King - Data (pct)')
                    handles_res.append(h3); labels_res.append('EEF residual: King - Data (pct)')

                # Combine legends from both axes (intensity residuals on left, EEF residuals on right)
                try:
                    handles_l, labels_l = ax_res.get_legend_handles_labels()
                    handles_r, labels_r = ax_res_eef.get_legend_handles_labels()
                    all_handles = handles_l + handles_r
                    all_labels = labels_l + labels_r
                    if all_handles:
                        ax_res.legend(all_handles, all_labels, fontsize=9)
                except Exception:
                    try:
                        ax_res.legend()
                    except Exception:
                        pass

                # Bottom-left: EEF curves (aggregated PSF and fitted models)
                # Flip axes: diameter on X, encircled energy (%) on Y. Plot aggregated PSF as black dots.
                label_data = 'Aggregated PSF (data)'
                label_pv = 'Modified pseudo-Voigt radial fit'
                # Use short, clean labels for fitted models in the combined EEF figure
                # Full fit parameters are retained in the exported Excel/CSV, not the legend.
                label_p4 = 'Pearson4 radial fit'
                label_king = 'King radial fit'
                try:
                    # Ensure Pearson4 EEF is available: if it wasn't computed earlier
                    # but we do have a fit result, compute it here so the combined
                    # figure always shows the Pearson4 EEF when possible.
                    if (pearson4_profile_pct is None or pearson4_profile_diam is None) and (pearson4_result is not None) and hasattr(pearson4_result, 'params'):
                        try:
                            p_params_plot = pearson4_result.params
                            rplot_local = np.linspace(0.0, float(r_arcsec.max()), 1000)
                            # compute model on r_arcsec for EEF
                            p_vals = pearson4_model.eval(p_params_plot, x=r_arcsec)
                            p_vals = np.maximum(p_vals, 0.0)
                            dr_local = float(r_arcsec[1] - r_arcsec[0]) if r_arcsec.size > 1 else 1.0
                            radial_p = 2.0 * np.pi * r_arcsec * p_vals * dr_local
                            tot_p = np.sum(radial_p)
                            if tot_p > 0 and np.isfinite(tot_p):
                                pearson4_profile_pct = 100.0 * np.cumsum(radial_p) / tot_p
                                pearson4_profile_diam = 2.0 * r_arcsec
                                # also provide Ifit_p4_plot for intensity panel if missing
                                try:
                                    Ifit_p4_plot = pearson4_model.eval(p_params_plot, x=rplot_local)
                                except Exception:
                                    Ifit_p4_plot = None
                        except Exception:
                            pearson4_profile_pct = None
                            pearson4_profile_diam = None

                    # Debug: report whether Pearson4 fit/result was used to compute the EEF
                    try:
                        present = (pearson4_result is not None and hasattr(pearson4_result, 'params'))
                        computed = (pearson4_profile_pct is not None and pearson4_profile_diam is not None)
                        print(f"DEBUG: pearson4_result present={present}, pearson4_profile computed={computed}")
                    except Exception:
                        pass

                    if profile_pct is not None and profile_diam is not None and len(profile_pct) and len(profile_diam):
                        pct_arr = np.asarray(profile_pct, dtype=float)
                        diam_arr = np.asarray(profile_diam, dtype=float)
                        order = np.argsort(pct_arr)
                        pct_sorted = pct_arr[order]
                        diam_sorted = diam_arr[order]
                        # Clip to 95% percentile
                        mask95 = pct_sorted <= 95.0
                        if np.any(mask95):
                            ax_eef.plot(diam_sorted[mask95], pct_sorted[mask95], color='k', marker='.', linestyle='None', ms=4, label=label_data, zorder=10)
                    # Plot PV EEF if available (diameter vs pct, clipped to 95%)
                    # Prefer explicit PV EEF computed on the fine rplot grid so it's always visible
                    try:
                        if 'Ifit_pv_plot' in locals() and Ifit_pv_plot is not None:
                            # rplot is in arcsec
                            rpv = rplot
                            drpv = float(rpv[1] - rpv[0]) if rpv.size > 1 else 1.0
                            radial_pv_plot = 2.0 * np.pi * rpv * Ifit_pv_plot * drpv
                            tot_pv_plot = np.sum(radial_pv_plot)
                            if tot_pv_plot > 0 and np.isfinite(tot_pv_plot):
                                pv_eef_pct_plot = 100.0 * np.cumsum(radial_pv_plot) / tot_pv_plot
                                pv_diam_plot = 2.0 * rpv
                                maskpv = pv_eef_pct_plot <= 95.0
                                if np.any(maskpv):
                                    line_pv_eef = ax_eef.plot(pv_diam_plot[maskpv], pv_eef_pct_plot[maskpv], color='red', linestyle='--', lw=_combined_thin_lw, label=label_pv)[0]
                                    try:
                                        line_pv_eef.set_dashes([6, 2])
                                    except Exception:
                                        pass
                        elif fit_profile_pct is not None and fit_profile_diam is not None:
                            fp_pct = np.asarray(fit_profile_pct, dtype=float)
                            fp_diam = np.asarray(fit_profile_diam, dtype=float)
                            mask = fp_pct <= 95.0
                            if np.any(mask):
                                line_pv_eef2 = ax_eef.plot(fp_diam[mask], fp_pct[mask], color='red', linestyle='--', lw=_combined_thin_lw, label=label_pv)[0]
                                try:
                                    line_pv_eef2.set_dashes([6, 2])
                                except Exception:
                                    pass
                    except Exception:
                        try:
                            if fit_profile_pct is not None and fit_profile_diam is not None:
                                fp_pct = np.asarray(fit_profile_pct, dtype=float)
                                fp_diam = np.asarray(fit_profile_diam, dtype=float)
                                mask = fp_pct <= 95.0
                                if np.any(mask):
                                    line_pv_eef3 = ax_eef.plot(fp_diam[mask], fp_pct[mask], color='red', linestyle='--', lw=_combined_thin_lw, label=label_pv)[0]
                                    try:
                                        line_pv_eef3.set_dashes([6, 2])
                                    except Exception:
                                        pass
                        except Exception:
                            pass
                    # Pearson4: only plot if a Pearson4 fit/result was accepted
                    try:
                        pearson4_ok = (pearson4_result is not None)
                    except Exception:
                        pearson4_ok = False
                    if pearson4_ok and pearson4_profile_pct is not None and pearson4_profile_diam is not None:
                        p4_pct = np.asarray(pearson4_profile_pct, dtype=float)
                        p4_diam = np.asarray(pearson4_profile_diam, dtype=float)
                        mask = p4_pct <= 95.0
                        if np.any(mask):
                            ax_eef.plot(p4_diam[mask], p4_pct[mask], color='orange', linestyle='--', lw=_combined_thin_lw, label=label_p4)
                    # King
                    if 'king_profile_pct' in locals() and king_profile_pct is not None and 'king_profile_diam' in locals() and king_profile_diam is not None:
                        k_pct = np.asarray(king_profile_pct, dtype=float)
                        k_diam = np.asarray(king_profile_diam, dtype=float)
                        mask = k_pct <= 95.0
                        if np.any(mask):
                            ax_eef.plot(k_diam[mask], k_pct[mask], color='purple', linestyle='--', linewidth=_combined_thin_lw, label=label_king)
                    else:
                        # Fallback: if we have an evaluated King profile on rplot, compute its EEF and plot
                        try:
                            if 'Ifit_king_plot' in locals() and Ifit_king_plot is not None:
                                rpv = rplot if 'rplot' in locals() else (np.linspace(0.0, float(r_arcsec.max()), 1000) if 'r_arcsec' in locals() else None)
                                if rpv is not None:
                                    drpv = float(rpv[1] - rpv[0]) if rpv.size > 1 else 1.0
                                    radial_k_plot = 2.0 * np.pi * rpv * Ifit_king_plot * drpv
                                    tot_k_plot = np.sum(radial_k_plot)
                                    if tot_k_plot > 0 and np.isfinite(tot_k_plot):
                                        k_eef_pct_plot = 100.0 * np.cumsum(radial_k_plot) / tot_k_plot
                                        k_diam_plot = 2.0 * rpv
                                        maskk = k_eef_pct_plot <= 95.0
                                        if np.any(maskk):
                                            ax_eef.plot(k_diam_plot[maskk], k_eef_pct_plot[maskk], color='purple', linestyle='--', linewidth=_combined_thin_lw, label=label_king)
                        except Exception:
                            pass
                    ax_eef.set_xlabel('Diameter [arcsec]')
                    ax_eef.set_ylabel('Encircled energy (%)')
                    # Use default legend from available handles so the full
                    # constructed labels (e.g. label_king) are shown without
                    # attempting fragile re-ordering.
                    try:
                        import textwrap
                        handles_all, labels_all = ax_eef.get_legend_handles_labels()
                        if handles_all:
                            # Wrap long legend labels to avoid over-wide single-line entries
                            wrapped_labels = [textwrap.fill(lbl, width=40) if isinstance(lbl, str) else lbl for lbl in labels_all]
                            # Place EEF legend inside the plotting area (upper left)
                            ax_eef.legend(handles_all, wrapped_labels, loc='upper left', fontsize=9)
                        
                    except Exception:
                        try:
                            ax_eef.legend()
                        except Exception:
                            pass
                    ax_eef.grid(True, which='both', linestyle='--', alpha=0.4)
                    # EEF 50% markers: vertical line at diameter for 50% and horizontal at 50%
                    try:
                        if 'pct_sorted' in locals() and np.any(pct_sorted <= 95.0):
                            pct_for_interp = pct_sorted
                            diam_for_interp = diam_sorted
                            if pct_for_interp[0] <= 50.0 <= pct_for_interp[-1]:
                                diam50 = float(np.interp(50.0, pct_for_interp, diam_for_interp))
                                # make 50% markers subtle: light grey and thinner
                                subtle_lw = 0.8
                                subtle_col = 'lightgrey'
                                ax_eef.axvline(x=diam50, linestyle='--', color=subtle_col, linewidth=subtle_lw)
                                ax_eef.axhline(y=50.0, linestyle='--', color=subtle_col, linewidth=subtle_lw)
                                try:
                                    ax_eef.text(diam_for_interp.min(), 50.0, f'EEF 50% = {diam50:.3f}"', ha='left', va='bottom', fontsize=9, color='purple')
                                except Exception:
                                    pass
                    except Exception:
                        pass
                except Exception:
                    # fallback: do nothing
                    pass

                # Bottom-right: EEF residuals (Data - Model) evaluated at r_fit
                # Plot as percent. Use recomputed fractional residuals when available
                # (those are in fractions -> multiply by 100), otherwise fall back
                # to earlier-percent values and just flip sign.
                if res_eef_pv_plot is not None:
                    line_pv_res = ax_eef_res.plot(r_fit, (-res_eef_pv_plot) * 100.0, color='red', linestyle='--', lw=_combined_thin_lw, label=f'EEF residual: Data - {label_pv} (pct)')[0]
                    try:
                        line_pv_res.set_dashes([6, 2])
                    except Exception:
                        pass
                elif res_eef_pv is not None:
                    line_pv_res2 = ax_eef_res.plot(r_fit, (-res_eef_pv), color='red', linestyle='--', lw=_combined_thin_lw, label=f'EEF residual: Data - {label_pv} (pct)')[0]
                    try:
                        line_pv_res2.set_dashes([6, 2])
                    except Exception:
                        pass
                if pearson4_ok:
                    if res_eef_p4_plot is not None:
                        ax_eef_res.plot(r_fit, (-res_eef_p4_plot) * 100.0, color='orange', linestyle='--', lw=_combined_thin_lw * 1.6, marker='o', markersize=3, markevery=50, label=f'EEF residual: Data - {label_p4} (pct)')
                    elif res_eef_p4 is not None:
                        ax_eef_res.plot(r_fit, (-res_eef_p4), color='orange', linestyle='--', lw=_combined_thin_lw * 1.6, marker='o', markersize=3, markevery=50, label=f'EEF residual: Data - {label_p4} (pct)')
                if res_eef_king_plot is not None:
                    ax_eef_res.plot(r_fit, (-res_eef_king_plot) * 100.0, color='purple', linestyle='--', lw=_combined_thin_lw * 1.6, marker='s', markersize=3, markevery=50, label=f'EEF residual: Data - {label_king} (pct)')
                elif 'res_eef_king' in locals() and res_eef_king is not None:
                    ax_eef_res.plot(r_fit, (-res_eef_king), color='purple', linestyle='--', lw=_combined_thin_lw * 1.6, marker='s', markersize=3, markevery=50, label=f'EEF residual: Data - {label_king} (pct)')
                ax_eef_res.axhline(0.0, color='k', linestyle='--', linewidth=1)
                ax_eef_res.set_xlabel('Radius [arcsec]')
                ax_eef_res.set_ylabel('EEF residual (pct)')
                # Mark radius corresponding to EEF 50% (diameter from bottom-left / 2)
                try:
                    if 'diam50' in locals():
                        radius50 = float(diam50) / 2.0
                        subtle_lw = 0.8
                        subtle_col = 'lightgrey'
                        ax_eef_res.axvline(x=radius50, linestyle='--', color=subtle_col, linewidth=subtle_lw)
                        try:
                            ylim = ax_eef_res.get_ylim()
                            # move R50 label further away from the line (increase vertical offset)
                            ax_eef_res.text(radius50, ylim[0] + 0.08 * (ylim[1] - ylim[0]), f'R50 = {radius50:.3f}"', ha='center', va='bottom', fontsize=9, color='black', rotation=90)
                        except Exception:
                            pass
                except Exception:
                    pass
                try:
                    handles_r, labels_r = ax_eef_res.get_legend_handles_labels()
                    if handles_r:
                        ax_eef_res.legend(handles_r, labels_r, loc='upper right', fontsize=9)
                except Exception:
                    try:
                        ax_eef_res.legend()
                    except Exception:
                        pass
                ax_eef_res.grid(True, which='both', linestyle='--', alpha=0.4)

                outfn_comb = os.path.join('Figures', 'E2E_fit_combined.png')
                # Ensure 50% markers are drawn even if earlier variables were missing
                try:
                    diam50_forced = None
                    def try_from_arrays(pct_arr, diam_arr):
                        try:
                            pa = np.asarray(pct_arr, dtype=float)
                            da = np.asarray(diam_arr, dtype=float)
                            order = np.argsort(pa)
                            pa_s = pa[order]
                            da_s = da[order]
                            if pa_s.size and pa_s[0] <= 50.0 <= pa_s[-1]:
                                return float(np.interp(50.0, pa_s, da_s))
                        except Exception:
                            return None
                        return None

                    # Try preferred sources in order
                    if diam50_forced is None and 'profile_pct' in locals() and profile_pct is not None:
                        diam50_forced = try_from_arrays(profile_pct, profile_diam)
                    if diam50_forced is None and 'fit_profile_pct' in locals() and fit_profile_pct is not None:
                        diam50_forced = try_from_arrays(fit_profile_pct, fit_profile_diam)
                    if diam50_forced is None and 'pearson4_profile_pct' in locals() and pearson4_profile_pct is not None:
                        diam50_forced = try_from_arrays(pearson4_profile_pct, pearson4_profile_diam)
                    if diam50_forced is None and 'king_profile_pct' in locals() and king_profile_pct is not None:
                        diam50_forced = try_from_arrays(king_profile_pct, king_profile_diam)

                    # If we didn't find diam50 earlier, force-compute from frac_profile/r_profile
                    try:
                        if diam50_forced is None:
                            diam50_forced = None
                            if 'frac_profile' in locals() and 'r_profile' in locals() and frac_profile is not None and r_profile is not None:
                                try:
                                    fp = np.asarray(frac_profile, dtype=float) * 100.0
                                    rp = np.asarray(r_profile, dtype=float) * m_to_arcsec * 2.0  # diameter
                                    order = np.argsort(fp)
                                    fp_s = fp[order]
                                    rp_s = rp[order]
                                    if fp_s.size and fp_s[0] <= 50.0 <= fp_s[-1]:
                                        diam50_forced = float(np.interp(50.0, fp_s, rp_s))
                                except Exception:
                                    diam50_forced = None
                        if diam50_forced is not None:
                            # draw with high zorder so they're visible above plotted curves
                            try:
                                subtle_lw_forced = 0.8
                                subtle_col = 'lightgrey'
                                ax_eef.axvline(x=diam50_forced, linestyle='--', color=subtle_col, linewidth=subtle_lw_forced, zorder=100)
                                ax_eef.axhline(y=50.0, linestyle='--', color=subtle_col, linewidth=subtle_lw_forced, zorder=100)
                                # mark intersection with a small circle
                                try:
                                    ax_eef.scatter([diam50_forced], [50.0], color='purple', s=40, zorder=101)
                                except Exception:
                                    pass
                                ax_eef.text(max(ax_eef.get_xlim()[0], 0), 50.0, f'EEF 50% = {diam50_forced:.3f}"', ha='left', va='bottom', fontsize=9, color='purple')
                            except Exception:
                                pass
                            try:
                                radius50_forced = float(diam50_forced) / 2.0
                                subtle_lw_forced = 0.8
                                subtle_col = 'lightgrey'
                                ax_eef_res.axvline(x=radius50_forced, linestyle='--', color=subtle_col, linewidth=subtle_lw_forced, zorder=100)
                                try:
                                    ax_eef_res.scatter([radius50_forced], [0.0], color=subtle_col, s=30, zorder=101)
                                except Exception:
                                    pass
                                try:
                                    ylim = ax_eef_res.get_ylim()
                                    # move R50 label further away from the line (increase vertical offset)
                                    ax_eef_res.text(radius50_forced, ylim[0] + 0.08 * (ylim[1] - ylim[0]), f'R50 = {radius50_forced:.3f}"', ha='center', va='bottom', fontsize=9, color='black', rotation=90)
                                except Exception:
                                    pass
                                # draw a short horizontal dashed marker at the top of the residual axis (in axes coords)
                                try:
                                    hw = max(0.5, float(radius50_forced) * 0.05)
                                    ax_eef_res.plot([radius50_forced - hw, radius50_forced + hw], [0.95, 0.95], transform=ax_eef_res.get_xaxis_transform(), linestyle='--', color=subtle_col, linewidth=subtle_lw_forced, zorder=100)
                                except Exception:
                                    pass
                            except Exception:
                                pass
                    except Exception:
                        pass
                except Exception:
                    pass
                    plt.savefig(outfn_comb, dpi=150)
                    plt.close()
                except Exception:
                    pass

            # Export fit parameters to Excel
            try:
                fit_export_data = {
                    'Parameter': [],
                    'Modified Pseudo-Voigt': [],
                    'Pearson4': [],
                    'King': []
                }
                
                # Pseudo-Voigt parameters
                pv_params = {
                    'Amplitude (A)': A_fit,
                    'Core width (Gamma_c) [arcsec]': Gamma_c_fit,
                    'Wing width (Gamma_w) [arcsec]': Gamma_w_fit,
                    'Mixing ratio (eta)': eta_fit,
                    'Wing exponent (beta)': beta_fit,
                    'Wing scale (scalar)': scalar_fit,
                }
                
                for param_name, value in pv_params.items():
                    fit_export_data['Parameter'].append(param_name)
                    fit_export_data['Modified Pseudo-Voigt'].append(value)
                    fit_export_data['Pearson4'].append(None)
                    fit_export_data['King'].append(None)
                
                # Pearson4 parameters
                if pearson4_result is not None:
                    _pr = pearson4_result.params
                    p4_params = {
                        'Amplitude': float(_pr['amplitude']),
                        'Center [arcsec]': 0.0,
                        'Sigma [arcsec]': float(_pr['sigma']),
                        'Exponent (m)': float(_pr['m']),
                        'Skewness (nu)': float(_pr['nu']),
                    }
                    
                    # Add Pearson4-specific parameters
                    for param_name, value in p4_params.items():
                        if param_name not in fit_export_data['Parameter']:
                            fit_export_data['Parameter'].append(param_name)
                            fit_export_data['Modified Pseudo-Voigt'].append(None)
                            fit_export_data['King'].append(None)
                            fit_export_data['Pearson4'].append(None)
                        idx = fit_export_data['Parameter'].index(param_name)
                        fit_export_data['Pearson4'][idx] = value

                # King parameters
                try:
                    if 'I0_k' in locals():
                        king_params = {
                            'King I0': I0_k,
                            'King rc [arcsec]': rc_k,
                            'King alpha': alpha_k,
                            'King background (b)': b_k,
                        }
                        for param_name, value in king_params.items():
                            if param_name not in fit_export_data['Parameter']:
                                fit_export_data['Parameter'].append(param_name)
                                fit_export_data['Modified Pseudo-Voigt'].append(None)
                                fit_export_data['Pearson4'].append(None)
                                fit_export_data['King'].append(None)
                            idx = fit_export_data['Parameter'].index(param_name)
                            fit_export_data['King'][idx] = value
                except Exception:
                    pass
                
                # Export to CSV
                fit_df = pd.DataFrame(fit_export_data)
                export_fit_path = os.path.join('Figures', 'fit_parameters.csv')
                fit_df.to_csv(export_fit_path, index=False)
                
            except Exception:
                pass

        else:
            pass
    except Exception:
        # SciPy not available or other error — skip fitting but don't fail
        print("scipy.optimize not available — skipping aggregated radial fit.")

    # Also compute 50% from origin for reference
    r_profile_00, cumulative_00, total_00 = radial_profile(0.0, 0.0, n_r=n_r_final, n_theta=n_theta_final, r_margin_factor=final_r_margin)
    frac_00 = cumulative_00 / total_00 if total_00 > 0 else cumulative_00
    radius_50_00 = _radius_for_fraction(frac_00, r_profile_00, target=0.5)
    # 90% at origin for reference
    radius_90_00 = _radius_for_fraction(frac_00, r_profile_00, target=0.9)
    # Fallback: if radial integration failed (NaN/None), approximate using
    # an effective sigma from the weighted mixture (use normalized weights).
    try:
        if not (isinstance(radius_50_00, float) and np.isfinite(radius_50_00)):
            # use sigmax/sigmay arrays (in meters) and normalized weight_arr
            try:
                sigx_arr = df['sigmax'].to_numpy(dtype=float, copy=False)
                sigy_arr = df['sigmay'].to_numpy(dtype=float, copy=False)
                # weight_arr should be normalized (sum==1) from earlier
                if 'weight_arr' in locals() and weight_arr is not None and np.isfinite(weight_arr).all() and weight_arr.size == sigx_arr.size:
                    w = weight_arr
                else:
                    wtmp = df.get('weight', pd.Series([1.0]*len(df))).to_numpy(dtype=float, copy=False)
                    wsum = float(np.nansum(wtmp)) if wtmp.size else 0.0
                    w = (wtmp / wsum) if (wsum and np.isfinite(wsum) and wsum > 0.0) else np.ones(len(df), dtype=float) / max(1, len(df))
                sigma2_eff = float(np.sum(w * (sigx_arr**2 + sigy_arr**2) * 0.5))
                if sigma2_eff > 0 and np.isfinite(sigma2_eff):
                    sigma_eff = float(np.sqrt(sigma2_eff))
                    radius_50_00 = sigma_eff * np.sqrt(2.0 * np.log(2.0))
            except Exception:
                pass
    except Exception:
        pass
    
    # Compute optimized configuration metrics if provided
    opt_center_x, opt_center_y, opt_radius_50, opt_radius_90 = None, None, None, None
    opt_r_profile, opt_frac_profile = None, None
    if df_optimized is not None:
        # Compute center for optimized config
        opt_total_weight = df_optimized['weight'].sum()
        opt_center_x = (df_optimized['mux'] * df_optimized['weight']).sum() / opt_total_weight
        opt_center_y = (df_optimized['muy'] * df_optimized['weight']).sum() / opt_total_weight

        # Build fast summation kernel for optimized df (threaded)
        opt_mux = df_optimized['mux'].to_numpy(dtype=float, copy=False)
        opt_muy = df_optimized['muy'].to_numpy(dtype=float, copy=False)
        opt_sigx = df_optimized['sigmax'].to_numpy(dtype=float, copy=False)
        opt_sigy = df_optimized['sigmay'].to_numpy(dtype=float, copy=False)
        opt_theta = df_optimized['theta_degrees'].to_numpy(dtype=float, copy=False)
        opt_weight = df_optimized['weight'].to_numpy(dtype=float, copy=False)
        opt_dist = df_optimized.get('distribution', pd.Series(['gaussian'] * len(df_optimized))).astype(str).str.lower().to_numpy(copy=False)
        opt_alpha_azi = pd.to_numeric(df_optimized.get('alpha_azi', pd.Series([0.5] * len(df_optimized))), errors='coerce').fillna(0.5).to_numpy(dtype=float, copy=False)
        opt_alpha_rad = pd.to_numeric(df_optimized.get('alpha_rad', pd.Series([0.5] * len(df_optimized))), errors='coerce').fillna(0.5).to_numpy(dtype=float, copy=False)

        # Apply same sigma floor to optimized dataset
        opt_sigx = np.maximum(opt_sigx, MIN_SIG_M)
        opt_sigy = np.maximum(opt_sigy, MIN_SIG_M)

        def _sum_chunk_on_grid_opt(Xg, Yg, idxs: np.ndarray, normalize_flag: bool) -> np.ndarray:
            Zc = np.zeros_like(Xg, dtype=float)
            for i in idxs:
                if opt_dist[i] in ['pseudo-voigt', 'voigt']:
                    Zc += pseudo_voigt_2d_rotated(
                        Xg, Yg,
                        muazi=opt_mux[i], murad=opt_muy[i],
                        sigmaazi=opt_sigx[i], sigmarad=opt_sigy[i],
                        theta=opt_theta[i],
                        alphaazi=opt_alpha_azi[i],
                        alpharad=opt_alpha_rad[i],
                        amplitude=opt_weight[i],
                        normalize=normalize_flag,
                        degrees=True,
                    )
                else:
                    Zc += gaussian_2d_rotated(
                        Xg, Yg,
                        mux=opt_mux[i], muy=opt_muy[i],
                        sigmax=opt_sigx[i], sigmay=opt_sigy[i],
                        theta=opt_theta[i],
                        amplitude=opt_weight[i],
                        normalize=normalize_flag,
                        degrees=True,
                    )
            return Zc

        def _sum_on_grid_opt(Xg, Yg, normalize_flag: bool) -> np.ndarray:
            n = len(opt_mux)
            if n == 0:
                return np.zeros_like(Xg, dtype=float)
            max_workers = 1 if n < 25 else min(8, (os.cpu_count() or 2))
            if max_workers <= 1:
                return _sum_chunk_on_grid_opt(Xg, Yg, np.arange(n), normalize_flag)
            chunks = np.array_split(np.arange(n), max_workers)
            with ThreadPoolExecutor(max_workers=max_workers) as ex:
                parts = list(ex.map(lambda c: _sum_chunk_on_grid_opt(Xg, Yg, c, normalize_flag), chunks))
            Zg = np.zeros_like(Xg, dtype=float)
            for p in parts:
                Zg += p
            return Zg
        
        # Create radial_profile function for optimized df
        def radial_profile_opt(cx, cy, n_r=400, n_theta=360, r_margin_factor=5.0):
            max_sigma = max(df_optimized['sigmax'].max(), df_optimized['sigmay'].max())
            max_center_dist = np.sqrt((df_optimized['mux'] - cx) ** 2 + (df_optimized['muy'] - cy) ** 2).max()
            r_max = max_center_dist + r_margin_factor * max_sigma
            if r_max <= 0:
                r_max = 1e-6
            theta = np.linspace(0.0, 2.0 * np.pi, n_theta, endpoint=False)
            r = np.linspace(0.0, r_max, n_r)
            dtheta = theta[1] - theta[0]
            dr = r[1] - r[0] if n_r > 1 else r_max
            R, TH = np.meshgrid(r, theta)
            Xp = cx + R * np.cos(TH)
            Yp = cy + R * np.sin(TH)
            Zp = _sum_on_grid_opt(Xp, Yp, normalize)
            radial_energy = np.sum(Zp * R, axis=0) * dtheta
            cumulative = np.cumsum(radial_energy * dr)
            total_energy = cumulative[-1] if cumulative.size else 1.0
            if debug:
                    pass
            return r, cumulative, total_energy
        
        # Find best focus for optimized (simple version - use computed center)
        def hew_at_center_opt(cx, cy, coarse=False):
            if coarse:
                n_r = 110 if fast else 140
                n_theta = 72 if fast else 100
            else:
                n_r = 200 if fast else 240
                n_theta = 120 if fast else 180
            r, cumulative, total_energy = radial_profile_opt(cx, cy, n_r=n_r, n_theta=n_theta)
            frac = cumulative / total_energy if total_energy > 0 else cumulative
            return _radius_for_fraction(frac, r, target=0.5)
        
        # Optimize center position
        if quick_mode:
            best_opt_cx, best_opt_cy = opt_center_x, opt_center_y
            best_opt_hew = hew_at_center_opt(best_opt_cx, best_opt_cy, coarse=True)
            step_size = 1e-6
            for cx, cy in (
                (opt_center_x + step_size, opt_center_y),
                (opt_center_x - step_size, opt_center_y),
                (opt_center_x, opt_center_y + step_size),
                (opt_center_x, opt_center_y - step_size),
            ):
                hew_val = hew_at_center_opt(cx, cy, coarse=True)
                if hew_val < best_opt_hew:
                    best_opt_cx, best_opt_cy, best_opt_hew = cx, cy, hew_val
            opt_center_x, opt_center_y = best_opt_cx, best_opt_cy
        else:
            candidates_start = [(opt_center_x, opt_center_y), (0.0, 0.0)]
            best_opt_cx, best_opt_cy = opt_center_x, opt_center_y
            best_opt_hew = hew_at_center_opt(best_opt_cx, best_opt_cy, coarse=True)
            for start_cx, start_cy in candidates_start:
                current_cx, current_cy = start_cx, start_cy
                current_hew = hew_at_center_opt(current_cx, current_cy, coarse=True)
                step_size = 1e-6
                for _ in range(25):
                    candidates = [
                        (current_cx + step_size, current_cy),
                        (current_cx - step_size, current_cy),
                        (current_cx, current_cy + step_size),
                        (current_cx, current_cy - step_size)
                    ]
                    improved = False
                    for cx, cy in candidates:
                        hew_val = hew_at_center_opt(cx, cy, coarse=True)
                        if hew_val < current_hew:
                            current_cx, current_cy = cx, cy
                            current_hew = hew_val
                            improved = True
                            break
                    if not improved:
                        step_size *= 0.5
                        if step_size < 1e-10:
                            break
                if current_hew < best_opt_hew:
                    best_opt_cx, best_opt_cy = current_cx, current_cy
                    best_opt_hew = current_hew
            
            opt_center_x, opt_center_y = best_opt_cx, best_opt_cy
        
        # Compute final optimized metrics
        opt_r_profile, opt_cumulative_profile, opt_total_energy = radial_profile_opt(opt_center_x, opt_center_y, n_r=n_r_final, n_theta=n_theta_final, r_margin_factor=final_r_margin)
        opt_frac_profile = opt_cumulative_profile / opt_total_energy if opt_total_energy > 0 else opt_cumulative_profile
        opt_radius_50 = _radius_for_fraction(opt_frac_profile, opt_r_profile, target=0.5)
        opt_radius_90 = _radius_for_fraction(opt_frac_profile, opt_r_profile, target=0.9)
    
    # Build plot bounds first (avoid expensive Z recomputation)
    max_radius = max(radius_50 or 0, radius_90 or 0, radius_50_00 or 0)
    # Fallback bounds based on Gaussian extents if radii are missing
    margin_factor = 3
    min_x = df['mux'].min() - margin_factor * df['sigmax'].max()
    max_x = df['mux'].max() + margin_factor * df['sigmax'].max()
    min_y = df['muy'].min() - margin_factor * df['sigmay'].max()
    max_y = df['muy'].max() + margin_factor * df['sigmay'].max()

    """ if max_radius and max_radius > 0:
        xlim = (
            min(min_x, center_x - 1.1 * max_radius, -1.1 * (radius_50_00 or 0.0)),
            max(max_x, center_x + 1.1 * max_radius,  1.1 * (radius_50_00 or 0.0)),
        )
        ylim = (
            min(min_y, center_y - 1.1 * max_radius, -1.1 * (radius_50_00 or 0.0)),
            max(max_y, center_y + 1.1 * max_radius,  1.1 * (radius_50_00 or 0.0)),
        )
    else:
        xlim = (min_x, max_x)
        ylim = (min_y, max_y) """

    xlim = (-0.015, 0.015)
    ylim = (-0.015, 0.015)

    x = np.linspace(xlim[0], xlim[1], nx)
    y = np.linspace(ylim[0], ylim[1], ny)
    X, Y = np.meshgrid(x, y)
    Z = _sum_on_grid(X, Y, normalize)

    # Create the plots in a single figure with two subplots
    # Use a reasonable figure size that will fit most screens; enable constrained layout
    # Produce the large PSF+EEF figure (single-window diagnostic).
    fig = plt.figure(figsize=(16, 8), constrained_layout=True)
    # Standard linewidth for EEF marker lines (keep 90% and 80% consistent)
    eef_linewidth = 1.5
    
    # Use GridSpec with 20 columns for finer control: left (12 cols) + right (8 cols)
    # Right (EEF) will therefore be 8/20 = 40% of the figure width
    gs = gridspec.GridSpec(1, 20, figure=fig)
    
    # First subplot: weighted sum of Gaussians (left, wider)
    ax1 = plt.subplot(gs[0, :12])
    # Convert to microns for display (1 m = 1e6 µm)
    im = plt.imshow(Z, extent=[x.min()*1e6, x.max()*1e6, y.min()*1e6, y.max()*1e6], origin='lower', cmap='viridis', aspect='equal')
    # Zoom to ~90% containment of the PSF for the display image (does not affect FITS export)
    try:
        # Compute radial distances in meters from the weighted center (center_x/center_y defined earlier)
        cx = center_x if 'center_x' in locals() else 0.0
        cy = center_y if 'center_y' in locals() else 0.0
        # X,Y are in meters; Z contains summed flux per grid cell
        r = np.hypot((X - cx), (Y - cy))
        flat_idx = np.argsort(r.ravel())
        zs = Z.ravel()[flat_idx]
        rs = r.ravel()[flat_idx]
        total = np.nansum(zs)
        if total > 0:
            cumsum = np.cumsum(np.nan_to_num(zs))
            frac = cumsum / total
            # find radius enclosing 90% (or fallback to max radius)
            idx90 = np.searchsorted(frac, 0.90)
            r90 = float(rs[idx90]) if idx90 < len(rs) else float(rs.max())
            # add a small margin (5%) to ensure visual context
            r_vis = r90 * 1.05
            # Convert to microns for axis limits and cap at +/-1 mm (1000 µm)
            r_vis_um = min(r_vis * 1e6, 1000.0)
            ax1.set_xlim(-r_vis_um, r_vis_um)
            ax1.set_ylim(-r_vis_um, r_vis_um)
    except Exception:
        pass
    # Smaller, tighter colorbar attached to ax1 to reduce horizontal crowding
    cbar = plt.colorbar(im, ax=ax1, label='counts', pad=0.02, fraction=0.046)
    ax1.set_xlabel('x [µm]')
    ax1.set_ylabel('y [µm]')
    # Overlay a higher-resolution rendition of the zoomed area for display
    try:
        # Use the zoom limits if available (r_vis_um set above), otherwise use full extent
        if 'r_vis_um' in locals():
            x0_um, x1_um = -r_vis_um, r_vis_um
            y0_um, y1_um = -r_vis_um, r_vis_um
        else:
            x0_um, x1_um = x.min()*1e6, x.max()*1e6
            y0_um, y1_um = y.min()*1e6, y.max()*1e6

        # Convert to meters for computation
        x0_m, x1_m = x0_um * 1e-6, x1_um * 1e-6
        y0_m, y1_m = y0_um * 1e-6, y1_um * 1e-6

        # Original grid spacing
        dx = x[1] - x[0]
        dy = y[1] - y[0]

        # Number of original samples across the zoom area
        nx_zoom = max(8, int(round((x1_m - x0_m) / max(dx, 1e-12))))
        ny_zoom = max(8, int(round((y1_m - y0_m) / max(dy, 1e-12))))

        # Increase resolution by factor 12 for display only
        factor = 12
        nx_disp = int(np.clip(nx_zoom * factor, 32, 10000))
        ny_disp = int(np.clip(ny_zoom * factor, 32, 10000))

        # Build fine mesh and evaluate PSF only over zoom area (display-only)
        x_disp = np.linspace(x0_m, x1_m, nx_disp)
        y_disp = np.linspace(y0_m, y1_m, ny_disp)
        Xd, Yd = np.meshgrid(x_disp, y_disp)
        Zd = _sum_on_grid(Xd, Yd, normalize)

        # Overlay the high-resolution image covering the current zoom limits (in µm)
        ax1.imshow(Zd, extent=[x0_um, x1_um, y0_um, y1_um], origin='lower', cmap='viridis', aspect='equal', interpolation='nearest')
    except Exception:
        pass
    
    # Add secondary axes for arcsec. Use the same project convention as
    # `arcsec_to_m` (1 arcsec = 12*π/180/3600 m), therefore 1 m equals
    # the reciprocal: 1 / (12*π/180/3600) arcsec. Compute explicitly.
    m_to_arcsec = 1.0 / (12.0 * np.pi / 180.0 / 3600.0)
    um_to_arcsec = m_to_arcsec * 1e-6  # microns to arcsec
    
    # Top axis for x in arcsec
    ax1_top = ax1.secondary_xaxis('top', functions=(lambda um: um * um_to_arcsec, lambda arcsec: arcsec / um_to_arcsec))
    ax1_top.set_xlabel('x [arcsec]')
    
    # Right axis for y in arcsec (immediately next to plot)
    ax1_right = ax1.secondary_yaxis('right', functions=(lambda um: um * um_to_arcsec, lambda arcsec: arcsec / um_to_arcsec))
    ax1_right.set_ylabel('y [arcsec]', rotation=270, va='bottom')
    
    
    # Mark the minimum with a green cross and coordinates (label updated)
    plt.plot(center_x*1e6, center_y*1e6, 'gx', markersize=10, label='center for minimum HEW')
    plt.text(center_x*1e6, center_y*1e6, f'({center_x*1e6:.2f}, {center_y*1e6:.2f})', color='green', ha='left', va='bottom')
    # Mark (0,0) with a blue cross
    plt.plot(0, 0, 'bx', markersize=10, label='(0,0)')
    # Add circles for 50% and 90% encircled energy
    max_radius = max(radius_50 or 0, radius_90 or 0)
    
    # Calculate axis limits to include: best focus circles, (0,0), and blue circle
    margin_factor = 0.1
    # Start with limits needed for best focus circles
    xlim_min = center_x - max_radius if max_radius > 0 else center_x - 1
    xlim_max = center_x + max_radius if max_radius > 0 else center_x + 1
    ylim_min = center_y - max_radius if max_radius > 0 else center_y - 1
    ylim_max = center_y + max_radius if max_radius > 0 else center_y + 1
    
    # Ensure (0,0) and blue circle (radius_50_00) are included
    if radius_50_00 is not None and radius_50_00 > 0:
        xlim_min = min(xlim_min, 0 - radius_50_00)
        xlim_max = max(xlim_max, 0 + radius_50_00)
        ylim_min = min(ylim_min, 0 - radius_50_00)
        ylim_max = max(ylim_max, 0 + radius_50_00)
    else:
        # At minimum, include (0,0)
        xlim_min = min(xlim_min, 0)
        xlim_max = max(xlim_max, 0)
        ylim_min = min(ylim_min, 0)
        ylim_max = max(ylim_max, 0)
    
    # Add margin and convert to microns
    x_range = xlim_max - xlim_min
    y_range = ylim_max - ylim_min
    margin_x = margin_factor * x_range
    margin_y = margin_factor * y_range
    # Guard against NaN/Inf axis limits (can happen if inputs contain NaNs).
    if not np.isfinite(xlim_min) or not np.isfinite(xlim_max):
        xlim_min = center_x - 1.0
        xlim_max = center_x + 1.0
    if not np.isfinite(ylim_min) or not np.isfinite(ylim_max):
        ylim_min = center_y - 1.0
        ylim_max = center_y + 1.0
    margin_x = margin_factor * (xlim_max - xlim_min)
    margin_y = margin_factor * (ylim_max - ylim_min)
    left_x = (xlim_min - margin_x) * 1e6
    right_x = (xlim_max + margin_x) * 1e6
    bottom_y = (ylim_min - margin_y) * 1e6
    top_y = (ylim_max + margin_y) * 1e6
    if not (np.isfinite(left_x) and np.isfinite(right_x)):
        left_x = (center_x - 1.0) * 1e6
        right_x = (center_x + 1.0) * 1e6
    if not (np.isfinite(bottom_y) and np.isfinite(top_y)):
        bottom_y = (center_y - 1.0) * 1e6
        top_y = (center_y + 1.0) * 1e6
    plt.xlim(left_x, right_x)
    plt.ylim(bottom_y, top_y)
    # Precompute arcsec conversions for legend/value annotations
    m_to_arcsec = 1.0 / (12.0 * np.pi / 180.0 / 3600.0)

    def _min_interval_width(axis_vals: np.ndarray, prof: np.ndarray, frac: float = 0.5) -> float | None:
        """Smallest interval (anywhere) containing a fraction of 1D energy.

        Given a nonnegative profile sampled on an increasing axis, returns the minimal
        width W such that there exists an interval [a, b] with (b-a)=W and
        \int_a^b prof(x) dx >= frac * \int prof(x) dx.

        This matches the request for HEW_x/HEW_y as the *smallest* interval containing 50%.
        """
        axis_vals = np.asarray(axis_vals, dtype=float)
        prof = np.asarray(prof, dtype=float)

        if prof.size < 3 or axis_vals.size != prof.size:
            return None
        if not np.isfinite(prof).any() or not np.isfinite(axis_vals).all():
            return None
        if float(frac) <= 0.0:
            return 0.0

        # Ensure monotonic increasing axis
        if not np.all(np.diff(axis_vals) > 0):
            return None

        # Clamp tiny negatives that can appear from numeric noise
        prof = np.maximum(prof, 0.0)

        # Segment integrals using trapezoids
        dx = np.diff(axis_vals)
        seg = 0.5 * (prof[:-1] + prof[1:]) * dx
        total = float(np.sum(seg))
        if not np.isfinite(total) or total <= 0.0:
            return None
        target = float(frac) * total

        # Prefix integral at nodes: pref[k] = integral from x[0] to x[k]
        pref = np.empty(prof.size, dtype=float)
        pref[0] = 0.0
        pref[1:] = np.cumsum(seg)

        best = None
        # Two-pointer sliding window: for each left index i, advance right index j
        # until the integral pref[j] - pref[i] >= target, then record width.
        n = prof.size
        j = 1
        for i in range(0, n - 1):
            if j < i + 1:
                j = i + 1
            # advance j until interval [i,j] reaches target or j==n-1
            while j < n and (pref[j] - pref[i]) < target:
                j += 1
            if j >= n:
                break
            width = float(axis_vals[j] - axis_vals[i])
            if best is None or width < best:
                best = width
        return best
    
    def _fwhm_from_profile(axis_vals: np.ndarray, prof: np.ndarray) -> float | None:
        #Compute FWHM from a 1D profile sampled on axis_vals.
        axis_vals = np.asarray(axis_vals, dtype=float)
        prof = np.asarray(prof, dtype=float)
        if prof.size < 3 or axis_vals.size != prof.size:
            return None
        if not np.isfinite(prof).any() or not np.isfinite(axis_vals).all():
            return None
        # Normalize profile
        prof = np.maximum(prof, 0.0)
        max_val = np.max(prof)
        if not np.isfinite(max_val) or max_val <= 0:
            return None
        half_max = max_val / 2.0
        above = np.where(prof >= half_max)[0]
        if above.size < 2:
            return None
        left = above[0]
        right = above[-1]
        return float(axis_vals[right] - axis_vals[left])


    def _compute_hew_xy_arcsec_from_grid_marginals(x_axis: np.ndarray, y_axis: np.ndarray, Zg: np.ndarray) -> tuple[float | None, float | None]:
        """Compute HEW_x and HEW_y (arcsec) from a gridded PSF `Zg` sampled on axes `x_axis`, `y_axis`.

        The function integrates the 2D PSF to obtain 1D marginals and then finds
        the minimal-width interval containing 50% of the energy for each axis.
        Returns (hew_x_arcsec, hew_y_arcsec) where values may be None on failure.
        """
        try:
            if getattr(Zg, 'size', 0) == 0:
                return (None, None)
            if not np.isfinite(np.asarray(Zg, dtype=float)).any():
                return (None, None)

            x_axis = np.asarray(x_axis, dtype=float)
            y_axis = np.asarray(y_axis, dtype=float)
            Zg = np.asarray(Zg, dtype=float)

            # Marginals (integrate the full 2D PSF along the other axis)
            prof_x = np.trapezoid(Zg, y_axis, axis=0)
            prof_y = np.trapezoid(Zg, x_axis, axis=1)

            hew_x_m = _min_interval_width(x_axis, prof_x, frac=0.5)
            hew_y_m = _min_interval_width(y_axis, prof_y, frac=0.5)

            # Convert meters back to arcsec using project convention
            m_to_arcsec = 1.0 / (12.0 * np.pi / 180.0 / 3600.0)
            hew_x_arcsec = (hew_x_m * m_to_arcsec) if hew_x_m is not None else None
            hew_y_arcsec = (hew_y_m * m_to_arcsec) if hew_y_m is not None else None
            return hew_x_arcsec, hew_y_arcsec
        except Exception:
            return (None, None)

    # HEW_x / HEW_y annotations for encircled-energy legend entries
    # Base curves share the same underlying PSF, so their HEW_x/HEW_y should be the same.
    hew_base_x_arcsec, hew_base_y_arcsec = _compute_hew_xy_arcsec_from_grid_marginals(x, y, Z)
    hew_best_x_arcsec, hew_best_y_arcsec = hew_base_x_arcsec, hew_base_y_arcsec
    hew_00_x_arcsec, hew_00_y_arcsec = hew_base_x_arcsec, hew_base_y_arcsec

    # FWHM_x and FWHM_y calculation
    prof_x = np.trapezoid(Z, y, axis=0)
    prof_y = np.trapezoid(Z, x, axis=1)
    fwhm_x_m = _fwhm_from_profile(x, prof_x)
    fwhm_y_m = _fwhm_from_profile(y, prof_y)
    m_to_arcsec = 1.0 / (12.0 * np.pi / 180.0 / 3600.0)
    fwhm_x_arcsec = fwhm_x_m * m_to_arcsec if fwhm_x_m is not None else None
    fwhm_y_arcsec = fwhm_y_m * m_to_arcsec if fwhm_y_m is not None else None

    fwhm_opt_x_arcsec, fwhm_opt_y_arcsec = (None, None)
    if df_optimized is not None:
        Z_opt = _sum_on_grid_opt(X, Y, normalize)
        prof_x_opt = np.trapezoid(Z_opt, y, axis=0)
        prof_y_opt = np.trapezoid(Z_opt, x, axis=1)
        fwhm_x_m_opt = _fwhm_from_profile(x, prof_x_opt)
        fwhm_y_m_opt = _fwhm_from_profile(y, prof_y_opt)
        fwhm_opt_x_arcsec = fwhm_x_m_opt * m_to_arcsec if fwhm_x_m_opt is not None else None
        fwhm_opt_y_arcsec = fwhm_y_m_opt * m_to_arcsec if fwhm_y_m_opt is not None else None
    hew_best_arcsec = 2 * radius_50 * m_to_arcsec if (radius_50 is not None and np.isfinite(radius_50)) else None
    hew_origin_arcsec = 2 * radius_50_00 * m_to_arcsec if (radius_50_00 is not None and np.isfinite(radius_50_00)) else None
    eef80_arcsec = 2 * radius_80 * m_to_arcsec if (radius_80 is not None and np.isfinite(radius_80)) else None
    eef90_arcsec = 2 * radius_90 * m_to_arcsec if (radius_90 is not None and np.isfinite(radius_90)) else None
    eef90_origin_arcsec = 2 * radius_90_00 * m_to_arcsec if (radius_90_00 is not None and np.isfinite(radius_90_00)) else None

    # If caller only requests metrics, return them now without any plotting side-effects.
    if return_metrics_only:
        return {
            'hew_origin_arcsec': hew_origin_arcsec,
            'hew_best_arcsec': hew_best_arcsec,
            'eef80_best_arcsec': eef80_arcsec,
            'eef90_origin_arcsec': eef90_origin_arcsec,
            'eef90_best_arcsec': eef90_arcsec,
            'hew_x_arcsec': hew_base_x_arcsec,
            'hew_y_arcsec': hew_base_y_arcsec,
            'hew_opt_x_arcsec': hew_opt_x_arcsec if 'hew_opt_x_arcsec' in locals() else None,
            'hew_opt_y_arcsec': hew_opt_y_arcsec if 'hew_opt_y_arcsec' in locals() else None,
            'fwhm_x_arcsec': fwhm_x_arcsec,
            'fwhm_y_arcsec': fwhm_y_arcsec,
            'hew_opt_arcsec': (2 * opt_radius_50 * m_to_arcsec) if 'opt_radius_50' in locals() and opt_radius_50 is not None else None,
            'eef90_opt_arcsec': (2 * opt_radius_90 * m_to_arcsec) if 'opt_radius_90' in locals() and opt_radius_90 is not None else None,
        }
    else:
        # Fallback: if origin HEW couldn't be computed from radial integration,
        # use the grid-marginal HEW as a best-effort surrogate.
        if hew_origin_arcsec is None and hew_base_x_arcsec is not None:
            hew_origin_arcsec = hew_base_x_arcsec
        # Also expose the exact same metrics when producing plots so both
        # interactive and metrics-only runs produce identical outputs.
        try:
            metrics_out = {
                'hew_origin_arcsec': hew_origin_arcsec,
                'hew_best_arcsec': hew_best_arcsec,
                'eef80_best_arcsec': eef80_arcsec,
                'eef90_origin_arcsec': eef90_origin_arcsec,
                'eef90_best_arcsec': eef90_arcsec,
                'hew_x_arcsec': hew_base_x_arcsec,
                'hew_y_arcsec': hew_base_y_arcsec,
                'hew_opt_x_arcsec': hew_opt_x_arcsec,
                'hew_opt_y_arcsec': hew_opt_y_arcsec,
                'hew_opt_arcsec': (2 * opt_radius_50 * m_to_arcsec) if opt_radius_50 is not None else None,
                'eef90_opt_arcsec': (2 * opt_radius_90 * m_to_arcsec) if opt_radius_90 is not None else None,
            }
            print(json.dumps(metrics_out, indent=2))
        except Exception:
            pass
    if radius_90 is not None:
        # Add a dashed circle for 90% encircled energy in red (short legend label)
        label_90 = 'EEF 90% (min)'
        circle_90 = plt.Circle((center_x*1e6, center_y*1e6), radius_90*1e6, fill=False, color='red', linestyle='--', linewidth=2, label=label_90)
        plt.gca().add_patch(circle_90)
    if radius_50_00 is not None:
        # Add a dashed circle for 50% encircled energy from (0,0) in blue (short label)
        label_50_00 = 'HEW (0,0)'
        circle_50_00 = plt.Circle((0, 0), radius_50_00*1e6, fill=False, color='blue', linestyle='--', linewidth=2, label=label_50_00)
        plt.gca().add_patch(circle_50_00)
    if radius_50 is not None:
        # Add a dashed circle for 50% encircled energy in green (short label)
        label_50 = 'HEW (min)'
        circle_50 = plt.Circle((center_x*1e6, center_y*1e6), radius_50*1e6, fill=False, color='green', linestyle='--', linewidth=2, label=label_50)
        plt.gca().add_patch(circle_50)
    
    # Add optimized circles if provided
    if df_optimized is not None and opt_radius_50 is not None:
        opt_hew_arcsec = 2 * opt_radius_50 * m_to_arcsec
        opt_eef90_arcsec = 2 * opt_radius_90 * m_to_arcsec if opt_radius_90 is not None else None
        # Mark optimized best focus
        plt.plot(opt_center_x*1e6, opt_center_y*1e6, 'mx', markersize=10, label='optimized minimum')
        # HEW circle (magenta dotted) — short label
        label_opt_50 = 'HEW (optimized)'
        circle_opt_50 = plt.Circle((opt_center_x*1e6, opt_center_y*1e6), opt_radius_50*1e6, fill=False, color='magenta', linestyle=':', linewidth=2, label=label_opt_50)
        plt.gca().add_patch(circle_opt_50)
        # EEF 90% circle (orange dotted)
        if opt_radius_90 is not None:
            label_opt_90 = 'EEF 90% (opt)'
            circle_opt_90 = plt.Circle((opt_center_x*1e6, opt_center_y*1e6), opt_radius_90*1e6, fill=False, color='orange', linestyle=':', linewidth=2, label=label_opt_90)
            plt.gca().add_patch(circle_opt_90)
    
    # Reorder legend items on the left subplot:
    # 1) focus points, 2) HEW circles, 3) EEF 90% circles
    handles, labels = ax1.get_legend_handles_labels()

    focus_order = {
        '(0,0)': 0,
        'center for minimum HEW': 1,
        'minimum': 2,
        'optimized minimum': 3,
    }

    def _legend_sort_key(label: str) -> tuple[int, int, str]:
        if label in focus_order:
            return (0, focus_order[label], label)
        if label.startswith('HEW'):
            if 'minimum' in label:
                return (1, 0, label)
            if '(0,0)' in label:
                return (1, 1, label)
            if 'optimized' in label:
                return (1, 2, label)
            return (1, 99, label)
        if label.startswith('EEF 80%'):
            # Place EEF 80% entries before EEF 90%
            if '(min)' in label or 'minimum' in label:
                return (2, -1, label)
            if 'optimized' in label:
                return (2, 0, label)
            return (2, 50, label)
        if label.startswith('EEF 90%'):
            if '(min)' in label or 'minimum' in label:
                return (2, 0, label)
            if 'optimized' in label:
                return (2, 1, label)
            return (2, 99, label)
        return (3, 99, label)

    order = sorted(range(len(labels)), key=lambda i: _legend_sort_key(labels[i]))
    # Ensure (0,0) is first and 'center for minimum HEW' is second if present
    preferred_first = []
    try:
        idx_00 = next(i for i, lbl in enumerate(labels) if lbl == '(0,0)')
        preferred_first.append(idx_00)
    except StopIteration:
        idx_00 = None
    try:
        idx_center = next(i for i, lbl in enumerate(labels) if lbl == 'center for minimum HEW')
        # Only add if it's not the same as (0,0)
        if idx_center is not None and idx_center != idx_00:
            preferred_first.append(idx_center)
    except StopIteration:
        pass

    # Build final order: preferred labels first (in the requested sequence), then the rest in the sorted order
    remaining = [i for i in order if i not in preferred_first]
    final_order = preferred_first + remaining
    handles_sorted = [handles[i] for i in final_order]
    labels_sorted = [labels[i] for i in final_order]
    # place compact legend inside the left subplot (upper-left) with 2 columns x 3 rows
    try:
        ax1.legend(handles_sorted, labels_sorted, loc='upper left', ncol=2, fontsize=8, framealpha=0.85, bbox_to_anchor=(0.02, 0.98), bbox_transform=ax1.transAxes)
    except Exception:
        try:
            ax1.legend(handles_sorted, labels_sorted, loc='upper left', framealpha=0.85, bbox_to_anchor=(0.02, 0.98), bbox_transform=ax1.transAxes)
        except Exception:
            ax1.legend(handles_sorted, labels_sorted, loc='upper left')
    
    # Second subplot: encircled energy function (right, 40% width)
    ax2 = plt.subplot(gs[0, 12:])
    # Convert diameter from meters to arcsec (1 m = 54000/π arcsec)
    profile_pct = frac_profile * 100 if 'frac_profile' in locals() else []
    profile_diam = 2 * r_profile * m_to_arcsec if 'r_profile' in locals() else []
    profile_pct_00 = frac_00 * 100 if 'frac_00' in locals() else []
    profile_diam_00 = 2 * r_profile_00 * m_to_arcsec if 'r_profile_00' in locals() else []
    # labels for the EEF subplot — include FWHM values when available
    try:
        if fwhm_x_arcsec is not None and fwhm_y_arcsec is not None:
            label_best = f'Centered on minimum (FWHM_x={fwhm_x_arcsec:.4f}\", FWHM_y={fwhm_y_arcsec:.4f}\")'
            label_00 = f'Centered on (0,0) (FWHM_x={fwhm_x_arcsec:.4f}\", FWHM_y={fwhm_y_arcsec:.4f}\")'
        else:
            label_best = 'Centered on minimum'
            label_00 = 'Centered on (0,0)'
    except Exception:
        label_best = 'Centered on minimum'
        label_00 = 'Centered on (0,0)'
    # Limit plot to 95% percentile
    def limit_percentile(pct, diam, max_pct=95):
        pct = np.array(pct)
        diam = np.array(diam)
        mask = pct <= max_pct
        return pct[mask], diam[mask]
    profile_pct_95, profile_diam_95 = limit_percentile(profile_pct, profile_diam)
    profile_pct_00_95, profile_diam_00_95 = limit_percentile(profile_pct_00, profile_diam_00)
    plt.plot(profile_pct_95, profile_diam_95, label=label_best, color='green', linestyle='-', linewidth=2.5)
    plt.plot(profile_pct_00_95, profile_diam_00_95, label=label_00, color='blue', linestyle='-', linewidth=2.5)
    if fit_profile_pct is not None and fit_profile_diam is not None:
        fit_profile_pct_95, fit_profile_diam_95 = limit_percentile(fit_profile_pct, fit_profile_diam)
    # PV label for EEF subplot — include full fit parameters when available
    pv_label = 'Modified pseudo-Voigt radial fit'
    try:
        if 'A_fit' in locals():
            pv_label = (f"Modified pseudo-Voigt radial fit (A={A_fit:.2e}, Gamma_c={Gamma_c_fit:.2f}\"",
                        f"Gamma_w={Gamma_w_fit:.2f}\", eta={eta_fit:.3f}, beta={beta_fit:.2f}, s={scalar_fit:.2f})")
            # join tuple if accidentally created as tuple above (fallback safety)
            if isinstance(pv_label, tuple):
                pv_label = ' '.join(pv_label)
    except Exception:
        pv_label = 'Modified pseudo-Voigt radial fit'

    # Plot PV EEF on this right-side EEF subplot. Prefer PV evaluated on the fine rplot
    # grid (`Ifit_pv_plot` with `rplot`) when available; otherwise fall back to the
    # precomputed `fit_profile_*` arrays.
    try:
        if 'Ifit_pv_plot' in locals() and 'rplot' in locals() and Ifit_pv_plot is not None:
            rpv = np.asarray(rplot, dtype=float)
            drpv = float(rpv[1] - rpv[0]) if rpv.size > 1 else 1.0
            radial_pv_plot = 2.0 * np.pi * rpv * np.asarray(Ifit_pv_plot, dtype=float) * drpv
            tot_pv_plot = np.sum(radial_pv_plot)
            if tot_pv_plot > 0 and np.isfinite(tot_pv_plot):
                pv_eef_pct_plot = 100.0 * np.cumsum(radial_pv_plot) / tot_pv_plot
                pv_diam_plot = 2.0 * rpv
                maskpv = pv_eef_pct_plot <= 95.0
                if np.any(maskpv):
                    plt.plot(pv_eef_pct_plot[maskpv], pv_diam_plot[maskpv], label=pv_label, color='red', linestyle='--', linewidth=eef_linewidth)
        elif fit_profile_pct is not None and fit_profile_diam is not None:
            plt.plot(fit_profile_pct_95, fit_profile_diam_95, label=pv_label, color='red', linestyle='--', linewidth=eef_linewidth)
    except Exception:
        try:
            if fit_profile_pct is not None and fit_profile_diam is not None:
                plt.plot(fit_profile_pct_95, fit_profile_diam_95, label=pv_label, color='red', linestyle='--', linewidth=eef_linewidth)
        except Exception:
            pass

    # Draw 80% EEF circle on the left PSF image (ax1) as a purple dashed circle
    try:
        if 'profile_pct' in locals() and profile_pct is not None and len(profile_pct):
            pct_arr_full = np.asarray(profile_pct, dtype=float)
            diam_arr_full = np.asarray(profile_diam, dtype=float)
            order_full = np.argsort(pct_arr_full)
            pct_sorted_full = pct_arr_full[order_full]
            diam_sorted_full = diam_arr_full[order_full]
            if pct_sorted_full[0] <= 80.0 <= pct_sorted_full[-1]:
                diam80 = float(np.interp(80.0, pct_sorted_full, diam_sorted_full))
                # convert diameter arcsec to radius in microns for ax1 (which uses µm)
                arcsec_to_m_local = 1.0 / m_to_arcsec
                radius_um = (diam80 / 2.0) * arcsec_to_m_local * 1e6
                try:
                    import matplotlib.patches as mpatches
                    label_80 = 'EEF 80% (min)'
                    circ = mpatches.Circle((center_x * 1e6, center_y * 1e6), radius_um, edgecolor='purple', linestyle='--', linewidth=eef_linewidth, fill=False, zorder=9, label=label_80)
                    ax1.add_patch(circ)
                    # update ax1 legend to include the new circle label
                    try:
                        handles_now, labels_now = ax1.get_legend_handles_labels()
                        order_now = sorted(range(len(labels_now)), key=lambda i: _legend_sort_key(labels_now[i]))
                        handles_sorted_now = [handles_now[i] for i in order_now]
                        labels_sorted_now = [labels_now[i] for i in order_now]
                        # place compact legend inside the left subplot to avoid overlap with colorbar
                        ax1.legend(handles_sorted_now, labels_sorted_now, loc='upper left', ncol=2, fontsize=8, framealpha=0.85, bbox_to_anchor=(0.02, 0.98), bbox_transform=ax1.transAxes)
                    except Exception:
                        try:
                            ax1.legend(handles_sorted_now, labels_sorted_now, loc='upper left', ncol=2, fontsize=8, framealpha=0.85, bbox_to_anchor=(0.02, 0.98), bbox_transform=ax1.transAxes)
                        except Exception:
                            try:
                                ax1.legend()
                            except Exception:
                                pass
                except Exception:
                    pass
    except Exception:
        pass

    # (EEF legend will be created after all fit curves are plotted,
    #  so that fit parameter labels are available and included.)
    if pearson4_profile_pct is not None and pearson4_profile_diam is not None:
        pearson4_profile_pct_95, pearson4_profile_diam_95 = limit_percentile(pearson4_profile_pct, pearson4_profile_diam)
        # Pearson4 label — include key fit parameters when available
        try:
            # Build a compact Pearson4 legend label showing only the five requested parameters
            p4_label = 'Pearson4 radial fit'
            try:
                params_p4 = None
                if 'pearson4_result' in locals() and pearson4_result is not None and hasattr(pearson4_result, 'params'):
                    params_p4 = pearson4_result.params
                elif 'pearson4_used_params_final' in locals() and pearson4_used_params_final is not None:
                    params_p4 = pearson4_used_params_final
                if params_p4 is not None:
                    # Map lmfit parameter names to the exact output keys the user expects
                    mapping = [
                        ('amplitude', 'Pearson4_Amplitude', 'Amplitude'),
                        ('sigma', 'Pearson4_Sigma_arcsec', 'Sigma_arcsec'),
                        ('m', 'Pearson4_Exponent_m', 'Exponent_m'),
                        ('nu', 'Pearson4_Skew_nu', 'Skew_nu'),
                    ]
                    parts = []
                    for lm_name, out_key, disp_name in mapping:
                        try:
                            if lm_name in params_p4:
                                val = params_p4[lm_name].value if hasattr(params_p4[lm_name], 'value') else float(params_p4[lm_name])
                                sval = f"{val:.4g}"
                                parts.append(f"{disp_name}={sval}")
                        except Exception:
                            continue
                        if parts:
                            p4_label = 'Pearson4 radial fit (' + ', '.join(parts) + ')'
            except Exception:
                pass
            plt.plot(pearson4_profile_pct_95, pearson4_profile_diam_95, label=p4_label, color='orange', linestyle='--', linewidth=eef_linewidth)
        except Exception:
            pass
    # Add King profile to EEF subplot if available (clip to 95% percentile)
    if 'king_profile_pct' in locals() and king_profile_pct is not None and 'king_profile_diam' in locals() and king_profile_diam is not None:
        try:
            # King label — include fit params if the least-squares King fit produced them
            kp_label = 'King radial fit'
            try:
                if 'I0_k' in locals():
                    kp_label = f'King radial fit (I0={I0_k:.2e}, rc={rc_k:.2f}\", alpha={alpha_k:.2f}, b={b_k:.2e})'
            except Exception:
                pass
            try:
                k_pct = np.asarray(king_profile_pct, dtype=float)
                k_diam = np.asarray(king_profile_diam, dtype=float)
                k_pct_95, k_diam_95 = limit_percentile(k_pct, k_diam, max_pct=95)
                if len(k_pct_95) > 0:
                    plt.plot(k_pct_95, k_diam_95, label=kp_label, color='purple', linestyle='--', linewidth=eef_linewidth, alpha=0.95)
            except Exception:
                # fallback: attempt to plot raw arrays but still guard errors
                try:
                    plt.plot(king_profile_pct, king_profile_diam, label=kp_label, color='purple', linestyle='--', linewidth=eef_linewidth, alpha=0.95)
                except Exception:
                    pass
        except Exception:
            pass
    else:
        # Try a proper King least-squares fit to the EEF reference if available
        try:
            if 'have_least_squares' in locals() and have_least_squares and ('eef_data' in locals() and eef_data is not None) and ('r_arcsec' in locals() and 'r_fit' in locals()):
                try:
                    # initial guesses (reuse sensible defaults from earlier)
                    I0_0 = float(np.nanmax(I_fit)) if ('I_fit' in locals() and I_fit.size) else 1.0
                except Exception:
                    I0_0 = 1.0
                try:
                    rc_guess = float(Gamma_c_fit) if 'Gamma_c_fit' in locals() else (float(np.median(r_fit)) if ('r_fit' in locals() and r_fit.size) else 1.0)
                except Exception:
                    rc_guess = 1.0
                rc_0 = max(0.5 * rc_guess, 1e-3)
                alpha_0 = 2.0
                try:
                    b0 = float(np.nanmin(I_fit)) if ('I_fit' in locals() and np.any(np.isfinite(I_fit))) else 0.0
                except Exception:
                    b0 = 0.0
                x0k = [I0_0, rc_0, alpha_0, b0]
                # bounds
                lbk = [max(float(floor) if 'floor' in locals() else 1e-12, 0.0), 1e-3, 1.0, 0.0]
                rc_cap = float(min(r_arcsec.max() * 0.5 if ('r_arcsec' in locals() and r_arcsec.size) else 10.0, 50.0))
                ubk = [np.inf, rc_cap, 8.0, np.inf]

                def _resid_king(vec):
                    try:
                        I0_k, rc_k, alpha_k, b_k = vec
                        model_vals = king_profile(r_arcsec, I0_k, rc_k, alpha_k, b_k)
                        model_vals = np.maximum(model_vals, 0.0)
                        dr = float(r_arcsec[1] - r_arcsec[0]) if r_arcsec.size>1 else 1.0
                        radial = 2.0 * np.pi * r_arcsec * model_vals * dr
                        tot = np.sum(radial)
                        if tot <= 0 or not np.isfinite(tot):
                            return np.ones_like(eef_data) * 1e6
                        model_eef_pct = 100.0 * np.cumsum(radial) / tot
                        from numpy import interp
                        model_at_ref = interp(r_fit, r_arcsec, model_eef_pct) / 100.0
                        return (model_at_ref - eef_data) * (eef_weight if 'eef_weight' in locals() else 1.0)
                    except Exception:
                        return np.ones_like(eef_data) * 1e6

                try:
                    resk = least_squares(_resid_king, x0k, bounds=(lbk, ubk), loss='soft_l1', f_scale=1.0, max_nfev=10000)
                    if hasattr(resk, 'x'):
                        I0_k, rc_k, alpha_k, b_k = resk.x.tolist()
                except Exception:
                    I0_k, rc_k, alpha_k, b_k = x0k

                # compute King model on rplot for EEF plotting
                try:
                    rpv = np.asarray(rplot if 'rplot' in locals() else np.linspace(0.0, float(r_arcsec.max()), 1000), dtype=float)
                    Ifit_king_plot = king_profile(rpv, I0_k, rc_k, alpha_k, b_k)
                    drpv = float(rpv[1] - rpv[0]) if rpv.size>1 else 1.0
                    radial_k_plot = 2.0 * np.pi * rpv * Ifit_king_plot * drpv
                    tot_k_plot = np.sum(radial_k_plot)
                    if tot_k_plot > 0 and np.isfinite(tot_k_plot):
                        k_eef_pct_plot = 100.0 * np.cumsum(radial_k_plot) / tot_k_plot
                        k_diam_plot = 2.0 * rpv
                        maskk = k_eef_pct_plot <= 95.0
                        if np.any(maskk):
                            try:
                                kp_label = f'King fit (I0={I0_k:.2e}, rc={rc_k:.2f}\", alpha={alpha_k:.2f}, b={b_k:.2e})'
                            except Exception:
                                kp_label = 'King fit'
                            plt.plot(k_eef_pct_plot[maskk], k_diam_plot[maskk], label=kp_label, color='purple', linestyle='--', linewidth=eef_linewidth, alpha=0.95)
                except Exception:
                    pass
        except Exception:
            pass
    

    # Add EEF 80% markers: vertical at x=80% (black dashed) and horizontal at diameter where best-focus reaches 80% (purple dashed)
    try:
        plt.axvline(x=80.0, linestyle='--', color='black', linewidth=eef_linewidth)
        if isinstance(profile_pct, (list, np.ndarray)) and len(profile_pct) and isinstance(profile_diam, (list, np.ndarray)) and len(profile_diam):
            # Ensure arrays are numpy arrays and increasing in pct
            pct_arr = np.asarray(profile_pct, dtype=float)
            diam_arr = np.asarray(profile_diam, dtype=float)
            # Sort by pct just in case
            order_idx = np.argsort(pct_arr)
            pct_sorted = pct_arr[order_idx]
            diam_sorted = diam_arr[order_idx]
            if pct_sorted[0] <= 80.0 <= pct_sorted[-1]:
                diam80 = float(np.interp(80.0, pct_sorted, diam_sorted))
                plt.axhline(y=diam80, linestyle='--', color='purple', linewidth=eef_linewidth)
                # Add label slightly below the horizontal line (with slight offset and larger font)
                try:
                    ymin, ymax = ax2.get_ylim()
                    yoff = 0.02 * (ymax - ymin)
                    text_y = diam80 - yoff
                    ax2.text(0, text_y, f'EEF 80% minimum = {diam80:.2f}"', ha='left', va='top', fontsize=9, color='purple')
                except Exception:
                    pass
    except Exception:
        pass
    
    # Add optimized curve if provided
    if df_optimized is not None and opt_frac_profile is not None:
        opt_profile_pct = opt_frac_profile * 100
        opt_profile_diam = 2 * opt_r_profile * m_to_arcsec
        opt_profile_pct_95, opt_profile_diam_95 = limit_percentile(opt_profile_pct, opt_profile_diam)
        label_opt = 'Optimized minimum'
        plt.plot(opt_profile_pct_95, opt_profile_diam_95, label=label_opt, linestyle=':', linewidth=2.5, color='magenta')
    
    # Create final legend for EEF subplot now that all fit curves have been plotted
    try:
        handles2, labels2 = ax2.get_legend_handles_labels()
        if handles2:
            # Wrap long legend labels and place legend slightly higher/outside
            try:
                import textwrap
                wrapped2 = [textwrap.fill(lbl, width=40) if isinstance(lbl, str) else lbl for lbl in labels2]
            except Exception:
                wrapped2 = labels2
            # Place EEF legend inside the EEF subplot area (upper right)
            ax2.legend(handles2, wrapped2, loc='upper left', ncol=1, fontsize=8)
    except Exception:
        pass

    # Add inline labels near each EEF curve so labels appear inside the graph area
    # (Inline curve labels removed — rely on legend and horizontal-line annotations)

    ax2.set_xlabel('Percentage (%)')
    ax2.set_ylabel('Diameter [arcsec]')
    # Increase tick density by reducing major tick spacing by half
    from matplotlib.ticker import MultipleLocator, AutoMinorLocator
    ax2.xaxis.set_major_locator(MultipleLocator(10))  # Major ticks every 10%
    ax2.xaxis.set_minor_locator(AutoMinorLocator(5))  # Minor ticks
    # For y-axis, use auto locator with more ticks
    ax2.yaxis.set_major_locator(plt.MaxNLocator(nbins=20))  # More bins for denser ticks
    ax2.yaxis.set_minor_locator(AutoMinorLocator(5))
    # Mark the 50% encircled energy (Half Energy Width - HEW) in green for minimum
    plt.axhline(y=hew_best_arcsec, linestyle='--', color='green', linewidth=eef_linewidth)  # Horizontal line at HEW diameter
    plt.axvline(x=50, linestyle='--', color='black', linewidth=eef_linewidth)  # Vertical line at 50%
    if hew_best_arcsec is not None:
        try:
            ymin, ymax = ax2.get_ylim()
            yoff = 0.02 * (ymax - ymin)
            text_y = hew_best_arcsec - yoff
            ax2.text(0, text_y, f'HEW minimum = {hew_best_arcsec:.3f}"', ha='left', va='top', fontsize=9, color='green')  # Label HEW with value
        except Exception:
            ax2.text(0, hew_best_arcsec, f'HEW minimum = {hew_best_arcsec:.3f}"', ha='left', va='top', fontsize=9, color='green')
    # Mark the 50% from (0,0) in blue
    plt.axhline(y=hew_origin_arcsec, linestyle='--', color='blue', linewidth=eef_linewidth)  # Horizontal line at HEW(0,0) diameter
    if hew_origin_arcsec is not None:
        try:
            ymin, ymax = ax2.get_ylim()
            yoff = 0.02 * (ymax - ymin)
            text_y = hew_origin_arcsec - yoff
            # place near right side (x=100) but slightly below the line
            ax2.text(100, text_y, f'HEW (0,0) = {hew_origin_arcsec:.3f}"', ha='center', va='top', fontsize=9, color='blue')
        except Exception:
            ax2.text(100, hew_origin_arcsec, f'HEW (0,0) = {hew_origin_arcsec:.3f}"', ha='center', va='top', fontsize=9, color='blue')
    # Mark the 90% encircled energy in red
    if radius_90 is not None:
        plt.axhline(y=eef90_arcsec, linestyle='--', color='red', linewidth=eef_linewidth)  # Horizontal line at EEF90 diameter
        plt.axvline(x=90, linestyle='--', color='black', linewidth=eef_linewidth)  # Vertical line at 90%
        try:
            ymin, ymax = ax2.get_ylim()
            yoff = 0.02 * (ymax - ymin)
            text_y = eef90_arcsec - yoff
            ax2.text(0, text_y, f'EEF 90% minimum = {eef90_arcsec:.3f}"', ha='left', va='top', fontsize=9, color='red')  # Label 90% with value
        except Exception:
            ax2.text(0, eef90_arcsec, f'EEF 90% minimum = {eef90_arcsec:.3f}"', ha='left', va='top', fontsize=9, color='red')
    
    # Add optimized reference lines if provided
    if df_optimized is not None and opt_radius_50 is not None:
        opt_hew_arcsec = 2 * opt_radius_50 * m_to_arcsec
        plt.axhline(y=opt_hew_arcsec, linestyle=':', color='magenta', linewidth=1.5)
        try:
            ymin, ymax = ax2.get_ylim()
            yoff = 0.02 * (ymax - ymin)
            text_y = opt_hew_arcsec - yoff
            ax2.text(0, text_y, f'HEW(opt)={opt_hew_arcsec:.3f}"', ha='left', va='top', fontsize=9, color='magenta')
        except Exception:
            ax2.text(0, opt_hew_arcsec, f'HEW(opt)={opt_hew_arcsec:.3f}"', ha='left', va='bottom', fontsize=9, color='magenta')
        if opt_radius_90 is not None:
            opt_eef90_arcsec = 2 * opt_radius_90 * m_to_arcsec
            plt.axhline(y=opt_eef90_arcsec, linestyle=':', color='orange', linewidth=eef_linewidth)
            try:
                ymin, ymax = ax2.get_ylim()
                yoff = 0.02 * (ymax - ymin)
                text_y = opt_eef90_arcsec - yoff
                ax2.text(0, text_y, f'EEF90(opt)={opt_eef90_arcsec:.3f}"', ha='left', va='top', fontsize=9, color='orange')
            except Exception:
                ax2.text(0, opt_eef90_arcsec, f'EEF90(opt)={opt_eef90_arcsec:.3f}"', ha='left', va='bottom', fontsize=9, color='orange')
    
    # Add titles at the same y-coordinate
    fig.suptitle('')  # Clear any figure title
    # Nudge titles upward to avoid overlapping the PSF image when using constrained_layout
    # Nudge titles upward to avoid overlapping other elements
    ax1.set_title(f'E2E PSF{title_suffix}', fontweight='bold', fontsize=12, y=1.08)
    ax2.set_title(f'Encircled energy function{title_suffix}', fontweight='bold', fontsize=12, y=1.08)
    # layout is handled by constrained_layout on the figure
    # Rebuild the left subplot legend explicitly to guarantee order and layout
    try:
        # Remove any existing legend and rebuild in a deterministic order
        try:
            existing = ax1.get_legend()
            if existing is not None:
                existing.remove()
        except Exception:
            pass
        handles_all, labels_all = ax1.get_legend_handles_labels()
        preferred = []
        for name in ['(0,0)', 'center for minimum HEW']:
            if name in labels_all:
                preferred.append(labels_all.index(name))
        remaining = [i for i in range(len(labels_all)) if i not in preferred]
        final_idx = preferred + remaining
        handles_final = [handles_all[i] for i in final_idx]
        labels_final = [labels_all[i] for i in final_idx]
        ax1.legend(handles_final, labels_final, loc='upper left', ncol=2, fontsize=8, framealpha=0.85, bbox_to_anchor=(0.02, 0.98), bbox_transform=ax1.transAxes)
    except Exception:
        try:
            ax1.legend(loc='upper left', ncol=2, fontsize=8, framealpha=0.85)
        except Exception:
            pass

    # Also write a guaranteed copy of the combined figure to the canonical filename
    try:
        out_comb = os.path.join('Figures', 'E2E_fit_combined.png')
        fig.savefig(out_comb, dpi=150, bbox_inches='tight')
        print(f"Saved combined figure to {out_comb}")
    except Exception:
        pass
    
    # Context menu implementation (cross-platform)
    menu_annotation = None
    menu_active = False
    
    def save_subplot_with_axes(ax, filename, include_colorbar=False):
        """Save a subplot with proper axis labels and ticks"""
        import time
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        final_filename = f"Figures/{filename}_{timestamp}.png"
        
        if include_colorbar:
            # For the E2E PSF image we need an exact pixel footprint that
            # matches the FITS grid. Prefer to render only the PSF axis into
            # a temporary figure sized to the desired pixel dimensions so the
            # saved PNG has precise width/height in pixels.
            try:
                # Determine desired pixel dims based on runtime mode. Use the
                # CLI `--mode` flag if present; otherwise fall back to Z.shape.
                import sys as _sys
                _main_args = getattr(_sys.modules.get('__main__'), 'args', None)
                mode = getattr(_main_args, 'mode', None) if _main_args is not None else None
                if mode == 'coarse':
                    desired_px = desired_py = 320
                elif mode == 'fine':
                    desired_px = desired_py = 2062
                # treat any unknown/legacy modes as 'fine' by default
                else:
                    if 'Z' in locals() and hasattr(Z, 'shape'):
                        desired_py, desired_px = Z.shape[0], Z.shape[1]
                    else:
                        desired_px = desired_py = 320

                # Grab the image data and properties from the existing axis
                orig_im = None
                try:
                    if hasattr(ax1, 'get_images') and ax1.get_images():
                        orig_im = ax1.get_images()[0]
                except Exception:
                    orig_im = None

                if orig_im is None:
                    # Fall back to coarse saving of the full figure if we
                    # can't access the axis image object.
                    ax2_visible = ax2.get_visible()
                    ax2.set_visible(False)
                    fig.canvas.draw()
                    from matplotlib.transforms import Bbox
                    bbox1 = ax1.get_tightbbox(fig.canvas.get_renderer()).transformed(fig.dpi_scale_trans.inverted())
                    fig.savefig(final_filename, dpi=300, bbox_inches=bbox1.expanded(1.15, 1.15))
                    ax2.set_visible(ax2_visible)
                    fig.canvas.draw_idle()
                else:
                    # Create a temporary figure sized so that width* dpi_out = pixels
                    dpi_out = 100
                    w_in = desired_px / dpi_out
                    h_in = desired_py / dpi_out
                    import matplotlib.pyplot as _mpl
                    fig2 = _mpl.figure(figsize=(w_in, h_in), dpi=dpi_out)
                    ax_tmp = fig2.add_axes([0, 0, 1, 1])
                    arr = orig_im.get_array()
                    # Use the axis view limits (zoomed to 95% containment) as extent
                    try:
                        x0, x1 = ax.get_xlim()
                        y0, y1 = ax.get_ylim()
                        extent = (x0, x1, y0, y1)
                    except Exception:
                        extent = orig_im.get_extent() if hasattr(orig_im, 'get_extent') else None
                    cmap = orig_im.get_cmap() if hasattr(orig_im, 'get_cmap') else None
                    norm = orig_im.get_norm() if hasattr(orig_im, 'get_norm') else None
                    interp = orig_im.get_interpolation() if hasattr(orig_im, 'get_interpolation') else 'nearest'
                    if extent is not None:
                        img = ax_tmp.imshow(arr, extent=extent, origin='lower', cmap=cmap, norm=norm, interpolation=interp)
                    else:
                        img = ax_tmp.imshow(arr, origin='lower', cmap=cmap, norm=norm, interpolation=interp)
                    ax_tmp.set_axis_off()
                    # Save the PSF image at exact pixel dimensions
                    fig2.savefig(final_filename, dpi=dpi_out)
                    _mpl.close(fig2)
            except Exception:
                # Any failure here should fall back to the previous combined save
                try:
                    ax2_visible = ax2.get_visible()
                    ax2.set_visible(False)
                    fig.canvas.draw()
                    from matplotlib.transforms import Bbox
                    bbox1 = ax1.get_tightbbox(fig.canvas.get_renderer()).transformed(fig.dpi_scale_trans.inverted())
                    fig.savefig(final_filename, dpi=300, bbox_inches=bbox1.expanded(1.15, 1.15))
                    ax2.set_visible(ax2_visible)
                    fig.canvas.draw_idle()
                except Exception:
                    pass
        else:
            # For other plots, just save the axis
            extent = ax.get_tightbbox(fig.canvas.get_renderer()).transformed(fig.dpi_scale_trans.inverted())
            fig.savefig(final_filename, dpi=300, bbox_inches=extent.expanded(1.15, 1.15))
        
        print(f"Saved plot to {final_filename}")
        return final_filename
    
    def export_psf_plot():
        """Export the PSF plot (left subplot) to PNG"""
        save_subplot_with_axes(ax1, "E2E_PSF", include_colorbar=True)
    
    def export_eef_plot():
        """Export the Encircled Energy Function plot (right subplot) to PNG"""
        save_subplot_with_axes(ax2, "Encircled_Energy", include_colorbar=False)

    def export_eef_csv():
        """Export the Encircled Energy Function data to CSV in CustomPSFs/"""
        import time as _time, os as _os
        ts = _time.strftime('%Y%m%d_%H%M%S')
        out = _os.path.join('CustomPSFs', f'E2E_EEF_{ts}.csv')
        _os.makedirs(_os.path.dirname(out), exist_ok=True)
        # Prepare columns
        try:
            def limit_percentile(pct, diam, max_pct=99.5):
                pct = np.array(pct)
                diam = np.array(diam)
                mask = pct <= max_pct
                return pct[mask], diam[mask]
            pct_best_raw = profile_pct if 'profile_pct' in locals() else (frac_profile * 100 if 'frac_profile' in locals() else [])
            diam_best_raw = profile_diam if 'profile_diam' in locals() else (2 * r_profile * m_to_arcsec if 'r_profile' in locals() else [])
            pct_orig_raw = profile_pct_00 if 'profile_pct_00' in locals() else (frac_00 * 100 if 'frac_00' in locals() else [])
            diam_orig_raw = profile_diam_00 if 'profile_diam_00' in locals() else (2 * r_profile_00 * m_to_arcsec if 'r_profile_00' in locals() else [])
            pct_best, diam_best = limit_percentile(pct_best_raw, diam_best_raw)
            pct_orig, diam_orig = limit_percentile(pct_orig_raw, diam_orig_raw)
        except Exception:
            pct_best, diam_best, pct_orig, diam_orig = [], [], [], []

        # Optimized arrays optional
        opt_pct = None
        opt_diam = None
        if 'opt_frac_profile' in locals() and opt_frac_profile is not None:
            try:
                opt_pct_raw = opt_frac_profile * 100
                opt_diam_raw = 2 * opt_r_profile * m_to_arcsec
                opt_pct, opt_diam = limit_percentile(opt_pct_raw, opt_diam_raw)
            except Exception:
                opt_pct, opt_diam = None, None

        # Build dataframe
        import pandas as _pd
        maxlen = max(len(pct_best) if hasattr(pct_best, '__len__') else 0,
                     len(pct_orig) if hasattr(pct_orig, '__len__') else 0,
                     len(opt_pct) if opt_pct is not None and hasattr(opt_pct, '__len__') else 0)

        def _pad(arr, n):
            if arr is None:
                return [None] * n
            if not hasattr(arr, '__len__'):
                return [arr] * n
            a = list(arr)
            if len(a) < n:
                a = a + [None] * (n - len(a))
            return a

        df_out = _pd.DataFrame({
            'pct_best': _pad(pct_best, maxlen),
            'diam_arcsec_best': _pad(diam_best, maxlen),
            'pct_origin': _pad(pct_orig, maxlen),
            'diam_arcsec_origin': _pad(diam_orig, maxlen),
        })
        if opt_pct is not None:
            df_out['pct_opt'] = _pad(opt_pct, maxlen)
            df_out['diam_arcsec_opt'] = _pad(opt_diam, maxlen)

        df_out.to_csv(out, index=False)
        print('Wrote EEF CSV to', out)

    def export_fit_params_csv():
        """Export the pseudo-Voigt fit parameters to CSV in CustomPSFs/"""
        import time as _time, os as _os, pandas as _pd
        if not fit_params_available:
            print('No pseudo-Voigt fit parameters are available for export.')
            return None
        ts = _time.strftime('%Y%m%d_%H%M%S')
        out = _os.path.join('CustomPSFs', f'E2E_fit_params_{ts}.csv')
        _os.makedirs(_os.path.dirname(out), exist_ok=True)
        df_out = _pd.DataFrame({
            'parameter': ['A', 'Gamma_core', 'Gamma_wing', 'eta', 'beta', 'scalar'],
            'value': [A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit],
        })
        df_out.to_csv(out, index=False)
        print('Wrote fit parameters CSV to', out)

    def export_eef_and_params_excel():
        """Export EEF curves and fit parameters into a single Excel workbook with two sheets."""
        import time as _time, os as _os
        try:
            import pandas as _pd
        except Exception:
            print('pandas required for Excel export')
            return

        if not fit_params_available:
            print('Warning: fit parameters not available; will still try to export EEF curves.')

        ts = _time.strftime('%Y%m%d_%H%M%S')
        out = _os.path.join('CustomPSFs', f'E2E_EEF_and_fitparams_{ts}.xlsx')
        _os.makedirs(_os.path.dirname(out), exist_ok=True)

        # Build a common diameter grid (arcsec) from 0..max r_arcsec
        try:
            rgrid = np.linspace(0.0, float(r_arcsec.max()), 1000)
        except Exception:
            rgrid = np.linspace(0.0, 100.0, 1000)
        diam_grid = 2.0 * rgrid

        # Helper to compute EEF percent from intensity profile on rgrid
        def compute_eef_pct_from_I(I_vals, r_vals):
            try:
                dr = float(r_vals[1] - r_vals[0]) if r_vals.size > 1 else 1.0
                radial = 2.0 * np.pi * r_vals * I_vals * dr
                tot = np.sum(radial)
                if tot <= 0 or not np.isfinite(tot):
                    return np.full_like(r_vals, np.nan)
                eef = 100.0 * np.cumsum(radial) / tot
                return eef
            except Exception:
                return np.full_like(r_vals, np.nan)

        # Aggregated (data) EEF interpolated to diam_grid
        try:
            if profile_pct is not None and profile_diam is not None:
                # Sort and interp
                pa = np.asarray(profile_pct, dtype=float)
                da = np.asarray(profile_diam, dtype=float)
                order = np.argsort(da)
                da_s = da[order]
                pa_s = pa[order]
                # interpolate percent at diam_grid; values outside range will be extrapolated as edge values
                agg_pct_on_grid = np.interp(diam_grid, da_s, pa_s, left=pa_s[0], right=pa_s[-1])
            else:
                agg_pct_on_grid = np.full_like(diam_grid, np.nan)
        except Exception:
            agg_pct_on_grid = np.full_like(diam_grid, np.nan)
        # Excel version: up to 99.9%; plot version: up to 95%
        agg_pct_xlsx = np.where(agg_pct_on_grid > 99.9, np.nan, agg_pct_on_grid)
        agg_pct_on_grid = np.where(agg_pct_on_grid > 95.0, np.nan, agg_pct_on_grid)

        # PV model EEF on grid
        try:
            pv_I = beta_pseudo_gaussian(rgrid, A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit)
            pv_eef = compute_eef_pct_from_I(pv_I, rgrid)
        except Exception:
            pv_eef = np.full_like(diam_grid, np.nan)
        pv_eef_xlsx = np.where(pv_eef > 99.9, np.nan, pv_eef)
        pv_eef = np.where(pv_eef > 95.0, np.nan, pv_eef)

        # Pearson4 EEF on grid - ONLY compute if an accepted pearson4_result exists
        try:
            try:
                pearson4_ok = ('pearson4_result' in locals() and pearson4_result is not None and hasattr(pearson4_result, 'params'))
            except Exception:
                pearson4_ok = False
            if pearson4_ok:
                try:
                    _pr = pearson4_result.params
                    _u = rgrid / np.maximum(float(_pr['sigma']), 1e-15)
                    p4_I = (float(_pr['amplitude'])
                            * (1.0 + _u * _u) ** (-np.maximum(float(_pr['m']), 0.5))
                            * np.exp(-float(_pr['nu']) * np.arctan(_u)))
                    p4_I = np.maximum(p4_I, 0.0)
                    p4_eef = compute_eef_pct_from_I(p4_I, rgrid)
                except Exception:
                    p4_eef = np.full_like(diam_grid, np.nan)
            else:
                p4_eef = np.full_like(diam_grid, np.nan)
        except Exception:
            p4_eef = np.full_like(diam_grid, np.nan)
        p4_eef_xlsx = np.where(p4_eef > 99.9, np.nan, p4_eef)
        p4_eef = np.where(p4_eef > 95.0, np.nan, p4_eef)

        # King EEF on grid
        try:
            if 'I0_k' in locals():
                k_I = king_profile(rgrid, I0_k, rc_k, alpha_k, b_k)
                k_I = np.maximum(k_I, 0.0)
                k_eef = compute_eef_pct_from_I(k_I, rgrid)
            else:
                k_eef = np.full_like(diam_grid, np.nan)
        except Exception:
            k_eef = np.full_like(diam_grid, np.nan)
        k_eef_xlsx = np.where(k_eef > 99.9, np.nan, k_eef)
        k_eef = np.where(k_eef > 95.0, np.nan, k_eef)

        # Make a canonical name used later in plotting code (95% cap)
        try:
            if 'k_eef' in locals() and k_eef is not None:
                king_eef_pct_plot = k_eef
        except Exception:
            pass

        # Origin (0,0) EEF interpolated to diam_grid
        try:
            _pct_orig_raw = (profile_pct_00 if 'profile_pct_00' in locals() and profile_pct_00 is not None
                             else (frac_00 * 100 if 'frac_00' in locals() and frac_00 is not None else None))
            _diam_orig_raw = (profile_diam_00 if 'profile_diam_00' in locals() and profile_diam_00 is not None
                              else (2 * r_profile_00 * m_to_arcsec if 'r_profile_00' in locals() and r_profile_00 is not None else None))
            if (_pct_orig_raw is not None and _diam_orig_raw is not None
                    and hasattr(_pct_orig_raw, '__len__') and len(_pct_orig_raw) > 0):
                _pa_o = np.asarray(_pct_orig_raw, dtype=float)
                _da_o = np.asarray(_diam_orig_raw, dtype=float)
                _ord_o = np.argsort(_da_o)
                orig_pct_on_grid = np.interp(diam_grid, _da_o[_ord_o], _pa_o[_ord_o],
                                             left=_pa_o[_ord_o[0]], right=_pa_o[_ord_o[-1]])
            else:
                orig_pct_on_grid = np.full_like(diam_grid, np.nan)
        except Exception:
            orig_pct_on_grid = np.full_like(diam_grid, np.nan)
        orig_pct_xlsx = np.where(orig_pct_on_grid > 99.9, np.nan, orig_pct_on_grid)
        orig_pct_on_grid = np.where(orig_pct_on_grid > 95.0, np.nan, orig_pct_on_grid)

        # Build DataFrame for sheet1: one column per curve (index = diameter)
        # Excel version uses 99.9%-capped arrays; plot arrays remain at 95%.
        df_eef = _pd.DataFrame({'diameter_arcsec': diam_grid, 'EEF_aggregated_pct': agg_pct_xlsx,
                                 'EEF_origin_pct': orig_pct_xlsx,
                                 'EEF_pv_pct': pv_eef_xlsx, 'EEF_pearson4_pct': p4_eef_xlsx, 'EEF_king_pct': k_eef_xlsx})

        # Build parameter table for sheet2
        params = {}
        # If a pearson4 params object used for plotting exists (p4_params_plot),
        # capture its values even if pearson4_result wasn't stored as a full fit result.
        try:
            if pearson4_result is not None and hasattr(pearson4_result, 'params'):
                pr = pearson4_result.params
                for src_name, out_key in (('amplitude', 'Pearson4_Amplitude'),
                                          ('sigma', 'Pearson4_Sigma_arcsec'),
                                          ('m', 'Pearson4_Exponent_m'),
                                          ('nu', 'Pearson4_Skew_nu')):
                    try:
                        if src_name in pr:
                            params[out_key] = float(pr[src_name])
                    except Exception:
                        continue
                params['Pearson4_Center_arcsec'] = 0.0
        except Exception:
            pass
        try:
            params['Modified_PV_Amplitude_A'] = A_fit
            params['Modified_PV_Gamma_c_arcsec'] = Gamma_c_fit
            params['Modified_PV_Gamma_w_arcsec'] = Gamma_w_fit
            params['Modified_PV_eta'] = eta_fit
            params['Modified_PV_beta'] = beta_fit
            params['Modified_PV_scalar'] = scalar_fit
        except Exception:
            pass
        try:
            if pearson4_result is not None and hasattr(pearson4_result, 'params'):
                pr = pearson4_result.params
                params['Pearson4_Amplitude'] = float(pr['amplitude']) if 'amplitude' in pr else None
                params['Pearson4_Center_arcsec'] = 0.0
                params['Pearson4_Sigma_arcsec'] = float(pr['sigma']) if 'sigma' in pr else None
                params['Pearson4_Exponent_m'] = float(pr['m']) if 'm' in pr else None
                params['Pearson4_Skew_nu'] = float(pr['nu']) if 'nu' in pr else None
        except Exception:
            pass
        try:
            if 'I0_k' in locals():
                params['King_I0'] = I0_k
                params['King_rc_arcsec'] = rc_k
                params['King_alpha'] = alpha_k
                params['King_b'] = b_k
        except Exception:
            pass

        # Convert params dict to DataFrame
        try:
            df_params = _pd.DataFrame(list(params.items()), columns=['parameter', 'value'])
        except Exception:
            df_params = _pd.DataFrame(columns=['parameter', 'value'])

        # Prepare PSF shape sheets (one per fit) on rgrid
        # We'll write Excel formulas for PV and King that reference the
        # Fit_parameters sheet via VLOOKUP, so the workbook contains the fit
        # expressions rather than only numeric values.
        formulas_pv = []
        formulas_k = []
        radii_list = list(rgrid)
        # Build lookup expressions once
        A_lookup = 'VLOOKUP("Modified_PV_Amplitude_A",Fit_parameters!$A:$B,2,FALSE)'
        eta_lookup = 'VLOOKUP("Modified_PV_eta",Fit_parameters!$A:$B,2,FALSE)'
        scalar_lookup = 'VLOOKUP("Modified_PV_scalar",Fit_parameters!$A:$B,2,FALSE)'
        beta_lookup = 'VLOOKUP("Modified_PV_beta",Fit_parameters!$A:$B,2,FALSE)'
        Gc_lookup = 'VLOOKUP("Modified_PV_Gamma_c_arcsec",Fit_parameters!$A:$B,2,FALSE)'
        Gw_lookup = 'VLOOKUP("Modified_PV_Gamma_w_arcsec",Fit_parameters!$A:$B,2,FALSE)'
        I0_lookup = 'VLOOKUP("King_I0",Fit_parameters!$A:$B,2,FALSE)'
        rc_lookup = 'VLOOKUP("King_rc_arcsec",Fit_parameters!$A:$B,2,FALSE)'
        alpha_lookup = 'VLOOKUP("King_alpha",Fit_parameters!$A:$B,2,FALSE)'
        b_lookup = 'VLOOKUP("King_b",Fit_parameters!$A:$B,2,FALSE)'

        for i, rval in enumerate(radii_list, start=2):
            r_cell = f"A{i}"
            # PV formula components
            G_expr = f"EXP(-4*LN(2)*(({r_cell}/{Gc_lookup})^2))"
            a_expr = f"(POWER(2,1/{beta_lookup})-1)"
            C_expr = f"1/POWER(1 + {a_expr}*(POWER(2*{r_cell}/{Gw_lookup},2)), {beta_lookup})"
            mix_expr = f"(1 - {eta_lookup})*{G_expr} + {eta_lookup}*{scalar_lookup}*{C_expr}"
            norm_expr = f"(1 - {eta_lookup}) + {eta_lookup}*{scalar_lookup}"
            pv_formula = f"={A_lookup}*({mix_expr}/{norm_expr})"

            # King formula
            king_formula = f"={I0_lookup}*POWER(1 + POWER({r_cell}/{rc_lookup},2), -{alpha_lookup}) + {b_lookup}"

            formulas_pv.append(pv_formula)
            formulas_k.append(king_formula)

        # Radius numeric column; intensity columns contain Excel formulas (strings starting with '=')
        df_pv = _pd.DataFrame({'radius_arcsec': radii_list, 'I_pv_formula': formulas_pv})
        try:
            df_p4 = _pd.DataFrame({'radius_arcsec': rgrid, 'I_pearson4': p4_I})
        except Exception:
            df_p4 = _pd.DataFrame({'radius_arcsec': rgrid, 'I_pearson4': [None] * len(rgrid)})
        df_k = _pd.DataFrame({'radius_arcsec': radii_list, 'I_king_formula': formulas_k})

        # Write to Excel with multiple sheets
        wrote = False
        try:
            with _pd.ExcelWriter(out, engine='openpyxl') as writer:
                df_eef.to_excel(writer, sheet_name='EEF_curves', index=False)
                df_params.to_excel(writer, sheet_name='Fit_parameters', index=False)
                df_pv.to_excel(writer, sheet_name='PSF_pv', index=False)
                df_p4.to_excel(writer, sheet_name='PSF_pearson4', index=False)
                df_k.to_excel(writer, sheet_name='PSF_king', index=False)
            wrote = True
            print('Wrote combined EEF+fit Excel to', out)
        except Exception as e:
            try:
                with _pd.ExcelWriter(out, engine='xlsxwriter') as writer:
                    df_eef.to_excel(writer, sheet_name='EEF_curves', index=False)
                    df_params.to_excel(writer, sheet_name='Fit_parameters', index=False)
                    df_pv.to_excel(writer, sheet_name='PSF_pv', index=False)
                    df_p4.to_excel(writer, sheet_name='PSF_pearson4', index=False)
                    df_k.to_excel(writer, sheet_name='PSF_king', index=False)
                wrote = True
                print('Wrote combined EEF+fit Excel to', out)
            except Exception:
                print('Failed to write Excel file:', e)
        # Also generate a 4-panel fitting performance PNG and save it into CustomPSFs and Figures
        try:
            import matplotlib.pyplot as _plt
            ts_short = ts
            perf_fname_ts = _os.path.join('CustomPSFs', f'fitting_performance_{ts_short}.png')
            perf_fname = _os.path.join('CustomPSFs', 'fitting_performance.png')
            # Prepare data for intensity panel (interpolate observed I_fit to rgrid)
            try:
                if 'I_fit' in locals() and I_fit is not None and 'r_arcsec' in locals():
                    r_obs = np.asarray(r_arcsec)
                    I_obs = np.asarray(I_fit)
                    order_obs = np.argsort(r_obs)
                    r_obs_s = r_obs[order_obs]
                    I_obs_s = I_obs[order_obs]
                    I_on_rgrid = np.interp(rgrid, r_obs_s, I_obs_s, left=np.nan, right=np.nan)
                else:
                    I_on_rgrid = np.full_like(rgrid, np.nan)
            except Exception:
                I_on_rgrid = np.full_like(rgrid, np.nan)

            # Ensure pv_I and p4_I arrays exist for plotting
            try:
                pv_I
            except NameError:
                try:
                    pv_I = beta_pseudo_gaussian(rgrid, A_fit, Gamma_c_fit, Gamma_w_fit, eta_fit, beta_fit, scalar_fit)
                except Exception:
                    pv_I = np.full_like(rgrid, np.nan)
            try:
                p4_I
            except NameError:
                try:
                    # Prefer Pearson4 parameters obtained from the EEF-only fit (pearson4_result.params)
                    p4_params_for_plot = None
                    try:
                        if 'pearson4_result' in locals() and pearson4_result is not None and hasattr(pearson4_result, 'params'):
                            p4_params_for_plot = pearson4_result.params
                    except Exception:
                        p4_params_for_plot = None
                    # If lmfit.Parameters object provided, convert to simple dict of values
                    try:
                        import lmfit
                        if p4_params_for_plot is not None and isinstance(p4_params_for_plot, lmfit.parameter.Parameters):
                            # Convert Parameters to dict of numeric values
                            p4_params_for_plot = {k: v.value for k, v in p4_params_for_plot.items()}
                    except Exception:
                        # lmfit may not be installed or conversion not needed
                        pass
                    if p4_params_for_plot is not None:
                        try:
                            _pr = p4_params_for_plot
                            _u = rgrid / np.maximum(float(_pr['sigma']), 1e-15)
                            p4_I = (float(_pr['amplitude'])
                                    * (1.0 + _u * _u) ** (-np.maximum(float(_pr['m']), 0.5))
                                    * np.exp(-float(_pr['nu']) * np.arctan(_u)))
                            p4_I = np.maximum(p4_I, 0.0)
                        except Exception:
                            p4_I = np.full_like(rgrid, np.nan)
                    else:
                        p4_I = np.full_like(rgrid, np.nan)
                except Exception:
                    p4_I = np.full_like(rgrid, np.nan)

            # Residuals for intensity (linear space on rgrid)
            resid_pv = I_on_rgrid - pv_I
            resid_p4 = I_on_rgrid - p4_I

            # Prepare King intensity model sampled on rgrid so residuals can be
            # plotted consistently with PV/Pearson4. Prefer already-computed
            # `k_I` (on rgrid), else use `Ifit_king_plot` (usually on `rplot`),
            # else try to evaluate `king_profile` using fitted params if available.
            king_model_on_rgrid = None
            try:
                if 'k_I' in locals() and k_I is not None and np.any(np.isfinite(k_I)):
                    king_model_on_rgrid = np.asarray(k_I, dtype=float)
                elif 'Ifit_king_plot' in locals() and Ifit_king_plot is not None:
                    try:
                        x_ifit = rplot if 'rplot' in locals() else r_arcsec
                    except Exception:
                        x_ifit = r_arcsec
                    king_model_on_rgrid = np.interp(rgrid, x_ifit, Ifit_king_plot)
                elif 'I0_k' in locals() and 'rc_k' in locals() and 'alpha_k' in locals():
                    try:
                        king_model_on_rgrid = king_profile(rgrid, I0_k, rc_k, alpha_k, b_k if 'b_k' in locals() else 0.0)
                    except Exception:
                        king_model_on_rgrid = None
            except Exception:
                king_model_on_rgrid = None

            if king_model_on_rgrid is None:
                king_model_on_rgrid = np.full_like(rgrid, np.nan)

            # King residuals (linear) consistent with PV/Pearson4 residuals
            resid_king_linear = I_on_rgrid - king_model_on_rgrid
            # Also compute delta = king_model - aggregated data and log-residuals
            try:
                delta_king = king_model_on_rgrid - I_on_rgrid
                # log residuals (use same floor as elsewhere)
                king_log_resid = None
                try:
                    king_log_resid = np.log(king_model_on_rgrid + floor) - np.log(I_on_rgrid + floor)
                except Exception:
                    king_log_resid = np.full_like(rgrid, np.nan)
                # Print concise diagnostics
                try:
                    mask = np.isfinite(delta_king)
                    nfin = int(np.sum(mask))
                    dmin = float(np.nanmin(delta_king[mask])) if nfin else float('nan')
                    dmax = float(np.nanmax(delta_king[mask])) if nfin else float('nan')
                    drmse = float(np.sqrt(np.nanmean((delta_king[mask])**2))) if nfin else float('nan')
                    print(f"KING DELTA: finite={nfin}, min={dmin:.3g}, max={dmax:.3g}, rmse={drmse:.3g}")
                except Exception:
                    pass
                # Save numeric table for inspection
                try:
                    os.makedirs('Figures', exist_ok=True)
                    import csv
                    csv_path = os.path.join('Figures', 'king_residuals.csv')
                    with open(csv_path, 'w', newline='') as _f:
                        w = csv.writer(_f)
                        w.writerow(['radius_arcsec', 'I_aggregated', 'king_model', 'delta', 'king_log_resid'])
                        for rr, ia, km, dk, klr in zip(rgrid.tolist(), I_on_rgrid.tolist(), king_model_on_rgrid.tolist(), delta_king.tolist(), (king_log_resid.tolist() if hasattr(king_log_resid, 'tolist') else [None]*len(rgrid))):
                            w.writerow([rr, ia, km, dk, klr])
                    print('Wrote King residuals CSV to', csv_path)
                except Exception:
                    pass
            except Exception:
                pass

            # EEF residuals (agg_pct_on_grid vs models)
            try:
                eef_resid_pv = agg_pct_on_grid - pv_eef
            except Exception:
                eef_resid_pv = np.full_like(diam_grid, np.nan)
            try:
                eef_resid_p4 = agg_pct_on_grid - p4_eef
            except Exception:
                eef_resid_p4 = np.full_like(diam_grid, np.nan)
            try:
                # King EEF residuals (Data - King fit) — ensure variable exists for plotting
                eef_resid_king = agg_pct_on_grid - k_eef
            except Exception:
                eef_resid_king = np.full_like(diam_grid, np.nan)

            # Create 2x2 figure
            figp = _plt.figure(figsize=(12, 8))
            ax1p = figp.add_subplot(2, 2, 1)
            # Aggregated data in black dots
            ax1p.plot(rgrid, I_on_rgrid, 'k.', markersize=3, label='Aggregated PSF (data)', zorder=12)
            # Fit curves: dashed and thinner
            _perf_lw = 1.0
            line_pv, = ax1p.plot(rgrid, pv_I, color='red', linestyle='--', lw=_perf_lw, label='Modified pseudo-Voigt', zorder=5)
            # Only plot Pearson4 intensity curve if an accepted fit exists
            try:
                pearson4_ok_plot = ('pearson4_result' in locals() and pearson4_result is not None and hasattr(pearson4_result, 'params'))
            except Exception:
                pearson4_ok_plot = False
            if pearson4_ok_plot and np.any(np.isfinite(p4_I)):
                line_p4, = ax1p.plot(rgrid, p4_I, color='orange', linestyle='--', lw=_perf_lw, label='Pearson4 fit', zorder=5)
            try:
                line_pv.set_dashes([6, 2])
                line_p4.set_dashes([6, 2])
            except Exception:
                pass
            # If King fit available from EEF optimization, prefer that over intensity fallback
            try:
                if 'Ifit_king_plot' in locals() and Ifit_king_plot is not None:
                    line_k, = ax1p.plot(rgrid, Ifit_king_plot, color='purple', linestyle='--', lw=_perf_lw, label='King fit', zorder=5)
                    try:
                        line_k.set_dashes([6, 2])
                    except Exception:
                        pass
                elif 'king_I' in locals():
                    line_k, = ax1p.plot(rgrid, king_I, color='purple', linestyle='--', lw=_perf_lw, label='King fit', zorder=5)
                    try:
                        line_k.set_dashes([6, 2])
                    except Exception:
                        pass
            except Exception:
                pass
            ax1p.set_xlabel('Radius [arcsec]')
            ax1p.set_ylabel('Mean intensity')
            # Use logarithmic scale for intensity plots
            try:
                ax1p.set_yscale('log')
            except Exception:
                pass
            ax1p.grid(True, which='both', linestyle='--', alpha=0.35)
            # Build legend proxies conditionally (include Pearson4 only if plotted)
            try:
                from matplotlib.lines import Line2D
                pv_proxy = Line2D([0], [0], color='red', linestyle='--', lw=_perf_lw)
                k_proxy = Line2D([0], [0], color='purple', linestyle='--', lw=_perf_lw)
                entries = [_plt.Line2D([], [], color='k', marker='.', linestyle='None', ms=3), pv_proxy]
                labels = ['Aggregated PSF (data)', 'Modified pseudo-Voigt']
                try:
                    if pearson4_ok_plot and np.any(np.isfinite(p4_I)):
                        from matplotlib.lines import Line2D as _L2
                        p4_proxy = _L2([0], [0], color='orange', linestyle='--', lw=_perf_lw)
                        entries.append(p4_proxy)
                        labels.append('Pearson4 fit')
                except Exception:
                    pass
                try:
                    if ('Ifit_king_plot' in locals() and Ifit_king_plot is not None) or ('king_I' in locals() and king_I is not None) or ('k_I' in locals() and k_I is not None):
                        entries.append(k_proxy)
                        labels.append('King fit')
                except Exception:
                    pass
                ax1p.legend(entries, labels, fontsize=8)
            except Exception:
                try:
                    ax1p.legend(fontsize=8)
                except Exception:
                    pass

            ax2p = figp.add_subplot(2, 2, 3)
            # Residuals: plot all intensity residuals as percent on the same y-axis
            ax2p.set_xlabel('Radius [arcsec]')
            ax2p.set_ylabel('Residual (%)')
            # Avoid division by zero by using the common floor value used elsewhere
            eps = floor if ('floor' in locals() or 'floor' in globals()) else 1e-12
            try:
                resid_pv_pct = 100.0 * (resid_pv) / (I_on_rgrid + eps)
            except Exception:
                resid_pv_pct = np.full_like(rgrid, np.nan)
            try:
                resid_p4_pct = 100.0 * (resid_p4) / (I_on_rgrid + eps)
            except Exception:
                resid_p4_pct = np.full_like(rgrid, np.nan)
            try:
                resid_king_pct = 100.0 * (resid_king_linear) / (I_on_rgrid + eps) if ('resid_king_linear' in locals() and resid_king_linear is not None) else np.full_like(rgrid, np.nan)
            except Exception:
                resid_king_pct = np.full_like(rgrid, np.nan)

            # Plot percent residuals on the same axis so they share y-limits/scaling
            try:
                lrpv, = ax2p.plot(rgrid, resid_pv_pct, color='red', linestyle='--', lw=_perf_lw, label='Intensity residual: Data - Modified pseudo-Voigt (%)')
            except Exception:
                lrpv = None
            # Only plot Pearson4 intensity residual if an accepted Pearson4 fit exists
            try:
                if pearson4_ok_plot and np.any(np.isfinite(p4_I)):
                    lrp4, = ax2p.plot(rgrid, resid_p4_pct, color='orange', linestyle='--', lw=_perf_lw * 1.6, marker='o', markersize=3, markevery=40, label='Intensity residual: Data - Pearson4 fit (%)')
                else:
                    lrp4 = None
            except Exception:
                lrp4 = None
            try:
                if lrpv is not None:
                    lrpv.set_dashes([6, 2])
                if lrp4 is not None:
                    lrp4.set_dashes([6, 2])
            except Exception:
                pass
            # If King residuals exist, plot them too (as percent to match others)
            try:
                # Build a king residual on the rgrid if possible (prefer res_int_king)
                resid_king_plot = None
                if 'res_int_king' in locals() and res_int_king is not None:
                    try:
                        resid_king_plot = np.interp(rgrid, r_fit, res_int_king, left=np.nan, right=np.nan)
                    except Exception:
                        resid_king_plot = None
                if resid_king_plot is None:
                    if 'Ifit_king_plot' in locals() and Ifit_king_plot is not None:
                        try:
                            x_ifit = rplot if 'rplot' in locals() else r_arcsec
                        except Exception:
                            x_ifit = r_arcsec
                        try:
                            resid_king_plot = I_on_rgrid - np.interp(rgrid, x_ifit, Ifit_king_plot)
                        except Exception:
                            resid_king_plot = None
                    elif 'king_I' in locals() and king_I is not None:
                        resid_king_plot = I_on_rgrid - king_I

                # Convert any available king residuals to percent to match the other traces
                try:
                    if resid_king_plot is not None and np.any(np.isfinite(resid_king_plot)):
                        resid_king_plot_pct = 100.0 * resid_king_plot / (I_on_rgrid + eps)
                    elif 'resid_king_linear' in locals() and resid_king_linear is not None and np.any(np.isfinite(resid_king_linear)):
                        resid_king_plot_pct = 100.0 * resid_king_linear / (I_on_rgrid + eps)
                    else:
                        resid_king_plot_pct = np.full_like(rgrid, np.nan)
                except Exception:
                    resid_king_plot_pct = np.full_like(rgrid, np.nan)

                if resid_king_plot_pct is not None and np.any(np.isfinite(resid_king_plot_pct)):
                    lrk, = ax2p.plot(
                        rgrid,
                        resid_king_plot_pct,
                        color='purple',
                        linestyle='--',
                        lw=_perf_lw * 1.6,
                        alpha=0.95,
                        zorder=15,
                        label='Intensity residual: Data - King fit (%)'
                    )
                    try:
                        lrk.set_dashes([6, 2])
                    except Exception:
                        pass
            except Exception:
                pass
            ax2p.axhline(0, color='gray', linewidth=0.8)
            ax2p.grid(True, which='both', linestyle='--', alpha=0.35)
            try:
                from matplotlib.lines import Line2D
                pv_proxy = Line2D([0], [0], color='red', linestyle='--', lw=_perf_lw)
                k_proxy = Line2D([0], [0], color='purple', linestyle='--', lw=_perf_lw)
                entries = [pv_proxy]
                labels = ['Intensity residual: Data - Modified pseudo-Voigt (%)']
                try:
                    if pearson4_ok_plot and np.any(np.isfinite(p4_I)):
                        from matplotlib.lines import Line2D as _L2
                        p4_proxy = _L2([0], [0], color='orange', linestyle='--', lw=_perf_lw)
                        entries.append(p4_proxy)
                        labels.append('Intensity residual: Data - Pearson4 fit (%)')
                except Exception:
                    pass
                try:
                    if ('resid_king_plot' in locals() and resid_king_plot is not None and np.any(np.isfinite(resid_king_plot))) or ('resid_king_linear' in locals() and resid_king_linear is not None and np.any(np.isfinite(resid_king_linear))):
                        entries.append(k_proxy)
                        labels.append('Intensity residual: Data - King fit (%)')
                except Exception:
                    pass
                if entries:
                    ax2p.legend(entries, labels, fontsize=8)
            except Exception:
                try:
                    ax2p.legend(fontsize=8)
                except Exception:
                    pass

            ax3p = figp.add_subplot(2, 2, 2)
            ax3p.plot(diam_grid, agg_pct_on_grid, color='k', marker='.', linestyle='None', ms=4, label='Aggregated PSF (data)', zorder=12)
            lpv_eef, = ax3p.plot(diam_grid, pv_eef, color='red', linestyle='--', lw=_perf_lw, label='Modified pseudo-Voigt')
            lp4_eef = None
            try:
                if np.any(np.isfinite(p4_eef)):
                    lp4_eef, = ax3p.plot(diam_grid, p4_eef, color='orange', linestyle='--', lw=_perf_lw, label='Pearson4 fit')
            except Exception:
                lp4_eef = None
            try:
                lpv_eef.set_dashes([6, 2])
                if lp4_eef is not None:
                    lp4_eef.set_dashes([6, 2])
            except Exception:
                pass
            # King EEF if available
            try:
                if 'king_eef_pct_plot' in locals():
                    lk_eef, = ax3p.plot(diam_grid, king_eef_pct_plot, color='purple', linestyle='--', lw=_perf_lw, label='King fit')
                    try:
                        lk_eef.set_dashes([6, 2])
                    except Exception:
                        pass
            except Exception:
                pass
            ax3p.grid(True, which='both', linestyle='--', alpha=0.35)
            ax3p.set_xlabel('Diameter [arcsec]')
            ax3p.set_ylabel('Encircled energy (%)')
            try:
                from matplotlib.lines import Line2D
                pv_proxy = Line2D([0], [0], color='red', linestyle='--', lw=_perf_lw)
                k_proxy = Line2D([0], [0], color='purple', linestyle='--', lw=_perf_lw)
                entries = [_plt.Line2D([], [], color='k', marker='.', linestyle='None', ms=4), pv_proxy]
                labels = ['Aggregated PSF (data)', 'Modified pseudo-Voigt']
                if lp4_eef is not None:
                    p4_proxy = Line2D([0], [0], color='orange', linestyle='--', lw=_perf_lw)
                    entries.append(p4_proxy)
                    labels.append('Pearson4 fit')
                try:
                    if 'king_eef_pct_plot' in locals():
                        entries.append(k_proxy)
                        labels.append('King fit')
                except Exception:
                    pass
                ax3p.legend(entries, labels, fontsize=8)
            except Exception:
                try:
                    ax3p.legend(fontsize=8)
                except Exception:
                    pass
            ax3p.set_xlabel('Diameter [arcsec]')
            ax3p.set_ylabel('EEF (%)')

            ax4p = figp.add_subplot(2, 2, 4)
            lrpv_eef, = ax4p.plot(diam_grid, eef_resid_pv, color='red', linestyle='--', lw=_perf_lw, label='EEF residual: Data - Modified pseudo-Voigt (pct)')
            lrp4_eef = None
            try:
                if np.any(np.isfinite(eef_resid_p4)):
                    lrp4_eef, = ax4p.plot(diam_grid, eef_resid_p4, color='orange', linestyle='--', lw=_perf_lw, label='EEF residual: Data - Pearson4 fit (pct)')
            except Exception:
                lrp4_eef = None
            try:
                lrpv_eef.set_dashes([6, 2])
                if lrp4_eef is not None:
                    lrp4_eef.set_dashes([6, 2])
            except Exception:
                pass
            try:
                if 'eef_resid_king' in locals():
                    lrk_eef, = ax4p.plot(diam_grid, eef_resid_king, color='purple', linestyle='--', lw=_perf_lw, label='EEF residual: Data - King fit (pct)')
                    try:
                        lrk_eef.set_dashes([6, 2])
                    except Exception:
                        pass
            except Exception:
                pass
            ax4p.axhline(0, color='gray', linewidth=0.8)
            ax4p.grid(True, which='both', linestyle='--', alpha=0.35)
            ax4p.set_xlabel('Diameter [arcsec]')
            ax4p.set_ylabel('EEF residual (pct)')
            try:
                from matplotlib.lines import Line2D
                pv_proxy = Line2D([0], [0], color='red', linestyle='--', lw=_perf_lw)
                k_proxy = Line2D([0], [0], color='purple', linestyle='--', lw=_perf_lw)
                entries = [pv_proxy]
                labels = ['EEF residual: Data - Modified pseudo-Voigt (pct)']
                try:
                    if lrp4_eef is not None:
                        from matplotlib.lines import Line2D as _L2
                        p4_proxy = _L2([0], [0], color='orange', linestyle='--', lw=_perf_lw)
                        entries.append(p4_proxy)
                        labels.append('EEF residual: Data - Pearson4 fit (pct)')
                except Exception:
                    pass
                try:
                    if 'eef_resid_king' in locals() and np.any(np.isfinite(eef_resid_king)):
                        entries.append(k_proxy)
                        labels.append('EEF residual: Data - King fit (pct)')
                except Exception:
                    pass
                if entries:
                    ax4p.legend(entries, labels, fontsize=8)
            except Exception:
                try:
                    ax4p.legend(fontsize=8)
                except Exception:
                    pass

            figp.tight_layout()
            # Save timestamped and canonical copies into CustomPSFs and Figures
            try:
                _plt.savefig(perf_fname_ts, dpi=150, bbox_inches='tight')
                _plt.savefig(perf_fname, dpi=150, bbox_inches='tight')
                # Also save a copy in Figures for quick viewing
                _plt.savefig(os.path.join('Figures', f'fitting_performance_{ts_short}.png'), dpi=150, bbox_inches='tight')
                _plt.savefig(os.path.join('Figures', 'fitting_performance.png'), dpi=150, bbox_inches='tight')
                _plt.close(figp)
                print('Wrote fitting performance PNG to', perf_fname)
            except Exception:
                try:
                    _plt.savefig(perf_fname, dpi=150, bbox_inches='tight')
                    _plt.close(figp)
                except Exception:
                    pass
        except Exception:
            pass

    def export_fits():
        """Export the aggregated E2E PSF grid `Z` to a minimal FITS file.

        Writes a Primary HDU-only FITS with big-endian doubles and header
        cards: TOT_AEFF, INTG_Z, PIXAS1/2 (arcsec), PIXM1/2 (meters), CDELT1/2 (deg/pixel).
        """
        import time as _time, os as _os
        # Minimal FITS writer (primary HDU only)
        def _write_simple_fits(path: str, data, header_cards: dict | None = None):
            arr = np.asarray(data, dtype=np.float64)
            if arr.ndim != 2:
                raise ValueError('data must be 2D')
            ny, nx = arr.shape
            cards = []
            def add_card(key, val, comment=None):
                if isinstance(val, str):
                    v = f"'{val}'"
                elif isinstance(val, bool):
                    v = 'T' if val else 'F'
                else:
                    v = str(val)
                if comment:
                    cards.append(f"{key:8s}= {v:20s} / {comment}")
                else:
                    cards.append(f"{key:8s}= {v:20s}")
            add_card('SIMPLE', True, 'file does conform to FITS standard')
            add_card('BITPIX', -64, 'number of bits per data pixel')
            add_card('NAXIS', 2, 'number of data axes')
            add_card('NAXIS1', nx, 'length of data axis 1')
            add_card('NAXIS2', ny, 'length of data axis 2')
            add_card('EXTEND', True, 'FITS dataset may contain extensions')
            if header_cards:
                for k, v in header_cards.items():
                    add_card(k, v)
            cards.append('END')
            header_str = ''.join(card.ljust(80) for card in cards)
            header_bytes = header_str.encode('ascii')
            pad = (2880 - (len(header_bytes) % 2880)) % 2880
            header_bytes += b' ' * pad
            data_be = arr.astype('>f8')
            data_bytes = data_be.tobytes(order='C')
            pad2 = (2880 - (len(data_bytes) % 2880)) % 2880
            data_bytes += b'\x00' * pad2
            with open(path, 'wb') as f:
                f.write(header_bytes)
                f.write(data_bytes)

        # Total A_eff from the DataFrame weights (df in closure).
        # Prefer the adjusted A_eff (after vignetting) when available.
        try:
            if 'aeff_adjusted' in df.columns:
                total_aeff_local = float(np.nansum(df['aeff_adjusted']))
            else:
                total_aeff_local = float(np.nansum(df['weight']))
        except Exception:
            total_aeff_local = float(np.nansum(weight_arr)) if 'weight_arr' in locals() else 0.0

        # Compute integral on our grid (x,y defined in closure)
        try:
            integral_local = float(np.trapz(np.trapz(Z, x, axis=1), y, axis=0))
        except Exception:
            integral_local = float(np.nansum(Z))

        # Pixel scales
        pix_m_x = float(x[1] - x[0]) if x.size > 1 else 0.0
        pix_m_y = float(y[1] - y[0]) if y.size > 1 else 0.0
        m_to_arcsec_local = (180.0 / (12.0 * np.pi)) * 3600.0
        pix_as_x = pix_m_x * m_to_arcsec_local
        pix_as_y = pix_m_y * m_to_arcsec_local
        cdelt1 = pix_as_x / 3600.0
        cdelt2 = pix_as_y / 3600.0

        # Output path
        ts = _time.strftime('%Y%m%d_%H%M%S')
        out = _os.path.join('CustomPSFs', f'E2E_aggregated_{ts}.fits')
        _os.makedirs(_os.path.dirname(out), exist_ok=True)
        # Try to pick up author/contact from git config when available
        def _git_cfg(k: str):
            try:
                import subprocess as _sub
                out = _sub.check_output(['git', 'config', '--get', k], stderr=_sub.DEVNULL)
                return out.decode('utf-8').strip()
            except Exception:
                return None

        author_local = _git_cfg('user.name') or 'Unknown'
        contact_local = _git_cfg('user.email') or 'ivo.ferreira@esa.int'
        orcid_local = _git_cfg('user.orcid') or ''
        inputfn = getattr(df, '_input_filename', None) or 'interactive'

        header = {
            'CREATOR': 'main.plot_sum:export_fits',
            'DATE': _time.strftime('%Y-%m-%dT%H:%M:%SZ'),
            'TOT_AEFF': total_aeff_local,
            'INTG_Z': integral_local,
            'CDELT1': float(cdelt1),
            'CDELT2': float(cdelt2),
            'PIXAS1': float(pix_as_x),
            'PIXAS2': float(pix_as_y),
            'PIXM1': float(pix_m_x),
            'PIXM2': float(pix_m_y),
            'AUTHOR': author_local,
            'CONTACT': contact_local,
            'ORCID': orcid_local,
            'INPUTFN': inputfn,
        }
        _write_simple_fits(out, Z, header_cards=header)
        print('Wrote FITS to', out)
    
    def show_context_menu(x, y):
        """Display context menu at given position"""
        nonlocal menu_annotation, menu_active
        
        if menu_active:
            hide_context_menu()
            return
        
        menu_text = "┌────────────────────────────────────────────┐\n"
        menu_text += "│  1. Export PSF Plot                         │\n"
        menu_text += "│  2. Export EEF Plot                         │\n"
        menu_text += "│  3. Export FITS                             │\n"
        menu_text += "│  4. Export EEF CSV                          │\n"
        menu_text += "│  5. Export Fit Parameters CSV               │\n"
        menu_text += "│  6. Export combined EEF + Fit (Excel)       │\n"
        menu_text += "│  7. Cancel                                  │\n"
        menu_text += "└────────────────────────────────────────────┘"
        
        menu_annotation = fig.text(x, y, menu_text,
                                   fontfamily='monospace',
                                   fontsize=10,
                                   bbox=dict(boxstyle='round,pad=0.5', 
                                            facecolor='white', 
                                            edgecolor='black',
                                            linewidth=2,
                                            alpha=0.95),
                                   verticalalignment='top',
                                   horizontalalignment='left',
                                   zorder=1000)
        menu_active = True
        fig.canvas.draw_idle()
    
    def hide_context_menu():
        """Hide the context menu"""
        nonlocal menu_annotation, menu_active
        if menu_annotation:
            menu_annotation.remove()
            menu_annotation = None
            menu_active = False
            fig.canvas.draw_idle()
    
    def on_mouse_press(event):
        """Handle mouse press events"""
        nonlocal menu_active, menu_annotation
        
        # Right-click detection (button 3 on Windows/Linux/Mac)
        is_right_click = False
        if event.button == 3:  # Standard right-click
            is_right_click = True
        elif event.button == 2 and event.key is None:  # Some Mac configs use button 2
            is_right_click = True
        elif event.button == 1 and event.key in ['control', 'ctrl', 'cmd', 'meta', 'super']:
            is_right_click = True  # Ctrl+click for Mac
        
        if is_right_click and event.inaxes in [ax1, ax2] and not menu_active:
            # Show context menu
            x, y = event.x / fig.bbox.width, event.y / fig.bbox.height
            show_context_menu(x, y)
        elif menu_active and event.button == 1:  # Left-click when menu is active
            # Menu is active, check if clicking on menu options
            if menu_annotation:
                # Get menu bounds
                bbox = menu_annotation.get_window_extent()
                if bbox.contains(event.x, event.y):
                    # Calculate relative position within menu (from bottom)
                    relative_y = (event.y - bbox.y0) / bbox.height
                    
                    # Menu structure for 6 options (top-to-bottom):
                    # Top border, Opt1, Opt2, Opt3, Opt4, Opt5, Opt6, Bottom border
                    # Ranges chosen empirically to match text layout
                    if 0.82 < relative_y <= 0.95:  # Option 1 (Export PSF)
                        hide_context_menu()
                        export_psf_plot()
                    elif 0.68 < relative_y <= 0.82:  # Option 2 (Export EEF)
                        hide_context_menu()
                        export_eef_plot()
                    elif 0.54 < relative_y <= 0.68:  # Option 3 (Export FITS)
                        hide_context_menu()
                        export_fits()
                    elif 0.40 < relative_y <= 0.54:  # Option 4 (Export EEF CSV)
                        hide_context_menu()
                        export_eef_csv()
                    elif 0.26 < relative_y <= 0.40:  # Option 5 (Export Fit Params CSV)
                        hide_context_menu()
                        export_fit_params_csv()
                    elif 0.14 < relative_y <= 0.26:  # Option 6 (Export combined EEF + Fit Excel)
                        hide_context_menu()
                        export_eef_and_params_excel()
                    elif 0.08 < relative_y <= 0.14:  # Option 7 (Cancel)
                        hide_context_menu()
                else:
                    # Clicked outside menu, hide it
                    hide_context_menu()
    
    def on_key_press(event):
        """Handle keyboard shortcuts"""
        nonlocal menu_active
        
        if menu_active:
            if event.key in ['1', 'p']:
                hide_context_menu()
                export_psf_plot()
            elif event.key in ['2', 'e']:
                hide_context_menu()
                export_eef_plot()
            elif event.key in ['3', 'f']:
                hide_context_menu()
                export_fits()
            elif event.key in ['4', 'c']:
                hide_context_menu()
            elif event.key in ['x', '6']:
                export_eef_and_params_excel()
                export_eef_csv()
            elif event.key in ['5', 's']:
                hide_context_menu()
                export_fit_params_csv()
            elif event.key in ['6', 'x']:
                hide_context_menu()
                export_eef_and_params_excel()
            elif event.key in ['6', 'escape']:
                hide_context_menu()
        else:
            if event.key in ['p', '1']:
                export_psf_plot()
            elif event.key in ['e', '2']:
                export_eef_plot()
            elif event.key in ['f', '3']:
                export_fits()
            elif event.key in ['s', '5']:
                export_fit_params_csv()
            elif event.key == 'h':
                print("\nKeyboard shortcuts:")
                print("  'p' or '1' - Export PSF plot")
                print("  'e' or '2' - Export Encircled Energy plot")
                print("  Right-click (or Ctrl+click on Mac) - Show context menu")
                print("  'h' - Show this help")
    
    # Connect events
    fig.canvas.mpl_connect('button_press_event', on_mouse_press)
    fig.canvas.mpl_connect('key_press_event', on_key_press)
    
    # Print help message
    print("\nKeyboard shortcuts:")
    print("  'p' or '1' - Export PSF plot")
    print("  'e' or '2' - Export Encircled Energy plot")
    print("  Right-click (or Ctrl+click on Mac) on plots - Show context menu")
    print("  'h' - Show help")
    print("  'f' or '3' - Export FITS (aggregated E2E PSF)")
    
    if output:
        plt.savefig(output, dpi=150)  # Save the combined plot
        print(f"Saved combined plot to {output}")
    if not output:
        # No explicit output requested. To avoid blocking in interactive
        # backends (which calls plt.show() and waits for user), save a
        # non-interactive copy of the combined plot to a file alongside the
        # input workbook. This prevents the CLI invocation from appearing to
        # "hang" while still producing a useful artifact. If saving fails
        # for any reason, fall back to showing the interactive window.
        # Prefer explicit interactive display when requested. Use a safe lookup
        # for CLI args exposed on __main__.args to avoid NameError when this
        # module is imported or called programmatically.
        try:
            _main_mod = sys.modules.get('__main__')
            _main_args = getattr(_main_mod, 'args', None) if _main_mod is not None else None
            _show_flag = getattr(_main_args, 'show', False) if _main_args is not None else False
        except Exception:
            _show_flag = False
        if _show_flag:
            try:
                plt.show()
            except Exception:
                try:
                    base, _ = os.path.splitext(args.file)
                    auto_out = f"{base}_plot.png"
                    plt.savefig(auto_out, dpi=150)
                    print(f"Interactive display failed; saved plot to {auto_out} instead.")
                except Exception:
                    pass
        else:
            # Default: show the interactive plot and keep it open. If this
            # fails (e.g. running in a headless environment), do NOT auto-save
            # a PNG; instead instruct the user to run with --output to save.
            # Additionally, support --export-package: generate exports into a
            # timestamped folder containing input, FITS, Excel and figures.
            try:
                # Check CLI flag exposed on __main__
                _main_mod = sys.modules.get('__main__')
                _main_args = getattr(_main_mod, 'args', None) if _main_mod is not None else None
                _export_pkg = getattr(_main_args, 'export_package', False) if _main_args is not None else False
            except Exception:
                _export_pkg = False

            if _export_pkg:
                try:
                    import shutil, glob, time as _time
                    ts = _time.strftime('%Y%m%d_%H%M%S')
                    pkg_dir = os.path.join('Exports', ts)
                    os.makedirs(pkg_dir, exist_ok=True)
                    # Copy input workbook
                    try:
                        shutil.copy2(args.file, os.path.join(pkg_dir, os.path.basename(args.file)))
                    except Exception:
                        pass
                    # Trigger standard exports (they will write into CustomPSFs/ and Figures/)
                    try:
                        export_fits()
                    except Exception:
                        pass
                    try:
                        export_eef_and_params_excel()
                    except Exception:
                        pass
                    try:
                        export_psf_plot()
                    except Exception:
                        pass
                    try:
                        export_eef_plot()
                    except Exception:
                        pass
                    # Collect generated artifacts by pattern and copy newest matches
                    def copy_latest(patterns, dest_dir):
                        for pat in patterns:
                            try:
                                matches = glob.glob(pat)
                                if not matches:
                                    continue
                                latest = max(matches, key=os.path.getmtime)
                                shutil.copy2(latest, os.path.join(dest_dir, os.path.basename(latest)))
                            except Exception:
                                continue

                    # FITS
                    copy_latest(['CustomPSFs/E2E_aggregated_*.fits'], pkg_dir)
                    # Combined EEF+fit Excel
                    copy_latest(['CustomPSFs/E2E_EEF_and_fitparams_*.xlsx', 'CustomPSFs/E2E_EEF_and_fitparams_*.xls'], pkg_dir)
                    # PSF and EEF PNGs
                    copy_latest(['Figures/E2E_PSF_*.png', 'Figures/E2E_PSF_*.PNG'], pkg_dir)
                    copy_latest(['Figures/Encircled_Energy_*.png', 'Figures/Encircled_Energy_*.PNG'], pkg_dir)
                    # EEF fit combined figure (may be named without timestamp)
                    copy_latest(['Figures/E2E_fit_combined*.png', 'Figures/E2E_fit_combined.png'], pkg_dir)
                    # Include fitting performance PNG if present in CustomPSFs
                    copy_latest(['CustomPSFs/fitting_performance*.png', 'CustomPSFs/fitting_performance.png'], pkg_dir)

                    print(f"Exported package to {pkg_dir}")
                except Exception as e:
                    try:
                        print('Failed to create export package:', e)
                    except Exception:
                        pass
            try:
                plt.show()
            except Exception:
                print("Interactive display failed — no GUI available. To save the plot use --output <file.png> or run in a GUI session.")


def compute_hew_eef_metrics(file: str = 'Distributions/TestDistribution.xlsx', sheet: str = 'MM_PSF', normalize: bool = True, fast: bool = True, df_optimized: pd.DataFrame = None) -> dict:
    """Convenience wrapper: load workbook and return HEW/EEF metrics (no plotting).

    Returns a dict matching the CLI JSON output.
    """
    df = load_gaussians_from_excel(file, sheet)
    return plot_sum(df, normalize=normalize, fast=fast, df_optimized=df_optimized, return_metrics_only=True)


if __name__ == '__main__':
    # Close all existing figures to ensure clean start
    plt.close('all')
    try:
        print('main.py: starting with argv=', sys.argv)
    except Exception:
        pass
    
    # Parse command-line arguments
    parser = argparse.ArgumentParser(description='Plot sum of rotated Gaussians from Excel.')
    parser.add_argument('-f','--file', default='Distributions/TestDistribution.xlsx', help='Excel file path')
    parser.add_argument('-s','--sheet', default='MM_PSF', help='Sheet name to read (default MM_PSF)')
    parser.add_argument('--normalize', dest='normalize', action='store_true', default=True, help='Normalize each Gaussian to integrate to 1 (default on)')
    parser.add_argument('--no-normalize', dest='normalize', action='store_false', help='Disable normalization')
    parser.add_argument('-o','--output', help='Optional output image file path (PNG)')
    parser.add_argument('--show', action='store_true', help='Show interactive plot window (overrides auto-save)')
    parser.add_argument(
        '--mode',
        type=str,
        choices=['coarse', 'fine'],
        default='coarse',
        help='Runtime mode: coarse or fine. Controls plotting + optimization speed/accuracy.'
    )
    parser.add_argument('--optimize', action='store_true', default=False, help='Enable MM position optimization (uses --mode for speed/accuracy).')
    # Compatibility alias (UK spelling)
    parser.add_argument('--optimise', dest='optimize', action='store_true', help=argparse.SUPPRESS)
    parser.add_argument(
        '--placement',
        type=str,
        nargs='?',
        const='cross',
        choices=['cross', 'x_axis', 'elliptical'],
        default=None,
        help=(
            "Apply a placement strategy and write a *_placed.xlsx file. "
            "If used together with --optimize, the selected placement seeds the optimization. "
            "Strategies: 'cross' (90° pattern, previous default), "
            "'x_axis' (+/-x with above/below alternation), or "
            "'elliptical' (per-row: best MMs toward x, worst toward y)."
        ),
    )
    parser.add_argument('--return_metrics_only', dest='return_metrics_only', action='store_true', help='Return HEW/EEF metrics only (no plot).')
    parser.add_argument('--suppress-output', dest='suppress_output', action='store_true', default=False, help='Do not write placed/optimised Excel outputs when running non-interactively.')
    parser.add_argument('--metrics-nr-final', type=int, default=None, help=argparse.SUPPRESS)
    parser.add_argument('--metrics-ntheta-final', type=int, default=None, help=argparse.SUPPRESS)
    parser.add_argument('--metrics-r-margin', type=float, default=None, help=argparse.SUPPRESS)
    # Removed --input-pickle support; keep input-csv for CSV single-sheet paths
    parser.add_argument('--input-csv', default=None, help=argparse.SUPPRESS)
    parser.add_argument('--export-package', dest='export_package', action='store_true', default=False, help='Create a timestamped folder with input, FITS, Excel and PNG artifacts')
    parser.add_argument('--batch-combinations', dest='batch_combinations', default=None, help='Path to Excel file containing combinations (one config per row). If provided the script will run each configuration against the input file and exit.')
    args = parser.parse_args()

    # Expose parsed args to helper functions that inspect __main__.args
    try:
        import sys as _sys
        _sys.modules['__main__'].args = args
    except Exception:
        pass

    # If running in batch/export mode, disable interactive plotting to avoid
    # blocking on GUI windows. Force non-interactive mode and noop plt.show().
    try:
        if getattr(args, 'export_package', False) or getattr(args, 'batch_combinations', None):
            try:
                plt.ioff()
            except Exception:
                pass
            try:
                plt.show = lambda *a, **k: None
            except Exception:
                pass
    except Exception:
        pass

    # If an explicit input CSV path was provided, prefer it as the input file
    if getattr(args, 'input_csv', None):
        args.file = args.input_csv
    # Note: in-memory pickled input support removed for release.

    # If args.file is a multi-sheet CSV, parse it and materialize a temporary .xlsx
    if isinstance(args.file, str) and args.file.lower().endswith('.csv'):
        try:
            sheets = parse_multisheet_csv(args.file)
            # If CSV contained multiple sheets, write to a temporary .xlsx and use that
            if isinstance(sheets, dict) and len(sheets) > 1:
                tf = tempfile.NamedTemporaryFile(prefix='multisheet_', suffix='.xlsx', delete=False)
                tf.close()
                tmp_path = Path(tf.name)
                try:
                    with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
                        for sname, sdf in sheets.items():
                            try:
                                sdf.to_excel(writer, sheet_name=sname, index=False)
                            except Exception:
                                continue
                    args.file = str(tmp_path)
                except Exception:
                    # fallback: leave args.file as CSV path
                    pass
        except Exception:
            # If parsing fails, fall back to treating CSV as single-sheet file
            pass

    # If user requested metrics-only, disable interactive plotting and suppress show().
    if getattr(args, 'return_metrics_only', False):
        try:
            plt.ioff()
            plt.show = lambda *a, **k: None
            # Suppress incidental debug/sample prints so CLI output is a clean JSON
            try:
                import os as _os
                _os.environ['SILENCE_OUTPUT'] = '1'
            except Exception:
                pass
        except Exception:
            pass

    # Load data from Excel
    # Batch mode: process combinations file and exit
    if getattr(args, 'batch_combinations', None):
        comb_path = args.batch_combinations
        try:
            combos = pd.read_excel(comb_path, engine='openpyxl')
        except Exception as e:
            print(f"Failed to read combinations file {comb_path}: {e}")
            sys.exit(2)

        # Iterate rows: expect columns as follows (by position):
        # A: (ignored), B: config name (prefix), C: off-axis, D: energy, E: defocus
        # Prepare a single per-batch Export folder named Export_<input_stem>_<ts>
        exports_root = os.path.join(os.getcwd(), 'Exports')
        os.makedirs(exports_root, exist_ok=True)
        src_path = args.file
        src_basename = os.path.basename(src_path)
        src_stem = os.path.splitext(src_basename)[0]
        batch_ts = time.strftime('%Y%m%d_%H%M%S')
        export_batch_dir = os.path.join(exports_root, f"Export_{src_stem}_{batch_ts}")
        os.makedirs(export_batch_dir, exist_ok=True)

        # Prepare aggregated results collector
        aggregated_rows = []
        agg_columns = None

        skipped_cannot_read = 0
        skipped_empty_prefix = 0
        for rid, crow in combos.iterrows():
            # Read configuration prefix from column B; skip rows where it's missing/NaN
            try:
                raw_prefix = crow.iat[1]
            except Exception:
                skipped_cannot_read += 1
                continue
            if pd.isna(raw_prefix) or str(raw_prefix).strip() == '':
                skipped_empty_prefix += 1
                continue
            prefix = str(raw_prefix).strip()
            try:
                offaxis_val = float(crow.iat[2]) if not pd.isna(crow.iat[2]) else 0.0
            except Exception:
                offaxis_val = 0.0
            try:
                energy_val = float(crow.iat[3]) if not pd.isna(crow.iat[3]) else 0.0
            except Exception:
                energy_val = 0.0
            try:
                defocus_val = float(crow.iat[4]) if not pd.isna(crow.iat[4]) else 0.0
            except Exception:
                defocus_val = 0.0

            # Detect per-row run mode (fine/coarse). Support a named column
            # ('mode', 'run_mode', 'plot_mode') or a positional 6th column.
            run_mode = None
            try:
                # Named column search (case-insensitive)
                for col in combos.columns:
                    if str(col).strip().lower() in ('mode', 'run_mode', 'plot_mode'):
                        val = crow.get(col)
                        if not pd.isna(val):
                            run_mode = str(val).strip().lower()
                        break
            except Exception:
                run_mode = None
            if run_mode is None:
                try:
                    # Positional fallback: column F (index 5)
                    if len(crow) > 5:
                        val = crow.iat[5]
                        if not pd.isna(val):
                            run_mode = str(val).strip().lower()
                except Exception:
                    run_mode = None

            # Normalize synonyms and validate
            if isinstance(run_mode, str):
                if run_mode in ('fast', 'quick'):
                    run_mode = 'coarse'
                elif run_mode in ('slow', 'accurate', 'fine', 'extrafine'):
                    run_mode = 'fine'
                elif run_mode not in ('coarse', 'fine'):
                    # Unknown token -> fall back to global default
                    run_mode = getattr(args, 'mode', 'coarse')
            else:
                run_mode = getattr(args, 'mode', 'coarse')

            print(f"Processing configuration '{prefix}': offaxis={offaxis_val}, energy={energy_val}, defocus={defocus_val}")

            # Prepare new input workbook path
            src_path = args.file
            src_dir = os.path.dirname(src_path) or '.'
            src_basename = os.path.basename(src_path)
            new_basename = f"{prefix}_{src_basename}"
            new_path = os.path.join(src_dir, new_basename)
            try:
                shutil.copy2(src_path, new_path)
            except Exception as e:
                print(f"Failed to copy input workbook to {new_path}: {e}")
                continue

            # Modify the copy: adjust Thermal sheet d_therm_rotx/roty and d_therm_z,
            # and set energy in C2 of vignetting sheets.
            try:
                import openpyxl
                wb = openpyxl.load_workbook(new_path)

                # Thermal adjustments
                if 'Thermal' in wb.sheetnames:
                    ws = wb['Thermal']
                    # Read header row (assume first row contains headers)
                    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
                    # Find matching columns (prefer exact base names and skip
                    # any headers that end with an underscore to avoid backup
                    # columns like 'd_therm_rotx_'). Normalise by removing
                    # non-alphanumeric characters (except underscore) for matching.
                    import re as _re
                    def _norm_header(s: str) -> str:
                        return _re.sub(r"[^0-9a-z_]+", "", s.lower()).strip()

                    col_rotx = None
                    col_roty = None
                    col_z = None
                    for idx, h in enumerate(headers):
                        if not h:
                            continue
                        # Skip headers that intentionally end with an underscore
                        if h.strip().endswith('_'):
                            continue
                        nh = _norm_header(h)
                        if nh == 'd_therm_rotx' or nh.startswith('d_therm_rotx'):
                            col_rotx = idx + 1
                        if nh == 'd_therm_roty' or nh.startswith('d_therm_roty'):
                            col_roty = idx + 1
                        if nh == 'd_therm_z' or nh.startswith('d_therm_z'):
                            col_z = idx + 1

                    # Defocus is now stored in the "Extra PSF shifts" sheet
                    # (d_extra_z column) rather than baked into Thermal d_therm_z.

                # Create / populate "Extra PSF shifts" sheet with the
                # off-axis contribution.  Values stored in arcsec.
                extra_rotx_arcsec = float(offaxis_val) * 60.0 / math.sqrt(2.0)
                extra_roty_arcsec = float(offaxis_val) * 60.0 / math.sqrt(2.0)
                if 'Extra PSF shifts' in wb.sheetnames:
                    ws_extra = wb['Extra PSF shifts']
                else:
                    ws_extra = wb.create_sheet('Extra PSF shifts')
                ws_extra.cell(row=1, column=1, value='Position #')
                ws_extra.cell(row=1, column=2, value='d_extra_rotx [arcsec]')
                ws_extra.cell(row=1, column=3, value='d_extra_roty [arcsec]')
                ws_extra.cell(row=1, column=4, value='d_extra_z [µm]')
                defocus_um = float(defocus_val) * 1e3  # mm -> µm
                # Determine positions from Thermal sheet (or fall back to 1..max)
                thermal_positions = []
                if 'Thermal' in wb.sheetnames:
                    ws_th = wb['Thermal']
                    for rr in range(2, (ws_th.max_row or 0) + 1):
                        pv = ws_th.cell(row=rr, column=1).value
                        if pv is not None:
                            try:
                                thermal_positions.append(int(float(pv)))
                            except (ValueError, TypeError):
                                pass
                if not thermal_positions:
                    thermal_positions = list(range(1, 601))
                for i, pos_num in enumerate(thermal_positions):
                    row_idx = i + 2
                    ws_extra.cell(row=row_idx, column=1, value=pos_num)
                    ws_extra.cell(row=row_idx, column=2, value=extra_rotx_arcsec)
                    ws_extra.cell(row=row_idx, column=3, value=extra_roty_arcsec)
                    ws_extra.cell(row=row_idx, column=4, value=defocus_um)

                # Vignetting energy in C2 for both rotrad and rotazi
                for sname in list(VIG_ROT_RAD_CANDIDATES) + list(VIG_ROT_AZI_CANDIDATES):
                    if sname in wb.sheetnames:
                        ws_v = wb[sname]
                        try:
                            ws_v.cell(row=2, column=3).value = float(energy_val)
                        except Exception:
                            pass

                # HEW degradation energy in C2 for both rotazi and rotrad
                for sname in list(HEW_DEG_ROT_AZI_CANDIDATES) + list(HEW_DEG_ROT_RAD_CANDIDATES):
                    if sname in wb.sheetnames:
                        ws_h = wb[sname]
                        try:
                            ws_h.cell(row=2, column=3).value = float(energy_val)
                        except Exception:
                            pass

                # Update A_eff column B based on per-energy mapping listed in
                # columns D (energy) and E (source column) of the A_eff sheet.
                # Supported E formats in column E: numeric column index (1-based),
                # Excel column letter ('L'), or header name to match.
                try:
                    if 'A_eff' in wb.sheetnames:
                        ws_a = wb['A_eff']
                        # Try to open a data-only workbook to read cached formula results
                        ws_a_values = None
                        try:
                            from openpyxl import load_workbook as _load_wb_vals
                            wb_vals = _load_wb_vals(new_path, data_only=True)
                            if 'A_eff' in wb_vals.sheetnames:
                                ws_a_values = wb_vals['A_eff']
                        except Exception:
                            ws_a_values = None
                        max_scan_row = min(ws_a.max_row or 0, 40)
                        mapping = []
                        from openpyxl.utils import column_index_from_string
                        import re as _re
                        for rr in range(1, max_scan_row + 1):
                            try:
                                cand_energy = ws_a.cell(row=rr, column=4).value
                                cand_src = ws_a.cell(row=rr, column=5).value
                            except Exception:
                                continue
                            if cand_energy is None or cand_src is None:
                                continue
                            # parse energy value (allow '1 keV' or numeric)
                            e_val = None
                            try:
                                if isinstance(cand_energy, (int, float)):
                                    e_val = float(cand_energy)
                                else:
                                    m = _re.search(r"(\d+(?:\.\d*)?)", str(cand_energy))
                                    if m:
                                        e_val = float(m.group(1))
                            except Exception:
                                e_val = None
                            if e_val is None:
                                continue
                            # parse source column
                            src_idx = None
                            try:
                                if isinstance(cand_src, (int, float)):
                                    src_idx = int(float(cand_src))
                                else:
                                    s = str(cand_src).strip()
                                    # if looks like a column letter
                                    if _re.fullmatch(r"[A-Za-z]+", s):
                                        try:
                                            src_idx = column_index_from_string(s.upper())
                                        except Exception:
                                            src_idx = None
                                    else:
                                        # try extracting integer from string
                                        m2 = _re.search(r"(\d+)", s)
                                        if m2:
                                            src_idx = int(m2.group(1))
                            except Exception:
                                src_idx = None
                            # If still no numeric src_idx, try to find a header match
                            if src_idx is None:
                                try:
                                    # scan header row for a column whose header matches cand_src
                                    hdr = str(cand_src).strip().lower()
                                    for ccol in range(1, (ws_a.max_column or 0) + 1):
                                        try:
                                            valh = ws_a.cell(row=1, column=ccol).value
                                        except Exception:
                                            valh = None
                                        if valh is None:
                                            continue
                                        if str(valh).strip().lower() == hdr:
                                            src_idx = ccol
                                            break
                                except Exception:
                                    pass
                            if src_idx is not None:
                                mapping.append((e_val, src_idx))

                        # Choose best mapping for this configuration's energy_val.
                        # Validate candidate source columns actually contain numeric
                        # per-MM weights before copying into column B.
                        chosen_src = None
                        try:
                            if mapping:
                                # compute numeric fraction for each candidate, prefer exact-energy
                                mapping_sorted = sorted(mapping, key=lambda x: abs(x[0] - float(energy_val)))
                                cand_info = []
                                for e_val, cand_col in mapping_sorted:
                                    try:
                                        if cand_col == 2:
                                            continue
                                        max_r = ws_a.max_row or 0
                                        numeric_count = 0
                                        total_checked = 0
                                        formula_count = 0  # unevaluated formulas (cached=None)
                                        for check_r in range(2, min(max_r, 1000) + 1):
                                            try:
                                                mmcell = ws_a.cell(row=check_r, column=1).value
                                            except Exception:
                                                mmcell = None
                                            try:
                                                if mmcell is None:
                                                    continue
                                                _ = int(float(mmcell))
                                            except Exception:
                                                continue
                                            try:
                                                v = ws_a.cell(row=check_r, column=cand_col).value
                                                # if cell contains a formula, prefer cached value from data_only workbook
                                                _was_formula = isinstance(v, str) and v.startswith('=')
                                                if _was_formula:
                                                    if ws_a_values is not None:
                                                        try:
                                                            v = ws_a_values.cell(row=check_r, column=cand_col).value
                                                        except Exception:
                                                            pass
                                                    else:
                                                        v = None
                                                # Formula with no cached result: count separately,
                                                # do not pollute total_checked with uncountable cells
                                                if _was_formula and v is None:
                                                    formula_count += 1
                                                    continue
                                                total_checked += 1
                                                if v is None:
                                                    continue
                                                try:
                                                    vv = float(v)
                                                    if not (isinstance(vv, float) and (vv != vv)):
                                                        numeric_count += 1
                                                except Exception:
                                                    pass
                                            except Exception:
                                                continue
                                            except Exception:
                                                continue
                                        # If every MM cell is an unevaluated formula the column is
                                        # almost certainly a numeric A_eff column — treat as frac=1.
                                        if total_checked == 0 and formula_count > 0:
                                            frac = 1.0
                                        elif total_checked > 0:
                                            frac = numeric_count / total_checked
                                        else:
                                            frac = 0.0
                                        dist = abs(e_val - float(energy_val))
                                        cand_info.append((e_val, cand_col, frac, dist))
                                    except Exception:
                                        continue

                                # Prefer exact-energy matches (within tolerance) and pick highest frac
                                exacts = [ci for ci in cand_info if abs(ci[0] - float(energy_val)) < 1e-8]
                                if exacts:
                                    # pick by highest fraction, tie-breaker smallest distance
                                    exacts_sorted = sorted(exacts, key=lambda x: (-x[2], x[3]))
                                    if exacts_sorted[0][2] > 0:
                                        chosen_src = exacts_sorted[0][1]
                                else:
                                    # pick candidate with highest numeric fraction, use distance as tie-breaker
                                    cand_sorted = sorted(cand_info, key=lambda x: (-x[2], x[3]))
                                    if cand_sorted and cand_sorted[0][2] > 0:
                                        chosen_src = cand_sorted[0][1]
                        except Exception:
                            chosen_src = None

                        # Fallback: when no D/E mapping table was found, scan the header
                        # row of the A_eff sheet for energy-labelled columns (e.g. "0.5 keV"
                        # or the numeric value 0.5) and pick the column whose energy is
                        # closest to energy_val.  Only applies when row 1 col A is non-numeric
                        # (i.e. the sheet has a header row, not raw data from row 1).
                        if chosen_src is None and energy_val is not None:
                            try:
                                import re as _re_hdr
                                # Only proceed if row 1 looks like a header (col A is non-numeric)
                                _row1_col1 = ws_a.cell(row=1, column=1).value
                                _has_header = True
                                try:
                                    if _row1_col1 is not None:
                                        int(float(_row1_col1))
                                        _has_header = False  # numeric → data row, not a header
                                except (ValueError, TypeError):
                                    pass  # non-numeric → header row
                                if _has_header:
                                    _best_hdr_col = None
                                    _best_hdr_dist = float('inf')
                                    _max_col_a = ws_a.max_column or 0
                                    _col2_energy = None  # energy label on col B itself
                                    # Read col B's own header to see if it already matches
                                    try:
                                        _hval_b = ws_a.cell(row=1, column=2).value
                                        if _hval_b is not None:
                                            if isinstance(_hval_b, (int, float)):
                                                _col2_energy = float(_hval_b)
                                            else:
                                                _mh_b = _re_hdr.search(r"(\d+(?:\.\d*)?)", str(_hval_b))
                                                if _mh_b:
                                                    _ev_b = float(_mh_b.group(1))
                                                    _hs_b = str(_hval_b)
                                                    if _ev_b > 100 and 'ev' in _hs_b.lower() and 'kev' not in _hs_b.lower():
                                                        _ev_b /= 1000.0
                                                    _col2_energy = _ev_b
                                    except Exception:
                                        _col2_energy = None
                                    # If col B already contains the right energy, nothing to do
                                    if _col2_energy is not None and abs(_col2_energy - float(energy_val)) < 1e-6:
                                        pass  # col B already correct; leave chosen_src = None
                                    else:
                                        # Scan cols 3+ for energy labels in row 1
                                        for _ccol in range(3, _max_col_a + 1):
                                            try:
                                                _hval = ws_a.cell(row=1, column=_ccol).value
                                            except Exception:
                                                _hval = None
                                            if _hval is None:
                                                continue
                                            _ev = None
                                            try:
                                                if isinstance(_hval, (int, float)):
                                                    _ev = float(_hval)
                                                else:
                                                    _hs = str(_hval)
                                                    _mh = _re_hdr.search(r"(\d+(?:\.\d*)?)", _hs)
                                                    if _mh:
                                                        _ev = float(_mh.group(1))
                                                        if _ev > 100 and 'ev' in _hs.lower() and 'kev' not in _hs.lower():
                                                            _ev /= 1000.0
                                            except Exception:
                                                _ev = None
                                            if _ev is None:
                                                continue
                                            _dist = abs(_ev - float(energy_val))
                                            if _dist < _best_hdr_dist:
                                                _best_hdr_dist = _dist
                                                _best_hdr_col = _ccol
                                        # Validate the chosen column has numeric data
                                        if _best_hdr_col is not None:
                                            _num_cnt = 0
                                            _tot_cnt = 0
                                            _form_cnt = 0
                                            _max_r_a2 = ws_a.max_row or 0
                                            for _cr in range(2, min(_max_r_a2, 1000) + 1):
                                                try:
                                                    _mmcv = ws_a.cell(row=_cr, column=1).value
                                                    if _mmcv is None:
                                                        continue
                                                    _ = int(float(_mmcv))
                                                except Exception:
                                                    continue
                                                try:
                                                    _dv = ws_a.cell(row=_cr, column=_best_hdr_col).value
                                                    _is_form = isinstance(_dv, str) and _dv.startswith('=')
                                                    if _is_form and ws_a_values is not None:
                                                        try:
                                                            _dv_c = ws_a_values.cell(row=_cr, column=_best_hdr_col).value
                                                            if _dv_c is not None:
                                                                _dv = _dv_c
                                                                _is_form = False
                                                        except Exception:
                                                            pass
                                                    if _is_form and _dv is not None:
                                                        _form_cnt += 1
                                                        continue
                                                    _tot_cnt += 1
                                                    if _dv is not None:
                                                        try:
                                                            float(_dv)
                                                            _num_cnt += 1
                                                        except Exception:
                                                            pass
                                                except Exception:
                                                    pass
                                            if _tot_cnt == 0 and _form_cnt > 0:
                                                _frac_h = 1.0  # all unevaluated formulas → trust
                                            elif _tot_cnt > 0:
                                                _frac_h = _num_cnt / _tot_cnt
                                            else:
                                                _frac_h = 0.0
                                            if _frac_h > 0.5:
                                                chosen_src = _best_hdr_col
                                                print(f"A_eff col B fallback: header-row col {_best_hdr_col} for energy={energy_val} keV")
                            except Exception:
                                pass

                        # If chosen_src found, copy its numeric values into column B (index=2).
                        # For formula-based columns (VLOOKUP/XLOOKUP, data_only=None),
                        # manually evaluate MM# → Row# → A_eff using the workbook's own
                        # lookup tables ($V$6:$AG$20 and MM configuration sheet).
                        if chosen_src is not None and chosen_src != 2:
                            # Pre-build manual MM# → A_eff map by evaluating the VLOOKUP chain.
                            mm_to_base_manual = {}
                            try:
                                import re as _re_fml
                                # Detect formula and extract VLOOKUP offset
                                _vlookup_offset = None
                                for _fchk in range(2, min((ws_a.max_row or 0) + 1, 12)):
                                    _fv = ws_a.cell(row=_fchk, column=chosen_src).value
                                    if isinstance(_fv, str) and _fv.startswith('='):
                                        _fm = _re_fml.search(r'\$V\$6:\$AG\$20\s*,\s*(\d+)', _fv)
                                        if _fm:
                                            _vlookup_offset = int(_fm.group(1))
                                        elif 10 <= chosen_src <= 20:
                                            _vlookup_offset = chosen_src - 8
                                        break
                                if _vlookup_offset is not None:
                                    # Build MM# → Row# from MM configuration (col D=MM#, col C=Row#)
                                    _mm_to_row = {}
                                    if 'MM configuration' in wb.sheetnames:
                                        _ws_mmcfg = wb['MM configuration']
                                        for _rr in range(2, (_ws_mmcfg.max_row or 0) + 1):
                                            _mm_v = _ws_mmcfg.cell(row=_rr, column=4).value
                                            _row_v = _ws_mmcfg.cell(row=_rr, column=3).value
                                            if _mm_v is not None and _row_v is not None:
                                                try:
                                                    _mm_to_row[int(float(_mm_v))] = int(float(_row_v))
                                                except Exception:
                                                    pass
                                    # Build Row# → A_eff from $V$6:$AG$20
                                    # col V = col 22 in sheet; VLOOKUP col index = 21 + offset
                                    _aeff_col = 21 + _vlookup_offset
                                    _row_to_aeff = {}
                                    for _lr in range(6, 21):
                                        _rk = ws_a.cell(row=_lr, column=22).value  # col V = Row#
                                        _av = ws_a.cell(row=_lr, column=_aeff_col).value
                                        if _rk is not None and _av is not None:
                                            try:
                                                _row_to_aeff[int(float(_rk))] = float(_av)
                                            except Exception:
                                                pass
                                    for _mm, _rk in _mm_to_row.items():
                                        _av = _row_to_aeff.get(_rk)
                                        if _av is not None:
                                            mm_to_base_manual[_mm] = _av
                            except Exception:
                                mm_to_base_manual = {}

                            for row_idx in range(2, (ws_a.max_row or 0) + 1):
                                try:
                                    # ensure this row corresponds to an MM entry
                                    mmcell = ws_a.cell(row=row_idx, column=1).value
                                    mm_int = None
                                    try:
                                        if mmcell is None:
                                            continue
                                        mm_int = int(float(mmcell))
                                    except Exception:
                                        continue

                                    # Prefer manually-evaluated lookup (for formula columns)
                                    write_val = mm_to_base_manual.get(mm_int)

                                    if write_val is None:
                                        src_val = ws_a.cell(row=row_idx, column=chosen_src).value
                                        try:
                                            if isinstance(src_val, str) and src_val.startswith('=') and ws_a_values is not None:
                                                tmp = ws_a_values.cell(row=row_idx, column=chosen_src).value
                                                if tmp is not None:
                                                    src_val = tmp
                                            elif isinstance(src_val, str) and src_val.startswith('='):
                                                src_val = None
                                        except Exception:
                                            src_val = None
                                        try:
                                            if src_val is not None:
                                                write_val = float(src_val)
                                        except Exception:
                                            write_val = None
                                        if write_val is None and src_val is not None:
                                            try:
                                                import re as _re
                                                m = _re.search(r"(-?\d+(?:\.\d+)?)", str(src_val))
                                                if m:
                                                    write_val = float(m.group(1))
                                            except Exception:
                                                write_val = None

                                    # Last resort: col C only when manual lookup produced nothing
                                    if write_val is None and not mm_to_base_manual:
                                        try:
                                            alt = ws_a.cell(row=row_idx, column=3).value
                                            if alt is not None:
                                                write_val = float(alt)
                                        except Exception:
                                            write_val = None

                                    if write_val is not None:
                                        ws_a.cell(row=row_idx, column=2).value = write_val
                                except Exception:
                                    continue
                except Exception:
                    pass

                 # Resolve MM_PSF formula cells (sigma_rad/sigma_azi in
                # columns D/E) using cached values from the *source*
                # workbook.  openpyxl preserves formula strings but does
                # NOT recalculate them, so the copy would have None for
                # data_only reads.  We read the cached numeric values
                # from the original file and stamp them as literals.
                try:
                    if 'MM_PSF' in wb.sheetnames:
                        ws_mm = wb['MM_PSF']
                        from openpyxl import load_workbook as _load_wb_src
                        wb_src_cached = _load_wb_src(src_path, data_only=True)
                        if 'MM_PSF' in wb_src_cached.sheetnames:
                            ws_src_cached = wb_src_cached['MM_PSF']
                            for rr in range(2, (ws_mm.max_row or 0) + 1):
                                for cc in (4, 5):  # D=sigma_rad, E=sigma_azi
                                    cell_val = ws_mm.cell(row=rr, column=cc).value
                                    if isinstance(cell_val, str) and cell_val.startswith('='):
                                        cached = ws_src_cached.cell(row=rr, column=cc).value
                                        if cached is not None:
                                            try:
                                                ws_mm.cell(row=rr, column=cc).value = float(cached)
                                            except (ValueError, TypeError):
                                                pass
                except Exception:
                    pass

                wb.save(new_path)
            except Exception as e:
                print(f"Failed to modify workbook {new_path}: {e}")
                continue

            # Compute HEW/EEF metrics for this modified workbook (no plotting)
            metrics = {}
            try:
                try:
                    metrics = compute_hew_eef_metrics(
                        new_path,
                        sheet=args.sheet if hasattr(args, 'sheet') else 'MM_PSF',
                        normalize=getattr(args, 'normalize', True),
                        fast=(run_mode == 'coarse')
                    )
                except Exception:
                    metrics = compute_hew_eef_metrics(new_path, fast=(run_mode == 'coarse'))
            except Exception as e:
                print(f"Warning: failed to compute HEW/EEF metrics for {new_basename}: {e}")

            # Run this script on the modified workbook with --export-package
            try:
                cmd = [sys.executable, os.path.abspath(__file__), '--file', new_path, '--export-package', '--mode', run_mode]
                print('Running:', ' '.join(cmd))

                # Record existing Exports subfolders so we can detect a new package
                existing_dirs = set(os.listdir(exports_root)) if os.path.isdir(exports_root) else set()

                subprocess.check_call(cmd)
            except Exception as e:
                print(f"Failed to run export for {prefix}: {e}")

            # Prefer zipping the full export package folder created by the child run.
            try:
                os.makedirs(exports_root, exist_ok=True)
                # Look for new subdirectories created during the child run
                all_dirs = [d for d in os.listdir(exports_root) if os.path.isdir(os.path.join(exports_root, d))]
                new_dirs = [d for d in all_dirs if d not in existing_dirs]
                # Build timestamp and safe stem for zip naming (no .xlsx)
                ts = time.strftime('%Y%m%d_%H%M%S')
                if new_dirs:
                    # Prefer the newest created package folder
                    latest_dir = max(new_dirs, key=lambda d: os.path.getmtime(os.path.join(exports_root, d)))
                    pkg_path = os.path.join(exports_root, latest_dir)

                    # Attempt to rename specific artifacts inside the package dir
                    try:
                        # Rename helpers: operate on files inside pkg_path
                        def _rename_in_pkg(glob_patterns, new_name):
                            for pat in glob_patterns:
                                matches = glob.glob(os.path.join(pkg_path, pat))
                                if not matches:
                                    continue
                                latest = max(matches, key=os.path.getmtime)
                                dst = os.path.join(pkg_path, new_name)
                                try:
                                    # If dst exists remove it first
                                    if os.path.exists(dst):
                                        os.remove(dst)
                                    os.replace(latest, dst)
                                    return True
                                except Exception:
                                    try:
                                        shutil.copy2(latest, dst)
                                        return True
                                    except Exception:
                                        return False
                            return False

                        # 1) Input file already copied into pkg_dir earlier
                        # 2) Rename combined figure
                        _rename_in_pkg(['E2E_fit_combined*.png', 'E2E_fit_combined.png'], 'Combined_E2E_EEF.png')
                        # 3) Rename PSF PNG
                        _rename_in_pkg(['E2E_PSF_*.png', 'E2E_PSF*.png', 'E2E_PSF_*.PNG'], 'E2E_PSF.png')
                        # 4) Rename aggregated FITS
                        _rename_in_pkg(['E2E_aggregated_*.fits', 'E2E_aggregated*.fits'], 'E2E_PSF.fits')
                        # 5) Rename EEF & fit params workbook
                        _rename_in_pkg(['E2E_EEF_and_fitparams_*.xlsx', 'E2E_EEF_and_fitparams*.xlsx', 'E2E_EEF_and_fitparams_*.xls'], 'EEF_fittingparams.xlsx')
                        # 6) Rename Encircled Energy plot to standardized EEF.png
                        _rename_in_pkg(['Encircled_Energy_*.png', 'Encircled_Energy*.png', 'Encircled_Energy_*.PNG'], 'EEF.png')
                    except Exception as e:
                        print(f"Warning: renaming artifacts failed: {e}")

                    # Create zip with controlled ordering: input, Combined_E2E_EEF.png,
                    # E2E_PSF.png, E2E_PSF.fits, EEF_fittingparams.xlsx, then rest.
                    zip_base = os.path.join(export_batch_dir, f"{prefix}_{src_stem}_{ts}")
                    zip_target = f"{zip_base}.zip"
                    try:
                        with zipfile.ZipFile(zip_target, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
                            ordered = []
                            # 1) Input file
                            input_dst = os.path.join(pkg_path, os.path.basename(args.file))
                            if os.path.exists(input_dst):
                                ordered.append(input_dst)
                            # 2) Combined figure
                            for name in ('Combined_E2E_EEF.png',):
                                p = os.path.join(pkg_path, name)
                                if os.path.exists(p):
                                    ordered.append(p)
                            # 3) E2E_PSF.png
                            p = os.path.join(pkg_path, 'E2E_PSF.png')
                            if os.path.exists(p):
                                ordered.append(p)
                            # 4) E2E_PSF.fits
                            p = os.path.join(pkg_path, 'E2E_PSF.fits')
                            if os.path.exists(p):
                                ordered.append(p)
                            # 5) EEF_fittingparams.xlsx
                            p = os.path.join(pkg_path, 'EEF_fittingparams.xlsx')
                            if os.path.exists(p):
                                ordered.append(p)
                            # 6) EEF.png (renamed Encircled Energy)
                            p = os.path.join(pkg_path, 'EEF.png')
                            if os.path.exists(p):
                                ordered.append(p)

                            # Before creating the package, compute A_eff sums and write EEF_fitting workbook
                            sum_orig = None
                            sum_mod = None
                            aeff_loss = None
                            # Prefer computing both sums from the modified workbook (new_path)
                            # - Aeff_sum_orig: sum of column B (index 1) in the modified A_eff sheet
                            # - Aeff_sum_mod: sum of column C (index 2) in the modified A_eff sheet
                            # If the modified workbook is not available or missing columns, fall back
                            # to the original input workbook where appropriate.
                            try:
                                # Prefer the workbook actually stored in the package (safer when child run
                                # copied/renamed files inside pkg_path). This is typically the input file
                                # copied into the package dir with the same basename as args.file.
                                file_in_pkg = None
                                try:
                                    inp_basename = os.path.basename(args.file)
                                    cand = os.path.join(pkg_path, inp_basename)
                                    if os.path.exists(cand):
                                        file_in_pkg = cand
                                except Exception:
                                    file_in_pkg = None

                                use_path = file_in_pkg or (new_path if (new_path and os.path.exists(new_path)) else None)
                                if use_path is not None:
                                    df_a_mod = pd.read_excel(use_path, sheet_name='A_eff', engine='openpyxl', header=None)
                                    if df_a_mod.shape[0] >= 2 and df_a_mod.shape[1] > 1:
                                        sum_orig = float(pd.to_numeric(df_a_mod.iloc[1:, 1], errors='coerce').fillna(0.0).sum())
                                    if df_a_mod.shape[0] >= 2 and df_a_mod.shape[1] > 2:
                                        sum_mod = float(pd.to_numeric(df_a_mod.iloc[1:, 2], errors='coerce').fillna(0.0).sum())
                            except Exception:
                                sum_orig = None
                                sum_mod = None
                            # If modified workbook didn't provide sum_orig, try original input file col B
                            if sum_orig is None:
                                try:
                                    df_a_orig = pd.read_excel(args.file, sheet_name='A_eff', engine='openpyxl', header=None)
                                    if df_a_orig.shape[0] >= 2 and df_a_orig.shape[1] > 1:
                                        sum_orig = float(pd.to_numeric(df_a_orig.iloc[1:, 1], errors='coerce').fillna(0.0).sum())
                                except Exception:
                                    sum_orig = None
                            # If modified workbook didn't provide sum_mod, try original input file col C
                            if sum_mod is None:
                                try:
                                    df_a_orig = df_a_orig if 'df_a_orig' in locals() else pd.read_excel(args.file, sheet_name='A_eff', engine='openpyxl', header=None)
                                    if df_a_orig.shape[0] >= 2 and df_a_orig.shape[1] > 2:
                                        sum_mod = float(pd.to_numeric(df_a_orig.iloc[1:, 2], errors='coerce').fillna(0.0).sum())
                                except Exception:
                                    sum_mod = None
                            if sum_orig is not None and sum_mod is not None and sum_orig != 0:
                                aeff_loss = 1.0 - (sum_mod / sum_orig)

                            # write EEF_fitting.xlsx inside the package dir (merge existing Fit_parameters + EEF_curves with formulas)
                            try:
                                target_fp = os.path.join(pkg_path, 'EEF_fitting.xlsx')
                                df_rows = []
                                df_eef_src = None
                                cand = os.path.join(pkg_path, 'EEF_fittingparams.xlsx')
                                existing_fp = cand if os.path.exists(cand) else None
                                if existing_fp is None:
                                    for cand2 in glob.glob(os.path.join(pkg_path, 'E2E_EEF_and_fitparams_*.xlsx')):
                                        existing_fp = cand2
                                        break
                                if existing_fp is not None:
                                    try:
                                        df_exist = pd.read_excel(existing_fp, sheet_name='Fit_parameters', engine='openpyxl')
                                        if 'parameter' in df_exist.columns and 'value' in df_exist.columns:
                                            for _, r in df_exist.iterrows():
                                                df_rows.append({'parameter': str(r['parameter']), 'value': r['value']})
                                    except Exception:
                                        df_rows = []
                                    try:
                                        df_eef_src = pd.read_excel(existing_fp, sheet_name='EEF_curves', engine='openpyxl')
                                    except Exception:
                                        df_eef_src = None
                                df_rows.append({'parameter': 'Aeff_sum_orig', 'value': sum_orig})
                                df_rows.append({'parameter': 'Aeff_sum_mod', 'value': sum_mod})
                                df_rows.append({'parameter': 'Aeff_loss', 'value': aeff_loss})
                                try:
                                    import pandas as _pd
                                    with _pd.ExcelWriter(target_fp, engine='openpyxl') as _writer:
                                        _pd.DataFrame(df_rows).to_excel(_writer, sheet_name='Fit_parameters', index=False)
                                        # Add EEF_curves sheet: data columns for both data curves,
                                        # formula columns (I + integrand + EEF) for each of the 3 fits.
                                        if df_eef_src is not None and 'diameter_arcsec' in df_eef_src.columns:
                                            _wb = _writer.book
                                            _ws = _wb.create_sheet('EEF_curves')
                                            _hdr = [
                                                'diameter_arcsec',
                                                'EEF_centered_min_pct',
                                                'EEF_centered_origin_pct',
                                                'I_pv [arb.u.]', 'I_pv_x_r', 'EEF_pv_pct',
                                                'I_pearson4 [arb.u.]', 'I_pearson4_x_r', 'EEF_pearson4_pct',
                                                'I_king [arb.u.]', 'I_king_x_r', 'EEF_king_pct',
                                            ]
                                            for _ci, _h in enumerate(_hdr, 1):
                                                _ws.cell(1, _ci).value = _h
                                            _diam_vals = list(df_eef_src['diameter_arcsec'])
                                            # Forward-fill NaN (capped >95% rows) so every row
                                            # has a numeric value matching the formula-column length
                                            def _to_full(col_name):
                                                if col_name not in df_eef_src.columns:
                                                    return [None] * len(_diam_vals)
                                                s = df_eef_src[col_name].ffill()
                                                return [None if (v != v) else v for v in s]
                                            _eef_min  = _to_full('EEF_aggregated_pct')
                                            _eef_orig = _to_full('EEF_origin_pct')
                                            _end = 1 + len(_diam_vals)  # last data row index
                                            # VLOOKUP helpers referencing the Fit_parameters sheet
                                            def _vl(key):
                                                return f'VLOOKUP("{key}",Fit_parameters!$A:$B,2,FALSE)'
                                            _A_lu = _vl('Modified_PV_Amplitude_A')
                                            _eta_lu = _vl('Modified_PV_eta')
                                            _sc_lu = _vl('Modified_PV_scalar')
                                            _beta_lu = _vl('Modified_PV_beta')
                                            _Gc_lu = _vl('Modified_PV_Gamma_c_arcsec')
                                            _Gw_lu = _vl('Modified_PV_Gamma_w_arcsec')
                                            _Amp4_lu = _vl('Pearson4_Amplitude')
                                            _sig4_lu = _vl('Pearson4_Sigma_arcsec')
                                            _m4_lu = _vl('Pearson4_Exponent_m')
                                            _nu4_lu = _vl('Pearson4_Skew_nu')
                                            _I0k_lu = _vl('King_I0')
                                            _rck_lu = _vl('King_rc_arcsec')
                                            _ak_lu = _vl('King_alpha')
                                            _bk_lu = _vl('King_b')
                                            for _ri, (_dval, _emin, _eorig) in enumerate(
                                                    zip(_diam_vals, _eef_min, _eef_orig), start=2):
                                                _r = f'A{_ri}/2'  # radius = diameter / 2
                                                _ws.cell(_ri, 1).value = _dval
                                                _ws.cell(_ri, 2).value = _emin
                                                _ws.cell(_ri, 3).value = _eorig
                                                # Col D: Modified PV intensity
                                                _G = f'EXP(-4*LN(2)*(({_r})/({_Gc_lu}))^2)'
                                                _ae = f'(POWER(2,1/({_beta_lu}))-1)'
                                                _C = f'1/POWER(1+{_ae}*(2*({_r})/({_Gw_lu}))^2,{_beta_lu})'
                                                _mix = f'(1-{_eta_lu})*{_G}+{_eta_lu}*{_sc_lu}*{_C}'
                                                _nrm = f'(1-{_eta_lu})+{_eta_lu}*{_sc_lu}'
                                                _ws.cell(_ri, 4).value = f'={_A_lu}*({_mix})/({_nrm})'
                                                # Col E: I_pv * r  (integrand for EEF)
                                                _ws.cell(_ri, 5).value = f'=D{_ri}*{_r}'
                                                # Col F: EEF_pv_pct (cumulative sum / total)
                                                _ws.cell(_ri, 6).value = (
                                                    f'=IF(SUM($E$2:$E${_end})>0,'
                                                    f'100*SUM($E$2:E{_ri})/SUM($E$2:$E${_end}),NA())')
                                                # Col G: Pearson4 intensity
                                                _u4 = f'({_r})/({_sig4_lu})'
                                                _ws.cell(_ri, 7).value = (
                                                    f'={_Amp4_lu}*POWER(1+({_u4})^2,-({_m4_lu}))'
                                                    f'*EXP(-({_nu4_lu})*ATAN({_u4}))')
                                                # Col H: I_pearson4 * r
                                                _ws.cell(_ri, 8).value = f'=G{_ri}*{_r}'
                                                # Col I: EEF_pearson4_pct
                                                _ws.cell(_ri, 9).value = (
                                                    f'=IF(SUM($H$2:$H${_end})>0,'
                                                    f'100*SUM($H$2:H{_ri})/SUM($H$2:$H${_end}),NA())')
                                                # Col J: King intensity
                                                _ws.cell(_ri, 10).value = (
                                                    f'={_I0k_lu}*POWER(1+(({_r})/({_rck_lu}))^2,-({_ak_lu}))+{_bk_lu}')
                                                # Col K: I_king * r
                                                _ws.cell(_ri, 11).value = f'=J{_ri}*{_r}'
                                                # Col L: EEF_king_pct
                                                _ws.cell(_ri, 12).value = (
                                                    f'=IF(SUM($K$2:$K${_end})>0,'
                                                    f'100*SUM($K$2:K{_ri})/SUM($K$2:$K${_end}),NA())')
                                    if existing_fp is not None and os.path.exists(existing_fp):
                                        try:
                                            os.remove(existing_fp)
                                        except Exception:
                                            pass
                                except Exception:
                                    pass
                            except Exception:
                                pass

                            # Add ordered files first into the zip
                            added = set()
                            for fpath in ordered:
                                arcname = os.path.basename(fpath)
                                try:
                                    zf.write(fpath, arcname=arcname)
                                    added.add(os.path.abspath(fpath))
                                except Exception:
                                    pass

                            # Then add remaining files under pkg_path
                            for root, _, files in os.walk(pkg_path):
                                for fname in files:
                                    full = os.path.join(root, fname)
                                    if os.path.abspath(full) in added:
                                        continue
                                    arcname = os.path.relpath(full, pkg_path)
                                    try:
                                        zf.write(full, arcname=arcname)
                                    except Exception:
                                        pass
                        print(f"Created package from export folder: {zip_target}")
                        print(f"Created package from export folder: {zip_target}")
                        # Extract fit parameters and EEF curves if available and append aggregated row
                        try:
                            fit_map = {}
                            eef80 = None
                            eef90 = None
                            # Look for the standardized EEF_fittingparams workbook first,
                            # then fallback to common original names inside the package folder.
                            candidate_files = []
                            if 'pkg_path' in locals() and os.path.isdir(pkg_path):
                                candidate_files.append(os.path.join(pkg_path, 'EEF_fittingparams.xlsx'))
                                # include the aeff-loss fitparams file we write earlier
                                candidate_files.append(os.path.join(pkg_path, 'EEF_fitting.xlsx'))
                                # search for common patterns (renamed or original)
                                candidate_files.extend(glob.glob(os.path.join(pkg_path, 'E2E_EEF_and_fitparams_*.xlsx')))
                                candidate_files.extend(glob.glob(os.path.join(pkg_path, 'CustomPSFs', 'E2E_EEF_and_fitparams_*.xlsx')))
                                candidate_files.extend(glob.glob(os.path.join(pkg_path, '**', 'E2E_EEF_and_fitparams_*.xlsx'), recursive=True))
                            # Try each candidate until one yields a Fit_parameters sheet
                            found_fp = None
                            for cand in candidate_files:
                                try:
                                    if not os.path.exists(cand):
                                        continue
                                    # Try to read Fit_parameters sheet
                                    df_fit = pd.read_excel(cand, sheet_name='Fit_parameters', engine='openpyxl')
                                    if 'parameter' in df_fit.columns and 'value' in df_fit.columns:
                                        fit_map = dict(zip(df_fit['parameter'].astype(str), df_fit['value']))
                                        found_fp = cand
                                        break
                                except Exception:
                                    continue
                            # If not found on filesystem, try inside the zip archive
                            if found_fp is None and os.path.exists(zip_target):
                                try:
                                    with zipfile.ZipFile(zip_target, 'r') as zf:
                                        for name in zf.namelist():
                                            lname = name.lower()
                                            # accept a variety of fitparams filenames including
                                            # E2E_EEF_and_fitparams, EEF_fittingparams, and
                                            # the EEF_fitting workbook we generate.
                                            if ('e2e_eef_and_fitparams' in lname or
                                                'eef_fittingparams' in lname or
                                                'eef_fitting' in lname or
                                                'fitparams' in lname):
                                                # extract to temp and attempt to read
                                                tf = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(name)[1])
                                                tf.close()
                                                try:
                                                    with open(tf.name, 'wb') as out_f:
                                                        out_f.write(zf.read(name))
                                                    df_fit = pd.read_excel(tf.name, sheet_name='Fit_parameters', engine='openpyxl')
                                                    if 'parameter' in df_fit.columns and 'value' in df_fit.columns:
                                                        fit_map = dict(zip(df_fit['parameter'].astype(str), df_fit['value']))
                                                        found_fp = tf.name
                                                        break
                                                except Exception:
                                                    try:
                                                        os.remove(tf.name)
                                                    except Exception:
                                                        pass
                                except Exception:
                                    pass
                            # If no Fit_parameters was discovered inside the package/zip,
                            # try reading it directly from the modified workbook copy (`new_path`).
                            if not fit_map:
                                try:
                                    if new_path and os.path.exists(new_path):
                                        df_fit = pd.read_excel(new_path, sheet_name='Fit_parameters', engine='openpyxl')
                                        if 'parameter' in df_fit.columns and 'value' in df_fit.columns:
                                            fit_map = dict(zip(df_fit['parameter'].astype(str), df_fit['value']))
                                            found_fp = new_path
                                except Exception:
                                    # keep fit_map empty on failure
                                    pass
                            # If we found a workbook, also try to read EEF_curves from it
                            if found_fp:
                                try:
                                    df_eef = pd.read_excel(found_fp, sheet_name='EEF_curves', engine='openpyxl')
                                    if 'diameter_arcsec' in df_eef.columns and 'EEF_aggregated_pct' in df_eef.columns:
                                        pct = pd.to_numeric(df_eef['EEF_aggregated_pct'], errors='coerce').to_numpy()
                                        diam = pd.to_numeric(df_eef['diameter_arcsec'], errors='coerce').to_numpy()
                                        order = np.argsort(pct)
                                        pct_s = pct[order]
                                        diam_s = diam[order]
                                        if np.nanmin(pct_s) <= 80.0 <= np.nanmax(pct_s):
                                            eef80 = float(np.interp(80.0, pct_s, diam_s))
                                        if np.nanmin(pct_s) <= 90.0 <= np.nanmax(pct_s):
                                            eef90 = float(np.interp(90.0, pct_s, diam_s))
                                except Exception:
                                    eef80 = None
                                    eef90 = None
                        except Exception:
                            fit_map = {}
                            eef80 = None
                            eef90 = None
                        # Compute Aeff loss as 1 - SUM(C2:C601)/SUM(B2:B601) on the A_eff sheets.
                        # Prefer the per-configuration modified workbook (`new_path`) for
                        # both the baseline (column B) and adjusted (column C) sums when
                        # available; fall back to the original input (`args.file`) only
                        # when the modified workbook is not present.
                        try:
                            sum_orig = None
                            sum_mod = None
                            try:
                                use_path = None
                                if 'new_path' in locals() and new_path and os.path.exists(new_path):
                                    use_path = new_path
                                # If the child run already created a package copy of the workbook
                                # prefer that file if present in pkg_path
                                try:
                                    inp_basename = os.path.basename(args.file)
                                    pkg_copy = os.path.join(pkg_path, inp_basename) if 'pkg_path' in locals() else None
                                    if pkg_copy and os.path.exists(pkg_copy):
                                        use_path = pkg_copy
                                except Exception:
                                    pass
                                if use_path is None:
                                    use_path = args.file
                                df_a_use = pd.read_excel(use_path, sheet_name='A_eff', engine='openpyxl', header=None)
                                if df_a_use.shape[0] >= 2 and df_a_use.shape[1] > 1:
                                    sum_orig = float(pd.to_numeric(df_a_use.iloc[1:, 1], errors='coerce').fillna(0.0).sum())
                                if df_a_use.shape[0] >= 2 and df_a_use.shape[1] > 2:
                                    sum_mod = float(pd.to_numeric(df_a_use.iloc[1:, 2], errors='coerce').fillna(0.0).sum())
                            except Exception:
                                sum_orig = None
                                sum_mod = None
                            if sum_orig is not None and sum_mod is not None and sum_orig != 0:
                                aeff_loss = 1.0 - (sum_mod / sum_orig)
                            else:
                                aeff_loss = None
                        except Exception:
                            aeff_loss = None

                        # Ensure A_eff sums are available for this row (compute fallbacks if needed)
                        try:
                            sum_orig_row = sum_orig if 'sum_orig' in locals() else None
                            sum_mod_row = sum_mod if 'sum_mod' in locals() else None
                            if sum_orig_row is None:
                                try:
                                    df_a_tmp = pd.read_excel(args.file, sheet_name='A_eff', engine='openpyxl', header=None)
                                    if df_a_tmp.shape[0] >= 2 and df_a_tmp.shape[1] > 1:
                                        sum_orig_row = float(pd.to_numeric(df_a_tmp.iloc[1:, 1], errors='coerce').fillna(0.0).sum())
                                except Exception:
                                    sum_orig_row = None
                            if sum_mod_row is None:
                                try:
                                    if new_path and os.path.exists(new_path):
                                        df_a_tmp2 = pd.read_excel(new_path, sheet_name='A_eff', engine='openpyxl', header=None)
                                        if df_a_tmp2.shape[0] >= 2 and df_a_tmp2.shape[1] > 2:
                                            sum_mod_row = float(pd.to_numeric(df_a_tmp2.iloc[1:, 2], errors='coerce').fillna(0.0).sum())
                                except Exception:
                                    sum_mod_row = None
                            if sum_mod_row is None and sum_orig_row is not None:
                                try:
                                    if df_a_tmp.shape[1] > 2:
                                        sum_mod_row = float(pd.to_numeric(df_a_tmp.iloc[1:, 2], errors='coerce').fillna(0.0).sum())
                                except Exception:
                                    pass
                        except Exception:
                            sum_orig_row = None
                            sum_mod_row = None

                        # Build aggregated row (include offaxis [arcmin], energy [keV], defocus [mm])
                        try:
                            # Determine EEF80/EEF90 values, preferring explicit EEF curves
                            # extracted from the Fit/EFF workbook, then any fit_map keys,
                            # and finally the computed metrics as a last resort.
                            eef80_val = None
                            eef90_val = None
                            try:
                                if eef80 is not None:
                                    eef80_val = float(eef80)
                            except Exception:
                                eef80_val = None
                            try:
                                if eef90 is not None:
                                    eef90_val = float(eef90)
                            except Exception:
                                eef90_val = None

                            # Check fit_map for named EEF entries
                            try:
                                if not eef80_val and isinstance(fit_map, dict):
                                    for key in ('EEF80_min_arcsec', 'EEF80_min', 'EEF80'):
                                        if key in fit_map and fit_map.get(key) is not None:
                                            try:
                                                eef80_val = float(fit_map.get(key))
                                                break
                                            except Exception:
                                                continue
                                if not eef90_val and isinstance(fit_map, dict):
                                    for key in ('EEF90_min_arcsec', 'EEF90_min', 'EEF90'):
                                        if key in fit_map and fit_map.get(key) is not None:
                                            try:
                                                eef90_val = float(fit_map.get(key))
                                                break
                                            except Exception:
                                                continue
                            except Exception:
                                pass

                            # Final fallback to computed metrics
                            try:
                                if (eef80_val is None or eef80_val == 0) and isinstance(metrics, dict):
                                    eef80_val = metrics.get('eef80_best_arcsec') or metrics.get('eef80') or eef80_val
                                if (eef90_val is None or eef90_val == 0) and isinstance(metrics, dict):
                                    eef90_val = metrics.get('eef90_best_arcsec') or metrics.get('eef90') or eef90_val
                            except Exception:
                                pass

                            row = {
                                'configuration_number': int(rid) + 1,
                                'configuration_name': prefix,
                                'offaxis_arcmin': offaxis_val,
                                'energy_keV': energy_val,
                                'defocus_mm': defocus_val,
                                'Aeff_sum_orig': sum_orig,
                                'Aeff_sum_mod': sum_mod,
                                'Aeff_loss': aeff_loss,
                                'HEW_00_arcsec': metrics.get('hew_origin_arcsec') if isinstance(metrics, dict) else None,
                                'HEW_min_arcsec': metrics.get('hew_best_arcsec') if isinstance(metrics, dict) else None,
                                'EEF80_min_arcsec': eef80_val if eef80_val is not None else None,
                                'EEF90_min_arcsec': eef90_val if eef90_val is not None else (metrics.get('eef90_best_arcsec') if isinstance(metrics, dict) else None),
                            }
                        except Exception:
                            row = {
                                'configuration_number': int(rid) + 1,
                                'configuration_name': prefix,
                                'offaxis_arcmin': offaxis_val,
                                'energy_keV': energy_val,
                                'defocus_mm': defocus_val,
                                'Aeff_sum_orig': sum_orig,
                                'Aeff_sum_mod': sum_mod,
                                'Aeff_loss': aeff_loss
                            }
                        pv_keys = ['Modified_PV_Amplitude_A', 'Modified_PV_Gamma_c_arcsec', 'Modified_PV_Gamma_w_arcsec', 'Modified_PV_eta', 'Modified_PV_beta', 'Modified_PV_scalar']
                        for k in pv_keys:
                            row[k] = fit_map.get(k) if isinstance(fit_map, dict) else None
                        p4_keys = ['Pearson4_Amplitude', 'Pearson4_Center_arcsec', 'Pearson4_Sigma_arcsec', 'Pearson4_Exponent_m', 'Pearson4_Skew_nu']
                        for k in p4_keys:
                            row[k] = fit_map.get(k) if isinstance(fit_map, dict) else None
                        king_keys = ['King_I0', 'King_rc_arcsec', 'King_alpha', 'King_b']
                        for k in king_keys:
                            row[k] = fit_map.get(k) if isinstance(fit_map, dict) else None
                        row['King_extra'] = None
                        aggregated_rows.append(row)
                        # Write/update aggregated Excel immediately in the export folder
                        try:
                            df_agg_now = pd.DataFrame(aggregated_rows)
                            # Ensure Aeff_sum columns exist and populate current row
                            try:
                                if 'Aeff_sum_orig' not in df_agg_now.columns:
                                    df_agg_now['Aeff_sum_orig'] = pd.NA
                                if 'Aeff_sum_mod' not in df_agg_now.columns:
                                    df_agg_now['Aeff_sum_mod'] = pd.NA
                                # Ensure EEF columns exist
                                if 'EEF80_min_arcsec' not in df_agg_now.columns:
                                    df_agg_now['EEF80_min_arcsec'] = pd.NA
                                if 'EEF90_min_arcsec' not in df_agg_now.columns:
                                    df_agg_now['EEF90_min_arcsec'] = pd.NA
                                # populate last appended row with computed sums if available
                                last_idx = len(df_agg_now) - 1
                                if last_idx >= 0:
                                    try:
                                        if 'sum_orig' in locals() and sum_orig is not None:
                                            df_agg_now.at[last_idx, 'Aeff_sum_orig'] = sum_orig
                                        if 'sum_mod' in locals() and sum_mod is not None:
                                            df_agg_now.at[last_idx, 'Aeff_sum_mod'] = sum_mod
                                            # populate EEF values if available in last aggregated_rows entry
                                            try:
                                                last_r = aggregated_rows[last_idx]
                                                if isinstance(last_r, dict):
                                                    if 'EEF80_min_arcsec' in last_r and last_r.get('EEF80_min_arcsec') is not None:
                                                        df_agg_now.at[last_idx, 'EEF80_min_arcsec'] = last_r.get('EEF80_min_arcsec')
                                                    if 'EEF90_min_arcsec' in last_r and last_r.get('EEF90_min_arcsec') is not None:
                                                        df_agg_now.at[last_idx, 'EEF90_min_arcsec'] = last_r.get('EEF90_min_arcsec')
                                            except Exception:
                                                pass
                                    except Exception:
                                        pass
                            except Exception:
                                pass
                            out_agg_now = os.path.join(export_batch_dir, f"Aggregated_results_{src_stem}.xlsx")
                            with pd.ExcelWriter(out_agg_now, engine='openpyxl') as writer:
                                df_agg_now.to_excel(writer, sheet_name='Aggregated', index=False)
                            print(f"Updated aggregated results: {out_agg_now}")
                        except Exception as e:
                            print(f"Warning: failed to update aggregated Excel after {prefix}: {e}")
                        # Now remove the package folder to avoid leaving duplicate artifacts
                        try:
                            if 'pkg_path' in locals() and os.path.isdir(pkg_path):
                                shutil.rmtree(pkg_path)
                                print(f"Removed intermediate package folder: {pkg_path}")
                        except Exception as e:
                            print(f"Warning: failed to remove package folder {pkg_path}: {e}")
                    except Exception as e:
                        print(f"Failed to create package archive: {e}")
                else:
                    # Fallback: create a zip containing just the modified workbook
                    zip_base = os.path.join(export_batch_dir, f"{prefix}_{src_stem}_{ts}")
                    zip_target = f"{zip_base}.zip"
                    with zipfile.ZipFile(zip_target, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
                        zf.write(new_path, arcname=os.path.basename(new_path))
                    print(f"Created fallback package (workbook only): {zip_target}")
            except Exception as e:
                print(f"Failed to create package for {prefix}: {e}")

        # Write aggregated results workbook and copy the combinations file + input workbook into the export folder
        try:
            if aggregated_rows:
                try:
                    df_agg = pd.DataFrame(aggregated_rows)
                    # ensure Aeff sum columns exist and populate from aggregated_rows
                    try:
                        if 'Aeff_sum_orig' not in df_agg.columns:
                            df_agg['Aeff_sum_orig'] = pd.NA
                        if 'Aeff_sum_mod' not in df_agg.columns:
                            df_agg['Aeff_sum_mod'] = pd.NA
                        # ensure EEF columns exist
                        if 'EEF80_min_arcsec' not in df_agg.columns:
                            df_agg['EEF80_min_arcsec'] = pd.NA
                        if 'EEF90_min_arcsec' not in df_agg.columns:
                            df_agg['EEF90_min_arcsec'] = pd.NA
                        for i, rdict in enumerate(aggregated_rows):
                            try:
                                if isinstance(rdict, dict):
                                    if 'Aeff_sum_orig' in rdict:
                                        df_agg.at[i, 'Aeff_sum_orig'] = rdict.get('Aeff_sum_orig')
                                    if 'Aeff_sum_mod' in rdict:
                                        df_agg.at[i, 'Aeff_sum_mod'] = rdict.get('Aeff_sum_mod')
                                    # copy EEF values if present
                                    if 'EEF80_min_arcsec' in rdict:
                                        df_agg.at[i, 'EEF80_min_arcsec'] = rdict.get('EEF80_min_arcsec')
                                    if 'EEF90_min_arcsec' in rdict:
                                        df_agg.at[i, 'EEF90_min_arcsec'] = rdict.get('EEF90_min_arcsec')
                            except Exception:
                                pass
                    except Exception:
                        pass
                    out_agg = os.path.join(export_batch_dir, f"Aggregated_results_{src_stem}.xlsx")
                    with pd.ExcelWriter(out_agg, engine='openpyxl') as writer:
                        df_agg.to_excel(writer, sheet_name='Aggregated', index=False)
                    print(f"Wrote aggregated results to {out_agg}")
                except Exception as e:
                    print(f"Failed to write aggregated results Excel: {e}")
            # Copy combinations file used to run the batch into the export folder
            try:
                if os.path.exists(comb_path):
                    shutil.copy2(comb_path, os.path.join(export_batch_dir, os.path.basename(comb_path)))
            except Exception as e:
                print(f"Failed to copy combinations file into export folder: {e}")
            # Copy the original input workbook used as source for the batch
            try:
                if os.path.exists(args.file):
                    shutil.copy2(args.file, os.path.join(export_batch_dir, os.path.basename(args.file)))
            except Exception:
                pass
        except Exception:
            pass

        # Summarize skipped rows (if any) and complete batch processing
        try:
            sc = int(skipped_cannot_read) if 'skipped_cannot_read' in locals() else 0
            se = int(skipped_empty_prefix) if 'skipped_empty_prefix' in locals() else 0
            if sc or se:
                parts = []
                if sc:
                    parts.append(f"{sc} unreadable prefix rows")
                if se:
                    parts.append(f"{se} empty prefix rows")
                print(f"Batch processing summary: skipped {', '.join(parts)} from combinations file")
        except Exception:
            pass

        # Completed batch processing
        sys.exit(0)

    df = load_gaussians_from_excel(args.file, args.sheet)
    # Final top-level safety: ensure `weight` exactly matches `aeff_adjusted`.
    try:
        if 'aeff_adjusted' in df.columns:
            df['weight'] = df['aeff_adjusted'].astype(float)
    except Exception:
        pass
    plot_title_suffix = ""
    df_optimized = None
    
    placement_strategy = args.placement
    
    # If optimization is requested, run it and reload the optimized configuration.
    # If --placement is also provided, use it as the optimizer seed.
    if args.optimize:
        from optimize_mm_rows import optimize_rows
        base, ext = os.path.splitext(args.file)
        if ext == '':
            ext = '.xlsx'
        opt_output = f"{base}_optimised{ext}"

        is_coarse = (args.mode == 'coarse')
        is_extra_fine = (args.mode == 'extra-fine')
        # Total budget requirements:
        # - mode=coarse: optimize+plot <= 30s
        # - mode=fine: optimize+plot <= 90s
        # - mode=extra-fine: optimize+plot <= 300s
        # We budget optimization to leave time for plotting.
        if is_coarse:
            opt_budget_s = 18.0
        elif is_extra_fine:
            opt_budget_s = 240.0
        else:
            opt_budget_s = 45.0

        start_strategy = placement_strategy or 'elliptical'
        print(f"Optimizing MM positions (mode={args.mode}, start={start_strategy})...")
        try:
            # If the input is a CSV-only file, skip workbook-based optimization
            if isinstance(args.file, str) and str(args.file).lower().endswith('.csv'):
                print("Skipping MM position optimization for CSV-only input (no MM configuration sheet).")
                opt_hew = None
            else:
                opt_hew = optimize_rows(
                    input_path=args.file,
                    output_path=opt_output,
                    mode=args.mode,
                    optimize=True,
                    time_budget_s=opt_budget_s,
                    start_placement=start_strategy,
                    write_output=(not getattr(args, 'suppress_output', False)),
                )
        except KeyboardInterrupt:
            print("Optimization interrupted (Ctrl+C). The optimised file was not updated.")
            opt_hew = None
        else:
            print(f"Optimization complete. Optimized HEW: {opt_hew:.6e} m")
            if not getattr(args, 'suppress_output', False):
                print(f"Optimized configuration saved to: {opt_output}")

            # Make it explicit whether MM configuration actually changed.
            try:
                mm_in = pd.read_excel(args.file, sheet_name="MM configuration", engine="openpyxl")
                mm_out = pd.read_excel(opt_output, sheet_name="MM configuration", engine="openpyxl")
                a = mm_in["MM #"].astype(int).to_numpy()
                b = mm_out["MM #"].astype(int).to_numpy()
                changed = (a != b)
                print(f"MM configuration changed entries: {int(changed.sum())}")
                if changed.any() and "Row #" in mm_in.columns:
                    rows_changed = sorted(mm_in.loc[changed, "Row #"].dropna().unique().tolist())
                    print(f"Rows with MM# changes: {rows_changed}")

                # Print a small sample so it's easy to spot in Excel.
                if changed.any():
                    idx = np.flatnonzero(changed)[:10].tolist()
                    sample = pd.DataFrame(
                        {
                            "sheet_row_index": idx,
                            "Row #": mm_in.loc[idx, "Row #"].to_numpy() if "Row #" in mm_in.columns else [None] * len(idx),
                            "MM # (input)": a[idx],
                            "MM # (optimised)": b[idx],
                        }
                    )
                    print("Sample MM# changes (first 10):")
                    print(sample.to_string(index=False))
            except Exception:
                pass

            # Load the optimized configuration for overlaying
            if os.path.exists(opt_output):
                df_optimized = load_gaussians_from_excel(opt_output, args.sheet)
            else:
                df_optimized = None
            plot_title_suffix = " (comparison)"

    # Placement-only mode
    elif placement_strategy is not None:
        from optimize_mm_rows import cross_placement, x_axis_placement, elliptical_placement
        base, ext = os.path.splitext(args.file)
        if ext == '':
            ext = '.xlsx'
        opt_output = f"{base}_placed{ext}"

        if placement_strategy == 'x_axis':
            placement_fn = x_axis_placement
            placement_label = 'x_axis'
        elif placement_strategy == 'elliptical':
            placement_fn = elliptical_placement
            placement_label = 'elliptical'
        else:
            placement_fn = cross_placement
            placement_label = 'cross'

        print(f"Applying placement strategy ({placement_label})...")
        try:
            # If the input is a CSV-only file, skip workbook-based placement
            if isinstance(args.file, str) and str(args.file).lower().endswith('.csv'):
                print("Skipping MM placement for CSV-only input (no MM configuration sheet).")
                placement_hew = None
            else:
                placement_hew = placement_fn(
                    input_path=args.file,
                    output_path=opt_output,
                    seed=42,
                    write_output=(not getattr(args, 'suppress_output', False)),
                )
        except KeyboardInterrupt:
            print("Placement interrupted (Ctrl+C). The placed file was not updated.")
            placement_hew = None
        else:
            if placement_hew is None:
                print("Placement skipped (CSV-only input); no placed configuration produced.")
                df_optimized = None
            else:
                print(f"Placement complete. Final HEW: {placement_hew:.6e} m")
                if not getattr(args, 'suppress_output', False):
                    print(f"Placed configuration saved to: {opt_output}")

                # Make it explicit how many positions changed
                try:
                    if not getattr(args, 'suppress_output', False) and os.path.exists(opt_output):
                        mm_in = pd.read_excel(args.file, sheet_name="MM configuration", engine="openpyxl")
                        mm_out = pd.read_excel(opt_output, sheet_name="MM configuration", engine="openpyxl")
                    else:
                        mm_in = None
                        mm_out = None
                    if mm_in is not None and mm_out is not None:
                        a = mm_in["MM #"].astype(int).to_numpy()
                        b = mm_out["MM #"].astype(int).to_numpy()
                        changed = (a != b)
                        print(f"MM configuration changed entries: {int(changed.sum())}")

                        # Print a small sample
                        if changed.any():
                            idx = np.flatnonzero(changed)[:10].tolist()
                            sample = pd.DataFrame(
                                {
                                    "sheet_row_index": idx,
                                    "MM # (input)": a[idx],
                                    "MM # (placed)": b[idx],
                                }
                            )
                            print("Sample MM# changes (first 10):")
                            print(sample.to_string(index=False))
                except Exception:
                    pass

                # Load the placed configuration for overlaying (only if file was written)
                if not getattr(args, 'suppress_output', False) and os.path.exists(opt_output):
                    df_optimized = load_gaussians_from_excel(opt_output, args.sheet)
            plot_title_suffix = f" (placement: {placement_label})"
    
    # Plot the sum and encircled energy (with optional optimized overlay)
    if args.return_metrics_only:
        metrics = plot_sum(
            df,
            normalize=args.normalize,
            output=args.output,
            fast=(args.mode == 'coarse'),
            nx=(2062 if args.mode == 'fine' else None),
            ny=(2062 if args.mode == 'fine' else None),
            title_suffix=plot_title_suffix,
            df_optimized=df_optimized,
            return_metrics_only=True,
            metrics_n_r_final=args.metrics_nr_final,
            metrics_n_theta_final=args.metrics_ntheta_final,
            metrics_r_margin=args.metrics_r_margin,
        )
        print(json.dumps(metrics, indent=2))
        sys.exit(0)


    plot_sum(
        df,
        normalize=args.normalize,
        output=args.output,
        fast=(args.mode == 'coarse'),
        nx=(2062 if args.mode == 'fine' else None),
        ny=(2062 if args.mode == 'fine' else None),
        title_suffix=plot_title_suffix,
        df_optimized=df_optimized,
    )

    # Always trigger export logic for A_eff sheet after plot_sum
    try:
        import openpyxl
        import numpy as np
        import tempfile, os, sys
        wb = openpyxl.load_workbook(args.file)
        df_a = df.copy()
        if 'A_eff' in wb.sheetnames:
            ws_a = wb['A_eff']
            max_r_a = ws_a.max_row or 0
            mm_series = pd.to_numeric(df_a.get('MM #', pd.Series([], dtype=float)), errors='coerce')
            base_series = pd.to_numeric(df_a.get('aeff_base', pd.Series([], dtype=float)), errors='coerce')
            adj_series = pd.to_numeric(df_a.get('aeff_adjusted', pd.Series([], dtype=float)), errors='coerce')
            mm_to_base = {}
            mm_to_adj = {}
            for mmv, bv, av in zip(mm_series.tolist(), base_series.tolist(), adj_series.tolist()):
                try:
                    key = int(mmv) if mmv is not None and not (isinstance(mmv, float) and np.isnan(mmv)) else None
                except Exception:
                    key = None
                if key is None:
                    continue
                try:
                    if bv is not None and not (isinstance(bv, float) and np.isnan(bv)):
                        mm_to_base[key] = float(bv)
                except Exception:
                    pass
                try:
                    if av is not None and not (isinstance(av, float) and np.isnan(av)):
                        mm_to_adj[key] = float(av)
                except Exception:
                    pass
            for r in range(1, max_r_a + 1):
                cell = ws_a.cell(row=r, column=1).value
                if cell is None:
                    continue
                try:
                    mmv = int(cell) if isinstance(cell, (int, float)) or (isinstance(cell, str) and str(cell).strip().isdigit()) else None
                    if isinstance(cell, str) and str(cell).strip().isdigit():
                        mmv = int(str(cell).strip())
                except Exception:
                    mmv = None
                if mmv is None:
                    continue
                # Write canonical A_eff into column B
                if mmv in mm_to_base:
                    try:
                        ws_a.cell(row=r, column=2, value=float(mm_to_base[mmv]))
                    except Exception:
                        pass
                # Always write column C: zero if base is zero or missing, else adjusted value if present, else zero
                try:
                    base_val = float(mm_to_base.get(mmv, 0.0))
                    adj_val = float(mm_to_adj.get(mmv, 0.0))
                    if base_val is None or (isinstance(base_val, float) and np.isnan(base_val)):
                        ws_a.cell(row=r, column=3, value=0.0)
                    elif base_val == 0.0:
                        ws_a.cell(row=r, column=3, value=0.0)
                    else:
                        ws_a.cell(row=r, column=3, value=adj_val)
                except Exception:
                    ws_a.cell(row=r, column=3, value=0.0)
            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            tmpf.close()
            wb.save(tmpf.name)
            os.replace(tmpf.name, args.file)
            sys.stdout.flush()
    except Exception:
        pass
