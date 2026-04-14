import pandas as pd
import sys, os
# ensure repo root is on sys.path so we can import main.py when run from scripts/
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from main import load_gaussians_from_excel
from openpyxl import load_workbook

p='Distributions/test1.xlsx'
print('Loading and computing...')
df=load_gaussians_from_excel(p)
# Build mm->pos mapping from MM configuration
mmcfg = pd.read_excel(p, sheet_name='MM configuration', engine='openpyxl')
mm_to_pos = {}
if 'MM #' in mmcfg.columns:
    for order_i, (_, row) in enumerate(mmcfg.iterrows()):
        mm_num = row.get('MM #')
        if pd.isna(mm_num):
            continue
        mm_num_i = int(mm_num)
        if 'Position #' in mmcfg.columns:
            pos_val = row.get('Position #')
            if not pd.isna(pos_val):
                try:
                    mm_to_pos[mm_num_i] = int(float(pos_val))
                except Exception:
                    pass
            if mm_num_i not in mm_to_pos:
                mm_to_pos[mm_num_i] = int(order_i) + 1
        else:
            mm_to_pos[mm_num_i] = int(order_i) + 1

wb = load_workbook(p,data_only=True)
# read vignette sheet B values keyed by position
vig_azi = {}
vig_rad = {}
for sname, vigmap in [('Vignetting rotazi', vig_azi), ('Vignetting rotrad', vig_rad)]:
    if sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, ws.max_row+1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            if a is None:
                continue
            try:
                key = int(a) if isinstance(a,(int,float)) or (isinstance(a,str) and a.strip().isdigit()) else None
            except Exception:
                key = None
            if key is not None and b is not None:
                try:
                    vigmap[key] = float(b)
                except Exception:
                    pass

print('Comparing first 40 MMs:')
print('MM, pos, base, adj, adj/base(df), vig_rad*vig_azi(sheet)')
count = 0
for _, row in df.iterrows():
    try:
        mm = int(row['MM #'])
    except Exception:
        continue
    base = float(row.get('aeff_base', 0.0) or 0.0)
    adj = float(row.get('aeff_adjusted', 0.0) or 0.0)
    ratio_df = (adj/base) if base>0 else None
    pos = mm_to_pos.get(mm)
    prod = None
    if pos is not None:
        r = vig_rad.get(pos, None)
        a = vig_azi.get(pos, None)
        if r is not None and a is not None:
            prod = float(r)*float(a)
    print(mm, pos, base, adj, ratio_df, prod)
    count += 1
    if count >= 40:
        break

# show any mismatches where both sides exist but differ by >1e-6
print('\nMISMATCHES (ratio_df vs prod)')
for _, row in df.iterrows():
    try:
        mm = int(row['MM #'])
    except Exception:
        continue
    base = float(row.get('aeff_base', 0.0) or 0.0)
    adj = float(row.get('aeff_adjusted', 0.0) or 0.0)
    if base <= 0:
        continue
    ratio_df = adj/base
    pos = mm_to_pos.get(mm)
    if pos is None:
        continue
    r = vig_rad.get(pos, None)
    a = vig_azi.get(pos, None)
    if r is None or a is None:
        continue
    prod = float(r)*float(a)
    if abs(ratio_df - prod) > 1e-9:
        print(f'MM {mm} pos {pos} ratio_df={ratio_df:.12f} prod={prod:.12f} diff={ratio_df-prod:.12e}')

print('\nDone')
