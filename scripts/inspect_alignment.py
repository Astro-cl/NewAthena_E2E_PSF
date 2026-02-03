from openpyxl import load_workbook

wb = load_workbook('Distributions/NewTest_Distribution.xlsx', data_only=True)
print('sheets:', wb.sheetnames)
ws = wb['Alignment']
rows = list(ws.iter_rows(values_only=True))
# find header row
header_row = None
for i, r in enumerate(rows[:40]):
    if r and r[0] == 'Position #':
        header_row = i
        break
if header_row is None:
    raise SystemExit('header not found')
header = rows[header_row]
print('Header:', header)
# collect numeric data rows
data = rows[header_row+1:]
# parse mm config
mmc = wb['MM configuration']
mmc_rows = list(mmc.iter_rows(values_only=True))
mmc_header = mmc_rows[0]
cols = {name: idx for idx, name in enumerate(mmc_header)}
mm_to_pos = {}
if 'MM #' in cols:
    for r in mmc_rows[1:]:
        if not r or r[cols['MM #']] is None:
            continue
        try:
            mm = int(r[cols['MM #']])
        except Exception:
            continue
        pos = None
        if 'Position #' in cols and r[cols['Position #']] is not None:
            try:
                pos = int(r[cols['Position #']])
            except Exception:
                pos = None
        mm_to_pos[mm] = pos
print('MM->pos sample:', {100: mm_to_pos.get(100), 300: mm_to_pos.get(300)})

# find data rows for those positions
for mm in (100, 300):
    pos = mm_to_pos.get(mm)
    print('\n--- MM', mm, 'pos', pos, '---')
    if pos is not None:
        for r in data:
            if r and isinstance(r[0], (int, float)) and int(r[0]) == pos:
                # print first 10 columns
                vals = r[:10]
                print(vals)
                break
    # also check if Alignment has MM # column
    if 'MM #' in header:
        mm_idx = header.index('MM #')
        for r in data:
            if r and len(r) > mm_idx and r[mm_idx] == mm:
                print('row by MM #:', r[:10])
                break

# show any rows with nonzero rotazi/rotrad
print('\nRows with nonzero rot values:')
for r in data:
    if not r or r[0] is None:
        continue
    try:
        pos = int(r[0])
    except Exception:
        continue
    rotazi = r[4] if len(r) > 4 else None
    rotrad = r[5] if len(r) > 5 else None
    rotz = r[6] if len(r) > 6 else None
    if any((isinstance(x, (int, float)) and x != 0) for x in (rotazi, rotrad, rotz)):
        print('pos', pos, 'rotazi', rotazi, 'rotrad', rotrad, 'rotz', rotz)
