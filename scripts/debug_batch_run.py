import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
import subprocess
import shutil


def make_files(workdir: Path):
    dist = workdir / 'Distributions'
    dist.mkdir(parents=True, exist_ok=True)
    sample_input = dist / 'sample_input.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Thermal'
    ws.append(['Position #', 'd_therm_rotx', 'd_therm_roty', 'd_therm_z'])
    ws.append([1, 0.0, 0.0, 0.0])
    ws2 = wb.create_sheet('Vignetting rotrad')
    ws2.append(['col1','col2','col3'])
    ws2.append([0,0,0])
    ws3 = wb.create_sheet('Vignetting rotazi')
    ws3.append(['col1','col2','col3'])
    ws3.append([0,0,0])
    # Add a minimal MM_PSF sheet and MM configuration so main.py can load the workbook
    ws4 = wb.create_sheet('MM_PSF')
    ws4.append(['MM #', 'm_rad [arcsec]', 'm_azi [arcsec]', 'sigma_rad [arcsec]', 'sigma_azi [arcsec]'])
    ws4.append([1, 0.0, 0.0, 1.0, 1.0])

    ws_cfg = wb.create_sheet('MM configuration')
    ws_cfg.append(['MM #', 'x_MM [m]', 'r_MM [m]'])
    ws_cfg.append([1, 0.0, 0.0])
    # Add A_eff sheet with base A_eff for MM #1
    ws_a = wb.create_sheet('A_eff')
    ws_a.append([1, 1.0])

    wb.save(sample_input)

    combos = pd.DataFrame(
        [
            [1, 'cfgA', 1.0, 5.0, 0.1],
            [2, 'cfgB', 2.5, 6.0, 0.2],
            [3, 'cfgC', 0.5, 7.0, 0.0],
        ], columns=['A','B','C','D','E']
    )
    combos_path = dist / 'combinations.xlsx'
    combos.to_excel(combos_path, index=False)
    return sample_input, combos_path


def run(workdir: Path):
    sample_input, combos_path = make_files(workdir)
    repo_root = Path(__file__).resolve().parents[1]
    main_py = repo_root / 'main.py'
    cmd = [sys.executable, str(main_py), '--batch-combinations', str(combos_path), '--file', str(sample_input)]
    print('Running:', ' '.join(cmd), 'cwd=', str(workdir))
    proc = subprocess.run(cmd, cwd=str(workdir), stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    print('RETURN CODE:', proc.returncode)
    print('STDOUT:\n', proc.stdout)
    print('STDERR:\n', proc.stderr)
    print('Exports dir exists?:', (workdir / 'Exports').exists())

if __name__ == '__main__':
    td = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('tmp_batch_debug')
    td = td.resolve()
    if td.exists():
        shutil.rmtree(td)
    td.mkdir(parents=True)
    run(td)
