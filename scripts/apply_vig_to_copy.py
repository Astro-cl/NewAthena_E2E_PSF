"""Copy Test_Distribution.xlsx -> Test_Distribution2.xlsx, compute vignette factors, write them and save.

Run: python3 scripts/apply_vig_to_copy.py
"""
from pathlib import Path
import shutil
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook

SRC = Path('Distributions/Test_Distribution.xlsx')
DST = Path('Distributions/Test_Distribution2.xlsx')

if not SRC.exists():
    print('Source not found:', SRC)
    sys.exit(2)

print('Copying', SRC, '->', DST)
shutil.copy2(SRC, DST)

print('Reading MM configuration')
mmc = pd.read_excel(DST, sheet_name='MM configuration', engine='openpyxl')
mm_to_pos = {}
pos_to_cfg_row = {}
mm_config_map = {}
for order_i, (_, row) in enumerate(mmc.iterrows()):
    mmnum = row.get('MM #')
    if pd.isna(mmnum):
        continue
    mmn = int(mmnum)
    if 'Position #' in mmc.columns:
        posv = row.get('Position #')
        if not pd.isna(posv):
            try:
                mm_to_pos[mmn] = int(float(posv))
            except Exception:
                pass
        if mmn not in mm_to_pos:
            mm_to_pos[mmn] = order_i + 1
    else:
        mm_to_pos[mmn] = order_i + 1
    cfg = order_i + 1
    pos = mm_to_pos.get(mmn, cfg)
    pos_to_cfg_row[pos] = cfg
    mm_config_map[mmn] = {'x_MM': row.get('x_MM [m]', 0)}

print('MM->pos count', len(mm_to_pos))

def read_pos_map(path, sheet, keys):
    out = {}
    try:
        df = pd.read_excel(path, sheet_name=sheet, engine='openpyxl')
    except Exception as e:
        print('Could not read', sheet, e)
        return out
    if 'Position #' in df.columns:
        for _, r in df.iterrows():
            pos = r.get('Position #')
            if pd.isna(pos):
                continue
            out[int(pos)] = {k: r.get(k, 0.0) for k in keys}
    elif 'MM #' in df.columns and mm_to_pos:
        for _, r in df.iterrows():
            mm = r.get('MM #')
            if pd.isna(mm):
                continue
            pos = mm_to_pos.get(int(mm))
            if pos is None:
                continue
            out[pos] = {k: r.get(k, 0.0) for k in keys}
    return out

print('Reading perturbations')
alignment = read_pos_map(DST, 'Alignment', ['d_align_rotrad', 'd_align_rotazi'])
gravity = read_pos_map(DST, 'Gravity offload', ['d_grav_rotrad', 'd_grav_rotazi'])
thermal = read_pos_map(DST, 'Thermal', ['d_therm_rotrad', 'd_therm_rotazi'])
print('Perturb maps sizes', len(alignment), len(gravity), len(thermal))

print('Computing rot projections via main.compute_total_rot_polar')
try:
    sys.path.insert(0, str(Path('.').resolve()))
    import main
    _, _, rot_rad_map, rot_azi_map = main.compute_total_rot_polar(mm_to_pos, mm_config_map, alignment, gravity, thermal)
    print('Rot maps sizes', len(rot_rad_map), len(rot_azi_map))
except Exception as e:
    print('Failed compute_total_rot_polar:', e)
    rot_rad_map = {}
    rot_azi_map = {}

print('Parsing vignetting sheets')
try:
    vdf_azi = pd.read_excel(DST, sheet_name='Vignetting rotazi', engine='openpyxl', header=None)
    vdf_rad = pd.read_excel(DST, sheet_name='Vignetting rotrad', engine='openpyxl', header=None)
    print('Vign shapes', vdf_azi.shape, vdf_rad.shape)
except Exception as e:
    print('Failed reading vignetting sheets', e)
    vdf_azi = vdf_rad = None

def build_vig(vdf):
    out = {}
    if vdf is None or vdf.shape[1] < 11:
        return out
    for _, r in vdf.iterrows():
        try:
            cfg = r.iloc[7]
            if pd.isna(cfg):
                continue
            cfg = int(float(cfg))
            # column I is in arcmin; convert to arcsec for interpolation
            x = r.iloc[8]
            try:
                x = float(x) * 60.0
            except Exception:
                x = None
            e = r.iloc[9]
            y = r.iloc[10]
            if pd.isna(x) or pd.isna(y):
                continue
            key_str = (cfg, str(e).strip())
            key_num = None
            try:
                key_num = (cfg, float(e))
            except Exception:
                key_num = None
            out.setdefault(key_str, {'xs': [], 'ys': []})
            out[key_str]['xs'].append(float(x))
            out[key_str]['ys'].append(float(y))
            if key_num is not None:
                out.setdefault(key_num, {'xs': [], 'ys': []})
                out[key_num]['xs'].append(float(x))
                out[key_num]['ys'].append(float(y))
        except Exception:
            continue
    for k, v in list(out.items()):
        order = np.argsort(v['xs'])
        out[k] = (np.array(v['xs'], dtype=float)[order], np.array(v['ys'], dtype=float)[order])
    return out

ys_azi = build_vig(vdf_azi)
ys_rad = build_vig(vdf_rad)
print('Built vig arrays counts:', len(ys_azi), len(ys_rad))

print('Computing vignette factors for each MM')
vig_vals_azi = {}
vig_vals_rad = {}
for mm, pos in mm_to_pos.items():
    cfg = pos_to_cfg_row.get(pos)
    f_azi = 1.0
    try:
        if cfg is not None:
            # try any match for that cfg
            matches = [k for k in ys_azi.keys() if k[0] == cfg]
                if matches:
                xs, ys = ys_azi[matches[0]]
                f_azi = float(np.interp(abs(float(rot_azi_map.get(pos, 0.0))), xs, ys))
    except Exception:
        f_azi = 1.0
    f_rad = 1.0
    try:
        if cfg is not None:
            matches = [k for k in ys_rad.keys() if k[0] == cfg]
                if matches:
                xs, ys = ys_rad[matches[0]]
                f_rad = float(np.interp(abs(float(rot_rad_map.get(pos, 0.0))), xs, ys))
    except Exception:
        f_rad = 1.0
    vig_vals_azi[pos] = float(f_azi)
    vig_vals_rad[pos] = float(f_rad)

print('Writing vignette values to workbook')
wb = load_workbook(DST)

def write_vig_sheet(wb_obj, sheet_name, vig_map):
    if sheet_name not in wb_obj.sheetnames:
        return 0
    ws = wb_obj[sheet_name]
    pos_to_row = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 40) + 1):
            cell = ws.cell(row=r, column=c).value
            try:
                if cell is None:
                    continue
                if isinstance(cell, (int, float)) and float(cell).is_integer():
                    v = int(cell)
                    if v not in pos_to_row:
                        pos_to_row[v] = r
                else:
                    s = str(cell).strip()
                    if s.isdigit():
                        v = int(s)
                        if v not in pos_to_row:
                            pos_to_row[v] = r
            except Exception:
                continue
    count = 0
    for pos_k, val in (vig_map or {}).items():
        if pos_k in pos_to_row:
            r = pos_to_row[pos_k]
            ws.cell(row=r, column=2, value=float(val))
            count += 1
    return count

c1 = write_vig_sheet(wb, 'Vignetting rotazi', vig_vals_azi)
c2 = write_vig_sheet(wb, 'Vignetting rotrad', vig_vals_rad)

print('Writing adjusted A_eff to column C')
if 'A_eff' in wb.sheetnames:
    ws_a = wb['A_eff']
    mm_to_row = {}
    for r in range(2, ws_a.max_row + 1):
        try:
            mm = ws_a.cell(row=r, column=1).value
            if mm is None:
                continue
            mm_to_row[int(float(mm))] = r
        except Exception:
            continue
    c3 = 0
    for mm, pos in mm_to_pos.items():
        r = mm_to_row.get(mm)
        if r is None:
            continue
        base = ws_a.cell(row=r, column=2).value
        try:
            base_f = float(base)
        except Exception:
            base_f = 1.0
        f_azi = vig_vals_azi.get(pos, 1.0)
        f_rad = vig_vals_rad.get(pos, 1.0)
        adj = base_f * float(f_azi) * float(f_rad)
        ws_a.cell(row=r, column=3, value=float(adj))
        c3 += 1
else:
    c3 = 0

print('Saving workbook', DST)
wb.save(DST)
print('Saved. rotazi_written=', c1, 'rotrad_written=', c2, 'aeff_adjusted=', c3)
