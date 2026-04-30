import argparse
import csv
import glob
import os
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd

# CLI example:
# .venv-1/bin/python scripts/fit_target_hew_batch.py \
#   --target-hew 8.964 \
#   --batch-file "batch_combinations.xlsx" \
#   --input-file "input_distributions.xlsx" \
#   --mode fine \
#   --metric-column HEW_min_arcsec \
#   --out-csv "tmp_mm_sigmas_target.csv" \
#   --out-workbook "tmp_patched_input.xlsx"


def _load_mm_to_row(workbook_path: Path) -> Tuple[Dict[int, int], List[int]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "MM configuration" not in wb.sheetnames:
        wb.close()
        raise RuntimeError("Workbook is missing 'MM configuration' sheet")

    ws = wb["MM configuration"]
    mm_to_row: Dict[int, int] = {}
    for r in range(2, (ws.max_row or 0) + 1):
        mm = ws.cell(r, 4).value
        rn = ws.cell(r, 3).value
        if mm is None:
            continue
        try:
            mm_i = int(float(mm))
            rn_i = int(float(rn)) if rn is not None else None
        except (TypeError, ValueError):
            continue
        if rn_i is None:
            continue
        mm_to_row[mm_i] = rn_i

    wb.close()
    if not mm_to_row:
        raise RuntimeError("No MM->row mapping found in 'MM configuration'")
    rows = sorted(set(mm_to_row.values()))
    return mm_to_row, rows


def _build_templates(
    rows_count: int,
    strategy: str,
    sr_min: float,
    sr_max: float,
    sa_min: float,
    sa_max: float,
    eps: float,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    sr_lo = np.array([sr_min + eps * (i + 1) for i in range(rows_count)], dtype=float)
    sa_lo = np.array([sa_min + eps * (i + 1) for i in range(rows_count)], dtype=float)

    if strategy == "spread":
        sr_hi = np.linspace(sr_min + 0.05, sr_max, rows_count)
        sa_hi = np.linspace(sa_max, sa_min + 0.05, rows_count)
    elif strategy == "max_all":
        sr_hi = np.array([sr_max - eps * (rows_count - i) for i in range(rows_count)], dtype=float)
        sa_hi = np.array([sa_max - eps * (rows_count - i) for i in range(rows_count)], dtype=float)
    else:
        raise ValueError(f"Unsupported strategy: {strategy}")

    return sr_lo, sa_lo, sr_hi, sa_hi


def _build_maps(
    t: float,
    mm_to_row: Dict[int, int],
    rows: List[int],
    sr_lo: np.ndarray,
    sa_lo: np.ndarray,
    sr_hi: np.ndarray,
    sa_hi: np.ndarray,
    sr_min: float,
    sr_max: float,
    sa_min: float,
    sa_max: float,
) -> Tuple[Dict[int, float], Dict[int, float], Dict[int, float], Dict[int, float]]:
    sr_by_row: Dict[int, float] = {}
    sa_by_row: Dict[int, float] = {}
    for i, row_id in enumerate(rows):
        sr = float(np.clip((1.0 - t) * sr_lo[i] + t * sr_hi[i], sr_min, sr_max))
        sa = float(np.clip((1.0 - t) * sa_lo[i] + t * sa_hi[i], sa_min, sa_max))
        # Keep tiny row-unique offsets to avoid accidental duplicates.
        sr_by_row[row_id] = round(sr + i * 1e-6, 6)
        sa_by_row[row_id] = round(sa + i * 1e-6, 6)

    sr_map = {mm: sr_by_row[mm_to_row[mm]] for mm in mm_to_row}
    sa_map = {mm: sa_by_row[mm_to_row[mm]] for mm in mm_to_row}
    return sr_map, sa_map, sr_by_row, sa_by_row


def _pick_aggregated_row(df_agg: pd.DataFrame, config_prefix: Optional[str]) -> pd.Series:
    if df_agg.empty:
        raise RuntimeError("Aggregated results workbook is empty")

    if config_prefix:
        for col in df_agg.columns:
            if str(col).strip().lower() in ("prefix", "config", "configuration", "name", "case", "label"):
                m = df_agg[col].astype(str).str.strip() == str(config_prefix).strip()
                if m.any():
                    return df_agg.loc[m].iloc[0]
        raise RuntimeError(
            f"Requested config prefix '{config_prefix}' not found in aggregated workbook columns {list(df_agg.columns)}"
        )

    return df_agg.iloc[0]


def _patch_sigmas_into_workbook(workbook_path: Path, sr_map: Dict[int, float], sa_map: Dict[int, float]) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    if "MM_PSF" not in wb.sheetnames:
        wb.close()
        raise RuntimeError("Workbook is missing 'MM_PSF' sheet")
    ws = wb["MM_PSF"]

    updated = 0
    for ri in range(2, (ws.max_row or 0) + 1):
        mm = ws.cell(ri, 1).value
        if mm is None:
            continue
        try:
            mm_i = int(float(mm))
        except (TypeError, ValueError):
            continue
        if mm_i in sr_map:
            ws.cell(ri, 4).value = sr_map[mm_i]
            ws.cell(ri, 5).value = sa_map[mm_i]
            updated += 1

    wb.save(workbook_path)
    wb.close()
    if updated == 0:
        raise RuntimeError("No MM_PSF rows were updated with sigma values")


def _evaluate_metric_with_exact_batch(
    input_workbook: Path,
    batch_file: Path,
    python_exe: Path,
    main_py: Path,
    sr_map: Dict[int, float],
    sa_map: Dict[int, float],
    mode: str,
    metric_column: str,
    config_prefix: Optional[str],
) -> Tuple[float, float, Path]:
    with tempfile.TemporaryDirectory(prefix="fit_target_hew_batch_") as td:
        tdp = Path(td)
        patched_wb = tdp / input_workbook.name
        shutil.copy2(input_workbook, patched_wb)
        _patch_sigmas_into_workbook(patched_wb, sr_map, sa_map)

        cmd = [
            str(python_exe),
            str(main_py),
            "--batch-combinations",
            str(batch_file),
            "-f",
            str(patched_wb),
            "--mode",
            str(mode),
        ]
        proc = subprocess.run(cmd, cwd=str(tdp), capture_output=True, text=True)
        if proc.returncode != 0:
            stderr_tail = "\n".join(proc.stderr.splitlines()[-40:])
            stdout_tail = "\n".join(proc.stdout.splitlines()[-40:])
            raise RuntimeError(
                "Batch run failed\n"
                f"Command: {' '.join(cmd)}\n"
                f"Return code: {proc.returncode}\n"
                f"STDOUT (tail):\n{stdout_tail}\n"
                f"STDERR (tail):\n{stderr_tail}"
            )

        agg_candidates = glob.glob(str(tdp / "Exports" / "Export_*" / "Aggregated_results_*.xlsx"))
        if not agg_candidates:
            raise RuntimeError("Aggregated results not found after batch run")
        agg_path = Path(sorted(agg_candidates)[-1])

        df_agg = pd.read_excel(agg_path, engine="openpyxl")
        row = _pick_aggregated_row(df_agg, config_prefix)

        if metric_column not in row.index:
            raise RuntimeError(
                f"Metric column '{metric_column}' not found in aggregated results columns: {list(df_agg.columns)}"
            )
        if "HEW_00_arcsec" not in row.index:
            raise RuntimeError("Expected 'HEW_00_arcsec' column missing from aggregated results")

        metric_val = float(row[metric_column])
        hew_00 = float(row["HEW_00_arcsec"])
        return metric_val, hew_00, agg_path


def _write_sigma_csv(out_csv: Path, mm_to_row: Dict[int, int], sr_map: Dict[int, float], sa_map: Dict[int, float]) -> None:
    with out_csv.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["MM_num", "Row", "sigma_rad", "sigma_azi"])
        for mm in sorted(mm_to_row):
            w.writerow([mm, mm_to_row[mm], sr_map[mm], sa_map[mm]])


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Seek a target HEW metric by fitting row-unique sigma values using exact batch execution."
    )
    parser.add_argument("--target-hew", type=float, required=True, help="Target metric value in arcsec (e.g. 8.964).")
    parser.add_argument("--batch-file", required=True, help="Path to batch combinations workbook.")
    parser.add_argument("--input-file", required=True, help="Path to input distribution workbook.")
    parser.add_argument("--metric-column", default="HEW_min_arcsec", help="Aggregated metric column to match (default: HEW_min_arcsec).")
    parser.add_argument("--config-prefix", default=None, help="Optional configuration prefix to select row in aggregated workbook.")
    parser.add_argument("--mode", default="fine", choices=["coarse", "fine"], help="Run mode for main.py batch execution.")
    parser.add_argument("--max-steps", type=int, default=20, help="Maximum bisection iterations.")
    parser.add_argument("--tolerance", type=float, default=5e-4, help="Absolute metric tolerance for convergence.")

    parser.add_argument("--sr-min", type=float, default=4.1303 * (7.5 / 8.0), help="Lower bound for sigma_rad.")
    parser.add_argument("--sr-max", type=float, default=4.74, help="Upper bound for sigma_rad.")
    parser.add_argument("--sa-min", type=float, default=1.2633 * (7.5 / 8.0), help="Lower bound for sigma_azi.")
    parser.add_argument("--sa-max", type=float, default=1.36, help="Upper bound for sigma_azi.")
    parser.add_argument("--eps", type=float, default=1e-4, help="Small offset used to preserve row uniqueness.")

    parser.add_argument("--python-exe", default=sys.executable if "sys" in globals() else None, help="Python executable to run main.py.")
    parser.add_argument("--main-py", default=None, help="Path to main.py (default: repo-root/main.py).")
    parser.add_argument("--out-csv", default=None, help="Output CSV path for winning MM sigmas.")
    parser.add_argument("--patch-input-in-place", action="store_true", help="Patch winning sigmas directly into --input-file.")
    parser.add_argument("--out-workbook", default=None, help="Write patched workbook copy to this path (if not patching in-place).")

    args = parser.parse_args()

    script_path = Path(__file__).resolve()
    repo_root = script_path.parents[1]

    input_file = Path(args.input_file).resolve()
    batch_file = Path(args.batch_file).resolve()
    if not input_file.exists():
        raise SystemExit(f"Input workbook not found: {input_file}")
    if not batch_file.exists():
        raise SystemExit(f"Batch file not found: {batch_file}")

    main_py = Path(args.main_py).resolve() if args.main_py else (repo_root / "main.py")
    if not main_py.exists():
        raise SystemExit(f"main.py not found: {main_py}")

    python_exe = Path(args.python_exe).resolve() if args.python_exe else Path(os.sys.executable).resolve()
    if not python_exe.exists():
        raise SystemExit(f"Python executable not found: {python_exe}")

    safe_target = str(args.target_hew).replace(".", "p")
    out_csv = Path(args.out_csv).resolve() if args.out_csv else (repo_root / f"row_unique_sigmas_target{safe_target}_{input_file.stem}_exact_batch.csv")

    print("=== Generic target-HEW batch fitter ===")
    print(f"Target metric: {args.metric_column} = {args.target_hew:.6f} arcsec")
    print(f"Input workbook: {input_file}")
    print(f"Batch file: {batch_file}")
    print(f"main.py: {main_py}")
    print(f"Python: {python_exe}")
    print(f"Mode: {args.mode}")
    if args.config_prefix:
        print(f"Config prefix filter: {args.config_prefix}")
    print()

    mm_to_row, rows = _load_mm_to_row(input_file)
    n = len(rows)
    print(f"Loaded MM->row map: {len(mm_to_row)} MMs across {n} rows")

    bracketed = False
    templates = None
    hew_lo = hew_hi = None
    print("Searching for bracketing template strategy...")
    for strategy in ("spread", "max_all"):
        print(f"  Strategy: {strategy}")
        sr_lo, sa_lo, sr_hi, sa_hi = _build_templates(
            rows_count=n,
            strategy=strategy,
            sr_min=args.sr_min,
            sr_max=args.sr_max,
            sa_min=args.sa_min,
            sa_max=args.sa_max,
            eps=args.eps,
        )

        lo_maps = _build_maps(0.0, mm_to_row, rows, sr_lo, sa_lo, sr_hi, sa_hi, args.sr_min, args.sr_max, args.sa_min, args.sa_max)
        hi_maps = _build_maps(1.0, mm_to_row, rows, sr_lo, sa_lo, sr_hi, sa_hi, args.sr_min, args.sr_max, args.sa_min, args.sa_max)

        hew_lo, hew00_lo, _ = _evaluate_metric_with_exact_batch(
            input_workbook=input_file,
            batch_file=batch_file,
            python_exe=python_exe,
            main_py=main_py,
            sr_map=lo_maps[0],
            sa_map=lo_maps[1],
            mode=args.mode,
            metric_column=args.metric_column,
            config_prefix=args.config_prefix,
        )
        hew_hi, hew00_hi, _ = _evaluate_metric_with_exact_batch(
            input_workbook=input_file,
            batch_file=batch_file,
            python_exe=python_exe,
            main_py=main_py,
            sr_map=hi_maps[0],
            sa_map=hi_maps[1],
            mode=args.mode,
            metric_column=args.metric_column,
            config_prefix=args.config_prefix,
        )

        print(f"    t=0.0 -> metric={hew_lo:.6f}, HEW_00={hew00_lo:.6f}")
        print(f"    t=1.0 -> metric={hew_hi:.6f}, HEW_00={hew00_hi:.6f}")
        if hew_lo <= args.target_hew <= hew_hi:
            bracketed = True
            templates = (sr_lo, sa_lo, sr_hi, sa_hi)
            print(f"    Bracketed with strategy '{strategy}'")
            break

    if not bracketed or templates is None:
        raise SystemExit(
            "Target not bracketed by tested template strategies. "
            f"Observed range: [{hew_lo:.6f}, {hew_hi:.6f}]"
        )

    print("\nRunning bisection...")
    sr_lo, sa_lo, sr_hi, sa_hi = templates
    t_lo, t_hi = 0.0, 1.0
    best = None
    best_abs_err = float("inf")

    for step in range(1, args.max_steps + 1):
        t_mid = 0.5 * (t_lo + t_hi)
        sr_map, sa_map, sr_by_row, sa_by_row = _build_maps(
            t_mid,
            mm_to_row,
            rows,
            sr_lo,
            sa_lo,
            sr_hi,
            sa_hi,
            args.sr_min,
            args.sr_max,
            args.sa_min,
            args.sa_max,
        )
        metric_val, hew_00, _ = _evaluate_metric_with_exact_batch(
            input_workbook=input_file,
            batch_file=batch_file,
            python_exe=python_exe,
            main_py=main_py,
            sr_map=sr_map,
            sa_map=sa_map,
            mode=args.mode,
            metric_column=args.metric_column,
            config_prefix=args.config_prefix,
        )

        err = metric_val - args.target_hew
        abs_err = abs(err)
        print(f"  step {step:02d}: t={t_mid:.8f} -> metric={metric_val:.6f} (err={err:+.6f}), HEW_00={hew_00:.6f}")

        if abs_err < best_abs_err:
            best_abs_err = abs_err
            best = (t_mid, metric_val, hew_00, sr_map, sa_map, sr_by_row, sa_by_row)

        if err < 0:
            t_lo = t_mid
        else:
            t_hi = t_mid

        if abs_err < args.tolerance:
            print(f"  Converged: |error| < {args.tolerance}")
            break

    if best is None:
        raise SystemExit("No valid solution found")

    t_best, metric_best, hew00_best, sr_best, sa_best, sr_row_best, sa_row_best = best
    print("\nFinal best solution:")
    print(f"  t={t_best:.8f}")
    print(f"  {args.metric_column}={metric_best:.6f} arcsec (target {args.target_hew:.6f})")
    print(f"  HEW_00_arcsec={hew00_best:.6f}")
    print(f"  |error|={best_abs_err:.6f}")
    print(f"  Row uniqueness sigma_rad={len(set(sr_row_best.values())) == len(rows)}")
    print(f"  Row uniqueness sigma_azi={len(set(sa_row_best.values())) == len(rows)}")

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    _write_sigma_csv(out_csv, mm_to_row, sr_best, sa_best)
    print(f"\nWrote sigma CSV: {out_csv}")

    if args.patch_input_in_place:
        _patch_sigmas_into_workbook(input_file, sr_best, sa_best)
        print(f"Patched input workbook in-place: {input_file}")
    elif args.out_workbook:
        out_wb = Path(args.out_workbook).resolve()
        out_wb.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(input_file, out_wb)
        _patch_sigmas_into_workbook(out_wb, sr_best, sa_best)
        print(f"Wrote patched workbook copy: {out_wb}")


if __name__ == "__main__":
    import sys

    main()