#!/usr/bin/env python3
import pandas as pd
from openpyxl import load_workbook, Workbook
import tempfile
import shutil
import os
import sys
from pathlib import Path
# ensure repo root is on sys.path so `import main` works
repo_root = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(repo_root))
import main
from pathlib import Path
import sys

p = sys.argv[1] if len(sys.argv) > 1 else 'Distributions/1__Ang__20__E__7__Def__0_MA_15042026.xlsx'
print('Loading DataFrame via load_gaussians_from_excel...')
df = main.load_gaussians_from_excel(p, sheet='MM_PSF')


def make_values_only_copy(src_path: str) -> str:
    """Create a copy of the workbook where all cells are values (no formulas).
    Try to use Excel via xlwings (preferred) to force evaluation and paste-values.
    If xlwings or Excel is not available, fall back to copying cached values using openpyxl's data_only read.
    Returns path to the values-only workbook copy (temporary file)."""
    src_path = str(src_path)
    dest_fd, dest_path = tempfile.mkstemp(prefix='values_only_', suffix='.xlsx', dir=os.getcwd())
    os.close(dest_fd)
    try:
        import xlwings as xw
        try:
            # Open Excel invisibly, paste values on each sheet and save to a new file
            app = xw.App(visible=False)
            wb_x = app.books.open(src_path)
            for sh in wb_x.sheets:
                # select and paste values for entire sheet
                sh.api.Cells.Copy()
                sh.api.Cells.PasteSpecial(-4163)  # xlPasteValues constant
            wb_x.api.SaveAs(dest_path)
            wb_x.close()
            app.quit()
            return dest_path
        except Exception:
            try:
                wb_x.close()
            except Exception:
                pass
            try:
                app.quit()
            except Exception:
                pass
            # fallthrough to openpyxl method
    except Exception:
        # xlwings not available or failed; fall back
        pass

    # Fallback: read cached values with openpyxl and write them into a new workbook
    wb = load_workbook(src_path, data_only=True)
    new_wb = Workbook()
    # remove default sheet created by Workbook()
    default = new_wb.active
    new_wb.remove(default)
    for name in wb.sheetnames:
        src = wb[name]
        dest = new_wb.create_sheet(title=name)
        for r in src.iter_rows(values_only=True):
            dest.append(list(r))
    new_wb.save(dest_path)
    return dest_path


values_only_path = make_values_only_copy(p)
print('Using values-only workbook copy:', values_only_path)
wb = load_workbook(values_only_path, data_only=True)

# find MM_PSF sheet
sheet = None
if 'MM_PSF' in wb.sheetnames:
    sheet = wb['MM_PSF']
else:
    for s in wb.sheetnames:
        if s.lower() == 'mm_psf':
            sheet = wb[s]
            break
if sheet is None:
    raise SystemExit('MM_PSF sheet not found')

# read rows: MM # col A, sigma_rad D, sigma_azi E
rows = []
start_row = 2
for r in range(start_row, sheet.max_row + 1):
    mm = sheet.cell(row=r, column=1).value
    if mm is None:
        continue
    try:
        mm_int = int(float(mm))
    except Exception:
        continue
    # Primary: columns 4/5 (D/E)
    d = sheet.cell(row=r, column=4).value
    e = sheet.cell(row=r, column=5).value
    rows.append({'MM #': mm_int, 'sigma_rad_cell': d, 'sigma_azi_cell': e, 'excel_row': r})

# Do not attempt to read alternate sigma columns; only use D/E values.

sheet_df = pd.DataFrame(rows)

# Merge on MM # if present
if 'MM #' in df.columns:
    try:
        df['MM #'] = pd.to_numeric(df['MM #'], errors='coerce').astype('Int64')
    except Exception:
        pass
    merged = pd.merge(df.reset_index(drop=True), sheet_df, on='MM #', how='left')
else:
    merged = df.copy().reset_index(drop=True)
    s2 = sheet_df.reset_index(drop=True)
    # align by index where possible
    merged = pd.concat([merged, s2[['sigma_rad_cell', 'sigma_azi_cell']].reindex(merged.index)], axis=1)

out_cols = ['MM #', 'aeff_base', 'aeff_vig_factor_rad', 'aeff_vig_factor_azi', 'aeff_vig_factor', 'aeff_adjusted', 'weight', 'sigma_rad [arcsec]', 'sigma_azi [arcsec]', 'sigma_rad', 'sigma_azi', 'sigma_rad_cell', 'sigma_azi_cell']
for c in out_cols:
    if c not in merged.columns:
        merged[c] = pd.NA

# Ensure CSV reflects final adjusted A_eff as authoritative weight
try:
    merged['aeff_adjusted'] = pd.to_numeric(merged.get('aeff_adjusted', 0.0), errors='coerce').fillna(0.0)
    merged['weight'] = merged['aeff_adjusted'].astype(float)
except Exception:
    # fallback: coerce existing weight to numeric
    merged['weight'] = pd.to_numeric(merged.get('weight', 0.0), errors='coerce').fillna(0.0)

# Do not fall back to DataFrame sigma columns; diagnostics must report
# the raw D/E cell cached values only (may be empty if workbook lacks caches).

out_path = 'tmp_mm_sigmas.csv'
merged[out_cols].to_csv(out_path, index=False)
print(f'Wrote {out_path} (rows={len(merged)})')
