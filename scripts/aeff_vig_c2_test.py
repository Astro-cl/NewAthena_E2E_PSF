import re, tempfile, pathlib
from openpyxl import load_workbook
import shutil

bp = pathlib.Path('Distributions/TestDistribution.xlsx')
if not bp.exists():
    print('TestDistribution.xlsx missing')
    raise SystemExit(2)

# parse preset name
preset = 'Variable 10% 1 keV'
# Prefer number immediately followed by 'keV', otherwise take last numeric token
m = re.search(r"(\d+(?:\.\d*)?)\s*(?:keV)", preset, flags=re.IGNORECASE)
if not m:
    all_nums = re.findall(r"(\d+(?:\.\d*)?)", preset)
    sval = all_nums[-1] if all_nums else None
else:
    sval = m.group(1)
sel = None
if sval is not None:
    try:
        sel = float(sval)
    except Exception:
        sel = None
print('Parsed sel_energy =', sel)

# copy workbook to temp and write C2
tmp = pathlib.Path(tempfile.mktemp(suffix='.xlsx'))
shutil.copy(bp, tmp)
wb2 = load_workbook(tmp)
if sel is not None:
    for s in ('Vignetting rotazi','Vignetting rotrad'):
        if s in wb2.sheetnames:
            wb2[s].cell(row=2, column=3, value=float(sel))
wb2.save(tmp)
# read back
wb3 = load_workbook(tmp)
for s in ('Vignetting rotazi','Vignetting rotrad'):
    if s in wb3.sheetnames:
        print(s, 'C2=', wb3[s].cell(row=2, column=3).value)
print('Saved file:', tmp)
