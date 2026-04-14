import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Distributions/test10kev.xlsx', data_only=False)
ws = wb['A_eff']

print("Exploring the lookup table range ($S$24:$AA$38):")
print("\n  Columns S-AA, Rows 24-38:")

# S is column 19, AA is column 27 in 1-based indexing (openpyxl)
for r in range(24, 39):
    row_data = []
    for c in range(19, 28):  # S to AA
        cell = ws.cell(r, c)
        val_str = str(cell.value)[:20] if cell.value else ""
        row_data.append(val_str)
    print(f"  Row {r}: {row_data}")

# Also check if there's any data in columns S-AA
print("\n  First 5 rows of columns S-AA:")
for r in range(1, 6):
    row_data = []
    for c in range(19, 28):
        cell = ws.cell(r, c)
        val_str = str(cell.value)[:20] if cell.value else ""
        row_data.append(val_str)
    print(f"  Row {r}: {row_data}")
