from pathlib import Path
import itertools
import pandas as pd
import traceback
import sys
import sys
import re
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from tools import run_sensitivity as rs
from main import load_gaussians_from_excel

SENS_PATH = rs.SENS_PATH
BASE_WORKBOOK = rs.BASE_WORKBOOK
OUT_XLSX = rs.OUT_XLSX

if not SENS_PATH.exists():
    print('Sensitivity file not found at', SENS_PATH)
    sys.exit(1)

sens = pd.read_excel(SENS_PATH, sheet_name=0, engine='openpyxl', header=0)
param_options = {}
for col in sens.columns:
    vals = sens[col].dropna().astype(str).map(str.strip)
    vals = [v for v in vals if v not in ['', '-', 'NaN', 'nan', 'None']]
    if vals:
        param_options[col] = vals

if 'A_eff' in param_options:
    vals = param_options['A_eff']
    expanded = []
    row_token_re = rs.re.compile(r'^(.*?)\s*\[row#\]\s*$', re.IGNORECASE)
    aeff_map_tmp = rs.build_aeff_mapping()
    if aeff_map_tmp.get('mm_row_map'):
        row_keys = sorted(aeff_map_tmp.get('mm_row_map').keys())
    else:
        row_keys = []
    for v in vals:
        m = row_token_re.match(v)
        if m and row_keys:
            base = m.group(1).strip()
            for r in row_keys:
                expanded.append(f"{base} [row{r}]")
        else:
            expanded.append(v)
    param_options['A_eff'] = expanded

standard_presets = rs.load_standard_mm_psf_presets(BASE_WORKBOOK)
keys = list(param_options.keys())
combos = list(itertools.product(*(param_options[k] for k in keys)))
print(f'Will run {len(combos)} combinations (one-shot)')
results = []
for idx, combo in enumerate(combos, start=1):
    combo_dict = dict(zip(keys, combo))
    print(f'[ {idx}/{len(combos)} ] {combo_dict}')
    try:
        df = load_gaussians_from_excel(str(BASE_WORKBOOK), sheet='MM_PSF')
        mapping = None
        if 'A_eff' in combo_dict:
            try:
                mapping = rs.find_aeff_weights_for_choice(combo_dict['A_eff'], rs.build_aeff_mapping())
                df['weight'] = df['MM #'].astype(int).map(mapping)
            except Exception as e:
                print('  Warning: could not apply A_eff choice:', e)
        if 'MM_PSF' in combo_dict:
            df = rs.apply_mm_psf_choice_to_df(df, combo_dict['MM_PSF'], rs.build_aeff_mapping(), rs.np.random.default_rng(hash(str(combo_dict))%2**32), standard_presets=standard_presets)
        # Create a temporary workbook by copying the BASE_WORKBOOK and
        # replacing the `MM_PSF` sheet with our modified `df`. This ensures
        # main.py can find all expected sheets (A_eff, MM configuration, etc.).
        import tempfile, shutil
        from openpyxl import load_workbook
        tmpdir = tempfile.mkdtemp(prefix='sensitivity_')
        tmp_in = str(Path(tmpdir) / Path(BASE_WORKBOOK).name)
        shutil.copy2(str(BASE_WORKBOOK), tmp_in)
        try:
            wb = load_workbook(tmp_in)
            if 'MM_PSF' in wb.sheetnames:
                std = wb['MM_PSF']
                wb.remove(std)
            wb.save(tmp_in)
            # Use pandas to write the MM_PSF sheet into the copied workbook
            with pd.ExcelWriter(tmp_in, engine='openpyxl', mode='a') as w:
                df.to_excel(w, sheet_name='MM_PSF', index=False)
        except Exception:
            # fallback: overwrite file entirely with only MM_PSF sheet
            with pd.ExcelWriter(tmp_in, engine='openpyxl') as w:
                df.to_excel(w, sheet_name='MM_PSF', index=False)

        # Run main.py on the temp workbook with quick metric overrides
        import subprocess, json, sys
        cmd = [sys.executable, str(Path(__file__).resolve().parent.parent / 'main.py'), '-f', tmp_in, '--return_metrics_only', '--metrics-nr-final', '300', '--metrics-ntheta-final', '24', '--metrics-r-margin', '6.0']
        try:
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
            if proc.returncode != 0:
                metrics = {'error': proc.stderr.strip(), 'stdout': proc.stdout, 'stderr': proc.stderr}
            else:
                try:
                    metrics = json.loads(proc.stdout)
                except Exception:
                    metrics = {'error': 'invalid-json', 'stdout': proc.stdout, 'stderr': proc.stderr}
        except subprocess.TimeoutExpired:
            metrics = {'error': 'timeout'}
        finally:
            try:
                shutil.rmtree(tmpdir)
            except Exception:
                pass
        row = {**combo_dict}
        # Normalize metrics capture: if subprocess returned a dict with an 'error'
        # field, record it; otherwise extract known metric keys.
        if isinstance(metrics, dict):
            if 'error' in metrics:
                row['error'] = metrics.get('error')
                # include raw stdout/stderr when available for debugging
                if 'stdout' in metrics:
                    row['proc_stdout'] = metrics.get('stdout')
                if 'stderr' in metrics:
                    row['proc_stderr'] = metrics.get('stderr')
            else:
                # extract expected metric keys, allow missing keys
                for k in ('hew_origin_arcsec', 'hew_best_arcsec', 'hew_opt_arcsec', 'eef90_origin_arcsec', 'eef90_best_arcsec', 'eef90_opt_arcsec'):
                    row[k] = metrics.get(k)
        else:
            row['error'] = 'no-metrics'
        results.append(row)
    except Exception as e:
        print('  ERROR in combination:', e)
        print(traceback.format_exc())
        results.append({**combo_dict, 'error': str(e)})

try:
    pd.DataFrame(results).to_excel(OUT_XLSX, index=False)
    print('Wrote results to', OUT_XLSX)
except Exception as e:
    print('Failed to write results file:', e)
    print(traceback.format_exc())
