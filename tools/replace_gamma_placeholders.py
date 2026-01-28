#!/usr/bin/env python3
import pandas as pd
import numpy as np
import re
import hashlib
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

fp = Path('sensitivity/input/20260125T112521Z_3_A_eff1_keV_MM_PSF50_Variable_Pseudo-Voigt_8_alpha_10_Alignment0_Gravity_offloadGZ_Thermal30_deg_FMS_Tilt.xlsx')
if not fp.exists():
    raise SystemExit(f'Target not found: {fp}')
mm = pd.read_excel(fp, sheet_name='MM_PSF', engine='openpyxl')

sr_col = 'sigma_rad_'
sa_col = 'sigma_azi_'

# seed
chosen = '50% Variable Pseudo-Voigt 8" (alpha 10%)'
h = int(hashlib.sha256((fp.name + chosen).encode()).hexdigest()[:8], 16)
rng = np.random.default_rng(h)

def parse_gamma_str(s):
    if not isinstance(s, str):
        return None
    m = re.search(r"gamma\s*\(\s*([0-9.]+)\s*,\s*([0-9.]+)%\*?([0-9.]+)\s*\)", s)
    if m:
        mu = float(m.group(1)); pct = float(m.group(2)); ref = float(m.group(3))
        sigma = (pct / 100.0) * mu
        return mu, sigma
    m2 = re.search(r"gamma\s*\(\s*([0-9.]+)\s*,\s*([0-9.]+)%\*?([^)]+)\)", s)
    if m2:
        mu = float(m2.group(1)); pct = float(m2.group(2))
        sigma = (pct / 100.0) * mu
        return mu, sigma
    return None

n = len(mm)
count = 0
for i in range(n):
    v = mm.at[i, sr_col]
    if isinstance(v, str) and 'gamma' in v.lower():
        parsed = parse_gamma_str(v)
        if parsed:
            mu, sigma = parsed
            if sigma <= 0 or mu <= 0:
                val = mu if mu > 0 else 1e-6
            else:
                k = (mu / sigma) ** 2
                theta = (sigma ** 2) / mu
                val = float(rng.gamma(shape=k, scale=theta, size=1)[0])
            mm.at[i, sr_col] = val
            count += 1
    w = mm.at[i, sa_col]
    if isinstance(w, str) and 'gamma' in w.lower():
        parsed = parse_gamma_str(w)
        if parsed:
            mu, sigma = parsed
            if sigma <= 0 or mu <= 0:
                val = mu if mu > 0 else 1e-6
            else:
                k = (mu / sigma) ** 2
                theta = (sigma ** 2) / mu
                val = float(rng.gamma(shape=k, scale=theta, size=1)[0])
            mm.at[i, sa_col] = val
            count += 1

print('replaced gamma placeholders for', count, 'entries')

# copy into canonical columns
if 'sigma_rad [arcsec]' in mm.columns and sr_col in mm.columns:
    mm['sigma_rad [arcsec]'] = mm[sr_col]
if 'sigma_azi [arcsec]' in mm.columns and sa_col in mm.columns:
    mm['sigma_azi [arcsec]'] = mm[sa_col]

# save
wb = load_workbook(fp)
if 'MM_PSF' in wb.sheetnames:
    del wb['MM_PSF']
ws = wb.create_sheet('MM_PSF')
for r in dataframe_to_rows(mm, index=False, header=True):
    ws.append(r)
wb.save(fp)
print('saved')
