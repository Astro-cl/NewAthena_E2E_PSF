import openpyxl as _openpyxl

# Load the workbook
wb = _openpyxl.load_workbook('Distributions/test10kev.xlsx', data_only=False)

# Check what sheets are available
print("Available sheets:", wb.sheetnames)

# Check the MM configuration sheet
if 'MM configuration' in wb.sheetnames:
    ws = wb['MM configuration']
    print(f"\nMM configuration sheet:")
    print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Show first few rows and columns
    print("\n  First 5 rows, columns A-H:")
    for r in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for c in range(1, 9):
            cell = ws.cell(r, c)
            row_data.append(cell.value)
        print(f"    Row {r}: {row_data}")

# Check the A_eff sheet structure
ws_aeff = wb['A_eff']
print(f"\nA_eff sheet:")
print(f"  Max row: {ws_aeff.max_row}, Max col: {ws_aeff.max_column}")

# Look at the formula columns and see what they reference
print("\n  Formula references (first formula in each column):")
for col in [10, 11, 12, 16]:  # Columns J, K, L, P (openpyxl 1-based)
    cell = ws_aeff.cell(2, col)  # Row 2, MM 1
    if cell.value and isinstance(cell.value, str):
        print(f"    Col {chr(64+col)}: {cell.value[:80]}")
