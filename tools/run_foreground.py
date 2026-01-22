#!/usr/bin/env python3
from pathlib import Path
import sys
ROOT=Path(__file__).resolve().parent.parent
sys.path.insert(0,str(ROOT))
from tools.run_sensitivity import build_aeff_mapping, apply_mm_psf_choice_to_df, load_standard_mm_psf_presets, SENS_PATH, OUT_XLSX, BASE_WORKBOOK, find_aeff_weights_for_choice
from main import load_gaussians_from_excel, plot_sum
import pandas as pd, numpy as np, itertools, hashlib, re, shutil, subprocess, os, sys, json
from numpy.random import default_rng

# Load sensitivity file
sens = pd.read_excel(SENS_PATH, sheet_name=0, engine='openpyxl', header=0)
param_options = {}
for col in sens.columns:
    vals = sens[col].dropna().astype(str).map(str.strip)
    vals = [v for v in vals if v not in ['', '-', 'NaN', 'nan', 'None']]
    if vals:
        param_options[col] = vals

# build aeff_map and attach mm_row_map for expansion
aeff_map = build_aeff_mapping()
try:
    from tools.run_sensitivity import load_mm_row_map
    aeff_map['mm_row_map'] = load_mm_row_map(BASE_WORKBOOK)
except Exception:
    aeff_map['mm_row_map'] = None

# Choose A_eff x MM_PSF if available
if 'A_eff' in param_options and 'MM_PSF' in param_options:
    keys = ['A_eff','MM_PSF']
else:
    keys = list(param_options.keys())[:2]

combos = list(itertools.product(*(param_options[k] for k in keys)))
# expand A_eff tokens in param_options if present
if 'A_eff' in param_options:
    vals = param_options['A_eff']
    expanded = []
    row_token_re = re.compile(r'^(.*?)\s*\[row#\]\s*$', re.IGNORECASE)
    if aeff_map.get('mm_row_map'):
        row_keys = sorted(aeff_map['mm_row_map'].keys())
    else:
        row_keys = []
    for v in vals:
        m = row_token_re.match(v)
        if m and row_keys:
            base = m.group(1).strip()
            for r in row_keys:
                expanded.append(f"{base} [row{r}]")
        else:
            expanded.append(v)
    param_options['A_eff'] = expanded

# Choose A_eff x MM_PSF if available
if 'A_eff' in param_options and 'MM_PSF' in param_options:
    keys = ['A_eff','MM_PSF']
else:
    keys = list(param_options.keys())[:2]

combos = list(itertools.product(*(param_options[k] for k in keys)))
print('Foreground sweep: columns=', keys, ' combos=', len(combos))

standard = load_standard_mm_psf_presets(BASE_WORKBOOK)
results = []

for idx, combo in enumerate(combos, start=1):
    combo_dict = dict(zip(keys, combo))
    print(f'[{idx}/{len(combos)}] {combo_dict}')
    try:
        h = hashlib.sha1(repr(combo_dict).encode('utf8')).hexdigest()
        seed = int(h[:16], 16) % (2**32)
        rng = default_rng(seed)

        df = load_gaussians_from_excel(str(BASE_WORKBOOK), sheet='MM_PSF')

        if 'A_eff' in combo_dict:
            try:
                mapvals = find_aeff_weights_for_choice(combo_dict['A_eff'], aeff_map)
                df['weight'] = df['MM #'].astype(int).map(mapvals)
            except Exception as e:
                print('  Warning: A_eff apply failed:', e)

        if 'MM_PSF' in combo_dict:
            df = apply_mm_psf_choice_to_df(df, combo_dict['MM_PSF'], aeff_map, rng, standard_presets=standard)

        # Optional placement/optimizer invocation (mirror run_sensitivity behaviour)
        # By default perform placement-only using 'elliptical' strategy and do not optimize.
        df_opt = None
        # Compute baseline metrics before placement for diagnostics
        try:
            pre_metrics = plot_sum(df, normalize=True, fast=True, df_optimized=None, return_metrics_only=True)
        except Exception:
            pre_metrics = None
        # Try to run placement in-process to avoid writing Excel files.
        placed_metrics = None
        df_opt = None
        try:
            from optimize_mm_rows import load_all_sheets, _load_base_params_from_workbook, _load_position_deltas, _elliptical_place_mm_config, rebuild_df
            sheets = load_all_sheets(str(BASE_WORKBOOK))
            if 'MM configuration' in sheets:
                mm_config = sheets['MM configuration'].copy()
                base_params = _load_base_params_from_workbook(str(BASE_WORKBOOK))
                # If A_eff mapping was applied to the runtime PSF dataframe, ensure
                # the optimizer uses the same per-MM weights for ranking/placement.
                try:
                    if 'mapvals' in locals() and mapvals is not None:
                        # Map weights onto base_params in the same way optimizer expects
                        mapped = base_params['MM #'].astype(int).map(mapvals).fillna(0.0)
                        total_mapped = float(mapped.sum())
                        MIN_TOTAL_WEIGHT = 1e-12
                        if not np.isfinite(total_mapped) or total_mapped <= MIN_TOTAL_WEIGHT:
                            print('    Warning: total A_eff weight for this combo is zero or too small; skipping optimization')
                            df_opt = None
                            raise RuntimeError('skip_optimization_due_to_small_total_weight')
                        # Otherwise apply mapped weights
                        base_params['weight'] = mapped
                except RuntimeError:
                    # Controlled skip: propagate to outer handler to fall back
                    raise
                except Exception:
                    pass
                alignment_by_pos, gravity_by_pos, thermal_by_pos = _load_position_deltas(str(BASE_WORKBOOK))
                # Respect combo overrides from Sensitivity.xlsx: zero position deltas when specified.
                def _is_zero_token(v):
                    try:
                        if v is None:
                            return False
                        s = str(v).strip()
                        if s == '':
                            return False
                        return float(s) == 0.0
                    except Exception:
                        return False

                def _zero_map_for_positions(mm_config_df):
                    if 'Position #' in mm_config_df.columns:
                        pos_series = pd.to_numeric(mm_config_df['Position #'], errors='coerce')
                        pos_list = [int(x) for x in pos_series.dropna().astype(int).unique().tolist()]
                    else:
                        pos_list = list(range(1, len(mm_config_df) + 1))
                    return pos_list

                combo_keys = {k.lower(): k for k in combo_dict.keys()}
                if 'alignment' in combo_keys or 'align' in combo_keys:
                    raw = combo_dict.get(combo_keys.get('alignment', combo_keys.get('align')))
                    if _is_zero_token(raw):
                        pos_list = _zero_map_for_positions(mm_config)
                        alignment_by_pos = {p: {'d_align_rad': 0.0, 'd_align_azi': 0.0, 'd_align_z': 0.0, 'd_align_rotz': 0.0} for p in pos_list}
                # gravity
                raw = None
                for candidate in ('gravity', 'gravity offload', 'grav'):
                    if candidate in combo_keys:
                        raw = combo_dict.get(combo_keys[candidate])
                        break
                if _is_zero_token(raw):
                    pos_list = _zero_map_for_positions(mm_config)
                    gravity_by_pos = {p: {'d_grav_x': 0.0, 'd_grav_y': 0.0, 'd_grav_z': 0.0, 'd_grav_rotz': 0.0} for p in pos_list}
                # thermal
                raw = None
                for candidate in ('thermal', 'therm'):
                    if candidate in combo_keys:
                        raw = combo_dict.get(combo_keys[candidate])
                        break
                if _is_zero_token(raw):
                    pos_list = _zero_map_for_positions(mm_config)
                    thermal_by_pos = {p: {'d_therm_x': 0.0, 'd_therm_y': 0.0, 'd_therm_z': 0.0, 'd_therm_rotz': 0.0} for p in pos_list}
                placed_mm = _elliptical_place_mm_config(
                    mm_config,
                    base_params,
                    alignment_by_pos=alignment_by_pos,
                    gravity_by_pos=gravity_by_pos,
                    thermal_by_pos=thermal_by_pos,
                    seed=int(seed),
                )
                final_df = rebuild_df(base_params, placed_mm)
                df_opt = final_df
            else:
                df_opt = None
        except Exception as e:
            print('  In-process placement failed; attempting subprocess fallback:', e)
            # Fallback: create a temporary workbook with the per-combo A_eff weights
            try:
                import tempfile, shutil
                from openpyxl import load_workbook

                tmpdir = tempfile.mkdtemp(prefix="sensitivity_")
                tmp_in = os.path.join(tmpdir, os.path.basename(str(BASE_WORKBOOK)))
                shutil.copy2(str(BASE_WORKBOOK), tmp_in)

                # Patch A_eff sheet weights (column B) based on mapvals (mm->weight)
                if 'mapvals' in locals() and mapvals is not None:
                    wb = load_workbook(tmp_in)
                    # find A_eff sheet case-insensitive
                    sheet_name = None
                    for name in wb.sheetnames:
                        if str(name).strip().lower() == 'a_eff' or str(name).strip().lower() == 'a_eff':
                            sheet_name = name
                            break
                    # fallback to literal 'A_eff'
                    if sheet_name is None and 'A_eff' in wb.sheetnames:
                        sheet_name = 'A_eff'
                    if sheet_name is None:
                        # try first sheet named like 'A eff' or similar
                        for name in wb.sheetnames:
                            if 'a' in name.lower() and 'eff' in name.lower():
                                sheet_name = name
                                break
                    if sheet_name is not None:
                        ws = wb[sheet_name]
                        # find header row for MM # in column A
                        header_row = None
                        for r in range(1, min(40, ws.max_row) + 1):
                            v = ws.cell(row=r, column=1).value
                            if isinstance(v, str) and v.strip().lower().replace(' ', '') in {'mm#', 'mm'}:
                                header_row = r
                                break
                        data_start = header_row + 1 if header_row is not None else 2
                        for r in range(data_start, ws.max_row + 1):
                            mm_cell = ws.cell(row=r, column=1).value
                            try:
                                mmn = int(mm_cell)
                            except Exception:
                                continue
                            if mmn in mapvals:
                                ws.cell(row=r, column=2).value = float(mapvals[mmn])
                        wb.save(tmp_in)

                # Run optimize_mm_rows.py in subprocess to perform placement on the patched workbook
                tmp_out = os.path.join(tmpdir, 'placed_output.xlsx')
                cmd = [sys.executable, str(ROOT / 'optimize_mm_rows.py'), '-f', tmp_in, '-o', tmp_out, '--optimize', '--mode', 'coarse']
                proc = subprocess.run(cmd, capture_output=True, text=True)
                if proc.returncode != 0:
                    print('    Subprocess placement failed:', proc.stderr)
                    df_opt = None
                else:
                    # Load placed mm_config and base_params from subprocess output
                    try:
                        from optimize_mm_rows import load_all_sheets, _load_base_params_from_workbook, rebuild_df
                        sheets2 = load_all_sheets(tmp_out)
                        if 'MM configuration' in sheets2:
                            mm_config2 = sheets2['MM configuration'].copy()
                            base_params2 = _load_base_params_from_workbook(tmp_out)
                            placed_mm2 = mm_config2
                            final_df2 = rebuild_df(base_params2, placed_mm2)
                            df_opt = final_df2
                        else:
                            df_opt = None
                    except Exception as ee:
                        print('    Failed to load placed output:', ee)
                        df_opt = None
                try:
                    shutil.rmtree(tmpdir)
                except Exception:
                    pass
            except Exception as ex:
                print('    Subprocess fallback failed:', ex)
                df_opt = None

        # If a placement dataframe was produced, compute post-placement metrics as diagnostics
        try:
            metrics = plot_sum(df, normalize=True, fast=True, df_optimized=(df_opt if df_opt is not None else None), return_metrics_only=True)
        except Exception:
            metrics = None
        # Post-placement raw metrics (before clamping)
        post_metrics_raw = None
        if df_opt is not None:
            try:
                post_metrics_raw = plot_sum(df, normalize=True, fast=True, df_optimized=df_opt, return_metrics_only=True)
            except Exception:
                post_metrics_raw = None
        # If placement returned metrics via stdout, merge optimized metrics
        if isinstance(metrics, dict) and 'placed_metrics' in locals() and placed_metrics:
            for kk, vv in placed_metrics.items():
                metrics[kk] = vv
        keep = [
            'hew_origin_arcsec','hew_best_arcsec','hew_opt_arcsec',
            'eef90_origin_arcsec','eef90_best_arcsec','eef90_opt_arcsec'
        ]
        trimmed = {k: (metrics.get(k) if isinstance(metrics, dict) else None) for k in keep}
        # Ensure optimized metrics are not worse than the 'best' metrics; clamp if necessary
        try:
            if trimmed.get('hew_opt_arcsec') is not None and trimmed.get('hew_best_arcsec') is not None:
                if trimmed['hew_opt_arcsec'] > trimmed['hew_best_arcsec'] * 1.0:
                    # If optimization produced worse HEW, keep the best (no-op placement assumed)
                    trimmed['hew_opt_arcsec'] = trimmed['hew_best_arcsec']
            if trimmed.get('eef90_opt_arcsec') is not None and trimmed.get('eef90_best_arcsec') is not None:
                if trimmed['eef90_opt_arcsec'] > trimmed['eef90_best_arcsec'] * 1.0:
                    trimmed['eef90_opt_arcsec'] = trimmed['eef90_best_arcsec']
        except Exception:
            pass
        # Parse MM_PSF into three detail columns: Sym/Asym, size token, Fixed/Variable
        def _parse_mm_psf_details(s):
            s0 = '' if s is None else str(s)
            ss = s0.strip()
            sym = ''
            size = ''
            fv = ''
            # Capture percentage variability if present (e.g. '10%') and remove all percent tokens from the string
            m_pct = re.search(r'(\d+(?:\.\d+)?)\s*%', ss)
            pct_token = None
            if m_pct:
                pct_token = m_pct.group(1).strip()
                ss = re.sub(r'\d+(?:\.\d+)?\s*%', '', ss)
            if re.search(r'\bSym\b|Symmetric|Symmetry', ss, re.IGNORECASE):
                sym = 'Sym'
            elif re.search(r'\bAsym\b|Asymmetric', ss, re.IGNORECASE):
                sym = 'Asym'
            # Fixed vs Variable detection
            if re.search(r'\bFixed\b', ss, re.IGNORECASE):
                fv = 'Fixed'
            elif pct_token is not None:
                fv = f'{pct_token}% Variable'
            elif re.search(r'\bVariable\b|gaussian\(|Variable Sym', ss, re.IGNORECASE):
                fv = 'Variable'
            else:
                if ss and not re.search(r'[0-9]', ss):
                    fv = 'Preset'
            # size token extraction
            m_pair = re.search(r'sigma_rad\s*=\s*([^,;\)]+).*sigma_azi\s*=\s*([^,;\)]+)', ss, re.IGNORECASE)
            if m_pair:
                size = f"{m_pair.group(1).strip()} * {m_pair.group(2).strip()}"
            else:
                # Match numbers that include an explicit unit (don't capture bare integers like '10' from '10%')
                m_nums = re.findall(r'([0-9]+\.?[0-9]*\s*(?:micron|µm|um|"|arcsec|arcsecond))', ss, re.IGNORECASE)
                if m_nums:
                    size = ' * '.join([m.strip() for m in m_nums])
                else:
                    q = re.search(r'"([^\"]+)"', ss)
                    if q:
                        size = q.group(1)
                    else:
                        size = ss
            return sym, size, fv

        sym, size_tok, fv = _parse_mm_psf_details(combo_dict.get('MM_PSF'))

        # extract row number from A_eff choices like '1 keV [row3]' and strip from A_eff
        row_num = '-'
        combo_out = dict(combo_dict)
        aeff_val = combo_out.get('A_eff')
        if aeff_val is not None:
            mrow = re.match(r'^(.*?)\s*\[row(\d+)\]\s*$', str(aeff_val), re.IGNORECASE)
            if mrow:
                combo_out['A_eff'] = mrow.group(1).strip()
                try:
                    row_num = int(mrow.group(2))
                except Exception:
                    row_num = '-'
        combo_out['row'] = row_num
        combo_out['MM_symmetry'] = sym
        combo_out['MM_size_token'] = size_tok
        combo_out['MM_fixed_variable'] = fv
        # Attach diagnostics: raw post-placement metrics and whether placement improved HEW
        try:
            if post_metrics_raw and isinstance(post_metrics_raw, dict):
                trimmed['hew_opt_raw_arcsec'] = post_metrics_raw.get('hew_opt_arcsec')
                trimmed['eef90_opt_raw_arcsec'] = post_metrics_raw.get('eef90_opt_arcsec')
                # placement_improved if raw optimized HEW is strictly smaller than best HEW
                trimmed['placement_improved'] = (post_metrics_raw.get('hew_opt_arcsec') is not None and trimmed.get('hew_best_arcsec') is not None and post_metrics_raw.get('hew_opt_arcsec') < trimmed.get('hew_best_arcsec'))
            else:
                trimmed['hew_opt_raw_arcsec'] = None
                trimmed['eef90_opt_raw_arcsec'] = None
                trimmed['placement_improved'] = False
        except Exception:
            trimmed['hew_opt_raw_arcsec'] = None
            trimmed['eef90_opt_raw_arcsec'] = None
            trimmed['placement_improved'] = False

        row = {**combo_out, **trimmed}
        results.append(row)
    except Exception as e:
        print('  ERROR:', e)
        results.append({**combo_dict, 'error': str(e)})

# save results
try:
    df_res = pd.DataFrame(results)
    if 'row' not in df_res.columns:
        df_res['row'] = '-'
    # Reorder columns: A_eff, row, MM_PSF, parsed MM cols, then rest
    cols = list(df_res.columns)
    desired_prefix = []
    for nm in ('A_eff','row','MM_PSF','MM_symmetry','MM_size_token','MM_fixed_variable'):
        if nm in cols:
            desired_prefix.append(nm)
    remaining = [c for c in cols if c not in desired_prefix]
    df_res = df_res[desired_prefix + remaining]
    df_res.to_excel(OUT_XLSX, index=False)
    print('\nWrote', OUT_XLSX)
    print(df_res.to_string(index=False))
except Exception as e:
    print('Failed to write results:', e)
