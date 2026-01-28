#!/usr/bin/env python3
from pathlib import Path
import numpy as np, pandas as pd, hashlib, datetime
import importlib.util

# load helper to parse standard defs
spec_path = Path('sensitivity/sensitivity_run.py')
spec = importlib.util.spec_from_file_location('srun', spec_path)
srun = importlib.util.module_from_spec(spec)
spec.loader.exec_module(srun)

baseline = Path('Distributions/Test_Distribution.xlsx')
std = srun.load_standard_mm_psf_defs(baseline)

fp = Path('sensitivity/input/20260125T112521Z_3_A_eff1_keV_MM_PSF50_Variable_Pseudo-Voigt_8_alpha_10_Alignment0_Gravity_offloadGZ_Thermal30_deg_FMS_Tilt.xlsx')
if not fp.exists():
    raise SystemExit(f'Target not found: {fp}')

# find matching preset key
chosen_guess = '50% Variable Pseudo-Voigt 8" (alpha 10%)'
chosen = None
if chosen_guess in std:
    chosen = chosen_guess
else:
    for k in std.keys():
        kl = k.lower()
        if 'pseudo' in kl and '50' in kl and '8' in kl:
            chosen = k; break
    if not chosen:
        for k in std.keys():
            if 'pseudo' in k.lower() or 'variable' in k.lower():
                chosen = k; break

print('Chosen preset:', chosen)
if not chosen:
    raise SystemExit('No matching preset found')
entry = std[chosen]

mm = pd.read_excel(fp, sheet_name='MM_PSF', engine='openpyxl')
# detect canonical sigma cols
sr_cols = [c for c in mm.columns if isinstance(c, str) and 'sigma_rad' in c.lower()]
sa_cols = [c for c in mm.columns if isinstance(c, str) and 'sigma_azi' in c.lower()]
if not sr_cols or not sa_cols:
    raise SystemExit('sigma columns not found')
sr_col = sr_cols[0]; sa_col = sa_cols[0]

# sampling RNG
h = int(hashlib.sha256((fp.name + str(chosen)).encode('utf-8')).hexdigest()[:8], 16)
rng = np.random.default_rng(h)

n = min(len(mm), 600)

def sample_from_def(defn, size):
    if defn is None:
        return np.zeros(size)
    dist = defn.get('dist')
    if dist == 'fixed':
        return np.full(size, float(defn.get('value', 0.0)))
    if dist == 'gaussian':
        mu = float(defn.get('mean', 0.0)); sigma = float(defn.get('sigma', 0.0))
        if sigma <= 0 or mu <= 0:
            return np.full(size, mu if mu > 0 else 1e-6)
        out = rng.normal(loc=mu, scale=sigma, size=size)
        mask = out <= 0
        attempts = 0
        while mask.any() and attempts < 100:
            out[mask] = rng.normal(loc=mu, scale=sigma, size=mask.sum())
            mask = out <= 0; attempts += 1
        out[out <= 0] = 1e-6
        return out
    if dist == 'gamma':
        mu = float(defn.get('mean', 0.0)); sigma = float(defn.get('sigma', 0.0))
        if sigma <= 0 or mu <= 0:
            return np.full(size, mu if mu > 0 else 1e-6)
        k = (mu / sigma) ** 2
        theta = (sigma ** 2) / mu
        out = rng.gamma(shape=k, scale=theta, size=size)
        out[out <= 0] = 1e-6
        return out
    if dist == 'uniform':
        lo = float(defn.get('min', 0.0)); hi = float(defn.get('max', lo))
        return rng.uniform(lo, hi, size=size)
    return np.zeros(size)

new_sr = sample_from_def(entry.get('sigma_rad'), n)
new_sa = sample_from_def(entry.get('sigma_azi'), n)

for i in range(n):
    mm.at[i, sr_col] = float(new_sr[i])
    mm.at[i, sa_col] = float(new_sa[i])

# alphas
ar_cols = [c for c in mm.columns if isinstance(c, str) and 'alpha_rad' in c.lower()]
aa_cols = [c for c in mm.columns if isinstance(c, str) and 'alpha_azi' in c.lower()]
if ar_cols and aa_cols:
    def sample_alpha(defn, size):
        if defn is None:
            return np.full(size, 0.5)
        dist = defn.get('dist')
        if dist == 'fixed':
            return np.full(size, float(defn.get('value', 0.5)))
        if dist == 'gaussian':
            mu = float(defn.get('mean', 0.5)); sigma = float(defn.get('sigma', 0.1))
            out = rng.normal(loc=mu, scale=sigma, size=size)
            mask = out < 0
            attempts = 0
            while mask.any() and attempts < 100:
                out[mask] = rng.normal(loc=mu, scale=sigma, size=mask.sum())
                mask = out < 0; attempts += 1
            out[out < 0] = 0.0
            return np.clip(out, 0.0, 1.0)
        if dist == 'gamma':
            mu = float(defn.get('mean', 0.5)); sigma = float(defn.get('sigma', 0.1))
            if sigma <= 0 or mu <= 0:
                return np.full(size, mu if mu > 0 else 0.0)
            k = (mu / sigma) ** 2; theta = (sigma ** 2) / mu
            out = rng.gamma(shape=k, scale=theta, size=size)
            return np.clip(out, 0.0, 1.0)
        if dist == 'uniform':
            lo = float(defn.get('min', 0.0)); hi = float(defn.get('max', 1.0))
            return rng.uniform(lo, hi, size=size)
        return np.full(size, 0.5)
    new_ar = sample_alpha(entry.get('alpha_rad'), n)
    new_aa = sample_alpha(entry.get('alpha_azi'), n)
    for i in range(n):
        mm.at[i, ar_cols[0]] = float(new_ar[i])
        mm.at[i, aa_cols[0]] = float(new_aa[i])

# write back
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
wb = load_workbook(fp)
if 'MM_PSF' in wb.sheetnames: del wb['MM_PSF']
ws = wb.create_sheet('MM_PSF')
for r in dataframe_to_rows(mm, index=False, header=True): ws.append(r)
# sampling log
sname = 'MM_PSF_SAMPLING'
if sname in wb.sheetnames: ws2 = wb[sname]
else:
    ws2 = wb.create_sheet(sname)
    ws2.append(['timestamp_utc','preset_name','seed','n','sr_mean','sr_std','sr_shape','sr_scale','sa_mean','sa_std','sa_shape','sa_scale'])

sr_mean = float(new_sr.mean()); sr_std = float(new_sr.std()); sa_mean = float(new_sa.mean()); sa_std = float(new_sa.std())
sr_shape = sr_scale = sa_shape = sa_scale = None
try:
    if isinstance(entry.get('sigma_rad'), dict) and entry.get('sigma_rad').get('dist') == 'gamma':
        mu = float(entry.get('sigma_rad').get('mean',0.0)); sigma = float(entry.get('sigma_rad').get('sigma',0.0))
        if sigma > 0 and mu > 0:
            sr_shape = (mu / sigma) ** 2
            sr_scale = (sigma ** 2) / mu
    if isinstance(entry.get('sigma_azi'), dict) and entry.get('sigma_azi').get('dist') == 'gamma':
        mu = float(entry.get('sigma_azi').get('mean',0.0)); sigma = float(entry.get('sigma_azi').get('sigma',0.0))
        if sigma > 0 and mu > 0:
            sa_shape = (mu / sigma) ** 2
            sa_scale = (sigma ** 2) / mu
except Exception:
    pass
ws2.append([datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'), str(chosen), int(h), int(n), sr_mean, sr_std, sr_shape, sr_scale, sa_mean, sa_std, sa_shape, sa_scale])
wb.save(fp)
print('Completed sampling and wrote to', fp)
