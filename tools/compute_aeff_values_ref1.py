#!/usr/bin/env python3
"""Compute and cache VLOOKUP formula results - now corrected."""

import openpyxl
import sys

def compute_aeff_values(excel_path):
    """Compute VLOOKUP results and write them as values to the A_eff sheet."""
    
    print(f"Processing {excel_path}...")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Get the MM configuration sheet to map MM -> Row #
    mm_config = wb['MM configuration']
    mm_to_row = {}
    for row in mm_config.iter_rows(min_row=2, max_row=mm_config.max_row, min_col=3, max_col=4):
        row_cell = row[0]  # Column C (Row #)
        mm_cell = row[1]   # Column D (MM #)
        
        if mm_cell.value is not None and row_cell.value is not None:
            try:
                mm = int(float(mm_cell.value))
                row_num = int(float(row_cell.value))
                mm_to_row[mm] = row_num
            except Exception:
                pass
    
    print(f"  Found {len(mm_to_row)} MM to Row mappings")
    
    # Get the A_eff sheet
    aeff = wb['A_eff']
    
    # Read source data from rows 5-6 and 6-20, columns T-AA (20-27)
    # Row 5: System Margin values for each energy
    # Rows 6-20: A_eff values and focal lengths
    # The lookup table at S24:AA38 contains calculated values keyed by row number (1-15)
    
    system_margins = {}  # col_idx -> margin
    for col_idx in range(20, 28):  # T to AA (columns 20-27)
        val = aeff.cell(5, col_idx).value
        if val is not None:
            try:
                system_margins[col_idx] = float(val)
            except Exception:
                system_margins[col_idx] = 0
    
    print(f"  Found {len(system_margins)} energy columns with system margins")
    
    # Build lookup table from rows 6-20
    lookup_data = {}  # (lookup_row_num, col_idx) -> value
    
    for source_row_num in range(6, 21):
        lookup_row = source_row_num - 5
        focal_len_cell = aeff.cell(source_row_num, 19)
        if focal_len_cell.value is None:
            break
        try:
            focal_len = float(focal_len_cell.value)
        except Exception:
            focal_len = 0
        
        for col_idx in range(20, 28):
            base_val = aeff.cell(source_row_num, col_idx).value
            if base_val is None:
                continue
            try:
                base_val = float(base_val)
                margin = system_margins.get(col_idx, 0)
                
                if focal_len != 0:
                    adjusted = base_val * (1.0 - margin) / focal_len
                else:
                    adjusted = 0
                
                lookup_data[(lookup_row, col_idx)] = adjusted
            except Exception:
                pass
    
    print(f"  Calculated {len(lookup_data)} adjusted A_eff values from lookup table")
    
    computed_count = 0
    skipped_count = 0
    
    for row_num in range(2, aeff.max_row + 1):
        mm_cell = aeff.cell(row_num, 1)
        
        if mm_cell.value is None or mm_cell.value == '':
            break
        
        try:
            mm = int(float(mm_cell.value))
        except Exception:
            break
        
        if mm not in mm_to_row:
            skipped_count += 1
            continue
        
        lookup_row = mm_to_row[mm]
        
        for col_idx in range(10, 18):
            energy_col = col_idx + 10
            cell = aeff.cell(row_num, col_idx)
            if cell.data_type == 'f':
                if (lookup_row, energy_col) in lookup_data:
                    value = lookup_data[(lookup_row, energy_col)]
                    cell.value = round(value, 10)
                    cell.data_type = 'n'
                    computed_count += 1
                else:
                    cell.value = 0
                    cell.data_type = 'n'
                    computed_count += 1
    
    print(f"  Replaced {computed_count} formula values with computed values")
    if skipped_count > 0:
        print(f"  Skipped {skipped_count} MM rows (not in configuration)")
    
    wb.save(excel_path)
    print(f"  ✓ Saved {excel_path}")

if __name__ == '__main__':
    try:
        path = 'Distributions/20260209_Reference1.xlsx'
        compute_aeff_values(path)
        print("\n✓ Successfully computed and cached all A_eff preset values!")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
