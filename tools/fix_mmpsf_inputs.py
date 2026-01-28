"""
Fix MM_PSF per-MM rows in generated workbooks under `sensitivity/input`.

This script:
- Loads standard MM_PSF definitions from the baseline workbook
- Scans each .xlsx in `sensitivity/input`
- If a workbook's `MM_PSF` sheet contains a preset that matches a standard
  definition that provides fixed sigma/alpha values, it enforces those
  fixed numeric values into the first N per-MM rows (N inferred from the
  MM configuration in the baseline or from the sheet length).
- Saves a backup of the original as `<name>.orig.xlsx` and overwrites the
  original file with the fixed workbook.

Usage:
  python3 tools/fix_mmpsf_inputs.py

"""
from pathlib import Path
import shutil
import pandas as pd
import re
import sys

ROOT = Path(__file__).resolve().parents[1]
SENS = ROOT / 'sensitivity'
INPUT_DIR = SENS / 'input'
BASELINE = ROOT / 'Distributions' / 'Test_Distribution.xlsx'

# import helper from sensitivity_run if possible
try:
    from sensitivity.sensitivity_run import load_standard_mm_psf_defs
except Exception:
    # fallback: try to import by path
    sys.path.insert(0, str(ROOT))
    from sensitivity.sensitivity_run import load_standard_mm_psf_defs

std_defs = {}
if BASELINE.exists():
    try:
        std_defs = load_standard_mm_psf_defs(BASELINE)
    except Exception:
        std_defs = {}

# helper normalization
def _norm(s):
    if s is None:
        return ''
    return re.sub(r"\s+"," ", str(s)).strip().lower()

files = sorted(INPUT_DIR.glob('*.xlsx'))
if not files:
    print('No .xlsx files found in', INPUT_DIR)
    sys.exit(0)

fixed_count = 0
for fp in files:
    try:
        print('Processing', fp.name)
        # load MM_PSF sheet
        try:
            mm_df = pd.read_excel(fp, sheet_name='MM_PSF', engine='openpyxl')
        except Exception:
            print('  - no MM_PSF sheet, skipping')
            continue
        # detect candidate preset name by scanning cells for any std_defs key
        chosen = None
        for k in std_defs.keys():
            kn = _norm(k)
            # look for the key string anywhere in the sheet (header and cells)
            found = False
            for col in mm_df.columns:
                try:
                    ser = mm_df[col].astype(str).fillna('')
                    if any(ser.str.strip().str.lower() == kn) or any(ser.str.lower().str.contains(kn)):
                        found = True
                        break
                except Exception:
                    continue
            if found:
                chosen = k
                break
        if not chosen:
            # fallback: try to infer preset from filename: look for 'MM_PSF' token
            m = re.search(r'MM_PSF([^_]*)', fp.name)
            if m:
                cand = m.group(1).replace('_',' ').strip()
                if cand:
                    # try to match by substring
                    for k in std_defs.keys():
                        if cand.lower() in k.lower() or k.lower() in cand.lower():
                            chosen = k
                            break
        if not chosen:
            print('  - no matching preset found, skipping')
            continue

        entry = std_defs.get(chosen)
        if not entry:
            print('  - no std entry for', chosen, 'skip')
            continue

        # find sigma/alpha canonical columns
        col_map = {}
        for c in mm_df.columns:
            cl = _norm(c)
            if 'sigma_rad' in cl and 'sigma_rad_' not in cl:
                col_map['sr'] = c
            if 'sigma_azi' in cl and 'sigma_azi_' not in cl:
                col_map['sa'] = c
            if 'sigma_rad_' in cl and 'sr_tail' not in col_map:
                col_map['sr_tail'] = c
            if 'sigma_azi_' in cl and 'sa_tail' not in col_map:
                col_map['sa_tail'] = c
            if 'alpha_rad' in cl:
                col_map['ar'] = c
            if 'alpha_azi' in cl:
                col_map['aa'] = c

        # determine number of MM rows to enforce: prefer 26 or mm_df length
        num_enforce = min(26, len(mm_df)) if len(mm_df) >= 1 else 26

        def _val(key):
            v = entry.get(key)
            if isinstance(v, dict):
                if v.get('dist') == 'fixed':
                    return float(v.get('value', 0.0))
                if v.get('dist') == 'gaussian':
                    return float(v.get('mean', 0.0))
            return None

        v_sr = _val('sigma_rad')
        v_sa = _val('sigma_azi')
        v_ar = _val('alpha_rad')
        v_aa = _val('alpha_azi')

        # infer overall PSF type from the standard preset: if any alpha is fixed,
        # treat as a pseudo-voigt (has alpha), otherwise gaussian.
        def _is_alpha_fixed(pname):
            pv = entry.get(pname)
            return isinstance(pv, dict) and pv.get('dist') == 'fixed'

        is_voigt = _is_alpha_fixed('alpha_rad') or _is_alpha_fixed('alpha_azi')
        psf_type_str = 'pseudo-voigt' if is_voigt else 'gaussian'

        if v_sr is None and v_sa is None and v_ar is None and v_aa is None:
            print('  - no fixed numeric params for preset, skipping')
            continue

        # Backup original
        bak = fp.with_suffix(fp.suffix + '.orig')
        if not bak.exists():
            shutil.copy2(fp, bak)

        # Open workbook with openpyxl to write canonical cells
        from openpyxl import load_workbook
        wb = load_workbook(fp)
        if 'MM_PSF' not in wb.sheetnames:
            print('  - MM_PSF sheet not present in workbook (unexpected), skipping')
            continue
        ws = wb['MM_PSF']
        # Attempt to locate header row by searching for a row that contains header tokens
        header_row = 1
        max_scan = min(10, ws.max_row)
        found_hr = False
        key_frags = ['sigma_rad', 'sigma_azi', 'mm']
        for r in range(1, max_scan + 1):
            row_vals = [str(ws.cell(row=r, column=c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
            matches = sum(1 for kf in key_frags if any(kf in v for v in row_vals))
            if matches >= 1:
                header_row = r
                found_hr = True
                break
        if not found_hr:
            header_row = 1

        # map header names -> column indices for tolerant matching
        hdr_map = {}
        for c in range(1, ws.max_column + 1):
            hv = str(ws.cell(row=header_row, column=c).value or '').strip().lower()
            hdr_map[c] = hv
        def _col_idx_for(fragment):
            for c, hv in hdr_map.items():
                if fragment in hv:
                    return c
            for c, hv in hdr_map.items():
                if hv.replace(' ', '').startswith(fragment.replace(' ', '')):
                    return c
            return None

        sr_idx = _col_idx_for('sigma_rad')
        sa_idx = _col_idx_for('sigma_azi')
        ar_idx = _col_idx_for('alpha_rad')
        aa_idx = _col_idx_for('alpha_azi')

        if not any((sr_idx, sa_idx, ar_idx, aa_idx)):
            print('  - could not find any sigma/alpha columns in header, skipping')
            continue

        # Apply fixed values into first num_enforce data rows (assume header_row + 1..)
        applied = False
        for i in range(header_row + 1, min(ws.max_row, header_row + num_enforce) + 1):
            if sr_idx and v_sr is not None:
                ws.cell(row=i, column=sr_idx, value=float(v_sr))
                applied = True
            if sa_idx and v_sa is not None:
                ws.cell(row=i, column=sa_idx, value=float(v_sa))
                applied = True

            # Set PSF type column (F / index 6) according to inferred preset type
            try:
                ws.cell(row=i, column=6, value=psf_type_str)
                applied = True
            except Exception:
                pass

            # For alpha columns: write numeric value only if the standard preset
            # defines a fixed alpha; otherwise write '-' to indicate not-applicable.
            if ar_idx:
                aentry = entry.get('alpha_rad')
                if isinstance(aentry, dict) and aentry.get('dist') == 'fixed':
                    try:
                        ws.cell(row=i, column=ar_idx, value=float(aentry.get('value', 0.0)))
                    except Exception:
                        ws.cell(row=i, column=ar_idx, value='-')
                else:
                    ws.cell(row=i, column=ar_idx, value='-')
                applied = True

            if aa_idx:
                aentry = entry.get('alpha_azi')
                if isinstance(aentry, dict) and aentry.get('dist') == 'fixed':
                    try:
                        ws.cell(row=i, column=aa_idx, value=float(aentry.get('value', 0.0)))
                    except Exception:
                        ws.cell(row=i, column=aa_idx, value='-')
                else:
                    ws.cell(row=i, column=aa_idx, value='-')
                applied = True

        if not applied:
            print('  - nothing applied (no matching columns), skipping save')
            continue

        wb.save(fp)
        fixed_count += 1
        print('  - fixed and saved', fp.name)

    except Exception as e:
        print('  - error processing', fp.name, str(e))

print(f"Completed. Fixed {fixed_count} workbook(s).")
