#!/usr/bin/env python3
"""Run sensitivity combinations defined in Distributions/Sensitivity.xlsx.

This script:
- Parses Sensitivity.xlsx sheet1 for columns of options.
- For each combination, loads the base `Distributions/Test_Distribution.xlsx` MM_PSF data,
  then applies A_eff and MM_PSF modifications (fixed and variable gaussian cases supported).
- Computes HEW/EEF metrics using `plot_sum(..., return_metrics_only=True)`.
- Optionally runs optimizer if the combination specifies `optimize` with a mode.
- Writes results to `Distributions/sensitivity_results.xlsx`.

Notes:
- Heuristics are used to locate A_eff numeric columns and standard MM_PSF parameter tables.
- Variable gaussian draws are reproducible per-combination using a deterministic RNG seed.
"""

from pathlib import Path
import itertools
                # Process combos in parallel using threads to speed up runs.
                # Each worker receives a combo tuple and returns a result dict.
                def _process_single(combo_tuple):
                    combo_dict = dict(zip(keys, combo_tuple))
                    try:
                        h = hashlib.sha1(repr(combo_dict).encode('utf8')).hexdigest()
                        seed = int(h[:16], 16) % (2**32)
                        rng = np.random.default_rng(seed)

                        df_local = load_gaussians_from_excel(str(BASE_WORKBOOK), sheet='MM_PSF')

                        if 'A_eff' in combo_dict:
                            aeff_choice = combo_dict['A_eff']
                            try:
                                mapping = find_aeff_weights_for_choice(aeff_choice, aeff_map)
                                df_local['weight'] = df_local['MM #'].astype(int).map(mapping)
                            except Exception:
                                mapping = None

                        if 'MM_PSF' in combo_dict:
                            df_local = apply_mm_psf_choice_to_df(df_local, combo_dict['MM_PSF'], aeff_map, rng, standard_presets=standard_presets)

                        # For coarse mode we rely on the fast subprocess which now
                        # always applies an elliptical placement during metrics run.
                        metrics = None
                        try:
                            metrics = _run_plot_sum_subprocess(df_local, mode=SENS_MODE, timeout_s=10.0)
                            if not isinstance(metrics, dict):
                                metrics = None
                        except Exception:
                            metrics = None

                        # If metrics are present, treat them as post-placement results
                        post_metrics_raw = metrics if isinstance(metrics, dict) else None

                        keep = [
                            'hew_origin_arcsec',
                            'hew_best_arcsec',
                            'hew_opt_arcsec',
                            'eef90_origin_arcsec',
                            'eef90_best_arcsec',
                            'eef90_opt_arcsec',
                        ]
                        trimmed = {k: (metrics.get(k) if isinstance(metrics, dict) else None) for k in keep}

                        try:
                            raw_hew = post_metrics_raw.get('hew_opt_arcsec') if (post_metrics_raw and isinstance(post_metrics_raw, dict)) else None
                            raw_eef90 = post_metrics_raw.get('eef90_opt_arcsec') if (post_metrics_raw and isinstance(post_metrics_raw, dict)) else None
                        except Exception:
                            raw_hew = None
                            raw_eef90 = None

                        clamped_hew = None
                        clamped_eef90 = None
                        try:
                            if raw_hew is not None and trimmed.get('hew_best_arcsec') is not None:
                                clamped_hew = raw_hew if raw_hew <= trimmed['hew_best_arcsec'] else trimmed['hew_best_arcsec']
                            elif trimmed.get('hew_opt_arcsec') is not None:
                                clamped_hew = trimmed.get('hew_opt_arcsec')
                            if raw_eef90 is not None and trimmed.get('eef90_best_arcsec') is not None:
                                clamped_eef90 = raw_eef90 if raw_eef90 <= trimmed['eef90_best_arcsec'] else trimmed['eef90_best_arcsec']
                            elif trimmed.get('eef90_opt_arcsec') is not None:
                                clamped_eef90 = trimmed.get('eef90_opt_arcsec')
                        except Exception:
                            clamped_hew = trimmed.get('hew_opt_arcsec') if trimmed.get('hew_opt_arcsec') is not None else None
                            clamped_eef90 = trimmed.get('eef90_opt_arcsec') if trimmed.get('eef90_opt_arcsec') is not None else None

                        combo_out = dict(combo_dict)
                        # Normalize A_eff label and extract row number
                        row_num = '-'
                        aeff_val = combo_out.get('A_eff')
                        if aeff_val is not None:
                            mrow = re.match(r'^(.*?)\s*\[row(\d+)\]\s*$', str(aeff_val), re.IGNORECASE)
                            if mrow:
                                combo_out['A_eff'] = normalize_aeff_label(mrow.group(1).strip(), aeff_map)
                                try:
                                    row_num = int(mrow.group(2))
                                except Exception:
                                    row_num = '-'
                        if 'A_eff' in combo_out:
                            try:
                                combo_out['A_eff'] = normalize_aeff_label(combo_out.get('A_eff'), aeff_map)
                            except Exception:
                                pass
                        combo_out['row'] = row_num

                        try:
                            trimmed['hew_opt_raw_arcsec'] = raw_hew
                            trimmed['eef90_opt_raw_arcsec'] = raw_eef90
                            trimmed['hew_opt_arcsec'] = clamped_hew
                            trimmed['eef90_opt_arcsec'] = clamped_eef90
                            trimmed['placement_improved'] = (raw_hew is not None and trimmed.get('hew_best_arcsec') is not None and raw_hew < trimmed.get('hew_best_arcsec'))
                        except Exception:
                            trimmed['hew_opt_raw_arcsec'] = None
                            trimmed['eef90_opt_raw_arcsec'] = None
                            trimmed['placement_improved'] = False

                        # Parse MM_PSF details for reporting
                        def _parse_mm_psf_details(s):
                            s0 = '' if s is None else str(s)
                            ss = s0.strip()
                            sym = ''
                            size = ''
                            fv = ''
                            m_pct = re.search(r'(\d+(?:\.\d+)?)\s*%', ss)
                            pct_token = None
                            if m_pct:
                                pct_token = m_pct.group(1).strip()
                                ss = re.sub(r'\d+(?:\.\d+)?\s*%', '', ss)
                            if re.search(r'\bSym\b|Symmetric|Symmetry', ss, re.IGNORECASE):
                                sym = 'Sym'
                            elif re.search(r'\bAsym\b|Asymmetric', ss, re.IGNORECASE):
                                sym = 'Asym'
                            if re.search(r'\bFixed\b', ss, re.IGNORECASE):
                                fv = 'Fixed'
                            elif pct_token is not None:
                                fv = f'{pct_token}% Variable'
                            elif re.search(r'\bVariable\b|gaussian\(|Variable Sym', ss, re.IGNORECASE):
                                fv = 'Variable'
                            else:
                                if ss and not re.search(r'[0-9]', ss):
                                    fv = 'Preset'
                            m_pair = re.search(r'sigma_rad\s*=\s*([^,;\)]+).*sigma_azi\s*=\s*([^,;\)]+)', ss, re.IGNORECASE)
                            if m_pair:
                                size = f"{m_pair.group(1).strip()} * {m_pair.group(2).strip()}"
                            else:
                                m_nums = re.findall(r'([0-9]+\.?[0-9]*\s*(?:micron|µm|um|"|arcsec|arcsecond))', ss, re.IGNORECASE)
                                if m_nums:
                                    size = ' * '.join([m.strip() for m in m_nums])
                                else:
                                    q = re.search(r'"([^"]+)"', ss)
                                    if q:
                                        size = q.group(1)
                                    else:
                                        size = ss
                            return sym, size, fv

                        sym, size_tok, fv = _parse_mm_psf_details(combo_out.get('MM_PSF'))
                        combo_out['MM_symmetry'] = sym
                        mmpsf_txt = str(combo_out.get('MM_PSF','') or '')
                        if 'voigt' in mmpsf_txt.lower() or 'pseudo-voigt' in mmpsf_txt.lower():
                            combo_out['MM_symmetry'] = combo_out.get('MM_symmetry') or 'Asym'
                        if 'voigt' in mmpsf_txt.lower() or 'pseudo-voigt' in mmpsf_txt.lower():
                            combo_out['MM_PSF_type'] = 'Pseudo-Voigt'
                        elif 'gaussian' in mmpsf_txt.lower():
                            combo_out['MM_PSF_type'] = 'Gaussian'
                        else:
                            combo_out['MM_PSF_type'] = 'Custom'
                        combo_out['MM_size_token'] = size_tok
                        combo_out['MM_fixed_variable'] = fv

                        row = {**combo_out, **trimmed}
                        return {'success': True, 'result': row}
                    except Exception as e:
                        tb = traceback.format_exc()
                        return {'success': False, 'error': str(e), 'traceback': tb, 'combo': combo_tuple}

                max_workers = min(8, max(1, (os.cpu_count() or 1) - 1))
                print(f'Running {len(combos)} combinations using {max_workers} threads')
                futures = []
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as exe:
                    for combo in combos:
                        futures.append(exe.submit(_process_single, combo))
                    completed = 0
                    for fut in concurrent.futures.as_completed(futures):
                        res = fut.result()
                        completed += 1
                        print(f'[ {completed}/{len(combos)} ] processed')
                        if res.get('success'):
                            results.append(res['result'])
                        else:
                            err_row = {**dict(zip(keys, res.get('combo', []))), 'error': res.get('error')}
                            results.append(err_row)
                        # autosave periodically
                        if (len(results) % SAVE_EVERY) == 0:
                            try:
                                pd.DataFrame(results).to_excel(OUT_XLSX, index=False)
                                print(f'  Autosaved {len(results)} rows to {OUT_XLSX}')
                            except Exception as e:
                                print('  Warning: failed to autosave partial results:', e)
            'after_sigmay': a_sigy,
            'before_mux': b_mux,
            'after_mux': a_mux,
            'before_muy': b_muy,
            'after_muy': a_muy,
            'changed': changed,
        })

    try:
        pd.DataFrame(rows).to_csv(outdir / f'apply_mm_psf_perrow_{_apply_ts}.csv', index=False)
    except Exception:
        pass

    # Post-write assertion: detect unexpected overwrites where a previously non-nan
    # centroid (mux/muy) becomes exact zero in the after snapshot.
    try:
        for r in rows:
            b_mux = r.get('before_mux')
            a_mux = r.get('after_mux')
            b_muy = r.get('before_muy')
            a_muy = r.get('after_muy')
            # treat NaN as missing; only assert when before was finite and non-zero magnitude
            if not pd.isna(b_mux) and not pd.isna(a_mux):
                if abs(float(b_mux)) > 1e-15 and float(a_mux) == 0.0:
                    raise AssertionError(f'Unexpected mux zeroing during apply_mm_psf for MM {r.get("MM_or_index")}: before={b_mux} after={a_mux}')
            if not pd.isna(b_muy) and not pd.isna(a_muy):
                if abs(float(b_muy)) > 1e-15 and float(a_muy) == 0.0:
                    raise AssertionError(f'Unexpected muy zeroing during apply_mm_psf for MM {r.get("MM_or_index")}: before={b_muy} after={a_muy}')
    except AssertionError:
        # re-raise so caller can observe and debug; files already written
        raise
    return df


def _sample_param_array(spec, rng, size: int):
    """Return an array of length `size` sampled according to `spec`.

    Spec may be None, a numeric value, or a string like '0.05 micron' or '1.2 arcsec'.
    Returns None when spec is None or cannot be parsed.
    """
    if spec is None:
        return None
    # simple numeric
    try:
        if isinstance(spec, (int, float)):
            return np.full(size, float(spec))
        if isinstance(spec, (list, tuple, np.ndarray)):
            arr = np.asarray(spec)
            if arr.size == 1:
                return np.full(size, float(arr.flatten()[0]))
            # if array matches size, return, else broadcast/truncate
            if arr.size == size:
                return arr.astype(float)
            return np.resize(arr.astype(float), size)
    except Exception:
        pass
    s = str(spec).strip()
    m = re.search(r'([0-9]+\.?[0-9]*)(?:\s*(micron|µm|um|micro|micrometer|arcsec|arcsecond|"|\'))?', s, re.IGNORECASE)
    if m:
        val = float(m.group(1))
        unit = m.group(2)
        arcsec_to_m = 12 * np.pi / 180 / 3600
        if unit is None:
            # assume micron
            return np.full(size, val * 1e-6)
        u = unit.strip().lower()
        if u in ('micron', 'µm', 'um', 'micro', 'micrometer'):
            return np.full(size, val * 1e-6)
        else:
            return np.full(size, val * arcsec_to_m)
    # fallback: could not parse
    return None


def _run_plot_sum_subprocess(df, mode='coarse', timeout_s=5.0):
    """Run main.py on a temporary copy of BASE_WORKBOOK with `MM_PSF` replaced by `df`.

    Returns parsed JSON dict on success or raises/returns None on failure.
    """
    import tempfile, shutil, json
    from openpyxl import load_workbook

    tmpdir = tempfile.mkdtemp(prefix='sensitivity_')
    tmp_in = os.path.join(tmpdir, os.path.basename(str(BASE_WORKBOOK)))
    try:
        shutil.copy2(str(BASE_WORKBOOK), tmp_in)
        try:
            wb = load_workbook(tmp_in)
            if 'MM_PSF' in wb.sheetnames:
                std = wb['MM_PSF']
                wb.remove(std)
            wb.save(tmp_in)
            with pd.ExcelWriter(tmp_in, engine='openpyxl', mode='a') as w:
                df.to_excel(w, sheet_name='MM_PSF', index=False)
        except Exception:
            with pd.ExcelWriter(tmp_in, engine='openpyxl') as w:
                df.to_excel(w, sheet_name='MM_PSF', index=False)

        cmd = [
            sys.executable,
            str(Path(__file__).resolve().parent.parent / 'main.py'),
            '-f', tmp_in,
            '--return_metrics_only',
            '--placement', 'elliptical',
            '--metrics-nr-final', '300',
            '--metrics-ntheta-final', '24',
            '--metrics-r-margin', '6.0'
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout_s)
        if proc.returncode != 0:
            try:
                return {'error': proc.stderr.strip(), 'stdout': proc.stdout, 'stderr': proc.stderr}
            finally:
                try:
                    shutil.rmtree(tmpdir)
                except Exception:
                    pass
        try:
            out = json.loads(proc.stdout)
        except Exception:
            out = {'error': 'invalid-json', 'stdout': proc.stdout, 'stderr': proc.stderr}
        return out
    finally:
        try:
            shutil.rmtree(tmpdir)
        except Exception:
            pass


def build_aeff_mapping(workbook_path: str | Path = BASE_WORKBOOK) -> dict:
    """Build a lightweight A_eff mapping used by sensitivity tools.

    Returns a dict with keys:
      - 'mm_vals': numpy array of MM # (ints)
      - 'col_names': list of column names (str) for numeric A_eff columns
      - 'col_values': list of numpy arrays, one per numeric A_eff column
      - 'letter_map': maps letters 'A','B',... to column indices (0-based)
    """
    try:
        raw = pd.read_excel(workbook_path, sheet_name='A_eff', engine='openpyxl', header=None)
    except Exception:
        return {'mm_vals': np.array([], dtype=int), 'col_names': [], 'col_values': [], 'letter_map': {}}
    header_row = 0
    for r in range(min(20, raw.shape[0])):
        v = raw.iloc[r, 0]
        if isinstance(v, str) and v.strip().lower().replace(' ', '') in {'mm#', 'mm'}:
            header_row = r
            break
    data = raw.iloc[header_row + 1 :].copy()
    if data.shape[1] < 2:
        return {'mm_vals': np.array([], dtype=int), 'col_names': [], 'col_values': [], 'letter_map': {}}
    cols = raw.iloc[header_row, :].fillna('').astype(str).tolist()
    mm_vals = pd.to_numeric(data.iloc[:, 0], errors='coerce')
    valid_idx = mm_vals.notna()
    mm_vals = mm_vals[valid_idx].astype(int).to_numpy()
    col_names = []
    col_values = []
    for j in range(1, data.shape[1]):
        col = pd.to_numeric(data.iloc[:, j], errors='coerce')
        col = col.fillna(0.0).to_numpy()[valid_idx]
        name = cols[j] if j < len(cols) else f'col{j}'
        col_names.append(str(name).strip())
        col_values.append(np.asarray(col))
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    letter_map = {letters[i]: i for i in range(min(len(col_values), len(letters)))}
    # identify numeric columns (have finite, non-zero entries)
    numeric_indices = []
    for i, col in enumerate(col_values):
        try:
            arr = np.asarray(col, dtype=float)
        except Exception:
            continue
        if arr.size == 0:
            continue
        finite = np.isfinite(arr)
        if not finite.any():
            continue
        if np.sum(np.abs(arr[finite])) <= 0.0:
            continue
        numeric_indices.append(i)

    return {
        'mm_vals': mm_vals,
        'col_names': col_names,
        'col_values': col_values,
        'letter_map': letter_map,
        'numeric_col_indices': numeric_indices,
    }


def _extract_energy_from_string(s: str) -> str | None:
    """Try to extract an energy token like '1 keV' from a header or token string.

    Returns a normalized token like '1 keV' or None when not found.
    """
    if not s:
        return None
    s = str(s)
    # common energies: keV, eV, MeV (case-insensitive)
    m = re.search(r'([0-9]+(?:\.[0-9]+)?)\s*(keV|eV|MeV)', s, re.IGNORECASE)
    if m:
        val = m.group(1)
        unit_raw = m.group(2).lower()
        unit_map = {'kev': 'keV', 'ev': 'eV', 'mev': 'MeV'}
        unit = unit_map.get(unit_raw, m.group(2))
        return f"{val} {unit}"
    # sometimes headers use '@1 keV' or 'A_eff @1 keV'
    m2 = re.search(r'@\s*([0-9]+(?:\.[0-9]+)?\s*(?:keV|eV|MeV))', s, re.IGNORECASE)
    if m2:
        return m2.group(1).strip()
    return None


def normalize_aeff_label(label: str, aeff_map: dict) -> str:
    """Return a human-friendly A_eff label for display in results.

    Strategy:
    - strip any trailing '[rowN]' tokens
    - prefer extracting an energy token from matching A_eff column headers
    - fallback to extracting an energy token from the label itself
    - otherwise return the cleaned label
    """
    if label is None:
        return ''
    s = str(label).strip()
    # remove explicit row token if present
    s = re.sub(r'\s*\[row\d+\]\s*$', '', s, flags=re.IGNORECASE).strip()
    # Try exact/contains match against known A_eff column names
    names = aeff_map.get('col_names', []) if isinstance(aeff_map, dict) else []
    for n in names:
        if not n:
            continue
        if s.lower() == str(n).lower():
            e = _extract_energy_from_string(n)
            return e or str(n).strip()
    for n in names:
        if not n:
            continue
        if s.lower() in str(n).lower():
            e = _extract_energy_from_string(n)
            return e or s
    # try to extract energy directly
    e = _extract_energy_from_string(s)
    if e:
        return e
    # if label contains an '@' use the rhs as a friendly label
    if '@' in s:
        rhs = s.split('@', 1)[1].strip()
        if rhs:
            return rhs
    # otherwise return cleaned token
    return s


def find_aeff_weights_for_choice(choice: str, aeff_map: dict) -> dict:
    """Return mapping {MM #: weight} for a given A_eff `choice` using `aeff_map`.

    Attempts exact and case-insensitive column name matching and letter tokens.
    """
    s = str(choice or '').strip()
    m = re.match(r'^(.*?)\s*\[row(\d+)\]\s*$', s, re.IGNORECASE)
    row_token = None
    if m:
        base = m.group(1).strip()
        try:
            row_token = int(m.group(2))
        except Exception:
            row_token = None
    else:
        base = s
    names = aeff_map.get('col_names', [])
    # If a row token is present, return a mapping that keeps only the weight for that
    # specific MM row (setting all others to zero) using the requested A_eff column.
    for i, n in enumerate(names):
        if n and base.lower() == n.lower():
            col = aeff_map['col_values'][i]
            mm_vals = aeff_map.get('mm_vals', [])
            if row_token is not None and 'mm_row_map' in aeff_map:
                # map row index -> MM value
                mm_for_row = aeff_map['mm_row_map'].get(row_token)
                if mm_for_row is None:
                    return {}
                # find the value for that mm in the column
                try:
                    idx = list(mm_vals).index(mm_for_row)
                    val = float(col[idx])
                except Exception:
                    val = 0.0
                return {int(m): (val if int(m) == int(mm_for_row) else 0.0) for m in mm_vals}
            return dict(zip(mm_vals, col.tolist()))
    for i, n in enumerate(names):
        if n and base.lower() in n.lower():
            col = aeff_map['col_values'][i]
            mm_vals = aeff_map.get('mm_vals', [])
            if row_token is not None and 'mm_row_map' in aeff_map:
                mm_for_row = aeff_map['mm_row_map'].get(row_token)
                if mm_for_row is None:
                    return {}
                try:
                    idx = list(mm_vals).index(mm_for_row)
                    val = float(col[idx])
                except Exception:
                    val = 0.0
                return {int(m): (val if int(m) == int(mm_for_row) else 0.0) for m in mm_vals}
            return dict(zip(mm_vals, col.tolist()))
    if len(base) == 1 and base.upper() in aeff_map.get('letter_map', {}):
        idx = aeff_map['letter_map'][base.upper()]
        col = aeff_map['col_values'][idx]
        return dict(zip(aeff_map.get('mm_vals', []), col.tolist()))
    return {}


def load_standard_mm_psf_presets(workbook_path: str | Path = BASE_WORKBOOK) -> dict:
    """Load named MM_PSF presets from the workbook if present.

    Minimal implementation: if no presets sheet found, return empty dict.
    This keeps callers tolerant when presets aren't provided.
    """
    try:
        wb = pd.read_excel(workbook_path, sheet_name=None, engine='openpyxl')
    except Exception:
        return {}
    # Heuristic: look for a sheet named 'MM_PSF_presets' or 'MM Presets'
    for name in wb.keys():
        if 'preset' in name.lower() or 'mm_psf' in name.lower() and 'preset' in name.lower():
            try:
                df = pd.read_excel(workbook_path, sheet_name=name, engine='openpyxl')
                # Convert to simple dict: row -> dict of params
                out = {}
                for _, r in df.iterrows():
                    key = str(r.iloc[0]) if len(r) > 0 else None
                    if not key or pd.isna(key):
                        continue
                    out[str(key).strip()] = r.dropna().to_dict()
                return out
            except Exception:
                continue
    return {}


def load_mm_row_map(workbook_path: str | Path = BASE_WORKBOOK) -> dict:
    """Return a mapping of row indices (1-based) to MM # values from the A_eff sheet.

    This is used to expand tokens like '1 keV [row#]' into explicit per-row
    choices. The mapping keys are simple 1..N indices corresponding to the
    data rows found under the 'MM' header in the A_eff sheet.
    """
    try:
        raw = pd.read_excel(workbook_path, sheet_name='A_eff', engine='openpyxl', header=None)
    except Exception:
        return {}
    header_row = None
    for r in range(min(40, raw.shape[0])):
        v = raw.iloc[r, 0]
        if isinstance(v, str) and v.strip().lower().replace(' ', '') in {'mm#', 'mm'}:
            header_row = r
            break
    if header_row is None:
        return {}
    data = raw.iloc[header_row + 1 :].copy()
    mm_vals = pd.to_numeric(data.iloc[:, 0], errors='coerce')
    mm_vals = mm_vals.dropna().astype(int).tolist()
    return {i + 1: mm for i, mm in enumerate(mm_vals)}


def main():
    if not SENS_PATH.exists():
        print('Sensitivity file not found at', SENS_PATH)
        sys.exit(1)

    last_mtime = None
    print('Entering watch loop for', SENS_PATH)
    try:
        while True:
            try:
                mtime = SENS_PATH.stat().st_mtime
            except Exception as e:
                print('Could not stat sensitivity file:', e)
                time.sleep(SLEEP_SECONDS)
                continue

            if last_mtime is None or mtime != last_mtime:
                last_mtime = mtime
                print('\nDetected Sensitivity.xlsx change; reloading combinations...')

                sens = pd.read_excel(SENS_PATH, sheet_name=0, engine='openpyxl', header=0)
                # Collect options per column: explicitly iterate rows and keep every
                # non-empty, non-placeholder cell (preserve order and duplicates).
                param_options = {}
                for col in sens.columns:
                    col_vals = []
                    for cell in sens[col].tolist():
                        if pd.isna(cell):
                            continue
                        s = str(cell).strip()
                        if s == '':
                            continue
                        if s.lower() in {'-', 'nan', 'none'}:
                            continue
                        col_vals.append(s)
                    if col_vals:
                        param_options[col] = col_vals
                if not param_options:
                    print('No options found in sensitivity file; waiting for changes...')
                    time.sleep(SLEEP_SECONDS)
                    continue

                print('Parameters found:', list(param_options.keys()))
                aeff_map = build_aeff_mapping()
                # attach mm_row_map to aeff_map for row-specific A_eff handling
                aeff_map['mm_row_map'] = load_mm_row_map(BASE_WORKBOOK)
                # Expand any A_eff '[row#]' tokens into explicit per-row choices
                if 'A_eff' in param_options:
                    vals = param_options['A_eff']
                    expanded = []
                    row_token_re = re.compile(r'^(.*?)\s*\[row#\]\s*$', re.IGNORECASE)
                    # prefer numeric A_eff column expansion when the user provided a
                    # single token like '1 keV [row#]' — users typically intend to
                    # sweep available energy columns rather than all 600 MM rows.
                    # Fall back to per-MM-row expansion only when necessary.
                    if len(vals) == 1:
                        m0 = row_token_re.match(vals[0])
                        if m0:
                            # User provided a base like '1 keV [row#]'. Expand that
                            # into per-row tokens '1 keV [row1]', '1 keV [row2]',...
                            # but only for numeric A_eff columns that match the base
                            # string (e.g. 'A_eff @1 keV'). If no matching column
                            # name is found, fall back to expanding by MM rows using
                            # the given base token.
                            base = m0.group(1).strip()
                            numeric_idx = aeff_map.get('numeric_col_indices', [])
                            cols_raw = aeff_map.get('col_names', [])
                            # find numeric columns whose header contains the base token
                            matching_idx = [i for i in numeric_idx if i < len(cols_raw) and base.lower() in str(cols_raw[i]).lower()]
                            mm_row_map = aeff_map.get('mm_row_map', {})
                            total_mm_rows = len(mm_row_map) if mm_row_map else 0
                            # Limit expansions to the first N rows where N equals the
                            # number of numeric A_eff columns (user expectation: 1..15).
                            n_numeric = len(numeric_idx)
                            max_row = n_numeric if n_numeric > 0 else total_mm_rows
                            if max_row <= 0:
                                max_row = total_mm_rows
                            row_keys_limited = list(range(1, max_row + 1))

                            if matching_idx and row_keys_limited:
                                # produce base [rowN] tokens (retain user-friendly base)
                                for i in matching_idx:
                                    for r in row_keys_limited:
                                        expanded.append(f"{base} [row{r}]")
                            elif row_keys_limited:
                                # no matching numeric column name — still expand by limited rows
                                for r in row_keys_limited:
                                    expanded.append(f"{base} [row{r}]")
                    if not expanded:
                        # default behavior: keep each explicit non-empty value as-is
                        for v in vals:
                            expanded.append(v)
                    param_options['A_eff'] = expanded
                standard_presets = load_standard_mm_psf_presets(BASE_WORKBOOK)

                keys = list(param_options.keys())
                combos = list(itertools.product(*(param_options[k] for k in keys)))
                print(f'Running {len(combos)} combinations')

                results = []

                # Process combos
                for idx, combo in enumerate(combos, start=1):
                    combo_dict = dict(zip(keys, combo))
                    print(f'[ {idx}/{len(combos)} ] {combo_dict}')
                    try:
                        h = hashlib.sha1(repr(combo_dict).encode('utf8')).hexdigest()
                        seed = int(h[:16], 16) % (2**32)
                        rng = np.random.default_rng(seed)

                        df = load_gaussians_from_excel(str(BASE_WORKBOOK), sheet='MM_PSF')

                        if 'A_eff' in combo_dict:
                            aeff_choice = combo_dict['A_eff']
                            try:
                                mapping = find_aeff_weights_for_choice(aeff_choice, aeff_map)
                                df['weight'] = df['MM #'].astype(int).map(mapping)
                            except Exception as e:
                                print('  Warning: could not apply A_eff choice:', e)

                        if 'MM_PSF' in combo_dict:
                            df = apply_mm_psf_choice_to_df(df, combo_dict['MM_PSF'], aeff_map, rng, standard_presets=standard_presets)

                        # Prefer in-process placement (no Excel write). Fallback to subprocess if it fails.
                        # When running sensitivity in 'coarse' mode, skip expensive
                        # in-process placement/optimization entirely and rely on the
                        # fast subprocess metric runner to keep per-combo time bounded.
                        df_opt = None
                        placed_metrics = None
                        if SENS_MODE != 'coarse':
                            try:
                                from optimize_mm_rows import load_all_sheets, _load_base_params_from_workbook, _load_position_deltas, _elliptical_place_mm_config, rebuild_df
                                sheets = load_all_sheets(str(BASE_WORKBOOK))
                                if 'MM configuration' in sheets:
                                    mm_config = sheets['MM configuration'].copy()
                                    base_params = _load_base_params_from_workbook(str(BASE_WORKBOOK))
                                    # If A_eff mapping was computed for this combo, override
                                    # the optimizer base_params weights so placement depends
                                    # on the selected A_eff choice.
                                    try:
                                        if 'mapping' in locals() and mapping is not None:
                                            mapped = base_params['MM #'].astype(int).map(mapping).fillna(0.0)
                                            total_mapped = float(mapped.sum())
                                            MIN_TOTAL_WEIGHT = 1e-12
                                            if not np.isfinite(total_mapped) or total_mapped <= MIN_TOTAL_WEIGHT:
                                                print('    Warning: total A_eff weight for this combo is zero or too small; skipping optimization')
                                                df_opt = None
                                                raise RuntimeError('skip_optimization_due_to_small_total_weight')
                                            base_params['weight'] = mapped
                                            # Note: per-MM runtime fields are preserved inside rebuild_df;
                                            # no need to copy mux/sigmax/etc. into base_params here.
                                    except Exception:
                                        pass
                                    alignment_by_pos, gravity_by_pos, thermal_by_pos = _load_position_deltas(str(BASE_WORKBOOK))
                                    # Allow Sensitivity.xlsx to override position deltas per combo.
                                    # If the combo specifies zero for Alignment/Gravity/Thermal,
                                    # replace the corresponding deltas with zeroed mappings.
                                    def _is_zero_token(v):
                                        try:
                                            if v is None:
                                                return False
                                            s = str(v).strip()
                                            if s == '':
                                                return False
                                            return float(s) == 0.0
                                        except Exception:
                                            return False

                                    def _zero_map_for_positions(mm_config_df):
                                        # determine position numbers from mm_config
                                        if mm_config_df is None:
                                            return []
                                        if 'Position #' in mm_config_df.columns:
                                            pos_series = pd.to_numeric(mm_config_df['Position #'], errors='coerce')
                                            pos_list = [int(x) for x in pos_series.dropna().astype(int).unique().tolist()]
                                        else:
                                            pos_list = list(range(1, len(mm_config_df) + 1))
                                        return pos_list

                                    # case-insensitive key matching
                                    combo_keys = {k.lower(): k for k in combo_dict.keys()}
                                    # Alignment
                                    if any(tok in combo_keys for tok in ('alignment', 'align')):
                                        raw = combo_dict.get(combo_keys.get('alignment', combo_keys.get('align')))
                                        if _is_zero_token(raw):
                                            pos_list = _zero_map_for_positions(mm_config)
                                            alignment_by_pos = {p: {'d_align_rad': 0.0, 'd_align_azi': 0.0, 'd_align_z': 0.0, 'd_align_rotz': 0.0} for p in pos_list}
                                    # Gravity
                                    if any(tok in combo_keys for tok in ('gravity', 'gravity offload', 'grav')):
                                        # choose key name present
                                        for candidate in ('gravity', 'gravity offload', 'grav'):
                                            if candidate in combo_keys:
                                                raw = combo_dict.get(combo_keys[candidate])
                                                break
                                        else:
                                            raw = None
                                        if _is_zero_token(raw):
                                            pos_list = _zero_map_for_positions(mm_config)
                                            gravity_by_pos = {p: {'d_grav_x': 0.0, 'd_grav_y': 0.0, 'd_grav_z': 0.0, 'd_grav_rotz': 0.0} for p in pos_list}
                                    # Thermal
                                    if any(tok in combo_keys for tok in ('thermal', 'therm')):
                                        raw = None
                                        for candidate in ('thermal', 'therm'):
                                            if candidate in combo_keys:
                                                raw = combo_dict.get(combo_keys[candidate])
                                                break
                                        if _is_zero_token(raw):
                                            pos_list = _zero_map_for_positions(mm_config)
                                            thermal_by_pos = {p: {'d_therm_x': 0.0, 'd_therm_y': 0.0, 'd_therm_z': 0.0, 'd_therm_rotz': 0.0} for p in pos_list}
                                    placed_mm = _elliptical_place_mm_config(
                                        mm_config,
                                        base_params,
                                        alignment_by_pos=alignment_by_pos,
                                        gravity_by_pos=gravity_by_pos,
                                        thermal_by_pos=thermal_by_pos,
                                        seed=int(seed),
                                    )
                                    final_df = rebuild_df(base_params, placed_mm)
                                    df_opt = final_df
                            except Exception as e:
                                print('  In-process placement failed; attempting subprocess fallback:', e)
                                # Subprocess fallback: create temp workbook with mapping, run optimize_mm_rows.py
                                try:
                                    import tempfile, shutil
                                    from openpyxl import load_workbook

                                    tmpdir = tempfile.mkdtemp(prefix='sensitivity_')
                                    tmp_in = os.path.join(tmpdir, os.path.basename(str(BASE_WORKBOOK)))
                                    shutil.copy2(str(BASE_WORKBOOK), tmp_in)

                                    if 'mapping' in locals() and mapping is not None:
                                        wb = load_workbook(tmp_in)
                                        sheet_name = None
                                        for name in wb.sheetnames:
                                            if str(name).strip().lower() == 'a_eff':
                                                sheet_name = name
                                                break
                                        if sheet_name is None and 'A_eff' in wb.sheetnames:
                                            sheet_name = 'A_eff'
                                        if sheet_name is None:
                                            for name in wb.sheetnames:
                                                if 'a' in name.lower() and 'eff' in name.lower():
                                                    sheet_name = name
                                                    break
                                        if sheet_name is not None:
                                            ws = wb[sheet_name]
                                            header_row = None
                                            for r in range(1, min(40, ws.max_row) + 1):
                                                v = ws.cell(row=r, column=1).value
                                                if isinstance(v, str) and v.strip().lower().replace(' ', '') in {'mm#', 'mm'}:
                                                    header_row = r
                                                    break
                                            data_start = header_row + 1 if header_row is not None else 2
                                            for r in range(data_start, ws.max_row + 1):
                                                mm_cell = ws.cell(row=r, column=1).value
                                                try:
                                                    mmn = int(mm_cell)
                                                except Exception:
                                                    continue
                                                if mmn in mapping:
                                                    ws.cell(row=r, column=2).value = float(mapping[mmn])
                                            wb.save(tmp_in)

                                    tmp_out = os.path.join(tmpdir, 'placed_output.xlsx')
                                    cmd = [
                                        sys.executable,
                                        str(ROOT / 'optimize_mm_rows.py'),
                                        '-f', tmp_in,
                                        '-o', tmp_out,
                                        '--optimize',
                                        '--mode', SENS_MODE,
                                        '--placement', 'elliptical',
                                    ]
                                    proc = subprocess.run(cmd, capture_output=True, text=True)
                                    if proc.returncode != 0:
                                        print('    Subprocess placement failed:', proc.stderr)
                                        df_opt = None
                                    else:
                                        try:
                                            from optimize_mm_rows import load_all_sheets, _load_base_params_from_workbook, rebuild_df
                                            sheets2 = load_all_sheets(tmp_out)
                                            if 'MM configuration' in sheets2:
                                                mm_config2 = sheets2['MM configuration'].copy()
                                                base_params2 = _load_base_params_from_workbook(tmp_out)
                                                final_df2 = rebuild_df(base_params2, mm_config2)
                                                df_opt = final_df2
                                            else:
                                                df_opt = None
                                        except Exception as ee:
                                            print('    Failed to load placed output:', ee)
                                            df_opt = None
                                    try:
                                        shutil.rmtree(tmpdir)
                                    except Exception:
                                        pass
                                except Exception as ex:
                                    print('    Subprocess fallback failed:', ex)
                                    df_opt = None

                        # If in-process placement produced an optimized dataframe, pass
                        # it to plot_sum so optimized HEW/EEF metrics are computed.
                        try:
                            # Run plot_sum in a subprocess on a temporary workbook with a 5s timeout
                            metrics = _run_plot_sum_subprocess(df, mode=SENS_MODE, timeout_s=5.0)
                            if not isinstance(metrics, dict):
                                metrics = None
                        except Exception:
                            metrics = None
                        if isinstance(metrics, dict) and 'placed_metrics' in locals() and placed_metrics:
                            for kk, vv in placed_metrics.items():
                                metrics[kk] = vv

                        keep = [
                            'hew_origin_arcsec',
                            'hew_best_arcsec',
                            'hew_opt_arcsec',
                            'eef90_origin_arcsec',
                            'eef90_best_arcsec',
                            'eef90_opt_arcsec',
                        ]
                        trimmed = {k: (metrics.get(k) if isinstance(metrics, dict) else None) for k in keep}

                        # Post-placement raw metrics (before clamping)
                        post_metrics_raw = None
                        if df_opt is not None:
                            try:
                                post_metrics_raw = _run_plot_sum_subprocess(df_opt, mode=SENS_MODE, timeout_s=5.0)
                                if not isinstance(post_metrics_raw, dict):
                                    post_metrics_raw = None
                            except Exception:
                                post_metrics_raw = None

                        # Use raw post-placement metrics to decide improvement, but store a clamped value
                        try:
                            raw_hew = post_metrics_raw.get('hew_opt_arcsec') if (post_metrics_raw and isinstance(post_metrics_raw, dict)) else None
                            raw_eef90 = post_metrics_raw.get('eef90_opt_arcsec') if (post_metrics_raw and isinstance(post_metrics_raw, dict)) else None
                        except Exception:
                            raw_hew = None
                            raw_eef90 = None

                        # Clamp optimized metrics to be no worse than 'best' metrics for reporting
                        clamped_hew = None
                        clamped_eef90 = None
                        try:
                            if raw_hew is not None and trimmed.get('hew_best_arcsec') is not None:
                                clamped_hew = raw_hew if raw_hew <= trimmed['hew_best_arcsec'] else trimmed['hew_best_arcsec']
                            elif trimmed.get('hew_opt_arcsec') is not None:
                                clamped_hew = trimmed.get('hew_opt_arcsec')
                            else:
                                clamped_hew = None
                            if raw_eef90 is not None and trimmed.get('eef90_best_arcsec') is not None:
                                clamped_eef90 = raw_eef90 if raw_eef90 <= trimmed['eef90_best_arcsec'] else trimmed['eef90_best_arcsec']
                            elif trimmed.get('eef90_opt_arcsec') is not None:
                                clamped_eef90 = trimmed.get('eef90_opt_arcsec')
                            else:
                                clamped_eef90 = None
                        except Exception:
                            clamped_hew = trimmed.get('hew_opt_arcsec') if trimmed.get('hew_opt_arcsec') is not None else None
                            clamped_eef90 = trimmed.get('eef90_opt_arcsec') if trimmed.get('eef90_opt_arcsec') is not None else None

                        # extract row number from A_eff choices like '1 keV [row3]'
                        row_num = '-'
                        combo_out = dict(combo_dict)
                        # Attach diagnostics: raw post-placement metrics and whether placement improved HEW
                        try:
                            trimmed['hew_opt_raw_arcsec'] = raw_hew
                            trimmed['eef90_opt_raw_arcsec'] = raw_eef90
                            trimmed['hew_opt_arcsec'] = clamped_hew
                            trimmed['eef90_opt_arcsec'] = clamped_eef90
                            trimmed['placement_improved'] = (raw_hew is not None and trimmed.get('hew_best_arcsec') is not None and raw_hew < trimmed.get('hew_best_arcsec'))
                        except Exception:
                            trimmed['hew_opt_raw_arcsec'] = None
                            trimmed['eef90_opt_raw_arcsec'] = None
                            trimmed['placement_improved'] = False
                        # Parse MM_PSF into three detail columns: Sym/Asym, size token, Fixed/Variable
                        def _parse_mm_psf_details(s):
                            s0 = '' if s is None else str(s)
                            ss = s0.strip()
                            sym = ''
                            size = ''
                            fv = ''
                            # Capture percentage variability if present (e.g. '10%') and remove all percent tokens
                            m_pct = re.search(r'(\d+(?:\.\d+)?)\s*%', ss)
                            pct_token = None
                            if m_pct:
                                pct_token = m_pct.group(1).strip()
                                ss = re.sub(r'\d+(?:\.\d+)?\s*%', '', ss)
                            if re.search(r'\bSym\b|Symmetric|Symmetry', ss, re.IGNORECASE):
                                sym = 'Sym'
                            elif re.search(r'\bAsym\b|Asymmetric', ss, re.IGNORECASE):
                                sym = 'Asym'
                            # Fixed vs Variable detection
                            if re.search(r'\bFixed\b', ss, re.IGNORECASE):
                                fv = 'Fixed'
                            elif pct_token is not None:
                                fv = f'{pct_token}% Variable'
                            elif re.search(r'\bVariable\b|gaussian\(|Variable Sym', ss, re.IGNORECASE):
                                fv = 'Variable'
                            else:
                                # If it's a named preset, mark as Preset
                                if ss and not re.search(r'[0-9]', ss):
                                    fv = 'Preset'
                            # size token extraction: prefer explicit sigma params, else any numeric+unit tokens, else the whole string
                            m_pair = re.search(r'sigma_rad\s*=\s*([^,;\)]+).*sigma_azi\s*=\s*([^,;\)]+)', ss, re.IGNORECASE)
                            if m_pair:
                                size = f"{m_pair.group(1).strip()} * {m_pair.group(2).strip()}"
                            else:
                                # numeric tokens excluding percentage (we already stripped percent)
                                # Match numbers that include an explicit unit (don't capture bare integers like '10' from '10%')
                                m_nums = re.findall(r'([0-9]+\.?[0-9]*\s*(?:micron|µm|um|"|arcsec|arcsecond))', ss, re.IGNORECASE)
                                if m_nums:
                                    size = ' * '.join([m.strip() for m in m_nums])
                                else:
                                    # fallback to quoted value or entire token
                                    q = re.search(r'"([^\"]+)"', ss)
                                    if q:
                                        size = q.group(1)
                                    else:
                                        size = ss
                            return sym, size, fv

                        sym, size_tok, fv = _parse_mm_psf_details(combo_out.get('MM_PSF'))
                        combo_out['MM_symmetry'] = sym
                        # If the choice explicitly mentions a Voigt/pseudo-Voigt, treat as Asym by default
                        mmpsf_txt = str(combo_out.get('MM_PSF','') or '')
                        if 'voigt' in mmpsf_txt.lower() or 'pseudo-voigt' in mmpsf_txt.lower():
                            combo_out['MM_symmetry'] = combo_out.get('MM_symmetry') or 'Asym'
                        # record distribution type for reporting
                        if 'voigt' in mmpsf_txt.lower() or 'pseudo-voigt' in mmpsf_txt.lower():
                            combo_out['MM_PSF_type'] = 'Pseudo-Voigt'
                        elif 'gaussian' in mmpsf_txt.lower():
                            combo_out['MM_PSF_type'] = 'Gaussian'
                        else:
                            combo_out['MM_PSF_type'] = 'Custom'
                        combo_out['MM_size_token'] = size_tok
                        combo_out['MM_fixed_variable'] = fv
                        aeff_val = combo_out.get('A_eff')
                        if aeff_val is not None:
                            mrow = re.match(r'^(.*?)\s*\[row(\d+)\]\s*$', str(aeff_val), re.IGNORECASE)
                            if mrow:
                                combo_out['A_eff'] = normalize_aeff_label(mrow.group(1).strip(), aeff_map)
                                try:
                                    row_num = int(mrow.group(2))
                                except Exception:
                                    row_num = '-'
                        # Ensure any non-row A_eff tokens are also normalized for display
                        if 'A_eff' in combo_out:
                            try:
                                combo_out['A_eff'] = normalize_aeff_label(combo_out.get('A_eff'), aeff_map)
                            except Exception:
                                pass
                        combo_out['row'] = row_num
                        row = {**combo_out, **trimmed}
                        results.append(row)

                        if (idx % SAVE_EVERY) == 0:
                            try:
                                df_partial = pd.DataFrame(results)
                                df_partial.to_excel(OUT_XLSX, index=False)
                                print(f'  Autosaved {len(results)} rows to {OUT_XLSX}')
                            except Exception as e:
                                print('  Warning: failed to autosave partial results:', e)
                    except Exception as e:
                        tb = traceback.format_exc()
                        print('  ERROR in combination:', e)
                        print(tb)
                        results.append({**combo_dict, 'error': str(e)})

                # finished sweep
                try:
                    df_res = pd.DataFrame(results)
                    # Ensure 'row' column exists (should be populated per-combination earlier)
                    if 'row' not in df_res.columns:
                        df_res['row'] = '-'
                    # Reorder columns so that 'A_eff' (col A), 'row' (col B), 'MM_PSF' (col C)
                    # and the parsed MM_PSF detail columns occupy cols D-F if present.
                    cols = list(df_res.columns)
                    desired_prefix = []
                    if 'A_eff' in cols:
                        desired_prefix.append('A_eff')
                    if 'row' in cols:
                        desired_prefix.append('row')
                    if 'MM_PSF' in cols:
                        desired_prefix.append('MM_PSF')
                    # parsed detail columns in order
                    parsed_cols = ['MM_symmetry', 'MM_size_token', 'MM_fixed_variable', 'MM_PSF_type']
                    for pc in parsed_cols:
                        if pc in cols:
                            desired_prefix.append(pc)
                    # build final column order: desired_prefix first, then remaining columns in original order
                    remaining = [c for c in cols if c not in desired_prefix]
                    cols = desired_prefix + remaining
                    df_res = df_res[cols]
                    df_res.to_excel(OUT_XLSX, index=False)
                    print('Completed sweep — wrote results to', OUT_XLSX)
                except Exception as e:
                    print('Failed to write results file after sweep:', e)

                print('Waiting for changes to Sensitivity.xlsx...')

            # no change detected: sleep a bit
            time.sleep(SLEEP_SECONDS)
    except KeyboardInterrupt:
        print('\nWatcher interrupted by user; exiting.')
        


if __name__ == '__main__':
    main()
